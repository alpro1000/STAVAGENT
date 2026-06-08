#!/usr/bin/env python3
"""
Atomic worklist Excel generator — HK212 style, mezonet rekonstrukce.

Reads  outputs/atomic_decomposition_map.json (produced by the UWO sandbox
exporter: sandbox/uwo-interier-mezonet/export-atomic-map.mjs).
Writes  outputs/Vykaz_vymer_Mezonet_ATOMIC_WORKLIST.xlsx with sheets:

  1. Souhrn            — struktura, stav kódů (4 tiers), náklady orientačně, headline gap
  2. Atomic_worklist   — všech 33 atomic operací, seskupeno po kapitolách (HK212 sloupce + náklad)
  3. Decomposition_Map — parent (řádek mistra) → atomic children
  4. GAPS_vs_mistr     — atomic operace, které mistr v nabídce vynechal
  5. Sanity_flagy      — false-plausible katalogové kódy z reálného ÚRS proba

Columns (HK212 style + náklad): Poř. / Kapitola / Atomic operace (popis) / MJ /
Množství / Vzorec / Zdroj výměry / URS kód kandidát / Status /
Cena/MJ (orient.) / Náklad (orient.) Kč / Zdroj sazby / Parent item_id /
Realizuje skladbu / Pozn.

Pattern 26 honest blanks (no 999/TBD). Work-First: seznam prací = primary
deliverable; URS kód = secondary, doplní se z app.urs.cz / KROS.
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
MAP_PATH = ROOT / "outputs" / "atomic_decomposition_map.json"
TARGET = ROOT / "outputs" / "Vykaz_vymer_Mezonet_ATOMIC_WORKLIST.xlsx"

THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
KAP_FILL = PatternFill("solid", fgColor="D6E0F0")
KAP_FONT = Font(name="Calibri", size=10, bold=True)
BODY_FONT = Font(name="Calibri", size=9)
GAP_FILL = PatternFill("solid", fgColor="FFE0E0")     # gap / blank — red tint
DECOMP_FILL = PatternFill("solid", fgColor="FFF6E0")  # decomposed child — amber
CAND_FILL = PatternFill("solid", fgColor="E0F0E0")    # candidate code — green
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="top")
CENTER = Alignment(horizontal="center", vertical="top")
TITLE_FONT = Font(name="Calibri", size=14, bold=True)
SECTION_FONT = Font(name="Calibri", size=11, bold=True)
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color="555555")

COLS = ["Poř.", "Kapitola", "Atomic operace (popis)", "MJ", "Množství",
        "Vzorec / Zdroj výměry", "URS kód kandidát", "Status",
        "Cena/MJ (orient.)", "Náklad (orient.) Kč", "Zdroj sazby",
        "Parent item_id", "Realizuje skladbu", "Pozn."]
WIDTHS = [5, 22, 46, 7, 9, 34, 14, 22, 12, 15, 13, 30, 18, 40]

KAP_ORDER = ["HSV-6 Bourací práce", "PSV-71 Izolace HI", "PSV-72 ZTI",
             "PSV-73 Vytápění", "PSV-76 SDK konstrukce", "PSV-76 Truhlář",
             "PSV-76 Výplně otvorů", "PSV-77 Podlahy", "PSV-78 Povrchové úpravy",
             "M-21 ELI silnoproud", "VRN — Doprava + odpad", "VRN — Společné"]

STATUS_4TIER = {
    "candidate": "⚠ kandidát-verify",
    "group_only": "⚠ família ??? (ÚRS online)",
    "not_verified": "❌ blank — ÚRS online",
    "exact": "✓ verified",
}
URS_ONLINE_NOTE = "Doplnit leaf z CS ÚRS online (app.urs.cz) dle família + popis + Vzorec."


def kc(n):
    try:
        return f"{int(round(n)):,}".replace(",", " ")
    except (TypeError, ValueError):
        return ""


def write_header(ws, cols, widths, freeze="A2"):
    for c, (name, w) in enumerate(zip(cols, widths), start=1):
        cell = ws.cell(1, c, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


def status_4tier(op):
    return STATUS_4TIER.get(op.get("binding_status"), "❌ blank — ÚRS online")


def build_worklist(wb, ops):
    ws = wb.create_sheet("Atomic_worklist")
    write_header(ws, COLS, WIDTHS)
    by_kap = {}
    for o in ops:
        by_kap.setdefault(o["kapitola"], []).append(o)
    ordered = [k for k in KAP_ORDER if k in by_kap] + [k for k in by_kap if k not in KAP_ORDER]
    row = 2
    seq = 0
    for kap in ordered:
        kap_ops = by_kap[kap]
        kap_cost = sum(o.get("cost_czk") or 0 for o in kap_ops)
        cell = ws.cell(row, 1, value=f"━━ {kap}  ({len(kap_ops)} operací · {kc(kap_cost)} Kč) ━━")
        cell.font = KAP_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLS))
        for c in range(1, len(COLS) + 1):
            ws.cell(row, c).fill = KAP_FILL
        row += 1
        for o in kap_ops:
            seq += 1
            code = o.get("urs_kod_kandidat")
            fam = o.get("urs_code_family_6digit")
            code_disp = code if code else (f"{fam} ???" if fam else "")
            pozn = o.get("pozn") or ""
            if o.get("binding_status") in ("not_verified", "group_only"):
                pozn = (pozn + "  " + URS_ONLINE_NOTE).strip()
            vals = [
                seq, o["kapitola"], o["atomic_operace_popis"], o["mj"], o.get("mnozstvi"),
                o.get("qty_formula") or "", code_disp, status_4tier(o),
                kc(o.get("rate_czk")), kc(o.get("cost_czk")), o.get("rate_source") or "",
                o.get("parent_frozen_item_id") or "", o.get("realizuje_skladbu") or "", pozn,
            ]
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(row, c, value=v)
                cell.font = BODY_FONT
                cell.border = BORDER
                cell.alignment = RIGHT if c in (5, 9, 10) else (CENTER if c in (1, 4, 8) else LEFT)
            if o.get("gap_vs_master"):
                for c in range(1, len(COLS) + 1):
                    ws.cell(row, c).fill = GAP_FILL
            elif o.get("decomposition_type"):
                for c in range(1, len(COLS) + 1):
                    ws.cell(row, c).fill = DECOMP_FILL
            elif o.get("binding_status") == "candidate":
                ws.cell(row, 7).fill = CAND_FILL
            row += 1
    return ws, seq


def build_souhrn(wb, data):
    ws = wb.create_sheet("Souhrn")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 56
    summ = data["_summary"]
    cost = summ["cost"]
    r = 1

    def title(t):
        nonlocal r
        ws.cell(r, 1, value=t).font = TITLE_FONT
        r += 1

    def note(t):
        nonlocal r
        ws.cell(r, 1, value=t).font = NOTE_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1

    def section(t):
        nonlocal r
        c = ws.cell(r, 1, value=t)
        c.font = SECTION_FONT
        c.fill = KAP_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1

    def kv(k, v, n=""):
        nonlocal r
        ws.cell(r, 1, value=k).font = BODY_FONT
        ws.cell(r, 2, value=v).font = Font(name="Calibri", size=10, bold=True)
        if n:
            ws.cell(r, 3, value=n).font = NOTE_FONT
        r += 1

    title("ATOMIC WORKLIST — Rekonstrukce mezonetu (interiér/PSV)")
    note("HK212 atomic-operation princip · Work-First → Catalog-Last · Pattern 26 honest blanks (no 999/TBD)")
    r += 1

    section("PŘEHLED STRUKTURY")
    kv("Atomic operací celkem", summ["atomic_operations_total"])
    kv("Decomposed parents (řádky mistra → children)", summ["items_decomposed"])
    kv("Atomic children z dekompozice", summ["atomic_children_from_decomposition"])
    kv("Carried 1:1", summ["items_carried_1to1"])
    r += 1

    section("STAV KÓDŮ (4 tiers — práce kompletní bez ohledu na stav kódu)")
    tiers = {"⚠ kandidát-verify": 0, "⚠ família ??? (ÚRS online)": 0, "❌ blank — ÚRS online": 0, "✓ verified": 0}
    for o in data["atomic_operations"]:
        tiers[status_4tier(o)] = tiers.get(status_4tier(o), 0) + 1
    for k, v in tiers.items():
        kv(k, v)
    note("Privátní zakázka → ÚRS primary. find_urs nevrací leaf u licencovaných dat → família/blank; leaf doplní člověk.")
    r += 1

    section("NÁKLADY (ORIENTAČNĚ ±10–15 % — detail u dodavatele)")
    kv("Nabídka mistra (baseline)", kc(cost["master_offer_total"]) + " Kč")
    kv("UWO grand ORIENTAČNÍ", kc(cost["grand_orientacni"]) + " Kč")
    kv("  ↳ sazby mistra", kc(cost["by_rate_source"]["master"]) + " Kč")
    kv("  ↳ rule-of-thumb (doplněné mezery)", kc(cost["by_rate_source"]["rule_of_thumb"]) + " Kč")
    kv("Δ vs mistr (podhodnoceno)", f"+{kc(cost['delta_vs_master'])} Kč", f"+{cost['delta_vs_master_pct']} %")
    kv("not_verified hodnota (NE přesná cifra)", kc(cost["not_verified_value"]) + " Kč", "ceny orientační, kód nepotvrzen")
    r += 1

    section("KAPITOLY")
    for kap, n in summ["atomic_per_kapitola"].items():
        kv(kap, n)
    r += 1

    section("HEADLINE — co mistr vynechal")
    for g in summ["gaps"]:
        ws.cell(r, 1, value="• " + g).font = BODY_FONT
        r += 1
    note("malba (stěny+podhledy), hydroizolace, montáž ZP, samonivelační stěrka, CELÁ výměna kotle, ochrana schodiště, odvoz suti, administrativa, hodinové.")
    r += 1
    note("DPH NENÍ zašito — rekonstrukce bytu obvykle snížená sazba 12 %, uplatnit po ověření podmínek.")
    return ws


def build_decomp(wb, data):
    ws = wb.create_sheet("Decomposition_Map")
    cols = ["Parent (řádek mistra)", "Kapitola", "N children", "Atomic children"]
    widths = [40, 22, 11, 60]
    write_header(ws, cols, widths)
    row = 2
    for dm in data["decomposition_map"]:
        vals = [dm["parent_frozen_item_id"], dm["parent_kapitola"], dm["n_atomic_children"], ", ".join(dm["atomic_children_ids"])]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row, c, value=v)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = CENTER if c == 3 else LEFT
            cell.fill = DECOMP_FILL
        row += 1
    return ws


def build_gaps(wb, ops):
    ws = wb.create_sheet("GAPS_vs_mistr")
    cols = ["Atomic operace", "Kapitola", "MJ", "Množství", "Náklad (orient.) Kč", "Pozn."]
    widths = [46, 22, 7, 9, 16, 44]
    write_header(ws, cols, widths)
    row = 2
    total = 0
    for o in ops:
        if not o.get("gap_vs_master"):
            continue
        total += o.get("cost_czk") or 0
        vals = [o["atomic_operace_popis"], o["kapitola"], o["mj"], o.get("mnozstvi"), kc(o.get("cost_czk")), o.get("pozn") or ""]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row, c, value=v)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = RIGHT if c in (4, 5) else LEFT
            cell.fill = GAP_FILL
        row += 1
    ws.cell(row, 1, value="CELKEM mezery vs mistr").font = KAP_FONT
    ws.cell(row, 5, value=kc(total)).font = KAP_FONT
    ws.cell(row, 5).alignment = RIGHT
    return ws


def build_sanity(wb, data):
    ws = wb.create_sheet("Sanity_flagy")
    cols = ["Atomic operace", "Podezřelý kód", "Confidence", "Typ", "Problém"]
    widths = [30, 16, 11, 18, 70]
    write_header(ws, cols, widths)
    note_row = 2
    ws.cell(note_row, 1, value="Z reálného ÚRS proba — false-plausible kódy, které NESMÍ být přijaty jako platné.").font = NOTE_FONT
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
    row = 3
    for f in data["_summary"]["sanity_flags"]:
        vals = [f["atom"], f["code"], f["confidence"], f["kind"], f["issue"]]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row, c, value=v)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = CENTER if c == 3 else LEFT
            cell.fill = GAP_FILL
        row += 1
    return ws


def main():
    with MAP_PATH.open(encoding="utf8") as f:
        data = json.load(f)
    ops = data["atomic_operations"]

    wb = Workbook()
    wb.remove(wb.active)
    build_souhrn(wb, data)
    _, n = build_worklist(wb, ops)
    build_decomp(wb, data)
    build_gaps(wb, ops)
    build_sanity(wb, data)

    TARGET.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(TARGET))

    fabricated = [o for o in ops if o.get("urs_kod_kandidat") in ("999999999", "TBD", "999")]
    traceable = all(o.get("parent_frozen_item_id") for o in ops)
    print(json.dumps({
        "file": str(TARGET.relative_to(ROOT)),
        "sheets": wb.sheetnames,
        "atomic_rows": n,
        "atomic_ops_in_map": len(ops),
        "rows_match": n == len(ops),
        "decomposition_parents": len(data["decomposition_map"]),
        "gaps": len(data["_summary"]["gaps"]),
        "sanity_flags": len(data["_summary"]["sanity_flags"]),
        "fabricated_codes": len(fabricated),
        "traceability_100pct": traceable,
        "grand_orientacni_czk": data["_summary"]["cost"]["grand_orientacni"],
    }, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
