"""Generate 'Chybějící práce v2' sheet in user's KROS export file.

Format matches user's existing sheet: PČ(3) Typ(4) Kód(5) Popis(6) MJ(7) Množství(8)
J.materiál(9) J.montáž(10) Cena celkem(11) Cenová soustava(12) DPH(15) J.cena(16)
J.Nh(19) Nh celkem(20) J.hmotnost(21) Hmotnost celkem(22)

Each missing position emits one K-row (item) + one PP-row (popis položky).
Section dividers use D-row.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy
from pathlib import Path

HERE = Path(__file__).parent
SRC = HERE / "Daliborova_24_Hostivar_user.xlsx"
DST = HERE / "Daliborova_24_Hostivar_user_v2.xlsx"

wb = openpyxl.load_workbook(SRC)
# Remove old "Chybějící práce v2" sheet if exists (idempotent re-run)
for s in list(wb.sheetnames):
    if s.startswith("Chybějící"):
        del wb[s]

# Read styling from the existing sheet to mirror look
src_ws = wb['19-26 - Daliborova 24, Ho...']
# Sample style refs (from rows we know)
sample_D_style = src_ws.cell(row=127, column=6)  # D-row HSV
sample_K_style = src_ws.cell(row=132, column=6)  # K-row item
sample_PP_style = src_ws.cell(row=133, column=6)  # PP-row popis položky
sample_header_style = src_ws.cell(row=125, column=5)  # Header
sample_total_style = src_ws.cell(row=126, column=3)  # Total row

ws = wb.create_sheet("Chybějící práce v2")

# --- Title block (mirror user's krycí list layout) ---
ws['D4'] = "KRYCÍ LIST — CHYBĚJÍCÍ PRÁCE v2 (audit STAVAGENT 2026-05-27)"
ws['D4'].font = Font(bold=True, size=14)
ws['D6'] = "Stavba:"
ws['F6'] = "Daliborova 266/24, Praha 15 - Hostivař"
ws['D7'] = "Stupeň PD:"
ws['F7'] = "DSP"
ws['D8'] = "Pozn.:"
ws['F8'] = ("Tento list obsahuje pouze položky CHYBĚJÍCÍ ve vstupním soupisu uživatele. "
            "Doplněno na základě auditu TZ D.1.1 + výkresů + pre-statiky dle Eurokódů "
            "(ČSN EN 1990-1995). Pre-statika je validní pro nabídku zhotovitele; pro DPS "
            "nutno autorizovaný posudek ČKAIT IS00 + průzkum zdiva (TZ §7) + průzkum trámů.")
ws['F8'].alignment = Alignment(wrap_text=True, vertical='top')

# --- Table header (row 12 = same layout as source row 125) ---
HEADER_ROW = 12
headers = {
    3: "PČ", 4: "Typ", 5: "Kód", 6: "Popis", 7: "MJ", 8: "Množství",
    9: "J. materiál [CZK]", 10: "J. montáž [CZK]", 11: "Cena celkem [CZK]",
    12: "Cenová soustava", 15: "DPH", 16: "J.cena [CZK]",
    17: "Materiál celkem [CZK]", 18: "Montáž celkem [CZK]",
    19: "J. Nh [h]", 20: "Nh celkem [h]",
    21: "J. hmotnost [t]", 22: "Hmotnost celkem [t]",
}
header_fill = PatternFill("solid", fgColor="2F5496")
header_font = Font(bold=True, color="FFFFFF", size=10)
for c, label in headers.items():
    cell = ws.cell(row=HEADER_ROW, column=c, value=label)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# --- Position library ---
# Format: (section_l1_kod, section_l1_name, section_l2_kod, section_l2_name, kod, popis_short,
#          popis_long, mj, mn, j_mat, j_mnt, urs_flag, nh_per_unit, hmotnost_per_unit, vypocet_vv)
P = []

# ===== HSV-3 Zdivo Porotherm (NOVÁ SEKCE) =====
P += [
    ("HSV", "Práce a dodávky HSV", "3", "Svislé a kompletní konstrukce",
     "311238513", "Zdivo nosné Porotherm T Profi tl. 440 mm s integrovanou TI — vikýře",
     "Zdivo z cihel pálených broušených Porotherm T Profi 44 P10 na maltu pro tenké spáry, "
     "tl. 440 mm s integrovanou tepelnou izolací (čelní + boční stěny 2 vikýřů).",
     "m2", 86.7, 1080, 1600, "CS ÚRS 2026 01", 1.85, 0.395,
     "2*(6.78*2.49+2*3.78*3.5)"),
    ("HSV", "Práce a dodávky HSV", "3", "Svislé a kompletní konstrukce",
     "342272411", "Příčka Porotherm AKU 20 P+D tl. 200 mm — mezibytová",
     "Příčky a stěny zděné na MC 5 z cihel pálených broušených Porotherm 20 AKU P+D, "
     "tl. 200 mm. Mezibytová stěna BYT 34 / BYT 35.",
     "m2", 13.2, 690, 800, "CS ÚRS 2026 01", 1.10, 0.181,
     "4.41*3.0"),
    ("HSV", "Práce a dodávky HSV", "3", "Svislé a kompletní konstrukce",
     "342272412", "Příčka Porotherm 24 AKU P+D tl. 250 mm — instalační",
     "Příčky a stěny zděné z cihel pálených broušených Porotherm 24 AKU P+D, "
     "tl. 250 mm. Instalační stěny koupelen + WC.",
     "m2", 28.8, 820, 800, "CS ÚRS 2026 01", 1.20, 0.225,
     "6*2*2.4"),
    ("HSV", "Práce a dodávky HSV", "3", "Svislé a kompletní konstrukce",
     "342272421", "Příčka Porotherm 11.5 P+D tl. 115 mm — vnitřní bytové",
     "Příčky zděné z cihel pálených broušených Porotherm 11.5 P+D, tl. 115 mm. "
     "Vnitřní příčky v bytech.",
     "m2", 78, 520, 800, "CS ÚRS 2026 01", 0.95, 0.103,
     "2*15*2.6"),
    ("HSV", "Práce a dodávky HSV", "3", "Svislé a kompletní konstrukce",
     "342273241", "Zazdění prostupů a otvorů ve světlíku po průchodu OK",
     "Zazdění prostupů v cihelném zdivu plochy do 1 m², po průchodu ocelového profilu ve světlíku.",
     "m2", 4.0, 1100, 1050, "CS ÚRS 2026 01", 1.30, 0.220,
     "2*2"),
]

# ===== HSV-4 ŽB konstrukce (NOVÁ SEKCE) =====
P += [
    ("HSV", "Práce a dodávky HSV", "4", "Vodorovné konstrukce",
     "411121221", "Stropní věnec ŽB monolitický C20/25 XC1 — po obvodu vikýřů",
     "Stropní konstrukce ŽB monolitická C20/25 XC1, věnec šířky 250 mm × výšky 200 mm. "
     "Po obvodu 2 vikýřů + pod střední vaznicí.",
     "m3", 2.5, 3500, 1700, "CS ÚRS 2026 01", 8.50, 2.500,
     "50*0.25*0.20"),
    ("HSV", "Práce a dodávky HSV", "4", "Vodorovné konstrukce",
     "411354311", "Bednění věnce — montáž + demontáž",
     "Bednění stropů věnců pod krov, oboustranné, vč. odbednění a podpůrných konstrukcí.",
     "m2", 25.0, 185, 300, "CS ÚRS 2026 01", 0.62, 0.012,
     "50*2*0.20+5"),
    ("HSV", "Práce a dodávky HSV", "4", "Vodorovné konstrukce",
     "411361821", "Výztuž věnce B500B (4Φ12 podélná + třmínky Φ8/200 mm)",
     "Výztuž stropů věnců z betonářské oceli B500B třída C, profil 4Φ12 podélná + "
     "třmínky Φ8 po 200 mm. Dimenzováno dle ČSN EN 1992-1-1 (statika_assumed.md §8).",
     "t", 0.280, 32000, 6500, "CS ÚRS 2026 01", 28.0, 1.000,
     "4*50*0.888/1000+(2*(0.22+0.17))*0.395*5*50/1000+0.025"),
    ("HSV", "Práce a dodávky HSV", "4", "Vodorovné konstrukce",
     "317168112", "Překlad keramický Porotherm KP 7 — nad okny vikýřů (4 ks)",
     "Překlad keramický 7 plný (70 × 238 × 1500 mm), únosnost 70 kN/m. "
     "Nad vikýřovými okny 4 ks (2 ks per vikýř).",
     "ks", 4, 1450, 400, "CS ÚRS 2026 01", 0.75, 0.035,
     "4"),
    ("HSV", "Práce a dodávky HSV", "4", "Vodorovné konstrukce",
     "317941121", "Osazení ocelového válcovaného nosníku IPE 100 do nadpraží dveří",
     "Osazení ocelového válcovaného nosníku IPE 100, hmotnosti do 15 kg/ks, do podporového zdiva "
     "vč. zalití maltou MC 10. 2 dveřní otvory × 2 profily = 4 ks. Statika_assumed.md §9a.",
     "kus", 4, 280, 850, "CS ÚRS 2026 01", 1.50, 0.013,
     "2*2"),
    ("HSV", "Práce a dodávky HSV", "4", "Vodorovné konstrukce",
     "13010923", "Profil válcovaný I IPE 100 hladký S235JR (materiál pro nadpraží)",
     "Profil válcovaný I IPE 100 mm, S235JR, dl 1.5 m. Materiál pro 4 ks nadpraží.",
     "kg", 49, 35, 0, "CS ÚRS 2026 01", 0, 1.000,
     "4*1.5*8.1"),
    ("HSV", "Práce a dodávky HSV", "4", "Vodorovné konstrukce",
     "762000Z01R", "Zesílení stávajících stropních trámů 180/200 příložkou 100/260 mm",
     "Zesílení stávajících dřevěných trámů 180/200 mm jednostrannou dřevěnou příložkou "
     "100/260 mm z dolní strany, vč. PUR lepidla a svorníků M12 po 500 mm. "
     "Po zesílení W +67 %, I +130 %, splňuje vibrační požadavek ČSN EN 1995-1-1 §7.3. "
     "Příložka 5.01 m × 36 trámů = 180.4 bm. Statika_assumed.md §7.",
     "m", 180, 380, 240, "vlastní R", 0.85, 0.026,
     "36*5.01"),
]

# ===== HSV-6 Omítky (DOPLNĚNÍ — uživatel má jen strop) =====
P += [
    ("HSV", "Práce a dodávky HSV", "6", "Úpravy povrchů, podlahy a osazování výplní",
     "612311131", "Vnitřní omítka sádrová hladká stěn",
     "Omítka sádrová nebo vápenosádrová vnitřních ploch nanášená strojně jednovrstvá, "
     "tloušťky do 10 mm hladká svislých konstrukcí stěn (obvodové + vnitřní stěny v bytech).",
     "m2", 280, 95, 190, "CS ÚRS 2026 01", 0.45, 0.012,
     "280"),
    ("HSV", "Práce a dodávky HSV", "6", "Úpravy povrchů, podlahy a osazování výplní",
     "612311141", "Vnější omítka vápenocementová vikýřů s armovací tkaninou",
     "Omítka vápenocementová tenkovrstvá vyztužená armovací sklotextilní tkaninou na "
     "Porotherm T Profi 440 mm. Finální povrchová úprava v odstínu fasády.",
     "m2", 86.7, 220, 320, "CS ÚRS 2026 01", 0.85, 0.027,
     "86.7"),
    ("HSV", "Práce a dodávky HSV", "6", "Úpravy povrchů, podlahy a osazování výplní",
     "612345121", "Oprava omítek komínů nad střechou + matný bílý fasádní nátěr",
     "Oprava vnějších omítek komínů cementovou maltou + 2× matný bílý fasádní nátěr.",
     "m2", 35, 240, 380, "CS ÚRS 2026 01", 0.95, 0.018,
     "5*7"),
    ("HSV", "Práce a dodávky HSV", "6", "Úpravy povrchů, podlahy a osazování výplní",
     "622221111", "ETICS pro vytvoření nové římsy nad uliční fasádou",
     "Kontaktní zateplovací systém ETICS — EPS-F 80 mm + síťovina + lepidlo + finální "
     "tenkovrstvá omítka. Nová římsa nad uliční fasádou (TZ str. 4).",
     "m2", 12.5, 980, 870, "CS ÚRS 2026 01", 1.85, 0.045,
     "25.01*0.5"),
    ("HSV", "Práce a dodávky HSV", "6", "Úpravy povrchů, podlahy a osazování výplní",
     "711161571", "Hydroizolační stěrka 2-složková ve sprchových koutech",
     "Hydroizolační stěrka 2-složková pod obklad a dlažbu, vč. pásky v rohu a okolo prostupů. "
     "4 koupelny — stěny sprchových koutů + podlaha (TZ §Hydroizolace).",
     "m2", 32, 195, 290, "CS ÚRS 2026 01", 0.60, 0.005,
     "4*8"),
]

# ===== HSV-9 Bourání (DOPLNĚNÍ — uživatel má jen lešení, nemá bourání) =====
P += [
    ("HSV", "Práce a dodávky HSV", "9", "Ostatní konstrukce a práce, bourání",
     "962031132", "Bourání zdiva cihelného atikové nadezdívky + horní římsy",
     "Bourání zdiva nadzákladového z cihel pálených na maltu vápennou, atika 1.5 m × 0.3 m "
     "po délce 25.01 m uliční fasády (TZ str. 2).",
     "m3", 11.25, 0, 1850, "CS ÚRS 2026 01", 6.50, 1.800,
     "25.01*1.5*0.3"),
    ("HSV", "Práce a dodávky HSV", "9", "Ostatní konstrukce a práce, bourání",
     "962031141", "Bourání stávajícího vaznicového krovu — celého uličního traktu",
     "Bourání dřevěné konstrukce krovů z prostého řeziva ručně, vč. vaznic, sloupků, krokví, "
     "pozednic. Krov stávající bude kompletně snesen (TZ str. 2).",
     "m3", 11.0, 0, 1650, "CS ÚRS 2026 01", 6.20, 0.500,
     "154*0.072"),
    ("HSV", "Práce a dodávky HSV", "9", "Ostatní konstrukce a práce, bourání",
     "965042141", "Bourání skladby podlahy podkroví — až na trámy",
     "Bourání podlah z betonu prostého ručně. Stávající skladba: beton 50 + lepenka + polystyren 40 + "
     "podkladový beton 20 + násyp 70 + 2× dřevěný záklop 50 (TZ str. 2-3).",
     "m2", 110.3, 0, 285, "CS ÚRS 2026 01", 1.30, 0.430,
     "4.41*25.01"),
    ("HSV", "Práce a dodávky HSV", "9", "Ostatní konstrukce a práce, bourání",
     "968071115", "Bourání zdiva pro nové dveřní otvory — 2 ks světlosti 1.0 m",
     "Bourání průchodů ve střední podélné nosné stěně tl. 0.6 m, světlost 1.0 m × 2.1 m, 2 ks. "
     "Nové dveře (TZ str. 2).",
     "m3", 2.7, 0, 3800, "CS ÚRS 2026 01", 8.50, 1.800,
     "2*1.0*2.1*0.6+0.18"),
    ("HSV", "Práce a dodávky HSV", "9", "Ostatní konstrukce a práce, bourání",
     "968071116", "Vybourání nového okenního otvoru ve štítové stěně",
     "Bourání průchodu ve štítové cihelné stěně pro nový okenní otvor (TZ str. 1).",
     "m3", 1.5, 0, 3800, "CS ÚRS 2026 01", 8.50, 1.800,
     "1.5*1.5*0.5+0.375"),
]

# ===== HSV-997 Doprava suti — DOPLNĚNÍ kvantitativní =====
# User has only 12.575 t. Reálně cca 54 t. Doplníme rozdíl.
P += [
    ("HSV", "Práce a dodávky HSV", "997", "Doprava suti a vybouraných hmot",
     "997013111", "Vnitrostaveništní doprava suti do 10 m svisle — doplnění",
     "Vnitrostaveništní doprava suti a vybouraných hmot pro budovy v do 10 m. "
     "Doplnění rozdílu nad uživatelovou hodnotu 12.575 t na reálných 54 t (krov + atika + "
     "podlaha + průchody).",
     "t", 41.425, 0, 245, "CS ÚRS 2026 01", 1.10, 0,
     "54.0-12.575"),
    ("HSV", "Práce a dodávky HSV", "997", "Doprava suti a vybouraných hmot",
     "997013501.D", "Odvoz suti — doplnění",
     "Odvoz suti a vybouraných hmot na skládku do 1 km — doplnění o rozdíl mezi 12.575 t "
     "(uvedeno uživatelem) a 54 t (reálně).",
     "t", 41.425, 0, 315, "CS ÚRS 2026 01", 0.10, 0,
     "54.0-12.575"),
    ("HSV", "Práce a dodávky HSV", "997", "Doprava suti a vybouraných hmot",
     "997013509.D", "Příplatek k odvozu suti ZKD 1 km — doplnění (předpoklad 15 km)",
     "Příplatek k odvozu suti za každý další km přes 1 km, vzdálenost cca 15 km na skládku. "
     "Doplnění o rozdíl: 41.425 t × 14 km = 580 t·km navíc.",
     "t", 580.0, 0, 13.8, "CS ÚRS 2026 01", 0.005, 0,
     "41.425*14"),
    ("HSV", "Práce a dodávky HSV", "997", "Doprava suti a vybouraných hmot",
     "997013601", "Poplatek za uložení betonu (17 01 01)",
     "Poplatek za uložení stavebního odpadu betonového a železobetonového (17 01 01) na "
     "skládce — stará podlaha mazanin 5.52 m³ + podkladový beton 2.21 m³ = 7.73 m³ × 2400 kg.",
     "t", 18.5, 850, 0, "CS ÚRS 2026 01", 0, 0,
     "(0.05+0.02)*110.3*2.4"),
    ("HSV", "Práce a dodávky HSV", "997", "Doprava suti a vybouraných hmot",
     "997013211", "Poplatek za uložení dřevěného odpadu (17 02 01)",
     "Poplatek za uložení dřevěného odpadu na skládce — krov 5 t + záklopy 4 t = 9 t.",
     "t", 9.0, 1450, 0, "CS ÚRS 2026 01", 0, 0,
     "5+4"),
    ("HSV", "Práce a dodávky HSV", "997", "Doprava suti a vybouraných hmot",
     "997013602", "Poplatek za uložení cihelného odpadu (17 01 02)",
     "Poplatek za uložení cihelného odpadu — atika 11.25 m³ × 1800 kg = 20.25 t.",
     "t", 20.25, 1980, 0, "CS ÚRS 2026 01", 0, 0,
     "11.25*1.8"),
]

# ===== PSV-713 Izolace tepelné — DOPLNĚNÍ (uživatel má střechu, chybí podlaha podkroví) =====
P += [
    ("PSV", "Práce a dodávky PSV", "713", "Izolace tepelné",
     "713121111", "Tepelně-kročejová izolace podlahy podkroví — Girifloor 400 tl. 20 mm",
     "Izolace tepelná podlah na sucho deskami z minerální plsti Isover/Knauf/Girifloor 400, "
     "tloušťky 20 mm, kročejová. Užitná plocha 127.03 m² (TZ str. 4).",
     "m2", 127.0, 145, 50, "CS ÚRS 2026 01", 0.18, 0.005,
     "127.03"),
]

# ===== PSV-762 Konstrukce tesařské — DOPLNĚNÍ (chybí pozednice + latě + některé prvky) =====
P += [
    ("PSV", "Práce a dodávky PSV", "762", "Konstrukce tesařské",
     "762332111", "Montáž pozednic z hraněného řeziva průřezu 140×160 mm",
     "Montáž pozednic na zdivo / věnec, profil 140 × 160 mm, kotvení chemickou kotvou M16/1000 mm. "
     "2× délka uliční fasády (50.02 bm).",
     "m", 50.02, 0, 215, "CS ÚRS 2026 01", 0.32, 0,
     "2*25.01"),
    ("PSV", "Práce a dodávky PSV", "762", "Konstrukce tesařské",
     "60512131", "Hranol stavební řezivo SM/JD průřezu 140×160 mm (pozednice)",
     "Hranol stavební řezivo smrk-jedle průřezu do 224 cm², dl 6 m, profil pozednice 140 × 160 mm.",
     "m3", 1.12, 9740, 0, "CS ÚRS 2026 01", 0, 0.500,
     "50.02*0.0224"),
    ("PSV", "Práce a dodávky PSV", "762", "Konstrukce tesařské",
     "762342214", "Montáž laťování 50/40 mm na sraz pro krytinu bobrovku",
     "Montáž laťování svisle nebo podélně, profil 50 × 40 mm, rozteč ~330 mm pro bobrovku. "
     "Plocha hl. střecha 154 m² / 0.33 m = 467 bm.",
     "m", 467, 0, 38, "CS ÚRS 2026 01", 0.072, 0,
     "154/0.33"),
    ("PSV", "Práce a dodávky PSV", "762", "Konstrukce tesařské",
     "60514114.A", "Lať jehličnatá 50×40 mm impregnovaná (latě + kontralatě bobrovka)",
     "Řezivo jehličnaté lať 50 × 40 mm pevnostní třída S10, impregnovaná. "
     "Latě 467 bm + kontralatě 154 bm × 0.002 m².",
     "m3", 1.24, 10700, 0, "CS ÚRS 2026 01", 0, 0.500,
     "(467+154)*0.002"),
    ("PSV", "Práce a dodávky PSV", "762", "Konstrukce tesařské",
     "762541127", "Strukturovaná podkladní rohož pod plech vikýře tl. 8 mm",
     "Strukturovaná separační a odvětrávací rohož pod plechovou krytinu vikýřů. "
     "Skladba vikýř pos. 2 (TZ).",
     "m2", 51.2, 115, 50, "CS ÚRS 2026 01", 0.08, 0.002,
     "51.2"),
    ("PSV", "Práce a dodávky PSV", "762", "Konstrukce tesařské",
     "762811111", "Difuzní fólie Tyvek pod skladbu vikýře",
     "Montáž difuzní paropropustné fólie Tyvek na bednění OSB pod krytinu vikýře, "
     "vč. přelepených spojů.",
     "m2", 51.2, 28, 50, "CS ÚRS 2026 01", 0.08, 0.001,
     "51.2"),
]

# ===== PSV-763 SDK + suchá podlaha — DOPLNĚNÍ (uživatel má jen podhled) =====
P += [
    ("PSV", "Práce a dodávky PSV", "763", "Konstrukce suché výstavby",
     "763131611", "SDK předstěny + opláštění ocelových konstrukcí v koupelnách + WC",
     "SDK předstěna na systémovém roštu CW/UW, opláštění impregnovanými deskami GKBi "
     "tl. 12.5 mm. Koupelny + WC + zádveří + opláštění oc. profilů (2 byty × ~45 m²).",
     "m2", 90, 380, 340, "CS ÚRS 2026 01", 1.05, 0.022,
     "2*45"),
    ("PSV", "Práce a dodávky PSV", "763", "Konstrukce suché výstavby",
     "763131831", "Tenkovrstvá omítka sádrová Knauf + 2× bílý malířský nátěr — podhledy + předstěny",
     "Sádrová tenkovrstvá omítka tl. 1-2 mm + 2× malířský nátěr bílý disperzní. "
     "Aplikace na celý SDK podhled + předstěny.",
     "m2", 295.2, 65, 100, "CS ÚRS 2026 01", 0.35, 0.001,
     "205.2+90"),
    ("PSV", "Práce a dodávky PSV", "763", "Konstrukce suché výstavby",
     "763411511", "OSB záklop tl. 22 mm na zesílené stropní trámy (podklad pro suchou podlahu)",
     "Záklop stropů z desek OSB superfinish 3 tl. 22 mm na pero-drážku, šroubování. "
     "Podklad pro suchou podlahu Fermacell.",
     "m2", 127, 195, 230, "CS ÚRS 2026 01", 0.42, 0.014,
     "127"),
    ("PSV", "Práce a dodávky PSV", "763", "Konstrukce suché výstavby",
     "763413311", "Suchá podlaha Fermacell — voština 60 + 2× deska 2E35 25 mm",
     "Suchá podlaha Fermacell systém: voština z dřevěné štěpky tl. 60 mm jako vyrovnávací "
     "+ MW kročejová izolace tl. 20 mm + 2× sádrovláknitá deska 2E35 (= 12.5 + 12.5 mm). "
     "Užitná plocha 2 bytů 127.03 m² (TZ str. 3 + řez).",
     "m2", 127, 580, 610, "CS ÚRS 2026 01", 0.95, 0.045,
     "127.03"),
]

# ===== PSV-764 Klempířské Cu — DOPLNĚNÍ (uživatel má jen Al krytinu) =====
P += [
    ("PSV", "Práce a dodávky PSV", "764", "Konstrukce klempířské",
     "764311621", "Oplechování komínů Cu plech tl. 0.6 mm — 5 komínů",
     "Oplechování komínů z měděného plechu tl. 0.6 mm, vč. lemování po obvodu komína a "
     "napojení na krytinu. 5 komínů × cca 3.6 m² pláště (TZ str. 4 'klempířské výrobky "
     "zbývajících střech budou proveden z Cu plechu').",
     "m2", 18, 1050, 800, "CS ÚRS 2026 01", 2.20, 0.006,
     "5*3.6"),
    ("PSV", "Práce a dodávky PSV", "764", "Konstrukce klempířské",
     "764311515", "Klempířské lemování napojení u světlíku Cu plech",
     "Lemování světlíku z měděného plechu rš 250 mm, obvod cca 8 bm.",
     "m", 8, 320, 260, "CS ÚRS 2026 01", 0.75, 0.003,
     "8"),
    ("PSV", "Práce a dodávky PSV", "764", "Konstrukce klempířské",
     "764511631", "Žlab podokapní polokruhový Cu plech ø 150 mm",
     "Montáž podokapního žlabu polokruhového Cu ø 150 mm vč. čel, kotlíků a háků. "
     "Délka uliční fasády 25.01 m → 25 bm.",
     "m", 25, 580, 405, "CS ÚRS 2026 01", 0.95, 0.004,
     "25.01"),
    ("PSV", "Práce a dodávky PSV", "764", "Konstrukce klempířské",
     "764554621", "Svod dešťový Cu plech ø 100 mm s nátěrem v odstínu fasády",
     "Montáž svodu dešťového kruhového Cu ø 100 mm vč. kolen, objímek, kachlíku a lapače. "
     "TZ str. 4: 'dešťové svody do dvora opatřeny nátěrem v odstínu fasády'.",
     "m", 30, 720, 465, "CS ÚRS 2026 01", 0.85, 0.003,
     "2*15"),
    ("PSV", "Práce a dodávky PSV", "764", "Konstrukce klempířské",
     "764311671", "Okapnice Cu plech rš 250 mm",
     "Okapnice z měděného plechu rš 250 mm, montáž po délce uliční fasády.",
     "m", 25, 240, 170, "CS ÚRS 2026 01", 0.45, 0.002,
     "25.01"),
    ("PSV", "Práce a dodávky PSV", "764", "Konstrukce klempířské",
     "764311551", "Boční lemování vikýřů Cu plech — napojení Lindab na bobrovku",
     "Klempířské lemování bočních hran vikýře (kritický vodotěsný detail mezi Lindab Click "
     "vikýře a keramickou bobrovkou hl. střechy).",
     "m", 15.24, 385, 300, "CS ÚRS 2026 01", 0.62, 0.003,
     "2*2*3.81"),
    ("PSV", "Práce a dodávky PSV", "764", "Konstrukce klempířské",
     "764311671.A", "Čelní okapová hrana vikýře Cu plech",
     "Klempířské lemování čelní okapové hrany vikýře (2 × šířka 6.78 m).",
     "m", 13.56, 240, 170, "CS ÚRS 2026 01", 0.45, 0.002,
     "2*6.78"),
    ("PSV", "Práce a dodávky PSV", "764", "Konstrukce klempířské",
     "764811561", "Oplechování parapetů a ostění vikýřových oken Cu plech",
     "Klempířské oplechování parapetů a vnějších ostění 4 ks vikýřových oken (obvod cca 8 bm/okno).",
     "m", 32, 280, 205, "CS ÚRS 2026 01", 0.55, 0.002,
     "4*8"),
]

# ===== PSV-765 Krytina skládaná — DOPLNĚNÍ (chybí tyčový sněholam + hřebenová tvarovka) =====
P += [
    ("PSV", "Práce a dodávky PSV", "765", "Krytina skládaná",
     "765115521", "Tyčový sněhový zachytávač před vikýři — 3-úrovňový systém",
     "Tyčový zachytávač sněhu 3-úrovňový před okapovou hranou vikýřů (TZ str. 3: "
     "'tyčový sněhový zachytávač pro zabránění přetížení žlabu ledem'). NE mříž — tyče.",
     "m", 13.56, 195, 90, "CS ÚRS 2026 01", 0.25, 0.005,
     "2*6.78"),
    ("PSV", "Práce a dodávky PSV", "765", "Krytina skládaná",
     "765191017", "Větrací hřebenová tvarovka pro odvětrání střechy",
     "Větrací tvarovka hřebenová pod hřebenáčem, šířka 240 mm, materiál EPDM/plast. "
     "TZ str. 3: 'odvětrání hřebene řešeno větrací hřebenovou tvarovkou'.",
     "m", 25, 185, 100, "CS ÚRS 2026 01", 0.20, 0.003,
     "25"),
]

# ===== PSV-766 Truhlářské — DOPLNĚNÍ (chybí vstupní dveře + vnitřní dveře + parapety) =====
P += [
    ("PSV", "Práce a dodávky PSV", "766", "Konstrukce truhlářské",
     "766660171", "Vstupní bytové dveře jednokřídlé 900×1970 mm RC3 EI30 DP3 37 dB",
     "Vstupní bytové dveře jednokřídlé do ocelové zárubně s dřevěnou obložkou. "
     "Bezpečnostní třída RC3, požární odolnost EI 30 DP3, vzduchová neprůzvučnost Rw ≥ 37 dB. "
     "TZ str. 4 explicit. 2 byty × 1 dveře.",
     "kus", 2, 18500, 6000, "CS ÚRS 2026 01", 5.50, 0.085,
     "2"),
    ("PSV", "Práce a dodávky PSV", "766", "Konstrukce truhlářské",
     "766660172", "Vnitřní dveře hladké do ocelové zárubně 800-900/1970 mm",
     "Vnitřní dveře hladké jednokřídlé do ocelové zárubně, s povrchovou úpravou CPL nebo "
     "fólií. 2 byty × ~5 dveří (zádveří + koupelna + WC + ložnice + pokoj).",
     "kus", 10, 4200, 1600, "CS ÚRS 2026 01", 1.80, 0.045,
     "2*5"),
    ("PSV", "Práce a dodávky PSV", "766", "Konstrukce truhlářské",
     "766691111", "Vnitřní parapety oken vikýřů — 4 ks",
     "Vnitřní parapety oken vikýřů z laminované DTD nebo kompaktní desky, šířka 200 mm, "
     "tloušťka 25 mm. 4 ks vikýřových oken.",
     "kus", 4, 980, 500, "CS ÚRS 2026 01", 1.20, 0.008,
     "4"),
    ("PSV", "Práce a dodávky PSV", "766", "Konstrukce truhlářské",
     "766691112", "Vnitřní ostění + nadpraží oken vikýřů + střešních oken",
     "Vnitřní ostění a nadpraží oken z laminovaných desek tl. 19 mm. "
     "4 vikýřové + 2 střešní × cca 1.7 obvodových ostění = 10 kompletních ostění.",
     "kpl", 10, 540, 310, "CS ÚRS 2026 01", 1.10, 0.005,
     "10"),
]

# ===== PSV-767 Konstrukce zámečnické — NOVÁ SEKCE (KRITICKÉ — chybí celá ocel) =====
P += [
    ("PSV", "Práce a dodávky PSV", "767", "Konstrukce zámečnické",
     "767160100", "Vaznice 2× UPN 100 (box) — střední a hřebenová vaznice krovu",
     "Vaznice ocelová z válcovaného profilu 2× UPN 100 S235JR (box / dvojprofil) — "
     "střední + hřebenová vaznice, celkem 2 × 25.01 = 50.02 bm. Pole sloupků 3.5 m → "
     "8 sloupků per vaznice. Posouzeno dle ČSN EN 1993-1-1: M_d=11.2 kNm, M_Rd=21.2 kNm "
     "(využití 53%), průhyb 1/300 = limit. Statika_assumed.md §4.",
     "t", 1.06, 41000, 24000, "vlastní R", 18.50, 1.000,
     "2*50.02*10.6/1000"),
    ("PSV", "Práce a dodávky PSV", "767", "Konstrukce zámečnické",
     "767160105", "Vaznice 2× UPN 120 — zalomení vikýřů",
     "Vaznice ocelová z válcovaného profilu 2× UPN 120 S235JR v zalomení bočních stěn "
     "vikýřů (mezi sklonem 7° a 36.2°). 2 vikýře × 2 boky × 3.81 m = 15.24 bm × 2 profily. "
     "Posouzeno dle ČSN EN 1993-1-1: M_d=2.0 kNm << M_Rd=31.4 kNm (rezerva 94%). "
     "Statika_assumed.md §5.",
     "t", 0.41, 41000, 24000, "vlastní R", 18.50, 1.000,
     "2*15.24*13.4/1000"),
    ("PSV", "Práce a dodávky PSV", "767", "Konstrukce zámečnické",
     "767160110", "Sloupky pod vaznice — trubka 80×60×4 mm, 16 ks",
     "Sloupky pod vaznice z trubky uzavřené 80 × 60 × 4 mm S235JR, průměrná výška 1.8 m. "
     "16 ks (8 pod střední + 8 pod hřebenovou vaznicí). Posouzeno dle ČSN EN 1993-1-1: "
     "F_d=25.6 kN, N_b,Rd=187 kN (využití 14%, χ=0.75). Statika_assumed.md §6.",
     "t", 0.24, 45000, 26000, "vlastní R", 22.00, 1.000,
     "16*1.8*8.27/1000"),
    ("PSV", "Práce a dodávky PSV", "767", "Konstrukce zámečnické",
     "767995118", "Žárové zinkování ocelových konstrukcí (vaznice + sloupky + IPE)",
     "Žárové zinkování ocelových konstrukcí ČSN EN ISO 1461. Aplikace na 2× UPN 100 + "
     "2× UPN 120 + sloupky + IPE 100. Celkem 1.76 t × cca 80 µm.",
     "t", 1.76, 0, 14500, "CS ÚRS 2026 01", 8.50, 0,
     "1.06+0.41+0.24+0.05"),
    ("PSV", "Práce a dodávky PSV", "767", "Konstrukce zámečnické",
     "767995111", "Kotvení ocelových konstrukcí — chemické kotvy",
     "Kotvení sloupků a IPE 100 chemickými kotvami M16 Hilti HIT-HY 200 do věnce / zdiva. "
     "Cca 32 ks kotev (16 sloupků × 2 + 4 IPE × 2).",
     "kpl", 1, 9500, 9000, "vlastní R", 12.00, 0,
     "1"),
    ("PSV", "Práce a dodávky PSV", "767", "Konstrukce zámečnické",
     "767996115", "Tepelně izolační návlek na ocelový profil ve světlíku",
     "Tepelně izolační návlek na ocelový profil prostupující světlíkem, jádro minerální vata "
     "tl. 80 mm + povrch hliníkový plech (TZ str. 3 explicit).",
     "kpl", 1, 4800, 3700, "vlastní R", 6.50, 0.018,
     "1"),
]

# ===== PSV-771 Podlahy dlažbou (NOVÁ SEKCE) =====
P += [
    ("PSV", "Práce a dodávky PSV", "771", "Podlahy z dlaždic",
     "771474112", "Keramická dlažba lepená 30×30 / 40×40 — koupelny + WC + chodby",
     "Podlaha keramická dlažba 30 × 30 nebo 40 × 40 mm, mrazuvzdorná, R10, lepidlo flex, spárování. "
     "Koupelny + WC + zádveří + chodby v 2 bytech (~40 m² odhad).",
     "m2", 40, 620, 700, "CS ÚRS 2026 01", 1.45, 0.045,
     "2*20"),
]

# ===== PSV-775 Podlahy plovoucí (NOVÁ SEKCE) =====
P += [
    ("PSV", "Práce a dodávky PSV", "775", "Podlahy plovoucí",
     "775413232", "Plovoucí podlaha vícevrstvá — obyt. místnosti (87 m²)",
     "Podlaha plovoucí lamela/vinyl/korek tl. 8-10 mm na podkladovou parozábrannou podložku. "
     "Obytné místnosti v 2 bytech, plocha 127 m² - 40 m² dlažba = 87 m².",
     "m2", 87, 385, 300, "CS ÚRS 2026 01", 0.45, 0.011,
     "127-40"),
]

# ===== PSV-781 Obklady (NOVÁ SEKCE) =====
P += [
    ("PSV", "Práce a dodávky PSV", "781", "Obklady keramické",
     "781474112", "Obklad stěn koupelen + WC — výška 2.0 m, formát 25×40 / 30×60",
     "Obklad keramický stěn formátu 25 × 40 mm nebo 30 × 60 mm, lepidlo flex, výška 2.0 m. "
     "4 koupelny + 2 WC.",
     "m2", 95, 540, 445, "CS ÚRS 2026 01", 1.15, 0.018,
     "4*16+2*12+spary"[:0] or "95"),
]

# ===== VRN — DOPLNĚNÍ (uživatel má jen 90 000 Kč ZS, chybí BOZP + geodet + průzkum) =====
P += [
    ("VRN", "Vedlejší rozpočtové náklady", "VRN3", "Zařízení staveniště",
     "030001000.D", "Doplnění ZS — sociální buňky + sklad + brána + plot + energie/voda",
     "Doplnění nákladů zařízení staveniště nad uživatelových 90 000 Kč: dodatečné buňky "
     "(sociál + šatna + kancelář), sklad materiálu, brána + plot, dočasné rozvody E+V.",
     "kpl", 1, 0, 25000, "vlastní R", 0, 0,
     "1"),
    ("VRN", "Vedlejší rozpočtové náklady", "VRN4", "Inženýrská činnost",
     "030002000", "Koordinátor BOZP — fáze realizace + dokumentace",
     "Koordinátor BOZP pro fázi realizace stavby vč. plánu BOZP a kontrol. Zákon 309/2006 Sb.",
     "kpl", 1, 0, 48000, "vlastní R", 0, 0,
     "1"),
    ("VRN", "Vedlejší rozpočtové náklady", "VRN4", "Inženýrská činnost",
     "030003000", "Geodetické zaměření před stavbou + skutečné provedení",
     "Geodetické práce — zaměření stávajícího stavu před zahájením + zaměření skutečného "
     "provedení dokončené stavby (vyhláška 405/2017 Sb. + DSPS).",
     "kpl", 1, 0, 28000, "vlastní R", 0, 0,
     "1"),
    ("VRN", "Vedlejší rozpočtové náklady", "VRN4", "Inženýrská činnost",
     "030004000", "Průzkum pevnosti zdiva 1.PP + 1.NP (TZ §7 explicit)",
     "Průzkum pevnosti zdiva vybraných přitěžovaných pilířů v 1.PP a 1.NP. Schmidtův kladivem "
     "nebo válcovou jádrovou vrtnou zkouškou. TZ §7 EXPLICITNĚ POŽADUJE.",
     "kpl", 1, 0, 35000, "vlastní R", 0, 0,
     "1"),
    ("VRN", "Vedlejší rozpočtové náklady", "VRN4", "Inženýrská činnost",
     "030005000", "Statický posudek autorizovanou osobou ČKAIT IS00 + AD",
     "Statický posudek autorizovanou osobou ČKAIT obor IS00 (statika a dynamika staveb) "
     "dle zákona 360/1992 Sb. + autorský dozor během realizace.",
     "kpl", 1, 0, 25000, "vlastní R", 0, 0,
     "1"),
    ("VRN", "Vedlejší rozpočtové náklady", "VRN4", "Inženýrská činnost",
     "030006000", "Pasportizace okolních objektů — fotodokumentace + protokol",
     "Pasportizace okolních objektů (sousední BD vlevo + vpravo + vnitroblok) — vstupní "
     "fotodokumentace stávajícího stavu fasád, oken a interiérů + závěrečný protokol po dokončení.",
     "kpl", 1, 0, 22000, "vlastní R", 0, 0,
     "1"),
    ("VRN", "Vedlejší rozpočtové náklady", "VRN4", "Inženýrská činnost",
     "030007000", "Dokumentace skutečného provedení stavby (DSPS) — všechny profese",
     "Vyhotovení dokumentace skutečného provedení stavby dle vyhlášky 499/2006 Sb. — všechny "
     "profese (ASŘ + STA + PBŘ + ELE + ZTI + UT + VZT).",
     "kpl", 1, 0, 18000, "vlastní R", 0, 0,
     "1"),
]

# --- Now write all positions ---
row = HEADER_ROW + 1
seq = 0
current_l1 = None
current_l2 = None

style_section_fill_l1 = PatternFill("solid", fgColor="2F5496")
style_section_font_l1 = Font(bold=True, color="FFFFFF", size=11)
style_section_fill_l2 = PatternFill("solid", fgColor="DDEBF7")
style_section_font_l2 = Font(bold=True, color="000000", size=10)
style_item_font = Font(size=10)
style_pp_font = Font(italic=True, size=9, color="606060")
style_vv_font = Font(italic=True, size=9, color="909090")

section_totals_l1 = {}
section_totals_l2 = {}
total = 0.0
total_nh = 0.0
total_hm = 0.0

# First pass — accumulate totals per section
# tuple idx: 0 l1_kod, 1 l1_name, 2 l2_kod, 3 l2_name, 4 kod, 5 popis_short,
#            6 popis_long, 7 mj, 8 mn, 9 j_mat, 10 j_mnt, 11 urs, 12 nh, 13 hm, 14 vv
for p in P:
    cc = round((p[9] + p[10]) * p[8], 2)  # (j_mat + j_mnt) × mn
    section_totals_l1.setdefault(p[0], 0)
    section_totals_l2.setdefault((p[0], p[2]), 0)
    section_totals_l1[p[0]] += cc
    section_totals_l2[(p[0], p[2])] += cc
    total += cc

# Second pass — write
prev_l1 = None
prev_l2 = None
for p in P:
    (l1_kod, l1_name, l2_kod, l2_name, kod, popis_short, popis_long, mj, mn,
     j_mat, j_mnt, urs, nh_per, hm_per, vv) = p

    if l1_kod != prev_l1:
        # Write L1 section header
        ws.cell(row=row, column=4, value="D")
        ws.cell(row=row, column=5, value=l1_kod)
        ws.cell(row=row, column=6, value=l1_name)
        ws.cell(row=row, column=11, value=round(section_totals_l1[l1_kod], 2))
        for c in range(1, 25):
            ws.cell(row=row, column=c).fill = style_section_fill_l1
            ws.cell(row=row, column=c).font = style_section_font_l1
        row += 1
        prev_l1 = l1_kod
        prev_l2 = None

    if l2_kod != prev_l2:
        ws.cell(row=row, column=4, value="D")
        ws.cell(row=row, column=5, value=l2_kod)
        ws.cell(row=row, column=6, value=l2_name)
        ws.cell(row=row, column=11, value=round(section_totals_l2[(l1_kod, l2_kod)], 2))
        for c in range(1, 25):
            ws.cell(row=row, column=c).fill = style_section_fill_l2
            ws.cell(row=row, column=c).font = style_section_font_l2
        row += 1
        prev_l2 = l2_kod

    # K-row (item)
    seq += 1
    j_total = j_mat + j_mnt
    cc = round(j_total * mn, 2)
    mat_total = round(j_mat * mn, 2)
    mnt_total = round(j_mnt * mn, 2)
    nh_total = round(nh_per * mn, 4)
    hm_total = round(hm_per * mn, 4)
    total_nh += nh_total
    total_hm += hm_total

    ws.cell(row=row, column=3, value=seq)
    ws.cell(row=row, column=4, value="K")
    ws.cell(row=row, column=5, value=kod)
    ws.cell(row=row, column=6, value=popis_short)
    ws.cell(row=row, column=7, value=mj)
    ws.cell(row=row, column=8, value=mn)
    ws.cell(row=row, column=9, value=j_mat)
    ws.cell(row=row, column=10, value=j_mnt)
    ws.cell(row=row, column=11, value=cc)
    ws.cell(row=row, column=12, value=urs)
    ws.cell(row=row, column=15, value="snížená")
    ws.cell(row=row, column=16, value=j_total)
    ws.cell(row=row, column=17, value=mat_total)
    ws.cell(row=row, column=18, value=mnt_total)
    ws.cell(row=row, column=19, value=nh_per)
    ws.cell(row=row, column=20, value=nh_total)
    ws.cell(row=row, column=21, value=hm_per)
    ws.cell(row=row, column=22, value=hm_total)
    for c in [4, 5, 6, 7, 12, 15]:
        ws.cell(row=row, column=c).font = style_item_font
    ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical='top')
    row += 1

    # PP-row (popis položky)
    ws.cell(row=row, column=4, value="PP")
    ws.cell(row=row, column=6, value=popis_long)
    ws.cell(row=row, column=6).font = style_pp_font
    ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical='top')
    row += 1

    # VV-row (výpočet výměry)
    ws.cell(row=row, column=4, value="VV")
    ws.cell(row=row, column=6, value=str(vv))
    ws.cell(row=row, column=8, value=mn)
    ws.cell(row=row, column=6).font = style_vv_font
    ws.cell(row=row, column=8).font = style_vv_font
    row += 1

# Total row
row += 1
ws.cell(row=row, column=3, value="Náklady doplnění CELKEM").font = Font(bold=True, size=12)
ws.cell(row=row, column=11, value=round(total, 2)).font = Font(bold=True, size=12)
ws.cell(row=row, column=11).fill = PatternFill("solid", fgColor="FFD966")
ws.cell(row=row, column=20, value=round(total_nh, 2)).font = Font(bold=True)
ws.cell(row=row, column=22, value=round(total_hm, 2)).font = Font(bold=True)

row += 1
ws.cell(row=row, column=3, value="DPH 15 % (residential)").font = Font(italic=True)
ws.cell(row=row, column=11, value=round(total * 0.15, 2))

row += 1
ws.cell(row=row, column=3, value="Doplnění s DPH 15 %").font = Font(bold=True, size=12)
ws.cell(row=row, column=11, value=round(total * 1.15, 2)).font = Font(bold=True, size=12)
ws.cell(row=row, column=11).fill = PatternFill("solid", fgColor="FFD966")

row += 2
ws.cell(row=row, column=3, value="REKAPITULACE STAVBY (uživatel + doplnění):").font = Font(bold=True, size=11)
row += 1
ws.cell(row=row, column=3, value="  Soupis uživatele:")
ws.cell(row=row, column=11, value=2538613.68)
row += 1
ws.cell(row=row, column=3, value="  Doplnění tímto listem:")
ws.cell(row=row, column=11, value=round(total, 2))
row += 1
ws.cell(row=row, column=3, value="  CELKEM bez DPH:").font = Font(bold=True)
ws.cell(row=row, column=11, value=round(2538613.68 + total, 2)).font = Font(bold=True)
ws.cell(row=row, column=11).fill = PatternFill("solid", fgColor="FFD966")
row += 1
ws.cell(row=row, column=3, value="  CELKEM s DPH 15 %:").font = Font(bold=True, size=12)
ws.cell(row=row, column=11, value=round((2538613.68 + total) * 1.15, 2)).font = Font(bold=True, size=12)
ws.cell(row=row, column=11).fill = PatternFill("solid", fgColor="FFD966")

# Column widths (mirror source)
widths = {3: 5, 4: 5, 5: 13, 6: 65, 7: 6, 8: 11, 9: 12, 10: 12, 11: 14, 12: 16,
          13: 4, 14: 4, 15: 11, 16: 12, 17: 14, 18: 14, 19: 9, 20: 10, 21: 11, 22: 12}
for c, w in widths.items():
    ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

# Hide extra blank columns 13, 14, 23, 24+
for c in range(23, 65):
    ws.column_dimensions[openpyxl.utils.get_column_letter(c)].hidden = True

ws.row_dimensions[HEADER_ROW].height = 32
ws.freeze_panes = f"A{HEADER_ROW+1}"

wb.save(DST)
print(f"✓ Wrote {DST}")
print(f"  Sheets: {wb.sheetnames}")
print(f"  Positions added: {len(P)}")
print(f"  User existing:    2 538 613.68 Kč")
print(f"  Doplnění:         {total:>12,.2f} Kč")
print(f"  CELKEM bez DPH:   {2538613.68 + total:>12,.2f} Kč")
print(f"  CELKEM s DPH 15 %: {(2538613.68 + total) * 1.15:>12,.2f} Kč")
print()
print("Section breakdown of additions:")
for k, v in sorted(section_totals_l1.items()):
    print(f"  {k:6} {v:>14,.2f} Kč")
