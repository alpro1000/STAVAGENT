#!/usr/bin/env python3
"""Generate KROS-compatible Excel from positions_v2.json — same column structure as v1."""
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
data = json.loads((HERE / "positions_v2.json").read_text(encoding="utf-8"))

wb = openpyxl.Workbook()

# ===== Sheet 1: Soupis (KROS layout) =====
ws = wb.active
ws.title = "Soupis"
headers = ["O", "P", "Úroveň", "TC", "ČP", "TV", "Typ položky", "Kód položky",
           "Popis", "TOV", "MJ", "Množství", "Jednotková cena", "Index",
           "J. cena indexovaná", "Montáž jedn.", "Dodávka jedn.", "Celková cena",
           "Confidence", "Zdroj", "Výpočet/Pozn."]
ws.append(headers)

hf = Font(bold=True, color="FFFFFF", size=10)
hfill = PatternFill("solid", fgColor="2F5496")
section_fill = PatternFill("solid", fgColor="DDEBF7")
warn_fill = PatternFill("solid", fgColor="FFF2CC")
border = Border(*[Side(style="thin", color="C0C0C0")] * 4)

for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = hf
    cell.fill = hfill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# Group positions by section, emit "D" (řádek skupiny) + "K" (řádky položek)
positions = data["positions_v2"]
sections_order = []
seen = set()
for p in positions:
    if p["section"] not in seen:
        seen.add(p["section"])
        sections_order.append((p["section"], p["section_name"]))

row_idx = 2
position_seq = 0
section_totals = {}
grand_total = 0.0

for section, section_name in sections_order:
    section_pos = [p for p in positions if p["section"] == section]
    section_total = sum(p.get("cc", 0) for p in section_pos)
    section_totals[section] = section_total
    grand_total += section_total

    # Header row for section
    ws.cell(row=row_idx, column=3, value=" >2").font = Font(bold=True)
    ws.cell(row=row_idx, column=5, value=0)
    ws.cell(row=row_idx, column=6, value="D")
    ws.cell(row=row_idx, column=8, value=section)
    ws.cell(row=row_idx, column=9, value=section_name)
    ws.cell(row=row_idx, column=18, value=round(section_total, 2))
    for c in range(1, len(headers) + 1):
        ws.cell(row=row_idx, column=c).fill = section_fill
        ws.cell(row=row_idx, column=c).font = Font(bold=True)
    row_idx += 1

    for p in section_pos:
        position_seq += 1
        row_type = "M" if (p["kod"][:1].isdigit() and len(p["kod"]) > 6 and p["kod"][:2] in ("28","55","59","60","61","63")) else "K"
        typ_polozky = "PSV" if section.startswith("PSV") else ("HSV" if section.startswith("HSV") else "VRN")
        index_cell = 1
        j_indx = p.get("jc", 0) * index_cell
        ws.cell(row=row_idx, column=3, value="  >3")
        ws.cell(row=row_idx, column=5, value=position_seq)
        ws.cell(row=row_idx, column=6, value=row_type)
        ws.cell(row=row_idx, column=7, value=typ_polozky)
        ws.cell(row=row_idx, column=8, value=p["kod"])
        ws.cell(row=row_idx, column=9, value=p["popis"])
        ws.cell(row=row_idx, column=11, value=p["mj"])
        ws.cell(row=row_idx, column=12, value=p["mn"])
        ws.cell(row=row_idx, column=13, value=p.get("jc", 0))
        ws.cell(row=row_idx, column=14, value=index_cell)
        ws.cell(row=row_idx, column=15, value=j_indx)
        ws.cell(row=row_idx, column=18, value=p.get("cc", 0))
        ws.cell(row=row_idx, column=19, value=p.get("confidence", ""))
        ws.cell(row=row_idx, column=20, value=p.get("zdroj", ""))
        notes = p.get("vypocet", "")
        if p.get("warning"):
            notes += (" ⚠ " + p["warning"]) if notes else ("⚠ " + p["warning"])
            for c in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=c).fill = warn_fill
        ws.cell(row=row_idx, column=21, value=notes)
        ws.cell(row=row_idx, column=9).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row_idx, column=21).alignment = Alignment(wrap_text=True, vertical="top")
        row_idx += 1

# Grand total
row_idx += 1
ws.cell(row=row_idx, column=9, value="CELKEM v2 (bez DPH)").font = Font(bold=True, size=12)
ws.cell(row=row_idx, column=18, value=round(grand_total, 2)).font = Font(bold=True, size=12)
ws.cell(row=row_idx, column=18).fill = PatternFill("solid", fgColor="FFD966")

row_idx += 1
ws.cell(row=row_idx, column=9, value="DPH 15 % (residential)")
ws.cell(row=row_idx, column=18, value=round(grand_total * 0.15, 2))

row_idx += 1
ws.cell(row=row_idx, column=9, value="CELKEM s DPH 15 %").font = Font(bold=True, size=12)
ws.cell(row=row_idx, column=18, value=round(grand_total * 1.15, 2)).font = Font(bold=True, size=12)
ws.cell(row=row_idx, column=18).fill = PatternFill("solid", fgColor="FFD966")

# Column widths
widths = [4, 4, 7, 4, 5, 5, 8, 14, 65, 6, 7, 10, 13, 6, 14, 12, 12, 14, 8, 30, 50]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# ===== Sheet 2: Souhrn po sekcich =====
ws2 = wb.create_sheet("Souhrn po sekcích")
ws2.append(["Sekce", "Název", "Počet položek", "Celkem Kč bez DPH", "% z celku"])
for c in range(1, 6):
    ws2.cell(row=1, column=c).font = hf
    ws2.cell(row=1, column=c).fill = hfill
    ws2.cell(row=1, column=c).alignment = Alignment(horizontal="center", wrap_text=True)

r = 2
for section, section_name in sections_order:
    section_pos = [p for p in positions if p["section"] == section]
    section_total = section_totals[section]
    ws2.cell(row=r, column=1, value=section)
    ws2.cell(row=r, column=2, value=section_name)
    ws2.cell(row=r, column=3, value=len(section_pos))
    ws2.cell(row=r, column=4, value=round(section_total, 2))
    ws2.cell(row=r, column=5, value=round(section_total / grand_total * 100, 1))
    r += 1
ws2.cell(row=r, column=2, value="CELKEM").font = Font(bold=True)
ws2.cell(row=r, column=3, value=len(positions)).font = Font(bold=True)
ws2.cell(row=r, column=4, value=round(grand_total, 2)).font = Font(bold=True)
ws2.cell(row=r, column=4).fill = PatternFill("solid", fgColor="FFD966")
ws2.cell(row=r, column=5, value=100.0).font = Font(bold=True)
for i, w in enumerate([8, 50, 14, 18, 12], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ===== Sheet 3: Audit v1 — fatální chyby =====
ws3 = wb.create_sheet("Audit v1 - fatální chyby")
ws3.append(["ID", "Sekce", "v1 položky", "Co v1 říká", "Co říká TZ/výkres", "Dopad", "Oprava"])
for c in range(1, 8):
    ws3.cell(row=1, column=c).font = hf
    ws3.cell(row=1, column=c).fill = hfill
    ws3.cell(row=1, column=c).alignment = Alignment(horizontal="center", wrap_text=True)

r = 2
for f in data["audit_v1"]["fatal_errors"]:
    ws3.cell(row=r, column=1, value=f["id"])
    ws3.cell(row=r, column=2, value=f["section"])
    ws3.cell(row=r, column=3, value=", ".join(map(str, f["v1_pos_ids"])))
    ws3.cell(row=r, column=4, value=f["v1_says"])
    ws3.cell(row=r, column=5, value=f["tz_says"])
    ws3.cell(row=r, column=6, value=f["impact"])
    ws3.cell(row=r, column=7, value=f["fix"])
    for c in range(1, 8):
        ws3.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
for i, w in enumerate([6, 22, 12, 45, 45, 45, 45], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ===== Sheet 4: Audit v1 — chybějící sekce =====
ws4 = wb.create_sheet("Audit v1 - chybějící")
ws4.append(["ID", "Sekce", "Co chybí", "v1 má", "Odhad nákladů (Kč)"])
for c in range(1, 6):
    ws4.cell(row=1, column=c).font = hf
    ws4.cell(row=1, column=c).fill = hfill
    ws4.cell(row=1, column=c).alignment = Alignment(horizontal="center", wrap_text=True)
r = 2
for m in data["audit_v1"]["missing_sections"]:
    ws4.cell(row=r, column=1, value=m["id"])
    ws4.cell(row=r, column=2, value=m["section"])
    ws4.cell(row=r, column=3, value=m["missing_what"])
    ws4.cell(row=r, column=4, value=m["v1_has_only"])
    ws4.cell(row=r, column=5, value=m["estimate_czk"])
    for c in range(1, 6):
        ws4.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
ws4.cell(row=r, column=4, value="CELKEM odhad chybějících sekcí:").font = Font(bold=True)
ws4.cell(row=r, column=5, value=sum(m["estimate_czk"] for m in data["audit_v1"]["missing_sections"])).font = Font(bold=True)
for i, w in enumerate([8, 40, 60, 40, 18], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

# ===== Sheet 5: Blockery =====
ws5 = wb.create_sheet("Blockery pre Fáze 2")
ws5.append(["Blocker"])
ws5.cell(row=1, column=1).font = hf
ws5.cell(row=1, column=1).fill = hfill
for i, b in enumerate(data["blockers_for_phase_2"], 2):
    ws5.cell(row=i, column=1, value=b)
    ws5.cell(row=i, column=1).alignment = Alignment(wrap_text=True)
ws5.column_dimensions["A"].width = 100

out = HERE / "Vykaz_BD_Daliborova_v2_KROS.xlsx"
wb.save(out)
print(f"✓ Wrote {out}")
print(f"  Sheets: {wb.sheetnames}")
print(f"  Positions: {len(positions)}")
print(f"  Sections: {len(sections_order)}")
print(f"  Grand total (bez DPH): {grand_total:,.2f} Kč")
print(f"  Grand total (s DPH 15 %): {grand_total * 1.15:,.2f} Kč")
