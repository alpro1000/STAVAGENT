#!/usr/bin/env python3
"""
build_soupis.py — Generuje soupis prací pro Žihle 2062-1.

Vstup:  otskp_mapping.yaml (per-element OTSKP kódy + ceny)
Výstup: soupis_praci_zihle_2062-1.xml  (UNIXML 1.2 KROS format)
        soupis_praci_zihle_2062-1.xlsx (duplikát pro screen view)

Schema UNIXML per Kfely template (inputs/reference/20 Rekonstrukce mostu Kfely.xml):
  <unixml format="unixml_cz" verze="1.2">
    <typ_souboru>SOUPIS_PRACI</typ_souboru>
    <stavba>
      <kod_stavby>...</kod_stavby>
      <nazev_stavby>...</nazev_stavby>
      <objekty>
        <objekt>
          <kod_objektu>SO 001</kod_objektu>
          <nazev_objektu>Demolice...</nazev_objektu>
          <polozky>
            <polozka>
              <typ_vety>K</typ_vety>
              <typ_polozky>HSV</typ_polozky>
              <kod_polozky>014102</kod_polozky>
              <cislo_polozky>1</cislo_polozky>
              <mnozstvi>5766,70</mnozstvi>
              <typ_mnozstvi>ZADANE</typ_mnozstvi>
              <jednotkova_cena>0,00</jednotkova_cena>
              ...
"""

import os
import sys
import yaml
from pathlib import Path
from xml.etree.ElementTree import Element, SubElement, ElementTree, tostring
from xml.dom import minidom

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).parent
MAPPING_PATH = ROOT / "otskp_mapping.yaml"
XML_OUT = ROOT / "soupis_praci_zihle_2062-1.xml"
XLSX_OUT = ROOT / "soupis_praci_zihle_2062-1.xlsx"

# ─────────────────────────────────────────────────────────────────────
# Project metadata
# ─────────────────────────────────────────────────────────────────────

STAVBA = {
    "kod_stavby": "2062-1/26-DZMS",
    "nazev_stavby": "Most ev.č. 2062-1 u obce Žihle, přestavba",
    "verejna_zakazka": "ANO",
    "datum": "05.05.2026",
    "misto": "Žihle, Plzeň-sever",
    "oblast": "Plzeňský kraj",
    "zpracoval": "Sandbox Phase D",
    "objednatel": "Správa a údržba silnic Plzeňského kraje, p.o. (IČO 72053119)",
    "projektant": "TBD — zhotovitel D&B",
    "zhotovitel": "TBD",
    "investor": "Plzeňský kraj",
    "kso": "823 26",
    "cpv": "45221111-3 Výstavba silničních mostů",
    "poznamka": "Sandbox výstup — NENÍ pro odevzdání tendru. Phase D Žihle 2062-1.",
}

OBJEKTY_META = {
    "SO_001": {
        "kod_objektu": "SO 001",
        "nazev_objektu": "Demolice stávajícího mostu",
        "typ_zakazky": "STAVEBNI_OBJEKT",
        "cpv": "45111100-9",
    },
    "SO_180": {
        "kod_objektu": "SO 180",
        "nazev_objektu": "Mostní provizorium (Mabey C200 nebo ekvivalent)",
        "typ_zakazky": "STAVEBNI_OBJEKT",
        "cpv": "45221111-3",
    },
    "SO_201": {
        "kod_objektu": "SO 201",
        "nazev_objektu": "Most ev.č. 2062-1 (nová stavba — integrální rám)",
        "typ_zakazky": "STAVEBNI_OBJEKT",
        "cpv": "45221111-3",
    },
    "SO_290": {
        "kod_objektu": "SO 290",
        "nazev_objektu": "Směrová úprava silnice III/206 2 km 0,600 – 0,900",
        "typ_zakazky": "STAVEBNI_OBJEKT",
        "cpv": "45233140-2",
    },
    "ZS": {
        "kod_objektu": "ZS",
        "nazev_objektu": "Zařízení staveniště + VRN",
        "typ_zakazky": "ZARIZENI_STAVENISTE",
        "cpv": "45100000-8",
    },
}

ORDER = ["SO_001", "SO_180", "SO_201", "SO_290", "ZS"]


# ─────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────

def cz_number(value, decimals=2):
    """Format float as Czech number (comma decimal separator)."""
    if value is None:
        return "0,00"
    return f"{value:.{decimals}f}".replace(".", ",")


def is_hsv(otskp_kod: str) -> str:
    """Return 'HSV' / 'PSV' / 'M' per OTSKP code prefix.
    HSV = stavební práce hlavní (1xx zemní, 2-4xx beton/zdivo, 9xx ost.)
    PSV = specializované práce (7xx izolace + 6xx střechy)
    M = montáže (svodidla, dopr. značení 91xx)
    """
    if not otskp_kod:
        return "HSV"
    p2 = otskp_kod[:2]
    if p2 in ("71", "72", "76"):
        return "PSV"
    if p2 in ("91", "94"):  # svodidla, dopr. značení
        return "M"
    return "HSV"


# ─────────────────────────────────────────────────────────────────────
# Load mapping
# ─────────────────────────────────────────────────────────────────────

def load_mapping():
    with open(MAPPING_PATH, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


# ─────────────────────────────────────────────────────────────────────
# Build XML
# ─────────────────────────────────────────────────────────────────────

def build_xml(mapping: dict) -> Element:
    root = Element("unixml", format="unixml_cz", verze="1.2")
    SubElement(root, "typ_souboru").text = "SOUPIS_PRACI"
    SubElement(root, "rozsah_exportu").text = "CELA_STAVBA"
    SubElement(root, "zdroj").text = "STAVAGENT_SANDBOX"

    stavba = SubElement(root, "stavba")
    for k, v in STAVBA.items():
        SubElement(stavba, k).text = str(v)

    objekty = SubElement(stavba, "objekty")
    cislo_polozky_global = 1

    for so_key in ORDER:
        meta = OBJEKTY_META[so_key]
        items = mapping.get(so_key, {}) or {}

        objekt = SubElement(objekty, "objekt")
        for k, v in meta.items():
            SubElement(objekt, k).text = v
        SubElement(objekt, "datum").text = STAVBA["datum"]
        SubElement(objekt, "objednatel").text = STAVBA["objednatel"]
        SubElement(objekt, "investor").text = STAVBA["investor"]

        polozky_el = SubElement(objekt, "polozky")
        cislo_polozky_local = 1

        for item_id, item in items.items():
            kod = (item.get("otskp_kod") or "").strip()
            nazev = item.get("nazev", "")
            mj = item.get("mj", "")
            mnozstvi = item.get("mnozstvi", 0) or 0
            cena = item.get("jedn_cena_kc", 0) or 0
            confidence = item.get("confidence", 1.0)
            note = item.get("note", "")
            source = item.get("source", "")

            # Cena adjustments — for items where OTSKP catalog price is 0 but we
            # know orient. cost (from mapping cena_celkem_kc), back-calculate per-unit
            cena_celkem = item.get("cena_celkem_kc", 0) or 0
            if cena == 0 and mnozstvi > 0 and cena_celkem > 0:
                cena_per_unit = cena_celkem / mnozstvi
            else:
                cena_per_unit = cena

            polozka = SubElement(polozky_el, "polozka")
            SubElement(polozka, "typ_vety").text = "K"
            SubElement(polozka, "typ_polozky").text = is_hsv(kod)
            SubElement(polozka, "kod_polozky").text = kod
            SubElement(polozka, "cislo_polozky").text = str(cislo_polozky_local)
            SubElement(polozka, "nazev_polozky").text = nazev
            SubElement(polozka, "mj").text = mj
            SubElement(polozka, "mnozstvi").text = cz_number(mnozstvi)
            SubElement(polozka, "typ_mnozstvi").text = "ZADANE"
            SubElement(polozka, "jednotkova_cena").text = cz_number(cena_per_unit)
            SubElement(polozka, "jednotkova_cena_dodavky").text = cz_number(0)
            SubElement(polozka, "typ_sazby_dph").text = "ZAKLADNI"
            # Annotate confidence + source in poznamka_polozky for audit trail
            poznamka = []
            if confidence < 1.0:
                poznamka.append(f"Confidence: {confidence}")
            if source:
                poznamka.append(f"Source: {source}")
            if note:
                poznamka.append(f"Note: {note}")
            SubElement(polozka, "poznamka_polozky").text = " | ".join(poznamka)[:500]
            SubElement(polozka, "uzivatelske_zarazeni").text = item_id

            cislo_polozky_local += 1
            cislo_polozky_global += 1

    return root


def write_xml(root: Element, path: Path):
    raw = tostring(root, encoding="utf-8")
    parsed = minidom.parseString(raw)
    pretty = parsed.toprettyxml(indent="  ", encoding="utf-8")
    path.write_bytes(pretty)


# ─────────────────────────────────────────────────────────────────────
# Build XLSX (duplicate)
# ─────────────────────────────────────────────────────────────────────

def build_xlsx(mapping: dict, path: Path):
    wb = Workbook()
    wb.remove(wb.active)

    bold = Font(bold=True)
    header_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
    so_total_fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
    custom_fill = PatternFill(start_color="FFFCE4D6", end_color="FFFCE4D6", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    # Sheet 1: Per-položka přehled (full detail)
    ws = wb.create_sheet("Soupis_polozek")
    headers = ["SO", "Č.", "Typ", "OTSKP kód", "Název", "MJ",
               "Množství", "Jedn. cena Kč", "Cena celkem Kč",
               "Confidence", "Source", "Note"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = bold
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 30

    grand_total = 0
    cislo = 1
    for so_key in ORDER:
        meta = OBJEKTY_META[so_key]
        items = mapping.get(so_key, {}) or {}
        so_total = 0
        for item_id, item in items.items():
            kod = item.get("otskp_kod", "")
            nazev = item.get("nazev", "")
            mj = item.get("mj", "")
            mnozstvi = item.get("mnozstvi", 0) or 0
            cena = item.get("jedn_cena_kc", 0) or 0
            cena_celkem = item.get("cena_celkem_kc", 0) or 0
            if cena == 0 and mnozstvi > 0 and cena_celkem > 0:
                cena_per_unit = cena_celkem / mnozstvi
            else:
                cena_per_unit = cena
            confidence = item.get("confidence", 1.0)
            source = item.get("source", "")
            note = item.get("note", "")

            ws.append([
                meta["kod_objektu"], cislo, is_hsv(kod), kod, nazev, mj,
                round(mnozstvi, 3), round(cena_per_unit, 2), round(cena_celkem, 2),
                confidence, source[:60], note[:80],
            ])
            row = ws.max_row
            if confidence == 0.0:
                for cell in ws[row]:
                    cell.fill = custom_fill
            so_total += cena_celkem
            cislo += 1
        # SO total row
        ws.append([meta["kod_objektu"], "", "", "", f"CELKEM {meta['kod_objektu']}", "",
                   "", "", round(so_total, 2), "", "", ""])
        for cell in ws[ws.max_row]:
            cell.font = bold
            cell.fill = so_total_fill
        grand_total += so_total

    ws.append([""] * 12)
    ws.append(["", "", "", "", "CELKEM BEZ DPH", "", "", "", round(grand_total, 2), "", "", ""])
    last = ws[ws.max_row]
    for cell in last:
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="FF70AD47", end_color="FF70AD47", fill_type="solid")

    # Column widths
    widths = [10, 6, 6, 12, 60, 7, 12, 14, 16, 11, 35, 50]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = width
    ws.freeze_panes = "A2"

    # Sheet 2: Souhrn per SO
    ws2 = wb.create_sheet("Souhrn_per_SO")
    ws2.append(["SO", "Název", "Počet položek", "Cena celkem [Kč]", "% z celku", "Pozn."])
    for cell in ws2[1]:
        cell.font = bold
        cell.fill = header_fill

    for so_key in ORDER:
        meta = OBJEKTY_META[so_key]
        items = mapping.get(so_key, {}) or {}
        so_total = sum((it.get("cena_celkem_kc", 0) or 0) for it in items.values())
        n_low_conf = sum(1 for it in items.values() if (it.get("confidence", 1) or 1) < 0.7)
        pct = (so_total / grand_total * 100) if grand_total > 0 else 0
        note = ""
        if so_key == "SO_180":
            note = f"7 z {len(items)} pol. = custom non-OTSKP (provizorium, vendor RFQ)"
        elif n_low_conf > 0:
            note = f"{n_low_conf} pol. confidence < 0.7"
        ws2.append([meta["kod_objektu"], meta["nazev_objektu"],
                    len(items), round(so_total, 2), round(pct, 1), note])

    ws2.append(["", "CELKEM BEZ DPH", "", round(grand_total, 2), 100.0, ""])
    last = ws2[ws2.max_row]
    for cell in last:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FF70AD47", end_color="FF70AD47", fill_type="solid")

    for col_idx, width in enumerate([10, 50, 16, 18, 12, 60], start=1):
        ws2.column_dimensions[chr(64 + col_idx)].width = width
    ws2.freeze_panes = "A2"

    # Sheet 3: Krycí list
    ws3 = wb.create_sheet("Kryci_list")
    ws3["A1"] = "KRYCÍ LIST SOUPISU PRACÍ"
    ws3["A1"].font = Font(bold=True, size=14)
    ws3.merge_cells("A1:D1")

    krycí = [
        ("Stavba:", STAVBA["nazev_stavby"]),
        ("Kód stavby:", STAVBA["kod_stavby"]),
        ("Místo:", STAVBA["misto"]),
        ("Investor:", STAVBA["investor"]),
        ("Objednatel:", STAVBA["objednatel"]),
        ("Datum:", STAVBA["datum"]),
        ("Veřejná zakázka:", STAVBA["verejna_zakazka"]),
        ("KSO:", STAVBA["kso"]),
        ("CPV:", STAVBA["cpv"]),
        ("Zpracoval:", STAVBA["zpracoval"]),
        ("Předpokládaná hodnota (ZD §5.5):", "30 000 000 Kč bez DPH"),
        ("Max. doba realizace (ZD §5.3):", "30 měsíců"),
        ("Lhůta podání (ZD §26.1):", "02.07.2026 10:00"),
        ("CELKEM BEZ DPH:", f"{grand_total:,.2f} Kč".replace(",", " ")),
        ("Headroom proti budgetu:", f"{30_000_000 - grand_total:,.2f} Kč ({(30_000_000 - grand_total)/30_000_000*100:.1f}%)".replace(",", " ")),
    ]
    for i, (k, v) in enumerate(krycí, start=3):
        ws3[f"A{i}"] = k
        ws3[f"A{i}"].font = bold
        ws3[f"B{i}"] = v
        ws3.merge_cells(f"B{i}:D{i}")

    ws3.column_dimensions["A"].width = 36
    ws3.column_dimensions["B"].width = 56

    wb.save(path)


# ─────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────

def main():
    print(f"📂 Loading {MAPPING_PATH}")
    mapping = load_mapping()

    print(f"⚙️  Building UNIXML...")
    root = build_xml(mapping)
    write_xml(root, XML_OUT)
    print(f"✅ {XML_OUT} ({XML_OUT.stat().st_size:,} B)")

    print(f"⚙️  Building XLSX...")
    build_xlsx(mapping, XLSX_OUT)
    print(f"✅ {XLSX_OUT} ({XLSX_OUT.stat().st_size:,} B)")

    # Sanity check counts
    n_polozek = sum(len(mapping.get(k, {})) for k in ORDER)
    grand = sum(
        (it.get("cena_celkem_kc", 0) or 0)
        for k in ORDER
        for it in mapping.get(k, {}).values()
    )
    print(f"\n📊 Stats: {n_polozek} položek, total {grand:,.2f} Kč bez DPH")


if __name__ == "__main__":
    main()
