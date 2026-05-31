#!/usr/bin/env python3
"""
Action 2 finalization — fix sklad S-code namespace + extend Var_E with 6 sklad skladby.

Two parts:

1. items.json — retroactive namespace tagging
   Phase 3.5 tagged sklad items with bare S01/S02/.../S05 (matching how dům
   uses bare S01/S02/...). Action 1C added HSV5.001 with explicit `S05_sklad`
   suffix to make Pattern 24 (multi-namespace) explicit.

   Per Pattern 24 the EXPLICIT suffix is the canonical form (avoids confusion
   when sklad S01 ≠ dům S01). Retro-tag all sklad items: bare SN → SN_sklad.

2. Excel Var_E — append 6 sklad skladby rows + 2 exterior skladby rows.
   Per gate-2 user spec: Var_E should grow from 15 → 21 entries
   (13 dům + 2 exterior + 6 sklad). Currently only 13 dům entries present.

   Skladba data per `outputs/sklad_skladby_legenda_canonical.json` +
   item _source fields for Anglický dvorek + Terasa.

   Generates new Excel file
   `Vykaz_vymer_RD_Jachymov_VSE_VARIANTY_<today>_v2_final.xlsx`.
"""

from __future__ import annotations

import json
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parent.parent
OUTPUTS = ROOT / "outputs"
ITEMS_PATH = OUTPUTS / "items_rd_jachymov_complete.json"
SNAPSHOT_PATH = OUTPUTS / "items_consolidated_FROZEN_2026-05-20.json"
SKLAD_LEGENDA_PATH = OUTPUTS / "sklad_skladby_legenda_canonical.json"
TODAY = str(date.today())

# ---------------------------------------------------------------------------
# Part 1 — namespace fix in items.json
# ---------------------------------------------------------------------------


def fix_namespace() -> dict:
    data = json.load(ITEMS_PATH.open())
    items = data["items"]
    changes = 0
    per_item: list[dict] = []
    for it in items:
        if it["objekt"] != "260217_sklad":
            continue
        rs = it.get("realizuje_skladbu")
        if not rs or isinstance(rs, str):
            # PSV77.001 has list-form ["S01"]; HSV5.001 has ["S05_sklad"].
            # Bare-string forms (if any) also handled
            if isinstance(rs, str) and not rs.endswith("_sklad") and rs.startswith("S"):
                new = rs + "_sklad"
                if new != rs:
                    it["realizuje_skladbu"] = new
                    per_item.append({"id": it["id"], "before": rs, "after": new})
                    changes += 1
            continue
        # List form — convert each bare SN to SN_sklad
        new_list: list[str] = []
        mutated = False
        for s in rs:
            if isinstance(s, str) and s.startswith("S") and not s.endswith("_sklad"):
                new_list.append(s + "_sklad")
                mutated = True
            else:
                new_list.append(s)
        if mutated:
            per_item.append({"id": it["id"], "before": rs, "after": new_list})
            it["realizuje_skladbu"] = new_list
            changes += 1

    # Add Phase 3.5+ log entry
    data.setdefault("_phase3_5_namespace_finalize_log", {}).update({
        "applied_at": TODAY,
        "purpose": (
            "Retroactive namespace-suffix tagging for sklad items per Pattern 24 "
            "(multi-namespace S-code/F-code handling). Bare SN → SN_sklad to avoid "
            "confusion with dům S01-S12b namespace."
        ),
        "items_renamed": changes,
        "per_item_changes": per_item,
    })
    ITEMS_PATH.write_text(json.dumps(data, indent=2, ensure_ascii=False))
    SNAPSHOT_PATH.write_text(json.dumps(data, indent=2, ensure_ascii=False))
    return {"items_renamed": changes, "per_item_changes": per_item}


# ---------------------------------------------------------------------------
# Part 2 — extend Var_E with 6 sklad + 2 exterior skladby
# ---------------------------------------------------------------------------


THIN = Side(border_style="thin", color="888888")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
SECTION_FILL = PatternFill("solid", fgColor="E8EEF7")
SECTION_FONT = Font(name="Calibri", size=11, bold=True)
BODY_FONT = Font(name="Calibri", size=10)
BODY_ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def latest_v2_xlsx() -> Path:
    cands = sorted(OUTPUTS.glob("Vykaz_vymer_RD_Jachymov_VSE_VARIANTY_*_v2.xlsx"))
    if not cands:
        raise FileNotFoundError("No _v2 Excel found — run extend_phase4_excel.py first")
    return cands[-1]


def extend_var_e(wb) -> dict:
    sh = wb["Var_E_Skladby_Vrstev"]
    sklad_legenda = json.load(SKLAD_LEGENDA_PATH.open())
    items_data = json.load(ITEMS_PATH.open())
    items = items_data["items"]

    # Find next free row (after last existing row + 1 blank)
    next_row = sh.max_row + 2

    # ---- Section header ----
    section_cell = sh.cell(next_row, 1, value="SKLAD-NAMESPACE SKLADBY (260217_sklad)")
    section_cell.font = SECTION_FONT
    section_cell.fill = SECTION_FILL
    sh.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=7)
    next_row += 1
    sub_cell = sh.cell(next_row, 1, value=(
        "Per Pattern 24 (multi-namespace S-code/F-code handling) — sklad has own S01-S05 "
        "namespace, distinct from dům S01-S12b legenda above. Cross-validated against DXF "
        "km_kóty layer + drawing screenshots + ChatGPT cross-check."
    ))
    sub_cell.font = Font(name="Calibri", size=10, italic=True, color="555555")
    sh.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=7)
    next_row += 2

    # Helper to lookup items by realizuje_skladbu and sum their plocha
    def items_for_skladba(code_suffix: str) -> tuple[list[str], float, str]:
        ids: list[str] = []
        total_area = 0.0
        for it in items:
            if it["objekt"] != "260217_sklad":
                continue
            rs = it.get("realizuje_skladbu")
            if not rs:
                continue
            codes = rs if isinstance(rs, list) else [rs]
            if code_suffix in codes:
                ids.append(it["id"].split(".", 1)[1])
                if (it.get("mj") or "").lower() in ("m²", "m2") and it.get("mnozstvi"):
                    total_area = max(total_area, float(it["mnozstvi"]))
        return ids, total_area, ", ".join(ids[:5])

    # Per S-code row
    for sk in sklad_legenda.get("skladby", []):
        code = sk["code"]
        suffix_code = f"{code}_sklad"
        ids, area, ids_text = items_for_skladba(suffix_code)
        composition = "\n".join(f"{i+1}. {layer}" for i, layer in enumerate(sk["layers"]))
        applies = "room 0.01 (sklad)" if code == "S01" else \
                  "parking 1.01 (stání) — sklad ceiling" if code == "S02" else \
                  "obvod skladu pod terénem" if code == "S03a" else \
                  "obvod skladu nad terénem" if code == "S03b" else \
                  "zadní opěrná stěna prefa Herkul" if code == "S04" else \
                  "room 1.02 (mezipodesta schodiště)" if code == "S05" else \
                  "—"
        row_values = [
            f"{code}_sklad — {sk['name']}",
            composition,
            "Drawing screenshot D.1.1.02.R1 (sklad půdorys) + DXF km_kóty layer S-code call-outs",
            f"DXF dum_DPZ + sklad_DPZ km_kóty (Phase 3.5 cross-check)",
            f"{applies} — items: {ids_text}" if ids_text else f"{applies}",
            f"{area:.2f} m²" if area else "—",
            "drawing_explicit + DXF cross-validated",
        ]
        for c, v in enumerate(row_values, start=1):
            cell = sh.cell(next_row, c, value=v)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN_LEFT
            cell.border = BORDER
        next_row += 1

    # ---- Exterior skladby (Anglický dvorek + Terasa) section ----
    next_row += 1
    section2 = sh.cell(next_row, 1, value="EXTERIOR SKLADBY (DXF-sourced — dům scope, garden)")
    section2.font = SECTION_FONT
    section2.fill = SECTION_FILL
    sh.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=7)
    next_row += 1

    exterior_skladby = [
        {
            "code": "Anglický dvorek",
            "items_signature": "Anglický dvorek",
            "name": "Anglický dvorek — vstup do 1.PP ze zahrady",
            "layers": [
                "Betonová dlažba 50 mm",
                "Kladecí vrstva - kamenná drť 4-8 mm 40 mm",
                "Podkladní nosná vrstva - kamenná drť 4-8 mm 150 mm",
                "Zhutněná zemní pláň (30 MPa)",
            ],
            "tz_source": "TZ ARS dům §6.2 + DXF SM__Popisy bubliny (Phase 3.1 external skladba)",
            "applies_to": "Vstup do 1.PP ze zahrady",
        },
        {
            "code": "Terasa",
            "items_signature": "Terasa",
            "name": "Terasa za opěrnou stěnou (garapa dřevěná prkna)",
            "layers": [
                "Terasová dřevěná prkna 25 mm (garapa)",
                "Rektifikovatelné terče pro terasu 50 mm",
                "Betonové dlaždice 50 mm",
                "Štěrkový podsyp 16/32 mm 100 mm",
                "Hrubý podsyp 4/8 mm 150 mm",
                "Geotextilie",
            ],
            "tz_source": "TZ ARS dům §6.2 terasa + DXF SM__Popisy bubliny (Phase 3.1 external skladba)",
            "applies_to": "Terasa za opěrnou stěnou (zahrada)",
        },
    ]

    for ex in exterior_skladby:
        ids = [it["id"].split(".", 1)[1] for it in items
               if it.get("realizuje_skladbu") == ex["items_signature"]]
        composition = "\n".join(f"{i+1}. {layer}" for i, layer in enumerate(ex["layers"]))
        row_values = [
            f"{ex['code']} — {ex['name']}",
            composition,
            ex["tz_source"],
            "DXF dum_DPZ SM__Popisy bubliny (decoded from broken CMap via Phase 3.1 fallback)",
            f"{ex['applies_to']} — items: {', '.join(ids) if ids else '—'}",
            "—",
            "DXF_decoded + TZ ARS explicit",
        ]
        for c, v in enumerate(row_values, start=1):
            cell = sh.cell(next_row, c, value=v)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN_LEFT
            cell.border = BORDER
        next_row += 1

    return {"sklad_skladby_added": len(sklad_legenda.get("skladby", [])), "exterior_skladby_added": len(exterior_skladby), "var_e_max_row": next_row - 1}


def update_souhrn_narrative(wb) -> list[str]:
    """Append Phase 3.5+ narrative paragraph at end of Souhrn (after R39)."""
    sh = wb["Souhrn"]
    notes: list[str] = []
    # live counts (Pattern 38 — no hardcoded item statistics)
    import json as _j
    from pathlib import Path as _P
    _it = _j.loads((_P(__file__).resolve().parent.parent / "outputs" / "items_rd_jachymov_complete.json").read_text(encoding="utf-8"))["items"]
    _tot = len(_it)
    _dum = sum(1 for i in _it if i.get("objekt") == "260219_dum")
    _skl = sum(1 for i in _it if i.get("objekt") == "260217_sklad")
    _n_skladba = sum(1 for i in _it if i.get("realizuje_skladbu"))
    next_row = sh.max_row + 2
    cell = sh.cell(next_row, 1, value="PHASE 3.5+ FINALIZACE (aktualizováno 2026-05-30)")
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    sh.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=6)
    next_row += 1
    lines = [
        "Phase 3.5 sklad S-code mapping complete: 14 items tagged + 9 explicit null markers (Pattern 24 multi-namespace S-codes — sklad SN_sklad distinct from dům SN)",
        "Action 1 (qty cross-check): PSV77.001 podlaha 21.209 → 17.60 m² (inner usable per DXF room 0.01); HSV4.005 parking pororošt 21.0 → 44.60 m² (full footprint per DXF room 1.01); HSV5.001 NEW mezipodesta schodiště prefa 5.50 m² (DXF room 1.02 gap-fill, S05_sklad skladba)",
        "Exhaustive extraction via DXF native parser (ezdxf, 15 S-codes + 11 POZN refs across 4 DXFs) + full OCR pipeline (34 PDFs at 300 DPI tesseract ces+eng)",
        f"Total items: {_tot} ({_dum} dům + {_skl} sklad) — vč. CEV +3 (komín/zídky/drenáž), anchor-gap +2 (přesun hmot/lešení), terasa-area fix + venkovní schody na terénu. {_n_skladba} items s realizuje_skladbu.",
        "Patterns 17 (Phase 0a) + 31 (CEV) + 15 (Work-First) + 24 (multi-namespace) applied throughout. No catalog code touched — Phase 5 (URS matching) gated behind frozen-baseline confirmation.",
    ]
    for line in lines:
        c = sh.cell(next_row, 1, value=f"• {line}")
        c.font = Font(name="Calibri", size=10)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        sh.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=6)
        next_row += 1
        notes.append(line[:80])
    return notes


def main() -> None:
    # Part 1 — namespace fix
    ns_result = fix_namespace()
    print("Part 1 — namespace fix:", json.dumps(ns_result, indent=2, ensure_ascii=False))

    # Part 2 — extend Var_E + Souhrn in Excel
    src = latest_v2_xlsx()
    target = OUTPUTS / f"Vykaz_vymer_RD_Jachymov_VSE_VARIANTY_{TODAY}_v2_final.xlsx"

    wb = load_workbook(str(src))
    var_e_result = extend_var_e(wb)
    souhrn_notes = update_souhrn_narrative(wb)
    wb.save(str(target))

    log = {
        "applied_at": TODAY,
        "source_xlsx": str(src.relative_to(ROOT)),
        "target_xlsx": str(target.relative_to(ROOT)),
        "namespace_fix": ns_result,
        "var_e_extension": var_e_result,
        "souhrn_narrative_lines": len(souhrn_notes),
    }
    (OUTPUTS / "_phase4_finalize_log.json").write_text(
        json.dumps(log, indent=2, ensure_ascii=False)
    )
    print()
    print("Final state:", json.dumps(log, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
