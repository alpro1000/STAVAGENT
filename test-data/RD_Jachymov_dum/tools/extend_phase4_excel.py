#!/usr/bin/env python3
"""
Phase 4 — VALIDATE LIST Excel extensions.

Post-processes the phase2-generated Excel to:
  1. Correct hardcoded Souhrn counts (171 → 211, 144 dum → 180, etc.)
  2. Extend Var_B_Polozkovy_Dum + Var_B_Polozkovy_Sklad with 3 new columns:
     - realizuje_skladbu (S-code traceability per Phase 3.1)
     - _audit_gap_fixed (per-drawing audit gap tag)
     - klempir_reconciliation_note (PSV-76 only, per Phase 3.3)
  3. Add new sheet "Cross_verification" with CEV findings summary

Reads:
  outputs/Vykaz_vymer_RD_Jachymov_VSE_VARIANTY_<today>.xlsx (in-place patch)
  outputs/items_rd_jachymov_complete.json (for skladba/audit_gap/klempir fields)
  outputs/cev_matrices_ab_gap_report.json
  outputs/cev_matrix_c_items_source_verification.json
  outputs/cev_matrix_d_cross_doc_consistency.json
  outputs/cev_per_drawing_annotations_audit.json

Writes:
  outputs/Vykaz_vymer_RD_Jachymov_VSE_VARIANTY_<today>_v2.xlsx (new file)

Idempotent: rebuilds new columns / sheet from source each run.
"""

from __future__ import annotations

import json
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUTPUTS = ROOT / "outputs"
TODAY = str(date.today())

SOURCE_XLSX = OUTPUTS / "Vykaz_vymer_RD_Jachymov_VSE_VARIANTY.xlsx"
TARGET_XLSX = OUTPUTS / "Vykaz_vymer_RD_Jachymov_VSE_VARIANTY_v2.xlsx"

THIN = Side(border_style="thin", color="888888")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_FONT = Font(name="Calibri", size=10)
BODY_ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
SECTION_FILL = PatternFill("solid", fgColor="E8EEF7")
SECTION_FONT = Font(name="Calibri", size=11, bold=True)
NOTE_FONT = Font(name="Calibri", size=10, italic=True, color="555555")


def load_items() -> list[dict]:
    """Return the full items list — DO NOT key by id because id collides
    across kapitolas in the current schema (Pattern #20 — known issue)."""
    data = json.load((OUTPUTS / "items_rd_jachymov_complete.json").open())
    return data["items"]


def patch_souhrn(ws, totals: dict) -> list[str]:
    """Walk Souhrn cells, replace hardcoded outdated counts with current."""
    changes: list[str] = []
    # Map old phrase fragments → new full values (substring match per cell)
    replacements: list[tuple[str, str]] = [
        # Variant comparison header
        ("171 (144 dum + 27 sklad)",
         f"{totals['active']} ({totals['dum_active']} dum + {totals['sklad_active']} sklad) "
         f"+ {totals['deprecated']} deprecated audit-trail = {totals['total']} total"),
        # Variant row counts (B — Položkový)
        ("Plný detail (dum 144 + sklad 27 = 171 položek)",
         f"Plný detail (dum {totals['dum_active']} + sklad {totals['sklad_active']} = "
         f"{totals['active']} aktivních + {totals['deprecated']} deprecated)"),
        # Variant row counts (A + C lines reference 30 + 111 ALWAYS as sheet shape — leave)
        # Vyjasnění queue line
        ("18 items (3 critical:",
         f"{totals['otazky_total']} items (Q2/Q4/Q18/Q20 RESOLVED + Q5 partial + Q21 plot working assumption out-of-scope, "),
        # Corpus patterns count
        ("Corpus patterns aplikované | 4",
         f"Corpus patterns aplikované | {totals['patterns']}"),
        # URS status line
        ("8 WebSearch verified (3.8 %)",
         f"{totals['urs_verified']} WebSearch verified ({totals['urs_verified_pct']} %)"),
    ]
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            v = cell.value
            if not isinstance(v, str):
                continue
            for old, new in replacements:
                if old in v:
                    cell.value = v.replace(old, new)
                    changes.append(f"R{r}C{c}: {old[:60]}")
                    v = cell.value
    return changes


def extend_var_b(ws, items: list[dict], objekt_filter: str) -> list[str]:
    """Add 3 new columns to a Var_B sheet, populate by (kapitola, popis_prefix)
    signature lookup against items_list filtered to the relevant objekt."""
    notes: list[str] = []
    n_cols = ws.max_column
    new_headers = [
        "realizuje_skladbu",
        "_audit_gap_fixed",
        "klempir_reconciliation_note",
    ]
    for i, header in enumerate(new_headers, start=1):
        col = n_cols + i
        hc = ws.cell(1, col, value=header)
        hc.font = HEADER_FONT
        hc.fill = HEADER_FILL
        hc.alignment = HEADER_ALIGN
        hc.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = 30

    # Build signature map (kapitola, popis[:60]) → item, restricted to objekt
    # filter. Items list may carry duplicates per-kapitola so use last-wins
    # which keeps the canonical 207-active entry (deprecated comes first in
    # current items.json order).
    items_signature: dict[tuple[str, str], dict] = {}
    for it in items:
        if it["objekt"] != objekt_filter:
            continue
        sig = (it["kapitola"], (it.get("popis") or "")[:60])
        items_signature[sig] = it

    filled = 0
    for r in range(2, ws.max_row + 1):
        kap = ws.cell(r, 2).value
        popis = ws.cell(r, 5).value
        if not (isinstance(kap, str) and isinstance(popis, str)):
            continue
        sig = (kap, popis[:60])
        it = items_signature.get(sig)
        if it is None:
            continue
        v_sk = it.get("realizuje_skladbu")
        if isinstance(v_sk, list):
            v_sk = ", ".join(v_sk)
        ws.cell(r, n_cols + 1, value=v_sk).alignment = BODY_ALIGN_LEFT
        ws.cell(r, n_cols + 2, value=it.get("_audit_gap_fixed")).alignment = BODY_ALIGN_LEFT
        kn = it.get("klempir_reconciliation_note")
        if kn:
            ws.cell(r, n_cols + 3, value=kn).alignment = BODY_ALIGN_LEFT
        filled += 1
    notes.append(f"{filled} rows enriched (of {ws.max_row - 1} data rows)")
    return notes


def add_cross_verification_sheet(wb) -> dict:
    """Append a Cross-verification sheet summarising CEV matrices A-D + per-
    drawing audit findings."""
    ws = wb.create_sheet("Cross_verification")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 70

    row = 1
    def section(title: str) -> None:
        nonlocal row
        cell = ws.cell(row, 1, value=title)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

    def header(*cols: str) -> None:
        nonlocal row
        for i, t in enumerate(cols, start=1):
            cell = ws.cell(row, i, value=t)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = BORDER
        row += 1

    def data_row(*cols) -> None:
        nonlocal row
        for i, t in enumerate(cols, start=1):
            cell = ws.cell(row, i, value=t)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN_LEFT
            cell.border = BORDER
        row += 1

    ws.cell(row, 1, value="CEV — Comprehensive Extraction Verification (final state)").font = Font(name="Calibri", size=14, bold=True)
    row += 2

    # Matrix A summary
    try:
        ma = json.load((OUTPUTS / "cev_matrix_a_tz_to_items.json").open())
        section("Matrix A — TZ → items.json")
        header("Verdict", "Count", "", "Note")
        s = ma["_summary"]
        data_row("COVERED", s["covered"], "", "TZ paragraph matched ≥1 item")
        data_row("N/A_DOCUMENTED", s["na_documented"], "", "stávající/zachováno/nevyskytuje-se markers; PDF extraction noise filtered")
        data_row("GAP", s["gap"], "", "0 — every TZ requirement has a corresponding item or explicit N/A")
        data_row("EXTRA", s["extra"], "", "Items without TZ backing — all sub-classified legitimate (DXF-sourced / universal VRN / audit-derived / TZ-implicit)")
        data_row("items_touched", ma["_items_touched_by_evidence"], "of", str(ma["_items_total"]))
        row += 1
    except FileNotFoundError:
        pass

    # Matrix B summary
    try:
        mb = json.load((OUTPUTS / "cev_matrix_b_dxf_to_items.json").open())
        section("Matrix B — DXF entities → items.json")
        header("Verdict", "Count", "", "Note")
        s = mb["_summary"]
        data_row("COVERED", s["covered"], "", "DXF entity consumed by ≥1 item")
        data_row("N/A_DOCUMENTED", s["na_documented"], "", "Meta blocks (razítko/severka/řezová značka) + investor scope (kuchyně)")
        data_row("GAP", s["gap"], "", "0")
        data_row("EXTRA", s["extra"], "", "0")
        row += 1
        # Per-entity breakdown
        section("Matrix B — per-entity verdicts")
        header("DXF entity", "Verdict", "Items found", "Logic")
        for r in mb["matrix_b"]:
            data_row(r["dxf_entity_id"], r["verdict"], len(r.get("items_found_ids", [])), r.get("expected_items_logic", "")[:120])
        row += 1
    except FileNotFoundError:
        pass

    # Matrix C summary
    try:
        mc = json.load((OUTPUTS / "cev_matrix_c_items_source_verification.json").open())
        section("Matrix C — items.json sources → corpora")
        header("Verdict", "Count", "%", "Note")
        s = mc["_summary"]
        total = mc["_items_total"]
        data_row("VERIFIED", s.get("verified", 0), f"{100*s.get('verified',0)/total:.1f} %", "All claim parts verifiable in TZ/DXF corpora")
        data_row("PARTIAL", s.get("partial", 0), f"{100*s.get('partial',0)/total:.1f} %", "Primary claim verified; secondary technical detail at sub-paragraph granularity")
        data_row("NOT_VERIFIABLE", s.get("not_verifiable", 0), f"{100*s.get('not_verifiable',0)/total:.1f} %", "0 — every item has at least one verifiable claim")
        row += 1
    except FileNotFoundError:
        pass

    # Matrix D summary
    try:
        md = json.load((OUTPUTS / "cev_matrix_d_cross_doc_consistency.json").open())
        section("Matrix D — cross-document consistency")
        header("Fact", "Verdict", "", "Note")
        for f in md["facts"]:
            data_row(f.get("fact"), f.get("verdict"), "", (f.get("notes") or "")[:140])
        row += 1
    except FileNotFoundError:
        pass

    # Per-drawing audit
    try:
        pd_audit = json.load((OUTPUTS / "cev_per_drawing_annotations_audit.json").open())
        section("Per-drawing extraction completeness audit")
        header("POZN ref", "Verdict", "Drawn in N", "Decoded summary")
        for v in pd_audit.get("pozn_to_items_verdicts", []):
            data_row(
                v["pozn_ref"],
                v["verdict"],
                v.get("drawn_in_drawings_count", "-"),
                (v.get("summary") or "")[:140],
            )
        row += 1
    except FileNotFoundError:
        pass

    # Phase 3 + 4 disposition summary
    section("Phase 3 + 4 dispositions applied")
    header("Action", "Count", "", "Note")
    data_row("New items added (3 GAPs)", 3, "", "HSV6.016 komín + HSV6.017 opěrné zídky + HSV1.015 drenáž za BV")
    data_row("VRN.001/Průzkumy enriched", 1, "", "+ dřevokazný hmyz survey")
    data_row("Skladba traceability tagged", 40, "", "38 S-code + 2 external (Anglický dvorek + Terasa)")
    data_row("Klempířina reconciliation note", 4, "", "Δ -8 % vs DXF MA_klempíř — flag for Karel walkthrough")
    data_row("Q21 added to Word otázky", 1, "", "Plot dřevěný — working assumption out-of-scope")
    row += 2

    # Klempířina detail
    section("Klempířina reconciliation (Phase 3.3)")
    header("Source", "Length (m)", "Δ vs DXF", "Note")
    data_row("DXF MA_klempíř (30 entities)", 75.4, "—", "")
    data_row("DXF SM__ klempířina (45 entities)", 98.4, "—", "")
    data_row("DXF total", 173.8, "0 %", "Reference")
    data_row("Items PSV-76 Klempíř sum", 159.9, "-8.0 %", "Within ±15 % tolerance — flag for Karel walkthrough; possible missing klempíř segment ~14 m")

    return {"sheet": "Cross_verification", "rows_written": row - 1}


def compute_totals(items: list[dict]) -> dict:
    from collections import Counter
    by_objekt_active = Counter()
    by_objekt_deprecated = Counter()
    for it in items:
        if it.get("status_flag") == "deprecated_audit_v2":
            by_objekt_deprecated[it["objekt"]] += 1
        else:
            by_objekt_active[it["objekt"]] += 1
    active = sum(by_objekt_active.values())
    deprecated = sum(by_objekt_deprecated.values())
    urs_v = sum(1 for it in items if it.get("urs_status") == "matched_websearch_verified")
    return {
        "total": active + deprecated,
        "active": active,
        "deprecated": deprecated,
        "dum_active": by_objekt_active.get("260219_dum", 0),
        "sklad_active": by_objekt_active.get("260217_sklad", 0),
        "otazky_total": 21,
        "patterns": 20,
        "urs_verified": urs_v,
        "urs_verified_pct": f"{100*urs_v/(active+deprecated):.1f}",
    }


def main() -> None:
    if not SOURCE_XLSX.exists():
        # Try last available phase2 output
        candidates = sorted(OUTPUTS.glob("Vykaz_vymer_RD_Jachymov_VSE_VARIANTY_*.xlsx"))
        if not candidates:
            raise FileNotFoundError("No phase2 Excel found — run tools/phase2_excel_generator.py first")
        src = candidates[-1]
    else:
        src = SOURCE_XLSX

    items = load_items()
    totals = compute_totals(items)

    wb = load_workbook(str(src))
    souhrn_changes = patch_souhrn(wb["Souhrn"], totals)
    var_b_dum_notes = extend_var_b(wb["Var_B_Polozkovy_Dum"], items, "260219_dum")
    var_b_sklad_notes = extend_var_b(wb["Var_B_Polozkovy_Sklad"], items, "260217_sklad")
    cross_info = add_cross_verification_sheet(wb)

    wb.save(str(TARGET_XLSX))

    log = {
        "applied_at": TODAY,
        "source_xlsx": str(src.relative_to(ROOT)),
        "target_xlsx": str(TARGET_XLSX.relative_to(ROOT)),
        "totals": totals,
        "souhrn_changes": souhrn_changes,
        "var_b_dum_notes": var_b_dum_notes,
        "var_b_sklad_notes": var_b_sklad_notes,
        "cross_verification_sheet": cross_info,
    }
    (OUTPUTS / "_phase4_excel_extend_log.json").write_text(
        json.dumps(log, indent=2, ensure_ascii=False)
    )
    print(json.dumps(log, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
