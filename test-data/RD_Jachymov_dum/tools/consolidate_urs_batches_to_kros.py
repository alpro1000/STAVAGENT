#!/usr/bin/env python3
"""
Consolidate 15 STAVAGENT URS_MATCHER batch files into KROS-format rozpočet.

Input:
  test-data/RD_Jachymov_dum/inputs/_superseded/2026-05-16_unsorted_audit/
    URS_STAVAGENT_batch_001_rows_001_012.xlsx ... batch_015_rows_169_181.xlsx
    01-2025-AP - RD Valcha SO1+ SO2.xlsx (structural template reference)

Output:
  outputs/Vykaz_vymer_RD_Jachymov_KROS_format_2026-05-19.xlsx (5 sheets):
    1. Rekapitulace stavby (krycí list referencing SO sheets)
    2. SO 260219 — Dům (krycí list + rekapitulace + soupis prací)
    3. SO 260217 — Sklad+parking (same structure, fewer items)
    4. Stav položek — URS verification (status breakdown + audit table)
    5. URS selection audit (full multi-factor scoring transparency)

Multi-factor candidate selection per derived item:
    score = 0.30 × confidence_raw
          + 0.25 × (urs_matcher_service ? 0.25 : perplexity ? 0.15 : 0)
          + 0.20 × unit_exact_match (or 0.10 compat, −0.15 mismatch)
          + 0.15 × Jaccard(tokens_query, tokens_candidate)
          + 0.15 × note_explicit_mention (or −0.10 if note prefers other candidate)

Verdict per item:
    clear_winner if score ≥ 0.5 AND Δ(top, 2nd) ≥ 0.15
    close_call if score ≥ 0.5 AND Δ < 0.15 → flag review
    low_confidence if score < 0.5 → flag review
    no_candidates if all candidates empty/missing → blank kód + manual lookup
"""

from __future__ import annotations

import json
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJ = Path(__file__).resolve().parent.parent
BATCH_DIR = PROJ / "inputs" / "_superseded" / "2026-05-16_unsorted_audit"
ITEMS_JSON = PROJ / "outputs" / "items_rd_jachymov_complete.json"
OUT_XLSX = PROJ / "outputs" / "Vykaz_vymer_RD_Jachymov_KROS_format_2026-05-19.xlsx"

TODAY = date.today().isoformat()

# ── Header aliases (batch files have inconsistent column naming) ───────────
ALIASES = {
    "source_row_id":    ["source_row_id"],
    "derived_item_id":  ["derived_item_id", "split_item_id"],
    "kapitola":         ["source_kapitola", "Kapitola", "old_kapitola"],
    "subkapitola":      ["subkapitola", "old_subkapitola", "source_subkapitola"],
    "old_name":         ["old_name"],
    "old_code":         ["old_urs_code", "old_code_navrh"],
    "old_popis":        ["old_description", "old_popis"],
    "old_mj":           ["old_mj", "old_MJ"],
    "old_qty":          ["old_qty", "old_quantity", "old_mnozstvi"],
    "split_item":       ["split_item", "split_description"],
    "query":            ["STAVAGENT_query", "stavegent_query"],
    "status":           ["status", "quality_status"],
    "note":             ["note", "status_note"],
}
for n in (1, 2, 3, 4):
    ALIASES[f"c{n}_code"] = [f"candidate_{n}_code"]
    ALIASES[f"c{n}_desc"] = [f"candidate_{n}_description", f"candidate_{n}_name"]
    ALIASES[f"c{n}_unit"] = [f"candidate_{n}_unit"]
    ALIASES[f"c{n}_conf"] = [f"candidate_{n}_confidence"]
    ALIASES[f"c{n}_src"]  = [f"candidate_{n}_source"]


def strip_dia(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", str(s or "")) if not unicodedata.combining(c))


def norm(s: str) -> str:
    return re.sub(r"\s+", " ", strip_dia(s).lower()).strip()


def normalize_mj(s: str) -> str:
    s = str(s or "").strip().lower()
    s = s.replace("²", "2").replace("³", "3")
    s = s.replace("m 2", "m2").replace("m 3", "m3")
    if s in ("m", "metr"): return "m"
    if s in ("bm", "m'"): return "bm"
    if s in ("kus", "ks."): return "ks"
    return s


def compatible_mj(a: str, b: str) -> bool:
    if not a or not b:
        return False
    if a == b:
        return True
    groups = [{"m", "bm"}, {"ks", "kpl", "sada", "set", "soubor", "kus"}, {"m2"}, {"m3"}, {"kg", "t"}]
    for g in groups:
        if a in g and b in g:
            return True
    return False


def tokenize_czech(s: str) -> set[str]:
    """Strip diacritics + tokenize on whitespace + drop stopwords/short tokens."""
    n = norm(s)
    stops = {"a", "i", "v", "z", "na", "pro", "do", "se", "po", "od", "k", "u", "s",
             "nebo", "ano", "ne", "je", "tj", "tzn", "tzv"}
    return {t for t in re.split(r"[^a-z0-9]+", n) if t and t not in stops and len(t) >= 3}


def build_header_map(ws) -> dict[str, int]:
    """Map canonical field → column index based on row 1 headers."""
    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if not h:
            continue
        h_clean = str(h).strip()
        for canon, alts in ALIASES.items():
            if h_clean in alts and canon not in headers:
                headers[canon] = c
                break
    return headers


def parse_qty(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("\xa0", " ").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def parse_conf(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        f = float(v)
        return f if 0 <= f <= 1 else f / 100
    except (ValueError, TypeError):
        return None


# ── Status normalization — some batches use sentences in status column ────
STATUS_ALIASES = {
    "match": "MATCH",
    "partial": "PARTIAL",
    "no_match": "NO_MATCH",
    "no match": "NO_MATCH",
    "bad_match": "BAD_MATCH",
    "bad match": "BAD_MATCH",
}


def normalize_status(s, c1_code: str | None) -> str:
    """Best-effort status normalization. If status field has weird text, infer from candidate availability."""
    if not s:
        return "NO_MATCH" if not c1_code else "PARTIAL"
    s_lower = str(s).lower().strip()
    for key, canon in STATUS_ALIASES.items():
        if key in s_lower:
            return canon
    # Fallback: if status text mentions "nenašel" / "n/a" → NO_MATCH
    if any(k in s_lower for k in ("nenašel", "n/a", "neexistuj", "nevrátil")):
        return "NO_MATCH"
    if any(k in s_lower for k in ("vhodný", "přesný", "exactly", "matches")):
        return "MATCH"
    # Otherwise treat as PARTIAL if candidates exist, else NO_MATCH
    return "PARTIAL" if c1_code else "NO_MATCH"


# ── Multi-factor candidate scoring ────────────────────────────────────────
def score_candidate(cand: dict, source: dict) -> dict:
    score = 0.0
    breakdown = {}

    conf = cand.get("confidence")
    if conf is not None:
        breakdown["conf"] = conf * 0.30
        score += breakdown["conf"]

    src = (cand.get("source") or "").strip().lower()
    if "urs_matcher" in src:
        breakdown["src"] = 0.25
    elif "perplex" in src:
        breakdown["src"] = 0.15
    else:
        breakdown["src"] = 0.0
    score += breakdown["src"]

    src_mj = normalize_mj(source.get("old_mj"))
    cand_mj = normalize_mj(cand.get("unit"))
    if src_mj and cand_mj:
        if src_mj == cand_mj:
            breakdown["mj"] = 0.20
        elif compatible_mj(src_mj, cand_mj):
            breakdown["mj"] = 0.10
        else:
            breakdown["mj"] = -0.15
    else:
        breakdown["mj"] = 0.0
    score += breakdown["mj"]

    src_words = tokenize_czech((source.get("query") or "") + " " + (source.get("old_popis") or ""))
    cand_words = tokenize_czech(cand.get("description") or "")
    if src_words and cand_words:
        jaccard = len(src_words & cand_words) / len(src_words | cand_words)
        breakdown["jaccard"] = jaccard * 0.15
    else:
        breakdown["jaccard"] = 0.0
    score += breakdown["jaccard"]

    note = (source.get("note") or "").lower()
    cand_code = (cand.get("code") or "").strip()
    other_codes = [c["code"] for c in source.get("candidates", [])
                   if c.get("code") and c["code"] != cand_code]
    if cand_code and cand_code in note:
        breakdown["note"] = 0.15
    elif cand_code and any(oc and oc in note for oc in other_codes):
        breakdown["note"] = -0.10
    else:
        breakdown["note"] = 0.0
    score += breakdown["note"]

    return {"score": round(score, 3), "breakdown": breakdown}


def select_best(item: dict) -> dict:
    cands = [c for c in item.get("candidates", []) if c.get("code")]
    if not cands:
        return {
            "selected_code": None,
            "selected_desc": None,
            "selected_unit": None,
            "selection_reason": "no_candidates",
            "needs_review": True,
            "alternative_code": None,
            "scored": [],
        }
    scored = []
    for c in cands:
        s = score_candidate(c, item)
        scored.append({**c, **s})
    scored.sort(key=lambda x: -x["score"])
    best = scored[0]
    best_score = best["score"]
    second = scored[1] if len(scored) > 1 else None
    second_score = second["score"] if second else -1

    if best_score < 0.5:
        reason = f"low_confidence_review_needed (score {best_score:.2f})"
        needs_review = True
        alt = second["code"] if second else None
    elif second and (best_score - second_score) < 0.15:
        reason = f"close_call_top_2 (Δ {best_score - second_score:.2f})"
        needs_review = True
        alt = second["code"]
    else:
        reason = f"clear_winner (score {best_score:.2f}, Δ {best_score - second_score:.2f})"
        needs_review = False
        alt = None

    return {
        "selected_code": best["code"],
        "selected_desc": best.get("description"),
        "selected_unit": best.get("unit"),
        "selected_confidence": best.get("confidence"),
        "selected_source": best.get("source"),
        "selected_score": best_score,
        "selection_reason": reason,
        "needs_review": needs_review,
        "alternative_code": alt,
        "alternative_desc": second.get("description") if alt else None,
        "scored": scored,
    }


# ── Batch parser ──────────────────────────────────────────────────────────
def parse_batch(path: Path) -> list[dict]:
    """Return list of derived items from a single batch."""
    wb = load_workbook(path, data_only=True)
    cand_sheets = [s for s in wb.sheetnames if "candidat" in s.lower() or "batch" in s.lower()]
    if not cand_sheets:
        return []
    # Prefer the one with most data
    ws_name = max(cand_sheets, key=lambda s: wb[s].max_row)
    ws = wb[ws_name]
    hmap = build_header_map(ws)
    out = []
    for r in range(2, ws.max_row + 1):
        def g(field):
            c = hmap.get(field)
            return ws.cell(row=r, column=c).value if c else None
        # Skip empty rows
        if g("derived_item_id") is None and g("source_row_id") is None:
            continue
        item = {
            "batch": path.stem,
            "source_row_id": g("source_row_id"),
            "derived_item_id": g("derived_item_id") or g("source_row_id"),
            "kapitola": g("kapitola") or "",
            "subkapitola": g("subkapitola") or "",
            "old_name": g("old_name") or "",
            "old_code": g("old_code") or "",
            "old_popis": g("old_popis") or "",
            "old_mj": g("old_mj") or "",
            "old_qty": parse_qty(g("old_qty")) or 0,
            "split_item": g("split_item") or "",
            "query": g("query") or "",
            "status_raw": g("status") or "",
            "note": g("note") or "",
            "candidates": [],
        }
        c1_code = (g("c1_code") or "").strip() if g("c1_code") else ""
        item["status"] = normalize_status(item["status_raw"], c1_code)
        for n in (1, 2, 3, 4):
            code = g(f"c{n}_code")
            if code and str(code).strip() and str(code).strip().upper() not in ("N/A", "NA", "NONE"):
                item["candidates"].append({
                    "rank": n,
                    "code": str(code).strip(),
                    "description": g(f"c{n}_desc") or "",
                    "unit": g(f"c{n}_unit") or "",
                    "confidence": parse_conf(g(f"c{n}_conf")),
                    "source": g(f"c{n}_src") or "",
                })
        out.append(item)
    return out


# ── Objekt assignment (dum vs sklad) — cross-ref to items.json ────────────
def assign_objekt_to_items(items: list[dict], items_json: list[dict]) -> None:
    """Match each derived item to items.json popis to infer dum vs sklad."""
    # Build lookup: norm(popis_words first 5) → objekt
    popis_to_objekt = {}
    for it in items_json:
        for popis_field in ("popis", "popis_was_fabricated"):
            p = it.get(popis_field)
            if not p:
                continue
            tokens = tokenize_czech(p)
            if tokens:
                key = " ".join(sorted(tokens)[:5])
                popis_to_objekt[key] = it.get("objekt", "260219_dum")
    # Match each derived item
    for d in items:
        # Try direct kapitola check first — some batches have "(sklad)" or distinct sklad markers in old_name
        kap_l = (d.get("kapitola") or "").lower() + " " + (d.get("old_name") or "").lower()
        if "sklad" in kap_l:
            d["objekt"] = "260217_sklad"
            continue
        # Fallback to popis match
        candidate_text = (d.get("old_popis") or "") + " " + (d.get("old_name") or "") + " " + (d.get("split_item") or "")
        tok = tokenize_czech(candidate_text)
        if not tok:
            d["objekt"] = "260219_dum"
            continue
        key = " ".join(sorted(tok)[:5])
        d["objekt"] = popis_to_objekt.get(key, "260219_dum")
        # If still ambiguous, look for sklad / parking keywords in popis
        if "sklad" in norm(candidate_text) or "parking" in norm(candidate_text):
            d["objekt"] = "260217_sklad"


# ── TKP family / section name mapping for KROS layout ─────────────────────
def kros_section(item: dict) -> tuple[str, str, str]:
    """Return (parent_section, sub_section_code, sub_section_name) per KROS taxonomy."""
    code = (item.get("selected_code") or item.get("old_code") or "").strip()
    digit1 = code[0] if code and code[0].isdigit() else "?"
    digit3 = code[:3] if len(code) >= 3 and code[:3].isdigit() else "?"
    kap = (item.get("kapitola") or "").lower()
    # HSV first digits 1-9
    if digit1 in "12345679":
        hsv_subs = {
            "1": "Zemní práce",
            "2": "Zakládání",
            "3": "Svislé a kompletní konstrukce",
            "4": "Vodorovné konstrukce",
            "5": "Komunikace pozemní",
            "6": "Úpravy povrchů, podlahy a osazování výplní",
            "9": "Ostatní konstrukce a práce-bourání",
        }
        if digit1 in hsv_subs:
            return ("HSV - Práce a dodávky HSV", digit1, hsv_subs[digit1])
        if digit1 == "7":
            # 7xx codes — PSV. Sub-section by 3-digit family
            psv_subs = {
                "711": "Izolace proti vodě, vlhkosti a plynům",
                "712": "Izolace povlakové střech",
                "713": "Izolace tepelné",
                "721": "Vnitřní kanalizace",
                "722": "Vnitřní vodovod",
                "725": "Zařizovací předměty",
                "731": "Ústřední vytápění - kotelny",
                "732": "Ústřední vytápění - strojovny",
                "733": "Ústřední vytápění - rozvody",
                "734": "Ústřední vytápění - armatury",
                "735": "Ústřední vytápění - otopná tělesa",
                "736": "Ústřední vytápění - plošné vytápění",
                "741": "Elektroinstalace - silnoproud",
                "742": "Elektroinstalace - slaboproud",
                "743": "Elektroinstalace",
                "751": "Vzduchotechnika",
                "762": "Konstrukce tesařské",
                "763": "Konstrukce suché výstavby",
                "764": "Konstrukce klempířské",
                "765": "Krytina skládaná",
                "766": "Konstrukce truhlářské",
                "767": "Konstrukce zámečnické",
                "771": "Podlahy z dlaždic",
                "775": "Podlahy skládané",
                "776": "Podlahy povlakové",
                "781": "Obklady keramické",
                "783": "Dokončovací práce - nátěry",
                "784": "Dokončovací práce - malby",
                "789": "Povrchové úpravy ocelových konstrukcí",
            }
            sub_name = psv_subs.get(digit3, f"{digit3} - Práce PSV")
            return ("PSV - Práce a dodávky PSV", digit3, sub_name)
    # If no URS code OR can't determine, fall back to kapitola text
    if "vrn" in kap or item.get("kapitola", "").startswith("VRN"):
        return ("VRN - Vedlejší rozpočtové náklady", "0", "Vedlejší rozpočtové náklady")
    if "psv-7" in kap:
        # Try to derive 7xx from kapitola text
        m = re.search(r"PSV-(\d{2})", item.get("kapitola", ""))
        if m:
            sub_digits = "7" + m.group(1)
            return ("PSV - Práce a dodávky PSV", sub_digits, item.get("kapitola", ""))
    if "m-21" in kap or "elektroinstalac" in kap:
        return ("PSV - Práce a dodávky PSV", "741", "Elektroinstalace - silnoproud")
    # Last resort
    return ("Ostatní", "?", item.get("kapitola", "Bez zařazení"))


# ── Excel styling ─────────────────────────────────────────────────────────
THIN = Side(border_style="thin", color="888888")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F3A5F")
HDR_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="E8EEF7")
SUBSECTION_FILL = PatternFill("solid", fgColor="F5F8FB")
REVIEW_FILL = PatternFill("solid", fgColor="FFF1CC")
NOMATCH_FILL = PatternFill("solid", fgColor="FFE4E4")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")


def write_krycí_list(ws, stavba: str, objekt: str, row_offset: int = 4) -> int:
    """Write a krycí list block. Returns next available row."""
    r = row_offset
    ws.cell(row=r, column=4, value="KRYCÍ LIST SOUPISU PRACÍ").font = Font(name="Arial", size=12, bold=True, color="1F3A5F")
    r += 2
    ws.cell(row=r, column=4, value="Stavba:").font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=5, value=stavba).font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=4, value="Objekt:").font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=5, value=objekt).font = Font(bold=True)
    r += 2
    ws.cell(row=r, column=4, value="Místo:"); ws.cell(row=r, column=6, value="Jáchymov")
    ws.cell(row=r, column=9, value="Datum:"); ws.cell(row=r, column=10, value=TODAY)
    r += 2
    ws.cell(row=r, column=4, value="Zadavatel:").font = Font(bold=True)
    ws.cell(row=r, column=5, value="Mgr. Jindřich Volný, Fibichova 733, 362 51 Jáchymov")
    r += 2
    ws.cell(row=r, column=4, value="Zhotovitel:").font = Font(bold=True)
    ws.cell(row=r, column=5, value="Ing. Karel Šmíd")
    r += 2
    ws.cell(row=r, column=4, value="Projektant:").font = Font(bold=True)
    ws.cell(row=r, column=5, value="SMASH architekti / TeAnau s.r.o. (statika) / TUSPO (PBŘ)")
    r += 2
    ws.cell(row=r, column=4, value="Zpracovatel:").font = Font(bold=True)
    ws.cell(row=r, column=5, value="Aleksandr Pro (STAVAGENT pipeline)")
    r += 3
    return r


def write_so_sheet(wb: Workbook, so_name: str, so_label: str, items: list[dict]) -> dict:
    """Write a complete SO sheet (krycí list + rekapitulace + soupis). Returns stats."""
    ws = wb.create_sheet(so_name)
    ws.sheet_view.showGridLines = False
    r = write_krycí_list(ws, "RD Jáchymov Fibichova 733", so_label)
    # Cena placeholder rows
    r += 1
    ws.cell(row=r, column=4, value="Cena bez DPH").font = Font(bold=True)
    ws.cell(row=r, column=10, value="(doplňte po vyplnění J.cen)").font = Font(italic=True, color="888888")
    r += 4

    # ── REKAPITULACE ČLENĚNÍ ──
    ws.cell(row=r, column=3, value="REKAPITULACE ČLENĚNÍ SOUPISU PRACÍ").font = Font(name="Arial", size=11, bold=True, color="1F3A5F")
    r += 2
    # Group items
    grouped = defaultdict(lambda: defaultdict(list))
    for it in items:
        parent, sub_code, sub_name = kros_section(it)
        grouped[parent][(sub_code, sub_name)].append(it)
    parent_order = ["HSV - Práce a dodávky HSV", "PSV - Práce a dodávky PSV",
                    "VRN - Vedlejší rozpočtové náklady", "Ostatní"]
    rekap_start = r
    for parent in parent_order:
        if parent not in grouped:
            continue
        ws.cell(row=r, column=4, value=parent).font = Font(bold=True)
        ws.cell(row=r, column=4).fill = SECTION_FILL
        r += 1
        for (sub_code, sub_name), sub_items in sorted(grouped[parent].items()):
            ws.cell(row=r, column=4, value=f"    {sub_code} - {sub_name}")
            ws.cell(row=r, column=10, value=f"({len(sub_items)} položek)")
            ws.cell(row=r, column=10).font = Font(italic=True, color="888888")
            r += 1
        r += 1

    r += 2
    # ── SOUPIS PRACÍ ──
    ws.cell(row=r, column=3, value="SOUPIS PRACÍ").font = Font(name="Arial", size=11, bold=True, color="1F3A5F")
    r += 2
    # Header row
    hdr_row = r
    headers = ["PČ", "Typ", "Kód", "Popis", "MJ", "Množství", "J.cena [CZK]", "Cena celkem [CZK]", "Pozn."]
    for ci, h in enumerate(headers):
        cell = ws.cell(row=r, column=3 + ci, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    r += 1

    pos = 1
    stats = {"clear": 0, "close_call": 0, "low_conf": 0, "no_match": 0, "total": 0}
    for parent in parent_order:
        if parent not in grouped:
            continue
        # Parent section divider
        ws.cell(row=r, column=4, value="D")
        ws.cell(row=r, column=5, value=parent.split(" - ")[0]).font = Font(bold=True)
        ws.cell(row=r, column=6, value=parent.split(" - ", 1)[1] if " - " in parent else parent).font = Font(bold=True)
        for c in range(3, 12):
            ws.cell(row=r, column=c).fill = SECTION_FILL
        r += 1
        for (sub_code, sub_name), sub_items in sorted(grouped[parent].items()):
            # Sub-section divider
            ws.cell(row=r, column=4, value="D")
            ws.cell(row=r, column=5, value=sub_code).font = Font(bold=True)
            ws.cell(row=r, column=6, value=sub_name).font = Font(bold=True)
            for c in range(3, 12):
                ws.cell(row=r, column=c).fill = SUBSECTION_FILL
            r += 1
            for it in sub_items:
                sel = it.get("selection", {})
                ws.cell(row=r, column=3, value=pos).alignment = Alignment(horizontal="center")
                ws.cell(row=r, column=4, value="K")
                ws.cell(row=r, column=5, value=sel.get("selected_code") or "")
                # Popis: prefer split_item + old_popis (cleaner)
                popis = it.get("split_item") or it.get("old_popis") or it.get("old_name")
                ws.cell(row=r, column=6, value=popis).alignment = Alignment(wrap_text=True, vertical="top")
                ws.cell(row=r, column=7, value=normalize_mj(it.get("old_mj"))).alignment = Alignment(horizontal="center")
                ws.cell(row=r, column=8, value=it.get("old_qty") or 0).number_format = "#,##0.00"
                ws.cell(row=r, column=8).alignment = Alignment(horizontal="right")
                # J.cena + cena celkem — empty for Karel to fill (with formula for cena celkem)
                ws.cell(row=r, column=9).number_format = "#,##0.00"
                ws.cell(row=r, column=10, value=f"=H{r}*I{r}")
                ws.cell(row=r, column=10).number_format = "#,##0.00"
                # Poznámka — review flag
                if sel.get("needs_review"):
                    if "no_candidates" in sel.get("selection_reason", ""):
                        ws.cell(row=r, column=11, value="MANUAL LOOKUP — žádný kandidát").fill = NOMATCH_FILL
                        ws.cell(row=r, column=11).font = Font(color="9C0006", bold=True)
                        stats["no_match"] += 1
                    elif "low_confidence" in sel.get("selection_reason", ""):
                        ws.cell(row=r, column=11, value=f"REVIEW — low conf ({sel.get('selected_score',0):.2f})").fill = REVIEW_FILL
                        ws.cell(row=r, column=11).font = Font(color="9C5700", bold=True)
                        stats["low_conf"] += 1
                    else:  # close_call
                        alt = sel.get("alternative_code")
                        ws.cell(row=r, column=11, value=f"REVIEW — alt: {alt}").fill = REVIEW_FILL
                        ws.cell(row=r, column=11).font = Font(color="9C5700", bold=True)
                        stats["close_call"] += 1
                else:
                    ws.cell(row=r, column=11, value="").fill = GREEN_FILL
                    stats["clear"] += 1
                for c in range(3, 12):
                    ws.cell(row=r, column=c).border = BORDER
                stats["total"] += 1
                pos += 1
                r += 1

    # Column widths
    widths = [3, 5, 5, 13, 60, 8, 12, 13, 16, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return stats


def write_rekapitulace_stavby(wb: Workbook, so_stats: dict) -> None:
    ws = wb.create_sheet("Rekapitulace stavby", 0)
    ws.sheet_view.showGridLines = False
    r = 1
    ws.cell(row=r, column=4, value="Export STAVAGENT pipeline").font = Font(italic=True, color="888888")
    r += 3
    ws.cell(row=r, column=4, value="REKAPITULACE STAVBY").font = Font(name="Arial", size=14, bold=True, color="1F3A5F")
    r += 2
    rows = [
        ("Kód:",           "260219-RD-Jachymov-2026-05-19"),
        ("Stavba:",        "RD Jáchymov Fibichova 733"),
        ("Místo:",         "Jáchymov, parc. č. 1094/16 + st. 1022, k.ú. Jáchymov 656437"),
        ("Datum:",         TODAY),
        ("Zadavatel:",     "Mgr. Jindřich Volný, Fibichova 733, 362 51 Jáchymov"),
        ("Zhotovitel:",    "Ing. Karel Šmíd (smid.karell@gmail.com, +420 608 930 914)"),
        ("Projektant:",    "SMASH architekti s.r.o. (Ing. arch. M. Smolka, ČKA 05394) / TeAnau s.r.o. (statika) / TUSPO (PBŘ)"),
        ("Zpracovatel:",   "Aleksandr Pro (STAVAGENT pipeline v1.x — 15 batches URS_MATCHER + multi-factor scoring)"),
        ("Poznámka:",      "Rozpočet vygenerován z items.json (208 items) cross-referenced s URS_MATCHER batches. Audit trail v Sheet 4-5."),
    ]
    for label, value in rows:
        ws.cell(row=r, column=4, value=label).font = Font(bold=True)
        ws.cell(row=r, column=5, value=value)
        r += 1
    r += 2

    ws.cell(row=r, column=4, value="REKAPITULACE OBJEKTŮ").font = Font(name="Arial", size=11, bold=True, color="1F3A5F")
    r += 2
    headers = ["Kód", "Popis", "Položek", "Cena bez DPH [CZK]", "DPH 21%", "Cena s DPH [CZK]"]
    for ci, h in enumerate(headers):
        cell = ws.cell(row=r, column=3 + ci, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    r += 1
    for so_label, count in [
        ("SO 260219 - Dům",         so_stats.get("dum_total", 0)),
        ("SO 260217 - Sklad+parking", so_stats.get("sklad_total", 0)),
    ]:
        ws.cell(row=r, column=3, value="SO" if "260219" in so_label else "SO")
        ws.cell(row=r, column=4, value=so_label)
        ws.cell(row=r, column=5, value=count).alignment = Alignment(horizontal="right")
        ws.cell(row=r, column=6, value="(doplňte)").font = Font(italic=True, color="888888")
        ws.cell(row=r, column=7, value="(auto 21%)").font = Font(italic=True, color="888888")
        ws.cell(row=r, column=8, value="(doplňte)").font = Font(italic=True, color="888888")
        for c in range(3, 9):
            ws.cell(row=r, column=c).border = BORDER
        r += 1
    r += 1
    ws.cell(row=r, column=4, value="CELKEM").font = Font(bold=True, size=11)
    ws.cell(row=r, column=5, value=so_stats.get("dum_total", 0) + so_stats.get("sklad_total", 0)).font = Font(bold=True)

    widths = [3, 5, 14, 35, 9, 18, 15, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_stav_polozek(wb: Workbook, items: list[dict]) -> None:
    ws = wb.create_sheet("Stav polozek - URS verification")
    ws.sheet_view.showGridLines = False
    r = 1
    ws.cell(row=r, column=2, value="STAV POLOŽEK — URS verification").font = Font(name="Arial", size=14, bold=True, color="1F3A5F")
    r += 2

    status_counts = Counter(it.get("status") for it in items)
    sel_counts = Counter(it["selection"]["selection_reason"].split(" ")[0] for it in items if it.get("selection"))
    review_count = sum(1 for it in items if it.get("selection", {}).get("needs_review"))
    no_cand = sum(1 for it in items if not it.get("selection", {}).get("selected_code"))

    rows = [
        ("Total derived items", len(items)),
        ("", ""),
        ("Status z batches (URS_MATCHER):", ""),
        ("  MATCH",      status_counts.get("MATCH", 0)),
        ("  PARTIAL",    status_counts.get("PARTIAL", 0)),
        ("  NO_MATCH",   status_counts.get("NO_MATCH", 0)),
        ("  BAD_MATCH",  status_counts.get("BAD_MATCH", 0)),
        ("", ""),
        ("Multi-factor selection (composite score):", ""),
        ("  Clear winner (no review needed)",    sel_counts.get("clear_winner", 0)),
        ("  Close call top-2 (Δ < 0.15)",        sel_counts.get("close_call_top_2", 0)),
        ("  Low confidence (score < 0.5)",       sel_counts.get("low_confidence_review_needed", 0)),
        ("  No candidates (blank kód)",          sel_counts.get("no_candidates", 0)),
        ("", ""),
        ("Total needing manual review",          review_count),
        ("Items with blank kód (manual lookup)", no_cand),
    ]
    for label, value in rows:
        ws.cell(row=r, column=2, value=label).font = Font(bold=label.endswith(":") or label.startswith("Total"))
        if isinstance(value, int) and label:
            ws.cell(row=r, column=4, value=value).alignment = Alignment(horizontal="right")
            ws.cell(row=r, column=4).font = Font(bold=True)
        r += 1
    r += 2

    # Per-item table
    ws.cell(row=r, column=2, value="DETAIL — Items needing review or with blanks").font = Font(name="Arial", size=11, bold=True, color="1F3A5F")
    r += 2
    headers = ["derived_id", "kapitola", "popis (truncated)", "selected_code", "selection_reason", "alternative_code"]
    for ci, h in enumerate(headers):
        cell = ws.cell(row=r, column=2 + ci, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    r += 1
    for it in items:
        sel = it.get("selection", {})
        if not sel.get("needs_review"):
            continue
        ws.cell(row=r, column=2, value=str(it.get("derived_item_id", "")))
        ws.cell(row=r, column=3, value=it.get("kapitola", "")[:30])
        ws.cell(row=r, column=4, value=(it.get("split_item") or it.get("old_popis") or "")[:70])
        ws.cell(row=r, column=5, value=sel.get("selected_code") or "—")
        ws.cell(row=r, column=6, value=sel.get("selection_reason", ""))
        ws.cell(row=r, column=7, value=sel.get("alternative_code") or "—")
        if not sel.get("selected_code"):
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = NOMATCH_FILL
        else:
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = REVIEW_FILL
        r += 1

    widths = [3, 15, 20, 60, 14, 35, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_urs_audit(wb: Workbook, items: list[dict]) -> None:
    ws = wb.create_sheet("URS selection audit")
    ws.sheet_view.showGridLines = False
    r = 1
    ws.cell(row=r, column=2, value="URS SELECTION AUDIT — full multi-factor scoring per item").font = Font(name="Arial", size=12, bold=True, color="1F3A5F")
    r += 1
    ws.cell(row=r, column=2, value="Composite score = 0.30×conf + 0.25×source + 0.20×mj_match + 0.15×jaccard + 0.15×note_hint").font = Font(italic=True, color="555555")
    r += 2
    headers = ["derived_id", "popis", "rank", "code", "description", "unit", "conf", "source",
               "score", "selected?"]
    for ci, h in enumerate(headers):
        cell = ws.cell(row=r, column=2 + ci, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    r += 1
    for it in items:
        sel = it.get("selection", {})
        scored = sel.get("scored", [])
        if not scored:
            ws.cell(row=r, column=2, value=str(it.get("derived_item_id", "")))
            ws.cell(row=r, column=3, value=(it.get("split_item") or it.get("old_popis") or "")[:60])
            ws.cell(row=r, column=4, value="—")
            ws.cell(row=r, column=5, value="NO CANDIDATES")
            for c in range(2, 12):
                ws.cell(row=r, column=c).fill = NOMATCH_FILL
            r += 1
            continue
        for i, c in enumerate(scored[:4]):
            ws.cell(row=r, column=2, value=str(it.get("derived_item_id", "")) if i == 0 else "")
            ws.cell(row=r, column=3, value=(it.get("split_item") or it.get("old_popis") or "")[:60] if i == 0 else "")
            ws.cell(row=r, column=4, value=c.get("rank"))
            ws.cell(row=r, column=5, value=c.get("code") or "")
            ws.cell(row=r, column=6, value=(c.get("description") or "")[:70])
            ws.cell(row=r, column=7, value=c.get("unit") or "")
            ws.cell(row=r, column=8, value=c.get("confidence"))
            ws.cell(row=r, column=9, value=(c.get("source") or "")[:20])
            ws.cell(row=r, column=10, value=c.get("score")).number_format = "0.000"
            is_selected = (c.get("code") == sel.get("selected_code"))
            is_alt = (c.get("code") == sel.get("alternative_code"))
            if is_selected:
                ws.cell(row=r, column=11, value="✓ SELECTED")
                for cc in range(2, 12):
                    ws.cell(row=r, column=cc).fill = GREEN_FILL
            elif is_alt:
                ws.cell(row=r, column=11, value="⚠ ALT")
                for cc in range(2, 12):
                    ws.cell(row=r, column=cc).fill = YELLOW_FILL
            r += 1
        r += 1  # spacer between items

    widths = [3, 10, 50, 5, 13, 50, 6, 6, 18, 8, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main() -> int:
    # ── 1. Load batches ────────────────────────────────────────────────────
    print(f"[1/4] Loading 15 batches from {BATCH_DIR.name}/...", file=sys.stderr)
    batch_files = sorted(BATCH_DIR.glob("URS_STAVAGENT_batch_*.xlsx"))
    if len(batch_files) != 15:
        print(f"WARN: expected 15 batches, found {len(batch_files)}", file=sys.stderr)
    all_items = []
    for bf in batch_files:
        items = parse_batch(bf)
        all_items.extend(items)
        print(f"  {bf.name}: {len(items)} derived items", file=sys.stderr)
    print(f"  TOTAL: {len(all_items)} derived items", file=sys.stderr)

    # ── 2. Assign objekt (dum vs sklad) via items.json cross-ref ───────────
    print(f"[2/4] Assigning objekt (dum vs sklad)...", file=sys.stderr)
    items_json = json.loads(ITEMS_JSON.read_text())["items"]
    assign_objekt_to_items(all_items, items_json)
    objekt_dist = Counter(it.get("objekt") for it in all_items)
    print(f"  Distribution: {dict(objekt_dist)}", file=sys.stderr)

    # ── 3. Multi-factor candidate selection ────────────────────────────────
    print(f"[3/4] Multi-factor candidate scoring...", file=sys.stderr)
    for it in all_items:
        it["selection"] = select_best(it)
    sel_dist = Counter(it["selection"]["selection_reason"].split(" ")[0] for it in all_items)
    print(f"  Selection: {dict(sel_dist)}", file=sys.stderr)
    review_count = sum(1 for it in all_items if it["selection"]["needs_review"])
    print(f"  Needing review: {review_count}", file=sys.stderr)

    # ── 4. Generate KROS Excel ─────────────────────────────────────────────
    print(f"[4/4] Writing KROS-format Excel...", file=sys.stderr)
    dum_items = [it for it in all_items if it.get("objekt") == "260219_dum"]
    sklad_items = [it for it in all_items if it.get("objekt") == "260217_sklad"]

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    stats_dum = write_so_sheet(wb, "SO 260219 - Dum", "SO 260219 - Dům", dum_items)
    print(f"  ✓ SO 260219 — Dům: {stats_dum['total']} items", file=sys.stderr)
    stats_sklad = write_so_sheet(wb, "SO 260217 - Sklad+parking", "SO 260217 - Sklad+parking", sklad_items)
    print(f"  ✓ SO 260217 — Sklad: {stats_sklad['total']} items", file=sys.stderr)

    write_rekapitulace_stavby(wb, {"dum_total": stats_dum["total"], "sklad_total": stats_sklad["total"]})
    print(f"  ✓ Rekapitulace stavby (sheet 1)", file=sys.stderr)

    write_stav_polozek(wb, all_items)
    print(f"  ✓ Stav položek (sheet 4)", file=sys.stderr)

    write_urs_audit(wb, all_items)
    print(f"  ✓ URS selection audit (sheet 5)", file=sys.stderr)

    wb.save(str(OUT_XLSX))
    size = OUT_XLSX.stat().st_size
    print(f"\n✓ Wrote {OUT_XLSX.relative_to(PROJ)} ({size:,} bytes)", file=sys.stderr)
    print(f"\n=== Summary ===", file=sys.stderr)
    print(f"  Total derived items: {len(all_items)}", file=sys.stderr)
    print(f"  Dum: {len(dum_items)} | Sklad: {len(sklad_items)}", file=sys.stderr)
    print(f"  Selection: clear_winner={sel_dist.get('clear_winner', 0)}, "
          f"close_call={sel_dist.get('close_call_top_2', 0)}, "
          f"low_conf={sel_dist.get('low_confidence_review_needed', 0)}, "
          f"no_match={sel_dist.get('no_candidates', 0)}", file=sys.stderr)
    print(f"  Needing review: {review_count}/{len(all_items)} ({review_count/max(len(all_items),1)*100:.0f}%)", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
