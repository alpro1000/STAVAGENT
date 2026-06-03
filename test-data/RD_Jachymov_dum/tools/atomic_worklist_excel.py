#!/usr/bin/env python3
"""
§6 — Atomic worklist Excel generator (HK212 6-sheet hybrid).

Reads outputs/atomic_decomposition_map.json (238 atomic ops). Generates
outputs/Vykaz_vymer_RD_Jachymov_ATOMIC_WORKLIST_2026-05-27.xlsx with 6 sheets:

  1. Souhrn               — structure overview + distributions
  2. 260219_DUM_HSV       — phase-ordered HSV-1..7
  3. 260219_DUM_PSV       — PSV-71..95 + M-21 folded
  4. 260219_DUM_VRN       — VRN lump-sum 1:1
  5. 260217_SKLAD         — HSV + PSV + VRN combined
  6. Composite_Decomposition_Map — parent → atomic children

Columns (HK212 style): Poř. / Kapitola / Atomic operace popis / MJ /
Množství / URS kód kandidát / Status / Parent frozen item_id /
Realizuje skladbu / Pozn.

items.json frozen NOT touched. Pattern 26 honest blanks (no 999/TBD).
"""

from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
MAP_PATH = ROOT / "outputs" / "atomic_decomposition_map.json"
TARGET = ROOT / "outputs" / "Vykaz_vymer_RD_Jachymov_ATOMIC_WORKLIST.xlsx"

# Styles
THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
KAP_FILL = PatternFill("solid", fgColor="D6E0F0")
KAP_FONT = Font(name="Calibri", size=10, bold=True)
BODY_FONT = Font(name="Calibri", size=9)
DECOMP_FILL = PatternFill("solid", fgColor="FFF6E0")   # decomposed children — amber tint
BLANK_FILL = PatternFill("solid", fgColor="FFE0E0")    # blank URS code — red tint
VERIFIED_FILL = PatternFill("solid", fgColor="E0F0E0") # verified — green tint
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="top")
CENTER = Alignment(horizontal="center", vertical="top")
TITLE_FONT = Font(name="Calibri", size=14, bold=True)
SECTION_FONT = Font(name="Calibri", size=11, bold=True)
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color="555555")

COLS = ["Poř.", "Kapitola", "Atomic operace (popis)", "MJ", "Množství",
        "Vzorec / Zdroj výměry", "URS kód kandidát", "Status",
        "Parent frozen item_id", "Realizuje skladbu", "Pozn."]
COL_WIDTHS = [6, 22, 52, 7, 10, 46, 15, 20, 22, 15, 44]

# Phase order within HSV/PSV per HK212 construction sequence
HSV_ORDER = ["HSV-1 Zemní práce", "HSV-2 Základové a ŽB", "HSV-3 Svislé konstrukce",
             "HSV-4 Vodorovné", "HSV-5 Krov + střecha", "HSV-5 Komunikace + schodiště",
             "HSV-6 Bourací práce", "HSV-7 Fasáda ETICS"]
PSV_ORDER = ["PSV-71 Izolace HI", "PSV-71 Izolace TI", "PSV-72 ZTI", "PSV-73 Vytápění",
             "PSV-76 Klempíř", "PSV-76 Truhlář", "PSV-76 Výplně otvorů", "PSV-76 Zámečnictví",
             "PSV-77 Podlahy", "PSV-78 Povrchové úpravy", "PSV-95 Detekce požární",
             "M-21 ELI silnoproud"]
VRN_ORDER = ["VRN — Zařízení staveniště", "VRN — Doprava + odpad", "VRN — BOZP",
             "VRN — Pojištění + zábory", "VRN — Průzkumy", "VRN — Geodet",
             "VRN — Dokumentace", "VRN — Revize", "VRN — Kolaudace", "VRN — Společné"]


def _code_digits(code) -> int:
    if not code:
        return 0
    return len(re.sub(r"[^\d]", "", str(code)))


def classify_status(op: dict) -> str:
    """4-tier finalized status (Alexander principle — work always remains,
    code state is orthogonal):
      ✓ verified              — leaf code verified (Phase 5 / carried / consensus leaf)
      ⚠ kandidát-verify       — 9-digit leaf from generator, NOT yet verified
      ⚠ família ???           — family resolved, leaf pending
      ❌ blank — ÚRS online    — no code; bind later from app.urs.cz
    """
    s = (op.get("status") or "").lower()
    code = op.get("urs_kod_kandidat")
    digits = _code_digits(code)
    family = op.get("urs_code_family_6digit")

    if not code and not family:
        return "❌ blank — ÚRS online"
    if not code and family:
        return "⚠ família ??? (ÚRS online)"
    # has a code
    if digits >= 9:
        if ("verified" in s and "family" not in s) or "carried_verified" in s \
           or s == "matched_websearch_verified" or "cross_discipline" in s \
           or "reconciled_596" in s:
            return "✓ verified"
        return "⚠ kandidát-verify"
    # family-only code (≤6 digit)
    return "⚠ família ??? (ÚRS online)"


URS_ONLINE_NOTE = "Doplnit leaf z CS ÚRS online (app.urs.cz, cen. úroveň 2026 01) dle família + popis + Vzorec."


def status_icon(status: str) -> str:
    s = (status or "").lower()
    if "verified" in s and "cross_discipline" not in s and "carried" not in s:
        return "✓ verified"
    if "carried_verified" in s or status == "matched_websearch_verified":
        return "✓ verified"
    if "cross_discipline" in s:
        return "✓ cross-disc"
    if "family" in s:
        return "⚠ family_only"
    if "wrong_leaf" in s:
        return "⚠ wrong_leaf"
    if "manual_lookup" in s:
        return "❌ MANUAL LOOKUP"
    if "needs_lookup" in s or "needs_production" in s:
        return "? needs_lookup"
    return f"? {status}"


def write_header(ws, row=1):
    for c, (name, w) in enumerate(zip(COLS, COL_WIDTHS), start=1):
        cell = ws.cell(row, c, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"


def write_op_row(ws, row, op, poradi_override=None):
    rs = op.get("realizuje_skladbu")
    if isinstance(rs, list):
        rs = ", ".join(rs)
    code = op.get("urs_kod_kandidat")
    is_decomp = op.get("decomposition_type") in ("skladba_vrstva", "fixture")
    is_blank = not code
    status_4tier = classify_status(op)
    # Show família in code column when leaf blank but family known
    family = op.get("urs_code_family_6digit")
    code_display = code if code else (f"{family} ???" if family else "")
    # Append ÚRS-online note to Pozn for blank/family ops (work always remains)
    base_pozn = op.get("pozn") or ""
    if ("família" in status_4tier) or ("blank" in status_4tier):
        base_pozn = (base_pozn + "  " + URS_ONLINE_NOTE).strip() if base_pozn else URS_ONLINE_NOTE
    vals = [
        poradi_override if poradi_override is not None else op["poradi"],
        op["kapitola"],
        op["atomic_operace_popis"],
        op["mj"],
        op["mnozstvi"],
        op.get("qty_formula") or "",
        code_display,
        status_4tier,
        f"{op['parent_frozen_item_id']} → {op['atomic_id']}" if is_decomp else op["parent_frozen_item_id"],
        rs or "",
        base_pozn,
    ]
    # Column indices (1-based): 1=Poř 4=MJ 5=Množství 8=Status
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(row, c, value=v)
        cell.font = BODY_FONT
        cell.border = BORDER
        cell.alignment = RIGHT if c == 5 else (CENTER if c in (1, 4, 8) else LEFT)
    # Row tint
    if is_blank:
        for c in range(1, len(COLS) + 1):
            ws.cell(row, c).fill = BLANK_FILL
    elif is_decomp:
        for c in range(1, len(COLS) + 1):
            ws.cell(row, c).fill = DECOMP_FILL
    elif "verified" in (op.get("status") or "").lower():
        ws.cell(row, 7).fill = VERIFIED_FILL


def write_kapitola_divider(ws, row, kapitola, n_ops):
    cell = ws.cell(row, 1, value=f"━━ {kapitola}  ({n_ops} atomic operací) ━━")
    cell.font = KAP_FONT
    cell.fill = KAP_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLS))
    for c in range(1, len(COLS) + 1):
        ws.cell(row, c).fill = KAP_FILL


def build_profession_sheet(wb, sheet_name, ops, kapitola_order):
    ws = wb.create_sheet(sheet_name)
    write_header(ws)
    row = 2
    by_kap = defaultdict(list)
    for op in ops:
        by_kap[op["kapitola"]].append(op)
    # Order kapitolas per construction sequence; unknown kapitolas appended
    ordered_kaps = [k for k in kapitola_order if k in by_kap]
    for k in by_kap:
        if k not in ordered_kaps:
            ordered_kaps.append(k)
    seq = 0
    for kap in ordered_kaps:
        kap_ops = by_kap[kap]
        write_kapitola_divider(ws, row, kap, len(kap_ops))
        row += 1
        for op in kap_ops:
            seq += 1
            write_op_row(ws, row, op, poradi_override=seq)
            row += 1
    return ws, seq


def build_souhrn(wb, data):
    ws = wb.create_sheet("Souhrn")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 60
    summ = data["_summary"]
    r = 1
    ws.cell(r, 1, value="ATOMIC WORKLIST — RD Jáchymov Fibichova 733").font = TITLE_FONT
    r += 1
    ws.cell(r, 1, value="HK212 atomic-operation principle — composite skladby/fixtures decomposed to codeable operations").font = NOTE_FONT
    r += 2

    def section(title):
        nonlocal r
        c = ws.cell(r, 1, value=title)
        c.font = SECTION_FONT
        c.fill = KAP_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1

    def kv(k, v, note=""):
        nonlocal r
        ws.cell(r, 1, value=k).font = BODY_FONT
        ws.cell(r, 2, value=v).font = Font(name="Calibri", size=10, bold=True)
        if note:
            ws.cell(r, 3, value=note).font = NOTE_FONT
        r += 1

    section("PŘEHLED STRUKTURY")
    kv("Frozen items (zdroj)", summ["frozen_items_total"], "items_rd_jachymov_complete.json — FROZEN, nedotčeno")
    kv("Atomic operací celkem", summ["atomic_operations_total"], "= carried 1:1 + decomposed children")
    kv("Items decomposed", summ["items_decomposed"], "composite skladby + multi-fixture")
    kv("Atomic children z dekompozice", summ["atomic_children_from_decomposition"], "")
    kv("Items carried 1:1", summ["items_carried_1to1"], "atomic + VRN lump-sum")
    r += 1

    section("ATOMIC OPERACE PER KAPITOLA")
    ws.cell(r, 1, value="Kapitola").font = HEADER_FONT
    ws.cell(r, 1).fill = HEADER_FILL
    ws.cell(r, 2, value="Počet").font = HEADER_FONT
    ws.cell(r, 2).fill = HEADER_FILL
    r += 1
    for kap, n in summ["atomic_per_kapitola"].items():
        kv(kap, n)
    r += 1

    # STAV KÓDŮ — 4-tier code-state summary computed from ops
    from collections import Counter as _C
    tier = _C(classify_status(o) for o in data["atomic_operations"])
    section("STAV KÓDŮ (4 tiers — práce kompletní bez ohledu na stav kódu)")
    for label, key in [("✓ verified (leaf ověřen)", "✓ verified"),
                       ("⚠ kandidát-verify (leaf generátor, neověřen)", "⚠ kandidát-verify"),
                       ("⚠ família ??? (leaf z ÚRS online)", "⚠ família ??? (ÚRS online)"),
                       ("❌ blank — ÚRS online", "❌ blank — ÚRS online")]:
        kv(label, tier.get(key, 0))
    r += 1
    note_lines = [
        "Kódy: leaf binding doplnit z CS ÚRS online (app.urs.cz, cenová úroveň 2026 01) dle família + popis + Vzorec.",
        "Reconciliation consensus: HSV1.004 dvorek 596811220 · HSV1.005 terasa = podkladní skladba 564 (dlaždice ROZNÁŠECÍ POD terče, ŘEZ C-C) · dřevo 762 = PSV76.002 Truhlář · HSV2.003/008 bednění 274.",
        "PRÁCE JE KOMPLETNÍ bez ohledu na stav kódu — seznam (postup prací) = primary deliverable. Code = secondary, doplní se v ÚRS online / KROS.",
    ]
    for line in note_lines:
        ws.cell(r, 1, value="• " + line).font = NOTE_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1
    r += 1

    section("URS FAMÍLIA DISTRIBUCE (54 distinct)")
    ws.cell(r, 1, value="Família").font = HEADER_FONT
    ws.cell(r, 1).fill = HEADER_FILL
    ws.cell(r, 2, value="Počet ops").font = HEADER_FONT
    ws.cell(r, 2).fill = HEADER_FILL
    r += 1
    fam_sorted = sorted(summ["familia_distribution"].items(), key=lambda x: -x[1])
    for fam, n in fam_sorted:
        note = ""
        if fam == "(blank)":
            note = "❌ MANUAL LOOKUP — Pattern 26 honest blank (NE fabrikováno)"
        kv(fam, n, note)
    r += 1

    section("METODIKA + PATTERN COMPLIANCE")
    for line in [
        "HK212 028a..f princip: composite → atomic codeable operace (per vrstva / per fixture)",
        "Pattern 15 Work-First: atomic ops = codeable units, catalog matching downstream",
        "Pattern 26: família-level URS kde leaf neznámý + status flag; 9 BLANK ops (NE 999/TBD)",
        "Pattern 28: (id, kapitola) compound key — řeší PSV76.003 ×3 + VRN.001 ×9 collisions",
        "Pattern 32: separate deliverable — items.json frozen + File A audit NEDOTČENO",
        "Traceability: každá atomic op → parent_frozen_item_id (100% coverage)",
        "Terasa ŘEZ C-C oprava: dlaždice = roznášecí vrstva POD terče (564, NE '636311 na terče'); dřevěná pochozí vrstva (762) = PSV76.002 Truhlář (split-by-trade). Plochy situace: terasa 9.23 / dvorek 16.54 m².",
    ]:
        ws.cell(r, 1, value="• " + line).font = BODY_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1
    return ws


def build_decomp_map(wb, data):
    ws = wb.create_sheet("Composite_Decomposition_Map")
    headers = ["Parent frozen item_id", "Kapitola", "Parent popis", "Parent MJ",
               "Parent qty", "N children", "Atomic children IDs"]
    widths = [24, 22, 50, 8, 10, 11, 40]
    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(1, c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    row = 2
    for dm in data["decomposition_map"]:
        vals = [dm["parent_frozen_item_id"], dm["parent_kapitola"], dm["parent_popis"],
                dm["parent_mj"], dm["parent_qty"], dm["n_atomic_children"],
                ", ".join(dm["atomic_children_ids"])]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row, c, value=v)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = RIGHT if c == 5 else (CENTER if c in (4, 6) else LEFT)
            cell.fill = DECOMP_FILL
        row += 1
    return ws


def build_pending_decisions(wb):
    """PENDING DECISIONS sheet — surfaces open vyjasnění + skipped anchor gaps
    so they're not forgotten during Stage 3 catalog binding. Reads
    vyjasneni_queue.json (open items) — these works are NOT in items.json
    (verified-not-in-TZ per Stage 1B), but Karel must decide on them."""
    queue_path = ROOT / "inputs" / "meta" / "vyjasneni_queue.json"
    try:
        with queue_path.open() as f:
            q = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        q = {"items": []}
    open_vyj = [v for v in q.get("items", []) if v.get("status") == "open"]

    ws = wb.create_sheet("PENDING_DECISIONS")
    headers = ["#", "Severity", "Téma (čeká rozhodnutí)", "Stav v rozpočtu", "Next action"]
    widths = [6, 12, 50, 30, 60]
    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(1, c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

    # Intro note row
    intro = ws.cell(2, 1, value=(
        "⚠ Tyto práce NEJSOU v items.json ani v atomic worklistu — byly ověřeny "
        "jako NEzmíněné v ARS dům TZ (Stage 1B-verify). Před cenotvorbou / katalogovým "
        "mapováním (Stage 3) je nutné rozhodnutí projektanta/investora. NE zapomenout."
    ))
    intro.font = NOTE_FONT
    intro.alignment = LEFT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    row = 3

    for v in open_vyj:
        vals = [
            v.get("id"),
            v.get("severity", ""),
            v.get("title", ""),
            "NEPŘIDÁNO (verify-first)",
            (v.get("next_action") or "")[:300],
        ]
        for c, val in enumerate(vals, start=1):
            cell = ws.cell(row, c, value=val)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = CENTER if c == 1 else LEFT
            cell.fill = BLANK_FILL  # red tint — needs attention
        row += 1

    # PM05 was SKIP-silent (not even a vyjasnění) — surface it too
    pm05 = [
        "PM05", "medium",
        "Okapový chodník + obvodová drenáž domu — NOT_IN_TZ (žádná zmínka v TZ)",
        "SKIP (rekonstrukce — možná existuje)",
        "Volitelné: ověřit zda okapový chodník po obvodu domu je v scope (mimo BV drenáž HSV1.015). Pokud ano → HSV-1 položka.",
    ]
    for c, val in enumerate(pm05, start=1):
        cell = ws.cell(row, c, value=val)
        cell.font = BODY_FONT
        cell.border = BORDER
        cell.alignment = CENTER if c == 1 else LEFT
        cell.fill = DECOMP_FILL  # amber — informational skip
    row += 1

    return ws, len(open_vyj) + 1


def main() -> None:
    # Defensive load — decomposition map may be missing / malformed / lack key
    try:
        with MAP_PATH.open() as f:
            data = json.load(f)
    except FileNotFoundError:
        raise SystemExit(f"ERROR: decomposition map not found: {MAP_PATH} (run atomic_decomposition.py first)")
    except json.JSONDecodeError as e:
        raise SystemExit(f"ERROR: invalid JSON in {MAP_PATH}: {e}")
    if "atomic_operations" not in data or not isinstance(data["atomic_operations"], list):
        raise SystemExit(f"ERROR: missing or malformed 'atomic_operations' list in {MAP_PATH}")
    ops = data["atomic_operations"]

    # Partition ops — tolerate missing objekt/kapitola via .get() (defensive)
    def _obj(o):
        return o.get("objekt", "")

    def _kap(o):
        return o.get("kapitola", "") or ""

    dum_hsv = [o for o in ops if _obj(o) == "260219_dum" and _kap(o).startswith("HSV")]
    dum_psv = [o for o in ops if _obj(o) == "260219_dum" and (_kap(o).startswith("PSV") or _kap(o).startswith("M-21"))]
    dum_vrn = [o for o in ops if _obj(o) == "260219_dum" and _kap(o).startswith("VRN")]
    sklad = [o for o in ops if _obj(o) == "260217_sklad"]

    wb = Workbook()
    wb.remove(wb.active)  # drop default

    build_souhrn(wb, data)
    _, n_hsv = build_profession_sheet(wb, "260219_DUM_HSV", dum_hsv, HSV_ORDER)
    _, n_psv = build_profession_sheet(wb, "260219_DUM_PSV", dum_psv, PSV_ORDER)
    _, n_vrn = build_profession_sheet(wb, "260219_DUM_VRN", dum_vrn, VRN_ORDER)
    _, n_sklad = build_profession_sheet(wb, "260217_SKLAD", sklad, HSV_ORDER + PSV_ORDER + VRN_ORDER)
    build_decomp_map(wb, data)
    _, n_pending = build_pending_decisions(wb)

    TARGET.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(str(TARGET))
    except (OSError, PermissionError) as e:
        raise SystemExit(f"ERROR: failed to save {TARGET}: {e}")

    # Validation
    total_rows = n_hsv + n_psv + n_vrn + n_sklad
    traceability_ok = all(o.get("parent_frozen_item_id") for o in ops)
    fabricated = [o for o in ops if o.get("urs_kod_kandidat") in ("999999999", "TBD", "999")]

    print(json.dumps({
        "file": str(TARGET.relative_to(ROOT)),
        "sheets": wb.sheetnames,
        "sheet_count": len(wb.sheetnames),
        "row_counts": {
            "DUM_HSV": n_hsv, "DUM_PSV": n_psv, "DUM_VRN": n_vrn, "SKLAD": n_sklad,
            "total_atomic_rows": total_rows,
        },
        "atomic_ops_in_map": len(ops),
        "rows_match_map": total_rows == len(ops),
        "decomp_map_parents": len(data["decomposition_map"]),
        "decomp_children": sum(d["n_atomic_children"] for d in data["decomposition_map"]),
        "traceability_100pct": traceability_ok,
        "fabricated_codes": len(fabricated),
        "pattern_26_honest_blanks": data["_summary"]["familia_distribution"].get("(blank)", 0),
        "pending_decisions_rows": n_pending,
    }, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
