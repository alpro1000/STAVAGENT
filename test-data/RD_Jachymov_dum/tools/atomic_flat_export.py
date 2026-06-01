#!/usr/bin/env python3
"""
ATOMIC FLAT export (2026-06-01) — THE single non-aggregated deliverable.

One sheet, ONE row per atomic operation (montáž / materiál split, P41). No aggregation,
no per-objekt tabs. Joins _source + urs_kapitola_hint from the parent item.

Output: outputs/ATOMIC_FLAT_<date>.xlsx
Columns: Pořadí | Atomic ID | Objekt | Kapitola | ÚRS kapitola hint | Popis | MJ |
         Množství | Vzorec | ÚRS kód kandidát | Zdroj | Status | Realizuje skladbu
Source of truth: atomic_decomposition_map.json (atomic_operations[]) + items.json (parent).
"""
from __future__ import annotations
import json
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "outputs"
MAP = OUT / "atomic_decomposition_map.json"
ITEMS = OUT / "items_rd_jachymov_complete.json"

COLS = ["Pořadí", "Atomic ID", "Objekt", "Kapitola", "ÚRS kapitola hint", "Popis",
        "MJ", "Množství", "Vzorec", "ÚRS kód kandidát", "Zdroj", "Status", "Realizuje skladbu"]


def main() -> None:
    amap = json.loads(MAP.read_text(encoding="utf-8"))
    ops = amap["atomic_operations"]
    items = {i["id"]: i for i in json.loads(ITEMS.read_text(encoding="utf-8"))["items"]}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ATOMIC_FLAT"

    hdr_fill = PatternFill("solid", fgColor="2F3640")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    null_fill = PatternFill("solid", fgColor="FFF3CD")  # amber for qty=null/neurčeno

    for c, name in enumerate(COLS, 1):
        cell = ws.cell(1, c, name)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for r, op in enumerate(ops, start=2):
        parent = items.get(op.get("parent_frozen_item_id"), {})
        src = parent.get("source", "")
        hint = parent.get("urs_kapitola_hint") or parent.get("kapitola", "")
        rsk = op.get("realizuje_skladbu") or parent.get("realizuje_skladbu") or []
        if isinstance(rsk, list):
            rsk = ", ".join(rsk)
        qty = op.get("mnozstvi")
        row = [
            op.get("poradi", r - 1), op.get("atomic_id", ""), op.get("objekt", ""),
            op.get("kapitola", ""), hint, op.get("atomic_operace_popis", ""),
            op.get("mj", ""), ("neurčeno" if qty is None else qty),
            op.get("qty_formula", ""), op.get("urs_kod_kandidat") or "",
            src, op.get("status", ""), rsk,
        ]
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (5, 6, 9, 11)))
            if c == 8 and qty is None:
                cell.fill = null_fill

    widths = [7, 22, 14, 24, 26, 50, 7, 11, 26, 14, 34, 22, 16]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(COLS))}{len(ops)+1}"

    out = OUT / f"ATOMIC_FLAT_{date.today().isoformat()}.xlsx"
    wb.save(out)
    n_null = sum(1 for o in ops if o.get("mnozstvi") is None)
    print(f"OK — {out.name}: {len(ops)} atomic operations (1 row each), {n_null} qty=neurčeno.")


if __name__ == "__main__":
    main()
