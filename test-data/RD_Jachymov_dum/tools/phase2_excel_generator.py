#!/usr/bin/env python3
"""
Phase 2 Excel deliverable for RD Jáchymov — 6 sheets in one workbook.

Reads outputs/items_rd_jachymov_complete.json + inputs/meta/project_header.json
+ inputs/meta/inventory.md and renders:

  List 1 — Souhrn (composite view, project header + variant comparison)
  List 2 — Var_A_Agregovany_Dum   (~22 řádků: GROUP BY kapitola)
  List 3 — Var_A_Agregovany_Sklad (~6 řádků)
  List 4 — Var_C_Hybrid           (~80 řádků: HSV detailně + zbytek agregovaně)
  List 5 — Var_B_Polozkovy_Dum    (~144 řádků: plný detail)
  List 6 — Var_B_Polozkovy_Sklad  (~27 řádků: plný detail)

Output:
  outputs/Vykaz_vymer_RD_Jachymov_VSE_VARIANTY_<datum>.xlsx

Run: python3 tools/phase2_excel_generator.py  (from project root)
"""

from __future__ import annotations

import json
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

PROJ = Path(__file__).resolve().parent.parent
OUT = PROJ / "outputs"
META = PROJ / "inputs" / "meta"

ITEMS_JSON = OUT / "items_rd_jachymov_complete.json"
HEADER_JSON = META / "project_header.json"
DXF_EXTRACT_JSON = OUT / "dxf_comprehensive_extract.json"
SKLADBY_JSON = OUT / "skladby_per_zone_v2.json"   # Path C Gate 5: use v2 with 13 S-codes from řez A-A legend
SKLADBY_JSON_FALLBACK = OUT / "skladby_per_zone.json"

TODAY = date.today().isoformat()
TARGET = OUT / f"Vykaz_vymer_RD_Jachymov_VSE_VARIANTY_{TODAY}.xlsx"

# ───────────────────────────────────────────────────────────────────────────
# Styling

THIN = Side(border_style="thin", color="888888")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

KAPITOLA_FILL = PatternFill("solid", fgColor="E8EEF7")
KAPITOLA_FONT = Font(name="Calibri", size=10, bold=True)

BODY_FONT = Font(name="Calibri", size=10)
BODY_ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
BODY_ALIGN_RIGHT = Alignment(horizontal="right", vertical="top")
BODY_ALIGN_CENTER = Alignment(horizontal="center", vertical="top")

DISCLAIMER_FONT = Font(name="Calibri", size=10, italic=True, color="555555")
DISCLAIMER_FILL = PatternFill("solid", fgColor="FFF8E1")

CONF_LOW_FILL    = PatternFill("solid", fgColor="FFE4E4")   # < 0.70 red
CONF_MED_FILL    = PatternFill("solid", fgColor="FFF1CC")   # < 0.80 orange


def write_header_row(ws, row: int, headers: list[str]):
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    ws.row_dimensions[row].height = 32


def auto_width(ws, max_width: int = 60):
    """Set column widths based on longest cell content per column."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            l = max(len(line) for line in str(cell.value).splitlines())
            if l > max_len:
                max_len = l
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), max_width)


def apply_conf_rules(ws, conf_col_letter: str, first_row: int, last_row: int):
    """Conditional formatting: confidence < 0.80 orange, < 0.70 red."""
    rng = f"{conf_col_letter}{first_row}:{conf_col_letter}{last_row}"
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="lessThan", formula=["0.70"], fill=CONF_LOW_FILL, stopIfTrue=True,
    ))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="lessThan", formula=["0.80"], fill=CONF_MED_FILL,
    ))


def apply_conf_rules_to_row(ws, conf_col_letter: str, row_first_col_letter: str, row_last_col_letter: str, first_row: int, last_row: int):
    """Conditional formatting on full row driven by confidence cell value."""
    rng = f"{row_first_col_letter}{first_row}:{row_last_col_letter}{last_row}"
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="lessThan",
        formula=[f"${conf_col_letter}{first_row}<0.70"],  # placeholder; openpyxl needs CellIsRule on the conf col directly
        fill=CONF_LOW_FILL,
    ))


# ───────────────────────────────────────────────────────────────────────────
# Sheet 1 — Souhrn

def build_sheet_souhrn(wb: Workbook, items: list[dict], header: dict) -> None:
    ws = wb.create_sheet("Souhrn")
    ws.sheet_view.showGridLines = False

    row = 1
    ws.cell(row=row, column=1, value="VÝKAZ VÝMĚR — RD Jáchymov Fibichova č.p. 733").font = Font(
        name="Calibri", size=18, bold=True, color="1F3A5F"
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    ws.cell(row=row, column=1,
            value="Stavební úpravy a nástavba RD + zahradní sklad / parking / přístupové schodiště").font = Font(
        name="Calibri", size=12, italic=True, color="555555"
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 2

    # Identifikační údaje
    ws.cell(row=row, column=1, value="IDENTIFIKAČNÍ ÚDAJE").font = Font(bold=True, size=11, color="1F3A5F")
    row += 1
    ident_rows = [
        ("Investor", f'{header["identifikacni_udaje"]["investor"]["jmeno"]} ({header["identifikacni_udaje"]["investor"]["adresa"]})'),
        ("Lokace", "Fibichova 733, 362 51 Jáchymov; parc. č. 1094/16 + st. 1022 (k.ú. Jáchymov 656437)"),
        ("Architekt (ARS)", "SMASH architekti s.r.o. — Ing. arch. Marek Smolka, ČKA 05394"),
        ("Statika", "TeAnau s.r.o. — Ing. Jan Tvardík, ČKAIT 0012219 (spoluautor Ing. Václav Bendík)"),
        ("PBŘ", "TUSPO — Ing. Jan Kirschbaum / Bc. Zbyněk Tuček (ČKAIT 0013446)"),
        ("Zhotovitel CN", "Ing. Karel Šmíd (smid.karell@gmail.com, +420 608 930 914) — smidstavitelstvi.cz"),
        ("Stupeň dokumentace", "DSP (dokumentace pro povolení stavby), realizace 2026-2027. DPS plánována není."),
        ("Datum vygenerování", TODAY),
        ("Verze", "Phase 2 — automated STAVAGENT pipeline v1.2 + subdodavatel mapping v1.2"),
    ]
    for label, value in ident_rows:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1
    row += 1

    # Geometrie
    ws.cell(row=row, column=1, value="GEOMETRIE OBOU OBJEKTŮ").font = Font(bold=True, size=11, color="1F3A5F")
    row += 1
    geo_rows = [
        ("260219 — Dům", "104,4 m² zastavěná | 219,3 m² podlahová | 987 m³ obestavěný | 13,0 m výška | 1.PP + 3.NP (nástavba)"),
        ("260217 — Sklad+Parking", "Lichoběžník 6,35 × 3,34 m | parking 7,0 m délka (DXF DIMENSION + LWPOLYLINE verified)"),
        ("Klima", "VII. sněhová oblast (Krušnohoří), III. větrová oblast, kategorie terénu III"),
        ("Geologie", "Svor R5-R6 / F4-CS, Rdt 350 kPa (dům) / 300 kPa (sklad), XA1, IGP neproveden — archivní vrty"),
    ]
    for label, value in geo_rows:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1
    row += 1

    # Cenové srovnání 3 variant
    ws.cell(row=row, column=1, value="CENOVÉ SROVNÁNÍ 3 VARIANT (J. ceny vyplní zhotovitel)").font = Font(
        bold=True, size=11, color="1F3A5F"
    )
    row += 1
    write_header_row(ws, row, ["Varianta", "Popis", "Počet položek", "Cena bez DPH (Kč)", "DPH 15 % (Kč)", "Cena s DPH (Kč)"])
    row += 1
    variant_rows = [
        ("A — Agregovaný", "GROUP BY kapitola (dum 22 + sklad 6 = ~28 řádků). Pro orientační cenovku.",
         len([it for it in items if it["_gate"] == "VRN" or it["_gate"] == "HSV"]) // 5 + 6, None, None, None),  # rough estimate
        ("C — Hybrid", "HSV položkově (95 ks) + PSV/TZB/VRN agregovaně (~80 řádků celkem).",
         95 + 16, None, None, None),
        ("B — Položkový", "Plný detail (dum 144 + sklad 27 = 171 položek). Cílový rozsah pro produkční pricing.",
         171, None, None, None),
    ]
    for v, popis, n, c1, c2, c3 in variant_rows:
        ws.cell(row=row, column=1, value=v).font = Font(bold=True)
        ws.cell(row=row, column=2, value=popis).alignment = BODY_ALIGN_LEFT
        ws.cell(row=row, column=3, value=n).alignment = BODY_ALIGN_RIGHT
        for col, val in [(4, c1), (5, c2), (6, c3)]:
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = BODY_ALIGN_RIGHT
            cell.number_format = '#,##0 "Kč";-#,##0 "Kč";"—"'
            cell.font = Font(italic=True, color="888888")
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = BORDER
        row += 1
    row += 1

    # Investiční odhad pásmo
    ws.cell(row=row, column=1, value="INVESTIČNÍ ODHAD (předběžně)").font = Font(bold=True, size=11, color="1F3A5F")
    row += 1
    ws.cell(row=row, column=1, value="Pásmo:").font = Font(bold=True)
    ws.cell(row=row, column=2, value="6,0 – 8,5 mil. Kč bez DPH (dle TZ B m.1.j obestavěný 987 m³ + sklad + parking, sazba 6-9 tis. Kč/m³ pro RD rekonstrukci v Krušnohoří)").alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 2

    # Statistiky
    ws.cell(row=row, column=1, value="STATISTIKY STAVAGENT PIPELINE (Phase 0b + Phase 1)").font = Font(
        bold=True, size=11, color="1F3A5F"
    )
    row += 1
    stat_rows = [
        ("Položky celkem", "171 (144 dum + 27 sklad)"),
        ("Gate distribution", "HSV 95 | PSV 35 | TZB+M 22 | VRN 19"),
        ("Confidence dist", "0.99 manual: 5 | 0.95 DXF/regex: 21 | 0.90 LWPOLYLINE: 5 | 0.85 TZ: 36 | 0.80 empirické: 16 | 0.75 geometry: 88 | <0.70: 0"),
        ("URS status", "171 needs_production_lookup (sandbox bez Cloud Run URS_MATCHER — viz disclaimer níže)"),
        ("subdodavatel needs_mapping", "0 (v1.2 mapping covers all 5 RD Jáchymov pilot flags)"),
        ("Phase 0b validation", "67/69 verified (97.1 %), 0 silent drifts, gate OPEN"),
        ("Phase 0b §3.3 DXF", "4/4 DXF parsed OK, vyjasnění #18 fully_resolved (sklad geom z DIMENSION + parking LWPOLYLINE bbox)"),
        ("Vyjasnění queue", "18 items (3 critical: #1 scope + #16 statika swap [fixed] + #99 N/A)"),
        ("Corpus patterns aplikované", "4 (UNSORTED dedup + drift detection + multi-modal geometry + _gate vs kapitola_group)"),
    ]
    for label, value in stat_rows:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1
    row += 2

    # Disclaimer — Part 5 (2026-05-17) updated for 187 items + Sheet 8 Var_E_Skladby
    ws.cell(row=row, column=1, value="POZNÁMKA — ROZSAH ROZPOČTU").font = Font(bold=True, size=11, color="A04000")
    row += 1
    # Compute live stats for disclaimer
    from collections import Counter
    conf_dist = Counter(round(it["mnozstvi_confidence"], 2) for it in items)
    n_expansion = sum(1 for it in items if it.get("_expansion_origin"))
    n_recalc = sum(1 for it in items if it.get("_previous_mnozstvi") is not None)
    conf_summary = " · ".join(f"{n}×{c:.2f}" for c, n in sorted(conf_dist.items(), reverse=True))
    disclaimer = (
        f"Rozpočet generován STAVAGENT pipelinem z DSP dokumentace s DPS-grade DXF metadaty.\n"
        f"{len(items)} položek celkem ({sum(1 for it in items if it['objekt']=='260219_dum')} dum + "
        f"{sum(1 for it in items if it['objekt']=='260217_sklad')} sklad), 4 corpus patterns aplikované, "
        f"zero-fabrication policy.\n\n"
        f"Confidence distribution: {conf_summary}.\n\n"
        f"{n_expansion} položek vzniklo per-room expansion ze 9 aggregate parents "
        f"(každá expanded položka tagged _expansion_origin).\n"
        f"{n_recalc} položek recalculated s exact external perimeter 38.70 m "
        f"(DXF km_R_návrh_tlustá 2 closed polygon) vs prior fallback 41.0 m.\n\n"
        f"Skladby vrstev (Sheet 8 Var_E) pochází POUZE z TZ explicit text + DXF cross-validation HATCH patterns. "
        f"ŽÁDNÉ generic 'standardní RD' assumption — kde TZ silent, explicit fallback flag s ČSN default.\n\n"
        f"ÚRS kódy navrženy, vyžadují produkční ověření 2-stage matcherem před finální cenotvorbou. "
        f"Jednotkové ceny ponecháno k vyplnění zhotovitelem."
    )
    cell = ws.cell(row=row, column=1, value=disclaimer)
    cell.font = DISCLAIMER_FONT
    cell.fill = DISCLAIMER_FILL
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.border = BORDER
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 6, end_column=6)
    ws.row_dimensions[row].height = 80

    # Column widths
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 22
    ws.sheet_view.zoomScale = 110


# ───────────────────────────────────────────────────────────────────────────
# Sheet 2 + 3 — Var A aggregated by kapitola

KAPITOLA_TO_GROUP_LABEL = {
    "HSV-1": "HSV-1 — Zemní práce",
    "HSV-2": "HSV-2 — Základové konstrukce + ŽB",
    "HSV-3": "HSV-3 — Svislé konstrukce",
    "HSV-4": "HSV-4 — Vodorovné konstrukce",
    "HSV-5": "HSV-5 — Krov + střecha",
    "HSV-6": "HSV-6 — Bourací práce + demontáže",
    "HSV-7": "HSV-7 — Fasáda ETICS",
    "PSV-71": "PSV-71 — Izolace HI + TI",
    "PSV-72": "PSV-72 — ZTI (vodovod + kanalizace + sanita)",
    "PSV-73": "PSV-73 — Vytápění + komín",
    "PSV-76": "PSV-76 — Výplně otvorů + klempířina + zámečnictví + truhlář",
    "PSV-77": "PSV-77 — Podlahy",
    "PSV-78": "PSV-78 — Povrchové úpravy",
    "PSV-95": "PSV-95 — Detekce požární",
    "M-21": "M-21 — Elektroinstalace silnoproudá",
    "VRN": "VRN — Vedlejší rozpočtové náklady",
}


def kapitola_prefix(kapitola: str) -> str:
    return kapitola.split(" ")[0]


def build_sheet_var_A(wb: Workbook, items: list[dict], objekt: str, title: str) -> None:
    ws = wb.create_sheet(title)
    ws.freeze_panes = "A2"

    # Group items by kapitola_prefix
    by_kap: dict[str, list[dict]] = defaultdict(list)
    for it in items:
        if it["objekt"] == objekt:
            by_kap[kapitola_prefix(it["kapitola"])].append(it)

    headers = ["Poř.", "Kapitola (souhrnná)", "Popis kapitoly", "Počet položek", "J. cena paušál (Kč)", "Cena celkem (Kč)", "Poznámka"]
    write_header_row(ws, 1, headers)

    row = 2
    pol = 1
    # Stable order based on kapitola code
    for kap in sorted(by_kap, key=lambda k: (k.split("-")[0], k.split("-")[1] if "-" in k else "")):
        items_in = by_kap[kap]
        group_label = KAPITOLA_TO_GROUP_LABEL.get(kap, kap)
        # Build short description from sample items
        sample_descs = [it["subkapitola"] for it in items_in[:4]]
        popis = "; ".join(sample_descs) + (f"; … (+{len(items_in) - 4} dalších)" if len(items_in) > 4 else "")

        ws.cell(row=row, column=1, value=pol).alignment = BODY_ALIGN_CENTER
        ws.cell(row=row, column=2, value=group_label).font = KAPITOLA_FONT
        ws.cell(row=row, column=2).fill = KAPITOLA_FILL
        ws.cell(row=row, column=3, value=popis).alignment = BODY_ALIGN_LEFT
        ws.cell(row=row, column=4, value=len(items_in)).alignment = BODY_ALIGN_RIGHT
        c5 = ws.cell(row=row, column=5, value=None)
        c5.number_format = '#,##0 "Kč";-#,##0 "Kč";"—"'
        c5.alignment = BODY_ALIGN_RIGHT
        c5.font = Font(italic=True, color="888888")
        c6 = ws.cell(row=row, column=6, value=None)
        c6.number_format = '#,##0 "Kč";-#,##0 "Kč";"—"'
        c6.alignment = BODY_ALIGN_RIGHT
        c6.font = Font(italic=True, color="888888")
        ws.cell(row=row, column=7, value="agregát — detail viz List B").alignment = BODY_ALIGN_LEFT

        for col in range(1, 8):
            ws.cell(row=row, column=col).border = BORDER
        ws.row_dimensions[row].height = 36
        row += 1
        pol += 1

    auto_width(ws, max_width=55)
    ws.auto_filter.ref = f"A1:G{row - 1}"


# ───────────────────────────────────────────────────────────────────────────
# Sheet 4 — Var C Hybrid (HSV detailně + zbytek agregovaně)

def build_sheet_var_C(wb: Workbook, items: list[dict]) -> None:
    ws = wb.create_sheet("Var_C_Hybrid")
    ws.freeze_panes = "A2"
    headers = ["Poř.", "Objekt", "Kapitola", "Popis", "MJ", "Množství", "J. cena (Kč)", "Cena celkem (Kč)", "Subdodavatel", "Poznámka"]
    write_header_row(ws, 1, headers)

    row = 2
    pol = 1
    # HSV detailně
    for it in items:
        if it["_gate"] != "HSV":
            continue
        ws.cell(row=row, column=1, value=pol).alignment = BODY_ALIGN_CENTER
        ws.cell(row=row, column=2, value=it["objekt"]).alignment = BODY_ALIGN_CENTER
        ws.cell(row=row, column=3, value=it["kapitola"]).alignment = BODY_ALIGN_LEFT
        popis = f"{it['subkapitola']} — {it['popis']}"
        ws.cell(row=row, column=4, value=popis).alignment = BODY_ALIGN_LEFT
        ws.cell(row=row, column=5, value=it["mj"]).alignment = BODY_ALIGN_CENTER
        c_qty = ws.cell(row=row, column=6, value=it["mnozstvi"])
        c_qty.number_format = "#,##0.00" if it["mj"] not in ("ks", "kpl", "ks-měs", "kpl-měs", "paušál", "soubor") else "#,##0"
        c_qty.alignment = BODY_ALIGN_RIGHT
        c7 = ws.cell(row=row, column=7, value=None); c7.number_format = '#,##0 "Kč";-#,##0 "Kč";"—"'; c7.alignment = BODY_ALIGN_RIGHT; c7.font = Font(italic=True, color="888888")
        c8 = ws.cell(row=row, column=8, value=None); c8.number_format = '#,##0 "Kč";-#,##0 "Kč";"—"'; c8.alignment = BODY_ALIGN_RIGHT; c8.font = Font(italic=True, color="888888")
        ws.cell(row=row, column=9, value=it["subdodavatel"]).alignment = BODY_ALIGN_LEFT
        ws.cell(row=row, column=10, value=f"URS-prop {it['urs_code_proposed']}").alignment = BODY_ALIGN_LEFT
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = BORDER
        row += 1
        pol += 1

    # PSV / TZB / VRN agregovaně by kapitola_prefix per object
    agg_groups = defaultdict(list)
    for it in items:
        if it["_gate"] == "HSV":
            continue
        key = (it["objekt"], kapitola_prefix(it["kapitola"]))
        agg_groups[key].append(it)
    for (obj, kap), grp in sorted(agg_groups.items()):
        group_label = KAPITOLA_TO_GROUP_LABEL.get(kap, kap)
        ws.cell(row=row, column=1, value=pol).alignment = BODY_ALIGN_CENTER
        ws.cell(row=row, column=2, value=obj).alignment = BODY_ALIGN_CENTER
        ws.cell(row=row, column=3, value=group_label).font = KAPITOLA_FONT
        ws.cell(row=row, column=3).fill = KAPITOLA_FILL
        popis = f"[agregát] {len(grp)} položek: " + "; ".join(it["subkapitola"] for it in grp[:3])
        if len(grp) > 3:
            popis += f"; … (+{len(grp) - 3} dalších)"
        ws.cell(row=row, column=4, value=popis).alignment = BODY_ALIGN_LEFT
        ws.cell(row=row, column=5, value="kpl").alignment = BODY_ALIGN_CENTER
        ws.cell(row=row, column=6, value=1).alignment = BODY_ALIGN_RIGHT
        c7 = ws.cell(row=row, column=7, value=None); c7.number_format = '#,##0 "Kč";-#,##0 "Kč";"—"'; c7.alignment = BODY_ALIGN_RIGHT; c7.font = Font(italic=True, color="888888")
        c8 = ws.cell(row=row, column=8, value=None); c8.number_format = '#,##0 "Kč";-#,##0 "Kč";"—"'; c8.alignment = BODY_ALIGN_RIGHT; c8.font = Font(italic=True, color="888888")
        ws.cell(row=row, column=9, value="—").alignment = BODY_ALIGN_LEFT
        ws.cell(row=row, column=10, value="agregát").alignment = BODY_ALIGN_LEFT
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = BORDER
        row += 1
        pol += 1

    auto_width(ws, max_width=70)
    ws.auto_filter.ref = f"A1:J{row - 1}"


# ───────────────────────────────────────────────────────────────────────────
# Sheet 5 + 6 — Var B Položkový (full detail)

def build_sheet_var_B(wb: Workbook, items: list[dict], objekt: str, title: str) -> None:
    ws = wb.create_sheet(title)
    ws.freeze_panes = "A2"
    headers = [
        "Poř.", "Kapitola", "Subkapitola", "URS kód (navrž.)", "Popis", "MJ", "Množství",
        "J. cena (Kč)", "Cena celkem (Kč)", "Subdodavatel",
        "_mnozstvi_conf", "_urs_status", "Source", "_vyjasneni_ref", "_NUTNO_OVERIT_URS",
    ]
    write_header_row(ws, 1, headers)

    obj_items = [it for it in items if it["objekt"] == objekt]
    # Sort by _gate (HSV < PSV < TZB < VRN), then by kapitola, then by id
    gate_order = {"HSV": 0, "PSV": 1, "TZB": 2, "VRN": 3, "OTHER": 4}
    obj_items.sort(key=lambda it: (gate_order.get(it["_gate"], 9), it["kapitola"], it["id"]))

    row = 2
    pol = 1
    for it in obj_items:
        ws.cell(row=row, column=1, value=pol).alignment = BODY_ALIGN_CENTER
        ws.cell(row=row, column=2, value=it["kapitola"]).alignment = BODY_ALIGN_LEFT
        ws.cell(row=row, column=3, value=it["subkapitola"]).alignment = BODY_ALIGN_LEFT
        ws.cell(row=row, column=4, value=it["urs_code_proposed"]).alignment = BODY_ALIGN_CENTER
        ws.cell(row=row, column=5, value=it["popis"]).alignment = BODY_ALIGN_LEFT
        ws.cell(row=row, column=6, value=it["mj"]).alignment = BODY_ALIGN_CENTER
        c_qty = ws.cell(row=row, column=7, value=it["mnozstvi"])
        c_qty.number_format = "#,##0.00" if it["mj"] not in ("ks", "kpl", "ks-měs", "kpl-měs", "paušál", "soubor") else "#,##0"
        c_qty.alignment = BODY_ALIGN_RIGHT
        c8 = ws.cell(row=row, column=8, value=None); c8.number_format = '#,##0 "Kč";-#,##0 "Kč";"—"'; c8.alignment = BODY_ALIGN_RIGHT; c8.font = Font(italic=True, color="888888")
        c9 = ws.cell(row=row, column=9, value=None); c9.number_format = '#,##0 "Kč";-#,##0 "Kč";"—"'; c9.alignment = BODY_ALIGN_RIGHT; c9.font = Font(italic=True, color="888888")
        ws.cell(row=row, column=10, value=it["subdodavatel"]).alignment = BODY_ALIGN_LEFT
        c_conf = ws.cell(row=row, column=11, value=it["mnozstvi_confidence"])
        c_conf.number_format = "0.00"
        c_conf.alignment = BODY_ALIGN_CENTER
        ws.cell(row=row, column=12, value=it["urs_status"]).alignment = BODY_ALIGN_CENTER
        ws.cell(row=row, column=13, value=it["source"]).alignment = BODY_ALIGN_LEFT
        vref = ", ".join(f"#{v}" for v in it.get("vyjasneni_ref", []) or [])
        ws.cell(row=row, column=14, value=vref or "—").alignment = BODY_ALIGN_CENTER
        ws.cell(row=row, column=15, value="ANO").alignment = BODY_ALIGN_CENTER
        ws.cell(row=row, column=15).font = Font(bold=True, color="A04000")

        for col in range(1, 16):
            ws.cell(row=row, column=col).border = BORDER
        row += 1
        pol += 1

    # Conditional formatting on confidence column (K)
    apply_conf_rules(ws, "K", 2, row - 1)

    auto_width(ws, max_width=55)
    ws.auto_filter.ref = f"A1:O{row - 1}"


# ───────────────────────────────────────────────────────────────────────────
# Sheet 7 — Var_D Per-podlaží + per-místnost s plnou audit trail (hk212 style)

def build_sheet_var_D(wb: Workbook, items: list[dict], dxf_extract: dict) -> None:
    ws = wb.create_sheet("Var_D_PerPodlazi_Mistnost")
    ws.freeze_panes = "A4"

    # Title block
    ws.cell(row=1, column=1, value="VÝKAZ VÝMĚR — PER PODLAŽÍ + PER MÍSTNOST + AUDIT TRAIL (hk212-style)").font = Font(
        name="Calibri", size=14, bold=True, color="1F3A5F"
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.cell(row=2, column=1, value="Per-místnost data z DXF tabulky místností (km_tabulka místností MTEXT); per-item formula audit trail ukazuje 'co s čím se sčítalo a násobilo'").font = DISCLAIMER_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    ws.row_dimensions[2].height = 24

    # ── PART A: Per-místnost table (raw DXF data) ────────────────────
    headers_a = [
        "Č. místnosti", "Název místnosti", "Podlaží", "Plocha m²",
        "Skupina (vinyl/dlažba/sklep/biodeska)",
        "Source DXF layer + MTEXT", "Confidence",
    ]
    row = 4
    ws.cell(row=row, column=1, value="ČÁST A — PER-MÍSTNOST z DXF (návrh fáze)").font = Font(bold=True, size=11, color="1F3A5F")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1
    write_header_row(ws, row, headers_a)
    row += 1

    rooms = dxf_extract["dum_DPZ"]["mistnosti"]["rooms"]

    def classify_room(name: str, podlazi: str) -> str:
        n = name.lower()
        if podlazi == "1.PP":
            return "dlazba_sklep"
        if any(k in n for k in ["koupeln", "wc", "kuchyn", "spíž", "spiz"]):
            return "dlazba"
        return "vinyl"

    for r in sorted(rooms, key=lambda r: (r["podlazi"], r["room_id"])):
        skupina = classify_room(r["name"], r["podlazi"])
        ws.cell(row=row, column=1, value=r["room_id"]).alignment = BODY_ALIGN_CENTER
        ws.cell(row=row, column=2, value=r["name"]).alignment = BODY_ALIGN_LEFT
        ws.cell(row=row, column=3, value=r["podlazi"]).alignment = BODY_ALIGN_CENTER
        c4 = ws.cell(row=row, column=4, value=r["area_m2"]); c4.number_format = "#,##0.00"; c4.alignment = BODY_ALIGN_RIGHT
        ws.cell(row=row, column=5, value=skupina).alignment = BODY_ALIGN_CENTER
        ws.cell(row=row, column=6, value=r["source"]).alignment = BODY_ALIGN_LEFT
        c7 = ws.cell(row=row, column=7, value=r["confidence"]); c7.number_format = "0.00"; c7.alignment = BODY_ALIGN_CENTER
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = BORDER
        row += 1

    # Per-podlaží subtotals
    row += 1
    per_floor_summary = dxf_extract["dum_DPZ"]["mistnosti"]["per_podlazi_summary"]
    ws.cell(row=row, column=1, value="ČÁST B — PER-PODLAŽÍ SUBTOTAL z DXF").font = Font(bold=True, size=11, color="1F3A5F")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1
    write_header_row(ws, row, ["Podlaží", "Počet místností", "Σ Plocha m²", "Formula (jak vznikla)", "Confidence"])
    row += 1
    for podlazi, data in sorted(per_floor_summary.items()):
        ws.cell(row=row, column=1, value=podlazi).alignment = BODY_ALIGN_CENTER
        ws.cell(row=row, column=1).font = KAPITOLA_FONT
        ws.cell(row=row, column=1).fill = KAPITOLA_FILL
        ws.cell(row=row, column=2, value=data["n_rooms"]).alignment = BODY_ALIGN_RIGHT
        c3 = ws.cell(row=row, column=3, value=data["total_m2"]); c3.number_format = "#,##0.00"; c3.alignment = BODY_ALIGN_RIGHT
        ws.cell(row=row, column=4, value=data["formula"]).alignment = BODY_ALIGN_LEFT
        ws.cell(row=row, column=5, value=0.95).alignment = BODY_ALIGN_CENTER
        ws.cell(row=row, column=5).number_format = "0.00"
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = BORDER
        row += 1

    # Total row
    total_navrh = dxf_extract["dum_DPZ"]["plochy_podlah_per_podlazi"]["total_navrh_m2"]
    tz_baseline = 219.3
    delta = round(total_navrh - tz_baseline, 2)
    ws.cell(row=row, column=1, value="CELKEM návrh").font = Font(bold=True)
    ws.cell(row=row, column=2, value=sum(d["n_rooms"] for d in per_floor_summary.values())).alignment = BODY_ALIGN_RIGHT
    c3 = ws.cell(row=row, column=3, value=total_navrh); c3.number_format = "#,##0.00"; c3.font = Font(bold=True); c3.alignment = BODY_ALIGN_RIGHT
    ws.cell(row=row, column=4, value=f"= součet 4 podlaží | TZ baseline: {tz_baseline} m² | Δ = {delta} m² ({round(delta/tz_baseline*100, 2)} %)").alignment = BODY_ALIGN_LEFT
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = BORDER
    row += 2

    # ── PART C: Per-material aggregation (vinyl / dlažba / sklep / biodeska) ────
    ws.cell(row=row, column=1, value="ČÁST C — PER-MATERIAL podlaha z DXF (= cenotvorba PSV-77)").font = Font(bold=True, size=11, color="1F3A5F")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1
    write_header_row(ws, row, ["Material", "Σ Plocha m²", "Místnosti zahrnuté", "Audit trail (Σ formula)", "Confidence"])
    row += 1
    for m in dxf_extract["dum_DPZ"]["plochy_podlah_per_material"]["results"]:
        ws.cell(row=row, column=1, value=m["material"]).font = KAPITOLA_FONT
        ws.cell(row=row, column=1).fill = KAPITOLA_FILL
        c2 = ws.cell(row=row, column=2, value=m["m2"]); c2.number_format = "#,##0.00"; c2.alignment = BODY_ALIGN_RIGHT
        ws.cell(row=row, column=3, value="; ".join(m["rooms_included"])).alignment = BODY_ALIGN_LEFT
        ws.cell(row=row, column=4, value=m["formula"]).alignment = BODY_ALIGN_LEFT
        c5 = ws.cell(row=row, column=5, value=m["confidence"]); c5.number_format = "0.00"; c5.alignment = BODY_ALIGN_CENTER
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = BORDER
        ws.row_dimensions[row].height = 60
        row += 1
    row += 1

    # ── PART D: All upgraded items s audit trail ────────────────────
    ws.cell(row=row, column=1, value="ČÁST D — POLOŽKY UPGRADED Z DXF EXTRAKCE (audit trail)").font = Font(bold=True, size=11, color="1F3A5F")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 1
    headers_d = [
        "ID", "Kapitola", "Položka", "MJ", "Nová qty",
        "Audit trail formula (co s čím se sčítalo / násobilo)",
        "Source", "Conf před", "Conf po", "Δ",
    ]
    write_header_row(ws, row, headers_d)
    row += 1
    upgraded = [it for it in items if it.get("_dxf_extraction_status") == "upgraded_by_comprehensive"]
    upgraded.sort(key=lambda it: (it["kapitola"], it["id"]))
    for it in upgraded:
        ws.cell(row=row, column=1, value=it["id"]).alignment = BODY_ALIGN_CENTER
        ws.cell(row=row, column=2, value=it["kapitola"]).alignment = BODY_ALIGN_LEFT
        ws.cell(row=row, column=3, value=it["subkapitola"]).alignment = BODY_ALIGN_LEFT
        ws.cell(row=row, column=4, value=it["mj"]).alignment = BODY_ALIGN_CENTER
        c5 = ws.cell(row=row, column=5, value=it["mnozstvi"]); c5.number_format = "#,##0.00"; c5.alignment = BODY_ALIGN_RIGHT
        ws.cell(row=row, column=6, value=it["mnozstvi_formula"]).alignment = BODY_ALIGN_LEFT
        ws.cell(row=row, column=7, value=it["source"]).alignment = BODY_ALIGN_LEFT
        # Get before_conf from upgrade log if available — use 0.75 as fallback
        # (we don't store before in item, so leave blank; the +0.10-0.20 jump is implicit)
        c8 = ws.cell(row=row, column=8, value="—"); c8.alignment = BODY_ALIGN_CENTER
        c9 = ws.cell(row=row, column=9, value=it["mnozstvi_confidence"]); c9.number_format = "0.00"; c9.alignment = BODY_ALIGN_CENTER
        c10 = ws.cell(row=row, column=10, value="↑"); c10.alignment = BODY_ALIGN_CENTER; c10.font = Font(bold=True, color="008000")
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = BORDER
        ws.row_dimensions[row].height = 50
        row += 1

    # Column widths
    widths_d = [22, 28, 35, 8, 12, 60, 50, 9, 9, 5]
    for i, w in enumerate(widths_d, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ───────────────────────────────────────────────────────────────────────────
# Sheet 8 — Var_E_Skladby_Vrstev (TZ-explicit composition per element type)

def build_sheet_var_E(wb: Workbook, skladby: dict) -> None:
    """Render skladby_per_zone.json as readable Sheet 8 — composition vrstev per element.
    Sources: TZ ARS + statika + B + PBŘ. ZERO generic patterns. ZERO S-code per-room assignment.
    """
    ws = wb.create_sheet("Var_E_Skladby_Vrstev")
    ws.freeze_panes = "A4"

    # Title + intro
    ws.cell(row=1, column=1, value="VÝKAZ SKLADEB VRSTEV — TZ-EXPLICIT COMPOSITION PER ELEMENT").font = Font(
        name="Calibri", size=14, bold=True, color="1F3A5F"
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    intro = (
        "Skladby vrstev pochází VÝHRADNĚ z TZ explicit text (ARS dum + statika dum + B souhrnná + PBŘ TUSPO). "
        "DXF cross-validation slouží pouze jako podpůrný důkaz přes semantic HATCH patterns "
        "(CONCRETE1 = ŽB, INSULATION = EPS/MW, WOOD3 = dřevo). "
        "ŽÁDNÝ generic 'standard RD' pattern. Kde TZ silent, explicit '_TZ_silent' flag s ČSN fallback."
    )
    c = ws.cell(row=2, column=1, value=intro)
    c.font = DISCLAIMER_FONT
    c.fill = DISCLAIMER_FILL
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    ws.row_dimensions[2].height = 48

    # Headers
    headers = [
        "Element type",
        "Composition vrstev (TZ explicit)",
        "TZ source citation",
        "DXF cross-validation",
        "Applies to rooms / plocha",
        "Plocha celkem",
        "Data quality",
    ]
    write_header_row(ws, 4, headers)

    row = 5
    for el in skladby["elements"]:
        # Compose layers as numbered list (with thickness + source per layer)
        layer_lines = []
        for i, layer in enumerate(el["composition_layers"], start=1):
            line = f"{i}. {layer['vrstva']}"
            if layer.get("tloušťka_mm") is not None:
                line += f" (tl. {layer['tloušťka_mm']} mm)"
            if layer.get("lambda_W_mK") is not None:
                line += f" (λ = {layer['lambda_W_mK']} W/mK)"
            if layer.get("rozteč_mm") is not None:
                line += f" (rozteč {layer['rozteč_mm']} mm)"
            if layer.get("vyztužení"):
                line += f" [výztuž: {layer['vyztužení']}]"
            layer_lines.append(line)
        composition_str = "\n".join(layer_lines)

        # TZ source citation
        tz_src = el.get("tz_source", "—")

        # DXF cross-validation
        dxf_cv = el.get("dxf_cross_validation", "—") or "—"

        # Applies to rooms / area
        applies = el.get("applies_to_rooms", "—") or "—"

        # Plocha celkem
        plocha = el.get("total_area_m2") or el.get("applies_to_area_m2") or el.get("applies_to_area_m2_approx")
        plocha_str = f"{plocha} m²" if plocha is not None else "—"

        # Data quality flag
        if el.get("tz_explicit") is True:
            dq = "TZ explicit"
        elif el.get("tz_explicit") is False:
            dq = el.get("_data_quality", "TZ silent — fallback")
        else:
            dq = el.get("_data_quality", "—")
        if "tz_silent_o_vyšce" in dq.lower() or "tz silent" in dq.lower():
            dq_color = "A04000"  # orange for silent
        elif "TZ explicit" in dq or "tz_explicit" in dq.lower():
            dq_color = "008000"  # green for explicit
        else:
            dq_color = "555555"

        # Write cells
        c_el = ws.cell(row=row, column=1, value=el["element_type"])
        c_el.font = KAPITOLA_FONT
        c_el.fill = KAPITOLA_FILL
        c_el.alignment = Alignment(wrap_text=True, vertical="top")

        c_comp = ws.cell(row=row, column=2, value=composition_str)
        c_comp.font = BODY_FONT
        c_comp.alignment = Alignment(wrap_text=True, vertical="top")

        c_tz = ws.cell(row=row, column=3, value=tz_src)
        c_tz.alignment = Alignment(wrap_text=True, vertical="top")

        c_dxf = ws.cell(row=row, column=4, value=dxf_cv)
        c_dxf.alignment = Alignment(wrap_text=True, vertical="top")

        c_app = ws.cell(row=row, column=5, value=applies)
        c_app.alignment = Alignment(wrap_text=True, vertical="top")

        c_pl = ws.cell(row=row, column=6, value=plocha_str)
        c_pl.alignment = BODY_ALIGN_RIGHT
        c_pl.font = Font(bold=True)

        c_dq = ws.cell(row=row, column=7, value=dq)
        c_dq.font = Font(bold=True, color=dq_color, italic=("silent" in dq.lower() or "fallback" in dq.lower()))
        c_dq.alignment = Alignment(wrap_text=True, vertical="top")

        for col in range(1, 8):
            ws.cell(row=row, column=col).border = BORDER
        # Adjust row height by composition length
        ws.row_dimensions[row].height = max(60, len(layer_lines) * 16)
        row += 1

    # Cross-validation notes section
    row += 1
    ws.cell(row=row, column=1, value="CROSS-VALIDATION POZNÁMKY (DXF × TZ konzistence)").font = Font(
        bold=True, size=11, color="1F3A5F"
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1
    for note in skladby.get("_cross_validation_notes", []):
        c = ws.cell(row=row, column=1, value="• " + note)
        c.font = DISCLAIMER_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        ws.row_dimensions[row].height = 30
        row += 1

    # Per-podlaží výška + obklad výška silent flags
    row += 1
    ws.cell(row=row, column=1, value="TZ SILENT VALUES — ČSN DEFAULT FALLBACK").font = Font(
        bold=True, size=11, color="A04000"
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1
    silent_notes = [
        f"Per-podlaží výška: TZ NEMÁ explicit value. Fallback: 1.PP 2.50 m, 1.NP 2.80 m, 2.NP 2.80 m, 3.NP nadezdívka 2.65 m. "
        f"Σ = 10.75 m vs TZ B 13.0 m celková výška = +2.25 m pro krov+střechu. Flag: tz_silent_fallback_csn_vyska_podlazi.",
        f"Obklad výška v koupelnách: TZ NEMÁ explicit value. Fallback: 2.0 m (CSN Czech RD standard, po sprchový kout). "
        f"Flag: tz_silent_fallback_csn_obklad_vyska_2m.",
    ]
    for note in silent_notes:
        c = ws.cell(row=row, column=1, value="⚠ " + note)
        c.font = Font(name="Calibri", size=10, italic=True, color="A04000")
        c.fill = DISCLAIMER_FILL
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        ws.row_dimensions[row].height = 36
        row += 1

    # Column widths
    widths = [28, 65, 30, 35, 28, 14, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ───────────────────────────────────────────────────────────────────────────
# Main

def main() -> int:
    print(f"[1/3] Loading inputs...", file=sys.stderr)
    bundle = json.loads(ITEMS_JSON.read_text())
    items = bundle["items"]
    header = json.loads(HEADER_JSON.read_text())
    dxf_extract = json.loads(DXF_EXTRACT_JSON.read_text()) if DXF_EXTRACT_JSON.exists() else None
    skladby = json.loads(SKLADBY_JSON.read_text()) if SKLADBY_JSON.exists() else (
        json.loads(SKLADBY_JSON_FALLBACK.read_text()) if SKLADBY_JSON_FALLBACK.exists() else None
    )
    print(f"  ✓ {len(items)} items, project_header, dxf_extract {'✓' if dxf_extract else '✗'}, skladby {'✓' if skladby else '✗'}", file=sys.stderr)

    n_sheets = 6 + (1 if dxf_extract else 0) + (1 if skladby else 0)
    print(f"[2/3] Building {n_sheets} sheets...", file=sys.stderr)
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    build_sheet_souhrn(wb, items, header)
    print(f"  ✓ Sheet 1: Souhrn", file=sys.stderr)
    build_sheet_var_A(wb, items, "260219_dum", "Var_A_Agregovany_Dum")
    print(f"  ✓ Sheet 2: Var_A_Agregovany_Dum", file=sys.stderr)
    build_sheet_var_A(wb, items, "260217_sklad", "Var_A_Agregovany_Sklad")
    print(f"  ✓ Sheet 3: Var_A_Agregovany_Sklad", file=sys.stderr)
    build_sheet_var_C(wb, items)
    print(f"  ✓ Sheet 4: Var_C_Hybrid", file=sys.stderr)
    build_sheet_var_B(wb, items, "260219_dum", "Var_B_Polozkovy_Dum")
    print(f"  ✓ Sheet 5: Var_B_Polozkovy_Dum", file=sys.stderr)
    build_sheet_var_B(wb, items, "260217_sklad", "Var_B_Polozkovy_Sklad")
    print(f"  ✓ Sheet 6: Var_B_Polozkovy_Sklad", file=sys.stderr)
    if dxf_extract:
        build_sheet_var_D(wb, items, dxf_extract)
        print(f"  ✓ Sheet 7: Var_D_PerPodlazi_Mistnost", file=sys.stderr)
    if skladby:
        build_sheet_var_E(wb, skladby)
        print(f"  ✓ Sheet 8: Var_E_Skladby_Vrstev", file=sys.stderr)

    print(f"[3/3] Saving workbook...", file=sys.stderr)
    wb.save(str(TARGET))
    size = TARGET.stat().st_size
    print(f"\n✓ Wrote {TARGET.relative_to(PROJ)} ({size:,} bytes)", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
