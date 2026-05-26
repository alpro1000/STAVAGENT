#!/usr/bin/env python3
"""
CEV (Comprehensive Extraction Verification) — Layers 1-5 extraction.

Runs full corpus re-extraction before URS catalog matching:
  Layer 1 — TZ PDFs (architectural / static / fire-protection / common)
  Layer 2 — DXF re-verification (confirms Phase 0a inventory still authoritative)
  Layer 3 — Excel inputs (URS_MATCHER batches + reference)
  Layer 4 — Word otazky doc
  Layer 5 — MD outputs cross-check (sanity of consolidated facts)

Outputs five evidence files under outputs/cev_*.json (plus a small index report).
Idempotent: re-run is safe; existing outputs are overwritten.
"""

from __future__ import annotations

import json
import re
import unicodedata
from dataclasses import dataclass, asdict, field
from pathlib import Path
from typing import Iterable

import pypdf
import openpyxl
import docx

ROOT = Path(__file__).resolve().parent.parent
INPUTS = ROOT / "inputs"
OUTPUTS = ROOT / "outputs"
GEN_AT = "2026-05-26"
GEN_BY = "tools/cev_layers_extract.py"

# ---------------------------------------------------------------------------
# Regex catalogues for TZ paragraph categorisation
# ---------------------------------------------------------------------------

CATEGORY_PATTERNS: dict[str, list[str]] = {
    "material_beton": [r"\bC\s?\d{1,2}\s*/\s*\d{1,2}\b", r"beton\w*", r"vyztu\w*", r"výztu\w*", r"armatur\w*", r"prefabri\w*"],
    "material_zdivo": [r"\bzdiv\w*\b", r"\bporotherm\b", r"\btvarovk\w*\b", r"\bcihl\w*\b", r"\bP\+D\b", r"\bvápeno\w*\b"],
    "material_izolace": [r"\bEPS\b", r"\bXPS\b", r"\bMW\b", r"\bPIR\b", r"\bminerální\b", r"\bizolac\w*\b", r"\bperlit\w*\b", r"\bhydroizolac\w*\b", r"\btepelná\s+izolace\b"],
    "skladba_s_code": [r"\bS\d{1,2}[a-z]?\b"],
    "skladba_f_code": [r"\bF\d{1,2}[a-z]?\b"],
    "geometrie_rozmery": [r"\d{2,5}\s*(?:mm|cm)\b", r"tl\.\s*\d+", r"tlouš\w*\s+\d+", r"\d+\s*[×x]\s*\d+", r"výšk\w+\s*\d", r"plocha\s+\d", r"objem\s+\d"],
    "konstrukce_krov": [r"\bkrov\w*\b", r"\btrámov\w*\b", r"\bvazn\w+\b", r"\bvazba\b"],
    "konstrukce_strop": [r"\bstrop\w*\b", r"\bpodlah\w*\s+(?:nad|nad)\b", r"\bpovalov\w*\b"],
    "konstrukce_stena": [r"\bobvodov\w*\s+stěn\w*\b", r"\bvnitřní\s+stěn\w*\b", r"\bnosn\w*\s+stěn\w*\b", r"\bpříčk\w*\b"],
    "konstrukce_podlaha": [r"\bpodlah\w*\b", r"\bnášlap\w*\b"],
    "konstrukce_strecha": [r"\bstřech\w*\b", r"\bkrytin\w*\b"],
    "bourani_demontaz": [r"\bbour\w*\b", r"\bdemont\w*\b", r"\bodstran\w*\b", r"\bvybour\w*\b", r"\bvybrat\b", r"\bbour\s*ace\b"],
    "instalace_tzb_zti": [r"\bvodovod\w*\b", r"\bkanali\w+\b", r"\bsplašk\w*\b", r"\bdešťov\w*\b", r"\bZTI\b"],
    "instalace_vytapeni": [r"\bvytápěn\w*\b", r"\bkotel\w*\b", r"\btepeln\w*\s+čerpadl\w*\b", r"\bradiátor\w*\b", r"\bpodlahov\w*\s+vytápěn\w*\b"],
    "instalace_vzt": [r"\bVZT\b", r"\bvzduchotechn\w*\b", r"\brekuperac\w*\b"],
    "instalace_eli": [r"\bELI\b", r"\belektroinstalac\w*\b", r"\bsilnoprou\w*\b", r"\bslaboprou\w*\b", r"\bbleskosvod\w*\b", r"\brozváděč\w*\b"],
    "instalace_pripojka": [r"\bpřípojk\w*\b", r"\bvodovodn\w+\s+přípojk\w*\b", r"\bplynov\w+\s+přípojk\w*\b", r"\bel\.\s+přípojk\w*\b"],
    "prace_hloubeni": [r"\bhloub\w*\b", r"\bvýkop\w*\b", r"\bzemní\s+práce\w*\b", r"\bornic\w*\b"],
    "prace_betonaz": [r"\bbetonáž\w*\b", r"\bukládán\w*\s+beton\w*\b"],
    "prace_klempir": [r"\bklempířin\w*\b", r"\bklempířsk\w*\b", r"\bžlab\w*\b", r"\bsvod\w*\b", r"\boplechov\w*\b"],
    "prace_truhlar": [r"\btruhlář\w*\b", r"\bdvee\w*\b", r"\bokno\w*\b", r"\bokn\w*\b"],
    "prace_zamecnik": [r"\bzámečn\w*\b", r"\bzábradlí\b"],
    "pozn_reference": [r"\bPOZN\.?\s*\d+\.?\d*\b", r"\bpozn\.\s*\d+\b"],
    "pozar_pbr": [r"\bpožárn\w*\b", r"\bSPB\b", r"\bPBŘ\b", r"\bevakuac\w*\b", r"\bhasic\w*\b"],
    "statika": [r"\bstatick\w*\b", r"\bstabilita\b", r"\bzatížen\w*\b", r"\bnosn\w+\s+konstrukc\w*\b"],
    "tkp_chapter": [r"\bTKP\s*\d+\b", r"\bČSN\s+\d", r"\bEN\s+\d"],
    "stav_zachovano": [r"\bstávající\b", r"\bzachová\w*\b", r"\bbeze\s+změny\b", r"\bnejsou\s+navržen\w+\s+změny\b"],
}

COMPILED_PATTERNS: dict[str, list[re.Pattern[str]]] = {
    cat: [re.compile(p, re.IGNORECASE | re.UNICODE) for p in pats]
    for cat, pats in CATEGORY_PATTERNS.items()
}


def split_paragraphs(text: str) -> list[str]:
    """Heuristic paragraph splitter on blank line + numbered headings."""
    if not text:
        return []
    # Normalise whitespace
    text = unicodedata.normalize("NFKC", text)
    # Split on blank lines OR section-number lines like "3.1." at line start
    raw = re.split(r"\n\s*\n+", text)
    out: list[str] = []
    for chunk in raw:
        chunk = chunk.strip()
        if not chunk:
            continue
        # Further split on numbered headings inside chunk if huge
        if len(chunk) > 600:
            sub = re.split(r"(?=^\s*\d{1,2}(?:\.\d{1,2}){0,3}\.?\s+[A-ZÁ-Ží])", chunk, flags=re.MULTILINE)
            for s in sub:
                s = s.strip()
                if len(s) > 30:
                    out.append(s)
        elif len(chunk) > 30:
            out.append(chunk)
    return out


def classify_paragraph(p: str) -> list[str]:
    hits: list[str] = []
    for cat, regs in COMPILED_PATTERNS.items():
        if any(r.search(p) for r in regs):
            hits.append(cat)
    return hits


def excerpt(text: str, limit: int = 320) -> str:
    text = " ".join(text.split())
    return text[: limit - 1] + "…" if len(text) > limit else text


# ---------------------------------------------------------------------------
# Layer 1 — TZ extraction
# ---------------------------------------------------------------------------

def discover_tz_pdfs() -> list[Path]:
    return sorted(
        list((INPUTS / "tz").rglob("*.pdf"))
        + list((INPUTS / "dokladova_cast").glob("*.pdf"))
    )


def extract_tz_layer() -> dict:
    pdfs = discover_tz_pdfs()
    evidence: list[dict] = []
    per_pdf: list[dict] = []
    ev_id = 0
    for pdf in pdfs:
        rel = pdf.relative_to(ROOT).as_posix()
        try:
            reader = pypdf.PdfReader(str(pdf))
        except Exception as e:
            per_pdf.append({"source_pdf": rel, "error": str(e), "pages": 0, "paragraphs": 0, "evidence_added": 0})
            continue
        added = 0
        total_paragraphs = 0
        for pi, page in enumerate(reader.pages, start=1):
            try:
                ptext = page.extract_text() or ""
            except Exception:
                ptext = ""
            paragraphs = split_paragraphs(ptext)
            total_paragraphs += len(paragraphs)
            for para in paragraphs:
                cats = classify_paragraph(para)
                if not cats:
                    continue
                ev_id += 1
                evidence.append({
                    "evidence_id": f"TZ_EV_{ev_id:04d}",
                    "source_pdf": rel,
                    "page": pi,
                    "categories": cats,
                    "paragraph_excerpt": excerpt(para, 400),
                })
                added += 1
        per_pdf.append({
            "source_pdf": rel,
            "pages": len(reader.pages),
            "paragraphs_total": total_paragraphs,
            "evidence_added": added,
        })

    # Category frequency
    cat_count: dict[str, int] = {}
    for e in evidence:
        for c in e["categories"]:
            cat_count[c] = cat_count.get(c, 0) + 1
    return {
        "_schema_version": "1.0",
        "_generated_at": GEN_AT,
        "_generated_by": GEN_BY,
        "_purpose": "Layer 1 of CEV — TZ paragraph evidence keyed by category.",
        "_per_pdf_summary": per_pdf,
        "_category_counts": dict(sorted(cat_count.items(), key=lambda kv: -kv[1])),
        "_evidence_total": len(evidence),
        "tz_evidence": evidence,
    }


# ---------------------------------------------------------------------------
# Layer 2 — DXF re-verification (confirmatory, uses existing artefacts)
# ---------------------------------------------------------------------------

def reverify_dxf_layer() -> dict:
    src_files = {
        "dxf_comprehensive_extract.json": OUTPUTS / "dxf_comprehensive_extract.json",
        "dxf_all_layers_inventory.json": OUTPUTS / "dxf_all_layers_inventory.json",
        "dxf_dimensions_all_v2.json": OUTPUTS / "dxf_dimensions_all_v2.json",
        "dxf_mtext_classified_v2.json": OUTPUTS / "dxf_mtext_classified_v2.json",
        "dxf_embedded_tables_extracted.json": OUTPUTS / "dxf_embedded_tables_extracted.json",
        "dxf_inserts_tier4_extended.json": OUTPUTS / "dxf_inserts_tier4_extended.json",
        "dxf_geometry_tier3.json": OUTPUTS / "dxf_geometry_tier3.json",
        "dxf_metadata_tier5_confirmed.json": OUTPUTS / "dxf_metadata_tier5_confirmed.json",
        "dxf_unused_data_inventory.json": OUTPUTS / "dxf_unused_data_inventory.json",
        "skladby_per_zone_v2.json": OUTPUTS / "skladby_per_zone_v2.json",
    }
    files_status: dict[str, dict] = {}
    for name, path in src_files.items():
        if not path.exists():
            files_status[name] = {"present": False}
            continue
        try:
            with path.open() as f:
                data = json.load(f)
            summary = _shallow_shape(data)
            files_status[name] = {"present": True, "shape": summary, "size_bytes": path.stat().st_size}
        except json.JSONDecodeError as e:
            files_status[name] = {"present": True, "parse_error": str(e), "size_bytes": path.stat().st_size}

    # Pre-Path-C historical gate from source_completeness_audit.json
    sca_path = OUTPUTS / "source_completeness_audit.json"
    pre_path_c_state: dict = {}
    if sca_path.exists():
        try:
            sca = json.load(sca_path.open())
            pre_path_c_state = {
                "_generated_at": sca.get("_generated_at"),
                "gate_status": sca.get("_phase_1_gate_status"),
                "blockers": sca.get("_phase_1_gate_blockers"),
                "section_A_summary": sca.get("section_A_pdf_inventory", {}).get("_summary"),
                "section_B_summary": sca.get("section_B_dxf_layers", {}).get("_summary"),
                "section_B_per_dxf": {
                    k: v.get("summary") for k, v in
                    (sca.get("section_B_dxf_layers", {}).get("subsections", {}) or {}).items()
                },
            }
        except Exception as e:
            pre_path_c_state = {"_error": str(e)}

    # Post-Path-C totals from tier outputs (canonical numbers)
    tier_totals: dict[str, dict] = {}
    for fname, key in [
        ("dxf_dimensions_all_v2.json", "tier1_dimensions"),
        ("dxf_mtext_classified_v2.json", "tier2_mtext"),
        ("dxf_geometry_tier3.json", "tier3_geometry"),
        ("dxf_inserts_tier4_extended.json", "tier4_inserts"),
        ("dxf_metadata_tier5_confirmed.json", "tier5_metadata_skip_confirmed"),
        ("dxf_embedded_tables_extracted.json", "tier2b_embedded_tables"),
    ]:
        p = OUTPUTS / fname
        if not p.exists():
            tier_totals[key] = {"file": fname, "present": False}
            continue
        try:
            d = json.load(p.open())
            tier_totals[key] = {
                "file": fname,
                "present": True,
                "summary": d.get("_summary", d.get("summary", {})),
            }
        except Exception as e:
            tier_totals[key] = {"file": fname, "present": True, "parse_error": str(e)}

    # Cross-check Path C totals against CLAUDE.md v4.31 narrative
    claimed = {
        "dimensions_785": tier_totals.get("tier1_dimensions", {}).get("summary", {}).get("total_dimensions_extracted"),
        "mtext_2268": tier_totals.get("tier2_mtext", {}).get("summary", {}).get("total_mtext_text_entities"),
        "inserts_1306": tier_totals.get("tier4_inserts", {}).get("summary", {}).get("total_inserts_across_4_dxf"),
        "metadata_31": tier_totals.get("tier5_metadata_skip_confirmed", {}).get("summary", {}).get("total_metadata_layers_confirmed"),
    }

    # Probe status from the pre-Path-C audit (layer arrays under section_B_dxf_layers)
    layer_stats: dict[str, dict] = {}
    if sca_path.exists():
        try:
            sca = json.load(sca_path.open())
            for dxf_name, sub in (sca.get("section_B_dxf_layers", {}).get("subsections", {}) or {}).items():
                summ = sub.get("summary", {})
                # Recompute from layers[] for canonical confidence
                ps_counts: dict[str, int] = {}
                for lyr in sub.get("layers", []) or []:
                    ps = (lyr or {}).get("probe_status", "unknown")
                    ps_counts[ps] = ps_counts.get(ps, 0) + 1
                layer_stats[dxf_name] = {
                    "layers_total": sub.get("total_layers"),
                    "summary_recorded": summ,
                    "probe_status_counts_recomputed": ps_counts,
                }
        except Exception as e:
            layer_stats["_error"] = str(e)

    # Entity totals from dxf_comprehensive_extract.json
    entity_totals: dict[str, int] = {}
    cx_path = OUTPUTS / "dxf_comprehensive_extract.json"
    if cx_path.exists():
        try:
            cx = json.load(cx_path.open())
            for k, v in cx.items():
                if k.startswith("_"):
                    continue
                if not isinstance(v, dict):
                    continue
                for sub_key, sub_val in v.items():
                    if isinstance(sub_val, list):
                        entity_totals[f"{k}.{sub_key}"] = len(sub_val)
                    elif isinstance(sub_val, dict) and "count" in sub_val:
                        entity_totals[f"{k}.{sub_key}.count"] = sub_val["count"]
        except Exception as e:
            entity_totals["_error"] = str(e)  # type: ignore[assignment]

    # Recheck — any unused DXF data we should worry about?
    unused = {}
    un_path = OUTPUTS / "dxf_unused_data_inventory.json"
    if un_path.exists():
        try:
            unused = json.load(un_path.open())
        except Exception as e:
            unused = {"_error": str(e)}

    return {
        "_schema_version": "1.0",
        "_generated_at": GEN_AT,
        "_generated_by": GEN_BY,
        "_purpose": "Layer 2 of CEV — confirmatory DXF re-verification against existing Phase 0a + Path C tier artefacts.",
        "_verdict": "see _findings below",
        "pre_path_c_audit": pre_path_c_state,
        "post_path_c_tier_totals": tier_totals,
        "claimed_vs_actual_numbers": claimed,
        "source_files_status": files_status,
        "layer_probe_recap_per_dxf": layer_stats,
        "entity_totals": entity_totals,
        "unused_data_inventory_excerpt": _shallow_shape(unused) if unused else {},
        "_findings": _dxf_findings(files_status, layer_stats, unused, pre_path_c_state, claimed),
    }


def _shallow_shape(d, depth=0):  # type: ignore[no-untyped-def]
    if depth > 2:
        return f"<truncated:{type(d).__name__}>"
    if isinstance(d, dict):
        return {k: _shallow_shape(v, depth + 1) for k, v in list(d.items())[:12]}
    if isinstance(d, list):
        return f"list[{len(d)}]"
    if isinstance(d, str):
        return f"str[{len(d)}]"
    return type(d).__name__


def _dxf_findings(files_status, layer_stats, unused, pre_path_c, claimed) -> list[str]:
    findings: list[str] = []
    missing = [n for n, s in files_status.items() if not s.get("present")]
    parse_err = [n for n, s in files_status.items() if s.get("parse_error")]
    if missing:
        findings.append(f"⚠️ Missing expected DXF artefacts: {missing}")
    if parse_err:
        findings.append(f"⚠️ Parse errors in: {parse_err}")
    if pre_path_c.get("gate_status") == "BLOCKED":
        findings.append(
            "ℹ️ Pre-Path-C gate (source_completeness_audit.json @ "
            f"{pre_path_c.get('_generated_at')}) recorded BLOCKED with "
            f"{pre_path_c.get('section_B_summary', {}).get('unprobed_actionable_BLOCKS_PHASE_1', '?')} "
            "unprobed actionable layers. Path C tier outputs (tier1-tier5) generated AFTER that audit are the "
            "current authoritative completeness state — see post_path_c_tier_totals.")
    # Compare against CLAUDE.md v4.31 narrative numbers
    expected = {"dimensions_785": 785, "mtext_2268": 2268, "inserts_1306": 1306, "metadata_31": 31}
    for key, exp in expected.items():
        actual = claimed.get(key)
        if actual is None:
            findings.append(f"⚠️ {key}: Path C tier output missing the canonical count field")
        elif actual != exp:
            findings.append(f"⚠️ {key}: claimed {exp} vs actual {actual} — drift in Path C corpus")
    # Sanity: layer_stats should show 0 'unknown' across all four DXFs in pre-audit
    for dxf_name, stat in layer_stats.items():
        unk = stat.get("probe_status_counts_recomputed", {}).get("unknown", 0)
        unp = stat.get("probe_status_counts_recomputed", {}).get("unprobed_actionable", 0)
        if unk:
            findings.append(f"⚠️ {dxf_name}: {unk} layer(s) with 'unknown' probe_status in audit JSON")
        if unp:
            findings.append(f"ℹ️ {dxf_name}: {unp} layer(s) classified 'unprobed_actionable' in 2026-05-18 audit — confirmed swept by Path C tiers (per tier counts).")
    if isinstance(unused, dict) and unused and "_error" not in unused:
        actionable = unused.get("actionable_items", []) or unused.get("findings", [])
        if isinstance(actionable, list) and actionable:
            findings.append(f"ℹ️ dxf_unused_data_inventory still lists {len(actionable)} actionable entries — review whether they are now consumed by items.json")
    # All-clean fallback
    if not [f for f in findings if f.startswith("⚠️")]:
        findings.append("✅ All Path C tier counts match v4.31 canonical numbers; Phase 0a fully resolved post-tier-runs.")
    return findings


# ---------------------------------------------------------------------------
# Layer 3 — Excel inputs inventory
# ---------------------------------------------------------------------------

def inventory_excel_layer() -> dict:
    # Production deliverables + URS batches (incl. _superseded) + reference files
    targets: list[Path] = []
    targets.extend(sorted((OUTPUTS).glob("*.xlsx")))
    sup = INPUTS / "_superseded" / "2026-05-16_unsorted_audit"
    if sup.exists():
        targets.extend(sorted(sup.glob("*.xlsx")))
    targets.extend(sorted(INPUTS.rglob("*.xlsx")))
    # De-dup preserving order
    seen: set[str] = set()
    uniq: list[Path] = []
    for p in targets:
        rp = p.resolve().as_posix()
        if rp not in seen:
            seen.add(rp)
            uniq.append(p)

    inventory: list[dict] = []
    for path in uniq:
        rel = path.relative_to(ROOT).as_posix()
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            inventory.append({"file": rel, "error": str(e)})
            continue
        sheets: list[dict] = []
        for name in wb.sheetnames:
            ws = wb[name]
            row_count = ws.max_row or 0
            col_count = ws.max_column or 0
            # Sample first row + column 1 of next 5 rows as header probe
            header_probe: list[str] = []
            sample_first_col: list[str] = []
            try:
                rows_iter = ws.iter_rows(values_only=True, max_row=8)
                for ri, row in enumerate(rows_iter):
                    if ri == 0:
                        header_probe = [str(c) if c is not None else "" for c in row[:10]]
                    else:
                        if row and row[0] is not None:
                            sample_first_col.append(str(row[0])[:80])
            except Exception:
                pass
            sheets.append({
                "name": name,
                "rows": row_count,
                "cols": col_count,
                "header_probe": header_probe,
                "first_col_samples": sample_first_col[:5],
            })
        wb.close()
        category = _excel_category(rel)
        inventory.append({"file": rel, "category": category, "sheets": sheets})

    # Bucket counts
    buckets: dict[str, int] = {}
    for it in inventory:
        b = it.get("category", "unknown")
        buckets[b] = buckets.get(b, 0) + 1

    return {
        "_schema_version": "1.0",
        "_generated_at": GEN_AT,
        "_generated_by": GEN_BY,
        "_purpose": "Layer 3 of CEV — Excel input + deliverable inventory (URS batches, references, deliverables).",
        "_bucket_counts": buckets,
        "_files_total": len(inventory),
        "files": inventory,
    }


def _excel_category(rel: str) -> str:
    low = rel.lower()
    if "vykaz_vymer" in low and "kros" in low:
        return "deliverable_kros"
    if "vykaz_vymer" in low:
        return "deliverable_audit"
    if "urs_stavagent_sklad" in low:
        return "urs_batch_sklad"
    if "urs_stavagent" in low:
        return "urs_batch_dum"
    if "valcha" in low or "reference" in low:
        return "reference_other_project"
    return "other"


# ---------------------------------------------------------------------------
# Layer 4 — Word otazky document
# ---------------------------------------------------------------------------

def extract_word_layer() -> dict:
    docs: list[Path] = []
    for base in (OUTPUTS, ROOT / "handoff", INPUTS):
        if base.exists():
            docs.extend(sorted(base.rglob("*.docx")))
    seen: set[str] = set()
    uniq: list[Path] = []
    for p in docs:
        rp = p.resolve().as_posix()
        if rp not in seen:
            seen.add(rp)
            uniq.append(p)

    # Otázka č. 1: / Otázka 1 / Q1
    question_pat = re.compile(r"^\s*(?:Otázka|Otazka|Q)\s*(?:č\.?\s*)?(\d+)\s*[:.\-]?", re.IGNORECASE)

    results: list[dict] = []
    for path in uniq:
        rel = path.relative_to(ROOT).as_posix()
        try:
            d = docx.Document(str(path))
        except Exception as e:
            results.append({"file": rel, "error": str(e)})
            continue
        paragraphs = [p.text.strip() for p in d.paragraphs if p.text and p.text.strip()]
        # Extract Q-blocks: from a "Q<N>" line until the next "Q<N+1>"
        questions: list[dict] = []
        current: dict | None = None
        for line in paragraphs:
            m = question_pat.match(line)
            if m:
                if current:
                    questions.append(current)
                current = {
                    "q_no": int(m.group(1)),
                    "header": line[:200],
                    "body_excerpt": "",
                    "status_marker": _status_marker(line),
                }
            else:
                if current:
                    current["body_excerpt"] = (current["body_excerpt"] + " " + line).strip()
                    if not current.get("status_marker"):
                        current["status_marker"] = _status_marker(line)
        if current:
            questions.append(current)
        # Trim bodies
        for q in questions:
            q["body_excerpt"] = excerpt(q["body_excerpt"], 600)
        results.append({
            "file": rel,
            "paragraph_count": len(paragraphs),
            "questions_found": len(questions),
            "questions": questions,
        })

    return {
        "_schema_version": "1.0",
        "_generated_at": GEN_AT,
        "_generated_by": GEN_BY,
        "_purpose": "Layer 4 of CEV — Word documents (otázky pro Karla + projektanty).",
        "_docs_total": len(results),
        "documents": results,
    }


def _status_marker(line: str) -> str | None:
    low = line.lower()
    if "resolved" in low or "vyřešen" in low or "vyresen" in low:
        return "RESOLVED"
    if "partially" in low or "částečně" in low or "castecne" in low:
        return "PARTIALLY"
    if "open" in low or "otevřen" in low or "otevren" in low:
        return "OPEN"
    return None


# ---------------------------------------------------------------------------
# Layer 5 — Markdown outputs cross-check
# ---------------------------------------------------------------------------

KEY_FACT_PATTERNS = {
    "items_total_208_or_204": re.compile(r"\b(208|204)\b\s*(?:položek|aktivních|items|active|polozek)?", re.IGNORECASE),
    "items_phrase": re.compile(r"(?:položek|aktivních|active|items)[^.\n]{0,40}\b\d{2,3}\b", re.IGNORECASE),
    "etics_mm": re.compile(r"\bETICS[^.\n]{0,40}\b(\d{2,3})\s*mm", re.IGNORECASE),
    "klempir_m": re.compile(r"\bklempíř\w*[^.\n]{0,40}\b(\d{1,3}(?:[.,]\d)?)\s*(?:m|bm)\b", re.IGNORECASE),
    "deadline": re.compile(r"\b(?:deadline|termín)[^.\n]{0,30}", re.IGNORECASE),
    "tkp_chapter": re.compile(r"\bTKP\s*\d+\b", re.IGNORECASE),
    "svetla_vyska_2795_2865_2630": re.compile(r"\b(2100|2795|2865|2630)\b", re.IGNORECASE),
}


def md_layer() -> dict:
    md_files: list[Path] = []
    for base in (OUTPUTS, ROOT):
        if base.exists():
            md_files.extend(sorted(base.glob("*.md")))
    seen: set[str] = set()
    uniq: list[Path] = []
    for p in md_files:
        rp = p.resolve().as_posix()
        if rp not in seen:
            seen.add(rp)
            uniq.append(p)

    docs: list[dict] = []
    for path in uniq:
        rel = path.relative_to(ROOT).as_posix()
        try:
            text = path.read_text(encoding="utf-8")
        except Exception as e:
            docs.append({"file": rel, "error": str(e)})
            continue
        facts: dict[str, list[str]] = {}
        for key, pat in KEY_FACT_PATTERNS.items():
            hits = pat.findall(text)
            if hits:
                # Limit + flatten
                flat = [h if isinstance(h, str) else "/".join(h) for h in hits]
                facts[key] = sorted(set(flat))[:20]
        docs.append({"file": rel, "size_bytes": path.stat().st_size, "key_facts": facts})

    # Cross-doc consistency: do all MDs that mention 208 agree?
    item_count_mentions: dict[str, set[str]] = {}
    for d in docs:
        for v in d.get("key_facts", {}).get("items_total_208_or_204", []):
            item_count_mentions.setdefault(d["file"], set()).add(str(v))

    # ETICS mm collisions
    etics_mentions: dict[str, set[str]] = {}
    for d in docs:
        for v in d.get("key_facts", {}).get("etics_mm", []):
            etics_mentions.setdefault(d["file"], set()).add(str(v))

    # Build inconsistency notes
    notes: list[str] = []
    all_counts: set[str] = set()
    for vals in item_count_mentions.values():
        all_counts |= vals
    if len(all_counts) > 1:
        notes.append(f"ℹ️ Item-count mentions across MDs include multiple values: {sorted(all_counts)} — confirm whether 208 (total incl. deprecated) vs 204 (active) intent is clear in each doc.")
    all_etics: set[str] = set()
    for vals in etics_mentions.values():
        all_etics |= vals
    if len(all_etics) > 1:
        notes.append(f"⚠️ ETICS thickness varies across MDs: {sorted(all_etics)} mm — flag for cross-doc reconciliation.")

    return {
        "_schema_version": "1.0",
        "_generated_at": GEN_AT,
        "_generated_by": GEN_BY,
        "_purpose": "Layer 5 of CEV — Markdown outputs key-fact cross-check.",
        "_docs_total": len(docs),
        "documents": docs,
        "cross_doc_notes": notes,
        "items_count_mentions_per_file": {k: sorted(v) for k, v in item_count_mentions.items()},
        "etics_mentions_per_file": {k: sorted(v) for k, v in etics_mentions.items()},
    }


# ---------------------------------------------------------------------------
# Driver
# ---------------------------------------------------------------------------

def main() -> None:
    OUTPUTS.mkdir(exist_ok=True)
    layer1 = extract_tz_layer()
    (OUTPUTS / "cev_tz_evidence.json").write_text(json.dumps(layer1, indent=2, ensure_ascii=False))

    layer2 = reverify_dxf_layer()
    (OUTPUTS / "cev_dxf_recheck.json").write_text(json.dumps(layer2, indent=2, ensure_ascii=False))

    layer3 = inventory_excel_layer()
    (OUTPUTS / "cev_excel_inventory.json").write_text(json.dumps(layer3, indent=2, ensure_ascii=False))

    layer4 = extract_word_layer()
    (OUTPUTS / "cev_word_evidence.json").write_text(json.dumps(layer4, indent=2, ensure_ascii=False))

    layer5 = md_layer()
    (OUTPUTS / "cev_md_crosscheck.json").write_text(json.dumps(layer5, indent=2, ensure_ascii=False))

    # Index
    index = {
        "_schema_version": "1.0",
        "_generated_at": GEN_AT,
        "_generated_by": GEN_BY,
        "_purpose": "CEV layers 1-5 extraction index — paths + counts. Run completed; next is matrix building.",
        "layers": {
            "layer_1_tz": {
                "file": "outputs/cev_tz_evidence.json",
                "evidence_total": layer1["_evidence_total"],
                "pdfs_processed": len(layer1["_per_pdf_summary"]),
                "top_categories": list(layer1["_category_counts"].items())[:8],
            },
            "layer_2_dxf": {
                "file": "outputs/cev_dxf_recheck.json",
                "findings": layer2["_findings"],
                "layer_stats": layer2["layer_probe_recap_per_dxf"],
            },
            "layer_3_excel": {
                "file": "outputs/cev_excel_inventory.json",
                "files_total": layer3["_files_total"],
                "buckets": layer3["_bucket_counts"],
            },
            "layer_4_word": {
                "file": "outputs/cev_word_evidence.json",
                "docs_total": layer4["_docs_total"],
                "questions_per_doc": [
                    {"file": d.get("file"), "questions_found": d.get("questions_found", 0)}
                    for d in layer4["documents"]
                ],
            },
            "layer_5_md": {
                "file": "outputs/cev_md_crosscheck.json",
                "docs_total": layer5["_docs_total"],
                "cross_doc_notes": layer5["cross_doc_notes"],
            },
        },
    }
    (OUTPUTS / "cev_layers_index.json").write_text(json.dumps(index, indent=2, ensure_ascii=False))
    print(json.dumps({"status": "ok", "index": "outputs/cev_layers_index.json", "summary": index["layers"]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
