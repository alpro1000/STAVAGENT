#!/usr/bin/env python3
"""
VÝMĚRY-First souhrn (Pattern 45) — structured measurement register → XLSX deliverable.

Source: inputs/meta/vymery_souhrn.json (editable measurement base).
Output: outputs/VYMERY_SOUHRN_<date>.xlsx — one row per jednotka (místnost / prvek).
Every qty in work items should trace to a row here.
"""
from __future__ import annotations
import json
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "inputs" / "meta" / "vymery_souhrn.json"
OUT = ROOT / "outputs"

COLS = [("jednotka", "Jednotka", 30), ("objekt", "Objekt / Podlaží", 22),
        ("plocha_m2", "Plocha m²", 11), ("svetla_vyska_m", "Světlá výška m", 13),
        ("obvod_m", "Obvod m", 10), ("objem_m3", "Objem m³", 10),
        ("skladba", "Skladba", 20), ("zdroj", "Zdroj", 30), ("status", "Status", 12)]

STATUS_FILL = {"measured": "D5F4E6", "derived": "DDEBF7", "estimate": "FFF3CD", "blank": "F2F2F2"}


def main() -> None:
    data = json.loads(SRC.read_text(encoding="utf-8"))
    rows = data["rows"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VYMERY"
    thin = Side(style="thin", color="C0C0C0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.cell(1, 1, "VÝMĚRY-First — měřená báze (Pattern 45) · " + data.get("_generated_for", "")).font = Font(bold=True, size=11)
    hdr_fill = PatternFill("solid", fgColor="2F3640")
    for c, (_, label, w) in enumerate(COLS, 1):
        cell = ws.cell(2, c, label)
        cell.fill = hdr_fill
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    for r, row in enumerate(rows, start=3):
        st = row.get("status", "blank")
        for c, (key, _, _) in enumerate(COLS, 1):
            v = row.get(key)
            cell = ws.cell(r, c, "—" if v is None else v)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (1, 7, 8)))
            if c == 9:
                cell.fill = PatternFill("solid", fgColor=STATUS_FILL.get(st, "FFFFFF"))
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{openpyxl.utils.get_column_letter(len(COLS))}{len(rows)+2}"

    out = OUT / "VYMERY_SOUHRN.xlsx"
    wb.save(out)
    from collections import Counter
    st = Counter(r.get("status") for r in rows)
    print(f"OK — {out.name}: {len(rows)} jednotek ({dict(st)}).")


if __name__ == "__main__":
    main()
