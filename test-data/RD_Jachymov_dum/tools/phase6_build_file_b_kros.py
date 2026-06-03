#!/usr/bin/env python3
"""
Phase 6 — File B KROS production deliverable per Pattern 32.

Reads:  outputs/items_rd_jachymov_complete.json  (214 items canonical, FROZEN — busy-einstein base; PM01 přesun hmot + PM02 lešení included via status_flag filter)
Writes: outputs/Vykaz_vymer_RD_Jachymov_KROS_format_<date>_v3_final.xlsx

5 sheets:
  1. Rekapitulace stavby     — investor header + per-SO summary
  2. SO_260219_Dum           — krycí list + rekapitulace členění + soupis prací (180 items)
  3. SO_260217_Sklad         — krycí list + rekapitulace + soupis prací (28 items)
  4. Stav polozek            — URS verification audit (Karel-facing)
  5. URS vyber audit         — full transparency (per-item candidates, provenance trail)

Pattern compliance:
  - 15: items.json FROZEN-read-only; File B is downstream production
  - 25/26: status icons (✓ ⚠ ❌ ?) per Pattern 26 honest fallback hierarchy;
          blank Kód when code unknown (NO 999999999, NO TBD)
  - 27: cross_verification_evidence_url surfaced in Pozn. tooltip
  - 28: compound key (id, kapitola) for ordering stability
  - 32: TWO-FILE — File B has NO provenance columns (_source, _data_quality,
        _audit_gap_fixed, vyjasneni_ref, mnozstvi_formula as own column).
        mnozstvi_formula surfaces only as VV-row beneath each K-row (KROS standard).
        File A (Vykaz_vymer_VSE_VARIANTY) keeps the full provenance.

Excludes 4 items with status_flag='deprecated_audit_v2' (kept in File A for audit).
208 active = 180 dum + 28 sklad.
"""
from __future__ import annotations

import json
from datetime import date
from pathlib import Path
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
ITEMS_PATH = ROOT / "outputs" / "items_rd_jachymov_complete.json"
TODAY = date.today().isoformat()
OUT_PATH = ROOT / "outputs" / "Vykaz_vymer_RD_Jachymov_KROS_format_v3_final.xlsx"
LOG_PATH = ROOT / "outputs" / "phase6_file_b_validation_log.json"

# === Stavba header constants ===
STAVBA = "RD Jáchymov Fibichova 733 — rekonstrukce + nástavba 3.NP + sklad + parking"
MISTO = "Fibichova 733, 362 51 Jáchymov, Karlovarský kraj"
DATUM = "27.05.2026"
ZADAVATEL = "Mgr. Jindřich Volný"
ZHOTOVITEL = "Karel Šmíd (smid.karell@gmail.com, +420 608 930 914)"
PROJEKTANT = (
    "SMASH — Marek Smolka (ČKA 05394, ARS); "
    "TeAnau — Tvardík (ČKAIT 0012219, stavební část); "
    "TUSPO — Vodička (PBŘ)"
)
ZPRACOVATEL = "Berger Bohemia — Alexander (přípravář, STAVAGENT v4.33)"

# === SO definitions ===
OBJEKTY = [
    {
        "kod": "260219",
        "objekt": "260219_dum",
        "nazev": "Dům — rekonstrukce + nástavba 3.NP",
        "sheet": "SO_260219_Dum",
    },
    {
        "kod": "260217",
        "objekt": "260217_sklad",
        "nazev": "Sklad + parking",
        "sheet": "SO_260217_Sklad",
    },
]

# === Kapitola order (matches items.json insertion order) ===
KAPITOLA_ORDER = [
    "HSV-1 Zemní práce",
    "HSV-2 Základové a ŽB",
    "HSV-3 Svislé konstrukce",
    "HSV-4 Vodorovné",
    "HSV-5 Krov + střecha",
    "HSV-5 Komunikace + schodiště",
    "HSV-6 Bourací práce",
    "HSV-7 Fasáda ETICS",
    "M-21 ELI silnoproud",
    "PSV-71 Izolace HI",
    "PSV-71 Izolace TI",
    "PSV-72 ZTI",
    "PSV-73 Vytápění",
    "PSV-76 Klempíř",
    "PSV-76 Truhlář",
    "PSV-76 Výplně otvorů",
    "PSV-76 Zámečnictví",
    "PSV-77 Podlahy",
    "PSV-78 Povrchové úpravy",
    "PSV-95 Detekce požární",
    "VRN — Společné",
    "VRN — Zařízení staveniště",
    "VRN — Přesun hmot",
    "VRN — Lešení",
    "VRN — Doprava + odpad",
    "VRN — Geodet",
    "VRN — Průzkumy",
    "VRN — BOZP",
    "VRN — Dokumentace",
    "VRN — Kolaudace",
    "VRN — Pojištění + zábory",
    "VRN — Revize",
]

# === Status → Pozn. icon + fill ===
PATTERN_26_LEGEND = {
    "VERIFIED":                     {"icon": "✓",  "label": "ověřeno cs-urs.cz",                "fill": "C6EFCE"},
    "CROSS_DISCIPLINE_OK":          {"icon": "✓",  "label": "cross-discipline OK",              "fill": "C6EFCE"},
    "FAMILY_VERIFIED":              {"icon": "⚠",  "label": "family OK, leaf KROS lookup",     "fill": "FFEB9C"},
    "FAMILY_VERIFIED_CHAPTER_MATCH":{"icon": "⚠",  "label": "chapter OK, leaf KROS lookup",   "fill": "FFEB9C"},
    "WRONG_LEAF":                   {"icon": "❌", "label": "WRONG family — KROS leaf TBD",   "fill": "FFC7CE"},
    "MANUAL_LOOKUP":                {"icon": "❌", "label": "MANUAL — KROS catalog lookup",   "fill": "FFC7CE"},
    "NONE_NEEDS_LOOKUP":            {"icon": "?",  "label": "needs production lookup",         "fill": "D9D9D9"},
}

# === Styling ===
THIN = Side(border_style="thin", color="808080")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FONT_TITLE = Font(name="Calibri", size=14, bold=True)
FONT_HDR = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
FONT_DROW = Font(name="Calibri", size=10, bold=True)
FONT_BODY = Font(name="Calibri", size=10)
FONT_VV = Font(name="Calibri", size=9, italic=True, color="606060")
FONT_LEGEND = Font(name="Calibri", size=9, italic=True)

FILL_HDR = PatternFill("solid", fgColor="305496")
FILL_DROW = PatternFill("solid", fgColor="D9E1F2")
FILL_REKAP = PatternFill("solid", fgColor="FCE4D6")
FILL_TOTAL = PatternFill("solid", fgColor="FFD966")

ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def classify_status(it: dict) -> tuple[str, str, str, str]:
    """Returns (icon, label, fill_color, verbose_pozn) for an item."""
    cvs = it.get("cross_verification_status")
    code = it.get("urs_code_proposed")
    family = it.get("urs_code_family_6digit")
    urs_status = it.get("urs_status", "")
    hint = it.get("correct_code_hint") or ""
    evidence_url = it.get("cross_verification_evidence_url") or ""

    # Phase 5B-flagged WRONG_LEAF with blank code (HSV2.003 + HSV2.008 from A1)
    if cvs == "WRONG_LEAF" and not code:
        leg = PATTERN_26_LEGEND["WRONG_LEAF"]
        pozn = f"{leg['icon']} family {family or '?'}: {leg['label']}"
        if hint:
            pozn += f" — {hint[:120]}"
        return leg["icon"], leg["label"], leg["fill"], pozn

    # Pre-existing wrong_leaf_disambiguation_needed (5 items with blank code, no Phase 5B verdict)
    if urs_status == "wrong_leaf_disambiguation_needed" and not code:
        leg = PATTERN_26_LEGEND["MANUAL_LOOKUP"]
        pozn = f"{leg['icon']} {leg['label']}"
        return leg["icon"], leg["label"], leg["fill"], pozn

    # Phase 5B verdicts (VERIFIED, FAMILY_VERIFIED, CROSS_DISCIPLINE_OK, FAMILY_VERIFIED_CHAPTER_MATCH)
    if cvs in PATTERN_26_LEGEND:
        leg = PATTERN_26_LEGEND[cvs]
        pozn = f"{leg['icon']} {leg['label']}"
        if cvs.startswith("FAMILY") and family:
            pozn = f"{leg['icon']} family {family} OK — leaf KROS lookup"
        if evidence_url:
            pozn += f" [evidence: {evidence_url[:60]}]"
        return leg["icon"], leg["label"], leg["fill"], pozn

    # Default — not Phase 5B-touched, has code
    if code:
        leg = PATTERN_26_LEGEND["NONE_NEEDS_LOOKUP"]
        return leg["icon"], leg["label"], leg["fill"], f"{leg['icon']} {leg['label']}"

    # No code, no Phase 5B verdict — MANUAL
    leg = PATTERN_26_LEGEND["MANUAL_LOOKUP"]
    return leg["icon"], leg["label"], leg["fill"], f"{leg['icon']} {leg['label']}"


def _set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# =====================================================================
# Sheet 1 — Rekapitulace stavby
# =====================================================================
def build_sheet_rekapitulace(wb: Workbook, per_so_total_ranges: dict):
    ws = wb.create_sheet("Rekapitulace stavby", 0)

    # Header block
    ws["B2"] = "REKAPITULACE STAVBY"
    ws["B2"].font = FONT_TITLE

    rows = [
        ("Stavba:",        STAVBA),
        ("Místo:",         MISTO),
        ("Datum:",         DATUM),
        ("Zadavatel:",     ZADAVATEL),
        ("Zhotovitel:",    ZHOTOVITEL),
        ("Projektant:",    PROJEKTANT),
        ("Zpracovatel:",   ZPRACOVATEL),
    ]
    for i, (lbl, val) in enumerate(rows, start=4):
        ws.cell(i, 2, lbl).font = FONT_DROW
        ws.cell(i, 3, val).font = FONT_BODY
        ws.cell(i, 3).alignment = ALIGN_LEFT
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=8)

    # Per-SO summary table
    r = 13
    headers = ["Kód SO", "Název objektu", "Cena bez DPH", "DPH 21 %", "Cena s DPH"]
    for c, h in enumerate(headers, start=2):
        cell = ws.cell(r, c, h)
        cell.font = FONT_HDR
        cell.fill = FILL_HDR
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL
    r += 1

    so_data_rows = []
    for obj in OBJEKTY:
        sheet_name = obj["sheet"]
        # Pointer to Cena celkem column on per-SO sheet (sum of K-row cena)
        rng = per_so_total_ranges.get(sheet_name)
        if rng:
            sum_formula = f"=SUM('{sheet_name}'!{rng})"
        else:
            sum_formula = 0
        ws.cell(r, 2, obj["kod"]).font = FONT_BODY
        ws.cell(r, 3, obj["nazev"]).font = FONT_BODY
        ws.cell(r, 4, sum_formula).font = FONT_BODY
        ws.cell(r, 4).number_format = '#,##0.00 "Kč"'
        ws.cell(r, 5, f"=D{r}*0.21").number_format = '#,##0.00 "Kč"'
        ws.cell(r, 6, f"=D{r}+E{r}").number_format = '#,##0.00 "Kč"'
        for c in range(2, 7):
            ws.cell(r, c).border = BORDER_ALL
        so_data_rows.append(r)
        r += 1

    # Total row
    ws.cell(r, 2, "").border = BORDER_ALL
    ws.cell(r, 3, "Celkem stavba").font = FONT_DROW
    ws.cell(r, 3).fill = FILL_TOTAL
    ws.cell(r, 4, f"=SUM(D{so_data_rows[0]}:D{so_data_rows[-1]})")
    ws.cell(r, 5, f"=SUM(E{so_data_rows[0]}:E{so_data_rows[-1]})")
    ws.cell(r, 6, f"=SUM(F{so_data_rows[0]}:F{so_data_rows[-1]})")
    for c in range(2, 7):
        ws.cell(r, c).border = BORDER_ALL
        ws.cell(r, c).fill = FILL_TOTAL
        ws.cell(r, c).font = FONT_DROW
        if c >= 4:
            ws.cell(r, c).number_format = '#,##0.00 "Kč"'
    r += 2

    # Pattern 26 legend
    ws.cell(r, 2, "Legenda stavu URS kódů (sloupec Pozn. v listech soupisu):").font = FONT_DROW
    r += 1
    for key, leg in PATTERN_26_LEGEND.items():
        ws.cell(r, 2, leg["icon"]).alignment = ALIGN_CENTER
        ws.cell(r, 2).fill = PatternFill("solid", fgColor=leg["fill"])
        ws.cell(r, 3, leg["label"]).font = FONT_LEGEND
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        r += 1

    r += 1
    ws.cell(r, 2, "Note:").font = FONT_DROW
    note = (
        "Blank cells in Kód column are intentional per Pattern 26 (honest fallback) — "
        "Karel ověří v KROS s pomocí 'family hint' v Pozn. NIKDY nedoplňovat 999999999 "
        "ani TBD. Provenance trail (mnozstvi_formula, source, _audit_gap_fixed, vyjasneni_ref) "
        "je v souboru File A: Vykaz_vymer_RD_Jachymov_VSE_VARIANTY_2026-05-27_v2_final.xlsx."
    )
    ws.cell(r, 3, note).font = FONT_LEGEND
    ws.cell(r, 3).alignment = ALIGN_LEFT
    ws.merge_cells(start_row=r, start_column=3, end_row=r+3, end_column=8)
    ws.row_dimensions[r].height = 60

    _set_col_widths(ws, {"A": 2, "B": 16, "C": 50, "D": 18, "E": 16, "F": 18, "G": 12, "H": 12})


# =====================================================================
# Sheet 2/3 — per-SO soupis prací
# =====================================================================
def build_so_sheet(wb: Workbook, obj_meta: dict, items: list[dict]) -> tuple[str, str]:
    """Build SO sheet; return (sheet_name, J-column-cena-range) for Rekap pointer."""
    ws = wb.create_sheet(obj_meta["sheet"])

    # ---- Krycí list ----
    ws["B2"] = "KRYCÍ LIST SOUPISU PRACÍ"
    ws["B2"].font = FONT_TITLE
    krycí = [
        ("Stavba:",       STAVBA),
        ("Objekt:",       f"SO {obj_meta['kod']} — {obj_meta['nazev']}"),
        ("KSO:",          "RD — pozemní stavby + rekonstrukce + nástavba"),
        ("Místo:",        MISTO),
        ("Datum:",        DATUM),
        ("Zadavatel:",    ZADAVATEL),
        ("Zhotovitel:",   ZHOTOVITEL),
        ("Projektant:",   PROJEKTANT),
        ("Zpracovatel:",  ZPRACOVATEL),
    ]
    for i, (lbl, val) in enumerate(krycí, start=4):
        ws.cell(i, 2, lbl).font = FONT_DROW
        ws.cell(i, 3, val).font = FONT_BODY
        ws.cell(i, 3).alignment = ALIGN_LEFT
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=11)

    # ---- Rekapitulace členění (per-kapitola subtotals via SUMIFS over Kapitola column) ----
    # We'll populate kapitola SUMIFS pointers AFTER we know K-row range.
    rekap_start_row = 14
    ws.cell(rekap_start_row, 2, "REKAPITULACE ČLENĚNÍ SOUPISU").font = FONT_DROW
    ws.cell(rekap_start_row, 2).fill = FILL_REKAP
    ws.merge_cells(start_row=rekap_start_row, start_column=2, end_row=rekap_start_row, end_column=11)
    rekap_header_row = rekap_start_row + 1
    rekap_headers = ["", "Kapitola", "Cena bez DPH"]
    for c, h in enumerate(rekap_headers, start=2):
        cell = ws.cell(rekap_header_row, c, h)
        cell.font = FONT_HDR
        cell.fill = FILL_HDR
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL

    # Determine which kapitolas appear in this objekt (preserve KAPITOLA_ORDER)
    kaps_in_obj = [k for k in KAPITOLA_ORDER if any(it.get("kapitola") == k for it in items)]
    rekap_data_start = rekap_header_row + 1
    for idx, k in enumerate(kaps_in_obj):
        r = rekap_data_start + idx
        ws.cell(r, 3, k).font = FONT_BODY
        ws.cell(r, 3).border = BORDER_ALL
        # SUMIFS placeholder — we'll fill in after soupis is built
        # We use 'M' column (hidden kapitola tag) on K-rows below.
        ws.cell(r, 4).number_format = '#,##0.00 "Kč"'
        ws.cell(r, 4).border = BORDER_ALL
        ws.cell(r, 2).border = BORDER_ALL
    rekap_data_end = rekap_data_start + len(kaps_in_obj) - 1

    # Total
    rekap_total = rekap_data_end + 1
    ws.cell(rekap_total, 3, "Celkem objekt").font = FONT_DROW
    ws.cell(rekap_total, 3).fill = FILL_TOTAL
    ws.cell(rekap_total, 4, f"=SUM(D{rekap_data_start}:D{rekap_data_end})")
    ws.cell(rekap_total, 4).fill = FILL_TOTAL
    ws.cell(rekap_total, 4).font = FONT_DROW
    ws.cell(rekap_total, 4).number_format = '#,##0.00 "Kč"'
    for c in range(2, 5):
        ws.cell(rekap_total, c).border = BORDER_ALL
    ws.cell(rekap_total, 2).fill = FILL_TOTAL

    # ---- Soupis prací header ----
    soupis_header_row = rekap_total + 3
    ws.cell(soupis_header_row, 2, "SOUPIS PRACÍ").font = FONT_DROW
    ws.cell(soupis_header_row, 2).fill = FILL_REKAP
    ws.merge_cells(start_row=soupis_header_row, start_column=2, end_row=soupis_header_row, end_column=11)
    col_hdr_row = soupis_header_row + 1
    # KROS standard columns C..K (we shift to B..J then K = Cenová soustava, L = Pozn., M = hidden kapitola tag)
    col_headers = ["PČ", "Typ", "Kód", "Popis", "MJ", "Množství",
                   "J.cena [Kč]", "Cena celkem [Kč]", "Cenová soustava", "Pozn.", "_kap_tag"]
    for c, h in enumerate(col_headers, start=2):
        cell = ws.cell(col_hdr_row, c, h)
        cell.font = FONT_HDR
        cell.fill = FILL_HDR
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL

    # Hide tag column (M)
    ws.column_dimensions["L"].width = 50  # Pozn.
    ws.column_dimensions["M"].hidden = True

    # ---- K-rows by kapitola ----
    k_row = col_hdr_row + 1
    k_row_first = k_row
    pc_seq = 0  # sequential PČ per objekt (across all kapitolas)
    cena_k_rows = []  # collect row numbers for SUMIFS

    # Top-level HSV / PSV / M / VRN dividers
    cur_group = None

    for kap in kaps_in_obj:
        kap_items = [it for it in items if it.get("kapitola") == kap]
        # Group divider (HSV / PSV / M / VRN)
        grp = kap_items[0].get("kapitola_group", "")
        grp_label_map = {
            "HSV": "Práce a dodávky HSV",
            "PSV": "Práce a dodávky PSV",
            "M":   "Práce a dodávky M (montáže)",
            "VRN": "Vedlejší rozpočtové náklady",
        }
        if grp != cur_group:
            ws.cell(k_row, 3, "D").font = FONT_DROW
            ws.cell(k_row, 4, grp).font = FONT_DROW
            ws.cell(k_row, 5, grp_label_map.get(grp, grp)).font = FONT_DROW
            for c in range(2, 13):
                ws.cell(k_row, c).fill = FILL_DROW
                ws.cell(k_row, c).border = BORDER_ALL
            k_row += 1
            cur_group = grp

        # Kapitola divider
        ws.cell(k_row, 3, "D").font = FONT_DROW
        ws.cell(k_row, 4, kap.split(" ")[0]).font = FONT_DROW
        ws.cell(k_row, 5, kap).font = FONT_DROW
        for c in range(2, 13):
            ws.cell(k_row, c).fill = FILL_REKAP
            ws.cell(k_row, c).border = BORDER_ALL
        k_row += 1

        for it in kap_items:
            pc_seq += 1
            icon, label, fill_color, pozn = classify_status(it)

            ws.cell(k_row, 2, pc_seq).font = FONT_BODY
            ws.cell(k_row, 2).alignment = ALIGN_CENTER
            ws.cell(k_row, 3, "K").font = FONT_BODY
            ws.cell(k_row, 3).alignment = ALIGN_CENTER

            kod = it.get("urs_code_proposed")
            ws.cell(k_row, 4, kod if kod else "").font = FONT_BODY  # Kód (blank if null per Pattern 26)
            ws.cell(k_row, 4).alignment = ALIGN_LEFT

            popis = it.get("popis", "")
            ws.cell(k_row, 5, popis).font = FONT_BODY
            ws.cell(k_row, 5).alignment = ALIGN_LEFT

            mj = it.get("mj", "")
            ws.cell(k_row, 6, mj).font = FONT_BODY
            ws.cell(k_row, 6).alignment = ALIGN_CENTER

            mn = it.get("mnozstvi")
            ws.cell(k_row, 7, mn).font = FONT_BODY
            ws.cell(k_row, 7).number_format = "#,##0.000"
            ws.cell(k_row, 7).alignment = ALIGN_RIGHT

            # J.cena — empty for Karel (no fabricated prices)
            ws.cell(k_row, 8, None).number_format = "#,##0.00"
            ws.cell(k_row, 8).alignment = ALIGN_RIGHT

            # Cena celkem = množství × J.cena (works with blank J.cena → 0)
            ws.cell(k_row, 9, f"=IF(H{k_row}=\"\",0,G{k_row}*H{k_row})")
            ws.cell(k_row, 9).number_format = "#,##0.00"
            ws.cell(k_row, 9).alignment = ALIGN_RIGHT

            # Cenová soustava
            ws.cell(k_row, 10, "ÚRS 2026").font = FONT_BODY
            ws.cell(k_row, 10).alignment = ALIGN_CENTER

            # Pozn. (status icon + label + evidence)
            ws.cell(k_row, 11, pozn).font = FONT_BODY
            ws.cell(k_row, 11).alignment = ALIGN_LEFT
            ws.cell(k_row, 11).fill = PatternFill("solid", fgColor=fill_color)

            # Hidden kapitola tag for SUMIFS
            ws.cell(k_row, 12, kap).font = FONT_BODY

            for c in range(2, 13):
                ws.cell(k_row, c).border = BORDER_ALL

            cena_k_rows.append(k_row)
            k_row += 1

            # Optional VV-row with mnozstvi_formula (KROS standard)
            mn_formula = it.get("mnozstvi_formula")
            if mn_formula:
                ws.cell(k_row, 3, "VV").font = FONT_VV
                ws.cell(k_row, 3).alignment = ALIGN_CENTER
                ws.cell(k_row, 5, str(mn_formula)).font = FONT_VV
                ws.cell(k_row, 5).alignment = ALIGN_LEFT
                ws.cell(k_row, 7, mn).font = FONT_VV
                ws.cell(k_row, 7).number_format = "#,##0.000"
                for c in range(2, 13):
                    ws.cell(k_row, c).border = BORDER_ALL
                k_row += 1

    k_row_last = k_row - 1

    # ---- Now fill SUMIFS in rekapitulace členění ----
    # SUMIFS(Cena, kap_tag_col, kapitola) — only counts K-rows because VV-rows have no Cena
    for idx, k in enumerate(kaps_in_obj):
        r = rekap_data_start + idx
        ws.cell(r, 4, f"=SUMIFS(I{k_row_first}:I{k_row_last},L{k_row_first}:L{k_row_last},C{r})")
        # Actually wait — the kapitola is shown in COLUMN C of the rekap row.
        # Hmm but on the rekap row we put "kapitola" in column 3 = C. So criterion is C{r}. Right.

    # Sheet column widths
    _set_col_widths(ws, {
        "A": 2, "B": 6, "C": 5, "D": 14, "E": 60, "F": 8, "G": 12,
        "H": 14, "I": 16, "J": 14, "K": 36, "L": 1,  # M hidden via dimensions
    })
    ws.column_dimensions["L"].width = 50

    # Return Cena celkem range as I-col SUMIFS-ready range for Rekap stavby
    if cena_k_rows:
        cena_range = f"I{cena_k_rows[0]}:I{cena_k_rows[-1]}"
    else:
        cena_range = "I1:I1"

    return obj_meta["sheet"], cena_range


# =====================================================================
# Sheet 4 — Stav položek (URS verification audit, Karel-facing)
# =====================================================================
def build_sheet_stav_polozek(wb: Workbook, items: list[dict]):
    ws = wb.create_sheet("Stav polozek")
    ws["B2"] = "STAV URS KÓDŮ — PŘEHLED PRO KARLA"
    ws["B2"].font = FONT_TITLE

    # Group items by classify_status result
    groups = OrderedDict([
        ("✓ VERIFIED + CROSS_DISCIPLINE_OK",        []),
        ("⚠ FAMILY_VERIFIED (leaf KROS lookup)",    []),
        ("❌ WRONG_LEAF (HSV2.003 + HSV2.008 — 274 family)", []),
        ("❌ MANUAL LOOKUP (no Phase 5B verdict, blank code)", []),
        ("? needs_production_lookup (default)",     []),
    ])
    for it in items:
        cvs = it.get("cross_verification_status")
        code = it.get("urs_code_proposed")
        urs_status = it.get("urs_status", "")
        if cvs in ("VERIFIED", "CROSS_DISCIPLINE_OK"):
            groups["✓ VERIFIED + CROSS_DISCIPLINE_OK"].append(it)
        elif cvs in ("FAMILY_VERIFIED", "FAMILY_VERIFIED_CHAPTER_MATCH"):
            groups["⚠ FAMILY_VERIFIED (leaf KROS lookup)"].append(it)
        elif cvs == "WRONG_LEAF":
            groups["❌ WRONG_LEAF (HSV2.003 + HSV2.008 — 274 family)"].append(it)
        elif urs_status == "wrong_leaf_disambiguation_needed" and not code:
            groups["❌ MANUAL LOOKUP (no Phase 5B verdict, blank code)"].append(it)
        else:
            groups["? needs_production_lookup (default)"].append(it)

    # Summary table
    r = 4
    ws.cell(r, 2, "Skupina").font = FONT_HDR
    ws.cell(r, 2).fill = FILL_HDR
    ws.cell(r, 3, "Počet").font = FONT_HDR
    ws.cell(r, 3).fill = FILL_HDR
    ws.cell(r, 2).alignment = ALIGN_CENTER
    ws.cell(r, 3).alignment = ALIGN_CENTER
    for c in (2, 3):
        ws.cell(r, c).border = BORDER_ALL
    r += 1
    for label, lst in groups.items():
        ws.cell(r, 2, label).font = FONT_BODY
        ws.cell(r, 3, len(lst)).font = FONT_BODY
        ws.cell(r, 3).alignment = ALIGN_CENTER
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2)
        for c in (2, 3):
            ws.cell(r, c).border = BORDER_ALL
        r += 1
    ws.cell(r, 2, "CELKEM AKTIVNÍ").font = FONT_DROW
    ws.cell(r, 2).fill = FILL_TOTAL
    ws.cell(r, 3, sum(len(v) for v in groups.values())).font = FONT_DROW
    ws.cell(r, 3).fill = FILL_TOTAL
    ws.cell(r, 3).alignment = ALIGN_CENTER
    for c in (2, 3):
        ws.cell(r, c).border = BORDER_ALL
    r += 3

    # Per-group actionable list
    for label, lst in groups.items():
        if not lst:
            continue
        ws.cell(r, 2, label).font = FONT_DROW
        ws.cell(r, 2).fill = FILL_REKAP
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        r += 1
        # column headers
        col_h = ["ID", "Objekt", "Kapitola", "Kód proposed", "Family", "Popis", "Pozn. / hint"]
        for c, h in enumerate(col_h, start=2):
            ws.cell(r, c, h).font = FONT_HDR
            ws.cell(r, c).fill = FILL_HDR
            ws.cell(r, c).alignment = ALIGN_CENTER
            ws.cell(r, c).border = BORDER_ALL
        r += 1
        for it in lst:
            _, _, _, pozn = classify_status(it)
            ws.cell(r, 2, it["id"]).font = FONT_BODY
            ws.cell(r, 3, it.get("objekt", "")).font = FONT_BODY
            ws.cell(r, 4, it.get("kapitola", "")).font = FONT_BODY
            ws.cell(r, 5, it.get("urs_code_proposed") or "").font = FONT_BODY
            ws.cell(r, 6, it.get("urs_code_family_6digit") or "").font = FONT_BODY
            ws.cell(r, 7, it.get("popis", "")[:70]).font = FONT_BODY
            ws.cell(r, 8, pozn).font = FONT_BODY
            for c in range(2, 9):
                ws.cell(r, c).border = BORDER_ALL
                ws.cell(r, c).alignment = ALIGN_LEFT
            r += 1
        r += 1

    _set_col_widths(ws, {
        "A": 2, "B": 30, "C": 14, "D": 28, "E": 14, "F": 8, "G": 60, "H": 60,
    })


# =====================================================================
# Sheet 5 — URS výběr audit (full transparency per Pattern 25)
# =====================================================================
def build_sheet_urs_vyber_audit(wb: Workbook, items: list[dict]):
    ws = wb.create_sheet("URS vyber audit")
    ws["B2"] = "URS VÝBĚR — AUDIT PRO TRANSPARENTNOST (Pattern 25)"
    ws["B2"].font = FONT_TITLE

    r = 4
    col_h = [
        "ID", "Kapitola", "Popis", "Kód proposed", "Alternativy",
        "Family 6digit", "urs_status", "urs_confidence",
        "cross_verification_status", "cross_verification_evidence_url",
        "urs_code_proposed_was", "correct_code_hint", "urs_verification_note",
    ]
    for c, h in enumerate(col_h, start=2):
        ws.cell(r, c, h).font = FONT_HDR
        ws.cell(r, c).fill = FILL_HDR
        ws.cell(r, c).alignment = ALIGN_CENTER
        ws.cell(r, c).border = BORDER_ALL
    r += 1

    for it in items:
        ws.cell(r, 2, it["id"]).font = FONT_BODY
        ws.cell(r, 3, it.get("kapitola", "")).font = FONT_BODY
        ws.cell(r, 4, it.get("popis", "")[:80]).font = FONT_BODY
        ws.cell(r, 5, it.get("urs_code_proposed") or "").font = FONT_BODY
        ws.cell(r, 6, ", ".join(it.get("urs_alternatives", []) or [])).font = FONT_BODY
        ws.cell(r, 7, it.get("urs_code_family_6digit") or "").font = FONT_BODY
        ws.cell(r, 8, it.get("urs_status") or "").font = FONT_BODY
        ws.cell(r, 9, it.get("urs_confidence")).font = FONT_BODY
        ws.cell(r, 9).number_format = "0.00"
        ws.cell(r, 10, it.get("cross_verification_status") or "").font = FONT_BODY
        ws.cell(r, 11, it.get("cross_verification_evidence_url") or "").font = FONT_BODY
        ws.cell(r, 12, it.get("urs_code_proposed_was") or "").font = FONT_BODY
        hint = (it.get("correct_code_hint") or "")[:200]
        ws.cell(r, 13, hint).font = FONT_BODY
        note = (it.get("urs_verification_note") or "")[:200]
        ws.cell(r, 14, note).font = FONT_BODY
        for c in range(2, 15):
            ws.cell(r, c).border = BORDER_ALL
            ws.cell(r, c).alignment = ALIGN_LEFT
        # Color rows by status
        _, _, fill_color, _ = classify_status(it)
        ws.cell(r, 10).fill = PatternFill("solid", fgColor=fill_color)
        r += 1

    _set_col_widths(ws, {
        "A": 2, "B": 30, "C": 28, "D": 50, "E": 12, "F": 20, "G": 10,
        "H": 32, "I": 10, "J": 22, "K": 40, "L": 16, "M": 50, "N": 50,
    })
    ws.freeze_panes = "B5"


# =====================================================================
# Main
# =====================================================================
def main() -> None:
    data = json.loads(ITEMS_PATH.read_text())
    all_items = data["items"]

    # Filter to active subset for File B
    active = [it for it in all_items
              if it.get("status_flag") in ("ready_for_phase2", "needs_subdod_mapping")]
    excluded = [it for it in all_items
                if it.get("status_flag") == "deprecated_audit_v2"]

    print(f"Active items for File B: {len(active)} (excluded {len(excluded)} deprecated)")

    # Group items per objekt, preserving items.json order within
    items_per_obj = {obj["objekt"]: [it for it in active if it.get("objekt") == obj["objekt"]]
                     for obj in OBJEKTY}
    for obj in OBJEKTY:
        cnt = len(items_per_obj[obj["objekt"]])
        print(f"  {obj['kod']}  {obj['nazev']:50s}  {cnt} items")

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Sheet 2 + 3 first (need ranges for Sheet 1)
    per_so_ranges = {}
    for obj in OBJEKTY:
        sheet_name, cena_range = build_so_sheet(wb, obj, items_per_obj[obj["objekt"]])
        per_so_ranges[sheet_name] = cena_range

    # Sheet 1
    build_sheet_rekapitulace(wb, per_so_ranges)
    # Move Rekapitulace to position 0
    wb.move_sheet("Rekapitulace stavby", offset=-(len(wb.sheetnames) - 1))

    # Sheet 4
    build_sheet_stav_polozek(wb, active)

    # Sheet 5
    build_sheet_urs_vyber_audit(wb, active)

    wb.save(OUT_PATH)

    # Validation log
    pattern_26_check_fail = []
    for it in active:
        code = it.get("urs_code_proposed")
        if code and str(code).strip() in ("999999999", "999-999-999", "TBD"):
            pattern_26_check_fail.append({"id": it["id"], "code": code})

    log = {
        "_generated_at": TODAY,
        "output_file": str(OUT_PATH.relative_to(ROOT)),
        "pattern_compliance": {
            "pattern_15": "items.json untouched (read-only)",
            "pattern_25_26": "status icons (✓⚠❌?) per honest fallback hierarchy",
            "pattern_27": "evidence URLs surfaced in Pozn. + audit sheet",
            "pattern_28": "compound key (id, kapitola) ordering stable",
            "pattern_32": "File B = production deliverable; no provenance columns (mnozstvi_formula as VV-row only, KROS standard)",
        },
        "sheets": list(wb.sheetnames),
        "sheet_count": len(wb.sheetnames),
        "active_items": len(active),
        "items_per_objekt": {obj["kod"]: len(items_per_obj[obj["objekt"]]) for obj in OBJEKTY},
        "excluded_deprecated": [{"id": it["id"], "status_flag": it.get("status_flag")} for it in excluded],
        "pattern_26_fabricated_code_check": {
            "fail_count": len(pattern_26_check_fail),
            "violations": pattern_26_check_fail,
        },
        "per_so_cena_ranges": per_so_ranges,
    }
    LOG_PATH.write_text(json.dumps(log, indent=2, ensure_ascii=False))

    print()
    print(f"✓ File B written: {OUT_PATH.relative_to(ROOT)}")
    print(f"✓ Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
    print(f"✓ Pattern 26 fabricated-code check: {len(pattern_26_check_fail)} violations")
    print(f"✓ Validation log: {LOG_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
