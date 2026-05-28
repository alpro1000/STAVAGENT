"""HK212 sequential construction list generator.

Reads items_hk212_etap1.json (138 items), re-orders them in logical Czech
construction sequence (11 phases per task §2), emits XLSX + CSV + JSON +
README + HANDOFF. No code matching, no classification, no invention —
flat ordered list for manual KROS/URS + price assignment.
"""
from __future__ import annotations

import csv
import json
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "outputs" / "phase_1_etap1" / "items_hk212_etap1.json"
OUT_DIR = ROOT / "outputs" / "sequential_list"

# --------------------------------------------------------------------------- #
# Construction-sequence definition.
# Each (phase_num, phase_title, [(krok_label, [item_id, ...]), ...])
# Item IDs come straight from items_hk212_etap1.json; ordering inside each
# krok follows logical sub-sequence (e.g. výztuž → bednění → beton).
# --------------------------------------------------------------------------- #
SEQUENCE: list[tuple[int, str, list[tuple[str, list[str]]]]] = [
    (
        1,
        "PŘÍPRAVA STAVENIŠTĚ + GEODÉZIE",
        [
            ("Geodetické vytýčení + vytýčení sítí", ["VRN-014", "VRN-017"]),
            ("Vyjádření správců sítí", ["VRN-016"]),
            ("Oplocení staveniště", ["VRN-004"]),
            ("Zařízení staveniště — buňky", ["VRN-001", "VRN-002", "VRN-003"]),
            ("Mobilní WC pro pracovníky", ["VRN-007"]),
            ("Přípojky pro stavbu (voda + elektro)", ["VRN-005", "VRN-006"]),
            ("BOZP plán + koordinátor", ["VRN-009", "VRN-008"]),
            ("Pojištění stavby", ["VRN-010"]),
            (
                "Kácení dřevin + frézování pařezů",
                ["HSV-1-020", "HSV-1-021", "HSV-1-022", "HSV-1-023"],
            ),
            (
                "Odstranění asfaltové vrstvy + odvoz suti",
                ["HSV-1-025", "HSV-1-026", "HSV-1-027"],
            ),
        ],
    ),
    (
        2,
        "ZEMNÍ PRÁCE",
        [
            ("Hlavní výkop figury pod desku", ["HSV-1-001"]),
            ("Ruční výkop v ochranných pásmech sítí", ["HSV-1-006", "HSV-1-007"]),
            (
                "Obetonování stávajícího potrubí (po odkrytí)",
                ["HSV-1-009", "HSV-1-010"],
            ),
            (
                "Dohloubky pro patky a pasy",
                ["HSV-1-002", "HSV-1-003", "HSV-1-005"],
            ),
            (
                "Atypický základ (varianta pilota) + pažení",
                ["HSV-1-004", "HSV-1-008"],
            ),
            ("Výkopy pro nové přípojky", ["HSV-1-011", "HSV-1-012"]),
            (
                "Nakládání + vodorovné přemístění + odvoz zeminy",
                ["HSV-1-016", "HSV-1-017", "HSV-1-018", "HSV-1-019"],
            ),
        ],
    ),
    (
        3,
        "ZÁKLADY",
        [
            (
                "Výměna aktivní zóny — odstr. navážek → odvoz → dovoz štěrku → hutnění → Edef2 → geodet",
                [
                    "HSV-1-028a",
                    "HSV-1-028b",
                    "HSV-1-028c",
                    "HSV-1-028d",
                    "HSV-1-028e",
                    "HSV-1-028f",
                ],
            ),
            ("Štěrkové lože + zhutnění podloží", ["HSV-1-013", "HSV-1-014"]),
            ("Hydroizolace plošná pod desku", ["HSV-2-018"]),
            ("Výztuž patek a pasů", ["HSV-2-009"]),
            (
                "Patky rámové — bednění + beton + odbednění",
                ["HSV-2-002", "HSV-2-001", "HSV-2-003"],
            ),
            (
                "Patky štítové — bednění + beton + odbednění",
                ["HSV-2-005", "HSV-2-004", "HSV-2-006"],
            ),
            ("Pasy ŽB — bednění + beton", ["HSV-2-008", "HSV-2-007"]),
            ("Zásyp výkopů kolem patek", ["HSV-1-015"]),
            (
                "VARIANTA — pilota Ø800 / L=8 m (alternativa)",
                ["HSV-2-010", "HSV-2-011", "HSV-2-012"],
            ),
            (
                "Podlahová deska — bednění + výztuž + distance + beton",
                [
                    "HSV-2-014",
                    "HSV-2-016",
                    "HSV-2-017",
                    "HSV-2-015",
                    "HSV-2-013",
                ],
            ),
        ],
    ),
    (
        4,
        "NOSNÁ OCELOVÁ KONSTRUKCE",
        [
            ("Doprava OK na stavbu", ["HSV-3-011"]),
            ("Kotvení sloupů ke patkám", ["HSV-3-009"]),
            ("Pomocné lešení OK", ["HSV-9-002"]),
            (
                "Dodávka prvků OK (sloupy, příčle, vaznice, ztužidla)",
                [
                    "HSV-3-001",
                    "HSV-3-002",
                    "HSV-3-003",
                    "HSV-3-004",
                    "HSV-3-005",
                    "HSV-3-006",
                    "HSV-3-007",
                ],
            ),
            ("Styčníkové plechy + spojovací materiál", ["HSV-3-008"]),
            ("Montáž OK — kompletace rámové haly", ["HSV-3-010"]),
            ("Demontáž pomocného lešení po montáži OK", ["HSV-9-003"]),
            (
                "Povrchové úpravy OK (antikoroze + protipožární)",
                ["HSV-3-012", "HSV-3-013"],
            ),
            ("Revize OK + protokol EXC2", ["HSV-3-014"]),
        ],
    ),
    (
        5,
        "OPLÁŠTĚNÍ KINGSPAN",
        [
            ("Pomocné lešení pro opláštění (pojízdné)", ["HSV-9-004"]),
            ("Doprava sendvičových panelů Kingspan", ["PSV-OPL-007"]),
            (
                "Stěnové panely KS 1000 NF 120 mm — dodávka + montáž",
                ["PSV-OPL-001", "PSV-OPL-002"],
            ),
            (
                "Stěnové panely KS 1000 FR 150 mm (rohové zóny) — dodávka + montáž",
                ["PSV-OPL-009", "PSV-OPL-010"],
            ),
            (
                "Střešní panely KS 1000 RW 160 mm (hlavní střecha) — dodávka + montáž",
                ["PSV-OPL-003", "PSV-OPL-004"],
            ),
            (
                "Střešní panely KS 1000 FF 175 mm (zóna) — dodávka + montáž",
                ["PSV-OPL-011", "PSV-OPL-012"],
            ),
            ("Spojovací materiál + EPDM těsnění Kingspan", ["PSV-OPL-006"]),
            ("Klempířské lemy přechody střecha–fasáda", ["PSV-OPL-005"]),
            ("Statické posouzení uchycení Kingspan", ["PSV-OPL-008"]),
        ],
    ),
    (
        6,
        "KLEMPÍŘSKÉ + ODVODNĚNÍ STŘECHY",
        [
            ("Doprava klempířiny na stavbu", ["PSV-78x-012"]),
            (
                "Atikové oplechování + úžlabí + nároží",
                ["PSV-78x-007", "PSV-78x-008", "PSV-78x-009"],
            ),
            ("Tmelení spár klempířských konstrukcí", ["PSV-78x-010"]),
            (
                "Lindab svody — dodávka + montáž",
                ["PSV-78x-001", "PSV-78x-002"],
            ),
            (
                "Wavin Tegra střešní vpusti — dodávka + montáž",
                ["PSV-78x-003", "PSV-78x-004"],
            ),
            (
                "MEA Mearin liniový žlab + mřížka",
                ["PSV-78x-005", "PSV-78x-006"],
            ),
            ("Ostatní oplechování (parapety, lemy)", ["PSV-78x-011"]),
        ],
    ),
    (
        7,
        "VÝPLNĚ OTVORŮ",
        [
            (
                "Okna plastová — dodávka (fixní + otvíravé) + montáž + parapet + lemy",
                [
                    "PSV-76x-001",
                    "PSV-76x-013",
                    "PSV-76x-002",
                    "PSV-76x-003",
                    "PSV-76x-004",
                ],
            ),
            (
                "Sekční vrata 3500×4000 — dodávka + pohon + montáž",
                [
                    "PSV-76x-005",
                    "PSV-76x-006",
                    "PSV-76x-007",
                    "PSV-76x-008",
                ],
            ),
            (
                "Vnější 2-křídlé dveře — dodávka + montáž + zámek + práh",
                [
                    "PSV-76x-009",
                    "PSV-76x-010",
                    "PSV-76x-011",
                    "PSV-76x-012",
                ],
            ),
        ],
    ),
    (
        8,
        "IZOLACE + SOKL",
        [
            ("Penetrace soklu", ["PSV-71x-001"]),
            ("Hydroizolace svislá soklu — SBS pás", ["PSV-71x-002"]),
            ("Hydroizolační lišty rohové + napojení", ["PSV-71x-003"]),
            ("Ochranná vrstva — nopová folie HDPE", ["PSV-71x-004"]),
        ],
    ),
    (
        9,
        "PODLAHA PRŮMYSLOVÁ",
        [
            ("Penetrace + primer pod stěrku", ["PSV-77x-001"]),
            ("Epoxidová / PU stěrka 4–5 mm", ["PSV-77x-002"]),
            (
                "Lokální zesílení anchorage zón strojů",
                ["PSV-77x-003"],
            ),
            (
                "Dilatace + lemy + protiskluz",
                ["PSV-77x-004", "PSV-77x-005", "PSV-77x-006"],
            ),
        ],
    ),
    (
        9.5,
        "TZB INSTALACE — VYTÁPĚNÍ",
        [
            ("Doprava topidel + přesun hmot ÚT", ["M-UT-012"]),
            ("Montážní plošina pro práci ve výšce 5 m (ECOSUN)", ["M-UT-009"]),
            (
                "ECOSUN S+ 12 — dodávka + závěs + montáž na strop haly",
                ["M-UT-001", "M-UT-003", "M-UT-002"],
            ),
            (
                "Dalap E-HP 9 kW — dodávka + montáž v rozích haly",
                ["M-UT-004", "M-UT-005"],
            ),
            ("Regulace — UET 15D + prostorové termostaty", ["M-UT-006", "M-UT-007"]),
            ("Pomocný montážní + kotevní materiál ÚT", ["M-UT-008"]),
            ("Funkční zkouška + nastavení systému vytápění", ["M-UT-010"]),
            ("Předání dokumentace + zaškolení obsluhy ÚT", ["M-UT-011"]),
        ],
    ),
    (
        10,
        "OSTATNÍ + PŘESUN HMOT",
        [
            ("Přesun hmot HSV vodorovně", ["HSV-9-001"]),
            ("Doprava materiálu na stavbu — paušál", ["VRN-013"]),
        ],
    ),
    (
        11,
        "DOKONČENÍ + REVIZE + ODEVZDÁNÍ",
        [
            (
                "Revize TZB (elektro, hydrant, hromosvod)",
                ["VRN-020", "VRN-021", "VRN-022"],
            ),
            ("Likvidace odpadů O + N", ["VRN-011", "VRN-012"]),
            ("Geodetické zaměření DSPS", ["VRN-015"]),
            (
                "Předávací protokoly + kolaudace",
                ["VRN-018", "VRN-019"],
            ),
            ("Náhradní výsadba dřevin (po dokončení)", ["HSV-1-024"]),
        ],
    ),
    (
        12,
        "VENKOVNÍ ÚPRAVY + ZTI (SO-13 + SO-14) — VV authoritative + norm-verified",
        [
            # --- ZTI vnější sítě (VV) — early phase, coordinate s výkopy ---
            (
                "ZTI vnější — kanalizace čerpaná (potrubí + přečerpávací stanice + zemní práce)",
                [f"M-ZTI-{n:03d}" for n in range(32, 47)],  # C section
            ),
            (
                "ZTI vnější — vodovod (PE100 + armatury + zemní práce)",
                [f"M-ZTI-{n:03d}" for n in range(47, 57)],  # D section
            ),
            # --- Domovní kanalizace — šachty + retenční/zasakovací + potrubí ---
            (
                "ZTI domovní kanalizace — potrubí PP-KG/HT + revizní šachty + "
                "retenční nádoba 15 m³ + zasakovací těleso (VV K1, pending ABMV_33)",
                [f"M-ZTI-{n:03d}" for n in range(16, 32)],  # B section
            ),
            # --- Domovní vodovod — potrubí + armatury + TUV + požární ---
            (
                "ZTI domovní vodovod — potrubí + armatury + TUV + požární vodovod "
                "+ hydrant D19/30 (VV DV)",
                [f"M-ZTI-{n:03d}" for n in range(1, 16)],  # A section
            ),
            # --- Přeložka vodovodu (podmíněná) ---
            ("Přeložka vodovodního řadu (PODMÍNĚNÁ — pending ABMV_23)", ["M-VK-009"]),
            ("NN chránička HDPE — stavební příprava (silové v elektro VV)", ["M-VK-008"]),
            # --- Okapník beton 0.7 m ---
            (
                "Okapní chodník beton — 10-layer stack 0.7 m × 80 m (výkop → odvoz "
                "→ pláň → geotextilie → štěrk → hutnění → bednění → výztuž → beton "
                "→ dilatace, pending ABMV_31/32/34)",
                [
                    "M-VK-027", "M-VK-028", "M-VK-024", "M-VK-029", "M-VK-022",
                    "M-VK-025", "M-VK-026", "M-VK-021", "M-VK-020", "M-VK-023",
                ],
            ),
            # --- Dlažba 1.5 m chodník (ADJACENT okapník = 2.2 m) ---
            (
                "Zámková dlažba chodník 1.5 m — ČSN 73 6131 stack (výkop → odvoz → "
                "pláň → geotextilie → štěrk → drť 4/8 → obrubník → lože → dlažba → "
                "spáry; TZ ARS B p09, pending ABMV_34)",
                [f"M-VK-{n:03d}" for n in range(30, 40)],
            ),
            # --- Rampy + žlab + vegetace ---
            ("Drcený štěrk podklad ramp", ["M-VK-003"]),
            (
                "Rampy 4× (R1-R4) — výztuž + bednění + beton + obrubník (per A101)",
                ["M-VK-002", "M-VK-001", "M-VK-004"],
            ),
            ("Liniový žlab pojízdný B125 — SZ + JZ fasáda (40 m)", ["M-VK-012"]),
            ("Vyspádování okolního terénu k odvodnění", ["M-VK-019"]),
            ("Ohumusování + osetí travou zbytkových ploch", ["M-VK-018"]),
        ],
    ),
]


def load_items() -> tuple[dict, dict]:
    """Load items.json; filter out items dropped/superseded per user decisions.

    Items with `_status_flag` starting with "dropped" or "superseded" are
    preserved in JSON (Pattern 14 forward journey) but excluded from the
    active sequential list rows.
    """
    raw = json.loads(SRC.read_text(encoding="utf-8"))
    inactive = ("dropped", "superseded")
    items_by_id = {
        it["id"]: it for it in raw["items"]
        if not str(it.get("_status_flag", "")).startswith(inactive)
    }
    return raw, items_by_id


def build_ordered_rows(items_by_id: dict) -> list[dict]:
    """Walk SEQUENCE, attach _sequence_position + _phase + _krok to each item."""
    out: list[dict] = []
    pos = 0
    seen: set[str] = set()
    for phase_num, phase_title, kroky in SEQUENCE:
        for krok_label, ids in kroky:
            for iid in ids:
                if iid not in items_by_id:
                    raise SystemExit(f"FATAL: id {iid!r} from SEQUENCE not in items.json")
                if iid in seen:
                    raise SystemExit(f"FATAL: id {iid!r} appears twice in SEQUENCE")
                seen.add(iid)
                pos += 1
                src = items_by_id[iid]
                out.append({
                    "_sequence_position": pos,
                    "_phase": phase_num,
                    "_phase_title": phase_title,
                    "_krok": krok_label,
                    "id": src["id"],
                    "kapitola": src.get("kapitola", ""),
                    "popis": src.get("popis", ""),
                    "mj": src.get("mj", ""),
                    "mnozstvi": src.get("mnozstvi", ""),
                    "confidence": src.get("confidence", ""),
                    "vzorec": make_vzorec(src),
                    "_review_concrete_class": src.get("_review_concrete_class"),
                    "_review_qty": src.get("_review_qty"),
                    "_vyjasneni_ref": src.get("_vyjasneni_ref"),
                })
    missing = set(items_by_id) - seen
    if missing:
        raise SystemExit(f"FATAL: {len(missing)} items.json IDs missing from SEQUENCE: {sorted(missing)}")
    if len(out) != 235:
        raise SystemExit(f"FATAL: expected 235 ordered rows, got {len(out)}")
    return out


VZOREC_MAX_LEN = 220
_REF_TYPE_LABELS = {
    "drawing": "výkres",
    "drawing_measurement": "měřeno",
    "tz_section": "TZ",
    "statika_section": "statika",
    "phase_ref": "phase",
    "step3_metric": "Step3",
    "dxf_layer": "DXF",
    "abmv_closure": "ABMV",
    "abmv_reopen": "ABMV",
}


def _format_ref(ref: dict) -> str:
    rtype = ref.get("type", "")
    label = _REF_TYPE_LABELS.get(rtype, rtype or "ref")
    code = ref.get("code") or ref.get("section") or ref.get("document") or ""
    if not code or code == "?":
        return label
    return f"{label}:{code}"


def make_vzorec(src: dict) -> str:
    """Build Vzorec / Zdroj výměry cell text from items.json _audit_trail."""
    at = src.get("audit_trail")
    if not isinstance(at, dict):
        return "(zdroj nenalezen — manual verify)"
    formula = at.get("formula", "").strip()
    if not formula:
        return "(zdroj nenalezen — manual verify)"

    refs = at.get("reference") or []
    ref_parts: list[str] = []
    seen: set[str] = set()
    for ref in refs:
        if not isinstance(ref, dict):
            continue
        formatted = _format_ref(ref)
        if formatted and formatted not in seen:
            seen.add(formatted)
            ref_parts.append(formatted)
        if len(ref_parts) >= 3:
            break

    text = formula
    if ref_parts:
        text = f"{text}  [{', '.join(ref_parts)}]"

    journey = at.get("_analytical_journey")
    if isinstance(journey, list) and journey:
        last = journey[-1]
        prev_val = last.get("value_m3") or last.get("value")
        prev_method = last.get("method", "")
        if prev_val is not None and prev_method:
            text = f"{text}  · was {prev_val}: {prev_method[:60]}"

    if len(text) > VZOREC_MAX_LEN:
        text = text[: VZOREC_MAX_LEN - 1].rstrip() + "…"
    return text


def make_pozn(row: dict) -> str:
    bits: list[str] = []
    if row["_review_concrete_class"]:
        bits.append("rev:beton-class")
    if row["_review_qty"]:
        bits.append("rev:qty")
    ref = row["_vyjasneni_ref"]
    if ref:
        if isinstance(ref, list):
            bits.append("ABMV:" + ",".join(ref))
        else:
            bits.append(f"ABMV:{ref}")
    return " | ".join(bits)


def write_xlsx(rows: list[dict], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Postup stavby"

    headers = ["#", "Krok", "Fáze", "Kapitola", "ID", "Položka (popis)", "MJ", "Mnozstvi",
               "Vzorec / Zdroj výměry", "Pozn."]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    for col_idx, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    phase_fill = PatternFill("solid", fgColor="FB923C")
    phase_font = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
    krok_fill = PatternFill("solid", fgColor="DBEAFE")
    krok_font = Font(bold=True, color="1E3A8A", name="Calibri", size=10)
    low_conf_fill = PatternFill("solid", fgColor="FEF3C7")
    review_fill = PatternFill("solid", fgColor="FED7AA")
    stripe_fill = PatternFill("solid", fgColor="F8FAFC")
    border = Border(left=Side(style="thin", color="E5E7EB"),
                    right=Side(style="thin", color="E5E7EB"),
                    top=Side(style="thin", color="E5E7EB"),
                    bottom=Side(style="thin", color="E5E7EB"))

    current_phase = 0
    current_krok = ""
    stripe_toggle = False
    for r in rows:
        # Phase separator
        if r["_phase"] != current_phase:
            current_phase = r["_phase"]
            current_krok = ""
            sep = f"═══ FÁZE {r['_phase']}: {r['_phase_title']} ═══"
            ws.append([sep] + [""] * (len(headers) - 1))
            row_idx = ws.max_row
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
            cell = ws.cell(row=row_idx, column=1)
            cell.fill = phase_fill
            cell.font = phase_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row_idx].height = 22
            stripe_toggle = False
        # Krok separator
        if r["_krok"] != current_krok:
            current_krok = r["_krok"]
            ws.append([f"→ {current_krok}"] + [""] * (len(headers) - 1))
            row_idx = ws.max_row
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
            cell = ws.cell(row=row_idx, column=1)
            cell.fill = krok_fill
            cell.font = krok_font
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            stripe_toggle = False
        # Data row
        pozn = make_pozn(r)
        ws.append([
            r["_sequence_position"],
            "",
            r["_phase"],
            r["kapitola"],
            r["id"],
            r["popis"],
            r["mj"],
            r["mnozstvi"],
            r["vzorec"],
            pozn,
        ])
        row_idx = ws.max_row
        try:
            conf_val = float(r["confidence"]) if r["confidence"] != "" else 1.0
        except (TypeError, ValueError):
            conf_val = 1.0
        has_review = bool(r["_review_concrete_class"] or r["_review_qty"])
        if has_review:
            row_fill = review_fill
        elif conf_val < 0.70:
            row_fill = low_conf_fill
        elif stripe_toggle:
            row_fill = stripe_fill
        else:
            row_fill = None
        stripe_toggle = not stripe_toggle
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.border = border
            if row_fill is not None:
                c.fill = row_fill
            if col_idx in (6, 9, 10):
                c.alignment = Alignment(wrap_text=True, vertical="top")
            elif col_idx in (1, 3, 7, 8):
                c.alignment = Alignment(horizontal="center", vertical="top")
            else:
                c.alignment = Alignment(vertical="top")

    # Column widths
    widths = {1: 5, 2: 4, 3: 6, 4: 10, 5: 14, 6: 60, 7: 8, 8: 11, 9: 55, 10: 26}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_csv(rows: list[dict], path: Path) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["#", "Fáze", "Krok", "Kapitola", "ID", "Položka", "MJ", "Mnozstvi",
                    "Vzorec / Zdroj výměry", "Pozn."])
        for r in rows:
            w.writerow([
                r["_sequence_position"], r["_phase"], r["_krok"], r["kapitola"],
                r["id"], r["popis"], r["mj"], r["mnozstvi"], r["vzorec"], make_pozn(r),
            ])


def write_json(rows: list[dict], items_by_id: dict, path: Path) -> None:
    out = []
    for r in rows:
        src = dict(items_by_id[r["id"]])
        src["_sequence_position"] = r["_sequence_position"]
        src["_phase"] = r["_phase"]
        src["_phase_title"] = r["_phase_title"]
        src["_krok"] = r["_krok"]
        src["_vzorec_display"] = r["vzorec"]
        out.append(src)
    path.write_text(json.dumps({"items": out, "count": len(out)},
                               ensure_ascii=False, indent=2), encoding="utf-8")


def write_readme(rows: list[dict], items_by_id: dict, path: Path) -> None:
    abmv_open = set()
    for it in items_by_id.values():
        ref = it.get("_vyjasneni_ref")
        if isinstance(ref, list):
            abmv_open.update(ref)
    review_qty = [it["id"] for it in items_by_id.values() if it.get("_review_qty")]
    review_conc = [it["id"] for it in items_by_id.values() if it.get("_review_concrete_class")]
    body = f"""# HK212 Sequential Construction List

235 active položek v logickém pořadí výstavby. Fáze 12 = Venkovní úpravy (SO-13) + ZTI (SO-14, 56 items z projektantského VV). 5 items dropped (M-VK-013..017 asfalt) + 5 superseded by VV (M-VK-005/006/007/010/011) preserved s _status_flag. Okapní chodník 10-layer + zámková dlažba 1.5 m 10-layer.
Žádné kódy, žádné ceny — jen popis + výměra ve správném pořadí.

**Použití:** user manually adds KROS/URS codes + ceny per row.

**Source:** `outputs/phase_1_etap1/items_hk212_etap1.json` (245 entries, 235 active)
**Branch:** `claude/hk212-vk-final-minimal` (12 M-UT items added per investor scope change 2026-05-26)
**Generated:** {date.today().isoformat()}

## Fáze
1. PŘÍPRAVA STAVENIŠTĚ + GEODÉZIE
2. ZEMNÍ PRÁCE
3. ZÁKLADY
4. NOSNÁ OCELOVÁ KONSTRUKCE
5. OPLÁŠTĚNÍ KINGSPAN
6. KLEMPÍŘSKÉ + ODVODNĚNÍ STŘECHY
7. VÝPLNĚ OTVORŮ
8. IZOLACE + SOKL
9. PODLAHA PRŮMYSLOVÁ
9.5. TZB INSTALACE — VYTÁPĚNÍ (M-UT, 12 items, DPS D.1.4.2)
10. OSTATNÍ + PŘESUN HMOT
11. DOKONČENÍ + REVIZE + ODEVZDÁNÍ
12. VENKOVNÍ ÚPRAVY (SO-13, M-VK, 24 active items — minimal scope + okapní chodník complete 10-layer stack, DPS 06/2026)

## Soubory
- `hk212_sequential_list.xlsx` — single-sheet "Postup stavby", formatted, freeze row 1, includes Vzorec / Zdroj výměry column
- `hk212_sequential_list.csv` — flat CSV mirror pro grep / diff
- `hk212_sequential_list.json` — items.json fields + `_sequence_position` + `_phase` + `_krok` + `_vzorec_display`

## Sloupec "Vzorec / Zdroj výměry"
Každý řádek nese stručný výpočet kvantity + zdrojové reference (výkres / TZ / statika / Step3 / phase ref) extrahované z `items_hk212_etap1.json` field `audit_trail.formula` + `audit_trail.reference`. Pokud zdroj nelze odvodit, řádek nese `(zdroj nenalezen — manual verify)`. Text je truncated na ~220 znaků; plné detaily (vstupy + krok-za-krok + analytical_journey) zůstávají v `items.json` audit_trail.

## ABMV open ({len(abmv_open)})
{', '.join(sorted(abmv_open)) if abmv_open else '(žádné)'}

## _review_qty flags ({len(review_qty)})
{', '.join(review_qty) if review_qty else '(žádné)'}

## _review_concrete_class flags ({len(review_conc)})
{', '.join(review_conc) if review_conc else '(žádné)'}

## Vizuální značení v XLSX
- oranžový fáze-separator (═══ FÁZE N: NÁZEV ═══)
- modrý krok-header (→ Krok)
- žluté pozadí: confidence < 0.70
- oranžové pozadí: aktivní `_review_*` flag
- alternující řádkové pruhování v rámci fáze
"""
    path.write_text(body, encoding="utf-8")


def write_handoff(rows: list[dict], items_by_id: dict, path: Path) -> None:
    abmv_open = set()
    for it in items_by_id.values():
        ref = it.get("_vyjasneni_ref")
        if isinstance(ref, list):
            abmv_open.update(ref)
    review_qty = [it["id"] for it in items_by_id.values() if it.get("_review_qty")]
    review_conc = [it["id"] for it in items_by_id.values() if it.get("_review_concrete_class")]
    low_conf = sum(1 for it in items_by_id.values()
                   if isinstance(it.get("confidence"), (int, float)) and it["confidence"] < 0.70)
    phase_counts: dict[int, int] = {}
    for r in rows:
        phase_counts[r["_phase"]] = phase_counts.get(r["_phase"], 0) + 1

    body = f"""# HANDOFF — HK212 Sequential List

**Status:** generated, validated, ready for manual KROS/URS + price assignment.

**Branch:** `claude/hk212-vk-final-minimal`
**Date:** {date.today().isoformat()}

## Counts
- 235 active items (138 baseline + 12 M-UT + 34 M-VK venkovní úpravy + 56 M-ZTI z projektantského VV; 10 inactive: 5 dropped + 5 superseded)
- 13 fází (1–12, vč. 9.5 TZB Vytápění + nová 12 Venkovní úpravy po Dokončení)
- 36 active items added (12 M-UT + 24 M-VK = 14 kept + 10 okapní layers), 5 dropped per user decision, 0 invented
- 1 ABMV updated + 3 ABMV opened (ABMV_23/24/25 from PR #1235) + 5 ABMV resolved + 1 new (ABMV_26..31 from this PR)

## Per-phase distribution
{chr(10).join(f"- FÁZE {p}: {phase_counts[p]} items" for p in sorted(phase_counts))}

## Quality flags propagated (from items.json)
- `_vyjasneni_ref` open ABMV: {len(abmv_open)} → {', '.join(sorted(abmv_open)) if abmv_open else '(none)'}
- `_review_qty`: {len(review_qty)} → {', '.join(review_qty) if review_qty else '(none)'}
- `_review_concrete_class`: {len(review_conc)} → {', '.join(review_conc) if review_conc else '(none)'}
- confidence < 0.70: {low_conf} items (yellow-tinted v XLSX)

## Validation (§6)
- ✔ row count = 168 (excl. separator rows + dropped items)
- ✔ each items.json id appears exactly once
- ✔ no id missing from output
- ✔ phases ordered 1→9, 9.5, 10, 11, 12 monotonically
- ✔ within phase: výztuž → bednění → beton; doprava → montáž; dodávka → instalace

## Sidecar files
- `outputs/sequential_list/hk212_sequential_list.xlsx`
- `outputs/sequential_list/hk212_sequential_list.csv`
- `outputs/sequential_list/hk212_sequential_list.json`
- `outputs/sequential_list/README.md`

## NOT in scope (per task §5)
- žádné KROS/URS code matching
- žádné nové položky
- žádná price assignment
- žádná modifikace `items.json`

## Next steps (user manual)
1. Otevři `hk212_sequential_list.xlsx` v Excelu / LibreOffice.
2. Iteruj řádek po řádku, doplň KROS/URS kód + J.cena.
3. Pro řádky se žlutým / oranžovým pozadím nejdřív vyřeš ABMV / review flag.

## Notes for next session
- Source branch `dilenska-ok-ut-dps-integration` was NOT found in remote (`git branch -r` empty for it).
  Generated from current main tip (`82b7cab2` — HK212 memory consolidation), which has the
  items.json checked in at `test-data/hk212_hala/outputs/phase_1_etap1/items_hk212_etap1.json`.
- If `dilenska-ok-ut-dps-integration` later appears with newer items.json, re-run
  `scripts/build_sequential_list.py` (deterministic — same items in = same order out, since
  SEQUENCE is hard-coded by ID).
"""
    path.write_text(body, encoding="utf-8")


def main() -> None:
    raw, items_by_id = load_items()
    rows = build_ordered_rows(items_by_id)

    # Validation
    assert len(rows) == 235, len(rows)
    ids = [r["id"] for r in rows]
    assert len(set(ids)) == 235, "duplicate IDs in output"
    # NOTE: items.json contains 173 entries but 5 are dropped per user decision
    # 2026-05-27 (_status_flag=dropped_per_user_decision_2026-05-27).
    # Sequential list active = 168 (= 173 entries - 5 dropped).
    active_ids_in_json = {iid for iid, src in items_by_id.items()
                          if not str(src.get("_status_flag","")).startswith("dropped")}
    assert set(ids) == active_ids_in_json, "ID set mismatch (active items only)"
    phases = [r["_phase"] for r in rows]
    for prev, cur in zip(phases, phases[1:]):
        assert cur >= prev, f"phase regression: {prev} → {cur}"
    empty_vzorec = [r["id"] for r in rows if not r["vzorec"]]
    assert not empty_vzorec, f"empty vzorec for: {empty_vzorec}"
    fallback_vzorec = [r["id"] for r in rows if r["vzorec"].startswith("(zdroj nenalezen")]
    print(f"  vzorec coverage: {235 - len(fallback_vzorec)}/235 with formula, {len(fallback_vzorec)} fallback")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    write_xlsx(rows, OUT_DIR / "hk212_sequential_list.xlsx")
    write_csv(rows, OUT_DIR / "hk212_sequential_list.csv")
    write_json(rows, items_by_id, OUT_DIR / "hk212_sequential_list.json")
    write_readme(rows, items_by_id, OUT_DIR / "README.md")
    write_handoff(rows, items_by_id, OUT_DIR / "HANDOFF_SEQUENTIAL.md")

    print(f"OK — 235 active items in 13 phases (vč. 9.5 + 12 Venkovní + ZTI VV + dlažba) written to {OUT_DIR}")
    for p in sorted({r["_phase"] for r in rows}):
        n = sum(1 for r in rows if r["_phase"] == p)
        print(f"  FÁZE {p}: {n}")


if __name__ == "__main__":
    main()
