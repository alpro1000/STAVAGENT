#!/usr/bin/env python3
"""
hk212 Phase 1 Etap 1 — Layer 3 helper: build smety corpus from example_vv/.

Walks test-data/hk212_hala/example_vv/*.{xlsx,xls}, extracts every K (položka)
and M (material) row from KROS "Export Komplet" sheets, and writes a flat
JSON corpus that rematch_layer3.py can fuzzy-match hk212 needs_review items
against. These reference smety are rozpočtář-curated descriptions paired with
real URS codes — strong signal for hala-style projects (Forestina × 4
analogous industrial halls inside the corpus).

Format (KROS Export Komplet, per app/parsers/xlsx_komplet_parser.py):
  col C = position number  col D = type (K|M|D|PP|VV|...)
  col E = URS code         col F = description (Czech)
  col G = MJ               col H = quantity
  col K = price source (e.g. "CS ÚRS 2023 02")

Run:
    python test-data/hk212_hala/scripts/phase_1_etap1/build_smety_corpus.py

Output:
    test-data/hk212_hala/outputs/phase_1_etap1/smety_corpus.json
"""

from __future__ import annotations

import json
import logging
import re
import sys
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("❌ Missing openpyxl  →  pip install openpyxl")
    sys.exit(2)


REPO_ROOT = Path(__file__).resolve().parents[3].parent
SMETY_DIR = REPO_ROOT / "test-data/hk212_hala/example_vv"
OUT_PATH = REPO_ROOT / "test-data/hk212_hala/outputs/phase_1_etap1/smety_corpus.json"

# KROS Export Komplet columns (0-based)
COL_TYPE = 3   # D
COL_CODE = 4   # E
COL_POPIS = 5  # F
COL_MJ = 6     # G
COL_QTY = 7    # H
COL_SOURCE = 10  # K — "CS ÚRS 2023 02"

# Accept K (main position) and M (material) row types
KEEP_TYPES = {"K", "M"}

# Sheet names to skip (cover/recapitulation, not item-bearing)
SKIP_SHEET_PATTERNS = re.compile(
    r"(rekap|krycí|stavba|pokyny|vzorpolozky|vrn |souhrn|n - návrh|a - příprava)",
    re.IGNORECASE,
)

# Code formats: 6-9 digit (122351104) or dotted (273.32.1611) or with prefix
RE_CODE = re.compile(r"^(?:R-)?[\d.]{5,15}$")

# "CS ÚRS 2023 02"  → 2023-II, "01" → I, "02" → II
RE_VINTAGE = re.compile(r"(?:CS\s+)?(?:ÚRS|URS)\s+(\d{4})\s+0?(\d)", re.IGNORECASE)


def setup_logging() -> logging.Logger:
    logger = logging.getLogger("smety_corpus")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    h = logging.StreamHandler(sys.stdout)
    h.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%H:%M:%S"))
    logger.addHandler(h)
    return logger


def deburr_lower(text: str) -> str:
    nfkd = unicodedata.normalize("NFKD", text)
    stripped = "".join(c for c in nfkd if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", stripped.lower()).strip()


def parse_vintage(source_marker: str | None) -> str | None:
    if not source_marker:
        return None
    m = RE_VINTAGE.search(str(source_marker))
    if not m:
        return None
    year, half = m.group(1), m.group(2)
    return f"{year}-I" if half == "1" else f"{year}-II"


def extract_items_from_sheet(ws, source_file: str, logger: logging.Logger) -> list[dict]:
    items: list[dict] = []
    seen_in_sheet: set[tuple[str, str]] = set()  # dedupe within sheet by (code, popis)
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) <= COL_SOURCE:
            continue
        type_cell = row[COL_TYPE]
        if type_cell not in KEEP_TYPES:
            continue
        code = str(row[COL_CODE]).strip() if row[COL_CODE] is not None else ""
        popis = str(row[COL_POPIS]).strip() if row[COL_POPIS] is not None else ""
        if not code or not popis or not RE_CODE.match(code):
            continue
        key = (code, deburr_lower(popis)[:80])
        if key in seen_in_sheet:
            continue
        seen_in_sheet.add(key)
        items.append({
            "code": code,
            "popis": popis,
            "popis_normalized": deburr_lower(popis),
            "mj": str(row[COL_MJ]).strip() if row[COL_MJ] else "",
            "qty": row[COL_QTY] if isinstance(row[COL_QTY], (int, float)) else None,
            "vintage": parse_vintage(row[COL_SOURCE]),
            "row_type": type_cell,
            "source_file": source_file,
            "source_sheet": ws.title[:40],
        })
    return items


def process_file(f: Path, logger: logging.Logger) -> list[dict]:
    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    except Exception as e:
        logger.warning(f"  skipping {f.name}: {e}")
        return []
    items: list[dict] = []
    sheets_scanned = 0
    sheets_with_items = 0
    for sheet_name in wb.sheetnames:
        if SKIP_SHEET_PATTERNS.search(sheet_name):
            continue
        sheets_scanned += 1
        ws = wb[sheet_name]
        sheet_items = extract_items_from_sheet(ws, f.name, logger)
        if sheet_items:
            sheets_with_items += 1
            items.extend(sheet_items)
    wb.close()
    logger.info(f"  ✓ {f.name[:55]:55s} → {len(items):>5} items "
                f"(across {sheets_with_items}/{sheets_scanned} sheets)")
    return items


def main() -> int:
    logger = setup_logging()
    if not SMETY_DIR.exists():
        logger.error(f"Smety dir missing: {SMETY_DIR}")
        return 2

    files = sorted([
        p for p in SMETY_DIR.iterdir()
        if p.is_file() and p.suffix.lower() in (".xlsx", ".xls")
    ])
    if not files:
        logger.error(f"No xlsx/xls files in {SMETY_DIR}")
        return 3

    logger.info(f"Found {len(files)} reference smety files")
    all_items: list[dict] = []
    per_file: list[dict] = []
    for f in files:
        before = len(all_items)
        items = process_file(f, logger)
        all_items.extend(items)
        per_file.append({"file": f.name, "items_count": len(all_items) - before})

    # Global dedupe — same code+popis across files keeps first occurrence
    seen: set[tuple[str, str]] = set()
    deduped: list[dict] = []
    for it in all_items:
        key = (it["code"], it["popis_normalized"][:80])
        if key in seen:
            continue
        seen.add(key)
        deduped.append(it)
    logger.info(f"Total: {len(all_items)} raw → {len(deduped)} after cross-file dedupe")

    # Vintage distribution
    from collections import Counter
    by_vintage = Counter(it["vintage"] for it in deduped)
    logger.info(f"By vintage: {dict(by_vintage)}")
    by_type = Counter(it["row_type"] for it in deduped)
    logger.info(f"By row type: {dict(by_type)}")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "built_at": datetime.now(timezone.utc).isoformat(),
        "source_dir": str(SMETY_DIR.relative_to(REPO_ROOT)),
        "files": per_file,
        "total_items": len(deduped),
        "by_vintage": dict(by_vintage),
        "by_row_type": dict(by_type),
        "items": deduped,
    }
    with open(OUT_PATH, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2, ensure_ascii=False)
    size_kb = OUT_PATH.stat().st_size / 1024
    logger.info(f"Wrote {OUT_PATH.relative_to(REPO_ROOT)} ({size_kb:.0f} KB)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
