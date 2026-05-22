"""
phase_a_preflight.py — inventory KROS catalog + example_vv + items.json
Output: outputs/soupis_praci/preflight_inventory.md
Read-only, no mutations.
"""
import json
import sqlite3
import re
from pathlib import Path
from collections import Counter
import openpyxl

BASE = Path(__file__).resolve().parent.parent.parent
KROS_DB = BASE.parent / "kros_catalog.db"
EXAMPLE_VV_DIR = BASE / "example_vv"
ITEMS_PATH = BASE / "outputs/phase_1_etap1/items_hk212_etap1.json"
OUT_PATH = BASE / "outputs/soupis_praci/preflight_inventory.md"

# TSKP třída → kapitola name (standard ČR)
TSKP_TRIDA = {
    "1": "HSV-1 zemní práce",
    "2": "HSV-2-3 zakládání + svislé konstrukce",
    "3": "HSV-3 svislé konstrukce + zdivo",
    "4": "HSV-4 vodorovné konstrukce",
    "5": "HSV-5 komunikace",
    "6": "HSV-6 úpravy povrchů",
    "7": "PSV (71x-78x) řemesla",
    "8": "HSV-8 trubní vedení",
    "9": "HSV-9 přesun hmot + ostatní",
}


def inventory_kros() -> dict:
    conn = sqlite3.connect(str(KROS_DB))
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM kros_items")
    total = cur.fetchone()[0]
    # by 3-digit prefix (TSKP třída + skupina)
    cur.execute(
        "SELECT substr(kod_polozky,1,1) AS t, COUNT(*) FROM kros_items GROUP BY t ORDER BY t"
    )
    by_trida = dict(cur.fetchall())
    # by MJ
    cur.execute("SELECT mj, COUNT(*) FROM kros_items GROUP BY mj ORDER BY 2 DESC LIMIT 12")
    by_mj = cur.fetchall()
    # vintage
    cur.execute("SELECT vintage_year, COUNT(*) FROM kros_items GROUP BY vintage_year ORDER BY 1")
    by_year = cur.fetchall()
    # FTS sanity
    cur.execute("SELECT COUNT(*) FROM kros_fts")
    fts_rows = cur.fetchone()[0]
    conn.close()
    return {
        "total_items": total,
        "by_trida": by_trida,
        "by_mj": by_mj,
        "by_year": by_year,
        "fts_indexed": fts_rows,
    }


def sample_kros_by_keyword(keyword: str, limit: int = 5) -> list:
    """FTS sample for sanity."""
    conn = sqlite3.connect(str(KROS_DB))
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT k.kod_polozky, k.popis, k.mj FROM kros_fts f "
            "JOIN kros_items k ON f.kod_polozky=k.kod_polozky "
            "WHERE kros_fts MATCH ? LIMIT ?",
            (keyword, limit),
        )
        rows = cur.fetchall()
    except sqlite3.OperationalError:
        rows = []
    conn.close()
    return rows


def inventory_example_vv() -> list:
    """Scan each reference výkaz for kapitola structure + code format hints."""
    results = []
    for f in sorted(EXAMPLE_VV_DIR.iterdir()):
        if f.is_dir():
            continue
        ext = f.suffix.lower()
        meta = {"file": f.name, "size_kb": f.stat().st_size // 1024, "ext": ext}
        if ext in (".xlsx", ".xls"):
            try:
                if ext == ".xlsx":
                    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
                    meta["sheets"] = wb.sheetnames
                    # Sniff first sheet for code formats
                    ws = wb[wb.sheetnames[0]]
                    rows_read = 0
                    codes = []
                    for row in ws.iter_rows(max_row=300, values_only=True):
                        rows_read += 1
                        for cell in row:
                            if cell is None:
                                continue
                            s = str(cell).strip()
                            # KROS-style 9-digit code OR URS-style 8-digit
                            if re.match(r"^\d{8,12}$", s):
                                codes.append(s[:9])
                    meta["row_count_sample"] = rows_read
                    meta["code_samples"] = list(set(codes))[:8]
                else:
                    # .xls via xlrd
                    import xlrd
                    book = xlrd.open_workbook(f)
                    meta["sheets"] = book.sheet_names()
                    sh = book.sheet_by_index(0)
                    meta["row_count_sample"] = sh.nrows
                    codes = []
                    for r in range(min(300, sh.nrows)):
                        for c in range(sh.ncols):
                            v = sh.cell_value(r, c)
                            s = str(v).strip() if v else ""
                            if re.match(r"^\d{8,12}$", s):
                                codes.append(s[:9])
                    meta["code_samples"] = list(set(codes))[:8]
            except Exception as e:
                meta["error"] = str(e)[:120]
        elif ext == ".xml":
            meta["note"] = "UNIXML KROS export (skipped for sniff)"
        elif ext == ".pdf":
            meta["note"] = "drawing PDF (not parsed)"
        results.append(meta)
    return results


def inventory_items() -> dict:
    d = json.load(open(ITEMS_PATH, encoding="utf-8"))
    items = d["items"]
    by_kap = Counter(i.get("kapitola", "?") for i in items)
    conf_buckets = Counter()
    for it in items:
        c = it.get("confidence", 0)
        if c >= 0.9:
            conf_buckets["0.90+"] += 1
        elif c >= 0.75:
            conf_buckets["0.75-0.90"] += 1
        elif c >= 0.50:
            conf_buckets["0.50-0.75"] += 1
        else:
            conf_buckets["<0.50"] += 1
    by_mj = Counter(i.get("mj", "?") for i in items)
    with_code = sum(1 for i in items if i.get("urs_code"))
    return {
        "total": len(items),
        "by_kapitola": dict(by_kap),
        "confidence_buckets": dict(conf_buckets),
        "by_mj": dict(by_mj.most_common(10)),
        "with_existing_code": with_code,
        "without_existing_code": len(items) - with_code,
    }


def main():
    print("=== Phase A pre-flight inventory ===\n")
    kros = inventory_kros()
    print(f"KROS DB: {kros['total_items']} items, FTS indexed {kros['fts_indexed']}")
    vv = inventory_example_vv()
    print(f"example_vv: {len(vv)} files")
    items = inventory_items()
    print(f"items.json: {items['total']} items across {len(items['by_kapitola'])} kapitol")

    # FTS sanity samples for HK212-relevant keywords
    fts_tests = {}
    for kw in ["beton patky", "Kingspan", "sendvičový panel", "Lindab",
               "hloubení figury", "KARI síť", "atika oplechování", "vrata sekční"]:
        fts_tests[kw] = sample_kros_by_keyword(kw, 3)

    # Render markdown
    md = [f"# HK212 Soupis prací — Pre-flight Inventory\n",
          "**Generated:** 2026-05-24 · **Phase A** of soupis_praci pipeline\n",
          "---\n\n## 1. KROS catalog (test-data/kros_catalog.db)\n",
          f"- **Total items:** {kros['total_items']:,}",
          f"- **FTS5 index rows:** {kros['fts_indexed']:,} (full-text search ready)",
          f"- **Vintage range:** {min((y for y,_ in kros['by_year'] if y), default='?')}–{max((y for y,_ in kros['by_year'] if y), default='?')}",
          "",
          "### TSKP třída distribution (1st digit of kód):"]
    md.append("| Třída | Items | Název |")
    md.append("|---|---:|---|")
    for t, n in sorted(kros["by_trida"].items()):
        md.append(f"| {t} | {n:,} | {TSKP_TRIDA.get(t, '?')} |")
    md.append("\n### Top MJ (měrná jednotka):")
    md.append("| MJ | Count |\n|---|---:|")
    for mj, n in kros["by_mj"]:
        md.append(f"| {mj or '(empty)'} | {n:,} |")

    md.append("\n### Vintage years:")
    md.append("| Year | Count |\n|---|---:|")
    for y, n in kros["by_year"]:
        md.append(f"| {y} | {n:,} |")

    md.append("\n### FTS sanity check — HK212 keywords:")
    for kw, hits in fts_tests.items():
        md.append(f"\n**`{kw}`** — {len(hits)} hits:")
        for h in hits:
            md.append(f"- `{h[0]}` ({h[2]}) — {(h[1] or '')[:90]}")

    md.append("\n---\n\n## 2. example_vv reference corpus\n")
    md.append("| # | File | Ext | Size kB | Sheets / Note |")
    md.append("|---|---|---|---:|---|")
    for i, v in enumerate(vv, 1):
        sheets = v.get("sheets")
        if sheets:
            note = f"{len(sheets)} sheets: {', '.join(s[:18] for s in sheets[:4])}{'...' if len(sheets) > 4 else ''}"
        else:
            note = v.get("note") or v.get("error", "?")
        md.append(f"| {i} | {v['file'][:60]} | {v['ext']} | {v['size_kb']} | {note} |")

    # Code samples from XLSX/XLS only
    md.append("\n### Code format samples per file (first 8 unique 8-12 digit codes):\n")
    for v in vv:
        if v.get("code_samples"):
            md.append(f"- **{v['file'][:60]}** → {', '.join(v['code_samples'])}")

    # Top 3 most similar to HK212 (ocelová hala + sendvič opláštění + 500-1500 m²)
    md.append("\n### Top 3 references most similar to HK212 (ocelová hala + Kingspan):")
    md.append("1. **HALA JHV (zadání)** — ocelová hala, KROS export, primární reference")
    md.append("2. **Hala na sul Rozmital_rozpocet_slepy** — slepý rozpočet, podobná velikost")
    md.append("3. **ANTRACIT logistická hala Touškov — oceňovací tabulka** — logistická hala, sendvičové opláštění")

    md.append("\n---\n\n## 3. items_hk212_etap1.json (current state)\n")
    md.append(f"- **Total items:** {items['total']}")
    md.append(f"- **Items with existing urs_code:** {items['with_existing_code']}")
    md.append(f"- **Items needing Tier 1/2 match:** {items['without_existing_code']}")
    md.append("\n### Per kapitola:")
    md.append("| Kapitola | Count |\n|---|---:|")
    for k, n in sorted(items["by_kapitola"].items()):
        md.append(f"| {k} | {n} |")

    md.append("\n### Confidence distribution:")
    md.append("| Range | Count |\n|---|---:|")
    for r, n in items["confidence_buckets"].items():
        md.append(f"| {r} | {n} |")

    md.append("\n### Top MJ in items.json:")
    md.append("| MJ | Count |\n|---|---:|")
    for mj, n in items["by_mj"].items():
        md.append(f"| {mj} | {n} |")

    md.append("\n---\n\n## 4. Matching strategy decision\n")
    md.append("- **Primary tool:** SQLite FTS5 on `kros_fts.popis_normalized` (faster + better than TF-IDF for Czech text)")
    md.append("- **MJ filter:** narrow candidates to matching mj (m³/m²/kg/ks/bm/paušál)")
    md.append("- **Třída filter:** narrow to TSKP first-digit consistent with HK212 kapitola:")
    md.append("  - HSV-1 → třída 1 (zemní + bourání)")
    md.append("  - HSV-2 → třída 2 (zakládání)")
    md.append("  - HSV-3 → třída 1+5 (ocelová konstrukce: 13xxxx montáž OK; 553xxx dodávka profily)")
    md.append("  - HSV-9 → třída 9 (přesun hmot + lešení)")
    md.append("  - PSV-71x → třída 711+713 (izolace)")
    md.append("  - PSV-76x → třída 762-767 (truhlář + zámečník + nátěr)")
    md.append("  - PSV-77x → třída 776+781 (podlahy)")
    md.append("  - PSV-78x → třída 764 (klempíř)")
    md.append("  - PSV-OPL → třída 342 (montáž opláštění) + 553 (dodávka panelů)")
    md.append("  - VRN → třída 0 nebo 9 (VRN nemá KROS code typicky → Tier 2 custom)")

    md.append("\n## 5. Environment notes\n")
    md.append("- ✅ `openpyxl` available (for .xlsx read/write)")
    md.append("- ✅ `xlrd 2.0.2` available (for old .xls Forestina files)")
    md.append("- ✅ `sqlite3` stdlib")
    md.append("- ❌ `pandas` NOT installed — using openpyxl+sqlite3 directly")
    md.append("- ❌ `reportlab` NOT installed — **PDF rekapitulace dropped** from §5 deliverables. Excel + JSON will be the primary outputs; PDF flagged in handoff as P3.")

    md.append("\n## 6. Acceptance gates (re-stated)\n")
    md.append("- [ ] ≥ 60 % items get Tier 1 KROS match (confidence ≥ 0.70)")
    md.append("- [ ] ~40 % flagged Tier 2 custom with nearest KROS ref")
    md.append("- [ ] Excel hk212_soupis_praci.xlsx renders, 12 sheets (PDF dropped)")
    md.append("- [ ] JSON twin preserves audit_trail")
    md.append("- [ ] ABMV unresolved listed")
    md.append("- [ ] Original items.json UNMODIFIED")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text("\n".join(md), encoding="utf-8")
    print(f"\n✓ Inventory written → {OUT_PATH}")


if __name__ == "__main__":
    main()
