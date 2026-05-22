"""
phase_c_generate_outputs.py — final soupis prací deliverables.

Inputs:
  - outputs/phase_1_etap1/items_hk212_etap1.json (128 items, READ-ONLY)
  - outputs/phase_1_etap1/project_header.json (geometric_summary + geotech)
  - outputs/abmv_email_queue.json (22 ABMV entries)
  - outputs/soupis_praci/kros_match_results.json (Phase B output)

Outputs:
  - outputs/soupis_praci/hk212_soupis_praci.json (enriched, all fields)
  - outputs/soupis_praci/hk212_soupis_praci.xlsx (12 sheets)
     Sheet 1: Hlavička
     Sheet 2: Rekapitulace
     Sheets 3-12: per kapitola detail (10 kapitol)
     Sheet 13: ABMV + Poznámky
  - (PDF rekapitulace dropped — reportlab missing)
"""
import json
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent.parent.parent
ITEMS_PATH = BASE / "outputs/phase_1_etap1/items_hk212_etap1.json"
HEADER_PATH = BASE / "outputs/phase_1_etap1/project_header.json"
ABMV_PATH = BASE / "outputs/abmv_email_queue.json"
MATCH_PATH = BASE / "outputs/soupis_praci/kros_match_results.json"

OUT_JSON = BASE / "outputs/soupis_praci/hk212_soupis_praci.json"
OUT_XLSX = BASE / "outputs/soupis_praci/hk212_soupis_praci.xlsx"

# Kapitola order per ČR convention (HSV → PSV → VRN)
KAPITOLA_ORDER = [
    ("HSV-1", "Zemní práce"),
    ("HSV-2", "Základy + deska"),
    ("HSV-3", "Ocelová konstrukce"),
    ("HSV-9", "Ostatní stavební + lešení"),
    ("PSV-71x", "Izolace (sokl HI)"),
    ("PSV-76x", "Výplně otvorů (okna, vrata, dveře)"),
    ("PSV-77x", "Podlahy průmyslové"),
    ("PSV-78x", "Klempířské konstrukce"),
    ("PSV-OPL", "Sendvičové opláštění Kingspan"),
    ("VRN", "Vedlejší rozpočtové náklady"),
]

# Custom code prefix per kapitola for Tier 2
TIER2_PREFIX = {
    "HSV-1": "HK212-Z",   # zemní
    "HSV-2": "HK212-ZK",  # základy
    "HSV-3": "HK212-OK",  # ocelová konstrukce
    "HSV-9": "HK212-OS",  # ostatní
    "PSV-71x": "HK212-IZ",
    "PSV-76x": "HK212-VY",
    "PSV-77x": "HK212-PD",
    "PSV-78x": "HK212-KL",
    "PSV-OPL": "HK212-OPL",
    "VRN": "HK212-VRN",
}

# Tier 2 reasons (template based on kapitola)
TIER2_REASON = {
    "VRN": "VRN položky nemají standardní KROS kód — vždy custom",
    "PSV-OPL": "Kingspan produkty (KS NF, KS FF-ROC) nejsou v KROS katalogu — produktově specifické",
    "PSV-78x": "Lindab + MEA Mearin produkty nejsou v KROS katalogu — produktově specifické",
    "HSV-3": "Specifické ocelové profily + montáž — KROS pokrývá generické položky",
    "HSV-9": "Specifické lešení pro Kingspan — KROS pokrývá generické typy",
}

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2E5C8A", end_color="2E5C8A", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D9E5F4", end_color="D9E5F4", fill_type="solid")
LOW_CONF_FILL = PatternFill(start_color="FFF3B0", end_color="FFF3B0", fill_type="solid")  # yellow
REVIEW_FILL = PatternFill(start_color="FFC9C9", end_color="FFC9C9", fill_type="solid")    # red-ish
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
THIN = Side(border_style="thin", color="888888")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_json(p):
    return json.load(open(p, encoding="utf-8"))


def build_enriched_items() -> tuple:
    """Combine items.json + kros_match_results.json → enriched soupis items."""
    items_doc = load_json(ITEMS_PATH)
    items = items_doc["items"]
    matches = {m["id"]: m for m in load_json(MATCH_PATH)["matches"]}

    # Tier 2 sequence counter per kapitola
    t2_seq = defaultdict(int)
    enriched = []

    for it in items:
        m = matches.get(it["id"], {})
        tier = m.get("tier", 2)
        kapitola = it.get("kapitola", "?")
        review_flags = [k for k, v in it.items() if k.startswith("_review") and v]

        if tier == 1:
            kod_soupis = m.get("kros_code")
            popis_soupis = m.get("kros_popis") or it.get("popis")
            custom = False
            ref_code = None
        else:
            # Tier 2 custom
            t2_seq[kapitola] += 1
            kod_soupis = f"{TIER2_PREFIX.get(kapitola, 'HK212-X')}-{t2_seq[kapitola]:03d}"
            popis_soupis = it.get("popis")
            custom = True
            # Reference nearest KROS code (top candidate even if below threshold)
            cands = m.get("kros_candidates") or []
            ref_code = cands[0]["code"] if cands else None

        soupis_row = {
            "id_internal": it["id"],
            "kapitola": kapitola,
            "kod_soupis": kod_soupis,
            "popis": popis_soupis,
            "popis_hk212_original": it.get("popis"),
            "mj": it.get("mj"),
            "mnozstvi": it.get("mnozstvi"),
            "j_cena": None,
            "cena_celkem": None,
            "_cena_note": "ceny neuvedeny (user directive — separate workflow)",
            "kros_code": m.get("kros_code"),
            "kros_popis_full": m.get("kros_popis"),
            "kros_match_confidence": m.get("kros_match_confidence", 0.0),
            "kros_match_method": m.get("kros_match_method"),
            "kros_candidates_top3": m.get("kros_candidates", [])[:3],
            "tier": tier,
            "_custom_position": custom,
            "_reference_kros_code": ref_code,
            "_custom_reason": TIER2_REASON.get(kapitola, "KROS catalog gap — produktově nebo procesně specifické")
            if custom else None,
            "items_json_confidence": it.get("confidence"),
            "review_flags": review_flags,
            "vyjasneni_ref": it.get("_vyjasneni_ref") or [],
            "audit_trail_ref": f"items_hk212_etap1.json :: {it['id']}",
            "source": it.get("source"),
        }
        enriched.append(soupis_row)

    return enriched, items_doc


def build_rekapitulace(enriched: list) -> list:
    """Aggregate per kapitola for sheet 2."""
    by_kap = defaultdict(lambda: {"count": 0, "tier1": 0, "tier2": 0, "by_mj": defaultdict(float)})
    for r in enriched:
        k = r["kapitola"]
        by_kap[k]["count"] += 1
        if r["tier"] == 1:
            by_kap[k]["tier1"] += 1
        else:
            by_kap[k]["tier2"] += 1
        if r["mnozstvi"] is not None:
            by_kap[k]["by_mj"][r["mj"]] += float(r["mnozstvi"])
    return by_kap


def style_header_row(ws, row_idx: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = CELL_BORDER


def write_hlavicka(wb, header_doc, items_doc):
    ws = wb.create_sheet("1_Hlavička", 0)
    meta = items_doc.get("metadata", {})
    geom = header_doc.get("geometric_summary", {})
    geotech = header_doc.get("geotechnical_summary", {})

    rows = [
        ("HK212 — SOUPIS PRACÍ", ""),
        ("", ""),
        ("Investor", "SOLAR DISPOREC s.r.o."),
        ("Investor sídlo", "Malostranské náměstí 5/28, Malá Strana, 118 00 Praha 1"),
        ("Stavba", "Hala Hradec Králové [212] — skladová hala FVE demontážní"),
        ("Místo stavby", "Hradec Králové, k.ú. Slezské Předměstí [646971], parc. č. 1939/1"),
        ("Stupeň PD", "DSP (Dokumentace pro povolení záměru / Dokumentace pro stavební povolení)"),
        ("Zpracovatel PD", "Ing. arch. Jakub Volka, ČKA 0003947 (Basepoint s.r.o.)"),
        ("Statika", "Ing. Jiří Plachý / Bc. M. Doležal, ČKAIT 0013051"),
        ("Požární bezpečnost", "Ing. Michal Netušil, Ph.D., ČKAIT 0012242"),
        ("IGP", "ALTAGEO s.r.o., Mgr. Jan Beneda (04/2026, zak. 526 026)"),
        ("Datum soupisu", "2026-05-24"),
        ("Zhotovitel", "___ (placeholder, vyplní investor po výběru)"),
        ("Měna", "CZK (bez cen v této verzi — separate price workflow)"),
        ("", ""),
        ("GEOMETRICKÉ PARAMETRY", ""),
        ("Zastavěná plocha", f"{geom.get('zastavena_plocha_m2', '?')} m²"),
        ("Podlahová plocha (TZ ARS)", "495 m² / 531 m² (A105 měřená deska)"),
        ("Obvod budovy", f"{geom.get('obvod_budovy_m', '?')} m"),
        ("Výška budovy (v hřebeni)", "7.1 m (TZ ARS) / 6.02 m (Step 3 měřená)"),
        ("Sklon střechy", f"{geom.get('sklon_strechy_deg', '?')}°"),
        ("Plocha střechy netto", f"{geom.get('strecha_netto_m2', '?')} m²"),
        ("Plocha fasády netto", f"{geom.get('fasada_netto_m2', '?')} m²"),
        ("Rozměry půdorysu", "28.18 × 19.59 m (TZ ARS) / 27.97 × 19.31 m (PBR)"),
        ("", ""),
        ("GEOTECHNICKÉ PARAMETRY (IGP)", ""),
        ("Geotechnická kategorie", f"{geotech.get('geotechnicka_kategorie', '?')}"),
        ("Rdt únosnost zeminy", f"{geotech.get('rdt_kPa', '?')} kPa (GT2 písčité štěrky)"),
        ("HPV ustálená", f"{geotech.get('hpv_ustalena_m_pt', '?')} m p.t. (neovlivní základové poměry)"),
        ("Navážky mocnost", f"{geotech.get('navazky_mocnost_m', '?')} m (GT1 — výměna aktivní zóny doporučena)"),
        ("Založení", "Plošné na patkách C16/20 XC0 do GT2 (primary per IGP §4.3.1)"),
        ("", ""),
        ("KAPITOLY A POLOŽKY", ""),
        ("Celkem položek", str(meta.get("total_items", "?"))),
        ("Kapitol", str(len(meta.get("kapitola_modules_loaded", [])))),
        ("KROS match tier 1 (≥0.70)", "viz Sheet 2 Rekapitulace"),
        ("KROS match tier 2 (custom)", "viz Sheet 2 Rekapitulace"),
        ("", ""),
        ("REFERENČNÍ DOKUMENTY", ""),
        ("TZ ARS D.1.1", "03_ars_d11_TZ.pdf (5 stran)"),
        ("TZ Statika D.1.2", "04_statika_d12_TZ_uplna.pdf (33 stran)"),
        ("PBR", "07_pbr_kpl.pdf (32 stran, II. SPB)"),
        ("IGP", "inputs/dokumentace/IGP_ALTAGEO_526026.md"),
        ("Výkres A105", "inputs/vykresy_dxf/A105_zaklady.dxf"),
        ("Výkres A201", "inputs/vykresy_dxf/A201_vykopy.dxf"),
        ("", ""),
        ("POZNÁMKY", ""),
        ("Ceny", "BEZ CEN v této verzi — sloupce J.cena + Celkem prázdné, ready pro manuální fill investorem"),
        ("KROS vintage", "2018-2026 mix — preferovat 2026 položky pro aktuální ceny"),
        ("ABMV otevřené", "viz Sheet 13 — Poznámky"),
    ]
    for ridx, (lbl, val) in enumerate(rows, 1):
        c1 = ws.cell(row=ridx, column=1, value=lbl)
        c2 = ws.cell(row=ridx, column=2, value=val)
        if ridx == 1:
            c1.font = Font(bold=True, size=14)
        if lbl and not val and lbl not in ("", "POZNÁMKY"):
            c1.font = Font(bold=True, size=11)
            c1.fill = SECTION_FILL
            c2.fill = SECTION_FILL
        else:
            c1.font = Font(bold=True)
        c1.alignment = LEFT
        c2.alignment = LEFT

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 90


def write_rekapitulace(wb, rekap, enriched):
    ws = wb.create_sheet("2_Rekapitulace")
    headers = ["#", "Kapitola", "Název", "Položek", "Tier 1 (KROS)", "Tier 2 (custom)",
               "Hlavní MJ", "Σ mnozstvi (hlavní MJ)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    total_items = 0
    total_t1 = 0
    total_t2 = 0
    for idx, (kap, nazev) in enumerate(KAPITOLA_ORDER, 1):
        info = rekap.get(kap, {"count": 0, "tier1": 0, "tier2": 0, "by_mj": {}})
        # Pick dominant MJ
        dominant_mj = max(info["by_mj"].items(), key=lambda x: x[1])[0] if info["by_mj"] else "—"
        dominant_sum = info["by_mj"].get(dominant_mj, 0) if dominant_mj != "—" else 0
        row = idx + 1
        ws.cell(row=row, column=1, value=idx).alignment = CENTER
        ws.cell(row=row, column=2, value=kap).alignment = LEFT
        ws.cell(row=row, column=3, value=nazev).alignment = LEFT
        ws.cell(row=row, column=4, value=info["count"]).alignment = RIGHT
        ws.cell(row=row, column=5, value=info["tier1"]).alignment = RIGHT
        ws.cell(row=row, column=6, value=info["tier2"]).alignment = RIGHT
        ws.cell(row=row, column=7, value=dominant_mj).alignment = CENTER
        ws.cell(row=row, column=8, value=round(dominant_sum, 2) if dominant_sum else "—").alignment = RIGHT
        total_items += info["count"]
        total_t1 += info["tier1"]
        total_t2 += info["tier2"]
        for c in range(1, 9):
            ws.cell(row=row, column=c).border = CELL_BORDER

    # Total row
    tr = len(KAPITOLA_ORDER) + 2
    ws.cell(row=tr, column=2, value="CELKEM").font = Font(bold=True)
    ws.cell(row=tr, column=4, value=total_items).font = Font(bold=True)
    ws.cell(row=tr, column=5, value=total_t1).font = Font(bold=True)
    ws.cell(row=tr, column=6, value=total_t2).font = Font(bold=True)
    for c in range(1, 9):
        cell = ws.cell(row=tr, column=c)
        cell.fill = SECTION_FILL
        cell.border = CELL_BORDER

    # Note
    nr = tr + 2
    ws.cell(row=nr, column=1, value="Pozn.").font = Font(bold=True, italic=True)
    ws.cell(
        row=nr,
        column=2,
        value=(
            f"Tier 1 ≥ 0.70 conf KROS/URS match. Tier 2 = custom položka s referencí na nejbližší "
            f"KROS code. Tier 1 % = {round(100*total_t1/total_items,1) if total_items else 0} % "
            "(target ≥ 60 %)."
        ),
    ).alignment = LEFT

    for col, w in zip("ABCDEFGH", [5, 11, 36, 9, 14, 16, 12, 18]):
        ws.column_dimensions[col].width = w


def write_kapitola_sheet(wb, kap: str, nazev: str, idx: int, items: list):
    sheet_name = f"{idx}_{kap}".replace("/", "_")[:31]
    ws = wb.create_sheet(sheet_name)

    # Header info row
    ws.cell(row=1, column=1, value=f"{kap} — {nazev}").font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    headers = ["Poř.", "Kód KROS/Custom", "Popis", "MJ", "Mnozstvi",
               "J.cena", "Celkem", "Tier", "Conf", "Pozn."]
    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    style_header_row(ws, 2, len(headers))

    for i, it in enumerate(items, 1):
        row = i + 2
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=it["kod_soupis"])
        ws.cell(row=row, column=3, value=it["popis"])
        ws.cell(row=row, column=4, value=it["mj"])
        ws.cell(row=row, column=5, value=it["mnozstvi"])
        ws.cell(row=row, column=6, value=None)  # J.cena empty
        ws.cell(row=row, column=7, value=None)  # Celkem empty
        ws.cell(row=row, column=8, value=f"T{it['tier']}{' (custom)' if it['_custom_position'] else ''}")
        ws.cell(row=row, column=9, value=round(it["kros_match_confidence"], 2))

        # Compose pozn
        notes = []
        if it["_custom_position"]:
            notes.append(f"Custom — ref KROS: {it['_reference_kros_code'] or 'n/a'}")
        if it["review_flags"]:
            notes.append("⚠ " + ", ".join(it["review_flags"]))
        if it["vyjasneni_ref"]:
            notes.append("ABMV: " + ", ".join(it["vyjasneni_ref"]))
        if it["items_json_confidence"] is not None and it["items_json_confidence"] < 0.70:
            notes.append(f"items conf {it['items_json_confidence']:.2f}")
        ws.cell(row=row, column=10, value=" | ".join(notes) if notes else "")

        # Highlight low confidence
        if it["kros_match_confidence"] < 0.70 or (
            it["items_json_confidence"] is not None and it["items_json_confidence"] < 0.70
        ):
            for c in range(1, 11):
                ws.cell(row=row, column=c).fill = LOW_CONF_FILL
        # Highlight active review flags
        if it["review_flags"]:
            for c in range(1, 11):
                ws.cell(row=row, column=c).fill = REVIEW_FILL

        for c in range(1, 11):
            cell = ws.cell(row=row, column=c)
            cell.border = CELL_BORDER
            if c == 3:
                cell.alignment = LEFT
            elif c == 10:
                cell.alignment = LEFT
            elif c in (5, 6, 7, 9):
                cell.alignment = RIGHT
            else:
                cell.alignment = CENTER

    # Column widths
    for col, w in zip("ABCDEFGHIJ", [6, 16, 60, 9, 12, 12, 14, 12, 7, 50]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"


def write_abmv_sheet(wb, abmv_doc):
    ws = wb.create_sheet("13_ABMV_Poznámky")
    items_a = abmv_doc.get("items", [])
    open_statuses = ("open", "needs_design_clarification", "working_assumption",
                     "blocking", "working_assumption_partial", "resolved_with_caveats")
    unresolved = [a for a in items_a if a.get("status") in open_statuses]
    resolved = [a for a in items_a if a.get("status") not in open_statuses]

    ws.cell(row=1, column=1, value="ABMV — nedořešené dotazy (PROJEKTANT FILL)").font = Font(bold=True, size=12)
    headers = ["ID", "Severity", "Status", "Title", "Working assumption / Resolution"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    style_header_row(ws, 2, len(headers))

    row = 3
    for a in unresolved:
        ws.cell(row=row, column=1, value=a["id"])
        ws.cell(row=row, column=2, value=a.get("severity", "?"))
        ws.cell(row=row, column=3, value=a.get("status", "?"))
        ws.cell(row=row, column=4, value=a.get("title", ""))
        wa = a.get("working_assumption") or a.get("resolution_note", "")
        ws.cell(row=row, column=5, value=(wa or "")[:200])
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = CELL_BORDER
            ws.cell(row=row, column=c).alignment = LEFT if c >= 4 else CENTER
        if a.get("severity") == "critical":
            for c in range(1, 6):
                ws.cell(row=row, column=c).fill = REVIEW_FILL
        elif a.get("severity") == "important":
            for c in range(1, 6):
                ws.cell(row=row, column=c).fill = LOW_CONF_FILL
        row += 1

    # Resolved table
    row += 2
    ws.cell(row=row, column=1, value=f"ABMV — vyřešené ({len(resolved)} ks, pro referenci)").font = Font(bold=True, size=11)
    row += 1
    for c, h in enumerate(["ID", "Status", "Title", "Resolution date"], 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, 4)
    row += 1
    for a in resolved:
        ws.cell(row=row, column=1, value=a["id"])
        ws.cell(row=row, column=2, value=a.get("status", ""))
        ws.cell(row=row, column=3, value=a.get("title", ""))
        ws.cell(row=row, column=4, value=a.get("resolution_date", ""))
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = CELL_BORDER
            ws.cell(row=row, column=c).alignment = LEFT if c >= 3 else CENTER
        row += 1

    # Poznámky doc list
    row += 2
    ws.cell(row=row, column=1, value="REFERENČNÍ DOKUMENTY").font = Font(bold=True, size=11)
    row += 1
    docs = [
        "inputs/tz/03_ars_d11_TZ.pdf — TZ ARS D.1.1 (architektonické řešení, 5 stran)",
        "inputs/tz/04_statika_d12_TZ_uplna.pdf — TZ statika D.1.2 (33 stran)",
        "inputs/tz/07_pbr_kpl.pdf — PBR (32 stran, II. SPB ČSN 73 0804)",
        "inputs/dokumentace/IGP_ALTAGEO_526026.md — IGP (geotechnická kategorie 1)",
        "inputs/vykresy_dxf/A105_zaklady.dxf — půdorys základů",
        "inputs/vykresy_dxf/A201_vykopy.dxf — výkopy + BILANCE ZEMINY (placeholder)",
        "outputs/dsp_geometry_extraction/step3_areas/area_aggregates.json — Step 3 plochy",
        "outputs/phase_1_etap1/items_hk212_etap1.json — source items",
    ]
    for d in docs:
        ws.cell(row=row, column=1, value=d).alignment = LEFT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1

    for col, w in zip("ABCDE", [10, 13, 28, 50, 80]):
        ws.column_dimensions[col].width = w


def main():
    print("=== Phase C: generate final outputs ===\n")
    enriched, items_doc = build_enriched_items()
    header_doc = load_json(HEADER_PATH)
    abmv_doc = load_json(ABMV_PATH)
    rekap = build_rekapitulace(enriched)

    # Write JSON
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    soupis_json = {
        "_meta": {
            "project": "HK212 hala Hradec Králové",
            "investor": "SOLAR DISPOREC s.r.o.",
            "stupen_pd": "DSP",
            "datum_soupisu": "2026-05-24",
            "total_items": len(enriched),
            "tier_counts": {
                "tier1_kros_match": sum(1 for e in enriched if e["tier"] == 1),
                "tier2_custom": sum(1 for e in enriched if e["tier"] == 2),
            },
            "ceny_status": "BEZ CEN (user directive — separate workflow)",
            "source_items": "outputs/phase_1_etap1/items_hk212_etap1.json (UNMODIFIED)",
            "source_matching": "outputs/soupis_praci/kros_match_results.json (Phase B)",
        },
        "hlavicka": {
            "stavba": "Hala Hradec Králové [212]",
            "misto": "Hradec Králové, k.ú. Slezské Předměstí [646971], parc. č. 1939/1",
            "investor": "SOLAR DISPOREC s.r.o., Malostranské náměstí 5/28, 118 00 Praha 1",
            "zpracovatel_pd": "Ing. arch. Jakub Volka, ČKA 0003947 (Basepoint s.r.o.)",
            "statika": "Ing. Jiří Plachý / Bc. M. Doležal, ČKAIT 0013051",
            "pbr": "Ing. Michal Netušil, Ph.D., ČKAIT 0012242",
            "igp": "ALTAGEO s.r.o., Mgr. Jan Beneda (04/2026, zak. 526 026)",
            "geometrie": header_doc.get("geometric_summary", {}),
            "geotechnika": header_doc.get("geotechnical_summary", {}),
        },
        "items": enriched,
        "abmv_unresolved": [
            {k: a.get(k) for k in ["id", "severity", "status", "title", "working_assumption"]}
            for a in abmv_doc.get("items", [])
            if a.get("status") in ("open", "needs_design_clarification", "working_assumption",
                                    "blocking", "working_assumption_partial", "resolved_with_caveats")
        ],
    }
    json.dump(soupis_json, open(OUT_JSON, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    print(f"✓ JSON: {OUT_JSON.name}")

    # Write XLSX
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    write_hlavicka(wb, header_doc, items_doc)
    write_rekapitulace(wb, rekap, enriched)

    # Per kapitola sheets
    by_kap = defaultdict(list)
    for e in enriched:
        by_kap[e["kapitola"]].append(e)
    for idx, (kap, nazev) in enumerate(KAPITOLA_ORDER, 3):
        kap_items = by_kap.get(kap, [])
        if not kap_items:
            continue
        write_kapitola_sheet(wb, kap, nazev, idx, kap_items)

    write_abmv_sheet(wb, abmv_doc)
    wb.save(OUT_XLSX)
    print(f"✓ XLSX: {OUT_XLSX.name} ({len(wb.sheetnames)} sheets)")

    # Summary
    t1 = sum(1 for e in enriched if e["tier"] == 1)
    t2 = sum(1 for e in enriched if e["tier"] == 2)
    print(f"\nTier 1: {t1} ({100*t1/len(enriched):.1f} %) · Tier 2: {t2} · Total: {len(enriched)}")


if __name__ == "__main__":
    main()
