#!/usr/bin/env python3
"""
hk212 Phase 2 — KROS-compatible Soupis prací Excel export.

Reads items_hk212_etap1.json (post Phase-1 rematch + merge_keep_best), writes
a procurement-style 12-column xlsx that Александр opens vedle KROS for manual
top-up of remaining needs_review codes, then submits as tender soupis.

Run:
    python test-data/hk212_hala/scripts/phase_2/export_kros.py
    # defaults:
    #   --items      test-data/hk212_hala/outputs/phase_1_etap1/items_hk212_etap1.json
    #   --output-dir test-data/hk212_hala/outputs/phase_2/
    # produces:
    #   HK212_Soupis_praci.xlsx       (all 141 items, color-coded)
    #   HK212_needs_review.xlsx       (only needs_review + top-3 urs_alternatives hints)
    #   HK212_summary.md              (per-kapitola counts, match rate, totals)

Flags:
    --no-companion   skip the two companion files
    --dry-run        compute counts, no files written
    --verbose
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    print("❌ Missing dependency: openpyxl  →  pip install openpyxl")
    sys.exit(2)


# ----------------------------------------------------------------------------
# Defaults & constants
# ----------------------------------------------------------------------------

REPO_ROOT = Path(__file__).resolve().parents[3].parent
DEFAULT_ITEMS = REPO_ROOT / "test-data/hk212_hala/outputs/phase_1_etap1/items_hk212_etap1.json"
DEFAULT_OUTPUT_DIR = REPO_ROOT / "test-data/hk212_hala/outputs/phase_2"

# 12 KROS-procurement columns + 1 audit (Confidence) + 5 audit-trail (Phase 2.1)
# = 18 cols total. KROS import strips columns past L, the extras are for audit only.
COLUMNS: list[tuple[str, int]] = [
    ("O", 4),                  # A — checkbox/marker, blank
    ("Ceník", 8),              # B — KROS ceník (mapped from kapitola)
    ("Část", 8),               # C — sub-section (blank unless parseable)
    ("Kód položky", 16),       # D — urs_code / Rpol-NNN / blank for needs_review
    ("Popis", 70),             # E — popis (wraps)
    ("MJ", 8),                 # F — měrná jednotka (m²→m2 normalized)
    ("Výrobce", 12),           # G — manufacturer, blank
    ("Orientační cena", 14),   # H — unit price, blank (cenař fills)
    ("PZN", 10),               # I — pricing zone, blank
    ("Dodávka", 12),           # J — supply, blank
    ("Množství", 12),          # K — quantity from mnozstvi
    ("Celkem", 14),            # L — =H{row}*K{row}
    ("Confidence", 10),        # M — urs_match_score (audit-only, KROS ignores extras)
    # Phase 2.1 audit-trail columns — populated only when items[].audit_trail exists
    ("Lokalizace", 30),        # N — audit_trail.lokalizace
    ("Výpočet", 50),           # O — audit_trail.formula
    ("Vstupy", 25),            # P — audit_trail.inputs (comma-sep operands)
    ("Reference", 30),         # Q — audit_trail.reference (flat string)
    ("Poznámka", 40),          # R — audit_trail.poznamka
]
# Index (1-based) of the first audit-trail column (so per-cell fills can branch)
AUDIT_TRAIL_FIRST_COL = 14  # N

# kapitola → Ceník (KROS convention). Authoritative fallback per task §2.
KAPITOLA_CENIK_MAP: dict[str, str] = {
    "HSV-1": "001",   # zemní práce
    "HSV-2": "011",   # základy a zvláštní zakládání
    "HSV-3": "013",   # ocelové konstrukce
    "HSV-9": "099",   # ostatní HSV
    "PSV-71x": "711", # izolace
    "PSV-76x": "766", # truhlářské/zámečnické
    "PSV-77x": "775", # podlahy
    "PSV-78x": "781", # dokončovací
    "M": "021",       # elektromontážní
    "VRN": "008",     # VRN
    "VZT": "024",     # vzduchotechnika
}

# Ordered kapitola sections (HSV → PSV → M → VRN → VZT)
KAPITOLA_ORDER = ["HSV-1", "HSV-2", "HSV-3", "HSV-9",
                  "PSV-71x", "PSV-76x", "PSV-77x", "PSV-78x",
                  "M", "VRN", "VZT"]

# Status → row fill (per task §2 visual styling)
STATUS_FILLS: dict[str, PatternFill] = {
    "needs_review":   PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid"),
    "matched_medium": PatternFill(start_color="FFFFD0", end_color="FFFFD0", fill_type="solid"),
    "custom_item":    PatternFill(start_color="D0E4FF", end_color="D0E4FF", fill_type="solid"),
    # matched_high → white (default)
}

HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11)
SECTION_FILL = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11, color="FFFFFF")
SUBTOTAL_FILL = PatternFill(start_color="EBEBEB", end_color="EBEBEB", fill_type="solid")
SUBTOTAL_FONT = Font(bold=True, size=10, italic=True)

# Phase 2.1 audit-trail tier fills (override status fill on cols N-R only)
AUDIT_FILL_GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
AUDIT_FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
AUDIT_FILL_RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

# Numeric columns for cell type
COL_QUANTITY = 11   # K
COL_PRICE = 8       # H
COL_TOTAL = 12      # L
COL_CONFIDENCE = 13 # M

# MJ ASCII-normalize map (popis keeps diacritics)
MJ_NORMALIZE = {"m²": "m2", "m³": "m3", "m^2": "m2", "m^3": "m3"}

# Excel formula injection guard (CVE-aware)
FORMULA_PREFIX_RE = re.compile(r"^[=+\-@]")


# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------

def normalize_mj(mj: str | None) -> str:
    if not mj:
        return ""
    s = mj.strip()
    return MJ_NORMALIZE.get(s, s)


def safe_text(text: str | None) -> str:
    """Excel formula injection guard — prefix with apostrophe if starts with = + - @."""
    if not text:
        return ""
    s = str(text)
    if FORMULA_PREFIX_RE.match(s):
        return "'" + s
    return s


def kros_code(item: dict) -> str:
    """Render the Kód položky cell value per status."""
    status = item.get("urs_status")
    code = item.get("urs_code")
    if status == "custom_item":
        return code or ""  # Rpol-NNN kept as-is per interview default
    if status == "needs_review":
        return ""  # blank → Александр fills manually in KROS
    return code or ""


def setup_logging(verbose: bool) -> logging.Logger:
    logger = logging.getLogger("export_kros")
    logger.setLevel(logging.DEBUG if verbose else logging.INFO)
    logger.handlers.clear()
    h = logging.StreamHandler(sys.stdout)
    h.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%H:%M:%S"))
    logger.addHandler(h)
    return logger


# ----------------------------------------------------------------------------
# Main soupis sheet
# ----------------------------------------------------------------------------

def write_header_row(ws: Worksheet, row: int = 1) -> None:
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_section_row(ws: Worksheet, row: int, kapitola: str, count: int) -> None:
    """Bold separator row: kapitola name in D column, span A-M filled."""
    label = f"{kapitola}  ·  Ceník {KAPITOLA_CENIK_MAP.get(kapitola, '???')}  ·  {count} položek"
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = SECTION_FILL
        cell.border = THIN_BORDER
    cell = ws.cell(row=row, column=4, value=label)  # D
    cell.font = SECTION_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")


def audit_tier(at: dict | None, tolerance: float = 0.05) -> str:
    """Map audit_trail block → tier label for fill colour.

    green:  match within tolerance AND confidence ≥ 0.85
    yellow: match within 2× tolerance OR confidence in [0.60, 0.85)
    red:    otherwise (or no audit_trail)
    """
    if not at:
        return "none"
    delta = at.get("match_delta_pct")
    conf = at.get("confidence") or 0.0
    if delta is not None and delta <= tolerance * 100 and conf >= 0.85:
        return "green"
    if (delta is not None and delta <= 2 * tolerance * 100) or 0.60 <= conf < 0.85:
        return "yellow"
    return "red"


AUDIT_FILLS = {
    "green": AUDIT_FILL_GREEN,
    "yellow": AUDIT_FILL_YELLOW,
    "red": AUDIT_FILL_RED,
}


def format_inputs(inputs: list[dict]) -> str:
    if not inputs:
        return ""
    return ", ".join(str(i.get("value", "")) for i in inputs)


def format_references(refs: list[dict]) -> str:
    if not refs:
        return ""
    bits: list[str] = []
    for r in refs:
        t = r.get("type")
        if t == "drawing":
            bits.append(r.get("code", ""))
        elif t == "tz_section":
            bits.append(f"TZ {r.get('section', '')}")
        elif t == "csn":
            bits.append(r.get("standard", ""))
        elif t == "phase_ref":
            bits.append(r.get("section", ""))
    return " · ".join(b for b in bits if b)[:200]


def write_item_row(ws: Worksheet, row: int, item: dict, ceník: str) -> None:
    status = item.get("urs_status", "")
    fill = STATUS_FILLS.get(status)
    at = item.get("audit_trail") or {}
    tier = audit_tier(at) if at else "none"
    audit_fill = AUDIT_FILLS.get(tier)

    values = [
        "",                              # A — O (checkbox)
        ceník,                           # B — Ceník
        "",                              # C — Část (parseable from urs_alternatives later)
        kros_code(item),                 # D — Kód položky
        safe_text(item.get("popis")),    # E — Popis
        normalize_mj(item.get("mj")),    # F — MJ
        "",                              # G — Výrobce
        None,                            # H — Orientační cena (empty)
        "",                              # I — PZN
        "",                              # J — Dodávka
        item.get("mnozstvi") or 0,       # K — Množství
        f"=H{row}*K{row}",               # L — Celkem (formula)
        item.get("urs_match_score") or "",  # M — Confidence (URS match)
        at.get("lokalizace") or "",      # N — Lokalizace (audit_trail)
        safe_text(at.get("formula")) or "",  # O — Výpočet
        format_inputs(at.get("inputs") or []),  # P — Vstupy
        format_references(at.get("reference") or []),  # Q — Reference
        safe_text(at.get("poznamka")) or "",  # R — Poznámka
    ]
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        # Status fill for cols 1..M (1..13). Audit-tier fill for cols N..R (14..18).
        if col_idx >= AUDIT_TRAIL_FIRST_COL and audit_fill is not None:
            cell.fill = audit_fill
        elif fill:
            cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(
            vertical="top",
            wrap_text=(col_idx in (5, 15, 17, 18)),  # wrap Popis, Výpočet, Reference, Poznámka
            horizontal=("right" if col_idx in (COL_QUANTITY, COL_PRICE, COL_TOTAL, COL_CONFIDENCE)
                        else "left"),
        )
    # Number formats
    ws.cell(row=row, column=COL_PRICE).number_format = "#,##0.00"
    ws.cell(row=row, column=COL_QUANTITY).number_format = "#,##0.00"
    ws.cell(row=row, column=COL_TOTAL).number_format = "#,##0.00"
    ws.cell(row=row, column=COL_CONFIDENCE).number_format = "0.00"


def write_subtotal_row(ws: Worksheet, row: int, first_item_row: int, last_item_row: int) -> None:
    """One row per kapitola at the end of its block."""
    if last_item_row < first_item_row:
        return
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = SUBTOTAL_FILL
        cell.font = SUBTOTAL_FONT
        cell.border = THIN_BORDER
    ws.cell(row=row, column=4, value="Mezisoučet").alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=COL_TOTAL, value=f"=SUM(L{first_item_row}:L{last_item_row})")
    ws.cell(row=row, column=COL_TOTAL).number_format = "#,##0.00 Kč"


def write_soupis_sheet(ws: Worksheet, items: list[dict]) -> dict[str, int]:
    """Returns per-kapitola row counts for summary."""
    ws.title = "Soupis prací"
    write_header_row(ws, row=1)

    groups: dict[str, list[dict]] = defaultdict(list)
    for it in items:
        groups[it.get("kapitola") or "(none)"].append(it)

    ordered_keys = [k for k in KAPITOLA_ORDER if k in groups]
    ordered_keys += [k for k in groups if k not in KAPITOLA_ORDER]

    row = 2
    per_kap_rows: dict[str, int] = {}
    for kap in ordered_keys:
        kap_items = groups[kap]
        if not kap_items:
            continue
        ceník = KAPITOLA_CENIK_MAP.get(kap, "")
        write_section_row(ws, row, kap, len(kap_items))
        row += 1
        first_item_row = row
        for it in sorted(kap_items, key=lambda x: x.get("id") or ""):
            write_item_row(ws, row, it, ceník)
            row += 1
        write_subtotal_row(ws, row, first_item_row, row - 1)
        per_kap_rows[kap] = len(kap_items)
        row += 1

    # Grand total (last row)
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    ws.cell(row=row, column=4, value="CELKEM").alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=COL_TOTAL,
            value=f"=SUMIF(D2:D{row - 1},\"Mezisoučet\",L2:L{row - 1})")
    ws.cell(row=row, column=COL_TOTAL).number_format = "#,##0.00 Kč"

    # Frozen pane: header
    ws.freeze_panes = "A2"
    # Autofilter on the data range
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{row}"
    return per_kap_rows


# ----------------------------------------------------------------------------
# Companion 1: needs_review.xlsx — focused for manual top-up
# ----------------------------------------------------------------------------

NEEDS_COLUMNS: list[tuple[str, int]] = [
    ("id", 14),
    ("kapitola", 10),
    ("Ceník", 8),
    ("Popis", 70),
    ("MJ", 8),
    ("Množství", 12),
    ("Alt 1 kód", 14),
    ("Alt 1 popis", 50),
    ("Alt 1 score", 10),
    ("Alt 2 kód", 14),
    ("Alt 2 popis", 50),
    ("Alt 2 score", 10),
    ("Alt 3 kód", 14),
    ("Alt 3 popis", 50),
    ("Alt 3 score", 10),
    ("→ vyplnit ÚRS kód zde", 22),  # blank for manual entry
]


def write_needs_review_sheet(ws: Worksheet, items: list[dict]) -> int:
    ws.title = "Needs review"
    for col_idx, (header, width) in enumerate(NEEDS_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    needs = [it for it in items if it.get("urs_status") == "needs_review"]
    # Sort by quantity desc — highest-impact first
    needs.sort(key=lambda x: -(x.get("mnozstvi") or 0))

    row = 2
    for it in needs:
        alts = it.get("urs_alternatives") or []
        kap = it.get("kapitola") or ""
        values: list[Any] = [
            it.get("id") or "",
            kap,
            KAPITOLA_CENIK_MAP.get(kap, ""),
            safe_text(it.get("popis")),
            normalize_mj(it.get("mj")),
            it.get("mnozstvi") or 0,
        ]
        for i in range(3):
            alt = alts[i] if i < len(alts) else {}
            values.extend([
                alt.get("code") or "",
                safe_text(alt.get("title") or ""),
                alt.get("score") or "",
            ])
        values.append("")  # blank entry column
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.fill = STATUS_FILLS["needs_review"]
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(col_idx in (4, 8, 11, 14)),  # popis + alt popis
                horizontal=("right" if col_idx in (6, 9, 12, 15) else "left"),
            )
        ws.cell(row=row, column=6).number_format = "#,##0.00"
        for c in (9, 12, 15):
            ws.cell(row=row, column=c).number_format = "0.00"
        row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(NEEDS_COLUMNS))}{row - 1}"
    return len(needs)


# ----------------------------------------------------------------------------
# Companion 2: summary.md
# ----------------------------------------------------------------------------

def write_summary_md(path: Path, items: list[dict], main_path: Path,
                      needs_path: Path | None) -> None:
    status_counts = Counter(it.get("urs_status") for it in items)
    n_total = len(items)
    n_custom = status_counts.get("custom_item", 0)
    n_matched = status_counts.get("matched_high", 0) + status_counts.get("matched_medium", 0)
    n_needs = status_counts.get("needs_review", 0)
    rate = n_matched / max(n_total - n_custom, 1) * 100

    per_kap: dict[str, Counter] = defaultdict(Counter)
    per_kap_qty_value: dict[str, float] = defaultdict(float)
    for it in items:
        kap = it.get("kapitola") or "(none)"
        per_kap[kap][it.get("urs_status")] += 1
        per_kap_qty_value[kap] += float(it.get("mnozstvi") or 0)

    lines: list[str] = []
    lines.append("# HK212 — Soupis prací Export Summary\n")
    lines.append(f"_Generated: {datetime.now(timezone.utc).isoformat()}_\n")
    lines.append("## Files\n")
    lines.append(f"- Main: `{main_path.name}` — all {n_total} items")
    if needs_path:
        lines.append(f"- Companion: `{needs_path.name}` — {n_needs} needs_review items "
                     f"(sorted by quantity, with top-3 alternatives + blank fill column)")
    lines.append("")
    lines.append("## Status distribution\n")
    lines.append("| Status | Count | Visual |")
    lines.append("|---|---:|---|")
    lines.append(f"| matched_high   | {status_counts.get('matched_high', 0)} | white |")
    lines.append(f"| matched_medium | {status_counts.get('matched_medium', 0)} | #FFFFD0 light yellow |")
    lines.append(f"| needs_review   | {n_needs} | #FFFF99 yellow (fill in KROS) |")
    lines.append(f"| custom_item    | {n_custom} | #D0E4FF light blue (Rpol-*) |")
    lines.append(f"| **Total**      | **{n_total}** | |")
    lines.append("")
    lines.append(f"**Match rate (excl. custom): {rate:.1f} %**\n")

    lines.append("## Per-kapitola\n")
    lines.append("| Kapitola | Ceník | Total | high | medium | needs_review | custom | Σ množství |")
    lines.append("|---|---|---:|---:|---:|---:|---:|---:|")
    ordered = [k for k in KAPITOLA_ORDER if k in per_kap] + \
              [k for k in per_kap if k not in KAPITOLA_ORDER]
    for k in ordered:
        c = per_kap[k]
        total = sum(c.values())
        lines.append(
            f"| {k} | {KAPITOLA_CENIK_MAP.get(k, '???')} | {total} | "
            f"{c.get('matched_high', 0)} | {c.get('matched_medium', 0)} | "
            f"{c.get('needs_review', 0)} | {c.get('custom_item', 0)} | "
            f"{per_kap_qty_value[k]:,.2f} |"
        )
    lines.append("")
    lines.append("## Workflow — manual top-up in KROS\n")
    lines.append(f"1. Otevři `{main_path.name}` v Excelu vedle KROS.")
    if needs_path:
        lines.append(f"2. Otevři `{needs_path.name}` — to je tvůj per-item TODO list (sortovaný podle "
                     "množství, top-3 alternativy z URS cache na jedné řádce).")
    lines.append("3. Pro každou žlutě podbarvenou položku v KROS: Ctrl+F popis nebo "
                 "část keyword → najdi kód → zkopíruj do sloupce **Kód položky** v Excelu.")
    lines.append(f"4. Estimate: 1-2 min/položka × {n_needs} = ~{n_needs * 1.5 / 60:.1f} h soustředěné práce.")
    lines.append("5. Po doplnění: re-save .xlsx + commit do repo.\n")

    path.write_text("\n".join(lines), encoding="utf-8")


# ----------------------------------------------------------------------------
# Companion 3 (Phase 2.1): audit_trail_review.xlsx — only yellow + red rows
# ----------------------------------------------------------------------------

AUDIT_REVIEW_COLUMNS: list[tuple[str, int]] = [
    ("id", 14),
    ("kapitola", 10),
    ("popis", 60),
    ("mj", 8),
    ("declared", 12),
    ("computed", 12),
    ("Δ%", 8),
    ("confidence", 11),
    ("method", 28),
    ("lokalizace", 30),
    ("formula", 50),
    ("reference", 30),
    ("tier", 8),
]


def write_audit_review_sheet(ws: Worksheet, items: list[dict]) -> int:
    ws.title = "Audit trail review"
    for col_idx, (header, width) in enumerate(AUDIT_REVIEW_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    review_items: list[tuple[str, dict]] = []
    for it in items:
        at = it.get("audit_trail")
        if not at:
            continue
        tier = audit_tier(at)
        if tier in ("yellow", "red"):
            review_items.append((tier, it))

    # Sort: red first, then yellow; within each, sort by absolute Δ%
    review_items.sort(key=lambda t: (
        0 if t[0] == "red" else 1,
        -(t[1]["audit_trail"].get("match_delta_pct") or 0),
    ))

    row = 2
    for tier, it in review_items:
        at = it["audit_trail"]
        fill = AUDIT_FILLS.get(tier)
        values = [
            it.get("id", ""),
            it.get("kapitola", ""),
            safe_text(it.get("popis")),
            normalize_mj(it.get("mj")),
            at.get("declared_quantity"),
            at.get("computed_quantity"),
            at.get("match_delta_pct"),
            at.get("confidence"),
            at.get("extraction_method", ""),
            at.get("lokalizace", ""),
            safe_text(at.get("formula")),
            format_references(at.get("reference") or []),
            tier,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(col_idx in (3, 10, 11, 12)),
                horizontal=("right" if col_idx in (5, 6, 7, 8) else "left"),
            )
        # Number formats
        for c in (5, 6):
            ws.cell(row=row, column=c).number_format = "#,##0.00"
        ws.cell(row=row, column=7).number_format = "0.0\"%\""
        ws.cell(row=row, column=8).number_format = "0.00"
        row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(AUDIT_REVIEW_COLUMNS))}{row - 1}"
    return len(review_items)


# ----------------------------------------------------------------------------
# Driver
# ----------------------------------------------------------------------------

def run(args, logger: logging.Logger) -> int:
    items_path: Path = args.items
    out_dir: Path = args.output_dir

    if not items_path.exists():
        logger.error(f"Items file not found: {items_path}")
        return 2

    with open(items_path, encoding="utf-8") as f:
        wrapper = json.load(f)
    items = wrapper if isinstance(wrapper, list) else wrapper.get("items", [])
    if not items:
        logger.error("No items in input")
        return 3

    n = len(items)
    status_counts = Counter(it.get("urs_status") for it in items)
    n_custom = status_counts.get("custom_item", 0)
    n_matched = status_counts.get("matched_high", 0) + status_counts.get("matched_medium", 0)
    rate = n_matched / max(n - n_custom, 1) * 100
    logger.info(f"Loaded {n} items from {items_path}")
    logger.info(f"Status: {dict(status_counts)}")
    logger.info(f"Match rate (excl. custom): {rate:.1f} %")

    if args.dry_run:
        logger.info("--dry-run — would write main + companions, exiting")
        return 0

    out_dir.mkdir(parents=True, exist_ok=True)

    # Main xlsx
    main_path = out_dir / "HK212_Soupis_praci.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    per_kap = write_soupis_sheet(ws, items)
    wb.save(main_path)
    logger.info(f"Wrote {main_path}")
    logger.info(f"  per kapitola: {dict(sorted(per_kap.items(), key=lambda x: KAPITOLA_ORDER.index(x[0]) if x[0] in KAPITOLA_ORDER else 99))}")

    # Companion: needs_review
    needs_path: Path | None = None
    if not args.no_companion:
        needs_path = out_dir / "HK212_needs_review.xlsx"
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        if ws2 is None:
            ws2 = wb2.create_sheet()
        n_needs = write_needs_review_sheet(ws2, items)
        wb2.save(needs_path)
        logger.info(f"Wrote {needs_path} ({n_needs} needs_review items)")

    # Companion (Phase 2.1): audit_trail_review.xlsx — only yellow + red rows
    has_audit = any(it.get("audit_trail") for it in items)
    if has_audit and not args.no_companion:
        audit_review_path = out_dir / "HK212_audit_trail_review.xlsx"
        wb3 = openpyxl.Workbook()
        ws3 = wb3.active
        if ws3 is None:
            ws3 = wb3.create_sheet()
        n_audit = write_audit_review_sheet(ws3, items)
        wb3.save(audit_review_path)
        logger.info(f"Wrote {audit_review_path} ({n_audit} yellow + red audit-trail rows)")

    # Companion: summary.md
    if not args.no_companion:
        summary_path = out_dir / "HK212_summary.md"
        write_summary_md(summary_path, items, main_path, needs_path)
        logger.info(f"Wrote {summary_path}")

    return 0


# ----------------------------------------------------------------------------
# CLI
# ----------------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser(description="hk212 Phase 2 — KROS soupis prací xlsx export")
    ap.add_argument("--items", type=Path, default=DEFAULT_ITEMS,
                    help=f"Input JSON (default {DEFAULT_ITEMS.name})")
    ap.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    ap.add_argument("--no-companion", action="store_true",
                    help="Skip needs_review.xlsx + summary.md")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--verbose", action="store_true")
    args = ap.parse_args()

    logger = setup_logging(args.verbose)
    logger.info(f"Items:  {args.items}")
    logger.info(f"Output: {args.output_dir}")

    try:
        return run(args, logger)
    except Exception as e:
        logger.exception(f"FATAL: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
