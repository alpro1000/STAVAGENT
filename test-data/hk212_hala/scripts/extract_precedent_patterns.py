"""§13.2 — Precedent pattern extraction.

Reads all 6 hala-typed precedent xlsx files in ``test-data/hk212_hala/example_vv/``
and produces aggregated structural metrics. Privacy: NEVER reads or dumps actual
prices, only structure markers (sheet count, column headers, item counts per
díl, granularity ratios, naming patterns, VRN composition).

Output:
- ``outputs/phase_0b_rerun/example_pattern_analysis.md`` — human-readable
- ``outputs/phase_0b_rerun/example_pattern_analysis.json`` — structured

Run from repo root::

    python3 test-data/hk212_hala/scripts/extract_precedent_patterns.py
"""
from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
import xlrd

REPO_ROOT = Path(__file__).resolve().parents[3]
EX_DIR = REPO_ROOT / "test-data" / "hk212_hala" / "example_vv"
OUT_DIR = REPO_ROOT / "test-data" / "hk212_hala" / "outputs" / "phase_0b_rerun"


# Discriminate per file: HALA precedents vs not-hala-typed (Forestina is sklad)
# Use partial-filename glob to dodge encoding differences (combining diacritic etc).
HALA_PATTERNS = [
    ("ROZMITAL_salt_hala",   "Rozmital"),                # PRIMARY (user-confirmed)
    ("HALA_JHV_deponace",    "HALA JHV"),
    ("KRALOVICE_skolska",    "KRALOVICE"),
    ("ANTRACIT_logistika",   "ANTRACIT"),
    ("SKLAD_SKROBU_P001",    "Slepý položkový rozpočet"),
    ("TREMOSNA_KD",          "Tremosna"),
]

# Czech URS/KROS-style code prefixes that map to HSV/PSV/VRN buckets
URS_BUCKET_RULES: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"^00[1-9]"), "VRN"),       # 00xxx = VRN (zařízení staveniště, geodet, zkoušky)
    (re.compile(r"^1[0-9]"), "HSV-1 Zemní práce"),
    (re.compile(r"^2[0-9]\b|^21-M"), "HSV-2 Základy / M-Elektro"),  # 21-M = M-elektro montáž
    (re.compile(r"^3[0-9]"), "HSV-3 Svislé a kompletní konstrukce"),
    (re.compile(r"^4[0-9]"), "HSV-4 Vodorovné konstrukce"),
    (re.compile(r"^5[0-9]"), "HSV-5 Komunikace"),
    (re.compile(r"^6[0-9]"), "HSV-6 Úpravy povrchů, podlahy"),
    (re.compile(r"^7[0-2]"), "HSV-7 Konstrukce a práce PSV"),
    (re.compile(r"^7[3-4]"), "PSV-73-74 Ústřední vytápění, OTK"),
    (re.compile(r"^75"), "PSV-75 Slaboproudé rozvody"),
    (re.compile(r"^76"), "PSV-76 Konstrukce truhlářské, zámečnické"),
    (re.compile(r"^77"), "PSV-77 Podlahy"),
    (re.compile(r"^78"), "PSV-78 Dokončovací práce"),
    (re.compile(r"^9[0-9]"), "HSV-9 Ostatní práce"),
    (re.compile(r"^Rpol|^R\d"), "Custom Rpol/Specifikace"),
]


def bucket_for_code(code: str) -> str:
    if not code:
        return "—"
    s = str(code).strip()
    for rx, label in URS_BUCKET_RULES:
        if rx.match(s):
            return label
    return "OTHER"


def read_xlsx_sheets(path: Path) -> list[dict]:
    """Read .xlsx — return list of sheet dicts with structured rows."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        out.append({"sheet": sn, "rows": rows})
    wb.close()
    return out


def read_xls_sheets(path: Path) -> list[dict]:
    """Read .xls (old format) — return same structure."""
    book = xlrd.open_workbook(str(path))
    out = []
    for ws in book.sheets():
        rows = []
        for r in range(ws.nrows):
            rows.append([ws.cell_value(r, c) for c in range(ws.ncols)])
        out.append({"sheet": ws.name, "rows": rows})
    return out


def detect_format(rows: list[list]) -> str:
    """Identify file format from first 5 rows."""
    head = " ".join(str(c) for row in rows[:5] for c in row if c)[:500].upper()
    if "#RTSROZP#" in head or "#TYPZAZNAMU#" in head:
        return "RTS_ROZPOCET"
    if "EXPORT KOMPLET" in head or "REKAPITULACE STAVBY" in head:
        return "URS_KROS_KOMPLET"
    if "KRYCÍ LIST SOUPISU PRACÍ" in head.upper().replace("Í", "I"):
        return "URS_KROS_KOMPLET"
    if "URS" in head and "CENA" in head:
        return "URS_KROS"
    if "POLOŽKOV" in head and "ROZPOČ" in head:
        return "POLOZKOVY_ROZPOCET_GENERIC"
    if "VÝKAZ VÝMĚR" in head or "VYKAZ VYMER" in head:
        return "VYKAZ_VYMER"
    return "UNKNOWN"


def extract_items_from_sheet(rows: list[list], fmt: str) -> dict:
    """Find header row, then collect items with code/popis/MJ + bucket."""
    if not rows:
        return {"items": [], "header_row": None, "columns": None}

    # Find header row: contains "Č. položky" / "Kód" / "MJ" / "Množství" / "PČ" / "Pol. č."
    # KROS Komplet headers can live as deep as row 80 (after krycí list + rekapitulace).
    header_idx = None
    header_cells = None
    for i, row in enumerate(rows[:120]):
        joined = " ".join(str(c).lower() for c in row if c is not None)
        if ("kód" in joined or "číslo položky" in joined or "kod" in joined or "pč" in joined) \
                and ("mj" in joined or "množ" in joined):
            header_idx = i
            header_cells = [str(c).strip() if c is not None else "" for c in row]
            break

    if header_idx is None:
        return {"items": [], "header_row": None, "columns": None, "row_count_total": len(rows)}

    items = []
    code_col = None
    popis_col = None
    mj_col = None
    type_col = None
    # Resolve column indices by header name
    for ci, h in enumerate(header_cells):
        hl = h.lower().strip()
        if code_col is None and ("číslo položky" in hl or hl in ("kód", "kod", "code")):
            code_col = ci
        if popis_col is None and ("název položky" in hl or hl == "popis" or "popis položky" in hl):
            popis_col = ci
        if mj_col is None and hl in ("mj", "měrná jednotka", "j.cena"):
            mj_col = ci
        if type_col is None and ("typ položky" in hl or hl.startswith("typ")):
            type_col = ci

    for r in rows[header_idx + 1:]:
        if not r or all(c is None or str(c).strip() == "" for c in r):
            continue
        code = str(r[code_col]).strip() if code_col is not None and code_col < len(r) and r[code_col] is not None else ""
        popis = str(r[popis_col]).strip()[:80] if popis_col is not None and popis_col < len(r) and r[popis_col] is not None else ""
        mj = str(r[mj_col]).strip() if mj_col is not None and mj_col < len(r) and r[mj_col] is not None else ""
        typ = str(r[type_col]).strip() if type_col is not None and type_col < len(r) and r[type_col] is not None else ""
        # Skip type tags that are not real items
        if typ in ("STA", "OBJ", "ROZ", "DIL", "POP", "VV"):
            continue
        if not code and not popis:
            continue
        # Filter out summary rows
        if popis.lower().startswith(("celkem", "součet", "rozpis ceny", "rozpočet")):
            continue
        items.append({
            "code": code,
            "popis": popis,
            "mj": mj,
            "bucket": bucket_for_code(code),
        })

    return {
        "items": items,
        "header_row": header_idx + 1,
        "columns": [h for h in header_cells if h],
        "column_count": sum(1 for h in header_cells if h),
        "row_count_total": len(rows),
    }


def analyze_file(path: Path) -> dict:
    print(f"  analyzing {path.name} ...")
    if path.suffix.lower() == ".xls":
        sheets = read_xls_sheets(path)
    else:
        sheets = read_xlsx_sheets(path)

    fmt = "UNKNOWN"
    sheet_analyses = []
    all_items: list[dict] = []
    for s in sheets:
        rows = s["rows"]
        if not rows:
            continue
        if fmt == "UNKNOWN":
            fmt = detect_format(rows)
        info = extract_items_from_sheet(rows, fmt)
        info["sheet"] = s["sheet"]
        info["item_count"] = len(info.get("items", []))
        # Strip detailed item list (we only care about counts + samples)
        info["sample_items"] = info["items"][:3]
        if info["items"]:
            all_items.extend(info["items"])
        info.pop("items", None)
        sheet_analyses.append(info)

    bucket_counts = Counter(it["bucket"] for it in all_items)
    mj_counts = Counter(it["mj"] for it in all_items if it["mj"])

    # Detect VRN section: count items in VRN bucket
    vrn_items = [it for it in all_items if it["bucket"] == "VRN"]

    # Granularity probe — find items related to typical hala elements
    granularity_probes = {
        "patka_zaklad": [it for it in all_items if re.search(r"\bpatk|zákl|zaklad", it["popis"], re.IGNORECASE)],
        "sloup_ocel": [it for it in all_items if re.search(r"sloup", it["popis"], re.IGNORECASE)],
        "vaznice_pricel": [it for it in all_items if re.search(r"vaznic|příčel|pricel", it["popis"], re.IGNORECASE)],
        "kingspan_oplasten": [it for it in all_items if re.search(r"kingspan|panel|sendvi|opláštění|oplasteni", it["popis"], re.IGNORECASE)],
        "okna_dvere_vrata": [it for it in all_items if re.search(r"okn|dveř|dver|vrata", it["popis"], re.IGNORECASE)],
        "vykop_zem": [it for it in all_items if re.search(r"hloubě|hlouben|výkop|vykop|zemní|zemni", it["popis"], re.IGNORECASE)],
        "beton_konstr": [it for it in all_items if re.search(r"beton|žb|zb\b", it["popis"], re.IGNORECASE)],
    }

    # Sample (privacy-safe): strip popis to first 50 chars; no prices ever recorded
    granularity_samples = {
        k: [{"code": it["code"], "popis": it["popis"][:50], "mj": it["mj"]} for it in vs[:5]]
        for k, vs in granularity_probes.items() if vs
    }

    return {
        "file": path.name,
        "format": fmt,
        "sheet_count": len(sheets),
        "sheet_names": [s["sheet"] for s in sheets],
        "sheet_analyses": sheet_analyses,
        "total_items": len(all_items),
        "bucket_distribution": dict(bucket_counts.most_common()),
        "mj_distribution": dict(mj_counts.most_common(15)),
        "vrn_item_count": len(vrn_items),
        "vrn_samples": [{"code": it["code"], "popis": it["popis"][:60]} for it in vrn_items[:10]],
        "granularity_probes": {k: len(v) for k, v in granularity_probes.items()},
        "granularity_samples": granularity_samples,
    }


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    results = {}
    all_files = sorted(EX_DIR.iterdir())
    for tag, needle in HALA_PATTERNS:
        matches = [p for p in all_files if needle.lower() in p.name.lower() and p.suffix.lower() in (".xls", ".xlsx")]
        if not matches:
            print(f"  ⚠ no file matched pattern '{needle}'")
            continue
        # Prefer .xlsx over .xls if both exist
        matches.sort(key=lambda p: (p.suffix.lower() == ".xls", p.name))
        p = matches[0]
        try:
            results[tag] = analyze_file(p)
            results[tag]["matched_filename"] = p.name
        except Exception as e:
            print(f"  ⚠ error on {p.name}: {type(e).__name__}: {e}")
            results[tag] = {"file": p.name, "error": f"{type(e).__name__}: {e}"}

    out_json = OUT_DIR / "example_pattern_analysis.json"
    out_json.write_text(json.dumps(results, ensure_ascii=False, indent=2))
    print(f"\n  → {out_json.relative_to(REPO_ROOT)}")

    # Aggregate metrics
    primary = results.get("ROZMITAL_salt_hala", {})
    print(f"\n=== ROŽMITÁL (primary precedent) ===")
    print(f"  format: {primary.get('format')}")
    print(f"  sheets: {primary.get('sheet_count')}")
    print(f"  total items: {primary.get('total_items')}")
    print(f"  bucket distribution:")
    for k, v in (primary.get("bucket_distribution") or {}).items():
        print(f"    {k:50s} {v}")
    print(f"  granularity probes:")
    for k, v in (primary.get("granularity_probes") or {}).items():
        print(f"    {k:25s} {v} items")

    print(f"\n=== CROSS-PRECEDENT comparison ===")
    print(f"{'file':75s} {'fmt':25s} {'sheets':>7s} {'items':>6s} {'vrn':>4s}")
    for fname, info in results.items():
        if "error" in info:
            print(f"  {fname[:75]:75s}  ERROR {info['error'][:50]}")
            continue
        print(f"  {fname[:73]:73s} {(info.get('format') or '?')[:25]:25s} {info.get('sheet_count') or 0:7d} {info.get('total_items') or 0:6d} {info.get('vrn_item_count') or 0:4d}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
