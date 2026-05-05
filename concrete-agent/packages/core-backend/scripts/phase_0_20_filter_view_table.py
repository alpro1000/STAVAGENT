"""Phase 0.20 v2 — Replace 12_Filter_view with Excel Table (Cesta 2).

User feedback (RU): "Я думал что агент добавит фильтры макросы чтобы можно
было нажимая на кнопки все фильтровать."

V1 (static stacked tables) replaced by V2 = Excel Table with native
column-header dropdown filters. Each header cell has ▼ icon → user clicks
→ multi-checkbox filter UI. Universal Excel compatibility (2016+, Mac,
Online, iPad).

NO VBA, NO PivotTable, NO Slicers — pure Excel Table feature.

Layout:
  Row 1: Summary metadata (outside table)
  Row 2: Header (table starts here)
  Rows 3-3023: Data (3021 items × 13 cols)

13 columns = 12 source from 1_Vykaz_vymer + 1 derived Podlaží.
"""
from __future__ import annotations

import re
import warnings
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import (CellIsRule, ColorScaleRule, DataBarRule,
                                       FormulaRule)
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Scoped warning suppression — only openpyxl-internal user warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

EXCEL = Path(
    "test-data/libuse/outputs/"
    "Vykaz_vymer_Libuse_objekt_D_dokoncovaci_prace.xlsx"
)
SHEET_SRC = "1_Vykaz_vymer"
SHEET_NEW = "12_Filter_view"


def extract_podlazi(misto) -> str:
    if not isinstance(misto, str):
        return "unknown"
    parts = [p.strip() for p in misto.split("·")]
    return parts[1] if len(parts) >= 2 else "unknown"


def find_col(headers: list, name: str) -> int:
    """Return 1-based column index of header `name`, or raise."""
    try:
        return headers.index(name) + 1
    except ValueError as e:
        raise RuntimeError(
            f"Required column {name!r} missing from source headers: {headers}"
        ) from e


def main() -> None:
    print("=" * 72)
    print("PHASE 0.20 v2 — 12_Filter_view jako Excel Table")
    print("=" * 72)

    wb = load_workbook(EXCEL)
    src_sheets_pre = list(wb.sheetnames)
    print(f"\nSheets pre-edit ({len(src_sheets_pre)}): {src_sheets_pre}")

    # === Read source (read-only) ===
    ws_src = wb[SHEET_SRC]
    n_src_rows = ws_src.max_row
    n_src_cols = ws_src.max_column
    print(f"\nSource {SHEET_SRC}: {n_src_rows} rows × {n_src_cols} cols")

    headers_src = [ws_src.cell(1, c).value for c in range(1, n_src_cols + 1)]
    print(f"  Source headers: {headers_src}")

    # Hard requirements: must have ≥1 data row + required columns by NAME
    REQUIRED_COLS = ["#", "ÚRS kód", "Kapitola", "Popis položky", "MJ",
                     "Množství", "Místo", "Skladba/povrch", "Confidence",
                     "Status", "Poznámka", "Source"]
    missing = [c for c in REQUIRED_COLS if c not in headers_src]
    if missing:
        print(f"⛔ STOP — source missing required columns: {missing}")
        return
    if n_src_rows < 2:
        print(f"⛔ STOP — source has no data rows (only header)")
        return
    # Soft warn if dimensions differ from session baseline (3022×12)
    if n_src_rows != 3022 or n_src_cols != 12:
        print(f"⚠️  Source dimensions {n_src_rows}×{n_src_cols} differ from "
              f"session baseline 3022×12 — proceeding anyway "
              f"(items count is variable across phases).")

    # Resolve column indices by NAME (not hardcoded position)
    misto_col = find_col(headers_src, "Místo")
    status_col = find_col(headers_src, "Status")
    confidence_col = find_col(headers_src, "Confidence")
    print(f"  Resolved cols: Místo={misto_col}, Status={status_col}, "
          f"Confidence={confidence_col}")

    # === Delete existing 12_Filter_view (V1 static) ===
    if SHEET_NEW in wb.sheetnames:
        print(f"\nRemoving existing {SHEET_NEW} (V1 static)…")
        del wb[SHEET_NEW]

    # === Create new 12_Filter_view ===
    ws = wb.create_sheet(SHEET_NEW)

    # ROW 1: Summary metadata (outside table)
    ws.cell(1, 1, "Total items:").font = Font(bold=True, size=11)
    ws.cell(1, 2, n_src_rows - 1).font = Font(size=11)
    ws.cell(1, 4, "Klikněte ▼ v každém sloupci pro filtrování / multi-select").font = (
        Font(italic=True, color="555555", size=10))
    ws.cell(1, 11, "Source: '1_Vykaz_vymer' read-only").font = (
        Font(italic=True, color="888888", size=9))
    ws.row_dimensions[1].height = 22

    # ROW 2: Headers (start of Excel Table)
    HEADER_ROW = 2
    headers_all = headers_src + ["Podlaží"]  # 13 columns
    for c, h in enumerate(headers_all, 1):
        cell = ws.cell(HEADER_ROW, c, h)
        # Note: when wrapped in Table, openpyxl will apply table style
        # automatically — we don't manually style header here

    # ROWS 3+ : Data
    DATA_START = HEADER_ROW + 1
    podlazi_col = n_src_cols + 1  # derived column = last+1
    for src_row in range(2, n_src_rows + 1):  # source rows 2..N
        target_row = DATA_START + (src_row - 2)
        for c in range(1, n_src_cols + 1):
            v = ws_src.cell(src_row, c).value
            ws.cell(target_row, c, v)
        # Derived Podlaží via name-resolved column
        misto_val = ws_src.cell(src_row, misto_col).value
        ws.cell(target_row, podlazi_col, extract_podlazi(misto_val))

    last_data_row = DATA_START + (n_src_rows - 2)
    n_total_cols = n_src_cols + 1  # source + Podlaží
    last_col_letter = get_column_letter(n_total_cols)
    table_ref = f"A{HEADER_ROW}:{last_col_letter}{last_data_row}"
    print(f"\nTable ref: {table_ref} (header row {HEADER_ROW}, "
          f"data rows {DATA_START}..{last_data_row}, total "
          f"{last_data_row - HEADER_ROW + 1} rows × {n_total_cols} cols)")

    # === Convert to Excel Table (auto-filter enabled by default) ===
    table = Table(displayName="VykazFilter", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)
    print(f"✓ Excel Table 'VykazFilter' added with TableStyleMedium9")

    # === Column widths ===
    widths = [6, 12, 10, 50, 6, 11, 22, 32, 12, 18, 25, 18, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # === Freeze pane below header (so header sticks during scroll) ===
    ws.freeze_panes = f"A{DATA_START}"  # = A3

    # === Conditional formatting ===
    # Status column — resolved by name (not hardcoded J)
    status_letter = get_column_letter(status_col)
    status_range = f"{status_letter}{DATA_START}:{status_letter}{last_data_row}"
    # Conditional formatting formulas use RELATIVE row reference so each
    # row evaluates its own Status cell. Range top-left = {status_letter}{DATA_START};
    # Excel offsets the formula per cell (Qodo + Amazon Q review fix).
    sl = status_letter  # alias for f-string brevity
    ds = DATA_START
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=[f'OR({sl}{ds}="matched_high",{sl}{ds}="matched_medium")'],
            fill=PatternFill("solid", fgColor="C6EFCE"),
            font=Font(color="006100")))
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=[f'OR({sl}{ds}="needs_review",{sl}{ds}="OPRAVENO_OBJEM",'
                     f'{sl}{ds}="OPRAVENO_POPIS")'],
            fill=PatternFill("solid", fgColor="FFEB9C"),
            font=Font(color="9C5700")))
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=[f'OR({sl}{ds}="no_match",{sl}{ds}="VYNECHANE_KRITICKE")'],
            fill=PatternFill("solid", fgColor="FFC7CE"),
            font=Font(color="9C0006")))
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=[f'OR({sl}{ds}="VYNECHANE_DETAIL",{sl}{ds}="deprecated")'],
            fill=PatternFill("solid", fgColor="D9D9D9"),
            font=Font(color="595959")))
    print(f"✓ Conditional formatting on Status column ({status_range})")

    # Confidence column — resolved by name
    confidence_letter = get_column_letter(confidence_col)
    confidence_range = (f"{confidence_letter}{DATA_START}:"
                         f"{confidence_letter}{last_data_row}")
    ws.conditional_formatting.add(
        confidence_range,
        DataBarRule(
            start_type="num", start_value=0,
            end_type="num", end_value=1,
            color="638EC6", showValue=True))
    print(f"✓ Data bar on Confidence column ({confidence_range})")

    # === Save ===
    print("\nSaving Excel…")
    wb.save(EXCEL)
    print(f"  Saved {EXCEL}")

    # === FINAL AUDIT ===
    print("\n=== FINAL AUDIT ===")
    wb2 = load_workbook(EXCEL, data_only=True)
    sheets_post = list(wb2.sheetnames)
    print(f"  Sheets ({len(sheets_post)}): {sheets_post}")

    # Verify source unchanged (dimensions match what we read pre-edit)
    ws_src2 = wb2[SHEET_SRC]
    if ws_src2.max_row != n_src_rows or ws_src2.max_column != n_src_cols:
        print(f"⛔ ERROR: source modified "
              f"({ws_src2.max_row}×{ws_src2.max_column} vs "
              f"pre-edit {n_src_rows}×{n_src_cols})")
        return

    # Verify new sheet structure (computed from source dims)
    ws_new2 = wb2[SHEET_NEW]
    print(f"  {SHEET_NEW}: {ws_new2.max_row} rows × {ws_new2.max_column} cols")
    expected_rows = HEADER_ROW + (n_src_rows - 1)  # metadata + header + data
    expected_cols = n_src_cols + 1  # source + Podlaží
    if ws_new2.max_row != expected_rows:
        print(f"⛔ ERROR: new sheet rows {ws_new2.max_row} != {expected_rows}")
        return
    if ws_new2.max_column != expected_cols:
        print(f"⛔ ERROR: new sheet cols {ws_new2.max_column} != {expected_cols}")
        return

    # Verify table exists
    tables = list(ws_new2.tables.keys())
    print(f"  Tables in {SHEET_NEW}: {tables}")
    if "VykazFilter" not in tables:
        print(f"⛔ ERROR: VykazFilter table not registered")
        return

    # Verify autofilter (Excel Table auto-includes it; .ref may be None
    # if openpyxl didn't reflect Table-internal filter to sheet-level —
    # that is fine; functional filter is in Table XML)
    af = ws_new2.auto_filter
    if af is not None and af.ref:
        print(f"  AutoFilter ref (sheet level): {af.ref}")
    else:
        print(f"  AutoFilter ref (sheet level): None — filter lives "
              f"inside Table XML (Excel-side dropdowns OK)")

    # Spot-check 3 random data rows match source (cols by name)
    import random
    spot_cols = []
    for name in ("#", "Popis položky", "Množství", "Místo", "Status"):
        if name in headers_src:
            spot_cols.append(headers_src.index(name) + 1)
    for src_row in random.sample(range(2, n_src_rows + 1), min(3, n_src_rows - 1)):
        target_row = DATA_START + (src_row - 2)
        for c in spot_cols:
            src_val = ws_src2.cell(src_row, c).value
            new_val = ws_new2.cell(target_row, c).value
            if src_val != new_val:
                print(f"⛔ ERROR: data mismatch row {src_row} col {c}: "
                      f"src={src_val!r} vs new={new_val!r}")
                return
        misto = ws_src2.cell(src_row, misto_col).value
        derived = extract_podlazi(misto)
        new_pod = ws_new2.cell(target_row, podlazi_col).value
        if derived != new_pod:
            print(f"⛔ ERROR: Podlaží mismatch row {src_row}: "
                  f"derived={derived!r} vs stored={new_pod!r}")
            return

    print(f"✓ Spot-check 3 rows × {len(spot_cols)} fields all match source")
    print("\n✅ All checks pass.")


if __name__ == "__main__":
    main()
