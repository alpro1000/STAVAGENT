"""Tabulka 0041 (TABULKA DVEŘÍ) — full row absorption.

Step 3 of Π.0a (TASK_PHASE_PI_0_SPEC.md §5 step 3 — gap G1 + PROBE 7
closure). Lifts all 28 columns of the doors table into the canonical
master_extract.json `doors[]` section.

Before this step, only 6 of 28 columns made it into the pipeline (via
phase_0_9_extract_doors_ownership.py): cislo, typ, from_room, to_room,
sirka_otvoru_mm, vyska_otvoru_mm + a derived `is_garage_gate` flag.
The other 22 columns (RC class, Rw acoustic, fire rating, EMZ lock,
ACS access control, kování, samozavírač, etc.) were dropped — that's
the dropped data that surfaced as PROBE 7 (D10 underspec).

Schema follows TASK_PHASE_PI_0_SPEC.md §2.6: each cell becomes
`{value, source, confidence: 1.0}`. Every door row carries:
- 28 lifted cell fields
- derived `is_garage_gate` (from cislo + width)
- `objekt_filter` (which objekty this door belongs to — determined by
  from_room / to_room prefix; comma-separated for cross-objekt rows
  e.g. "C,D" for the 5 doors that are placed at the C–D boundary).
"""
from __future__ import annotations

import openpyxl
from pathlib import Path

# ---------------------------------------------------------------------------
# Column → schema-field map
# ---------------------------------------------------------------------------
# (col_index, schema_field, comment) — col 1-based per openpyxl convention.
COLUMN_MAP: list[tuple[int, str, str]] = [
    (1,  "cislo",                       "door instance number"),
    (2,  "typ",                         "type marker D## / W##"),
    (3,  "from_room",                   "z místnosti"),
    (4,  "to_room",                     "do místnosti"),
    (5,  "otvirani",                    "L (left) / P (right) / similar"),
    (6,  "pocet_kridel",                "No. of leaves: 1 / 2"),
    (7,  "celkova_svetla_sirka_mm",     "total clear width"),
    (8,  "sirka_aktivniho_kridla_mm",   "active leaf width"),
    (9,  "svetla_vyska_mm",             "clear height"),
    (10, "popis_kridla",                "leaf description code"),
    (11, "ral_kridla",                  "leaf RAL color"),
    (12, "sirka_otvoru_mm",             "rough opening width"),
    (13, "vyska_otvoru_mm",             "rough opening height"),
    (14, "popis_zarubne",               "frame description code (Z##)"),
    (15, "ral_zarubne",                 "frame RAL color"),
    (16, "pozarni_odolnost",            "fire resistance (e.g. EI 30 D3)"),
    (17, "rw_lab_db",                   "laboratory acoustic Rw, dB"),
    (18, "rw_site_db",                  "on-site acoustic R'w, dB"),
    (19, "bezpecn_odolnost",            "RC class (RC1/RC2/RC3 + ESG) — PROBE 7"),
    (20, "tepelne_vlastnosti_u",        "thermal U-value, W/m²·K"),
    (21, "typ_kovani",                  "hardware type (KP1, MM, etc.)"),
    (22, "typ_samozavirace",            "door closer type (SN1/SN2)"),
    (23, "zamek",                       "lock type (EMZ — PROBE 7)"),
    (24, "doplnky",                     "supplements (Z2, KL, CT1, …)"),
    (25, "eps_fas",                     "EPS / FAS (fire alarm) flag"),
    (26, "acs",                         "ACS / Access control (●) — PROBE 7"),
    (27, "vzt",                         "ventilation flag"),
    (28, "poznamka",                    "free-text note"),
]

# Header row 6, data rows 7..297. Data column count = 28 (rest are empty).
HEADER_ROW = 6
DATA_START_ROW = 7
SHEET_NAME = "tab dvere"

# is_garage_gate rule (per phase_0_9_extract_doors_ownership.py):
#   typ == 'D05' OR rough opening width > 3000 mm
GARAGE_GATE_TYPES = {"D05"}
GARAGE_GATE_WIDTH_THRESHOLD_MM = 3000

# Objekt prefix detection — how a from_room / to_room maps to objekt
OBJEKT_PREFIXES = {
    "A": ("A.", "S.A."),
    "B": ("B.", "S.B."),
    "C": ("C.", "S.C."),
    "D": ("D.", "S.D."),
}


def _detect_objekty(from_room: str | None, to_room: str | None) -> set[str]:
    """Which objekt(s) does this door belong to? Either side counts."""
    found: set[str] = set()
    for room in (from_room or "", to_room or ""):
        room = str(room).strip()
        if not room:
            continue
        for objekt, prefixes in OBJEKT_PREFIXES.items():
            if any(room.startswith(p) for p in prefixes):
                found.add(objekt)
                break
    return found


def _is_garage_gate(typ: str | None, sirka_otvoru_mm) -> bool:
    if typ and str(typ).strip().upper() in GARAGE_GATE_TYPES:
        return True
    if isinstance(sirka_otvoru_mm, (int, float)) and sirka_otvoru_mm > GARAGE_GATE_WIDTH_THRESHOLD_MM:
        return True
    return False


def _coerce(value):
    """Strip strings; pass numbers / None through. ``''`` → None."""
    if isinstance(value, str):
        v = value.strip()
        return v if v else None
    return value


def _wrap_cell(value, src: str) -> dict:
    """Wrap a literal Tabulka cell as `{value, source, confidence: 1.0}`."""
    return {"value": _coerce(value), "source": src, "confidence": 1.0}


def _wrap_derived(value, src: str, confidence: float = 0.95) -> dict:
    """Wrap a derived value with given confidence."""
    return {"value": value, "source": src, "confidence": confidence}


def extract_doors_for_objekt(tabulka_dveri_xlsx: Path, objekt: str) -> list[dict]:
    """Return all Tabulka 0041 rows that belong to `objekt` (A/B/C/D).

    Cross-objekt doors (e.g. C–D shared boundary) appear in BOTH
    objekty's lists — the `objekt_filter.value` field tells consumers
    which sides the door spans.

    Each entry is a dict matching TASK_PHASE_PI_0_SPEC.md §2.6 schema:

        {
          "id": str,                # composite "{objekt}.row{N}.{cislo}.{typ}"
          "tabulka_dveri_row": <wrapped int>,
          "objekt_filter": <wrapped "A"|"A,C"|...>,
          "is_garage_gate": <wrapped bool>,
          "<col_field>": <wrapped value>,   # for each of the 28 cols
        }

    Confidence: 1.0 for every literal cell (XLSX is the source of truth).
    `is_garage_gate` is derived = 0.95.
    """
    if objekt not in OBJEKT_PREFIXES:
        raise ValueError(f"objekt must be one of A/B/C/D, got {objekt!r}")
    if not tabulka_dveri_xlsx.exists():
        return []

    wb = openpyxl.load_workbook(tabulka_dveri_xlsx, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        return []
    ws = wb[SHEET_NAME]

    xlsx_rel = "shared/xlsx/" + tabulka_dveri_xlsx.name
    out: list[dict] = []

    for r, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True),
                             start=DATA_START_ROW):
        # row is a tuple of cell values; index 0 = col 1 = cislo
        if not row or len(row) < 4:
            continue
        cislo = _coerce(row[0])
        typ = _coerce(row[1])
        from_room = _coerce(row[2])
        to_room = _coerce(row[3])
        if not typ:
            # blank row in the middle of the table — skip
            continue

        belongs = _detect_objekty(from_room, to_room)
        if objekt not in belongs:
            continue

        # Lift all 28 columns
        entry: dict = {
            "id": f"{objekt}.row{r}.{cislo or '?'}.{typ}",
            "tabulka_dveri_row": _wrap_cell(r, f"XLSX|{xlsx_rel}|{SHEET_NAME}|row={r}"),
            "objekt_filter": _wrap_derived(",".join(sorted(belongs)),
                                           "DERIVED|from_room+to_room"),
            "is_garage_gate": _wrap_derived(_is_garage_gate(typ, _coerce(row[11])),
                                             "DERIVED|typ+sirka_otvoru_mm"),
        }
        for col_idx, field, _ in COLUMN_MAP:
            value = row[col_idx - 1] if (col_idx - 1) < len(row) else None
            src = f"XLSX|{xlsx_rel}|{SHEET_NAME}|row={r},col={col_idx}"
            entry[field] = _wrap_cell(value, src)

        out.append(entry)

    wb.close()
    return out
