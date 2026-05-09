"""Tabulka 0030 (TABULKA SKLADEB A POVRCHŮ) — full skladba absorption.

Step 4 of Π.0a (TASK_PHASE_PI_0_SPEC.md §5 step 4 — gap G6). Extends
the legacy `phase_0_8_extract_master_skladby.py` (which only handled
WF) to all 5 skladba kinds present in Tabulka 0030:

    F   — povrchy (single-layer surface finishes)
    FF  — skladby podlah (floor compositions)
    WF  — skladby stěn (wall compositions)
    RF  — skladby střech (roof compositions)
    CF  — podhledy (ceiling compositions)

Schema follows TASK_PHASE_PI_0_SPEC.md §2.5: skladby[<kind>][<code>] →
{kind, label, vrstvy[], celkova_tloustka_mm, source, confidence}.
The whole entry shares one source + confidence (1.0 — literal XLSX).

Note: TP / OP / LI / LP catalogues live in OTHER tables (0050
zámečnické, 0060 klempířské, 0080 ostatní) and are absorbed in
later steps.

Tabulka 0030 is **komplex-wide** (skladby don't differ per objekt
A/B/C/D — same wall composition standards apply to all). A single
extraction populates the same skladby{} for all 4 objekty.
"""
from __future__ import annotations

import openpyxl
from pathlib import Path

# ---------------------------------------------------------------------------
# Sheet → kind mapping
# ---------------------------------------------------------------------------
# (sheet_name, kind, has_celkova_col)
# Note: 'povrchy' sheet has 1 fewer column (no celková tl.).
SHEET_MAP: list[tuple[str, str, bool]] = [
    ("povrchy",        "F",  False),
    ("skladby_podlah", "FF", True),
    ("skladby sten",   "WF", True),
    ("skladby strech", "RF", True),
    ("podhledy",       "CF", True),
]

HEADER_ROW = 4
DATA_START_ROW = 5


# ---------------------------------------------------------------------------
# Layer schema — per row in the sheet, columns map to vrstva fields
# ---------------------------------------------------------------------------
# povrchy layout (7 cols, no celková):
#   col1=Kód, col2=order, col3=layer_name, col4=tloustka, col5=tech_spec, col6=produkt_ref
# Other sheets (8 cols):
#   col1=Kód, col2=order, col3=layer_name, col4=tloustka, col5=celkova, col6=tech_spec, col7=produkt_ref


def _coerce_str(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _normalize_code(kind: str, raw_order) -> str:
    """Build skladba code from kind + raw 'pořadí' value.

    F + '0' → F00; F + '1' → F01; WF + '01' → WF01; WF + '2' → WF02;
    RF + '10' → RF10. Always 2-digit zero-padded suffix.
    """
    if raw_order is None:
        return kind  # fallback — should not happen for real rows
    s = str(raw_order).strip()
    # Numeric? zero-pad to 2 digits.
    try:
        return f"{kind}{int(float(s)):02d}"
    except (ValueError, TypeError):
        return f"{kind}{s}"


def _extract_skladby_from_sheet(ws, kind: str, has_celkova: bool,
                                  source_prefix: str) -> dict[str, dict]:
    """Parse one sheet → dict[skladba_code → entry].

    Algorithm:
      - Track the most recent "label row" (col1 has text, col2 empty,
        no Kód) — describes the next skladba.
      - When a row has Kód = `kind` (e.g. "FF") and col2 has the order
        marker, start a new skladba at that row.
      - All subsequent rows (until the next code row or label row) are
        continuation layers.
    """
    skladby: dict[str, dict] = {}
    cur_label: str | None = None
    cur_entry: dict | None = None
    cur_first_row: int | None = None

    # Column indices (1-based)
    if has_celkova:
        col_idx = {"kod": 1, "order": 2, "layer_name": 3, "tloustka": 4,
                   "celkova": 5, "tech_spec": 6, "produkt_ref": 7}
    else:
        col_idx = {"kod": 1, "order": 2, "layer_name": 3, "tloustka": 4,
                   "celkova": None, "tech_spec": 5, "produkt_ref": 6}

    def _layer_from_row(r) -> dict | None:
        layer_name = _coerce_str(ws.cell(r, col_idx["layer_name"]).value)
        if not layer_name:
            return None
        tloustka_raw = ws.cell(r, col_idx["tloustka"]).value
        try:
            tloustka_mm = float(tloustka_raw) if tloustka_raw not in (None, "", "-") else None
        except (ValueError, TypeError):
            tloustka_mm = None
        return {
            "order": ws.cell(r, col_idx["order"]).value,  # may be None for continuation
            "label": layer_name,
            "tloustka_mm": tloustka_mm,
            "technicka_specifikace": _coerce_str(ws.cell(r, col_idx["tech_spec"]).value),
            "produkt_ref": _coerce_str(ws.cell(r, col_idx["produkt_ref"]).value),
        }

    for r in range(DATA_START_ROW, ws.max_row + 1):
        kod_cell = _coerce_str(ws.cell(r, col_idx["kod"]).value)
        order_cell = ws.cell(r, col_idx["order"]).value

        # CASE A: label row (col1 has descriptive text, col2 empty,
        # not the kind code itself)
        if kod_cell and kod_cell != kind and order_cell in (None, ""):
            cur_label = kod_cell
            continue

        # CASE B: code row (col1 = kind, col2 = order number)
        if kod_cell == kind and order_cell not in (None, ""):
            code = _normalize_code(kind, order_cell)
            celkova_mm = None
            if has_celkova and ws.cell(r, col_idx["celkova"]).value not in (None, ""):
                try:
                    celkova_mm = float(ws.cell(r, col_idx["celkova"]).value)
                except (ValueError, TypeError):
                    celkova_mm = None
            layer = _layer_from_row(r)
            cur_entry = {
                "code": code,
                "kind": kind,
                "label": cur_label,
                "vrstvy": [layer] if layer else [],
                "celkova_tloustka_mm": celkova_mm,
                "source": f"{source_prefix}|row={r}",
                "confidence": 1.0,
            }
            cur_first_row = r
            skladby[code] = cur_entry
            continue

        # CASE C: continuation layer (col1 empty, col2 empty/order, has layer_name)
        if cur_entry is not None and not kod_cell:
            layer = _layer_from_row(r)
            if layer:
                cur_entry["vrstvy"].append(layer)
                continue

        # CASE D: blank row or end of skladba — reset
        # (label may persist for the next skladba in same row group)

    return skladby


def extract_skladby(tabulka_skladeb_xlsx: Path) -> dict[str, dict[str, dict]]:
    """Read all 5 sheets of Tabulka 0030; return nested skladby{}.

    Output shape:
        {
          "F":  {"F00": {...}, "F01": {...}, ...},
          "FF": {"FF01": {...}, ...},
          "WF": {"WF01": {...}, ...},
          "RF": {"RF10": {...}, ...},
          "CF": {"CF10": {...}, ...},
        }

    Each entry is a dict with bare fields (kind, label, vrstvy[],
    celkova_tloustka_mm, source, confidence: 1.0) per SPEC §2.5.

    Komplex-wide: same skladby{} is appropriate for any of A/B/C/D.
    """
    if not tabulka_skladeb_xlsx.exists():
        return {}
    wb = openpyxl.load_workbook(tabulka_skladeb_xlsx, data_only=True, read_only=True)
    out: dict[str, dict[str, dict]] = {}
    rel = "shared/xlsx/" + tabulka_skladeb_xlsx.name

    for sheet_name, kind, has_celkova in SHEET_MAP:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        source_prefix = f"XLSX|{rel}|{sheet_name}"
        out[kind] = _extract_skladby_from_sheet(ws, kind, has_celkova, source_prefix)

    wb.close()
    return out
