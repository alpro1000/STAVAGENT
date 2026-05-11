"""Tabulka 0042 (TABULKA OKEN) — full row absorption.

Step 7 of Π.0a (TASK_PHASE_PI_0_SPEC.md §5 step 7 — gap G2). Closes
the 0% extraction gap on the windows table. Lifts all 20 columns
into `windows[]`:

    Označení / Type Mark, Typ okna, Šířka × Výška, Otvírání, Zasklení,
    Typ kování, Těsnění, Parapet vnitřní + vnější, Doplňky,
    Uw (thermal), SF g-value (solar gain), LT (light transmission),
    R'w (acoustic on-site), Bezpečnostní odolnost (RC class),
    EPS / FAS (fire alarm), ACS (access control), Poznámka, Počet.

Note: unlike Tabulka 0041 (one row per door instance with from_room /
to_room), Tabulka 0042 is a CATALOGUE of window types. Each W##
code appears once with full specs. Pi_0a includes the full catalogue
in every objekt's windows[] section (komplex-wide reference). Per-
instance window placement comes from DXF openings[] cross-linking.
"""
from __future__ import annotations

import openpyxl
from pathlib import Path

# ---------------------------------------------------------------------------
# Column → schema-field map
# ---------------------------------------------------------------------------
COLUMN_MAP: list[tuple[int, str, str]] = [
    (1,  "kod",                         "Označení / Type Mark (W01, W02, ...)"),
    (2,  "typ_okna",                    "Typ okna (AL1, AL2, ...)"),
    (3,  "sirka_mm",                    "Šířka mm"),
    (4,  "vyska_mm",                    "Výška mm"),
    (5,  "otvirani",                    "Otvírání code (O1, O2, ...)"),
    (6,  "zasklenie",                   "Zasklení glazing code (S1, S2, ...)"),
    (7,  "typ_kovani",                  "Typ kování (K1, K2, ...)"),
    (8,  "tesnenie",                    "Těsnění code (T1, T2, ...)"),
    (9,  "parapet_vnitrni",             "Parapet vnitřní (IP1, ...)"),
    (10, "parapet_vnejsi",              "Parapet vnější (EP1, ...)"),
    (11, "doplnky",                     "Doplňky / supplements"),
    (12, "uw",                          "Uw thermal W/m²·K"),
    (13, "sf_g_value",                  "SF / g-value solar gain factor"),
    (14, "lt_transmission",             "LT solar/light transmission"),
    (15, "rw_site_db",                  "R'w on-site acoustic dB"),
    (16, "bezpecn_odolnost",            "Bezpečnostní RC class (RC2, RC3, ...)"),
    (17, "eps_fas",                     "EPS / FAS fire alarm flag"),
    (18, "acs",                         "ACS access control flag"),
    (19, "poznamka",                    "Note / poznámka"),
    (20, "pocet",                       "Total instance count across komplex"),
]

HEADER_ROW = 6
DATA_START_ROW = 8  # row 7 is units (mm, %, etc.)
SHEET_NAME = "tabulka"


def _coerce(value):
    if isinstance(value, str):
        v = value.strip()
        return v if v else None
    return value


def _wrap_cell(value, src: str) -> dict:
    return {"value": _coerce(value), "source": src, "confidence": 1.0}


def extract_windows(tabulka_oken_xlsx: Path) -> list[dict]:
    """Read Tabulka 0042 → list of window-type entries.

    Returns one entry per W## row with all 20 columns wrapped in
    `{value, source, confidence: 1.0}`. Komplex-wide catalogue —
    same content for any objekt.
    """
    if not tabulka_oken_xlsx.exists():
        return []
    wb = openpyxl.load_workbook(tabulka_oken_xlsx, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        return []
    ws = wb[SHEET_NAME]

    xlsx_rel = "shared/xlsx/" + tabulka_oken_xlsx.name
    out: list[dict] = []

    for r, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True),
                             start=DATA_START_ROW):
        if not row:
            continue
        kod = _coerce(row[0])
        if not kod or not isinstance(kod, str) or not kod.startswith("W"):
            continue

        entry: dict = {
            "id": f"window.{kod}.row{r}",
            "tabulka_oken_row": _wrap_cell(r, f"XLSX|{xlsx_rel}|{SHEET_NAME}|row={r}"),
        }
        for col_idx, field, _ in COLUMN_MAP:
            value = row[col_idx - 1] if (col_idx - 1) < len(row) else None
            src = f"XLSX|{xlsx_rel}|{SHEET_NAME}|row={r},col={col_idx}"
            entry[field] = _wrap_cell(value, src)

        out.append(entry)

    wb.close()
    return out
