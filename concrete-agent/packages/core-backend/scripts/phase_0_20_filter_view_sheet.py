"""Phase 0.20 — Add 12_Filter_view sheet (6 sortovacích views).

Klient (VELTON) vyžádal flexibilní pohledy na výkaz výměr bez modifikace
original `1_Vykaz_vymer` sheet. Tento script přidá NEW sheet
`12_Filter_view` s 6 views (A–F) vertically stacked + section headers.

Constraints (hard):
  - 1_Vykaz_vymer = read-only source
  - Žádné výpočty navíc — jen agregace
  - Žádné dummy data
  - 11 existing sheets nezasaženy

Reality (per user Opce A):
  - Source má 3021 items (ne 3018 z task spec — drift +3 z Phase 0.19)
  - Sheets count: 11 → 12 (ne 12 → 13)
"""
from __future__ import annotations

import re
import warnings
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

EXCEL = Path(
    "test-data/libuse/outputs/"
    "Vykaz_vymer_Libuse_objekt_D_dokoncovaci_prace.xlsx"
)
SHEET_SRC = "1_Vykaz_vymer"
SHEET_NEW = "12_Filter_view"

# ===== Styling =====
FILL_SECTION = PatternFill("solid", fgColor="1F4E78")
FONT_SECTION_H = Font(bold=True, color="FFFFFF", size=14)
FILL_HEADER = PatternFill("solid", fgColor="4472C4")
FONT_HEADER = Font(bold=True, color="FFFFFF", size=10)
FONT_DATA = Font(size=10)
FILL_ALT = PatternFill("solid", fgColor="F2F2F2")
ALIGN_R = Alignment(horizontal="right", vertical="center")
ALIGN_C = Alignment(horizontal="center", vertical="center")
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(border_style="thin", color="888888")
BORDER = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)

# Status color coding
STATUS_FILL = {
    "matched_high": PatternFill("solid", fgColor="C6EFCE"),
    "matched_medium": PatternFill("solid", fgColor="C6EFCE"),
    "needs_review": PatternFill("solid", fgColor="FFEB9C"),
    "OPRAVENO_OBJEM": PatternFill("solid", fgColor="FFEB9C"),
    "OPRAVENO_POPIS": PatternFill("solid", fgColor="FFEB9C"),
    "no_match": PatternFill("solid", fgColor="FFC7CE"),
    "VYNECHANE_KRITICKE": PatternFill("solid", fgColor="FFC7CE"),
    "VYNECHANE_DETAIL": PatternFill("solid", fgColor="D9D9D9"),
}


def extract_podlazi(misto: str) -> str:
    if not isinstance(misto, str):
        return "unknown"
    parts = [p.strip() for p in misto.split("·")]
    if len(parts) >= 2:
        return parts[1]
    return "unknown"


def extract_skladba_kod(s) -> str | None:
    if not isinstance(s, str):
        return None
    m = re.search(r"\b(FF|CF|RF|WF|F)(\d{2})\b", s)
    if m:
        return m.group(0)
    return None


def load_source_df() -> pd.DataFrame:
    """Load 1_Vykaz_vymer as DataFrame (read-only)."""
    wb = load_workbook(EXCEL, data_only=True)
    ws = wb[SHEET_SRC]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    data = [
        [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        for r in range(2, ws.max_row + 1)
    ]
    df = pd.DataFrame(data, columns=headers)
    df["Podlaží"] = df["Místo"].apply(extract_podlazi)
    df["Skladba_kod"] = df["Skladba/povrch"].apply(extract_skladba_kod)
    return df


def write_section_header(ws, row: int, text: str, span_cols: int = 8) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row,
                    end_column=span_cols)
    cell = ws.cell(row, 1, text)
    cell.fill = FILL_SECTION
    cell.font = FONT_SECTION_H
    cell.alignment = ALIGN_L
    ws.row_dimensions[row].height = 22
    return row + 1


def write_table_header(ws, row: int, headers: list[str]) -> int:
    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i, h)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = ALIGN_C
        c.border = BORDER
    return row + 1


def write_data_row(ws, row: int, values: list, *,
                    align_per_col: list[str] | None = None,
                    alt: bool = False, status_color: str | None = None) -> int:
    for i, v in enumerate(values, 1):
        c = ws.cell(row, i, v)
        c.font = FONT_DATA
        c.border = BORDER
        if alt:
            c.fill = FILL_ALT
        if status_color and i == 1:
            sf = STATUS_FILL.get(status_color)
            if sf:
                c.fill = sf
        if align_per_col and i <= len(align_per_col):
            t = align_per_col[i - 1]
            c.alignment = (ALIGN_R if t == "r"
                            else ALIGN_C if t == "c"
                            else ALIGN_L)
    return row + 1


def view_a_per_druh(df: pd.DataFrame, ws, start_row: int) -> int:
    """Per-druh práce — group by Popis, sort by Σ Množství DESC."""
    g = df.groupby("Popis položky").agg(
        suma_mnozstvi=("Množství", "sum"),
        mj_unique=("MJ", lambda x: "|".join(sorted(set(str(v) for v in x.dropna())))),
        pocet_rooms=("Místo", "nunique"),
        min_qty=("Množství", "min"),
        median_qty=("Množství", "median"),
        max_qty=("Množství", "max"),
        kapitola_first=("Kapitola", lambda x: x.iloc[0] if len(x) else ""),
    ).reset_index()
    g = g.sort_values("suma_mnozstvi", ascending=False)
    print(f"  View A rows: {len(g)}")
    if not (50 < len(g) < 1000):
        print(f"⚠️  View A row count {len(g)} mimo expected 50-1000")

    r = write_section_header(
        ws, start_row,
        f"VIEW A — Per-druh práce ({len(g)} unikátních popisů, sort Σ Množství DESC)")
    r = write_table_header(ws, r, [
        "Popis práce", "Σ Množství", "MJ", "Počet rooms",
        "Min qty", "Median qty", "Max qty", "Hlavní kapitola"])
    align = ["l", "r", "c", "r", "r", "r", "r", "c"]
    for i, row in enumerate(g.itertuples(index=False)):
        vals = [row[0], round(row[1], 2), row[2], int(row[3]),
                round(row[4], 3), round(row[5], 3), round(row[6], 3), row[7]]
        r = write_data_row(ws, r, vals, align_per_col=align, alt=(i % 2 == 1))
    return r + 1


def view_b_per_podlazi(df: pd.DataFrame, ws, start_row: int) -> int:
    g = df.groupby("Podlaží").agg(
        pocet_items=("#", "count"),
        pocet_rooms=("Místo", "nunique"),
    ).reset_index()
    cat_patterns = {
        "stěny": r"omítk|štuk|malb|obklad|nátěr",
        "podlahy": r"pot[ěe]r|dlažb|vinyl|FF\d|epoxid|polystyren",
        "podhledy": r"podhled|SDK|CF\d",
        "fasáda": r"fasád|ETICS|cihel|pásk",
    }
    for kat, pat in cat_patterns.items():
        mask = (df["Popis položky"].str.contains(pat, case=False, regex=True, na=False)
                 & (df["MJ"] == "m2"))
        s = df[mask].groupby("Podlaží")["Množství"].sum()
        g[f"m2_{kat}"] = g["Podlaží"].map(s).fillna(0).round(2)

    order = ["1.PP", "1.NP", "2.NP", "3.NP", "střecha", "fasáda", "ALL", "unknown"]
    g["_s"] = g["Podlaží"].map({p: i for i, p in enumerate(order)}).fillna(99)
    g = g.sort_values("_s").drop("_s", axis=1)

    print(f"  View B rows: {len(g)}")
    if g["pocet_items"].sum() != len(df):
        print(f"⚠️  View B Σ items {g['pocet_items'].sum()} != source {len(df)}")

    r = write_section_header(
        ws, start_row,
        f"VIEW B — Per-podlaží ({len(g)} pater, kategorie m² breakdown)")
    r = write_table_header(ws, r, [
        "Podlaží", "Σ items", "Σ rooms", "m² stěny",
        "m² podlahy", "m² podhledy", "m² fasáda"])
    align = ["c", "r", "r", "r", "r", "r", "r"]
    for i, row in enumerate(g.itertuples(index=False)):
        vals = [row[0], int(row[1]), int(row[2]),
                float(row[3]), float(row[4]), float(row[5]), float(row[6])]
        r = write_data_row(ws, r, vals, align_per_col=align, alt=(i % 2 == 1))
    return r + 1


def view_c_per_kapitola(df: pd.DataFrame, ws, start_row: int) -> int:
    g = df.groupby("Kapitola").agg(
        pocet_items=("#", "count"),
    ).reset_index()
    mj_list = ["m2", "m3", "ks", "kg", "m", "h", "kpl"]
    for mj in mj_list:
        mask = df["MJ"] == mj
        s = df[mask].groupby("Kapitola")["Množství"].sum()
        g[f"Σ_{mj}"] = g["Kapitola"].map(s).fillna(0).round(2)

    g["_p"] = g["Kapitola"].apply(
        lambda k: (0 if str(k).startswith("HSV") else
                    1 if str(k).startswith("PSV") else 2, str(k)))
    g = g.sort_values("_p").drop("_p", axis=1)

    print(f"  View C rows: {len(g)}")
    if g["pocet_items"].sum() != len(df):
        print(f"⚠️  View C Σ items {g['pocet_items'].sum()} != source {len(df)}")

    r = write_section_header(
        ws, start_row, f"VIEW C — Per-Kapitola ({len(g)} kapitol, MJ breakdown)")
    r = write_table_header(ws, r,
                            ["Kapitola", "Σ items"] + [f"Σ_{mj}" for mj in mj_list])
    align = ["c", "r"] + ["r"] * len(mj_list)
    for i, row in enumerate(g.itertuples(index=False)):
        vals = [row[0], int(row[1])] + [float(v) for v in row[2:]]
        r = write_data_row(ws, r, vals, align_per_col=align, alt=(i % 2 == 1))
    return r + 1


def view_d_quality(df: pd.DataFrame, ws, start_row: int) -> int:
    s = df["Status"].value_counts(dropna=False).reset_index()
    s.columns = ["Status", "Počet items"]
    s["% z total"] = (s["Počet items"] / len(df) * 100).round(1)
    akce_map = {
        "matched_high": "OK — žádná akce",
        "matched_medium": "OK — sekundární kontrola",
        "no_match": "Manuální ÚRS lookup potřeba",
        "needs_review": "Per-row review",
        "VYNECHANE_KRITICKE": "Re-validovat (kritické)",
        "VYNECHANE_DETAIL": "Optional review",
        "OPRAVENO_OBJEM": "Track oprava",
        "OPRAVENO_POPIS": "Track oprava",
        "to_audit": "Audit pending",
        "deprecated": "Item zachován v historii (mnozstvi=0)",
    }
    s["Akce"] = s["Status"].astype(str).map(akce_map).fillna("—")
    print(f"  View D rows: {len(s)}, Σ items: {s['Počet items'].sum()}")

    r = write_section_header(
        ws, start_row, f"VIEW D — Quality dashboard ({len(s)} statuses)")
    r = write_table_header(ws, r, ["Status", "Počet items", "% z total", "Akce"])
    align = ["c", "r", "r", "l"]
    for i, row in enumerate(s.itertuples(index=False)):
        vals = [str(row[0]), int(row[1]), float(row[2]), row[3]]
        r = write_data_row(ws, r, vals, align_per_col=align,
                            alt=(i % 2 == 1), status_color=str(row[0]))
    return r + 1


def view_e_per_skladba(df: pd.DataFrame, ws, start_row: int) -> int:
    sub = df[df["Skladba_kod"].notna()].copy()
    g_count = sub.groupby("Skladba_kod").size().rename("pocet_items")
    g_rooms = sub.groupby("Skladba_kod")["Místo"].nunique().rename("pocet_rooms")
    sub_m2 = sub[sub["MJ"] == "m2"]
    g_m2 = sub_m2.groupby("Skladba_kod")["Množství"].sum().rename("suma_m2")
    g = pd.concat([g_count, g_m2, g_rooms], axis=1).reset_index()
    g["suma_m2"] = g["suma_m2"].fillna(0).round(2)
    g = g.sort_values("Skladba_kod")

    print(f"  View E rows: {len(g)}")
    r = write_section_header(
        ws, start_row, f"VIEW E — Per-skladba ({len(g)} unikátních F/FF/CF/RF/WF)")
    r = write_table_header(ws, r,
                            ["Skladba kód", "Σ items", "Σ m²", "Počet rooms"])
    align = ["c", "r", "r", "r"]
    for i, row in enumerate(g.itertuples(index=False)):
        vals = [row[0], int(row[1]), float(row[2]), int(row[3])]
        r = write_data_row(ws, r, vals, align_per_col=align, alt=(i % 2 == 1))
    return r + 1


def view_f_filterable(df: pd.DataFrame, ws, start_row: int) -> int:
    cols = [c for c in df.columns if c not in ("Podlaží", "Skladba_kod")]
    r = write_section_header(
        ws, start_row,
        f"VIEW F — Filterable raw list ({len(df)} items + Excel auto-filter)",
        span_cols=len(cols))
    header_row = r
    for i, col in enumerate(cols, 1):
        c = ws.cell(r, i, col)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = ALIGN_C
        c.border = BORDER
    r += 1

    for i, row in enumerate(df.itertuples(index=False)):
        rd = dict(zip(df.columns, row))
        for j, col in enumerate(cols, 1):
            v = rd.get(col)
            cell = ws.cell(r, j, v)
            cell.font = FONT_DATA
            if i % 2 == 1:
                cell.fill = FILL_ALT
        r += 1
    end_row = r - 1
    last_col_letter = get_column_letter(len(cols))
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{end_row}"
    return r + 1


def write_toc(ws, view_starts: dict[str, int]) -> None:
    ws.cell(1, 1, "FILTER VIEW — Multi-perspective sorting tool").font = Font(
        bold=True, size=14)
    ws.cell(3, 1, "Obsah:").font = Font(bold=True)
    descriptions = {
        "A": "Per-druh práce (rozpočtářský pohled)",
        "B": "Per-podlaží (timeline)",
        "C": "Per-Kapitola (HSV/PSV trade)",
        "D": "Quality dashboard (status)",
        "E": "Per-skladba (FF/F/CF cross-check)",
        "F": "Filterable raw list (3021 items + autofilter)",
    }
    for i, (k, v) in enumerate(view_starts.items()):
        ws.cell(4 + i, 1, f"  View {k} → row {v}: {descriptions.get(k, '')}")


def main() -> None:
    print("=" * 72)
    print("PHASE 0.20 — Add 12_Filter_view sheet (6 views)")
    print("=" * 72)

    print("\nLoading source…")
    df = load_source_df()
    print(f"  Source rows: {len(df)} (expected 3021 per user Opce A)")
    print(f"  Source columns: {list(df.columns[:12])}")
    print(f"  Distinct podlaží: {sorted(df['Podlaží'].unique())}")

    if len(df) != 3021:
        print(f"⛔ STOP — source row count {len(df)} != 3021")
        return

    print("\nOpening Excel + creating new sheet…")
    wb = load_workbook(EXCEL)
    if SHEET_NEW in wb.sheetnames:
        print(f"  Removing existing {SHEET_NEW}…")
        del wb[SHEET_NEW]
    ws = wb.create_sheet(SHEET_NEW)
    pre_sheet_count = len(wb.sheetnames)
    print(f"  Sheets after add: {pre_sheet_count} ({wb.sheetnames})")

    # Reserve top rows pro TOC (1-10), View A starts row 12
    start_a = 12
    start_a_actual = start_a
    print("\nWriting views…")
    next_row = view_a_per_druh(df, ws, start_a)
    start_b = next_row
    next_row = view_b_per_podlazi(df, ws, next_row)
    start_c = next_row
    next_row = view_c_per_kapitola(df, ws, next_row)
    start_d = next_row
    next_row = view_d_quality(df, ws, next_row)
    start_e = next_row
    next_row = view_e_per_skladba(df, ws, next_row)
    start_f = next_row
    next_row = view_f_filterable(df, ws, next_row)

    write_toc(ws, {"A": start_a_actual, "B": start_b, "C": start_c,
                    "D": start_d, "E": start_e, "F": start_f})

    # Column widths
    widths = [40, 14, 12, 14, 12, 12, 12, 18, 14, 14, 30, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A12"

    print("\nSaving Excel…")
    wb.save(EXCEL)
    print(f"  Saved {EXCEL}")

    # Final audit
    print("\n=== FINAL AUDIT ===")
    wb2 = load_workbook(EXCEL, data_only=True)
    print(f"  Sheets ({len(wb2.sheetnames)}): {wb2.sheetnames}")
    ws_src = wb2[SHEET_SRC]
    src_rows = ws_src.max_row - 1
    print(f"  1_Vykaz_vymer rows: {src_rows} (expected 3021)")
    if src_rows != 3021:
        print(f"⛔ ERROR: source modified!")
        return

    ws_new = wb2[SHEET_NEW]
    found_views = []
    for r in range(1, ws_new.max_row + 1):
        v = ws_new.cell(r, 1).value
        if isinstance(v, str) and v.startswith("VIEW "):
            found_views.append(v[:50])
    print(f"  Views in {SHEET_NEW}: {len(found_views)}")
    for fv in found_views:
        print(f"    {fv}")
    if len(found_views) < 6:
        print(f"⛔ ERROR: only {len(found_views)} views, need 6")
        return
    print("\n✅ All checks pass.")


if __name__ == "__main__":
    main()
