"""Phase 6 — Excel export of complete výkaz výměr for objekt D.

Output: test-data/libuse/outputs/Vykaz_vymer_Libuse_objekt_D_dokoncovaci_prace.xlsx

10 sheets:
  0_Souhrn                  — executive summary + recommendations
  1_Vykaz_vymer             — 2277 items table with color coding
  2_Audit_proti_staremu     — diff vs starý VV
  3_Critical_findings       — PROBE detail with cost impact
  4_Mistnosti               — 109 rooms reference table
  5_Skladby                 — 31 skladba codes layer breakdown
  6_Border_zone             — items to_be_clarified_with_collegues
  7_VRN                     — vedlejší rozpočtové náklady
  8_Carry_forward_findings  — full findings log
  9_Metadata                — generation info + provenance
"""
from __future__ import annotations

import datetime as dt
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT_DIR = Path("test-data/libuse/outputs")
COMBINED = OUT_DIR / "items_objekt_D_complete.json"
DS = OUT_DIR / "objekt_D_geometric_dataset.json"
DIFF = OUT_DIR / "phase_5_diff.json"
STARY = OUT_DIR / "stary_vv_normalized.json"
TAB = OUT_DIR / "tabulky_loaded.json"
OUT = OUT_DIR / "Vykaz_vymer_Libuse_objekt_D_dokoncovaci_prace.xlsx"


# Color palette
FILL_KRIT = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
FILL_DETAIL = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
FILL_NOVE = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
FILL_SHODA = PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")
FILL_OPRAVENO_OBJ = PatternFill(start_color="E0E0FF", end_color="E0E0FF", fill_type="solid")
FILL_OPRAVENO_POP = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
FILL_HEADER = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
FILL_SECTION = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")

FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BOLD = Font(name="Calibri", size=10, bold=True)
FONT_KRIT = Font(name="Calibri", size=10, bold=True, color="9C0000")
FONT_DETAIL = Font(name="Calibri", size=10, bold=True, color="C66800")
FONT_NORMAL = Font(name="Calibri", size=10)
FONT_TITLE = Font(name="Calibri", size=14, bold=True, color="003366")

BORDER_THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

STATUS_FILLS = {
    "VYNECHANE_KRITICKE": FILL_KRIT,
    "VYNECHANE_DETAIL": FILL_DETAIL,
    "NOVE": FILL_NOVE,
    "SHODA_SE_STARYM": FILL_SHODA,
    "OPRAVENO_OBJEM": FILL_OPRAVENO_OBJ,
    "OPRAVENO_POPIS": FILL_OPRAVENO_POP,
}
STATUS_FONTS = {
    "VYNECHANE_KRITICKE": FONT_KRIT,
    "VYNECHANE_DETAIL": FONT_DETAIL,
}


def autosize(ws, max_width=60):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value is None:
                continue
            v = str(cell.value)
            if len(v) > max_len:
                max_len = min(len(v), max_width)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def write_header(ws, row, headers, fill=FILL_HEADER, font=FONT_HEADER):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = fill
        c.font = font
        c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN


def write_misto(misto: dict) -> str:
    parts = [misto.get("objekt", ""), misto.get("podlazi", "")]
    mistnosti = misto.get("mistnosti", [])
    if mistnosti:
        parts.append(",".join(mistnosti))
    return " · ".join(p for p in parts if p)


def write_skladba(skl: dict) -> str:
    if not skl:
        return ""
    pairs = []
    for k, v in skl.items():
        if isinstance(v, (str, int, float)):
            pairs.append(f"{k}={v}")
    return "; ".join(pairs[:3])


# =================================================== Sheet builders


def build_sheet_souhrn(wb, items, audit, dataset):
    ws = wb.active
    ws.title = "0_Souhrn"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "VÝKAZ VÝMĚR — Bytový soubor Libuše, objekt D"
    ws["A1"].font = FONT_TITLE
    ws["A2"] = "Akce 185-01 / Klient VELTON REAL ESTATE / Generální projektant ABMV world s.r.o."
    ws["A2"].font = FONT_BOLD
    ws["A3"] = "DPS revize 01 (30/11/2021) — pouze dokončovací práce (hrubá stavba je hotová)"
    ws["A4"] = f"Vygenerováno: {dt.datetime.now().strftime('%Y-%m-%d %H:%M')}  · STAVAGENT Phase 0–6 pipeline"
    ws["A4"].font = Font(italic=True, size=9, color="666666")

    # Section: Headline metrics
    ws["A6"] = "Headline metrics"
    ws["A6"].font = Font(size=12, bold=True, color="003366")
    ws.merge_cells("A6:D6")

    rows = [
        ("Total items pro objekt D", len(items), ""),
        ("Stary VV položky processed", audit["metadata"]["old_items_processed"], ""),
        ("Match coverage (SHODA + OPRAVENO)",
         sum(1 for it in items if it.get("urs_status", "").startswith(("SHODA", "OPRAVENO"))),
         ""),
        ("VYNECHANE_KRITICKE (PROBE-flagged)",
         sum(1 for it in items if it.get("urs_status") == "VYNECHANE_KRITICKE"),
         "PROBE 1 cement screed + PROBE 2 hydroizolace pod obklad"),
        ("VYNECHANE_DETAIL (Detaily/OP/LI/PSV-768)",
         sum(1 for it in items if it.get("urs_status") == "VYNECHANE_DETAIL"),
         "Stykové detaily z Knihy detailů + Tabulek prvků"),
        ("NOVE (granular vs collapsed)",
         sum(1 for it in items if it.get("urs_status") == "NOVE"),
         "Náš generator emit více vrstva-items per skladba"),
        ("Orphan staré VV (likely hrubá stavba)",
         audit["metadata"]["orphan_old_count"],
         "Manual review needed — confirm out-of-scope"),
    ]
    for i, (label, val, note) in enumerate(rows, 7):
        ws.cell(row=i, column=1, value=label).font = FONT_BOLD
        ws.cell(row=i, column=2, value=val).alignment = ALIGN_RIGHT
        ws.cell(row=i, column=3, value=note).font = Font(italic=True, size=9, color="666666")

    # Section: Critical findings
    r = 16
    ws.cell(row=r, column=1, value="Critical findings (PROBE)").font = Font(size=12, bold=True, color="9C0000")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1
    cff = dataset.get("carry_forward_findings", [])
    for f in cff:
        ws.cell(row=r, column=1, value=f"⚠️ {f.get('from_phase', '')}").font = FONT_KRIT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1
        ws.cell(row=r, column=1, value=f.get("summary", ""))
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 45
        r += 1
        ws.cell(row=r, column=1, value=f"→ Next action: {f.get('next_action', '')}").font = Font(italic=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 2

    # Section: Cost impact
    ws.cell(row=r, column=1, value="Estimated cost impact (under-booking)").font = Font(size=12, bold=True, color="003366")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1
    cost_rows = [
        ("PROBE 1 — cement screed gap", "~2000 m² komplex × 700 Kč/m²", "~1,400,000 Kč"),
        ("PROBE 2 — hydroizolace pod obklad gap", "~1250 m² komplex × 400 Kč/m²", "~500,000 Kč"),
        ("Stykové detaily VYNECHANE_DETAIL (98 items)", "parapety + ostění + dilatace + …", "~200,000 – 400,000 Kč"),
        ("Total estimated under-booking", "", "~2,100,000 – 2,300,000 Kč"),
    ]
    for label, basis, cost in cost_rows:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=basis).font = Font(italic=True, size=9, color="666666")
        ws.cell(row=r, column=3, value=cost).font = FONT_BOLD
        if label.startswith("Total"):
            for col in range(1, 5):
                ws.cell(row=r, column=col).fill = FILL_KRIT
                ws.cell(row=r, column=col).font = FONT_KRIT
        r += 1

    # Section: Recommendations
    r += 1
    ws.cell(row=r, column=1, value="Recommendations pro investora").font = Font(size=12, bold=True, color="003366")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1
    recs = [
        "1. Doplnit PROBE 1 (cement screed ~2000 m² komplex) do revidovaného VV",
        "2. Doplnit PROBE 2 (hydroizolace pod obklad ~1250 m² komplex) do revidovaného VV",
        "3. Zařadit stykové detaily (vnitřní parapety, ostění, dilatace, mřížky)",
        "4. Vyjasnit border-zone items s elektro/VZT/ZTI collegues (List 6)",
        "5. Negotiovat VRN scope (List 7) — TDI hodiny, % pojištění, záruční rezerva",
        "6. Manual review 1055 orphan staré VV položky — confirm 'out-of-scope hrubá stavba'",
    ]
    for rec in recs:
        ws.cell(row=r, column=1, value=rec)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1

    # Column widths
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 18


def build_sheet_vykaz(wb, items):
    ws = wb.create_sheet("1_Vykaz_vymer")
    headers = ["#", "ÚRS kód", "Kapitola", "Popis položky", "MJ", "Množství",
                "Místo", "Skladba/povrch", "Confidence", "Status", "Poznámka", "Source"]
    write_header(ws, 1, headers)
    ws.freeze_panes = "A2"

    for i, it in enumerate(items, 2):
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=it.get("urs_code") or "[doplnit]")
        ws.cell(row=i, column=3, value=it.get("kapitola"))
        ws.cell(row=i, column=4, value=it.get("popis"))
        ws.cell(row=i, column=5, value=it.get("MJ"))
        ws.cell(row=i, column=6, value=it.get("mnozstvi"))
        ws.cell(row=i, column=7, value=write_misto(it.get("misto", {})))
        ws.cell(row=i, column=8, value=write_skladba(it.get("skladba_ref", {})))
        ws.cell(row=i, column=9, value=it.get("confidence"))
        ws.cell(row=i, column=10, value=it.get("urs_status", ""))
        poznamka = it.get("poznamka", "")
        if it.get("audit_note"):
            poznamka = (poznamka + " | " + it["audit_note"]).strip(" |")
        ws.cell(row=i, column=11, value=poznamka)
        ws.cell(row=i, column=12, value=it.get("category", ""))

        s = it.get("urs_status", "")
        fill = STATUS_FILLS.get(s)
        font = STATUS_FONTS.get(s)
        if fill:
            for col in range(1, 13):
                ws.cell(row=i, column=col).fill = fill
        if font:
            ws.cell(row=i, column=10).font = font
            ws.cell(row=i, column=4).font = font

    # Column widths (manual — autosize is too slow on 2k+ rows)
    widths = {"A": 6, "B": 15, "C": 13, "D": 60, "E": 6, "F": 10,
              "G": 26, "H": 28, "I": 8, "J": 22, "K": 50, "L": 22}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    # AutoFilter
    ws.auto_filter.ref = f"A1:L{len(items) + 1}"


def build_sheet_audit(wb, audit_diff, stary):
    ws = wb.create_sheet("2_Audit_proti_staremu")
    headers = ["#", "Old kód", "Položka starého VV", "MJ staré", "Množství komplex",
                "Status", "Best new popis", "New score", "Sekce starého VV"]
    write_header(ws, 1, headers)
    ws.freeze_panes = "A2"

    # Re-load match candidates for full per-old detail
    match_path = OUT_DIR / "match_candidates.json"
    candidates = json.loads(match_path.read_text(encoding="utf-8"))["candidates"]

    for i, c in enumerate(candidates, 2):
        score = c["best_match"]["composite_score"]
        if score >= 0.45:
            status = "MATCH_HIGH"
            fill = FILL_SHODA
        elif score >= 0.25:
            status = "MATCH_POSSIBLE"
            fill = FILL_OPRAVENO_POP
        else:
            status = "VYNECHANE_ZE_STAREHO"
            fill = FILL_NOVE

        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=c["old_code"])
        ws.cell(row=i, column=3, value=c["old_popis"])
        ws.cell(row=i, column=4, value=c["old_MJ"])
        ws.cell(row=i, column=5, value=c["old_mnozstvi_komplex"])
        ws.cell(row=i, column=6, value=status)
        ws.cell(row=i, column=7, value=c["best_match"]["new_popis"])
        ws.cell(row=i, column=8, value=score)
        ws.cell(row=i, column=9, value=c["old_section"])
        for col in range(1, 10):
            ws.cell(row=i, column=col).fill = fill
    widths = {"A": 6, "B": 14, "C": 60, "D": 8, "E": 12, "F": 22, "G": 60, "H": 10, "I": 38}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.auto_filter.ref = f"A1:I{len(candidates) + 1}"


def build_sheet_critical(wb, dataset, items):
    ws = wb.create_sheet("3_Critical_findings")
    cff = dataset.get("carry_forward_findings", [])
    r = 1
    ws.cell(row=r, column=1, value="Critical findings — PROBE detail")
    ws.cell(row=r, column=1).font = FONT_TITLE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 2

    for f in cff:
        sev = f.get("severity", "info").upper()
        ws.cell(row=r, column=1, value=f"{sev} — {f.get('from_phase', '')}")
        ws.cell(row=r, column=1).font = FONT_KRIT
        ws.cell(row=r, column=1).fill = FILL_KRIT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1
        # Detail rows
        rows = [
            ("Summary", f.get("summary", "")),
            ("Next action", f.get("next_action", "")),
            ("Parser D-side estimate (m²)", f.get("parser_d_side_m2") or "—"),
            ("Old komplex VV value (if known)", f.get("old_komplex_m2") or "—"),
        ]
        for k, v in rows:
            ws.cell(row=r, column=1, value=k).font = FONT_BOLD
            ws.cell(row=r, column=2, value=v)
            ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
            r += 1
        # List flagged items per finding
        kapitola_match = None
        if "HSV-631" in f.get("from_phase", "") or "cement" in f.get("summary", "").lower():
            kapitola_match = "HSV-631"
        elif "PSV-781" in f.get("from_phase", "") or "obklad" in f.get("summary", "").lower():
            kapitola_match = "PSV-781"
        if kapitola_match:
            flagged = [it for it in items
                       if it.get("urs_status") == "VYNECHANE_KRITICKE"
                       and it.get("kapitola") == kapitola_match]
            ws.cell(row=r, column=1, value=f"Items flagged ({len(flagged)} v {kapitola_match})")
            ws.cell(row=r, column=1).font = FONT_BOLD
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            r += 1
            ws.cell(row=r, column=1, value="Item ID")
            ws.cell(row=r, column=2, value="Popis")
            ws.cell(row=r, column=3, value="MJ × Množství")
            ws.cell(row=r, column=4, value="Místo")
            for col in range(1, 5):
                ws.cell(row=r, column=col).fill = FILL_HEADER
                ws.cell(row=r, column=col).font = FONT_HEADER
            r += 1
            for fl in flagged[:30]:
                ws.cell(row=r, column=1, value=fl["item_id"][:8] + "…")
                ws.cell(row=r, column=2, value=fl["popis"])
                ws.cell(row=r, column=3, value=f"{fl['mnozstvi']} {fl['MJ']}")
                ws.cell(row=r, column=4, value=write_misto(fl["misto"]))
                for col in range(1, 5):
                    ws.cell(row=r, column=col).fill = FILL_KRIT
                r += 1
            if len(flagged) > 30:
                ws.cell(row=r, column=1, value=f"… +{len(flagged) - 30} more")
                r += 1
        r += 2
    widths = {"A": 30, "B": 60, "C": 18, "D": 30, "E": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_sheet_mistnosti(wb, dataset):
    ws = wb.create_sheet("4_Mistnosti")
    headers = ["Kód", "Objekt", "Podlaží", "Sekce", "Místnost #",
                "Název", "Plocha m²", "Světlá výška mm", "Obvod m",
                "FF", "F povrch podlahy", "F povrch stěn", "CF typ podhledu", "F povrch podhledu"]
    write_header(ws, 1, headers)
    ws.freeze_panes = "A2"
    for i, r in enumerate(dataset["rooms"], 2):
        ws.cell(row=i, column=1, value=r["code"])
        ws.cell(row=i, column=2, value=r["objekt"])
        ws.cell(row=i, column=3, value=r["podlazi"])
        ws.cell(row=i, column=4, value=r["byt_or_section"])
        ws.cell(row=i, column=5, value=r["mistnost_num"])
        ws.cell(row=i, column=6, value=r.get("nazev", ""))
        ws.cell(row=i, column=7, value=r.get("plocha_podlahy_m2"))
        ws.cell(row=i, column=8, value=r.get("svetla_vyska_mm"))
        ws.cell(row=i, column=9, value=r.get("obvod_m"))
        ws.cell(row=i, column=10, value=r.get("FF", ""))
        ws.cell(row=i, column=11, value=r.get("F_povrch_podlahy", ""))
        ws.cell(row=i, column=12, value=r.get("F_povrch_sten", ""))
        ws.cell(row=i, column=13, value=r.get("CF", ""))
        ws.cell(row=i, column=14, value=r.get("F_povrch_podhledu", ""))
    autosize(ws, max_width=40)
    ws.auto_filter.ref = f"A1:N{len(dataset['rooms']) + 1}"


def build_sheet_skladby(wb, dataset):
    ws = wb.create_sheet("5_Skladby")
    headers = ["Kód skladby", "Kind", "Label", "Vrstva #", "Materiál",
                "Tloušťka mm", "Specifikace", "Referenční výrobek"]
    write_header(ws, 1, headers)
    ws.freeze_panes = "A2"
    r = 2
    for code, skl in sorted(dataset.get("skladby", {}).items()):
        if not skl.get("vrstvy"):
            ws.cell(row=r, column=1, value=code).font = FONT_BOLD
            ws.cell(row=r, column=4, value="(no vrstvy in master Tabulka)")
            ws.cell(row=r, column=4).font = Font(italic=True, color="C66800")
            r += 1
            continue
        for j, v in enumerate(skl["vrstvy"], 1):
            ws.cell(row=r, column=1, value=code if j == 1 else "")
            ws.cell(row=r, column=2, value=skl.get("kind") if j == 1 else "")
            ws.cell(row=r, column=3, value=skl.get("label", "") if j == 1 else "")
            ws.cell(row=r, column=4, value=v.get("poradi", j))
            ws.cell(row=r, column=5, value=v.get("nazev"))
            ws.cell(row=r, column=6, value=v.get("tloustka_mm"))
            ws.cell(row=r, column=7, value=v.get("specifikace"))
            ws.cell(row=r, column=8, value=v.get("referencni_vyrobek"))
            if j == 1:
                ws.cell(row=r, column=1).font = FONT_BOLD
            r += 1
    widths = {"A": 14, "B": 12, "C": 40, "D": 8, "E": 30, "F": 12, "G": 50, "H": 30}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_sheet_border_zone(wb, items):
    ws = wb.create_sheet("6_Border_zone")
    headers = ["Item ID", "Kapitola", "Popis", "MJ", "Množství",
                "Action — s kým mluvit", "Original poznámka"]
    write_header(ws, 1, headers)
    ws.freeze_panes = "A2"
    r = 2
    for it in items:
        if it.get("status") != "to_be_clarified_with_collegues":
            continue
        # Action heuristic from popis
        popis = it["popis"].lower()
        if "elektro" in popis or "drážek" in popis:
            action = "elektro (drážky pro kabely)"
        elif "zti" in popis or "vzt" in popis or "prostup" in popis:
            action = "ZTI / VZT (prostupy pro potrubí)"
        else:
            action = "TZB (general interface)"
        ws.cell(row=r, column=1, value=it["item_id"][:12])
        ws.cell(row=r, column=2, value=it["kapitola"])
        ws.cell(row=r, column=3, value=it["popis"])
        ws.cell(row=r, column=4, value=it["MJ"])
        ws.cell(row=r, column=5, value=it["mnozstvi"])
        ws.cell(row=r, column=6, value=action)
        ws.cell(row=r, column=7, value=it.get("poznamka", ""))
        for col in range(1, 8):
            ws.cell(row=r, column=col).fill = FILL_DETAIL
            ws.cell(row=r, column=col).font = FONT_DETAIL
        r += 1
    widths = {"A": 14, "B": 13, "C": 50, "D": 8, "E": 10, "F": 30, "G": 50}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_sheet_vrn(wb, items):
    ws = wb.create_sheet("7_VRN")
    headers = ["VRN code", "Popis", "MJ", "Množství", "Confidence",
                "Status", "Poznámka", "Warnings"]
    write_header(ws, 1, headers)
    ws.freeze_panes = "A2"
    r = 2
    for it in items:
        if not it["kapitola"].startswith("VRN-"):
            continue
        ws.cell(row=r, column=1, value=it["kapitola"])
        ws.cell(row=r, column=2, value=it["popis"])
        ws.cell(row=r, column=3, value=it["MJ"])
        ws.cell(row=r, column=4, value=it["mnozstvi"])
        ws.cell(row=r, column=5, value=it.get("confidence"))
        ws.cell(row=r, column=6, value=it.get("status", ""))
        ws.cell(row=r, column=7, value=it.get("poznamka", ""))
        ws.cell(row=r, column=8, value=" | ".join(it.get("warnings", [])))
        for col in range(1, 9):
            ws.cell(row=r, column=col).fill = FILL_NOVE
        r += 1
    # Sumační řádek
    ws.cell(row=r, column=1, value=f"Σ VRN items: {r - 2}").font = FONT_BOLD
    widths = {"A": 12, "B": 50, "C": 10, "D": 12, "E": 10, "F": 32, "G": 50, "H": 50}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_sheet_findings_log(wb, dataset):
    ws = wb.create_sheet("8_Carry_forward_findings")
    headers = ["#", "From phase", "Severity", "Summary",
                "Next action", "Parser D-side (m²)", "Old komplex VV (m²)"]
    write_header(ws, 1, headers)
    cff = dataset.get("carry_forward_findings", [])
    for i, f in enumerate(cff, 2):
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=f.get("from_phase"))
        ws.cell(row=i, column=3, value=f.get("severity", "info").upper())
        ws.cell(row=i, column=4, value=f.get("summary"))
        ws.cell(row=i, column=5, value=f.get("next_action"))
        ws.cell(row=i, column=6, value=f.get("parser_d_side_m2") or "—")
        ws.cell(row=i, column=7, value=f.get("old_komplex_m2") or "—")
        if f.get("severity") == "critical":
            for col in range(1, 8):
                ws.cell(row=i, column=col).fill = FILL_KRIT
                ws.cell(row=i, column=col).font = FONT_KRIT
    widths = {"A": 5, "B": 30, "C": 12, "D": 70, "E": 50, "F": 14, "G": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


def build_sheet_metadata(wb, items):
    ws = wb.create_sheet("9_Metadata")
    rows = [
        ("Generated", dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Tool", "STAVAGENT — concrete-agent pipeline"),
        ("Branch", "claude/phase-0-5-batch-and-parser"),
        ("", ""),
        ("Source files used", ""),
        ("  PDFs (DPS dokumentace)", "33 files (test-data/libuse/inputs/pdf/)"),
        ("  DWGs converted to DXF", "14 files (12 valid, 2 skipped — ARS desky + odvodneni teras)"),
        ("  Tabulky XLSX", "9 files (mistnosti, skladby, dveře, okna, prosklené, zámeč, klempíř, překlady, ostatní)"),
        ("  Starý VV", "1 file (Vykaz_vymer_stary.xlsx)"),
        ("", ""),
        ("Pipeline phases completed", ""),
        ("  Phase 0.0", "File reorganization (test-data/libuse/inputs/{pdf,dwg}/)"),
        ("  Phase 0.5", "DWG → DXF batch (libredwg dwg2dxf 0.13.4 from source)"),
        ("  Phase 0.7", "Cross-object validation D (footprint −0.43 % vs spec)"),
        ("  Phase 1", "Tabulky load + room enrichment + skladby decomposition + aggregates"),
        ("  Phase 3a", "1425 items vnitřní (HSV omítky/mazaniny/podlahy/obklady/malby + refinements)"),
        ("  Phase 3b", "104 items vnější + suterén (cihelné pásky, ETICS, Tondach, klempíř, F10/F11)"),
        ("  Phase 3c", "521 items SDK + truhlářské + zámečnické vnitřní + detaily + refinements"),
        ("  Phase 3d", "25 items lešení + pomocné + zařízení staveniště"),
        ("  Phase 3e", "202 items osazení + spec dveře + úklid + border-zone + VRN"),
        ("  Phase 4", "ÚRS RSPS lookup — DEFERRED (hybrid KROS + Perplexity + manual)"),
        ("  Phase 5", "Audit + diff vs starý VV (TF-IDF cosine matching, status assignment)"),
        ("  Phase 6", "Excel export (this file)"),
        ("", ""),
        ("Total items", str(len(items))),
        ("", ""),
        ("Author", "Claude Code (STAVAGENT pipeline)"),
        ("Contact", "alpro1000/STAVAGENT GitHub repo"),
        ("", ""),
        ("Notes pro investora", ""),
        ("  ÚRS kódy", "Sloupec B v Listu 1 obsahuje placeholder '[doplnit]'. Phase 4 (KROS + Perplexity) doplní."),
        ("  Komplex vs D", "Tento výkaz pokrývá pouze objekt D + společný 1.PP suterén. Pro objekty A/B/C je třeba PDF measurement nebo dodatečné DWG."),
        ("  Critical findings", "PROBE 1 + PROBE 2 vyžadují doplnění do revidovaného VV (~2.1 mil Kč under-booked)"),
        ("  Border-zone (List 6)", "2 items čekají na clarification s elektro/VZT/ZTI"),
        ("  VRN (List 7)", "11 items čekají na negociaci s investorem"),
    ]
    for i, (k, v) in enumerate(rows, 1):
        ws.cell(row=i, column=1, value=k).font = FONT_BOLD if k and not k.startswith("  ") else FONT_NORMAL
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 90


# =================================================== Main


def main() -> None:
    print("Loading data…")
    items = json.loads(COMBINED.read_text(encoding="utf-8"))["items"]
    audit = json.loads(DIFF.read_text(encoding="utf-8"))
    dataset = json.loads(DS.read_text(encoding="utf-8"))
    stary = json.loads(STARY.read_text(encoding="utf-8"))

    print("Building workbook…")
    wb = openpyxl.Workbook()

    print("  0_Souhrn")
    build_sheet_souhrn(wb, items, audit, dataset)
    print("  1_Vykaz_vymer")
    build_sheet_vykaz(wb, items)
    print("  2_Audit_proti_staremu")
    build_sheet_audit(wb, audit, stary)
    print("  3_Critical_findings")
    build_sheet_critical(wb, dataset, items)
    print("  4_Mistnosti")
    build_sheet_mistnosti(wb, dataset)
    print("  5_Skladby")
    build_sheet_skladby(wb, dataset)
    print("  6_Border_zone")
    build_sheet_border_zone(wb, items)
    print("  7_VRN")
    build_sheet_vrn(wb, items)
    print("  8_Carry_forward_findings")
    build_sheet_findings_log(wb, dataset)
    print("  9_Metadata")
    build_sheet_metadata(wb, items)

    print("Saving…")
    wb.save(str(OUT))
    print(f"Wrote {OUT} ({OUT.stat().st_size:,} bytes)")
    print(f"Sheets: {len(wb.sheetnames)} — {wb.sheetnames}")


if __name__ == "__main__":
    main()
