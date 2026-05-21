"""GATE 8.4 — 11c_AVK_smeta single-sheet polish (in-place patch).

ONLY touches 11c_AVK_smeta.  No pipeline regen.  No other sheets modified.

Steps:
  A. Add LOKACE rows for masters with status in SKIP_STATUSES (deprecated,
     interpretace_pending_ABMV, …) that pair_materials.py excluded but
     which physically represent rooms where the work IS performed.
  B. Delete G039 "Lepidlo flexibilní pod dlažbu" entirely — duplicates
     the "Dlažba kladení — Lepidlo (Cemix)" sub-item under PSV-771.
  C. Rate verification — flag MATERIÁL rows whose Sp./MJ differs > 10 %
     from KB (don't auto-fix; report to dashboard Block 20).
  D. Area completeness — verify Mn = Σ Mn × Sp./MJ math per LOKACE row.

Output:
  - Excel saved in place (backup pre_gate8_4.xlsx created first)
  - GATE_8_4_POLISH_SUMMARY.md written

Run from repo root:
    python concrete-agent/packages/core-backend/scripts/phase_6_6_gate8_4_polish.py
"""
from __future__ import annotations

import json
import re
import shutil
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[4]
OUTPUTS = REPO_ROOT / "test-data" / "libuse" / "outputs"
EXCEL = OUTPUTS / "Vykaz_vymer_Libuse_objekt_D_dokoncovaci_prace.xlsx"
BACKUP = OUTPUTS / "Vykaz_vymer_pre_gate8_4.xlsx"
ITEMS = OUTPUTS / "items_objekt_D_complete.json"
GROUPS = OUTPUTS / "urs_query_groups.json"
KB_RATES = REPO_ROOT / "test-data" / "libuse" / "knowledge_base" / "generic_consumption_rates.json"
REPORT = OUTPUTS / "GATE_8_4_POLISH_SUMMARY.md"

SKIP_STATUSES = {
    "deprecated", "WRONGLY_ATTRIBUTED_TO_D",
    "interpretace_pending_ABMV",
    "to_be_negotiated_with_investor",
    "to_be_clarified_with_collegues",
}

# G-groups to DELETE entirely (duplicates of sub-items under other masters)
DELETE_GROUPS = {"G039"}  # Lepidlo flexibilní pod dlažbu

NCOLS = 15
COL_POL = 1
COL_GKOD = 2
COL_TYP = 3
COL_KAPITOLA = 4
COL_POPIS = 5
COL_MJ_MASTER = 6
COL_SUM_MN = 7
COL_VSTUP = 8
COL_SPMJ = 9
COL_MN = 10
COL_MJ_MAT = 11
COL_CENA = 12
COL_STOIMOST = 13
COL_ZDROJ = 14
COL_STATUS = 15

LOC_FILL = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
STATUS_FILLS_LOCAL = {
    "OK":      PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Confirm": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "Odhad":   PatternFill(start_color="FFC09A", end_color="FFC09A", fill_type="solid"),
    "no_match": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "deprecated": PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),
    "interpretace_pending_ABMV": PatternFill(start_color="FFE4B5",
                                              end_color="FFE4B5", fill_type="solid"),
}
THIN = Side(style="thin", color="B0B0B0")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LEFT_NOWRAP = Alignment(horizontal="left", vertical="center", wrap_text=False)
NUMBER_FORMAT_QTY = "#,##0.000"
NUMBER_FORMAT_CZK = "#,##0.00"

_RATE_RE = re.compile(r"(\d+(?:[.,]\d+)?)\s*([a-zA-Z²]+)\s*/\s*([a-zA-Z²]+)")


def _parse_rate(s: str | None) -> tuple[float | None, str, str]:
    """Parse "0.2 l/m²" → (0.2, "l", "m²").  Returns (None, "", "") on failure."""
    if not s:
        return None, "", ""
    m = _RATE_RE.search(str(s))
    if not m:
        return None, "", ""
    try:
        return float(m.group(1).replace(",", ".")), m.group(2), m.group(3)
    except ValueError:
        return None, "", ""


def _write_misto_short(misto: dict) -> str:
    if not misto:
        return "—"
    parts = [misto.get("objekt") or "", misto.get("podlazi") or ""]
    mistnosti = misto.get("mistnosti") or []
    if mistnosti:
        parts.append(",".join(mistnosti))
    rendered = " · ".join(p for p in parts if p)
    return f"• {rendered}" if rendered else "—"


def _write_skladba_short(skl: dict) -> str:
    if not skl:
        return ""
    pairs = [f"{k}={v}" for k, v in skl.items()
              if isinstance(v, (str, int, float))]
    return "; ".join(pairs[:3])


def main() -> int:
    print("GATE 8.4 — 11c single-sheet polish")
    print("=" * 60)

    if not BACKUP.exists():
        shutil.copy2(EXCEL, BACKUP)
        print(f"[backup] {BACKUP.relative_to(REPO_ROOT)} ({BACKUP.stat().st_size:,} B)")
    else:
        print(f"[backup] preserved {BACKUP.relative_to(REPO_ROOT)}")

    items_blob = json.loads(ITEMS.read_text(encoding="utf-8"))
    items = items_blob["items"]
    by_id = {it["item_id"]: it for it in items}
    groups = json.loads(GROUPS.read_text(encoding="utf-8"))["groups"]
    g_by_id = {g["group_id"]: g for g in groups}
    kb = json.loads(KB_RATES.read_text(encoding="utf-8"))["rates"]

    print(f"\n[load] items={len(items)} groups={len(groups)} kb_rates={len(kb)}")

    # Build: per G-group → list of skip-status masters (need LOKACE add)
    # Skip VRN-* groups — they have their own "(služby — bez materiálu)"
    # layout and shouldn't get per-room expansion.
    skip_masters_per_g: dict[str, list[dict]] = defaultdict(list)
    for g in groups:
        kapitola = (g.get("kapitola") or "").upper()
        if kapitola.startswith("VRN"):
            continue
        for mid in g.get("items_ids") or []:
            m = by_id.get(mid)
            if not m:
                continue
            if (m.get("status") in SKIP_STATUSES
                    and float(m.get("mnozstvi") or 0) > 0):
                skip_masters_per_g[g["group_id"]].append(m)

    total_skip = sum(len(v) for v in skip_masters_per_g.values())
    print(f"[scope A] skip-status masters needing LOKACE: {total_skip} "
          f"across {len(skip_masters_per_g)} G-groups")

    wb = openpyxl.load_workbook(EXCEL)
    if "11c_AVK_smeta" not in wb.sheetnames:
        print("ERR: 11c_AVK_smeta sheet missing", flush=True)
        return 2
    ws = wb["11c_AVK_smeta"]

    # ----- Pass 1: read all existing 11c rows into list of dicts -----
    all_rows: list[dict] = []
    for r in range(2, ws.max_row + 1):
        cells = [ws.cell(r, c).value for c in range(1, NCOLS + 1)]
        all_rows.append({
            "orig_row": r,
            "cells": cells,
            "outline": (ws.row_dimensions[r].outline_level
                        if ws.row_dimensions.get(r) else 0),
            "hidden": (ws.row_dimensions[r].hidden
                       if ws.row_dimensions.get(r) else False),
        })
    print(f"[load] existing 11c data rows: {len(all_rows)}")

    # ----- Pass 2: index G-group structure in existing rows -----
    # Per G-kód: PRÁCE row index + list of MATERIÁL rows (with their
    # rate + LOKACE range)
    g_structure: dict[str, dict] = defaultdict(
        lambda: {"prace_idx": None, "materials": [], "rendered_master_ids": set()}
    )
    cur_g = None
    cur_mat_idx = None
    for i, rd in enumerate(all_rows):
        cells = rd["cells"]
        typ = cells[COL_TYP - 1]
        gid = cells[COL_GKOD - 1]
        if not gid:
            continue
        if typ == "PRÁCE":
            cur_g = gid
            g_structure[gid]["prace_idx"] = i
            cur_mat_idx = None
        elif typ == "MATERIÁL":
            cur_mat_idx = len(g_structure[gid]["materials"])
            rate_val, rate_num, rate_denom = _parse_rate(cells[COL_SPMJ - 1])
            g_structure[gid]["materials"].append({
                "idx": i, "pol": cells[COL_POL - 1],
                "rate_value": rate_val, "rate_num": rate_num,
                "rate_denom": rate_denom,
                "popis_vstup": cells[COL_VSTUP - 1],
                "mj": cells[COL_MJ_MAT - 1],
                "lokace_indices": [],
            })
        elif typ == "LOKACE" and cur_mat_idx is not None and gid in g_structure:
            g_structure[gid]["materials"][cur_mat_idx]["lokace_indices"].append(i)
            # Track which master_id this LOKACE rendered (col 8 has misto str;
            # we can't recover master_id directly, just count for now)

    # ----- Pass 3: Step B — collect rows to DELETE (G039 etc.) -----
    rows_to_delete: set[int] = set()
    for gid in DELETE_GROUPS:
        if gid in g_structure:
            s = g_structure[gid]
            if s["prace_idx"] is not None:
                rows_to_delete.add(s["prace_idx"])
            for mat in s["materials"]:
                rows_to_delete.add(mat["idx"])
                rows_to_delete.update(mat["lokace_indices"])
    print(f"[Step B] rows marked for deletion (G039 etc.): {len(rows_to_delete)}")

    # ----- Pass 4: Step A — build new LOKACE rows for skip-status masters -----
    # For each G-group with skip-status masters, for each existing MATERIÁL row
    # in that group, emit additional LOKACE rows.  Rendered masters tracked by
    # misto string match.
    new_lokace_per_mat: dict[int, list[dict]] = defaultdict(list)
    added_count = 0
    for gid, skip_masters in skip_masters_per_g.items():
        struct = g_structure.get(gid)
        if not struct or not struct["materials"]:
            continue
        # Existing LOKACE misto strings per material row
        existing_misto: dict[int, set] = {}
        for mat in struct["materials"]:
            misto_set = set()
            for loc_i in mat["lokace_indices"]:
                misto_set.add(all_rows[loc_i]["cells"][COL_VSTUP - 1])
            existing_misto[mat["idx"]] = misto_set

        for sm in skip_masters:
            misto_str = _write_misto_short(sm.get("misto") or {})
            m_area = float(sm.get("mnozstvi") or 0)
            for mat in struct["materials"]:
                if misto_str in existing_misto[mat["idx"]]:
                    continue  # already rendered
                # Compute LOKACE qty
                rate_v = mat["rate_value"]
                # If self-material (parent mat MJ == master MJ from items), 1:1
                if mat["mj"] and (sm.get("MJ") or "").lower() == str(mat["mj"]).lower():
                    loc_qty = m_area
                elif rate_v is not None:
                    loc_qty = m_area * rate_v
                else:
                    loc_qty = m_area
                new_lokace_per_mat[mat["idx"]].append({
                    "master": sm,
                    "misto_str": misto_str,
                    "loc_qty": loc_qty,
                })
                added_count += 1
                existing_misto[mat["idx"]].add(misto_str)

    print(f"[Step A] new LOKACE rows planned: {added_count} across "
          f"{len(new_lokace_per_mat)} MATERIÁL parent rows")

    # ----- Pass 5: Step C — rate verification (flag, don't fix) -----
    rate_flags: list[dict] = []
    kb_by_kind_keyword: dict[str, dict] = {}
    for key, entry in kb.items():
        for kw in entry.get("category_keywords") or []:
            kb_by_kind_keyword.setdefault(kw.lower(), entry)
    for i, rd in enumerate(all_rows):
        if rd["cells"][COL_TYP - 1] != "MATERIÁL":
            continue
        rate_v, rate_num, rate_denom = _parse_rate(rd["cells"][COL_SPMJ - 1])
        if rate_v is None:
            continue
        vstup = str(rd["cells"][COL_VSTUP - 1] or "").lower()
        # Find best KB match
        best_kb = None
        for kw, entry in kb_by_kind_keyword.items():
            if kw in vstup:
                best_kb = entry
                break
        if not best_kb:
            continue
        kb_rate = float(best_kb["rate"])
        if kb_rate == 0:
            continue
        delta_pct = abs(rate_v - kb_rate) / kb_rate * 100
        if delta_pct > 10:
            rate_flags.append({
                "pol": rd["cells"][COL_POL - 1],
                "gid": rd["cells"][COL_GKOD - 1],
                "vstup": rd["cells"][COL_VSTUP - 1],
                "current_rate": rate_v,
                "kb_rate": kb_rate,
                "delta_pct": round(delta_pct, 1),
                "kb_template": best_kb.get("popis_template"),
            })

    print(f"[Step C] rate-deviation flags: {len(rate_flags)} MATERIÁL rows >10% off KB")

    # ----- Pass 6: Step D — area completeness check -----
    area_flags: list[dict] = []
    for i, rd in enumerate(all_rows):
        if rd["cells"][COL_TYP - 1] != "LOKACE":
            continue
        sum_mn = rd["cells"][COL_SUM_MN - 1]
        mn = rd["cells"][COL_MN - 1]
        rate_v, _, _ = _parse_rate(rd["cells"][COL_SPMJ - 1])
        try:
            sum_mn_f = float(sum_mn or 0)
            mn_f = float(mn or 0)
        except (ValueError, TypeError):
            continue
        if sum_mn_f <= 0:
            area_flags.append({"pol": rd["cells"][COL_POL - 1],
                                "issue": "Σ Mn. = 0 or missing"})
            continue
        if rate_v is None or rate_v == 0:
            continue  # self-material or rate not parseable
        expected = sum_mn_f * rate_v
        if abs(expected - mn_f) > max(0.01, expected * 0.01):
            area_flags.append({
                "pol": rd["cells"][COL_POL - 1],
                "issue": f"Mn={mn_f} != Σ.Mn×rate={expected:.3f}",
            })
    print(f"[Step D] area-completeness flags: {len(area_flags)} LOKACE rows")

    # ----- Pass 7: Rebuild row list with insertions + deletions -----
    new_row_data: list[dict] = []
    cur_gid = None
    cur_mat_pol_n = 0
    cur_loc_pol_n = 0
    for i, rd in enumerate(all_rows):
        if i in rows_to_delete:
            continue
        cells = rd["cells"]
        typ = cells[COL_TYP - 1]
        gid = cells[COL_GKOD - 1]

        # Renumber Pol. č. — keep PRÁCE/MATERIÁL/LOKACE pattern
        if typ == "PRÁCE":
            cur_gid = gid
            cur_mat_pol_n = 0
        elif typ == "MATERIÁL":
            cur_mat_pol_n += 1
            cur_loc_pol_n = 0
            new_pol = f"{gid}.M{cur_mat_pol_n}"
            cells = list(cells)
            cells[COL_POL - 1] = new_pol
        elif typ == "LOKACE":
            cur_loc_pol_n += 1
            new_pol = f"{gid}.M{cur_mat_pol_n}.L{cur_loc_pol_n}"
            cells = list(cells)
            cells[COL_POL - 1] = new_pol
        new_row_data.append({
            "cells": cells, "typ": typ, "gid": gid,
            "outline": rd["outline"], "hidden": rd["hidden"],
            "is_new": False,
        })

        # If this was the last LOKACE of a MATERIÁL bucket, inject new
        # LOKACE rows for skip-status masters of that bucket.
        if typ == "LOKACE":
            # Check if next row is NOT LOKACE — means end of bucket
            next_typ = (all_rows[i + 1]["cells"][COL_TYP - 1]
                        if i + 1 < len(all_rows) else None)
            if next_typ != "LOKACE":
                # Find which MATERIÁL idx this LOKACE belongs to
                # (search g_structure)
                struct = g_structure.get(gid, {})
                target_mat = None
                for mat in struct.get("materials", []):
                    if i in mat["lokace_indices"]:
                        target_mat = mat
                        break
                if target_mat:
                    additions = new_lokace_per_mat.get(target_mat["idx"], [])
                    for add in additions:
                        cur_loc_pol_n += 1
                        sm = add["master"]
                        skl_str = _write_skladba_short(sm.get("skladba_ref") or {})
                        new_cells = [None] * NCOLS
                        new_cells[COL_POL - 1] = f"{gid}.M{cur_mat_pol_n}.L{cur_loc_pol_n}"
                        new_cells[COL_GKOD - 1] = gid
                        new_cells[COL_TYP - 1] = "LOKACE"
                        new_cells[COL_KAPITOLA - 1] = cells[COL_KAPITOLA - 1]
                        new_cells[COL_POPIS - 1] = cells[COL_POPIS - 1]
                        new_cells[COL_MJ_MASTER - 1] = cells[COL_MJ_MASTER - 1]
                        new_cells[COL_SUM_MN - 1] = round(
                            float(sm.get("mnozstvi") or 0), 3)
                        new_cells[COL_VSTUP - 1] = add["misto_str"]
                        new_cells[COL_SPMJ - 1] = cells[COL_SPMJ - 1]
                        new_cells[COL_MN - 1] = round(add["loc_qty"], 3)
                        new_cells[COL_MJ_MAT - 1] = cells[COL_MJ_MAT - 1]
                        new_cells[COL_CENA - 1] = None  # formula re-built after row assignment
                        new_cells[COL_STOIMOST - 1] = None
                        new_cells[COL_ZDROJ - 1] = (
                            f"{skl_str + ' · ' if skl_str else ''}"
                            f"📋 deprecated/pending added")
                        new_cells[COL_STATUS - 1] = sm.get("status") or "—"
                        new_row_data.append({
                            "cells": new_cells, "typ": "LOKACE", "gid": gid,
                            "outline": 2, "hidden": True,
                            "is_new": True,
                        })

    # ----- Pass 8: Wipe + write -----
    print(f"\n[write] {len(new_row_data)} data rows (was {len(all_rows)}, "
          f"delta {len(new_row_data) - len(all_rows):+d})")
    # Pre-build parent_mat_row lookup for Cena formula on new LOKACE
    parent_mat_excel_row: dict[int, int] = {}  # idx in new_row_data → Excel row
    for idx, nr in enumerate(new_row_data):
        excel_row = idx + 2  # +1 for header, +1 for 0-index
        if nr["typ"] == "MATERIÁL":
            parent_mat_excel_row[idx] = excel_row
    # Walk new_row_data to determine each LOKACE's parent MATERIÁL excel row
    last_mat_excel_row = {}
    cur_g = None
    cur_mat_excel = None
    for idx, nr in enumerate(new_row_data):
        excel_row = idx + 2
        if nr["typ"] == "PRÁCE":
            cur_g = nr["gid"]
            cur_mat_excel = None
        elif nr["typ"] == "MATERIÁL":
            cur_mat_excel = excel_row
        elif nr["typ"] == "LOKACE":
            last_mat_excel_row[idx] = cur_mat_excel

    # Delete all data rows + re-write
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    cena_letter = get_column_letter(COL_CENA)
    for idx, nr in enumerate(new_row_data):
        excel_row = idx + 2
        cells = nr["cells"]
        for c_idx, val in enumerate(cells, start=1):
            ws.cell(excel_row, c_idx, val)
        # Re-emit formulas for Cena + Stoimost
        if nr["typ"] == "MATERIÁL":
            ws.cell(excel_row, COL_CENA, "")
            ws.cell(excel_row, COL_STOIMOST,
                    f'=IF({cena_letter}{excel_row}="","",'
                    f'J{excel_row}*{cena_letter}{excel_row})')
        elif nr["typ"] == "LOKACE":
            parent_row = last_mat_excel_row.get(idx, excel_row - 1)
            ws.cell(excel_row, COL_CENA, f"={cena_letter}{parent_row}")
            ws.cell(excel_row, COL_STOIMOST,
                    f'=IF({cena_letter}{excel_row}="","",'
                    f'J{excel_row}*{cena_letter}{excel_row})')
        elif nr["typ"] == "PRÁCE":
            ws.cell(excel_row, COL_CENA, "")
            ws.cell(excel_row, COL_STOIMOST,
                    f'=IF({cena_letter}{excel_row}="","",'
                    f'G{excel_row}*{cena_letter}{excel_row})')
        # Fills + alignment + borders + outline
        fill = None
        if nr["typ"] == "LOKACE":
            fill = LOC_FILL
        for c in range(1, NCOLS + 1):
            cell = ws.cell(excel_row, c)
            if fill:
                cell.fill = fill
            cell.alignment = LEFT_NOWRAP
            cell.border = BORDER_ALL
        # Status fill for LOKACE
        if nr["typ"] == "LOKACE":
            status = nr["cells"][COL_STATUS - 1]
            if status in STATUS_FILLS_LOCAL:
                ws.cell(excel_row, COL_STATUS).fill = STATUS_FILLS_LOCAL[status]
        # Number formats
        ws.cell(excel_row, COL_SUM_MN).number_format = NUMBER_FORMAT_QTY
        ws.cell(excel_row, COL_MN).number_format = NUMBER_FORMAT_QTY
        ws.cell(excel_row, COL_CENA).number_format = NUMBER_FORMAT_CZK
        ws.cell(excel_row, COL_STOIMOST).number_format = NUMBER_FORMAT_CZK
        # Outline
        if nr["outline"]:
            ws.row_dimensions[excel_row].outline_level = nr["outline"]
        if nr["hidden"]:
            ws.row_dimensions[excel_row].hidden = True

    # Re-apply auto-filter
    last_col_letter = get_column_letter(NCOLS)
    ws.auto_filter.ref = f"A1:{last_col_letter}{len(new_row_data) + 1}"

    wb.save(EXCEL)
    print(f"[saved] {EXCEL.relative_to(REPO_ROOT)} ({EXCEL.stat().st_size:,} B)")

    # ----- Report -----
    lines = []
    lines.append("# GATE 8.4 — 11c_AVK_smeta polish summary\n")
    lines.append(f"**Generated:** {EXCEL.stat().st_mtime}\n")
    lines.append(f"## Row counts\n")
    lines.append(f"- Before: {len(all_rows):,} data rows")
    lines.append(f"- After: {len(new_row_data):,} data rows")
    lines.append(f"- Delta: {len(new_row_data) - len(all_rows):+d}\n")
    lines.append(f"## Step A — Added LOKACE for skip-status masters\n")
    lines.append(f"Total new LOKACE: **{added_count}** across "
                 f"{len(new_lokace_per_mat)} MATERIÁL parent rows\n")
    lines.append(f"| G-kód | Skip masters | Popis |")
    lines.append(f"|---|---:|---|")
    for gid in sorted(skip_masters_per_g.keys()):
        ms = skip_masters_per_g[gid]
        g = g_by_id.get(gid)
        popis = (g.get("popis_canonical") or "")[:60] if g else ""
        lines.append(f"| {gid} | {len(ms)} | {popis} |")
    lines.append(f"\n## Step B — Deleted G-groups\n")
    for gid in sorted(DELETE_GROUPS):
        g = g_by_id.get(gid)
        popis = (g.get("popis_canonical") or "?") if g else "?"
        lines.append(f"- **{gid}** {popis} (duplicates sub-item under "
                      f"another master)")
    lines.append(f"\n## Step C — Rate deviation flags (>10 % off KB)\n")
    lines.append(f"Total flagged: **{len(rate_flags)}**\n")
    if rate_flags[:20]:
        lines.append(f"| Pol. č. | G | Vstup | Current | KB | Δ% |")
        lines.append(f"|---|---|---|---:|---:|---:|")
        for rf in rate_flags[:20]:
            v = str(rf["vstup"] or "")[:40]
            lines.append(f"| {rf['pol']} | {rf['gid']} | {v} | "
                          f"{rf['current_rate']} | {rf['kb_rate']} | "
                          f"{rf['delta_pct']}% |")
        if len(rate_flags) > 20:
            lines.append(f"| … +{len(rate_flags) - 20} more flags | | | | | |")
    lines.append(f"\n## Step D — Area completeness flags\n")
    lines.append(f"Total flagged: **{len(area_flags)}**\n")
    if area_flags[:10]:
        lines.append(f"| Pol. č. | Issue |")
        lines.append(f"|---|---|")
        for af in area_flags[:10]:
            lines.append(f"| {af['pol']} | {af['issue']} |")
        if len(area_flags) > 10:
            lines.append(f"| … +{len(area_flags) - 10} more | |")

    REPORT.write_text("\n".join(lines), encoding="utf-8")
    print(f"\n[report] {REPORT.relative_to(REPO_ROOT)}")
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
