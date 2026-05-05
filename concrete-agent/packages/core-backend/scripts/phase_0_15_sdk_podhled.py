"""Phase 0.15 — SDK podhled Knauf D112 (CF20 + CF21).

User clarification: S.D.* všechny rooms patří objektu D (sektor suterénu
pod budovou D, vč. kóje pro byty B/C — construction work patří D, jen
užívací právo pro B/C). NEW SCOPE = 111 rooms (68 NP + 43 1.PP).

Per master Tabulka 0030:
  CF20 = podhled chodeb — Systémové zavěšení Knauf D112 (54mm) +
         Pevný SDK podhled 12.5mm + Povrchová úprava 1mm
  CF21 = podhled koupelen — Systémové zavěšení D112 + Pevný SDK do
         vlhka (impregnovaný) 12.5mm + Povrchová úprava

Per starý VV (Vykaz_vymer_stary.xlsx, sheet '100 — Architektonicko-
stavební'):
  Kód 606: "SDK podhled desky 2xA 12,5 bez izolace dvouvrstvá spodní
           kce profil CD+UD"
  Mj: m²; Cena: 898 Kč/m²; Komplex množství: 1590.5 m²

Note: starý VV used 2xA double-layer SDK; master Tabulka 0030 CF20/21
specifikuje pouze 1× SDK 12.5mm. Použit 898 Kč/m² as best-available
ÚRS reference; možný over-estimate ~2× ale validates direction.
"""
from __future__ import annotations

import json
import re
import uuid
import warnings
from pathlib import Path

from openpyxl import load_workbook

warnings.filterwarnings("ignore")

INPUTS = Path("test-data/libuse/inputs")
OUT_DIR = Path("test-data/libuse/outputs")
TAB_MISTNOSTI = INPUTS / "185-01_DPS_D_SO01_100_0020_R01_TABULKA MISTNOSTI.xlsx"
ITEMS = OUT_DIR / "items_objekt_D_complete.json"

URS_KOD = "763171311"  # ÚRS RSPS kapitola 763 SDK konstrukce, generic 9-digit
URS_LABEL_OLD_VV = "kód 606: SDK podhled desky 2xA 12,5 dvouvrstvá kce CD+UD"
UNIT_PRICE_KC = 898.0
MJ = "m2"


def parse_plocha(s) -> float:
    if s is None:
        return 0.0
    try:
        return float(str(s).replace("\xa0", "").replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return 0.0


def is_objekt_D(room_kod: str) -> bool:
    """User-clarified D scope: D.* (NP) + S.D.* (ALL 1.PP under D building)."""
    return (
        room_kod.startswith("D.1.")
        or room_kod.startswith("D.2.")
        or room_kod.startswith("D.3.")
        or room_kod.startswith("S.D.")
    )


def main() -> None:
    print("=" * 72)
    print("PHASE 0.15 — SDK podhled Knauf D112 (CF20 + CF21)")
    print("=" * 72)

    # PRE-FLIGHT
    wb = load_workbook(TAB_MISTNOSTI, data_only=True)
    ws = wb["tabulka místností"]
    target_rooms: list[dict] = []
    for r in ws.iter_rows(min_row=7, max_row=ws.max_row, values_only=True):
        if not r[0]:
            continue
        code = str(r[0])
        if not is_objekt_D(code):
            continue
        cf_typ = str(r[7]) if len(r) > 7 and r[7] else ""
        if "CF20" not in cf_typ and "CF21" not in cf_typ:
            continue
        plocha = parse_plocha(r[2])
        if plocha <= 0:
            continue
        # Determine podlazi
        if code.startswith("S.D."):
            podlazi = "1.PP"
        else:
            parts = code.split(".")
            podlazi = f"{parts[1]}.NP" if len(parts) >= 2 and parts[1].isdigit() else "?"
        target_rooms.append({
            "code": code, "podlazi": podlazi,
            "nazev": str(r[1]) if r[1] else "",
            "plocha_m2": plocha, "cf_typ": cf_typ.strip(),
        })

    print(f"\nD-rooms with CF20 or CF21: {len(target_rooms)}")
    cf20 = [r for r in target_rooms if "CF20" in r["cf_typ"]]
    cf21 = [r for r in target_rooms if "CF21" in r["cf_typ"]]
    print(f"  CF20 (chodby/schodiště): {len(cf20)}")
    print(f"  CF21 (koupelny): {len(cf21)}")
    total_plocha = sum(r["plocha_m2"] for r in target_rooms)
    print(f"  Total plocha: {total_plocha:.2f} m²")

    # SANITY CHECK pre-add
    expected_min, expected_max = 200, 700
    if not (expected_min <= total_plocha <= expected_max):
        print(f"\n⛔ STOP — total plocha {total_plocha:.2f} m² mimo expected "
              f"{expected_min}–{expected_max} m²")
        return

    # LOAD existing items + duplicate check
    items_blob = json.loads(ITEMS.read_text(encoding="utf-8"))
    items = items_blob["items"]
    existing_sdk_rooms: set[str] = set()
    for it in items:
        if (it.get("mnozstvi") or 0) <= 0:
            continue
        if it.get("status") == "deprecated":
            continue
        popis_low = it["popis"].lower()
        if "sdk podhled" in popis_low and "knauf" in popis_low:
            for rc in it.get("misto", {}).get("mistnosti", []):
                existing_sdk_rooms.add(rc)
    print(f"\nRooms with existing SDK podhled item: {len(existing_sdk_rooms)}")

    # ADD items
    new_items: list[dict] = []
    skipped = 0
    for room in target_rooms:
        if room["code"] in existing_sdk_rooms:
            skipped += 1
            continue
        impreg = "CF21" in room["cf_typ"]
        sdk_type = ("Pevný SDK do vlhka 12.5mm impregnovaný"
                     if impreg else "Pevný SDK podhled 12.5mm")
        popis = (f"SDK podhled Knauf D112 — systémové zavěšení 54mm + "
                  f"{sdk_type} + povrchová úprava ({room['cf_typ']})")
        item = {
            "item_id": str(uuid.uuid4()),
            "kapitola": "PSV-763",
            "popis": popis,
            "popis_canonical": popis,
            "MJ": MJ,
            "mnozstvi": round(room["plocha_m2"], 2),
            "misto": {"objekt": "D", "podlazi": room["podlazi"],
                       "mistnosti": [room["code"]]},
            "skladba_ref": {"CF": room["cf_typ"], "vrstva": "celá CF skladba"},
            "vyrobce_ref": "Knauf D112 / Knauf GKB / GKBi",
            "urs_code": URS_KOD,
            "urs_status": "needs_review",
            "urs_confidence": 0.85,
            "urs_alternatives": [],
            "unit_price_kc": UNIT_PRICE_KC,
            "total_price_kc": round(room["plocha_m2"] * UNIT_PRICE_KC, 2),
            "confidence": 0.85,
            "status": "to_audit",
            "source_method": "PHASE_0_15_SDK_podhled_CF20_CF21",
            "audit_note": (
                f"PHASE_0_15: SDK podhled gap fix per Phase 0.14b "
                f"triangulation. Plocha = plocha místnosti per Tabulka 0020. "
                f"ÚRS reference ze starého VV {URS_LABEL_OLD_VV} (Cena "
                f"{UNIT_PRICE_KC} Kč/m². Note: starý VV uvádí 2xA double-"
                f"layer; master Tabulka 0030 CF20/21 specifikuje single-layer "
                f"SDK 12.5mm — possible price over-estimate ~2×, but ÚRS "
                f"directive is best-available evidence)."
            ),
            "warnings": [],
        }
        new_items.append(item)

    # SPOT-CHECK 3 rooms (smallest, median, largest)
    new_items_sorted = sorted(new_items, key=lambda i: i["mnozstvi"])
    samples = []
    if len(new_items_sorted) >= 3:
        samples = [new_items_sorted[0],
                    new_items_sorted[len(new_items_sorted) // 2],
                    new_items_sorted[-1]]
    print(f"\nSpot-check 3 rooms:")
    for it in samples:
        rc = it["misto"]["mistnosti"][0]
        match_tab = next((r for r in target_rooms if r["code"] == rc), None)
        cf_match = match_tab["cf_typ"] in str(it["skladba_ref"]) if match_tab else False
        plocha_match = match_tab and abs(match_tab["plocha_m2"] - it["mnozstvi"]) < 0.5
        scope_ok = is_objekt_D(rc)
        print(f"  {rc:12s} {it['mnozstvi']:>7.2f} m² × {UNIT_PRICE_KC} = "
              f"{it['total_price_kc']:>10,.0f} Kč | "
              f"scope={scope_ok} cf_match={cf_match} plocha_match={plocha_match}")
        if not (scope_ok and cf_match and plocha_match):
            print(f"  ⛔ SPOT-CHECK FAIL: {rc}")
            return

    # APPLY
    items.extend(new_items)
    items_blob["items"] = items
    items_blob["metadata"]["items_count"] = len(items)
    items_blob["metadata"]["phase_0_15_applied"] = True
    ITEMS.write_text(json.dumps(items_blob, ensure_ascii=False, indent=2),
                     encoding="utf-8")

    total_price = sum(i["total_price_kc"] for i in new_items)
    print(f"\n✅ PHASE 0.15 RESULTS:")
    print(f"   Items added: {len(new_items)}")
    print(f"   Skipped (existing): {skipped}")
    print(f"   Total surface: {sum(i['mnozstvi'] for i in new_items):.2f} m²")
    print(f"   Total recovery: {total_price:>12,.0f} Kč")
    print(f"   Items in dataset: {len(items)}")


if __name__ == "__main__":
    main()
