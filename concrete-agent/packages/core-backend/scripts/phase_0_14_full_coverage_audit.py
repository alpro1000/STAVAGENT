"""Phase 0.14 — Full coverage audit napříč VŠEMI D-rooms × VŠEMI F-kódy.

Po Phase 0.10 (documentation), 0.11 (missing kóje), 0.12 (F15 tepelná
izolace), 0.13 (F11+F14+FF01 deprecate) je 1.PP plně pokryto. Tento
audit ověří že **1.NP, 2.NP, 3.NP** D-rooms také mají všechny očekávané
items dle Tabulky místností XLSX.

Přístup:
  1. Pro každý F-kód definujeme expected keyword patterns (sady items
     které by měly v dataset existovat).
  2. Pro každý D-room z Tabulky místností:
       a. extrahujeme F-kódy ze 5 sloupců (FF / F povrch podlahy /
          F povrch sten / CF typ podhledu / F povrch podhledu)
       b. zkontrolujeme že pro každý F-kód existuje aspoň 1 odpovídající
          item v items_objekt_D_complete.json
  3. Report:
       - per-F-code coverage rate (% rooms s alespoň 1 matching item)
       - per-room missing categories
       - top 20 rooms s nejvíc missing items
       - estimated under-billing per gap

Output: ``coverage_audit.json`` + console summary.
"""
from __future__ import annotations

import json
import re
import warnings
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

warnings.filterwarnings("ignore")

INPUTS = Path("test-data/libuse/inputs")
OUT_DIR = Path("test-data/libuse/outputs")
TAB_MISTNOSTI = INPUTS / "185-01_DPS_D_SO01_100_0020_R01_TABULKA MISTNOSTI.xlsx"
ITEMS = OUT_DIR / "items_objekt_D_complete.json"
OUT = OUT_DIR / "coverage_audit.json"

# Expected keyword patterns per F-code, derived from master Tabulka 0030.
# Each F-code → list of [pattern, category_label, est_kc_per_m2] tuples.
# Pattern matches items s mnozstvi > 0 (deprecated zeroed items don't count).
EXPECTED = {
    # ===== FF skladby (podlahy) =====
    # Patterns liberalized to match Phase 6 actual item naming
    "FF20": [
        (r"pot[ěe]r", "cementový potěr 50mm", 600),
        (r"kari s[íi]ť", "kari síť 150/150/4", 80),
        (r"kročej", "kročejová izolace 25mm", 120),
        (r"polysterén|polysteren|psb 50|liapor mix", "polysterén beton 40mm", 200),
    ],
    "FF21": [
        (r"pot[ěe]r", "cementový potěr 50mm", 600),
        (r"kari s[íi]ť", "kari síť 150/150/4", 80),
        (r"kročej", "kročejová izolace 25mm", 120),
        (r"polysterén|psb 50|liapor", "polysterén beton 40mm", 200),
    ],
    "FF30": [
        (r"pot[ěe]r", "cementový potěr 50mm", 600),
        (r"kari s[íi]ť", "kari síť", 80),
        (r"kročej", "kročejová izolace 25mm", 120),
        (r"polysterén|psb 50|liapor", "polysterén beton 40mm", 200),
    ],
    "FF31": [
        (r"pot[ěe]r.*58|cementový potěr 58", "cementový potěr 58mm", 700),
        (r"kari s[íi]ť", "kari síť", 80),
        (r"kročej", "kročejová izolace 25mm", 120),
        (r"polysterén|psb 50|liapor", "polysterén beton 40mm", 200),
    ],
    "FF03": [
        (r"asfalt.*pás|glastek|hydroizolac.*asfalt|radonov", "asfaltová HI proti radonu", 250),
        (r"xps|extrudovan|dekperimetr", "XPS izolace 80mm", 350),
    ],
    "FF01": [],  # pancéřová s vsypem; vlastní items by F11 / F10
    # ===== F povrch podlahy =====
    "F02": [
        (r"keramick[áa] dlažb|dlažba.*chodb", "dlažba bytové chodby", 850),
        (r"lepidlo.*flex|lepidlo flex", "lepidlo flex", 100),
    ],
    "F03": [
        (r"vinyl|gerfl", "vinyl 2.5mm", 950),
        (r"vyrovnávací|samonivelační", "vyrovnávací vrstva 4.5mm", 220),
    ],
    "F10": [
        (r"polyuret|sikafloor.*359|garáž.*podlah", "polyuretanový garážový systém", 1500),
    ],
    "F11": [
        (r"epoxid|sikafloor 2540", "epoxidový nátěr 2-coat", 250),
    ],
    "F18": [
        (r"keramick[áa] dlažb.*koupeln|dlažba.*koupeln|dlažba bytov", "dlažba koupelny", 950),
        (r"hydroizolac.*stěrk|cemelastik", "HI stěrka pod dlažbu", 180),
    ],
    # ===== F povrch stěn =====
    "F04": [
        (r"sádrov[áa] omítka", "sádrová omítka 10mm chodby", 320),
        (r"otěruvzdorný nátěr|primalex|disperzní nátěr|malba", "otěruvzdorný nátěr chodby", 80),
    ],
    "F05": [
        (r"sádrov[áa] omítka", "sádrová omítka 10mm bytové", 320),
        (r"otěruvzdorný nátěr|primalex|disperzní nátěr|malba", "otěruvzdorný nátěr bytové", 80),
    ],
    "F06": [
        (r"jádrová omítka|cemix 082", "jádrová omítka 10mm koupelny", 300),
        (r"hydroizolac.*stěrk|cemelastik", "HI stěrka koupelny", 180),
        (r"keramický obklad|obklad keramický", "keramický obklad", 950),
    ],
    "F14": [
        (r"sikagard|bezprašný nátěr", "bezprašný nátěr ŽB", 180),
    ],
    "F19": [
        (r"vápenocementov[áa] omítka|cemix 073", "vápenocementová omítka", 300),
        (r"otěruvzdorný nátěr|primalex|disperzní|malba", "otěruvzdorný nátěr sklep", 80),
    ],
    # ===== F povrch podhledu =====
    "F15": [
        (r"tepelná izolac.*strop", "tepelná izolace stropů 1PP", 480),
    ],
    "F17": [
        (r"otěruvzdorná výmalba|sdk.*(nátěr|malba|výmalb)|disperzní.*sdk", "SDK otěruvzdorná výmalba", 80),
    ],
    # ===== CF podhledy =====
    "CF20": [
        (r"sdk podhled|knauf d112|systémové zavěšení.*podhled", "SDK podhled chodeb", 750),
    ],
    "CF21": [
        (r"sdk podhled.*(koupel|impreg|vlhk)|knauf gkbi", "SDK podhled koupelny", 850),
    ],
}


def parse_plocha(s) -> float:
    if s is None:
        return 0.0
    try:
        return float(str(s).replace("\xa0", "").replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return 0.0


def extract_codes(value: str) -> list[str]:
    """Extract F-codes (F01–F99, FF01–FF99, CF01–CF99, RF01–RF99) from cell."""
    if not value:
        return []
    return re.findall(r"\b(?:FF|CF|RF|F)\d{1,2}\b", str(value))


def main() -> None:
    print("Loading sources…")
    wb = load_workbook(TAB_MISTNOSTI, data_only=True)
    ws = wb["tabulka místností"]

    rooms: list[dict] = []
    for r in ws.iter_rows(min_row=7, max_row=ws.max_row, values_only=True):
        if not r[0]:
            continue
        code = str(r[0])
        if not (code.startswith("D.") or code.startswith("S.D.")):
            continue
        plocha = parse_plocha(r[2])
        if plocha <= 0:
            continue
        all_codes = set()
        for col in (4, 5, 6, 7, 8):  # FF, F povrch podlahy, F povrch sten,
                                      # CF typ podhledu, F povrch podhledu
            if len(r) > col and r[col]:
                all_codes.update(extract_codes(str(r[col])))
        # Add F14 if poznámka mentions it
        poznamka = str(r[9]) if len(r) > 9 and r[9] else ""
        if "F14" in poznamka:
            all_codes.add("F14")

        # Determine podlazi
        if code.startswith("S.D."):
            podlazi = "1.PP"
        else:
            parts = code.split(".")
            podlazi = f"{parts[1]}.NP" if len(parts) >= 2 and parts[1].isdigit() else "?"

        rooms.append({
            "code": code,
            "podlazi": podlazi,
            "nazev": str(r[1])[:30] if r[1] else "",
            "plocha": plocha,
            "codes_referenced": sorted(all_codes),
            "poznamka": poznamka,
        })

    print(f"Total D-rooms in Tabulka místností: {len(rooms)}")

    # Load items
    items_blob = json.loads(ITEMS.read_text(encoding="utf-8"))
    items = items_blob["items"]
    # Build per-room item index (only items s mnozstvi > 0 a not deprecated)
    items_by_room: dict[str, list[str]] = defaultdict(list)
    for it in items:
        if (it.get("mnozstvi") or 0) <= 0:
            continue
        if it.get("status") == "deprecated":
            continue
        rc = (it.get("misto", {}).get("mistnosti") or [None])[0]
        if rc:
            items_by_room[rc].append(it["popis"].lower())

    # Coverage matrix
    code_total_rooms: Counter = Counter()  # how many rooms reference each code
    code_covered_rooms: dict[str, set] = defaultdict(set)  # rooms with ≥1 item per code
    code_missing_categories: list[dict] = []
    rooms_gaps: list[dict] = []

    for room in rooms:
        room_items_text = items_by_room.get(room["code"], [])
        gaps = []
        for code in room["codes_referenced"]:
            if code not in EXPECTED:
                continue
            patterns = EXPECTED[code]
            if not patterns:  # FF01 has no expected items (handled via F11)
                code_total_rooms[code] += 1
                code_covered_rooms[code].add(room["code"])
                continue
            code_total_rooms[code] += 1
            # Check each expected category
            covered_any = False
            missing_cats = []
            for pattern, label, est_kc in patterns:
                rgx = re.compile(pattern, re.IGNORECASE)
                hit = any(rgx.search(t) for t in room_items_text)
                if hit:
                    covered_any = True
                else:
                    missing_cats.append({"label": label, "est_kc_per_m2": est_kc})
            if covered_any:
                code_covered_rooms[code].add(room["code"])
            if missing_cats:
                gaps.append({
                    "code": code,
                    "missing_categories": missing_cats,
                })
        if gaps:
            rooms_gaps.append({
                "room": room["code"],
                "podlazi": room["podlazi"],
                "nazev": room["nazev"],
                "plocha": room["plocha"],
                "gaps": gaps,
            })

    # Coverage rate per F-code
    print()
    print("=" * 72)
    print("COVERAGE BY F-CODE")
    print("=" * 72)
    print(f"{'Code':6s} {'Used':>5s} {'Covered':>8s} {'Rate':>7s}  Description")
    print("-" * 72)
    coverage_summary = []
    for code in sorted(code_total_rooms, key=lambda c: code_total_rooms[c], reverse=True):
        total = code_total_rooms[code]
        covered = len(code_covered_rooms[code])
        rate = covered / total * 100
        desc = ", ".join(p[1] for p in EXPECTED.get(code, [])) or "(no patterns)"
        print(f"{code:6s} {total:>5d} {covered:>8d} {rate:>6.1f}%  {desc[:50]}")
        coverage_summary.append({
            "code": code, "rooms_using": total, "rooms_covered": covered,
            "coverage_pct": round(rate, 1),
        })

    # Per-room gap distribution
    print()
    print("=" * 72)
    print(f"ROOMS WITH GAPS: {len(rooms_gaps)} / {len(rooms)}")
    print("=" * 72)
    rooms_gaps.sort(key=lambda r: -sum(len(g["missing_categories"]) for g in r["gaps"]))
    print(f"{'Room':12s} {'Pl(m²)':>7s} {'Gap#':>5s}  {'Missing categories':45s}")
    print("-" * 72)
    for r in rooms_gaps[:25]:
        n_gap = sum(len(g["missing_categories"]) for g in r["gaps"])
        gap_summary = ", ".join(
            f"{g['code']}:{len(g['missing_categories'])}" for g in r["gaps"][:5])
        print(f"{r['room']:12s} {r['plocha']:>7.2f} {n_gap:>5d}  {gap_summary[:45]}")

    # Per-podlaží gap totals
    print()
    print("=" * 72)
    print("GAPS BY PODLAŽÍ")
    print("=" * 72)
    by_podlazi: dict[str, dict] = defaultdict(
        lambda: {"rooms": 0, "rooms_with_gaps": 0, "total_gap_categories": 0,
                  "total_plocha_with_gaps": 0.0})
    for room in rooms:
        by_podlazi[room["podlazi"]]["rooms"] += 1
    for r in rooms_gaps:
        p = r["podlazi"]
        by_podlazi[p]["rooms_with_gaps"] += 1
        by_podlazi[p]["total_gap_categories"] += sum(
            len(g["missing_categories"]) for g in r["gaps"])
        by_podlazi[p]["total_plocha_with_gaps"] += r["plocha"]
    for p in sorted(by_podlazi):
        d = by_podlazi[p]
        print(f"  {p:6s} rooms={d['rooms']:>3d}  with_gaps={d['rooms_with_gaps']:>3d}  "
              f"total_gap_categs={d['total_gap_categories']:>4d}  "
              f"plocha={d['total_plocha_with_gaps']:>7.2f} m²")

    # Estimated under-billing
    total_under = 0.0
    by_code_under: Counter = Counter()
    for r in rooms_gaps:
        for g in r["gaps"]:
            for cat in g["missing_categories"]:
                under = r["plocha"] * cat["est_kc_per_m2"]
                total_under += under
                by_code_under[g["code"]] += under
    print()
    print("=" * 72)
    print("ESTIMATED UNDER-BILLING (rough rates per ÚRS averages)")
    print("=" * 72)
    for code in sorted(by_code_under, key=by_code_under.get, reverse=True):
        print(f"  {code:6s}  ~{by_code_under[code]:>10,.0f} Kč")
    print(f"  {'TOTAL':6s}  ~{total_under:>10,.0f} Kč")

    # Persist
    output = {
        "summary": {
            "total_rooms": len(rooms),
            "rooms_with_gaps": len(rooms_gaps),
            "estimated_under_billing_czk": round(total_under, 0),
            "by_podlazi": {p: dict(d) for p, d in by_podlazi.items()},
        },
        "coverage_by_code": coverage_summary,
        "estimated_gap_value_by_code_czk":
            {c: round(v, 0) for c, v in by_code_under.items()},
        "rooms_gaps_top25": rooms_gaps[:25],
        "rooms_gaps_all": rooms_gaps,
    }
    OUT.write_text(json.dumps(output, ensure_ascii=False, indent=2),
                    encoding="utf-8")
    print(f"\nWrote {OUT}")


if __name__ == "__main__":
    main()
