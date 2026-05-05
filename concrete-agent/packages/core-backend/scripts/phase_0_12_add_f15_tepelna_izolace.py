"""Phase 0.12 — Generate missing F15 ``tepelná izolace stropů 1PP`` items.

User cross-check confirmed F15 is referenced in `Povrch podhledu` for
ALL 43 D-rooms in 1.PP per Tabulka místností XLSX. Per master Tabulka
0030 F15 = "Povrch stropů 1PP - tepelná izolace" with skladba:
  - Tepelná izolace minerální vlna 100 mm (ISOVER Top V Final, λ=0.040)
  - Lepidlo 5 mm bezkotevní cementová lepící hmota (Cemix Superfix)

Phase 6 generator never produced items for F15 (verified: 0 tepelná
izolace items across 2574 items). PROBE 4 finding: ~278.61 m² × cca 480 Kč/m²
= ~134 000 Kč under-billed.

This patch generates 2 items per F15-tagged D-room (43 rooms total),
totaling 86 new HSV-713 items.
"""
from __future__ import annotations

import json
import uuid
import warnings
from pathlib import Path

from openpyxl import load_workbook

warnings.filterwarnings("ignore")

INPUTS = Path("test-data/libuse/inputs")
OUT_DIR = Path("test-data/libuse/outputs")
TAB_MISTNOSTI = INPUTS / "185-01_DPS_D_SO01_100_0020_R01_TABULKA MISTNOSTI.xlsx"
ITEMS = OUT_DIR / "items_objekt_D_complete.json"
DS = OUT_DIR / "objekt_D_geometric_dataset.json"

# Per master Tabulka 0030 F15 skladba (přesná specifikace)
F15_LAYERS = [
    ("HSV-713", "Tepelná izolace stropů 1PP — minerální vlna 100 mm "
                "(ISOVER Top V Final, čedičová λ=0,040, TR 30 kPa)",
     1.0),  # qty multiplier (× plocha)
    ("HSV-713", "Lepidlo bezkotevní cementová lepící hmota tl. 5 mm "
                "(Cemix Superfix) — montáž tepelné izolace stropů 1PP",
     1.0),
]


def parse_plocha(s) -> float:
    if s is None:
        return 0.0
    try:
        return float(str(s).replace("\xa0", "").replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return 0.0


def make_item(room_code: str, podlazi: str, kapitola: str, popis: str,
              mnozstvi: float) -> dict:
    return {
        "item_id": str(uuid.uuid4()),
        "kapitola": kapitola,
        "popis": popis,
        "popis_canonical": popis,
        "MJ": "m2",
        "mnozstvi": round(mnozstvi, 3),
        "misto": {"objekt": "D", "podlazi": podlazi, "mistnosti": [room_code]},
        "skladba_ref": {"F_povrch_podhledu": "F15", "vrstva": "tepelná izolace"},
        "vyrobce_ref": "ISOVER Top V Final / Cemix Superfix",
        "urs_code": None,
        "urs_status": "needs_review",
        "urs_confidence": 0.0,
        "urs_alternatives": [],
        "confidence": 0.85,
        "status": "to_audit",
        "source_method": "PHASE_0_12_F15_tepelna_izolace_stropu_1PP",
        "audit_note": (
            "PHASE_0_12: F15 tepelná izolace stropů 1PP — Phase 6 generator "
            "missed this category. Plocha = plocha místnosti per Tabulka "
            "místností (strop 1PP rovná podlaze 1.NP nad ním)."
        ),
        "warnings": [],
    }


def main() -> None:
    print("Loading Tabulka místností…")
    wb = load_workbook(TAB_MISTNOSTI, data_only=True)
    ws = wb["tabulka místností"]

    # Find all D-rooms with F15 in Povrch podhledu
    f15_rooms: list[tuple[str, str, str, float]] = []
    for r in ws.iter_rows(min_row=7, max_row=ws.max_row, values_only=True):
        if not r[0]:
            continue
        code = str(r[0])
        if not (code.startswith("D.") or code.startswith("S.D.")):
            continue
        podhled = str(r[8]) if r[8] else ""
        if "F15" not in podhled:
            continue
        plocha = parse_plocha(r[2])
        if plocha <= 0:
            continue
        # Determine podlazi from code
        if code.startswith("S.D."):
            podlazi = "1.PP"
        else:
            # D.X.Y.ZZ → X is podlazi number
            parts = code.split(".")
            if len(parts) >= 2 and parts[1].isdigit():
                podlazi = f"{parts[1]}.NP"
            else:
                podlazi = "?"
        nazev = str(r[1])[:30] if r[1] else ""
        f15_rooms.append((code, nazev, podlazi, plocha))

    print(f"Rooms with F15 podhled: {len(f15_rooms)}")
    print(f"Total plocha: {sum(p for _,_,_,p in f15_rooms):.2f} m²")

    # Load existing items
    items_blob = json.loads(ITEMS.read_text(encoding="utf-8"))
    items = items_blob["items"]

    # Generate F15 items per room (2 layers each)
    new_items: list[dict] = []
    for code, nazev, podlazi, plocha in f15_rooms:
        for kapitola, popis, qty_mult in F15_LAYERS:
            new_items.append(make_item(code, podlazi, kapitola, popis,
                                        plocha * qty_mult))

    # De-dup safety: skip if any room already has F15 items
    existing_f15_room_codes = set()
    for it in items:
        if "F15" in str(it.get("skladba_ref", {})):
            for rc in it.get("misto", {}).get("mistnosti", []):
                existing_f15_room_codes.add(rc)

    if existing_f15_room_codes:
        print(f"Skipping {len(existing_f15_room_codes)} rooms with existing "
              f"F15 items: {sorted(existing_f15_room_codes)[:5]}…")
        new_items = [it for it in new_items
                      if it["misto"]["mistnosti"][0] not in existing_f15_room_codes]

    items.extend(new_items)
    items_blob["items"] = items
    items_blob["metadata"]["items_count"] = len(items)
    items_blob["metadata"]["phase_0_12_applied"] = True
    ITEMS.write_text(json.dumps(items_blob, ensure_ascii=False, indent=2),
                     encoding="utf-8")

    # Update carry_forward_findings → PROBE 4
    ds = json.loads(DS.read_text(encoding="utf-8"))
    cff = ds.setdefault("carry_forward_findings", [])
    total_m2 = sum(p for _, _, _, p in f15_rooms)
    cff.append({
        "phase": "0.12",
        "type": "PROBE_FINDING",
        "label": "PROBE 4 — F15 tepelná izolace stropů 1PP",
        "description": (
            f"Phase 6 generator missed F15 (tepelná izolace stropů 1PP, "
            f"minerální vlna 100mm + lepidlo 5mm) napříč {len(f15_rooms)} "
            f"D-rooms. Tabulka místností XLSX má F15 v Povrch podhledu "
            f"sloupci pro all S.D.* sklepní kóje + D-side společné prostory. "
            f"Phase 0.12 inject 2 items per room (HSV-713 izolace + lepidlo)."
        ),
        "rooms_affected": len(f15_rooms),
        "total_plocha_m2": round(total_m2, 2),
        "estimated_value_czk": round(total_m2 * 480, 0),  # ~480 Kč/m² avg
        "items_added": len(new_items),
        "source_specifikace": "Tabulka skladeb 0030 F15: tepelná izolace "
                              "minerální vlna ISOVER Top V Final 100mm + "
                              "Cemix Superfix bezkotevní lepidlo 5mm",
    })
    DS.write_text(json.dumps(ds, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"\n✅ Added {len(new_items)} F15 items pro {len(f15_rooms)} rooms")
    print(f"   Total plocha: {total_m2:.2f} m²")
    print(f"   Estimated under-billing closed: ~{total_m2 * 480:,.0f} Kč")
    print(f"   Total items: {len(items)}")


if __name__ == "__main__":
    main()
