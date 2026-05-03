"""Phase 1 step 1 — load all Tabulky XLSX into a single JSON.

Best-effort parsing per Tabulka. Each Tabulka has its own sheet
structure and header row position; we run a dedicated loader per file.
Warnings are collected but never crash the run.

Output: test-data/libuse/outputs/tabulky_loaded.json
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Iterable

import openpyxl
import warnings as w

# Suppress openpyxl header_footer warnings
w.filterwarnings("ignore", message=".*header or footer.*")

INP = Path("test-data/libuse/inputs")
OUT = Path("test-data/libuse/outputs/tabulky_loaded.json")


def _val(c) -> str:
    if c is None:
        return ""
    return str(c).strip()


def _num(c):
    if c is None or _val(c) == "":
        return None
    try:
        return float(str(c).replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Tabulka místností
# ---------------------------------------------------------------------------

def load_mistnosti() -> dict:
    p = INP / "185-01_DPS_D_SO01_100_0020_R01_TABULKA MISTNOSTI.xlsx"
    wb = openpyxl.load_workbook(str(p), data_only=True)
    ws = wb["tabulka místností"]
    rooms: dict[str, dict] = {}
    warnings: list[str] = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        code = _val(row[0])
        if not code:
            continue
        if not re.match(r"^[A-D]\.|^S\.", code):
            continue
        rooms[code] = {
            "code": code,
            "nazev": _val(row[1]),
            "plocha_m2": _num(row[2]),
            "svetla_vyska_mm": _num(row[3]),
            "skladba_podlahy": _val(row[4]),       # FF##
            "povrch_podlahy": _val(row[5]),        # F##
            "povrch_sten": _val(row[6]),           # F##
            "typ_podhledu": _val(row[7]),          # CF##
            "povrch_podhledu": _val(row[8]),       # F##
            "poznamka": _val(row[9]),
        }
    return {"rooms": rooms, "count": len(rooms), "source": str(p), "warnings": warnings}


# ---------------------------------------------------------------------------
# Tabulka skladeb a povrchů — multi-sheet (povrchy / skladby_podlah / skladby sten / skladby strech / podhledy)
# ---------------------------------------------------------------------------

def _load_skladby_sheet(ws, kind: str) -> tuple[dict[str, dict], list[str]]:
    """Skladby format (per Tabulka skladeb):
       - Column A holds the prefix letter(s): F, FF, WF, CF, RF
       - Column B holds the digit suffix: '1', '2', '20', '01', '03', '10' …
       - Column C+ hold the first vrstva name (and continuation rows have empty A+B)
       Code is normalized as f"{A}{int(B):02d}" for 1-2 digit B, or fX for X-digit B
       to match Tabulka místností conventions (FF20, WF03, CF21, F01).
    """
    out: dict[str, dict] = {}
    warns: list[str] = []
    current_code: str | None = None
    current_label: str = ""
    for row in ws.iter_rows(min_row=4, values_only=True):
        a = _val(row[0])
        b = _val(row[1])
        c = _val(row[2])
        d = _val(row[3])
        e = _val(row[4])
        f = _val(row[5]) if len(row) > 5 else ""

        # New code start: A = letter(s), B = digit(s)
        if re.match(r"^[A-Z]{1,3}$", a) and re.match(r"^\d{1,3}$", b):
            current_code = f"{a}{int(b):02d}"
            current_label = current_label or a
            out.setdefault(current_code, {
                "code": current_code,
                "kind": kind,
                "label": current_label,
                "vrstvy": [],
                "celkova_tloustka_mm": 0.0,
            })
            # The first vrstva is on the same row (column C is the vrstva name)
            if c:
                thick = _num(d)
                out[current_code]["vrstvy"].append({
                    "poradi": "1",
                    "nazev": c,
                    "tloustka_mm": thick,
                    "specifikace": e,
                    "referencni_vyrobek": f,
                })
                if thick is not None and thick > 0:
                    out[current_code]["celkova_tloustka_mm"] += thick
            continue

        # Group label row (text in A, no code in B)
        if a and not b and not re.match(r"^\d", c):
            current_label = a
            continue

        # vrstva continuation row (empty A+B, but C has vrstva name)
        if not a and not b and c and current_code is not None:
            thick = _num(d)
            out[current_code]["vrstvy"].append({
                "poradi": str(len(out[current_code]["vrstvy"]) + 1),
                "nazev": c,
                "tloustka_mm": thick,
                "specifikace": e,
                "referencni_vyrobek": f,
            })
            if thick is not None and thick > 0:
                out[current_code]["celkova_tloustka_mm"] += thick

    for v in out.values():
        v["celkova_tloustka_mm"] = round(v["celkova_tloustka_mm"], 1)
    return out, warns


def load_skladby() -> dict:
    p = INP / "185-01_DPS_D_SO01_100_0030_R01_TABULKA SKLADEB A POVRCHU_R01.xlsx"
    wb = openpyxl.load_workbook(str(p), data_only=True)
    out: dict[str, dict] = {}
    warnings: list[str] = []
    for sheet, kind in [
        ("povrchy", "povrch"),
        ("skladby_podlah", "podlaha"),
        ("skladby sten", "stena"),
        ("skladby strech", "strecha"),
        ("podhledy", "podhled"),
    ]:
        if sheet not in wb.sheetnames:
            warnings.append(f"Skladby: sheet {sheet!r} missing")
            continue
        ws = wb[sheet]
        sub, warns = _load_skladby_sheet(ws, kind)
        for code, v in sub.items():
            if code in out:
                # Same code on multiple sheets — keep the more detailed one (more vrstvy)
                if len(v["vrstvy"]) > len(out[code]["vrstvy"]):
                    out[code] = v
            else:
                out[code] = v
        warnings.extend(warns)
    return {"skladby": out, "count": len(out), "source": str(p), "warnings": warnings}


# ---------------------------------------------------------------------------
# Tabulka dveří
# ---------------------------------------------------------------------------

def load_dvere() -> dict:
    p = INP / "185-01_DPS_D_SO01_100_0041_TABULKA DVERI.xlsx"
    wb = openpyxl.load_workbook(str(p), data_only=True)
    items: dict[str, dict] = {}
    warnings: list[str] = []
    if "tab dvere" not in wb.sheetnames:
        return {"items": {}, "count": 0, "source": str(p), "warnings": ["sheet 'tab dvere' missing"]}
    ws = wb["tab dvere"]
    # Walk; pick rows where col A = 'Dxx'
    for row in ws.iter_rows(min_row=4, values_only=True):
        cells = [_val(c) for c in row]
        for c in cells[:4]:
            m = re.match(r"^D\s*(\d{1,3})$", c)
            if m:
                code = f"D{int(m.group(1)):02d}"
                # Find a name + dimensions in the row
                name = next((x for x in cells if 5 <= len(x) <= 80 and not x.startswith("D")), "")
                items.setdefault(code, {
                    "code": code,
                    "raw_row": cells[:14],
                    "name": name,
                })
                break
    return {"items": items, "count": len(items), "source": str(p), "warnings": warnings}


# ---------------------------------------------------------------------------
# Tabulka oken
# ---------------------------------------------------------------------------

def load_okna() -> dict:
    p = INP / "185-01_DPS_D_SO01_100_0042_TABULKA OKEN.xlsx"
    wb = openpyxl.load_workbook(str(p), data_only=True)
    items: dict[str, dict] = {}
    warnings: list[str] = []
    sheet = "tabulka" if "tabulka" in wb.sheetnames else None
    if sheet is None:
        return {"items": {}, "count": 0, "source": str(p), "warnings": ["sheet 'tabulka' missing"]}
    ws = wb[sheet]
    for row in ws.iter_rows(min_row=4, values_only=True):
        cells = [_val(c) for c in row]
        for c in cells[:4]:
            m = re.match(r"^W\s*(\d{1,3})$", c)
            if m:
                code = f"W{int(m.group(1)):02d}"
                name = next((x for x in cells if 5 <= len(x) <= 80 and not x.startswith("W")), "")
                items.setdefault(code, {"code": code, "raw_row": cells[:14], "name": name})
                break
    return {"items": items, "count": len(items), "source": str(p), "warnings": warnings}


# ---------------------------------------------------------------------------
# Tabulka prosklených příček
# ---------------------------------------------------------------------------

def load_prosklene_pricky() -> dict:
    """Codes are CW## (curtain wall) — e.g. CW01, CW07a, CW08b."""
    p = INP / "185-01_DPS_D_SO01_100_0043_TABULKA PROSKLENYCH PRICEK.xlsx"
    wb = openpyxl.load_workbook(str(p), data_only=True)
    items: dict[str, dict] = {}
    warnings: list[str] = []
    if "tabulka" not in wb.sheetnames:
        return {"items": {}, "count": 0, "source": str(p), "warnings": ["sheet 'tabulka' missing"]}
    ws = wb["tabulka"]
    for row in ws.iter_rows(min_row=4, values_only=True):
        cells = [_val(c) for c in row]
        a = cells[0] if cells else ""
        m = re.match(r"^CW\s*(\d{1,3}[a-z]?)$", a, re.IGNORECASE)
        if not m:
            continue
        code = re.sub(r"\s+", "", a.upper())
        items.setdefault(code, {
            "code": code,
            "type_okna": cells[1] if len(cells) > 1 else "",
            "width_mm": _num(cells[2]) if len(cells) > 2 else None,
            "height_mm": _num(cells[3]) if len(cells) > 3 else None,
            "otvirani": cells[4] if len(cells) > 4 else "",
            "zasklení": cells[5] if len(cells) > 5 else "",
            "kovani": cells[6] if len(cells) > 6 else "",
            "tesneni": cells[7] if len(cells) > 7 else "",
        })
    return {"items": items, "count": len(items), "source": str(p), "warnings": warnings}


# ---------------------------------------------------------------------------
# Tabulka zámečnických výrobků (LP##)
# ---------------------------------------------------------------------------

def _load_named_table(p: Path, sheet: str, code_prefix: str, columns: list[str]) -> dict:
    wb = openpyxl.load_workbook(str(p), data_only=True)
    items: dict[str, dict] = {}
    warnings: list[str] = []
    if sheet not in wb.sheetnames:
        return {"items": {}, "count": 0, "source": str(p), "warnings": [f"sheet {sheet!r} missing"]}
    ws = wb[sheet]
    code_pat = re.compile(rf"^{re.escape(code_prefix)}\s*(\d{{1,3}}[a-z]?)$", re.IGNORECASE)
    for row in ws.iter_rows(min_row=7, values_only=True):
        a = _val(row[0])
        m = code_pat.match(a)
        if not m:
            continue
        code = f"{code_prefix.upper()}{m.group(1).zfill(2 if m.group(1).isdigit() else 0)}".replace(" ", "")
        # normalize: 'LP 20' → 'LP20', 'TP 01a' → 'TP01a'
        code_clean = re.sub(r"\s+", "", a.upper())
        item = {"code": code_clean}
        for j, col in enumerate(columns):
            if j + 1 < len(row):
                v = row[j + 1]
                # Numeric quantity column?
                if col in ("mnozstvi", "mj_quantity", "rozvinuta_sirka_mm"):
                    item[col] = _num(v) if isinstance(v, (int, float, str)) else None
                else:
                    item[col] = _val(v)
        items[code_clean] = item
    return {"items": items, "count": len(items), "source": str(p), "warnings": warnings}


def load_zamecnicke() -> dict:
    p = INP / "185-01_DPS_D_SO01_100_0050_R01_TABULKA ZAMECNICKYCH VYROBKU.xlsx"
    cols = ["nazev", "umisteni", "tech_specifikace", "povrch", "ref_vyrobek", "mj", "mnozstvi", "poznamka"]
    return _load_named_table(p, "tabulka", "LP", cols)


def load_klempirske() -> dict:
    p = INP / "185-01_DPS_D_SO01_100_0060_R01_TABULKA KLEMPIRSKYCH PRVKU.xlsx"
    cols = ["nazev", "umisteni", "tech_specifikace", "povrch", "rozvinuta_sirka_mm", "mj", "mnozstvi", "poznamka"]
    return _load_named_table(p, "tabulka", "TP", cols)


def load_ostatni() -> dict:
    p = INP / "185-01_DPS_D_SO01_100_0080_R02 - TABULKA OSTATNICH PRVKU.xlsx"
    cols = ["nazev", "umisteni", "tech_specifikace", "povrch", "mj", "mnozstvi", "poznamka"]
    return _load_named_table(p, "tabulka", "OP", cols)


def load_preklady() -> dict:
    """Tabulka překladů — code split into 2 cells: 'LI' + '01'."""
    p = INP / "185-01_DPS_D_SO01_100_0070_R01_TABULKA PREKLADU.xlsx"
    wb = openpyxl.load_workbook(str(p), data_only=True)
    items: dict[str, dict] = {}
    warnings: list[str] = []
    if "tabulka LI" not in wb.sheetnames:
        return {"items": {}, "count": 0, "source": str(p), "warnings": ["sheet 'tabulka LI' missing"]}
    ws = wb["tabulka LI"]
    for row in ws.iter_rows(min_row=7, values_only=True):
        prefix = _val(row[0])
        suffix = _val(row[1])
        if prefix.upper() == "LI" and suffix:
            code = f"LI{suffix.zfill(2)}"
            items[code] = {
                "code": code,
                "nazev_vyrobku": _val(row[2]),
                "umisteni": _val(row[3]),
                "tech_specifikace": _val(row[4]),
                "mj": _val(row[5]),
                "povrch": _val(row[6]),
                "mnozstvi": _num(row[7]),
                "poznamka": _val(row[8]),
            }
    return {"items": items, "count": len(items), "source": str(p), "warnings": warnings}


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    print("Loading all Tabulky XLSX…")
    out = {
        "mistnosti": load_mistnosti(),
        "skladby":   load_skladby(),
        "dvere":     load_dvere(),
        "okna":      load_okna(),
        "prosklene_pricky": load_prosklene_pricky(),
        "zamecnicke":load_zamecnicke(),
        "klempirske":load_klempirske(),
        "preklady":  load_preklady(),
        "ostatni":   load_ostatni(),
    }
    print()
    for name, blob in out.items():
        n = blob.get("count", "?")
        warns = len(blob.get("warnings", []))
        print(f"  {name:18s} {n:>5} items  warnings={warns}")
    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nWrote {OUT} ({OUT.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
