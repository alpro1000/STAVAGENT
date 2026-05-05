"""Phase 0.16 — Kročejová izolace 25mm Isover N (FF20/21/30/31).

Per master Tabulka 0030 skladby_podlah:
  FF20 (podlaha nad suterénem - dlažba) row 24: Kročejová 25 mm
  FF21 (podlaha nad suterénem - vinyl)  row 32: Kročejová 25 mm
  FF30 (podlaha běžná - dlažba)         row 40: Kročejová 25 mm
  FF31 (podlaha běžná - vinyl)          row 48: Kročejová 25 mm
  + FF10 (podlaha na terénu) row 55: Kročejová 25 mm
  Material: Minerální hydrofobizovaná vlna, dynamická tuhost 23MN/m³
            (Isover N).

Per starý VV (sheet '100 — Architektonicko-stavební'):
  Row 1991 kód 567 / 631R200: "Kročejová izolace - minerální
  hydrofobizovaná vlna, dynamická tuhost 23MN/m³, tl. 25mm"
  Mj: m² (verified col 6); Cena: 120 Kč/m²; Komplex: 3745.94 m².

D-scope per user clarification (S.D.* all = D):
  68 NP rooms + 43 1.PP rooms = 111 rooms total in D-scope.
  Of these, rooms with FF20/21/30/31 = 56 rooms × per-room plocha
  Tabulky 0020 → real D total ~730 m² (~19.5% of komplex 3745.94 m²).
"""
from __future__ import annotations

import json
import re
import uuid
import warnings
from pathlib import Path

from openpyxl import load_workbook

warnings.filterwarnings("ignore")

INPUTS = Path("test-data/libuse/inputs")
OUT_DIR = Path("test-data/libuse/outputs")
TAB_MISTNOSTI = INPUTS / "185-01_DPS_D_SO01_100_0020_R01_TABULKA MISTNOSTI.xlsx"
TAB_SKLADEB = INPUTS / "185-01_DPS_D_SO01_100_0030_R01_TABULKA SKLADEB A POVRCHU_R01.xlsx"
ITEMS = OUT_DIR / "items_objekt_D_complete.json"

URS_KOD = "631R200"
UNIT_PRICE_KC = 120.0
MJ = "m2"
EXPECTED_THICKNESS_MM = 25


def parse_plocha(s) -> float:
    if s is None:
        return 0.0
    try:
        return float(str(s).replace("\xa0", "").replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return 0.0


def is_objekt_D(room_kod: str) -> bool:
    return (room_kod.startswith("D.1.") or room_kod.startswith("D.2.")
            or room_kod.startswith("D.3.") or room_kod.startswith("S.D."))


def verify_thickness_in_skladby() -> bool:
    """Verify all FF20/21/30/31 mají 25mm kročejovou per master Tabulka 0030."""
    wb = load_workbook(TAB_SKLADEB, data_only=True)
    ws = wb["skladby_podlah"]
    found = []
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        cells = [str(c) if c is not None else "" for c in r[:8]]
        if "kročej" in " ".join(cells).lower():
            try:
                tl_raw = cells[3]
                tl_mm = int(float(tl_raw)) if tl_raw else 0
            except (ValueError, TypeError):
                continue
            found.append(tl_mm)
    if not found:
        return False
    if any(t != EXPECTED_THICKNESS_MM for t in found):
        return False
    return True


def main() -> None:
    print("=" * 72)
    print("PHASE 0.16 — Kročejová izolace 25mm Isover N")
    print("=" * 72)

    if not verify_thickness_in_skladby():
        print("⛔ STOP — kročejová tloušťka NEJE 25mm v Tabulce skladeb!")
        return
    print(f"✓ Tabulka skladeb 0030: kročejová izolace = "
          f"{EXPECTED_THICKNESS_MM} mm (verified)")

    # Load D-rooms with FF20/21/30/31
    wb = load_workbook(TAB_MISTNOSTI, data_only=True)
    ws = wb["tabulka místností"]
    target_rooms: list[dict] = []
    ff_pattern = re.compile(r"FF(20|21|30|31)\b")
    for r in ws.iter_rows(min_row=7, max_row=ws.max_row, values_only=True):
        if not r[0]:
            continue
        code = str(r[0])
        if not is_objekt_D(code):
            continue
        ff = str(r[4]) if r[4] else ""
        if not ff_pattern.search(ff):
            continue
        plocha = parse_plocha(r[2])
        if plocha <= 0:
            continue
        if code.startswith("S.D."):
            podlazi = "1.PP"
        else:
            parts = code.split(".")
            podlazi = f"{parts[1]}.NP" if len(parts) >= 2 and parts[1].isdigit() else "?"
        target_rooms.append({
            "code": code, "podlazi": podlazi,
            "nazev": str(r[1])[:30] if r[1] else "",
            "plocha_m2": plocha, "ff_typ": ff.strip(),
        })

    print(f"\nD-rooms with FF20/21/30/31: {len(target_rooms)}")
    total = sum(r["plocha_m2"] for r in target_rooms)
    print(f"Total plocha (per-room sum from Tabulka 0020): {total:.2f} m²")
    # Per user clarification: scope expanded, expected range up
    if not (300 <= total <= 1000):
        print(f"⛔ STOP — total {total:.2f} m² mimo expected 300-1000")
        return
    print(f"✓ Total in expanded D-scope range")

    # Load existing items + duplicate check
    items_blob = json.loads(ITEMS.read_text(encoding="utf-8"))
    items = items_blob["items"]
    existing_rooms: set[str] = set()
    for it in items:
        if (it.get("mnozstvi") or 0) <= 0:
            continue
        if it.get("status") == "deprecated":
            continue
        if "kročej" in it["popis"].lower():
            for rc in it.get("misto", {}).get("mistnosti", []):
                existing_rooms.add(rc)
    print(f"\nRooms with existing kročejová item: {len(existing_rooms)}")

    new_items: list[dict] = []
    skipped = 0
    for room in target_rooms:
        if room["code"] in existing_rooms:
            skipped += 1
            continue
        popis = (f"Kročejová izolace 25mm — minerální hydrofobizovaná vlna "
                  f"(Isover N, dyn. tuhost 23MN/m³) v {room['ff_typ']}")
        new_items.append({
            "item_id": str(uuid.uuid4()),
            "kapitola": "PSV-713",
            "popis": popis,
            "popis_canonical": popis,
            "MJ": MJ,
            "mnozstvi": round(room["plocha_m2"], 2),
            "misto": {"objekt": "D", "podlazi": room["podlazi"],
                       "mistnosti": [room["code"]]},
            "skladba_ref": {"FF": room["ff_typ"],
                             "vrstva": "kročejová izolace 25mm"},
            "vyrobce_ref": "Isover N",
            "urs_code": URS_KOD,
            "urs_status": "needs_review",
            "urs_confidence": 0.95,
            "urs_alternatives": [],
            "unit_price_kc": UNIT_PRICE_KC,
            "total_price_kc": round(room["plocha_m2"] * UNIT_PRICE_KC, 2),
            "confidence": 0.95,
            "status": "to_audit",
            "source_method": "PHASE_0_16_krocejova_izolace_25mm",
            "audit_note": (
                f"PHASE_0_16: kročejová izolace gap fix per Phase 0.14b. "
                f"Plocha = plocha místnosti per Tabulka 0020 ({room['plocha_m2']} "
                f"m²). Thickness 25mm verified Tabulka 0030 (skladba "
                f"{room['ff_typ']}). ÚRS reference starý VV row 1991 kód "
                f"567 / 631R200 (mj=m², 120 Kč/m²). Original Phase 0.14b "
                f"expected ~302 m² @ 28% share — real D scope per user "
                f"clarification (S.D.* all = D) = 730 m² total."
            ),
            "warnings": [],
        })

    # Spot-check
    sorted_items = sorted(new_items, key=lambda i: i["mnozstvi"])
    samples = ([sorted_items[0], sorted_items[len(sorted_items) // 2],
                sorted_items[-1]] if len(sorted_items) >= 3 else sorted_items)
    print(f"\nSpot-check 3 rooms:")
    all_ok = True
    for it in samples:
        rc = it["misto"]["mistnosti"][0]
        match = next((r for r in target_rooms if r["code"] == rc), None)
        scope_ok = is_objekt_D(rc)
        ff_match = match and match["ff_typ"] == str(it["skladba_ref"]["FF"])
        plocha_ok = match and abs(match["plocha_m2"] - it["mnozstvi"]) < 0.5
        ok = scope_ok and ff_match and plocha_ok
        if not ok:
            all_ok = False
        print(f"  {rc:12s} {it['mnozstvi']:>7.2f} m² × {UNIT_PRICE_KC} = "
              f"{it['total_price_kc']:>9,.0f} Kč | scope={scope_ok} "
              f"ff={ff_match} pl={plocha_ok}")
    if not all_ok:
        print("⛔ SPOT-CHECK FAIL")
        return

    # Apply
    items.extend(new_items)
    items_blob["items"] = items
    items_blob["metadata"]["items_count"] = len(items)
    items_blob["metadata"]["phase_0_16_applied"] = True
    ITEMS.write_text(json.dumps(items_blob, ensure_ascii=False, indent=2),
                     encoding="utf-8")

    total_kc = sum(i["total_price_kc"] for i in new_items)
    print(f"\n✅ PHASE 0.16 RESULTS:")
    print(f"   Items added: {len(new_items)} | skipped: {skipped}")
    print(f"   Total surface: {sum(i['mnozstvi'] for i in new_items):.2f} m²")
    print(f"   Total recovery: {total_kc:>12,.0f} Kč")


if __name__ == "__main__":
    main()
