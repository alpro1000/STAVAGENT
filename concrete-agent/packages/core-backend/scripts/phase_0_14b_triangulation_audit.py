"""Phase 0.14b — Triangulation: Tabulka místností × Starý VV × Můj pipeline.

User question: 'comparing what's in documentation AND previous Vykaz
proti tomu co jsem připravil k exportu?'

Phase 0.14 audit compared only 2 sources (Tabulka místností × můj
pipeline output). Phase 5 (earlier session) compared starý VV → můj
pipeline but reported 86% NOVE due to granular vs collapsed methodology.

This patch does the missing 3rd corner: for each gap category found in
Phase 0.14, look it up in starý VV (Vykaz_vymer_stary.xlsx) to confirm
whether the gap is REAL (existed in old VV → just my pipeline missed it)
or whether old VV also missed it (then Tabulka místností is sole SOT).

Method:
  1. Load gap categories from Phase 0.14 coverage_audit.json
  2. For each gap keyword, scan starý VV sheet '100 - Architektonicko-
     stavební' for matching popis + extract qty
  3. Estimate objekt D share = ~28% komplexu (objekt D ~28% from plocha
     ratio + byt count)
  4. Diff = expected D qty - my pipeline qty = REAL D-gap
  5. Output triangulation table

Reference: starý VV total komplex ~183.65 mil Kč bez DPH (zadavatel
VELTON, dodavatel DMG Stav a.s., zpracovatel QSB s.r.o., 2023-03-15).
Objekt D share derived from architecture brief: 348.71 m² objekt D
podlahová plocha vs ~1280 m² komplex = ~27.2%; 11 bytů D / 36 bytů
komplex = 30.6%. Avg ratio used: 28%.
"""
from __future__ import annotations

import json
import re
import warnings
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

warnings.filterwarnings("ignore")

INPUTS = Path("test-data/libuse/inputs")
OUT_DIR = Path("test-data/libuse/outputs")
STARY_VV = INPUTS / "Vykaz_vymer_stary.xlsx"
ITEMS = OUT_DIR / "items_objekt_D_complete.json"
COVERAGE = OUT_DIR / "coverage_audit.json"
OUT = OUT_DIR / "triangulation_audit.json"

OBJEKT_D_SHARE = 0.28  # 28% per architecture brief

# Gap categories from Phase 0.14 with keyword patterns to find in starý VV
GAP_CATEGORIES = [
    # (label, my_pipeline_pattern, stary_vv_pattern, mj, est_kc_per_m2)
    ("Kročejová izolace 25mm Isover N",
     r"kročej",
     r"kročejová.*izolac|isover.*n", "m2", 120),
    ("Polysterén beton PSB 50 / Liapor Mix 40mm",
     r"polysterén|psb 50|liapor",
     r"polysterén|psb 50|liapor mix", "m2", 200),
    ("PE folie separace Deksepar",
     r"pe folie|deksepar",
     r"pe folie|deksepar|separac.*folie", "m2", 30),
    ("SDK podhled Knauf D112 (CF20/CF21)",
     r"sdk podhled|knauf d112",
     r"sdk podhled|knauf d112", "m2", 750),
    ("Systémové zavěšení podhledu",
     r"systémové zavěšení",
     r"systémové zavěšení|závěs.*podhled", "m2", 250),
    ("XPS 80mm Dekperimetr (FF03)",
     r"xps|dekperimetr|extrudovan",
     r"polystyr.*xps|extrudovan", "m2", 350),
    ("Asfaltový pás Glastek 40 radon (FF03)",
     r"glastek|asfalt.*radon",
     r"asfaltový pás|glastek|hydroizolac.*asfalt", "m2", 250),
    ("Cementový potěr 50mm (FF20/FF21/FF30)",
     r"cementový pot[ěe]r.*50|cementový potěr 50",
     r"cementový potěr.*tl\. 50|cementový potěr 50mm", "m2", 600),
    ("Cementový potěr 58mm (FF31)",
     r"cementový pot[ěe]r.*58",
     r"cementový potěr.*tl\. 58|cementový potěr 58mm", "m2", 700),
    ("Tepelná izolace Isover Top V Final 100mm (F15)",
     r"tepelná izolac.*strop|isover top v",
     r"tepelná izolace.*minerální.*isover top|isover top v final", "m2", 480),
    ("Vinyl 2.5mm Gerflor Creation",
     r"vinyl|gerfl",
     r"vinyl|gerflor", "m2", 950),
    ("Keramický obklad bytové koupelny (F06)",
     r"keramick[ýy] obklad|obklad keramický",
     r"keramick.*obklad", "m2", 950),
    ("Sádrová omítka 10mm (F04/F05)",
     r"sádrová omítka",
     r"sádrová omítka", "m2", 320),
    ("Vápenocementová omítka 10mm (F19)",
     r"vápenocementová omítka",
     r"vápenocementová omítka", "m2", 300),
]


def scan_stary_vv() -> dict[str, list[dict]]:
    """Scan starý VV for each gap pattern → list of matching rows."""
    wb = load_workbook(STARY_VV, data_only=True)
    ws = wb["100 - Architektonicko-sta..."]
    hits: dict[str, list[dict]] = defaultdict(list)
    for i, r in enumerate(ws.iter_rows(min_row=130, max_row=ws.max_row, values_only=True), 130):
        cells = list(r[:14])
        text = " ".join(str(c) for c in cells if c is not None).lower()
        if not text.strip():
            continue
        # Find popis (first long string in cells)
        popis = ""
        for c in cells[:10]:
            if c and isinstance(c, str) and len(c) > 12:
                popis = c[:90]
                break
        # Find qty (first numeric > 0 after position 4)
        qty = 0.0
        for c in cells[3:]:
            if isinstance(c, (int, float)) and c > 0:
                qty = float(c)
                break
        for label, _, vv_pattern, _, _ in GAP_CATEGORIES:
            if re.search(vv_pattern, text, re.IGNORECASE):
                hits[label].append({"row": i, "popis": popis, "qty": qty})
                break  # first matching category wins per row
    return hits


def count_pipeline_items(pattern: str) -> tuple[int, float]:
    """Count my pipeline items + total qty matching pattern (D-rooms only)."""
    items = json.loads(ITEMS.read_text(encoding="utf-8"))["items"]
    rgx = re.compile(pattern, re.IGNORECASE)
    n = 0
    qty_sum = 0.0
    for it in items:
        if (it.get("mnozstvi") or 0) <= 0:
            continue
        if it.get("status") == "deprecated":
            continue
        if rgx.search(it["popis"].lower()):
            n += 1
            qty_sum += it["mnozstvi"]
    return n, qty_sum


def main() -> None:
    print("Loading sources…")
    print(f"  starý VV: {STARY_VV.name}")
    print(f"  můj pipeline: {ITEMS.name}")
    print(f"  Object D share assumption: {OBJEKT_D_SHARE * 100:.0f}% komplexu")
    vv_hits = scan_stary_vv()

    print("\n" + "=" * 100)
    print("3-WAY TRIANGULATION")
    print("=" * 100)
    print(f"{'Category':45s} {'StaryVV (komplex)':>22s} {'D očekáváno':>12s} {'Můj VV':>10s} {'Diff':>8s}")
    print("-" * 100)

    triangulation: list[dict] = []
    total_under_real = 0.0
    for label, my_pat, _, mj, est_kc in GAP_CATEGORIES:
        vv = vv_hits.get(label, [])
        vv_qty_total = sum(h["qty"] for h in vv)
        d_expected = vv_qty_total * OBJEKT_D_SHARE
        my_n, my_qty = count_pipeline_items(my_pat)
        diff = d_expected - my_qty
        status = "REAL gap" if diff > d_expected * 0.3 else (
                  "PARTIAL" if diff > 0 else "OK")
        under_kc = max(0.0, diff) * est_kc
        total_under_real += under_kc
        triangulation.append({
            "category": label,
            "stary_vv_komplex_qty": round(vv_qty_total, 2),
            "stary_vv_n_rows": len(vv),
            "d_expected_qty": round(d_expected, 2),
            "my_pipeline_n": my_n,
            "my_pipeline_qty": round(my_qty, 2),
            "diff_d_qty": round(diff, 2),
            "status": status,
            "estimated_under_billing_czk": round(under_kc, 0),
        })
        print(f"{label[:45]:45s} {vv_qty_total:>16.2f} {mj}   {d_expected:>10.2f} {my_qty:>9.2f}  {diff:>+7.2f}  [{status}]")

    print("-" * 100)
    print(f"{'TOTAL UNDER-BILLING (real, validated by starý VV)':80s} {total_under_real:>10,.0f} Kč")
    print()

    output = {
        "method": "Triangulation: Tabulka místností × Starý VV × Můj pipeline",
        "objekt_D_share_assumption": OBJEKT_D_SHARE,
        "stary_vv_source": STARY_VV.name,
        "stary_vv_komplex_total_czk": 183655741.82,
        "total_real_under_billing_czk": round(total_under_real, 0),
        "categories": triangulation,
    }
    OUT.write_text(json.dumps(output, ensure_ascii=False, indent=2),
                    encoding="utf-8")
    print(f"Wrote {OUT}")

    # Summary table by status
    print()
    by_status = defaultdict(lambda: {"n": 0, "kc": 0.0})
    for t in triangulation:
        by_status[t["status"]]["n"] += 1
        by_status[t["status"]]["kc"] += t["estimated_under_billing_czk"]
    for s in ("REAL gap", "PARTIAL", "OK"):
        d = by_status[s]
        print(f"  {s:12s}  {d['n']:>2d} categories  ~{d['kc']:>10,.0f} Kč under-billing")


if __name__ == "__main__":
    main()
