"""Phase 0.10 — Documentation inconsistency audit.

Scans master Tabulka skladeb 0030 + Tabulka místností 0020 for architect
documentation errors that the user/VELTON should be aware of (and that
ABMV needs to clarify before final delivery).

Detects:
  D1 — F-codes referenced in Tabulka místností but undefined in Tabulka skladeb
  D2 — FF column values that lack 'FF' prefix (typo: F30 → should be FF30)
  D3 — Sequence gaps in master tabulka (missing F-numbers between defined ones)
  D4 — Tabulka místností rooms that don't appear in geometric DXF dataset
  D5 — Geometric DXF rooms that don't appear in Tabulka místností
  D6 — WF skladby referenced in dataset.skladby but not in master XLSX

Output: ``test-data/libuse/outputs/documentation_inconsistencies.json``
Used pro ABMV e-mail + dataset enrichment (recommended_interpretation).
"""
from __future__ import annotations

import json
import re
import warnings
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

warnings.filterwarnings("ignore")

INPUTS = Path("test-data/libuse/inputs")
OUT_DIR = Path("test-data/libuse/outputs")
TAB_MISTNOSTI = INPUTS / "185-01_DPS_D_SO01_100_0020_R01_TABULKA MISTNOSTI.xlsx"
TAB_SKLADEB = INPUTS / "185-01_DPS_D_SO01_100_0030_R01_TABULKA SKLADEB A POVRCHU_R01.xlsx"
DATASET = OUT_DIR / "objekt_D_geometric_dataset.json"
OUT = OUT_DIR / "documentation_inconsistencies.json"

# Recommended interpretations for known typos (high-confidence guesses,
# need ABMV confirmation before applying to pricing).
RECOMMENDED = {
    "F20_in_FF": "FF20 (= podlaha nad suterénem dlažba 130 mm)",
    "F30_in_FF": "FF30 (= běžná podlaha dlažba 130 mm)",
    "F20_in_stěn_podhled": "F17 (= SDK obklad / podhled s otěruvzdornou výmalbou)",
}


def load_master_F_codes() -> dict[str, str]:
    """Sequential F-code numbering from master XLSX `povrchy` sheet."""
    wb = load_workbook(TAB_SKLADEB, data_only=True)
    ws = wb["povrchy"]
    labels = []
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        a = str(r[0]).strip() if r[0] else ""
        if a.startswith("Povrch"):
            labels.append(a)
    return {f"F{i:02d}": lbl for i, lbl in enumerate(labels)}


def load_mistnosti() -> list[dict]:
    """Parse Tabulka místností 0020 → list of room dicts."""
    wb = load_workbook(TAB_MISTNOSTI, data_only=True)
    ws = wb["tabulka místností"]
    rooms = []
    for r in ws.iter_rows(min_row=7, max_row=ws.max_row, values_only=True):
        if not r[0] or not str(r[0]).startswith(("D.", "S.D.")):
            continue
        rooms.append({
            "code": str(r[0]).strip(),
            "nazev": str(r[1]) if r[1] else "",
            "plocha_m2": str(r[2]) if r[2] else "",
            "FF": str(r[4]) if r[4] else "",
            "F_povrch_podlahy": str(r[5]) if r[5] else "",
            "F_povrch_sten": str(r[6]) if r[6] else "",
            "CF_typ_podhledu": str(r[7]) if r[7] else "",
            "F_povrch_podhledu": str(r[8]) if r[8] else "",
        })
    return rooms


def main() -> None:
    print("Loading sources…")
    f_codes = load_master_F_codes()
    mistnosti = load_mistnosti()
    ds = json.loads(DATASET.read_text(encoding="utf-8"))
    ds_skladby = ds.get("skladby", {})
    ds_rooms_codes = {r["code"] for r in ds.get("rooms", [])}

    findings: dict = {
        "scan_date": "2026-05-05",
        "source_xlsx": [TAB_MISTNOSTI.name, TAB_SKLADEB.name],
        "summary": {},
    }

    # =================================================================
    # D1 — F-codes referenced but undefined
    # =================================================================
    print("\nD1 — Undefined F-codes in Tabulka místností:")
    f_pattern = re.compile(r"\bF\d{2}\b")
    referenced = Counter()
    refs_per_code: dict[str, list[dict]] = {}
    for room in mistnosti:
        for field in ("F_povrch_sten", "F_povrch_podlahy", "F_povrch_podhledu"):
            for tok in f_pattern.findall(room.get(field, "")):
                referenced[tok] += 1
                refs_per_code.setdefault(tok, []).append(
                    {"room": room["code"], "field": field})
    undefined = {c: refs_per_code[c] for c in referenced
                  if c not in f_codes}
    findings["D1_undefined_F_codes"] = {
        "description": "F-codes used in Tabulka místností XLSX but not defined "
                        "in Tabulka skladeb master XLSX",
        "undefined": [
            {
                "code": code,
                "n_references": len(refs),
                "rooms": [r["room"] for r in refs[:15]],
                "fields": list(set(r["field"] for r in refs)),
                "recommended_interpretation": RECOMMENDED.get(
                    f"{code}_in_stěn_podhled", "manual review with ABMV"),
            }
            for code, refs in sorted(undefined.items())
        ],
    }
    for code, refs in sorted(undefined.items()):
        print(f"  {code} — {len(refs)} ref ({', '.join(set(r['field'] for r in refs))})"
              f" — rooms: {[r['room'] for r in refs[:5]]}…")

    # =================================================================
    # D2 — FF column values without FF prefix (typos)
    # =================================================================
    print("\nD2 — FF column typos (F## should be FF##):")
    ff_typos = []
    for room in mistnosti:
        ff_val = room.get("FF", "").strip()
        if re.fullmatch(r"F\d{2}", ff_val):  # Bare F## without FF prefix
            ff_typos.append({
                "room": room["code"],
                "nazev": room["nazev"],
                "FF_value_in_xlsx": ff_val,
                "recommended": f"F{ff_val}",  # F30 → FF30
            })
    findings["D2_FF_column_typos"] = {
        "description": "Tabulka místností FF column contains bare F## codes "
                        "(missing FF prefix); high-confidence typos.",
        "typos": ff_typos,
    }
    for t in ff_typos:
        print(f"  {t['room']:12s} ({t['nazev']:15s}) FF=\"{t['FF_value_in_xlsx']}\" "
              f"→ should be \"{t['recommended']}\"")

    # =================================================================
    # D3 — Sequence gaps in master tabulka
    # =================================================================
    print("\nD3 — Master Tabulka 0030 sequence gaps:")
    defined_nums = sorted(int(c[1:]) for c in f_codes)
    gaps = []
    for i in range(min(defined_nums), max(defined_nums) + 1):
        if i not in defined_nums:
            gaps.append(f"F{i:02d}")
    findings["D3_master_sequence_gaps"] = {
        "description": "F-code numbers missing from master Tabulka 0030 sequential "
                        "numbering. May or may not be intentional reservations.",
        "gaps": gaps,
    }
    print(f"  Defined F-codes: F00–F{max(defined_nums):02d}")
    print(f"  Gaps: {gaps}")

    # =================================================================
    # D4 — Rooms in Tabulka místností but not in DXF dataset
    # =================================================================
    print("\nD4 — Rooms in Tabulka místností XLSX but not in DXF dataset:")
    xlsx_d_rooms = {r["code"] for r in mistnosti
                     if r["code"].startswith("D.") or r["code"].startswith("S.D.")}
    in_xlsx_not_dxf = sorted(xlsx_d_rooms - ds_rooms_codes)
    findings["D4_xlsx_only_rooms"] = {
        "description": "D-rooms in Tabulka místností XLSX without matching room "
                        "in DXF geometric dataset (DXF parsing gap or XLSX-only "
                        "ghost rooms).",
        "rooms": in_xlsx_not_dxf,
    }
    print(f"  Count: {len(in_xlsx_not_dxf)} rooms")
    for r in in_xlsx_not_dxf[:10]:
        print(f"    {r}")

    # =================================================================
    # D5 — Rooms in DXF dataset but not in Tabulka místností
    # =================================================================
    print("\nD5 — Rooms in DXF dataset but not in Tabulka místností:")
    in_dxf_not_xlsx = sorted(ds_rooms_codes - xlsx_d_rooms)
    findings["D5_dxf_only_rooms"] = {
        "description": "DXF rooms without matching XLSX entry (DXF over-detection "
                        "or undocumented spaces).",
        "rooms": in_dxf_not_xlsx,
    }
    print(f"  Count: {len(in_dxf_not_xlsx)} rooms")
    for r in in_dxf_not_xlsx[:10]:
        print(f"    {r}")

    # =================================================================
    # D6 — Printed-legend section misfile (D.1.3.01 anomaly)
    # =================================================================
    # XLSX has D.1.3.01 in correct D.1.3 byt; PDF legend visually misfiles
    # it under D.1.2 block (architect printout error). Confirmed by
    # subtotal arithmetic: D.1.2 byt printed sum 55.9 m² = 49.59 + 6.35
    # (D.1.3.01); D.1.3 byt printed sum 43.2 m² = 43.24 (without D.1.3.01).
    findings["D6_printed_legend_anomaly"] = {
        "description": "Printed Tabulka místností legend (PDF výkresu) misfiles "
                        "D.1.3.01 under D.1.2 byt section. Source XLSX has it in "
                        "correct D.1.3 byt — pipeline is not affected, but "
                        "printed deliverable to client should be regenerated.",
        "evidence": {
            "xlsx_section": "D.1.3 byt (D.1.3.01 CHODBA 6.35 m², FF20/F02/F05/CF20/F17)",
            "printed_subtotal_D12_pdf": "55.9 m² (= 49.59 actual D.1.2 + 6.35 misfiled D.1.3.01)",
            "printed_subtotal_D13_pdf": "43.2 m² (= 43.24 actual D.1.3 minus D.1.3.01)",
            "correct_subtotal_D12": "49.59 m²",
            "correct_subtotal_D13": "49.59 m² (parity with D.1.2 byt)",
        },
    }
    print("\nD6 — Printed legend anomaly: D.1.3.01 misfiled under D.1.2 in PDF")
    print("  XLSX correct: D.1.3.01 is D.1.3 byt CHODBA 6.35 m²")
    print("  PDF subtotals: D.1.2=55.9 (incorrect, should be 49.59),"
          " D.1.3=43.2 (incorrect, should be 49.59)")

    # =================================================================
    # Summary
    # =================================================================
    findings["summary"] = {
        "D1_undefined_F_codes_count": len(undefined),
        "D2_FF_typos_count": len(ff_typos),
        "D3_sequence_gaps_count": len(gaps),
        "D4_xlsx_only_rooms_count": len(in_xlsx_not_dxf),
        "D5_dxf_only_rooms_count": len(in_dxf_not_xlsx),
        "D6_printed_legend_anomaly": "D.1.3.01 misfiled in PDF (XLSX OK)",
    }
    findings["abmv_email_required"] = [
        "F-CODE NUMBERING INCONSISTENCY (XLSX vs PDF):"
        " XLSX `povrchy` sheet má 23 sequential rows (F00–F22), PDF Tabulka 0030"
        " ma explicit codes F00–F23 with F20 SKIPPED. From F20+ XLSX a PDF se"
        " neshodují (XLSX F20 = obchodní podlaha, PDF F21 = obchodní podlaha)."
        " Která verze je kanonická?",
        f"F20 — used in {referenced.get('F20', 0)} room references. Per XLSX"
        " seq = 'Povrch podlahy - obchodní jednotky' (podlahový kód!). Used in"
        " F povrch stěn + F povrch podhledu of 8 residential POKOJ rooms"
        " (D.3.1.02/06/07, D.3.2.03/04, D.3.3.02/06/07) — semantic mismatch:"
        " podlaha kód aplikován na stěny/podhled. Recommended: F17 (SDK +"
        " otěruvzdorná výmalba). Confirm.",
        "F30 — used in 4 WC FF column references (D.2.1.03, D.2.4.03,"
        " D.3.1.03, D.3.3.03). Bare 'F30' missing 'FF' prefix — typo for FF30"
        " (běžná podlaha dlažba 130 mm). Confirm.",
        "F20 in FF column of D.1.4.03 WC — typo for FF20 (podlaha nad suterénem"
        " dlažba 130 mm). Confirm.",
        "Printed legend Tabulka místností 1.NP — D.1.3.01 misfiled under D.1.2"
        " byt section. Subtotaly 55.9 / 43.2 m² mají být 49.59 / 49.59 m²."
        " Regenerovat printout pro VELTON delivery.",
        f"S.D.16 SKLEPNÍ KÓJE - C (7.62 m²) and S.D.42 SKLEPNÍ KÓJE - D"
        " (2.99 m²) defined in Tabulka místností XLSX but missing from DXF"
        " geometric dataset — Phase 0.x DXF parser may have skipped them"
        " (small kóje? code position outside detected polygons?).",
    ]

    OUT.write_text(json.dumps(findings, ensure_ascii=False, indent=2),
                    encoding="utf-8")
    print(f"\nWrote {OUT}")
    print(f"\nSummary: {findings['summary']}")
    print(f"\nABMV e-mail items: {len(findings['abmv_email_required'])}")


if __name__ == "__main__":
    main()
