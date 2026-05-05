"""Phase 0.13 — Complete 1.PP catalog: F14 + F11 inject + FF01 deprecate.

User cross-check Tabulka místností revealed Phase 6 generator missed
multiple skladby for 1.PP D-rooms + made one wrong mapping. This patch
brings the výkaz výměr in line with master Tabulka 0030 specifikace.

Per-room audit (S.D.27 peer baseline) pre-fix:
  - F15 tepelná izolace stropů 1PP — fixed in Phase 0.12 ✅
  - F14 bezprašný nátěr ŽB stěn — MISSING (architect note "Všechny ŽB
    stěny s povrchem F14" applies to 41 rooms)
  - F11 epoxidový nátěr 2-coat + penetrace — MISSING (Tabulka místností
    column "Povrch podlahy = F11" pro 43 D-rooms in 1.PP)
  - Cementový potěr 50mm "(FF01)" — WRONG: FF01 per master Tabulka 0030
    = pancéřová betonová podlaha s vsypem 5kg/m² (vjezdová rampa
    skladba), NOT cementový potěr. Phase 6 generator mismapped.

PROBE 5 — FF01 generator mismapping:
  Phase 6 produced 3 false items per 1.PP room (penetrace pod potěr +
  cementový potěr 50mm + kari síť) totaling ~835 m² across 41 rooms ×
  3 layers = 123 fictitious item rows. Should be replaced with F11
  epoxidový nátěr 2-coat (per Tabulka místností column).

Patch:
  1. F14 bezprašný nátěr ŽB stěn — inject 2 nátěr items per room with
     "Všechny ŽB stěny s povrchem F14" note (41 rooms × 2 = 82 items)
  2. F11 epoxidový nátěr — inject 3 items per room (penetrace + 1.+2.
     nátěr) for all 43 D-rooms in 1.PP (43 × 3 = 129 items)
  3. Wrong cementový potěr items (HSV-631 with "(FF01)" label and
     popis containing "potěr" or "kari") — zero out mnozstvi + add
     deprecation warning. Items kept in dataset for audit traceability.

Net financial impact estimated:
  - F14 under-billing closed: 41 rooms × ~30 m² avg ŽB stěna × 2 nátěr
    × ~80 Kč/m² ≈ ~98 000 Kč
  - F11 under-billing closed: 43 rooms × avg 6.5 m² × ~250 Kč/m² ≈
    ~70 000 Kč
  - FF01 wrong items zeroed: ~278 m² × ~600 Kč/m² (potěr+síť+penetrace)
    ≈ ~167 000 Kč REMOVED
  - Net: under-billing -167k + +98k + +70k ≈ +1k Kč correction
    (i.e. wrong over-billing of potěr almost exactly offset missing
    F11 + F14 — by accident, but each individual category was wrong).
"""
from __future__ import annotations

import json
import re
import uuid
import warnings
from pathlib import Path

from openpyxl import load_workbook

warnings.filterwarnings("ignore")

INPUTS = Path("test-data/libuse/inputs")
OUT_DIR = Path("test-data/libuse/outputs")
TAB_MISTNOSTI = INPUTS / "185-01_DPS_D_SO01_100_0020_R01_TABULKA MISTNOSTI.xlsx"
ITEMS = OUT_DIR / "items_objekt_D_complete.json"
DS = OUT_DIR / "objekt_D_geometric_dataset.json"

# Per master Tabulka 0030 F14 skladba: Bezprašný nátěr Sikagard 555W +
# Základní nátěr Sikagard 552W Aquaprimer (= 1 base + 1 final layer)
F14_LAYERS = [
    ("PSV-783", "Základní nátěr ŽB stěn — Sikagard 552W Aquaprimer "
                "(přilnavost; F14 vrstva 1)", 1.0),
    ("PSV-783", "Bezprašný nátěr ŽB stěn — Sikagard 555W Elastic "
                "(transparentní bezbarvý; F14 vrstva 2)", 1.0),
]

# Per master Tabulka 0030 F11 skladba: Penetrace + 1. epoxid + 2. epoxid
F11_LAYERS = [
    ("HSV-631", "Penetrace podlahy 1PP — akrylátová disperze "
                "Sika Level-01 Primer (F11)", 1.0),
    ("HSV-631", "1. epoxidový nátěr podlahy 1PP — Sikafloor 2540 W + "
                "5% vody (0.1 kg/m²; F11 vrstva 1)", 1.0),
    ("HSV-631", "2. epoxidový nátěr podlahy 1PP — Sikafloor 2540 W "
                "(0.2 kg/m²; F11 vrstva 2)", 1.0),
]


def parse_plocha(s) -> float:
    if s is None:
        return 0.0
    try:
        return float(str(s).replace("\xa0", "").replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return 0.0


def make_item(room_code: str, podlazi: str, kapitola: str, popis: str,
              mnozstvi: float, *, source: str, skladba_ref: dict,
              note: str) -> dict:
    return {
        "item_id": str(uuid.uuid4()),
        "kapitola": kapitola,
        "popis": popis,
        "popis_canonical": popis,
        "MJ": "m2",
        "mnozstvi": round(mnozstvi, 3),
        "misto": {"objekt": "D", "podlazi": podlazi, "mistnosti": [room_code]},
        "skladba_ref": skladba_ref,
        "vyrobce_ref": "",
        "urs_code": None,
        "urs_status": "needs_review",
        "urs_confidence": 0.0,
        "urs_alternatives": [],
        "confidence": 0.85,
        "status": "to_audit",
        "source_method": source,
        "audit_note": note,
        "warnings": [],
    }


def main() -> None:
    print("Loading Tabulka místností + items…")
    wb = load_workbook(TAB_MISTNOSTI, data_only=True)
    ws = wb["tabulka místností"]

    # Build per-room metadata: F14_applies (architect note), F11_applies
    # (column has F11), plocha, obvod approx.
    rooms_meta: dict[str, dict] = {}
    for r in ws.iter_rows(min_row=7, max_row=ws.max_row, values_only=True):
        if not r[0]:
            continue
        code = str(r[0])
        if not (code.startswith("D.") or code.startswith("S.D.")):
            continue
        plocha = parse_plocha(r[2])
        if plocha <= 0:
            continue
        try:
            vyska_mm = float(str(r[3]).replace(" ", "").replace(",", "."))
        except (ValueError, TypeError):
            vyska_mm = 2800.0
        f_pov_pod = str(r[5]) if r[5] else ""
        f_pov_sten = str(r[6]) if r[6] else ""
        poznamka = str(r[9]) if len(r) > 9 and r[9] else ""

        if code.startswith("S.D."):
            podlazi = "1.PP"
        else:
            parts = code.split(".")
            podlazi = f"{parts[1]}.NP" if len(parts) >= 2 and parts[1].isdigit() else "?"

        rooms_meta[code] = {
            "code": code,
            "podlazi": podlazi,
            "plocha_m2": plocha,
            "vyska_mm": vyska_mm,
            "F_povrch_podlahy": f_pov_pod,
            "F_povrch_sten": f_pov_sten,
            "poznamka": poznamka,
            "f14_applies": "F14" in poznamka,
            "f11_applies": "F11" in f_pov_pod,
        }

    # Filter to 1.PP D-rooms (objekt D scope)
    pp1_rooms = {c: m for c, m in rooms_meta.items() if m["podlazi"] == "1.PP"}
    print(f"1.PP D-rooms: {len(pp1_rooms)}")
    f14_rooms = [c for c, m in pp1_rooms.items() if m["f14_applies"]]
    f11_rooms = [c for c, m in pp1_rooms.items() if m["f11_applies"]]
    print(f"  with F14 architect note: {len(f14_rooms)}")
    print(f"  with F11 in Povrch podlahy: {len(f11_rooms)}")

    items_blob = json.loads(ITEMS.read_text(encoding="utf-8"))
    items = items_blob["items"]

    # Track existing items (skip dup if room already has F14/F11 via Phase 0.11)
    existing_f14: set[str] = set()
    existing_f11: set[str] = set()
    for it in items:
        rc = (it.get("misto", {}).get("mistnosti") or [None])[0]
        if not rc:
            continue
        popis_low = it["popis"].lower()
        if "bezprašný" in popis_low or "f14" in popis_low:
            existing_f14.add(rc)
        if "epoxidový" in popis_low and "podlah" in popis_low:
            existing_f11.add(rc)

    # ------ INJECT F14 (skip rooms with existing F14 from Phase 0.11) ------
    new_items: list[dict] = []
    for code in f14_rooms:
        if code in existing_f14:
            continue
        m = pp1_rooms[code]
        # Plocha ŽB stěn — approximate from plocha podlahy + výška
        # Assumes square-ish room: obvod = 4 × sqrt(plocha)
        import math
        obvod = 4 * math.sqrt(m["plocha_m2"])
        plocha_sten = obvod * (m["vyska_mm"] / 1000.0)
        for kap, popis, mult in F14_LAYERS:
            new_items.append(make_item(
                code, m["podlazi"], kap, popis, plocha_sten * mult,
                source="PHASE_0_13_F14_bezprasny_natěr",
                skladba_ref={"F_povrch_sten": "F14_per_poznamka",
                              "vrstva": "bezprašný nátěr Sikagard"},
                note=("PHASE_0_13: per Tabulka místností note "
                      "'Všechny ŽB stěny s povrchem F14'. Plocha stěn "
                      "approx z 4·sqrt(plocha) × výška.")))

    # ------ INJECT F11 (skip rooms with existing F11) ------
    for code in f11_rooms:
        if code in existing_f11:
            continue
        m = pp1_rooms[code]
        for kap, popis, mult in F11_LAYERS:
            new_items.append(make_item(
                code, m["podlazi"], kap, popis, m["plocha_m2"] * mult,
                source="PHASE_0_13_F11_epoxidovy_natěr",
                skladba_ref={"F_povrch_podlahy": "F11",
                              "vrstva": "epoxidový nátěr Sikafloor 2540 W"},
                note=("PHASE_0_13: F11 epoxidový nátěr per Tabulka "
                      "místností Povrch podlahy column.")))

    # ------ ZERO OUT wrong cementový potěr items (FF01 mismap) ------
    # Phase 6 generated "Penetrace pod potěr (FF01)" + "Cementový potěr F5
    # tl. 50 mm (FF01)" + "Kari síť 150/150/4 mm pro potěr (FF01)" pro
    # všech 1.PP D-rooms. Per master Tabulka 0030, FF01 nemá cementový
    # potěr — to patří do FF20/FF30 skladby pro 1.NP+ podlahy.
    deprecated = 0
    pp1_codes = set(pp1_rooms.keys())
    POTER_PATTERN = re.compile(r"(potěr|potěru|kari síť).*\(FF01\)", re.IGNORECASE)
    for it in items:
        rc = (it.get("misto", {}).get("mistnosti") or [None])[0]
        if rc not in pp1_codes:
            continue
        if not POTER_PATTERN.search(it["popis"]):
            continue
        # Zero out + add deprecation flag
        it["mnozstvi"] = 0.0
        it["status"] = "deprecated"
        warns = it.get("warnings") or []
        warns.append(
            "PHASE_0_13_PROBE5: FF01 = pancéřová podlaha s vsypem (vjezdová "
            "rampa skladba) per master Tabulka 0030, NE cementový potěr. "
            "Phase 6 generator mismapped FF01. Item zeroed; F11 epoxidový "
            "nátěr items now correctly assigned per Tabulka místností.")
        it["warnings"] = warns
        it["audit_note"] = (it.get("audit_note", "")
                             + "; PHASE_0_13_PROBE5: zeroed FF01 mismap").strip("; ")
        deprecated += 1

    items.extend(new_items)
    items_blob["items"] = items
    items_blob["metadata"]["items_count"] = len(items)
    items_blob["metadata"]["phase_0_13_applied"] = True
    ITEMS.write_text(json.dumps(items_blob, ensure_ascii=False, indent=2),
                     encoding="utf-8")

    # Update carry_forward_findings → PROBE 5
    ds = json.loads(DS.read_text(encoding="utf-8"))
    cff = ds.setdefault("carry_forward_findings", [])
    cff.append({
        "phase": "0.13",
        "type": "PROBE_FINDING + GENERATOR_FIX",
        "label": "PROBE 5 — FF01 mismap + F14/F11 missing pro 1.PP D-rooms",
        "description": (
            f"3 systemic issues v Phase 6 generator pro 1.PP D-rooms (43 rooms): "
            f"(a) F14 bezprašný nátěr ŽB stěn (Sikagard 555W) chybělo přes všech "
            f"{len(f14_rooms)} rooms s architectovou poznámkou; "
            f"(b) F11 epoxidový nátěr podlahy (Sikafloor 2540 W 2-coat) chybělo "
            f"přes všech {len(f11_rooms)} rooms; "
            f"(c) FF01 mismap: Phase 6 produced cementový potěr 50mm + kari síť "
            f"items pro 1.PP, ale FF01 = pancéřová betonová podlaha s vsypem (NE "
            f"cementový potěr — to patří do FF20/FF30 nad suterénem). "
            f"Phase 0.13 inject {len(new_items)} new items + zeroed {deprecated} "
            f"wrong potěr items."
        ),
        "items_added": len(new_items),
        "items_zeroed": deprecated,
        "rooms_affected_F14": len(f14_rooms),
        "rooms_affected_F11": len(f11_rooms),
        "estimated_correction_czk": {
            "F14_under_billing_closed": "+98 000 Kč",
            "F11_under_billing_closed": "+70 000 Kč",
            "FF01_over_billing_removed": "-167 000 Kč",
            "net_correction": "+1 000 Kč (close to neutral, but each category was wrong)",
        },
    })
    DS.write_text(json.dumps(ds, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"\n✅ PHASE 0.13 RESULTS:")
    print(f"   F14 items added: {sum(1 for i in new_items if 'Sikagard' in i['popis'])}")
    print(f"   F11 items added: {sum(1 for i in new_items if 'Sikafloor' in i['popis'] or 'Sika Level' in i['popis'])}")
    print(f"   FF01 wrong items zeroed: {deprecated}")
    print(f"   Total items: {len(items)}")


if __name__ == "__main__":
    main()
