"""Phase 5 step 1 — parse + normalize starý výkaz výměr.

Walks every sheet in Vykaz_vymer_stary.xlsx, extracts item rows
(code + popis + MJ + množství + jednotková cena + celková cena),
normalises text for fuzzy matching, and saves a flat JSON.
"""
from __future__ import annotations

import json
import re
import warnings
from pathlib import Path

import openpyxl
warnings.filterwarnings("ignore", message=".*header or footer.*")

VV = Path("test-data/libuse/inputs/Vykaz_vymer_stary.xlsx")
OUT = Path("test-data/libuse/outputs/stary_vv_normalized.json")

# Sheets that contain construction items (skip rekapitulace + krycí list)
RELEVANT_SHEET_PREFIXES = ("100", "D1.", "IO", "S0", "SO", "00")
# Patterns
CODE_PATTERNS = [
    re.compile(r"^[0-9]{6,12}$"),                  # ÚRS standard 12-digit
    re.compile(r"^R-?[0-9A-Z]{5,12}$", re.IGNORECASE),  # R-prefixed
    re.compile(r"^[A-Z]{2,5}_?[0-9A-Z]{3,12}$"),   # ROZPOCET, etc.
]
SECTION_HDR_RE = re.compile(r"^\s*([0-9]{2,3})\s*-\s*(.+)$")  # "712 - Povlakové krytiny"
NUM_RE = re.compile(r"^-?\d+([.,]\d+)?$")


def _val(c):
    if c is None:
        return ""
    return str(c).strip()


def _num(c):
    if c is None:
        return None
    s = str(c).strip()
    if not s:
        return None
    try:
        return float(s.replace(",", ".").replace(" ", "").replace("\xa0", ""))
    except ValueError:
        return None


def normalize_popis(s: str) -> str:
    """Lowercase, strip punctuation, collapse whitespace."""
    if not s:
        return ""
    out = s.lower()
    # remove non-letter/digit/space (Czech-aware: keep accents)
    out = re.sub(r"[^\w\sáčďéěíňóřšťúůýžÁČĎÉĚÍŇÓŘŠŤÚŮÝŽ]", " ", out)
    out = re.sub(r"\s+", " ", out).strip()
    return out


def is_code(s: str) -> bool:
    if not s:
        return False
    for p in CODE_PATTERNS:
        if p.match(s.strip()):
            return True
    return False


def parse_sheet(ws, sheet_name: str) -> tuple[list[dict], dict]:
    """Walk the sheet, identify item rows. Each item row has:
       - some column with code (6+ char alphanumeric)
       - a popis cell (string > 15 chars)
       - 5+ numeric columns (qty, unit_price, total, mass, …)
    Returns (items, diagnostics).
    """
    items = []
    current_section = None
    rows_scanned = 0
    rows_with_code = 0
    rows_kept = 0

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        rows_scanned += 1
        if not row:
            continue
        cells = [_val(c) for c in row]
        # Section header detection: long string row with no other content
        non_empty_strs = [c for c in cells if c and len(c) > 2]
        if len(non_empty_strs) <= 2:
            for c in non_empty_strs:
                m = SECTION_HDR_RE.match(c)
                if m:
                    current_section = f"{m.group(1)} - {m.group(2)[:40]}"
                    break

        # Find code in row (any cell)
        code = None
        code_col = None
        for j, c in enumerate(cells):
            if is_code(c):
                code = c
                code_col = j
                break
        if code is None:
            continue
        rows_with_code += 1

        # Find popis: first long string ≠ code, ≠ header keywords
        popis = ""
        for c in cells:
            s = c.strip()
            if s == code:
                continue
            if len(s) >= 12 and not s.startswith("Bytový") and not s.startswith("KRYCÍ") \
                and not s.startswith("Stavba") and not s.startswith("Objekt") \
                and not s.startswith("NEPLATIT") and not s.startswith("Soupis"):
                popis = s
                break
        if not popis:
            continue
        # MJ — tiny string token
        mj = ""
        mj_kw = ("m", "m2", "m³", "m3", "kus", "ks", "bm", "M", "M2", "M3",
                 "t", "kg", "l", "ks", "soubor", "soub", "h", "hod", "ha")
        for c in cells:
            s = c.strip()
            if s in mj_kw or (len(s) <= 4 and s.lower() in (k.lower() for k in mj_kw)):
                mj = s
                break

        # Numeric extraction
        nums = [n for n in (_num(c) for c in cells) if n is not None and n > 0]

        # Quantity = first positive number after the code column;
        # cena_jedn = next; cena_celkem = largest of remaining or third
        if not nums:
            continue
        mnozstvi = nums[0]
        cena_jedn = nums[1] if len(nums) > 1 else None
        cena_celkem = nums[2] if len(nums) > 2 else None

        items.append({
            "sheet": sheet_name,
            "row": row_idx,
            "section": current_section,
            "code": code,
            "popis": popis,
            "popis_normalized": normalize_popis(popis),
            "MJ": mj,
            "mnozstvi": mnozstvi,
            "cena_jedn_kc": cena_jedn,
            "cena_celkem_kc": cena_celkem,
        })
        rows_kept += 1

    diag = {
        "rows_scanned": rows_scanned,
        "rows_with_code": rows_with_code,
        "rows_kept": rows_kept,
    }
    return items, diag


def main() -> None:
    wb = openpyxl.load_workbook(str(VV), data_only=True)
    all_items: list[dict] = []
    per_sheet_diag: dict[str, dict] = {}
    for sn in wb.sheetnames:
        if not sn.startswith(RELEVANT_SHEET_PREFIXES):
            # Skip rekapitulace / sadové úpravy etc. Keep architektonicko first.
            if sn in ("Rekapitulace stavby", "Seznam figur"):
                continue
        ws = wb[sn]
        items, diag = parse_sheet(ws, sn)
        all_items.extend(items)
        per_sheet_diag[sn] = diag

    # Cumulative stats
    by_sheet = {}
    by_section = {}
    by_mj = {}
    for it in all_items:
        by_sheet[it["sheet"]] = by_sheet.get(it["sheet"], 0) + 1
        if it["section"]:
            by_section[it["section"]] = by_section.get(it["section"], 0) + 1
        if it["MJ"]:
            by_mj[it["MJ"]] = by_mj.get(it["MJ"], 0) + 1

    out = {
        "metadata": {
            "source": str(VV),
            "sheets_processed": list(per_sheet_diag.keys()),
            "items_count": len(all_items),
            "items_per_sheet": by_sheet,
            "items_per_section_top10": dict(sorted(by_section.items(), key=lambda x: -x[1])[:10]),
            "items_per_mj": dict(sorted(by_mj.items(), key=lambda x: -x[1])),
            "per_sheet_diagnostics": per_sheet_diag,
        },
        "items": all_items,
    }
    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUT} ({OUT.stat().st_size:,} bytes)")
    print(f"Total items: {len(all_items)}")
    print()
    print("Items per sheet:")
    for s, n in by_sheet.items():
        print(f"  {s:48s} {n:>5}")
    print()
    print("Top 10 sections:")
    for s, n in sorted(by_section.items(), key=lambda x: -x[1])[:10]:
        print(f"  {s[:60]:60s} {n:>4}")


if __name__ == "__main__":
    main()
