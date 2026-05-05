"""Phase 6.1 step 1a — Deep Tabulka structure analysis.

Looks for per-objekt count columns (or room-mapping columns we can use
to derive per-objekt counts) in every Tabulka. Documents findings.
"""
from __future__ import annotations

import json
import re
import warnings
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
warnings.filterwarnings("ignore", message=".*header or footer.*")

INP = Path("test-data/libuse/inputs")
OUT = Path("test-data/libuse/outputs/tabulky_structure_analysis.md")
OUT_JSON = Path("test-data/libuse/outputs/tabulky_perobjekt_counts.json")


def is_d_room(s: str) -> bool:
    """D-objekt room code: D.x.x.xx or S.D.xx (sklep)."""
    if not s:
        return False
    s = str(s).strip()
    if s.startswith("D."):
        return True
    parts = s.split(".")
    return len(parts) == 3 and parts[0] == "S" and parts[1] == "D"


def objekt_of_room(s: str) -> str | None:
    if not s:
        return None
    s = str(s).strip()
    if re.match(r"^[A-D]\.", s):
        return s[0]
    parts = s.split(".")
    if len(parts) == 3 and parts[0] == "S" and parts[1] in "ABCD":
        return parts[1]
    return None


lines: list[str] = []
lines.append("# Tabulky structure analysis — per-objekt count investigation")
lines.append("")
lines.append("Phase 6.1 bug fix — find which Tabulky carry per-objekt data we can "
             "exploit instead of uniform 0.25 D-share.")
lines.append("")

result = {}

# ---------------------------------------------------------------------------
# 1) Tabulka dveří — has from/to room columns!
# ---------------------------------------------------------------------------
lines.append("## Tabulka dveří")
lines.append("")
p = INP / "185-01_DPS_D_SO01_100_0041_TABULKA DVERI.xlsx"
wb = openpyxl.load_workbook(str(p), data_only=True)
ws = wb["tab dvere"]
# Header at row 6: cols A=#, B=Type Mark, C=from room, D=to room, etc.
# Data starts row 8
d_count_per_type: Counter = Counter()
total_per_type: Counter = Counter()
for row in ws.iter_rows(min_row=8, values_only=True):
    type_mark = str(row[1] or "").strip()
    if not re.match(r"^D\d{1,3}$", type_mark):
        continue
    from_room = str(row[2] or "").strip()
    to_room = str(row[3] or "").strip()
    total_per_type[type_mark] += 1
    if is_d_room(from_room) or is_d_room(to_room):
        d_count_per_type[type_mark] += 1

lines.append("**Structure**: tab dvere sheet has FROM-room (col C) + TO-room (col D) per door instance!")
lines.append("Each row = ONE door with explicit room mapping → exact per-objekt count.")
lines.append("")
lines.append(f"Total D## doors across komplex (rows 8+): **{sum(total_per_type.values())}**")
lines.append(f"Doors in objekt D (from_room or to_room is D.* / S.D.*): **{sum(d_count_per_type.values())}**")
lines.append("")
lines.append("| D-code | Komplex | Objekt D | D-share |")
lines.append("|---|---:|---:|---:|")
for t in sorted(total_per_type, key=lambda x: -total_per_type[x]):
    k = total_per_type[t]
    d = d_count_per_type.get(t, 0)
    share = d / k if k else 0
    lines.append(f"| `{t}` | {k} | {d} | {share:.2%} |")
lines.append("")
result["dvere"] = {
    "has_per_objekt_data": True,
    "method": "from_room / to_room columns → filter by D.* / S.D. prefix",
    "komplex_total_per_type": dict(total_per_type),
    "objekt_d_per_type": dict(d_count_per_type),
}

# ---------------------------------------------------------------------------
# 2) Tabulka oken — has Počet (komplex total), no per-objekt
# ---------------------------------------------------------------------------
lines.append("## Tabulka oken")
lines.append("")
p = INP / "185-01_DPS_D_SO01_100_0042_TABULKA OKEN.xlsx"
wb = openpyxl.load_workbook(str(p), data_only=True)
ws = wb["tabulka"]
total_per_w: Counter = Counter()
for row in ws.iter_rows(min_row=8, values_only=True):
    type_mark = str(row[0] or "").strip()
    if not re.match(r"^W\d{1,3}$", type_mark):
        continue
    pocet = row[19] if len(row) > 19 else None
    try:
        n = int(float(str(pocet)))
    except (ValueError, TypeError):
        n = None
    if n:
        total_per_w[type_mark] = n
lines.append("**Structure**: tabulka sheet col 20 (T) = 'Počet' komplex total. NO per-objekt column.")
lines.append("**Solution**: use DXF spatial counts from objekt D drawings (Phase 1 aggregate "
             "`windows_by_type_code`).")
lines.append("")
lines.append(f"Total W## komplex: **{sum(total_per_w.values())}**")
lines.append("")
lines.append("| W-code | Komplex | DXF objekt D count (Phase 1) |")
lines.append("|---|---:|---:|")
# Load Phase 1 windows aggregate for cross-reference
ds = json.loads(Path("test-data/libuse/outputs/objekt_D_geometric_dataset.json").read_text(encoding="utf-8"))
dxf_w = ds["aggregates"]["windows_by_type_code"]
for w in sorted(total_per_w, key=lambda x: -total_per_w[x]):
    lines.append(f"| `{w}` | {total_per_w[w]} | {dxf_w.get(w, 0)} |")
lines.append("")
result["okna"] = {
    "has_per_objekt_data": False,
    "method": "use DXF spatial counts from objekt D drawings (Phase 1 windows_by_type_code)",
    "komplex_total_per_type": dict(total_per_w),
    "objekt_d_per_type_from_dxf": dxf_w,
}

# ---------------------------------------------------------------------------
# 3) Tabulka klempířských TP## — only komplex total, no per-objekt
# ---------------------------------------------------------------------------
lines.append("## Tabulka klempířských TP##")
lines.append("")
p = INP / "185-01_DPS_D_SO01_100_0060_R01_TABULKA KLEMPIRSKYCH PRVKU.xlsx"
wb = openpyxl.load_workbook(str(p), data_only=True)
ws = wb["tabulka"]
tp_komplex: dict[str, dict] = {}
for row in ws.iter_rows(min_row=7, values_only=True):
    code = str(row[0] or "").strip()
    if not re.match(r"^TP\s*\d", code, re.IGNORECASE):
        continue
    code_clean = re.sub(r"\s+", "", code).upper()
    nazev = str(row[1] or "")
    umisteni = str(row[2] or "")
    qty = row[7]
    try:
        qty_n = float(str(qty).replace(",", ".")) if qty else None
    except (ValueError, TypeError):
        qty_n = None
    tp_komplex[code_clean] = {
        "nazev": nazev[:40],
        "umisteni": umisteni[:40],
        "mnozstvi_komplex": qty_n,
    }
lines.append("**Structure**: cols A=code, B=Název, C=Umístění (location text only — not per-objekt).")
lines.append(f"Total TP## codes: **{len(tp_komplex)}**")
lines.append("")
lines.append("Per-objekt data: ❌ not in Tabulka. Possible derivation:")
lines.append("- Some TP items are roof-located (TP01 zaatikové žlaby) → split by per-objekt roof obvod")
lines.append("- Some are facade (TP25 dešťový svod) → 4 svody / 4 objekty = 1/objekt heuristic")
lines.append("- Most are uniform 0.25 D-share fallback")
lines.append("")
result["klempirske"] = {
    "has_per_objekt_data": False,
    "method": "fallback uniform 0.25 D-share; refine via DXF roof drawing in future",
    "komplex": tp_komplex,
}

# ---------------------------------------------------------------------------
# 4) Tabulka zámečnických LP## — only komplex
# ---------------------------------------------------------------------------
p = INP / "185-01_DPS_D_SO01_100_0050_R01_TABULKA ZAMECNICKYCH VYROBKU.xlsx"
wb = openpyxl.load_workbook(str(p), data_only=True)
ws = wb["tabulka"]
lp_komplex: dict[str, dict] = {}
for row in ws.iter_rows(min_row=7, values_only=True):
    code = str(row[0] or "").strip()
    if not re.match(r"^LP\s*\d", code, re.IGNORECASE):
        continue
    code_clean = re.sub(r"\s+", "", code).upper()
    qty = row[7]
    try:
        qty_n = float(str(qty).replace(",", ".")) if qty else None
    except (ValueError, TypeError):
        qty_n = None
    lp_komplex[code_clean] = {
        "umisteni": str(row[2] or "")[:50],
        "mnozstvi_komplex": qty_n,
    }
lines.append("## Tabulka zámečnických LP##")
lines.append("")
lines.append("**Structure**: same as klempířské. Cols A=code, C=Umístění (text), H=Množství komplex.")
lines.append(f"Total LP## codes: **{len(lp_komplex)}**")
lines.append("")
lines.append("Per-objekt data: ❌ not in Tabulka. **DXF segment tag spatial counts** recoverable from Phase 1 + Phase 2 inventory.")
lines.append("")
result["zamecnicke"] = {
    "has_per_objekt_data": False,
    "method": "fallback uniform 0.25; refine via DXF segment tag spatial joining (LP## tags found on A-FLOR-HRAL-IDEN per Phase 1)",
    "komplex": lp_komplex,
}

# ---------------------------------------------------------------------------
# 5) Tabulka ostatních OP## — only komplex
# ---------------------------------------------------------------------------
p = INP / "185-01_DPS_D_SO01_100_0080_R02 - TABULKA OSTATNICH PRVKU.xlsx"
wb = openpyxl.load_workbook(str(p), data_only=True)
ws = wb["tabulka"]
op_komplex: dict[str, dict] = {}
for row in ws.iter_rows(min_row=7, values_only=True):
    code = str(row[0] or "").strip()
    if not re.match(r"^OP\s*\d", code, re.IGNORECASE):
        continue
    code_clean = re.sub(r"\s+", "", code).upper()
    qty = row[6]
    try:
        qty_n = float(str(qty).replace(",", ".")) if qty else None
    except (ValueError, TypeError):
        qty_n = None
    op_komplex[code_clean] = {
        "nazev": str(row[1] or "")[:40],
        "umisteni": str(row[2] or "")[:40],
        "mnozstvi_komplex": qty_n,
        "mj": str(row[5] or "").strip(),
    }
lines.append("## Tabulka ostatních OP##")
lines.append("")
lines.append(f"Total OP## codes: **{len(op_komplex)}**")
lines.append("**Structure**: cols A=code, B=Název, C=Umístění (text), F=MJ, G=Množství komplex.")
lines.append("")
lines.append("Per-objekt data: ❌ not in Tabulka. **DXF segment tag spatial counts** recoverable from Phase 1 + Phase 2 inventory (A-GENM-____-IDEN with 48 OP## split codes per 1.NP D půdorys).")
lines.append("")
result["ostatni"] = {
    "has_per_objekt_data": False,
    "method": "DXF spatial count per OP## occurrence on objekt-D drawings",
    "komplex": op_komplex,
}

# ---------------------------------------------------------------------------
# 6) Tabulka překladů LI## — only komplex
# ---------------------------------------------------------------------------
p = INP / "185-01_DPS_D_SO01_100_0070_R01_TABULKA PREKLADU.xlsx"
wb = openpyxl.load_workbook(str(p), data_only=True)
ws = wb["tabulka LI"]
li_komplex: dict[str, dict] = {}
for row in ws.iter_rows(min_row=7, values_only=True):
    prefix = str(row[0] or "").strip()
    suffix = str(row[1] or "").strip()
    if prefix.upper() == "LI" and suffix:
        code = f"LI{suffix.zfill(2)}"
        try:
            qty_n = float(str(row[7]).replace(",", "."))
        except (ValueError, TypeError):
            qty_n = None
        li_komplex[code] = {
            "nazev": str(row[2] or "")[:40],
            "umisteni": str(row[3] or "")[:40],
            "mnozstvi_komplex": qty_n,
        }
lines.append("## Tabulka překladů LI##")
lines.append("")
lines.append(f"Total LI## codes: **{len(li_komplex)}**")
lines.append("Per-objekt data: ❌ not in Tabulka. DXF spatial count per A-GENM-____-IDEN LI## tags (Phase 1 inventory: 32 LI tags in 1.NP D).")
lines.append("")
result["preklady"] = {
    "has_per_objekt_data": False,
    "method": "DXF spatial count per LI## occurrence on objekt-D drawings",
    "komplex": li_komplex,
}

OUT.write_text("\n".join(lines), encoding="utf-8")
OUT_JSON.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"Wrote {OUT}")
print(f"Wrote {OUT_JSON}")
print()
print("Headline findings:")
print(f"  Tabulka dveří: ✅ has per-objekt mapping (from_room/to_room cols) → "
       f"D-doors {sum(d_count_per_type.values())} of komplex {sum(total_per_type.values())}")
print(f"  Tabulka oken:  use DXF (Phase 1) — komplex {sum(total_per_w.values())}, "
       f"DXF D = {sum(dxf_w.values())}")
print(f"  Klempířské TP##: {len(tp_komplex)} codes, NO per-objekt → uniform fallback")
print(f"  Zámečnické LP##: {len(lp_komplex)} codes, NO per-objekt → DXF spatial fallback")
print(f"  Ostatní OP##:   {len(op_komplex)} codes, NO per-objekt → DXF spatial count")
print(f"  Překlady LI##:  {len(li_komplex)} codes, NO per-objekt → DXF spatial count")
