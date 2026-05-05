"""Phase 0.9 — Door per-room ownership from Tabulka dveří 0041 (Bug #2 fix).

Replaces nearest-room geometric clustering (which fails in 1.PP cellar
cluster: 11 doors → 4 m² S.D.27 instead of real 1).

Each row in ``tab dvere`` has ``z místnosti`` + ``do místnosti`` columns.
Door physically belongs to BOTH rooms (perimeter passes through both
walls), so it's added to both ownership lists.

Output: ``test-data/libuse/outputs/objekt_D_doors_ownership.json``
"""
from __future__ import annotations

import json
import re
import warnings
from pathlib import Path

from openpyxl import load_workbook

warnings.filterwarnings("ignore")

XLSX = Path(
    "test-data/libuse/inputs/"
    "185-01_DPS_D_SO01_100_0041_TABULKA DVERI.xlsx"
)
OUT = Path("test-data/libuse/outputs/objekt_D_doors_ownership.json")

# Garage gates / oversized non-špaleta openings → exclude from HSV-612 entirely.
# D05 (Hoermann garážová vrata 5700×2100), gates with explicit Hoermann/Gate keywords.
GARAGE_TYPES = {"D05"}
GARAGE_WIDTH_THRESHOLD_MM = 3000


def is_garage_gate(door: dict) -> bool:
    """Bug #3 — exclude garage gates from HSV-612 calculation."""
    if door.get("typ") in GARAGE_TYPES:
        return True
    sirka = door.get("sirka_otvoru_mm") or 0
    if sirka and sirka > GARAGE_WIDTH_THRESHOLD_MM:
        return True
    return False


def main() -> None:
    print(f"Loading XLSX: {XLSX.name}")
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["tab dvere"]

    # Header row at index 6 (1-based); data starts row 7 (placeholder) → 8.
    doors: list[dict] = []
    for i, row in enumerate(ws.iter_rows(min_row=7, max_row=ws.max_row, values_only=True), 7):
        cislo = row[0] if row else None
        if not cislo:
            continue
        # Skip placeholder row "xxx"
        if isinstance(cislo, str) and cislo.strip().lower() in ("xxx", ""):
            continue
        try:
            cislo_str = str(cislo).strip()
            if not re.fullmatch(r"\d{1,3}", cislo_str):
                continue
        except Exception:
            continue
        typ = str(row[1]).strip() if row[1] else None
        z_mistnoti = str(row[2]).strip() if row[2] else None
        do_mistnoti = str(row[3]).strip() if row[3] else None
        sirka_otvoru = row[11] if len(row) > 11 else None
        vyska_otvoru = row[12] if len(row) > 12 else None
        try:
            sirka_int = int(round(float(sirka_otvoru))) if sirka_otvoru else None
        except (TypeError, ValueError):
            sirka_int = None
        try:
            vyska_int = int(round(float(vyska_otvoru))) if vyska_otvoru else None
        except (TypeError, ValueError):
            vyska_int = None
        doors.append({
            "cislo": cislo_str,
            "typ": typ,
            "z_mistnoti": z_mistnoti,
            "do_mistnoti": do_mistnoti,
            "sirka_otvoru_mm": sirka_int,
            "vyska_otvoru_mm": vyska_int,
        })

    print(f"Parsed {len(doors)} doors from Tabulka 0041")

    # Build ownership: room_code → list[door], add to both z + do rooms
    ownership: dict[str, list[dict]] = {}
    garage_gate_count = 0
    for d in doors:
        for room_field in ("z_mistnoti", "do_mistnoti"):
            rc = d.get(room_field)
            if not rc:
                continue
            ownership.setdefault(rc, []).append({
                "cislo": d["cislo"],
                "typ": d["typ"],
                "sirka_otvoru_mm": d["sirka_otvoru_mm"],
                "vyska_otvoru_mm": d["vyska_otvoru_mm"],
                "is_garage_gate": is_garage_gate(d),
                "from_room": d["z_mistnoti"],
                "to_room": d["do_mistnoti"],
            })
        if is_garage_gate(d):
            garage_gate_count += 1

    # Filter to D-objekt rooms (S.D.* + D.x.x.xx)
    d_ownership = {k: v for k, v in ownership.items()
                    if k and (k.startswith("D.") or k.startswith("S.D."))}
    print(f"Ownership: {len(d_ownership)} D-rooms with at least 1 door")
    print(f"Garage gates flagged for HSV-612 exclusion: {garage_gate_count}")

    # Spot-check verification (Bug #2 ground truth)
    for room in ("S.D.27", "S.D.40"):
        cnt = len(d_ownership.get(room, []))
        print(f"  {room}: {cnt} door(s) — expected 1")
        for d in d_ownership.get(room, []):
            print(f"    {d['cislo']} {d['typ']} {d['sirka_otvoru_mm']}×{d['vyska_otvoru_mm']} "
                  f"({d['from_room']}→{d['to_room']}) garage={d['is_garage_gate']}")

    OUT.write_text(json.dumps({
        "source": "Tabulka 0041 z_mistnoti/do_mistnoti ownership",
        "method": "explicit_table_lookup",
        "garage_exclusion": {
            "rule": f"D05 OR sirka_otvoru_mm > {GARAGE_WIDTH_THRESHOLD_MM}",
            "count_flagged": garage_gate_count,
        },
        "ownership": d_ownership,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nWrote {OUT}")


if __name__ == "__main__":
    main()
