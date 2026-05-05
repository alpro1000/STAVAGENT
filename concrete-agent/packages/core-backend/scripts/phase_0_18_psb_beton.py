"""Phase 0.18 — Polystyrenbeton PSB 50 instalační vrstva 40mm.

⚠️ UNIT CONVERSION m² → m³ — kritické per task spec.

Per master Tabulka 0030 (verified rows 25, 33, 41, 49, 56):
  FF10 / FF20 / FF21 / FF30 / FF31 ALL contain:
    "Instalační vrstva 40mm — Polystyrenbeton PSB 50,
     max. 400 kg/m³ (v suchém stavu) — Liapor Mix"
  Thickness: 40 mm (= 0.04 m) ALL 5 skladby.

Per starý VV row 939 kód 128:
  Mj: m³ (verified col 6)
  Cena: 4830 Kč/m³
  Komplex: 146.899 m³

D scope per user clarification: 56 rooms with FF10/20/21/30/31 (same
set as Phase 0.16 kročejová).

UNIT CONVERSION (per-room):
  surface_m2 = plocha_m2 z Tabulky místností 0020
  thickness_m = 0.04 (verified Tabulka 0030)
  volume_m3 = surface_m2 × 0.04
  cena = volume_m3 × 4830 Kč/m³

PE folie separace (FF skladby vrstva "Separece - PE folie"):
  ⛔ SKIP per task STOP CONDITION — žádný separátní ÚRS kód v starém VV
  found (PE folie thickness uváděna jako "-" v Tabulce skladeb,
  pravděpodobně součást ceny PSB nebo cementového potěru). Manual
  review with ABMV recommended before adding standalone item.
"""
from __future__ import annotations

import json
import re
import uuid
import warnings
from pathlib import Path

from openpyxl import load_workbook

warnings.filterwarnings("ignore")

INPUTS = Path("test-data/libuse/inputs")
OUT_DIR = Path("test-data/libuse/outputs")
TAB_MISTNOSTI = INPUTS / "185-01_DPS_D_SO01_100_0020_R01_TABULKA MISTNOSTI.xlsx"
TAB_SKLADEB = INPUTS / "185-01_DPS_D_SO01_100_0030_R01_TABULKA SKLADEB A POVRCHU_R01.xlsx"
ITEMS = OUT_DIR / "items_objekt_D_complete.json"

URS_KOD = "631R128"
UNIT_PRICE_KC_PER_M3 = 4830.0
MJ = "m3"
THICKNESS_MM = 40
THICKNESS_M = 0.04


def parse_plocha(s) -> float:
    if s is None:
        return 0.0
    try:
        return float(str(s).replace("\xa0", "").replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return 0.0


def is_objekt_D(c: str) -> bool:
    return (c.startswith("D.1.") or c.startswith("D.2.")
            or c.startswith("D.3.") or c.startswith("S.D."))


def verify_thickness() -> bool:
    """Verify all FF20/21/30/31 mají 40mm PSB per master Tabulka 0030."""
    wb = load_workbook(TAB_SKLADEB, data_only=True)
    ws = wb["skladby_podlah"]
    found = []
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        cells = [str(c) if c is not None else "" for c in r[:8]]
        text = " ".join(cells).lower()
        if "polysterén" in text or "polysteren" in text:
            try:
                tl = int(float(cells[3])) if cells[3] else 0
            except (ValueError, TypeError):
                continue
            found.append(tl)
    return len(found) >= 4 and all(t == THICKNESS_MM for t in found)


def main() -> None:
    print("=" * 72)
    print("PHASE 0.18 — Polystyrenbeton PSB 50 instalační vrstva 40mm")
    print("=" * 72)

    if not verify_thickness():
        print("⛔ STOP — PSB tloušťka NEJE 40mm v Tabulce skladeb!")
        return
    print(f"✓ Tabulka skladeb: PSB 50 = {THICKNESS_MM} mm verified")
    print(f"✓ ÚRS reference: starý VV row 939 kód 128, mj=m³, "
          f"{UNIT_PRICE_KC_PER_M3} Kč/m³, komplex 146.899 m³")
    print(f"✓ PE folie separace: SKIP (no separate ÚRS in old VV)")

    wb = load_workbook(TAB_MISTNOSTI, data_only=True)
    ws = wb["tabulka místností"]
    target: list[dict] = []
    pat = re.compile(r"FF(10|20|21|30|31)\b")
    for r in ws.iter_rows(min_row=7, max_row=ws.max_row, values_only=True):
        if not r[0]: continue
        code = str(r[0])
        if not is_objekt_D(code): continue
        ff = str(r[4]) if r[4] else ""
        if not pat.search(ff): continue
        plocha = parse_plocha(r[2])
        if plocha <= 0: continue
        if code.startswith("S.D."):
            podlazi = "1.PP"
        else:
            parts = code.split(".")
            podlazi = (f"{parts[1]}.NP" if len(parts) >= 2 and parts[1].isdigit()
                       else "?")
        target.append({
            "code": code, "podlazi": podlazi,
            "nazev": str(r[1])[:30] if r[1] else "",
            "plocha_m2": plocha, "ff_typ": ff.strip(),
        })

    total_surface = sum(t["plocha_m2"] for t in target)
    total_volume = total_surface * THICKNESS_M
    total_kc = total_volume * UNIT_PRICE_KC_PER_M3
    print(f"\nD-rooms with FF10/20/21/30/31: {len(target)}")
    print(f"Total surface: {total_surface:.2f} m²")
    print(f"Total volume: {total_volume:.4f} m³ (= surface × 0.04)")
    print(f"Total recovery: {total_kc:,.0f} Kč")

    # Sanity check (aligned with task spec example math)
    if not (10 <= total_volume <= 80):
        print(f"⛔ STOP — volume {total_volume:.4f} m³ mimo expected 10-80 m³")
        return

    items_blob = json.loads(ITEMS.read_text(encoding="utf-8"))
    items = items_blob["items"]
    existing_rooms: set[str] = set()
    for it in items:
        if (it.get("mnozstvi") or 0) <= 0: continue
        if it.get("status") == "deprecated": continue
        p = it["popis"].lower()
        if ("polysterén" in p or "polystyren" in p or "psb 50" in p
                or "liapor" in p):
            for rc in it.get("misto", {}).get("mistnosti", []):
                existing_rooms.add(rc)
    print(f"\nRooms with existing PSB item: {len(existing_rooms)}")

    new_items: list[dict] = []
    for room in target:
        if room["code"] in existing_rooms:
            continue
        volume = round(room["plocha_m2"] * THICKNESS_M, 4)
        new_items.append({
            "item_id": str(uuid.uuid4()),
            "kapitola": "HSV-631",
            "popis": (f"Polystyrenbeton PSB 50 instalační vrstva 40mm "
                       f"(Liapor Mix; v {room['ff_typ']})"),
            "popis_canonical": "Polystyrenbeton PSB 50 40mm",
            "MJ": MJ,
            "mnozstvi": volume,
            "misto": {"objekt": "D", "podlazi": room["podlazi"],
                       "mistnosti": [room["code"]]},
            "skladba_ref": {"FF": room["ff_typ"],
                             "vrstva": "instalační vrstva PSB 50 40mm"},
            "vyrobce_ref": "Liapor Mix PSB 50",
            "urs_code": URS_KOD,
            "urs_status": "needs_review",
            "urs_confidence": 0.95,
            "urs_alternatives": [],
            "unit_price_kc": UNIT_PRICE_KC_PER_M3,
            "total_price_kc": round(volume * UNIT_PRICE_KC_PER_M3, 2),
            "confidence": 0.95,
            "status": "to_audit",
            "source_method": "PHASE_0_18_PSB_beton_40mm",
            "audit_note": (
                f"PHASE_0_18: PSB 50 instalační vrstva (Liapor Mix) — "
                f"konverze m² → m³ explicit:\n"
                f"  surface = {room['plocha_m2']} m² (Tabulka 0020)\n"
                f"  thickness = {THICKNESS_MM} mm = {THICKNESS_M} m "
                f"(Tabulka 0030 verified)\n"
                f"  volume = surface × thickness = {volume} m³\n"
                f"  cena = volume × {UNIT_PRICE_KC_PER_M3} Kč/m³ = "
                f"{volume * UNIT_PRICE_KC_PER_M3:.2f} Kč\n"
                f"ÚRS reference starý VV row 939 kód 128 (m³, "
                f"{UNIT_PRICE_KC_PER_M3} Kč/m³, komplex 146.899 m³)."
            ),
            "warnings": [],
        })

    # Spot-check
    sorted_items = sorted(new_items, key=lambda i: i["mnozstvi"])
    samples = ([sorted_items[0], sorted_items[len(sorted_items) // 2],
                sorted_items[-1]] if len(sorted_items) >= 3 else sorted_items)
    print(f"\nSpot-check 3 rooms (UNIT CONVERSION verification):")
    all_ok = True
    for it in samples:
        rc = it["misto"]["mistnosti"][0]
        match = next((r for r in target if r["code"] == rc), None)
        scope = is_objekt_D(rc)
        ff_ok = match and match["ff_typ"] == str(it["skladba_ref"]["FF"])
        # Volume formula check
        expected_vol = round(match["plocha_m2"] * THICKNESS_M, 4)
        vol_ok = match and abs(it["mnozstvi"] - expected_vol) < 0.001
        unit_ok = it["MJ"] == "m3"
        ok = scope and ff_ok and vol_ok and unit_ok
        if not ok: all_ok = False
        print(f"  {rc:12s} {match['plocha_m2']:>6.2f} m² × 0.04 = "
              f"{it['mnozstvi']:.4f} m³ × {UNIT_PRICE_KC_PER_M3} = "
              f"{it['total_price_kc']:>9,.2f} Kč | scope={scope} ff={ff_ok} "
              f"vol={vol_ok} unit={unit_ok}")
    if not all_ok:
        print("⛔ SPOT-CHECK FAIL"); return

    items.extend(new_items)
    items_blob["items"] = items
    items_blob["metadata"]["items_count"] = len(items)
    items_blob["metadata"]["phase_0_18_applied"] = True
    ITEMS.write_text(json.dumps(items_blob, ensure_ascii=False, indent=2),
                     encoding="utf-8")

    total_added_kc = sum(i["total_price_kc"] for i in new_items)
    print(f"\n✅ PHASE 0.18 RESULTS:")
    print(f"   Items added: {len(new_items)}")
    print(f"   Total volume: {sum(i['mnozstvi'] for i in new_items):.4f} m³")
    print(f"   Total recovery: {total_added_kc:>12,.0f} Kč")


if __name__ == "__main__":
    main()
