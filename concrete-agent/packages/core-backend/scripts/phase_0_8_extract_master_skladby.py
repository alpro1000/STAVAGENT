"""Phase 0.8 — Extract complete WF skladby from master XLSX (Bug #1 fix).

Phase 0.x extraction populated only 2/26 WF skladeb (parsing the wrong source).
This module reads master ``Tabulka skladeb a povrchu`` XLSX and patches
``objekt_D_geometric_dataset.json`` with all detected WF entries.

Each WF gets:
  - ``code`` (WF01, WF03, …)
  - ``celkova_tloustka_mm``
  - ``label`` (category caption from row above)
  - ``vrstvy`` (list of layer dicts: name, thickness, specifikace, vyrobek)
  - ``kind`` (``obvodova`` | ``vnitrni_nosna`` | ``pricka`` | ``sdk``)
  - ``specifikum`` (None | ``podezdivka_van`` | ``instalacni_sachta``
    | ``podezdivka_schodiste`` | ``ocel_HEB``)
  - ``in_tabulka_skladeb`` = True
"""
from __future__ import annotations

import json
import re
import warnings
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

warnings.filterwarnings("ignore")

XLSX = Path(
    "test-data/libuse/inputs/"
    "185-01_DPS_D_SO01_100_0030_R01_TABULKA SKLADEB A POVRCHU_R01.xlsx"
)
DS_PATH = Path("test-data/libuse/outputs/objekt_D_geometric_dataset.json")

# Classification by code prefix (verified against 26 WF entries in master XLSX)
KIND_BY_CODE = {
    # obvodové stěny (suterén + nadzemní + atiky + štíty + nadezdívky)
    **{c: "obvodova" for c in
       ("WF01", "WF03", "WF10", "WF11", "WF12", "WF13", "WF14",
        "WF15", "WF16", "WF17", "WF18", "WF19", "WF90")},
    # vnitřní nosné
    **{c: "vnitrni_nosna" for c in ("WF20", "WF21", "WF22", "WF23", "WF24", "WF25")},
    # příčky zděné
    **{c: "pricka" for c in ("WF30", "WF31", "WF32")},
    # SDK předstěny
    **{c: "sdk" for c in ("WF40", "WF41", "WF50", "WF51")},
}

# Per-code specifikum from documentation (Tabulka skladeb labels manually verified).
SPECIFIKUM = {
    "WF22": "ocel_HEB",            # ocel HEB220 + tepelná izolace pod střechou
    "WF23": "instalacni_sachta",   # železobeton instalační šachta
    "WF24": "instalacni_sachta",   # variant
    "WF25": "podezdivka_schodiste",  # Porotherm 17.5 podezdívka schodiště v 1.PP
    "WF31": "instalacni_sachta",   # Porotherm 11.5 instalační šachta
    "WF32": "podezdivka_van",      # Ytong Klasik 50 — podezdívka van
}


def parse_skladby_sten(ws) -> list[dict]:
    """Parse ``skladby sten`` sheet → list of WF entries with vrstvy.

    Block layout: code header row (col A=``WF``, B=``NN``, E=total mm)
    is preceded by 1+ category caption rows; layer rows follow until the
    next code header.
    """
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    code_rows: list[tuple[int, str, int]] = []
    for i, row in enumerate(rows):
        a = str(row[0]).strip() if row[0] else ""
        b = str(row[1]).strip() if row[1] else ""
        if a == "WF" and re.fullmatch(r"\d{2}", b):
            celkova = row[4] if len(row) > 4 else None
            try:
                tloustka = int(round(float(celkova))) if celkova not in (None, "") else None
            except (TypeError, ValueError):
                tloustka = None
            code_rows.append((i, f"WF{b}", tloustka))

    skladby: list[dict] = []
    for idx, (row_idx, code, tloustka) in enumerate(code_rows):
        # Label = nearest non-empty caption row above (skip blank rows + header rows)
        label = ""
        for back in range(row_idx - 1, max(row_idx - 6, -1), -1):
            cell = rows[back][0] if rows[back] else None
            if cell and isinstance(cell, str) and cell.strip() and cell.strip() not in ("WF", "Kód / Code"):
                label = cell.strip()
                break

        # Vrstvy = rows from row_idx to next code header (or +20)
        end = code_rows[idx + 1][0] if idx + 1 < len(code_rows) else min(row_idx + 20, len(rows))
        vrstvy: list[dict] = []
        for j in range(row_idx, end):
            r = rows[j]
            if not r or len(r) < 7:
                continue
            poradi = r[1] if r[1] else None
            nazev = r[2]
            tl = r[3]
            specifikace = r[5]
            vyrobek = r[6]
            if not nazev:
                continue
            # Skip header repeat
            if isinstance(nazev, str) and nazev.strip() == "Pořadí vrstev / Order of layers":
                continue
            try:
                tl_val = float(tl) if tl not in (None, "", "-") else None
            except (TypeError, ValueError):
                tl_val = None
            vrstvy.append({
                "poradi": str(poradi).strip() if poradi else None,
                "nazev": str(nazev).strip(),
                "tloustka_mm": tl_val,
                "specifikace": str(specifikace).strip() if specifikace else None,
                "referencni_vyrobek": str(vyrobek).strip() if vyrobek else None,
            })

        kind = KIND_BY_CODE.get(code)
        skladby.append({
            "code": code,
            "in_tabulka_skladeb": True,
            "referenced_in": [],
            "kind": kind,
            "label": label,
            "celkova_tloustka_mm": tloustka,
            "vrstvy": vrstvy,
            "specifikum": SPECIFIKUM.get(code),
        })

    return skladby


def main() -> None:
    print(f"Loading XLSX: {XLSX.name}")
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["skladby sten"]
    wf = parse_skladby_sten(ws)
    print(f"Extracted {len(wf)} WF skladeb from master XLSX")
    by_kind: dict[str, list[str]] = {}
    for s in wf:
        by_kind.setdefault(s["kind"] or "?", []).append(s["code"])
    for k, codes in by_kind.items():
        print(f"  {k:15s} ({len(codes)}): {sorted(codes)}")

    # Patch dataset
    print(f"\nPatching {DS_PATH}…")
    ds = json.loads(DS_PATH.read_text(encoding="utf-8"))
    sk_dict = ds.get("skladby") or {}
    patched = 0
    for s in wf:
        existing = sk_dict.get(s["code"])
        if existing and existing.get("in_tabulka_skladeb"):
            # Already populated — keep referenced_in but refresh kind/specifikum
            existing.setdefault("vrstvy", s["vrstvy"])
            existing.setdefault("celkova_tloustka_mm", s["celkova_tloustka_mm"])
            existing["kind"] = s["kind"]
            existing["specifikum"] = s["specifikum"]
            existing["label"] = s.get("label") or existing.get("label", "")
            existing.pop("warning", None)
        else:
            # Carry referenced_in from existing stub if any
            if existing:
                s["referenced_in"] = existing.get("referenced_in", [])
            sk_dict[s["code"]] = s
            patched += 1
    ds["skladby"] = sk_dict
    DS_PATH.write_text(json.dumps(ds, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  patched {patched} new WF entries; {len(wf) - patched} refreshed")
    print(f"  total WF in dataset: {sum(1 for c in sk_dict if c.startswith('WF'))}")


if __name__ == "__main__":
    main()
