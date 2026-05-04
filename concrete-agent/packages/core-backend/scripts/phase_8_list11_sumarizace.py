"""Phase 8 — Add List 11 Sumarizace_dle_kódu to existing Excel.

Single in-place edit:
- Adds sheet "11_Sumarizace_dle_kódu" (group-based summary, KROS-friendly)
- Updates List 0 "0_Souhrn" with note about List 11 + manual KROS workflow
- Updates List 9 "9_Metadata" with Phase 8 info
- Saves IN-PLACE to existing Excel file (all other 10 sheets preserved)
"""
from __future__ import annotations

import json
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT_DIR = Path("test-data/libuse/outputs")
GROUPS = OUT_DIR / "urs_query_groups.json"
ITEMS = OUT_DIR / "items_objekt_D_complete.json"
EXCEL = OUT_DIR / "Vykaz_vymer_Libuse_objekt_D_dokoncovaci_prace.xlsx"

# Color palette (matches existing Excel scheme)
FILL_KRIT = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
FILL_DETAIL = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
FILL_NOVE = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
FILL_SHODA = PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")
FILL_OPRAVENO_OBJ = PatternFill(start_color="E0E0FF", end_color="E0E0FF", fill_type="solid")
FILL_OPRAVENO_POP = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
FILL_HEADER = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
FILL_MASTER = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
FILL_DETAIL_ROW = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
FILL_URS_PLACEHOLDER = PatternFill(start_color="FFF2A8", end_color="FFF2A8", fill_type="solid")
FILL_MIX = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_MASTER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_DETAIL = Font(name="Calibri", size=10, color="555555")
FONT_BOLD = Font(name="Calibri", size=10, bold=True)
FONT_NORMAL = Font(name="Calibri", size=10)
FONT_KRIT = Font(name="Calibri", size=10, bold=True, color="9C0000")
FONT_DETAIL_KRIT = Font(name="Calibri", size=10, bold=True, color="9C0000")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
BORDER_THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

STATUS_FILLS = {
    "VYNECHANE_KRITICKE": FILL_KRIT,
    "VYNECHANE_DETAIL": FILL_DETAIL,
    "NOVE": FILL_NOVE,
    "SHODA_SE_STARYM": FILL_SHODA,
    "OPRAVENO_OBJEM": FILL_OPRAVENO_OBJ,
    "OPRAVENO_POPIS": FILL_OPRAVENO_POP,
}


def master_status_color(status_mix: dict) -> PatternFill:
    """Determine master row color based on status mix purity."""
    if not status_mix:
        return FILL_NOVE
    statuses = set(status_mix.keys())
    if statuses == {"VYNECHANE_KRITICKE"}:
        return FILL_KRIT
    if statuses == {"NOVE"}:
        return FILL_NOVE
    if statuses == {"SHODA_SE_STARYM"}:
        return FILL_SHODA
    if statuses == {"VYNECHANE_DETAIL"}:
        return FILL_DETAIL
    return FILL_MIX  # Mixed statuses


def write_misto(misto: dict) -> str:
    parts = [misto.get("objekt", ""), misto.get("podlazi", "")]
    mistnosti = misto.get("mistnosti") or []
    if mistnosti:
        parts.append(",".join(mistnosti))
    return " · ".join(p for p in parts if p)


def write_skladba(skl: dict) -> str:
    if not skl:
        return ""
    pairs = []
    for k, v in skl.items():
        if isinstance(v, (str, int, float)):
            pairs.append(f"{k}={v}")
    return "; ".join(pairs[:3])


def main() -> None:
    print("Loading…")
    groups_blob = json.loads(GROUPS.read_text(encoding="utf-8"))
    groups = groups_blob["groups"]
    items_data = json.loads(ITEMS.read_text(encoding="utf-8"))
    items_by_id: dict[str, dict] = {it["item_id"]: it for it in items_data["items"]}

    # Sort groups by group_id ASC (G001, G002, ...)
    groups_sorted = sorted(groups, key=lambda g: g["group_id"])

    print(f"Opening existing Excel ({EXCEL})…")
    wb = openpyxl.load_workbook(str(EXCEL))
    print(f"  Sheets present: {wb.sheetnames}")

    # Drop existing List 11 if rerun
    if "11_Sumarizace_dle_kódu" in wb.sheetnames:
        del wb["11_Sumarizace_dle_kódu"]
        print("  Removed pre-existing List 11 (rerun)")

    print("Building List 11…")
    ws = wb.create_sheet("11_Sumarizace_dle_kódu")

    # Header
    headers = ["#", "ÚRS kód", "Popis", "MJ", "Total množství",
                "Components", "Skladby", "Kapitola", "Status mix",
                "Group ID", "Note"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN
    ws.freeze_panes = "B2"

    row_idx = 2
    detail_total = 0
    for n, g in enumerate(groups_sorted, 1):
        # Master row
        status_mix = g.get("status_mix_in_group") or {}
        master_fill = master_status_color(status_mix)
        # Status mix string (e.g. "47× NOVE, 12× VYNECHANE_KRITICKE")
        status_str = ", ".join(
            f"{count}× {st}"
            for st, count in sorted(status_mix.items(), key=lambda x: -x[1])
        )

        ws.cell(row=row_idx, column=1, value=n)
        # ÚRS kód cell — empty placeholder, yellow highlight (manual KROS entry)
        urs_cell = ws.cell(row=row_idx, column=2, value="")
        urs_cell.fill = FILL_URS_PLACEHOLDER
        ws.cell(row=row_idx, column=3, value=g["popis_canonical"])
        ws.cell(row=row_idx, column=4, value=g["MJ"])
        ws.cell(row=row_idx, column=5, value=g["total_mnozstvi"])
        ws.cell(row=row_idx, column=6, value=g["items_count"])
        ws.cell(row=row_idx, column=7, value=", ".join(g.get("skladba_refs") or []))
        ws.cell(row=row_idx, column=8, value=g["kapitola"])
        ws.cell(row=row_idx, column=9, value=status_str)
        ws.cell(row=row_idx, column=10, value=g["group_id"])
        ws.cell(row=row_idx, column=11, value="")  # Note — KROS comments

        # Apply master row formatting (skip column 2 — keep yellow highlight)
        for col in range(1, 12):
            cell = ws.cell(row=row_idx, column=col)
            if col != 2:
                cell.fill = master_fill if master_fill is not FILL_HEADER else FILL_MASTER
            # If pure status fill is a "white-on-blue" master, override font
            if col == 1 or (col >= 3 and master_fill in (FILL_KRIT, FILL_NOVE,
                                                          FILL_SHODA, FILL_DETAIL,
                                                          FILL_MIX)):
                cell.font = FONT_BOLD
            cell.border = BORDER_THIN

        master_row_idx = row_idx
        row_idx += 1

        # Detail rows (level 1, collapsed)
        for item_id in g.get("items_ids") or []:
            it = items_by_id.get(item_id)
            if not it:
                continue
            misto_str = write_misto(it.get("misto") or {})
            skl_str = write_skladba(it.get("skladba_ref") or {})
            ds = it.get("data_source") or ""
            urs_status = it.get("urs_status") or ""

            ws.cell(row=row_idx, column=1, value="")
            ws.cell(row=row_idx, column=2, value="")
            ws.cell(row=row_idx, column=3, value=f"  • {misto_str}" if misto_str else "  • —")
            ws.cell(row=row_idx, column=4, value=it.get("MJ"))
            ws.cell(row=row_idx, column=5, value=it.get("mnozstvi"))
            ws.cell(row=row_idx, column=6, value="")
            ws.cell(row=row_idx, column=7, value=skl_str)
            ws.cell(row=row_idx, column=8, value=it.get("kapitola"))
            ws.cell(row=row_idx, column=9, value=urs_status)
            ws.cell(row=row_idx, column=10, value="")
            ws.cell(row=row_idx, column=11, value=f"Source: {ds}" if ds else "")

            # Apply detail row formatting
            for col in range(1, 12):
                cell = ws.cell(row=row_idx, column=col)
                cell.fill = FILL_DETAIL_ROW
                cell.font = FONT_DETAIL
                cell.border = BORDER_THIN
            # Color-code status column inline
            status_fill = STATUS_FILLS.get(urs_status)
            if status_fill:
                ws.cell(row=row_idx, column=9).fill = status_fill
                if urs_status == "VYNECHANE_KRITICKE":
                    ws.cell(row=row_idx, column=9).font = FONT_DETAIL_KRIT

            # Outline level 1 (collapsed by default)
            ws.row_dimensions[row_idx].outline_level = 1
            ws.row_dimensions[row_idx].hidden = True

            row_idx += 1
            detail_total += 1

    # Sheet outline properties — master ABOVE details
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False

    # AutoFilter on entire range
    last_col_letter = get_column_letter(11)
    ws.auto_filter.ref = f"A1:{last_col_letter}{row_idx - 1}"

    # Column widths per spec
    widths = {"A": 5, "B": 15, "C": 50, "D": 6, "E": 13,
                "F": 11, "G": 22, "H": 13, "I": 28, "J": 8, "K": 32}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    print(f"  Master rows: {len(groups_sorted)}")
    print(f"  Detail rows (collapsed): {detail_total}")

    # Update List 0 "0_Souhrn" — append note about List 11
    print("Updating List 0 (Souhrn)…")
    s0 = wb["0_Souhrn"]
    # Find first empty row at bottom of A column
    next_r = s0.max_row + 2
    s0.cell(row=next_r, column=1, value="List 11 — Sumarizace_dle_kódu (Phase 8)").font = Font(
        size=12, bold=True, color="003366"
    )
    s0.merge_cells(start_row=next_r, start_column=1, end_row=next_r, end_column=4)
    next_r += 1
    note_text = (
        "List 11 obsahuje sumarizaci práce podle group_id pro KROS workflow. "
        "ÚRS kódy v sloupci B jsou prázdné (žluté zvýraznění) — ručně vyplnit při KROS pricing."
    )
    s0.cell(row=next_r, column=1, value=note_text).alignment = Alignment(wrap_text=True, vertical="top")
    s0.merge_cells(start_row=next_r, start_column=1, end_row=next_r, end_column=4)
    s0.row_dimensions[next_r].height = 30

    # Update List 9 "9_Metadata" — append Phase 8 info
    print("Updating List 9 (Metadata)…")
    s9 = wb["9_Metadata"]
    next_r = s9.max_row + 2
    s9.cell(row=next_r, column=1, value="Phase 8 List 11").font = Font(size=11, bold=True)
    s9.cell(row=next_r, column=2,
             value=f"Group-based summary added — {len(groups_sorted)} master rows × "
                   f"{detail_total} collapsed detail rows. ÚRS kódy manual entry "
                   "via KROS workflow (no automation).")
    s9.cell(row=next_r, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    # Save in-place
    wb.save(str(EXCEL))
    print(f"\nSaved {EXCEL} ({EXCEL.stat().st_size:,} bytes)")
    print(f"Sheets after: {wb.sheetnames}")


if __name__ == "__main__":
    main()
