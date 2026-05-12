"""Phase 0.21 — List 13 Filter_view_plus (Subdodavatel + Discipline enrichment).

In-place edit of the existing D Excel that ADDS a new List 13 sheet
alongside the unchanged List 12. Zero-regression guarantee on List 12:
the script reads it but does not write to it.

List 13 layout = List 12's 13 columns + 2 enrichment columns:
  14. Discipline    — HSV / PSV / M (derived from kapitola prefix)
  15. Subdodavatel  — looked up from test-data/libuse/data/subdodavatel_mapping.json

Also enriches items_objekt_D_complete.json with a `subdodavatel` field
on each item (additive — no existing field changed). This makes the
mapping queryable by downstream consumers (List 11 sumarizace, future
ÚRS-pricing pipeline, etc.) without re-running this script.

Excel Table on List 13 is named `VykazFilterPlus` (parallels List 12's
`VykazFilter`); drop-down filter UI applies to all 15 columns.

Run from repo root:
    python concrete-agent/packages/core-backend/scripts/phase_0_21_list13_filter_view_plus.py
"""
from __future__ import annotations

import json
import re
import warnings
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

ROOT = Path(__file__).resolve().parents[4]
OUT = ROOT / "test-data" / "libuse" / "outputs"
EXCEL = OUT / "Vykaz_vymer_Libuse_objekt_D_dokoncovaci_prace.xlsx"
ITEMS = OUT / "items_objekt_D_complete.json"
MAPPING = ROOT / "test-data" / "libuse" / "data" / "subdodavatel_mapping.json"

SHEET_SRC = "1_Vykaz_vymer"
SHEET_12 = "12_Filter_view"
SHEET_13 = "13_Filter_view_plus"


# ---------------------------------------------------------------------------
# Subdodavatel resolution
# ---------------------------------------------------------------------------
def load_mapping() -> dict:
    raw = json.loads(MAPPING.read_text(encoding="utf-8"))
    # Compile popis-regex rules once
    raw["_popis_compiled"] = [
        (re.compile(r["popis_regex"]), r["subdodavatel"])
        for r in raw.get("_popis_prefix_special_cases", {}).get("rules", [])
    ]
    # Pre-lowercase intra-kapitola granular keywords for fast substring scan.
    # Schema: { kapitola: [{keywords: [...], subdodavatel: "..."}, ...] }
    granular_raw = raw.get("_kapitola_popis_granular", {}) or {}
    raw["_granular_compiled"] = {
        kap: [
            {
                "keywords_lc": [kw.lower() for kw in rule.get("keywords", [])],
                "subdodavatel": rule["subdodavatel"],
            }
            for rule in rules
            if isinstance(rule, dict) and "keywords" in rule
        ]
        for kap, rules in granular_raw.items()
        if not kap.startswith("_")
    }
    return raw


def discipline_of(kapitola: str | None) -> str:
    if not kapitola:
        return "?"
    if kapitola.startswith("HSV"):
        return "HSV"
    if kapitola.startswith("PSV"):
        return "PSV"
    if kapitola.startswith("M-"):
        return "M"
    if kapitola.startswith("VRN"):
        return "VRN"
    if kapitola.startswith(("OP-", "LI-", "Detail-")):
        return "Detail"
    return "?"


def subdodavatel_for(item: dict, mapping: dict) -> str:
    """Resolve subdodavatel per the priority chain documented in the
    mapping JSON: popis-regex override → intra-kapitola granular keyword
    rules → exact kapitola → kapitola prefix → discipline default →
    'vlastní'.
    """
    popis = item.get("popis", "") or ""
    for rx, sub in mapping.get("_popis_compiled", []):
        if rx.match(popis):
            return sub
    kapitola = item.get("kapitola", "") or ""
    # Intra-kapitola granular rules (v1.1): keyword scan over popis +
    # skladba_ref.vrstva. First matching rule wins. Used for kapitoly
    # whose scope covers multiple trades (e.g. PSV-783 Ochrana konstrukcí
    # = epoxid/PU/Sikagard/anti-graffiti/pancéřový/žárové zinkování).
    granular_rules = mapping.get("_granular_compiled", {}).get(kapitola)
    if granular_rules:
        skladba_vrstva = ((item.get("skladba_ref") or {}).get("vrstva", "") or "")
        haystack = (popis + " " + skladba_vrstva).lower()
        for rule in granular_rules:
            if any(kw in haystack for kw in rule["keywords_lc"]):
                return rule["subdodavatel"]
    exact = mapping.get("exact_kapitola_match", {})
    if kapitola in exact:
        return exact[kapitola]
    # Prefix fallback (longer prefix wins)
    prefixes = mapping.get("kapitola_prefix_fallback", {})
    best_prefix = ""
    best_value = None
    for prefix, value in prefixes.items():
        if kapitola.startswith(prefix) and len(prefix) > len(best_prefix):
            best_prefix = prefix
            best_value = value
    if best_value:
        return best_value
    disc = discipline_of(kapitola)
    return mapping.get("discipline_default", {}).get(disc, "vlastní")


# ---------------------------------------------------------------------------
# Items enrichment (in-place additive field)
# ---------------------------------------------------------------------------
def enrich_items() -> tuple[int, Counter]:
    """Add subdodavatel field to every item in items_objekt_D_complete.json.

    Idempotent — re-runs produce byte-identical output for the same
    mapping + items state.
    """
    mapping = load_mapping()
    data = json.loads(ITEMS.read_text(encoding="utf-8"))
    counts: Counter = Counter()
    for item in data["items"]:
        sub = subdodavatel_for(item, mapping)
        item["subdodavatel"] = sub
        counts[sub] += 1
    # Metadata trail
    meta = data.setdefault("metadata", {})
    meta["subdodavatel_enriched"] = True
    meta["subdodavatel_mapping_version"] = mapping.get("_doc", {}).get("version")
    meta["subdodavatel_breakdown"] = dict(counts)
    ITEMS.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return len(data["items"]), counts


# ---------------------------------------------------------------------------
# List 13 generation (in-place add, List 12 untouched)
# ---------------------------------------------------------------------------
def build_list_13() -> dict:
    """Open existing Excel, add List 13 alongside List 12, save.

    Returns a stats dict for downstream reporting.
    """
    mapping = load_mapping()
    items_data = json.loads(ITEMS.read_text(encoding="utf-8"))
    # Build lookup: row index 1..N in source aligns to item position
    items_list = items_data["items"]

    wb = load_workbook(EXCEL)
    pre_sheets = list(wb.sheetnames)
    assert SHEET_SRC in pre_sheets, f"{SHEET_SRC!r} missing"
    assert SHEET_12 in pre_sheets, (
        f"{SHEET_12!r} missing — Phase 0.20 must run before this script"
    )

    # Read source rows 1..N from List 1 (List 12 = same content + Podlazi
    # derived col). We re-derive everything from List 1 here, NOT from
    # List 12 — that way our regen is deterministic regardless of List 12
    # state, AND we don't accidentally read stale List 12 data if anyone
    # re-ran Phase 0.20 with subtle differences.
    ws_src = wb[SHEET_SRC]
    n_src_rows = ws_src.max_row
    n_src_cols = ws_src.max_column
    headers_src = [ws_src.cell(1, c).value for c in range(1, n_src_cols + 1)]
    REQUIRED = ["#", "ÚRS kód", "Kapitola", "Popis položky", "MJ",
                "Množství", "Místo", "Skladba/povrch", "Confidence",
                "Status", "Poznámka", "Source"]
    missing = [c for c in REQUIRED if c not in headers_src]
    if missing:
        raise RuntimeError(f"{SHEET_SRC} missing required cols: {missing}")

    # Resolve source col indices by header name
    col_idx = {h: i + 1 for i, h in enumerate(headers_src) if h is not None}

    # Drop pre-existing List 13 if rerun
    if SHEET_13 in wb.sheetnames:
        del wb[SHEET_13]
    ws = wb.create_sheet(SHEET_13)

    # ROW 1: summary metadata banner (outside table — same convention as List 12)
    n_data_rows = n_src_rows - 1
    ws.cell(1, 1, "Total items:").font = Font(bold=True, size=11)
    ws.cell(1, 2, n_data_rows).font = Font(size=11)
    ws.cell(1, 4, "List 13 = List 12 + Discipline + Subdodavatel. Klikněte ▼ na header pro filtrování.").font = Font(italic=True, color="555555", size=10)
    ws.cell(1, 13, "Source: '1_Vykaz_vymer' + mapping data/subdodavatel_mapping.json").font = Font(italic=True, color="888888", size=9)
    ws.row_dimensions[1].height = 22

    # ROW 2: headers (15 cols)
    HEADER_ROW = 2
    headers_all = headers_src + ["Podlaží", "Discipline", "Subdodavatel"]
    for c, h in enumerate(headers_all, 1):
        ws.cell(HEADER_ROW, c, h)

    DATA_START = HEADER_ROW + 1
    podlazi_col = n_src_cols + 1
    discipline_col = n_src_cols + 2
    subdodavatel_col = n_src_cols + 3

    # Build a fast lookup item_id → subdodavatel from enriched items
    # Phase 0.20 places source rows row 2..N+1 in List 1, items[i] →
    # row i+2. We rely on the same ordering convention here.
    if len(items_list) != n_data_rows:
        raise RuntimeError(
            f"items count {len(items_list)} != List 1 data rows {n_data_rows} "
            "— item order/Excel divergence; re-run Phase 6 first."
        )

    misto_col = col_idx["Místo"]
    for src_row in range(2, n_src_rows + 1):
        target_row = DATA_START + (src_row - 2)
        item = items_list[src_row - 2]
        # Copy 12 source columns
        for c in range(1, n_src_cols + 1):
            ws.cell(target_row, c, ws_src.cell(src_row, c).value)
        # Derived Podlaží (same convention as List 12)
        misto_val = ws_src.cell(src_row, misto_col).value
        ws.cell(target_row, podlazi_col, _extract_podlazi(misto_val))
        # Discipline
        ws.cell(target_row, discipline_col, discipline_of(item.get("kapitola")))
        # Subdodavatel — pre-enriched at items level
        ws.cell(target_row, subdodavatel_col,
                item.get("subdodavatel") or subdodavatel_for(item, mapping))

    last_data_row = DATA_START + n_data_rows - 1
    n_total_cols = subdodavatel_col
    last_col_letter = get_column_letter(n_total_cols)
    table_ref = f"A{HEADER_ROW}:{last_col_letter}{last_data_row}"

    # Convert to Excel Table
    table = Table(displayName="VykazFilterPlus", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium7",  # different from List 12's Medium9 for visual cue
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)

    # Column widths (mirror List 12 + add 2 trailing cols)
    widths_src = [6, 12, 10, 50, 6, 11, 22, 32, 12, 18, 25, 18]
    widths_extra = [12, 14, 26]  # Podlaží, Discipline, Subdodavatel
    for i, w in enumerate(widths_src + widths_extra, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"A{DATA_START}"

    # Conditional formatting — Status + Confidence (mirror List 12)
    status_col_letter = get_column_letter(col_idx["Status"])
    confidence_col_letter = get_column_letter(col_idx["Confidence"])

    status_range = (f"{status_col_letter}{DATA_START}:"
                    f"{status_col_letter}{last_data_row}")
    sl = status_col_letter
    ds = DATA_START
    ws.conditional_formatting.add(status_range, FormulaRule(
        formula=[f'OR({sl}{ds}="matched_high",{sl}{ds}="matched_medium")'],
        fill=PatternFill("solid", fgColor="C6EFCE"), font=Font(color="006100")))
    ws.conditional_formatting.add(status_range, FormulaRule(
        formula=[f'OR({sl}{ds}="needs_review",{sl}{ds}="OPRAVENO_OBJEM",'
                 f'{sl}{ds}="OPRAVENO_POPIS")'],
        fill=PatternFill("solid", fgColor="FFEB9C"), font=Font(color="9C5700")))
    ws.conditional_formatting.add(status_range, FormulaRule(
        formula=[f'OR({sl}{ds}="no_match",{sl}{ds}="VYNECHANE_KRITICKE")'],
        fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(color="9C0006")))
    ws.conditional_formatting.add(status_range, FormulaRule(
        formula=[f'OR({sl}{ds}="VYNECHANE_DETAIL",{sl}{ds}="deprecated")'],
        fill=PatternFill("solid", fgColor="D9D9D9"), font=Font(color="595959")))

    confidence_range = (f"{confidence_col_letter}{DATA_START}:"
                        f"{confidence_col_letter}{last_data_row}")
    ws.conditional_formatting.add(confidence_range, DataBarRule(
        start_type="num", start_value=0, end_type="num", end_value=1,
        color="638EC6", showValue=True))

    # Sanity: List 12 must still be present + identical row count
    ws12 = wb[SHEET_12]
    assert ws12.max_row == DATA_START + n_data_rows - 1, (
        f"List 12 row drift {ws12.max_row} vs expected "
        f"{DATA_START + n_data_rows - 1}"
    )

    wb.save(EXCEL)

    return {
        "n_items": n_data_rows,
        "n_total_cols": n_total_cols,
        "sheets_pre": pre_sheets,
        "sheets_post": list(load_workbook(EXCEL).sheetnames),
        "list_13_rows": last_data_row,
        "table_name": "VykazFilterPlus",
    }


def _extract_podlazi(misto) -> str:
    if not isinstance(misto, str):
        return "unknown"
    parts = [p.strip() for p in misto.split("·")]
    return parts[1] if len(parts) >= 2 else "unknown"


def main() -> int:
    print("Phase 0.21 — List 13 Filter_view_plus")
    print("=" * 60)

    print("\nStep 1: enrich items_objekt_D_complete.json…")
    n_items, breakdown = enrich_items()
    print(f"  ✓ {n_items} items enriched with `subdodavatel` field")
    print(f"  Breakdown (top 10):")
    for sub, n in breakdown.most_common(10):
        print(f"    {sub:<45} {n:>5}")
    if len(breakdown) > 10:
        rest = sum(n for sub, n in breakdown.most_common()[10:])
        print(f"    (+ {len(breakdown) - 10} other categories, {rest} items)")

    print("\nStep 2: add List 13 to Excel (in-place; List 12 untouched)…")
    stats = build_list_13()
    print(f"  ✓ Sheets pre: {len(stats['sheets_pre'])} {stats['sheets_pre']}")
    print(f"  ✓ Sheets post: {len(stats['sheets_post'])} {stats['sheets_post']}")
    print(f"  ✓ List 13 rows × cols: {stats['list_13_rows']} × {stats['n_total_cols']}")
    print(f"  ✓ Excel Table: {stats['table_name']}")
    print(f"\n  → wrote {EXCEL.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
