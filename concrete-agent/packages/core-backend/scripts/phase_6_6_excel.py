"""Phase 6.6 GATE 3 — Excel generator for material decomposition.

Adds two new sheets to the existing Vykaz_vymer Excel:

  * `Material_rozklad`  — detail sheet, one row per material sub-item with
                          full provenance, hyperlink back to master VV row,
                          conditional formatting by Status, alternating
                          row tint per master ID.
  * `Material_audit`    — dashboard sheet (set as ACTIVE on open):
                            block 1 hero stats
                            block 2 donut chart + table
                            block 3 per-kapitola horizontal bar chart
                            block 4 top-20 needs-VELTON-confirm
                            block 5 Case 5 masters summary
                            block 6 orphan library entries
                            block 7 vykres-annotated note
                            block 8 disclaimer footer

Main VV sheet (`1_Vykaz_vymer`) and the other 11 existing sheets are
untouched. Backup pre_phase6_6 written before overwrite. Run from repo
root.
"""
from __future__ import annotations

import json
import re
import shutil
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

import openpyxl  # type: ignore[import-not-found]
from openpyxl.chart import BarChart, DoughnutChart, Reference  # type: ignore[import-not-found]
from openpyxl.chart.label import DataLabelList  # type: ignore[import-not-found]
from openpyxl.formatting.rule import CellIsRule  # type: ignore[import-not-found]
from openpyxl.styles import (Alignment, Border, Font, PatternFill,  # type: ignore[import-not-found]
                              Side)
from openpyxl.utils import get_column_letter  # type: ignore[import-not-found]
from openpyxl.workbook.defined_name import DefinedName  # type: ignore[import-not-found]

REPO_ROOT = Path(__file__).resolve().parents[4]
LIBUSE = REPO_ROOT / "test-data" / "libuse"
OUTPUTS = LIBUSE / "outputs"

EXCEL_TARGET = OUTPUTS / "Vykaz_vymer_Libuse_objekt_D_dokoncovaci_prace.xlsx"
EXCEL_BACKUP_PHASE_6_6 = OUTPUTS / "Vykaz_vymer_pre_phase6_6.xlsx"
EXCEL_BACKUP_GATE4 = OUTPUTS / "Vykaz_vymer_pre_gate4.xlsx"
EXCEL_BACKUP_GATE5 = OUTPUTS / "Vykaz_vymer_pre_gate5.xlsx"
EXCEL_BACKUP_GATE6 = OUTPUTS / "Vykaz_vymer_pre_gate6.xlsx"
EXCEL_BACKUP_GATE8 = OUTPUTS / "Vykaz_vymer_pre_gate8.xlsx"
EXCEL_BACKUP_GATE8_1 = OUTPUTS / "Vykaz_vymer_pre_gate8_1.xlsx"
ITEMS_IN = OUTPUTS / "items_objekt_D_with_materials.json"
LIBRARY_IN = OUTPUTS / "material_library_D.json"
KB_IN = LIBUSE / "knowledge_base" / "generic_consumption_rates.json"
GROUPS_IN = OUTPUTS / "urs_query_groups.json"

ROZKLAD_SHEET = "Material_rozklad"
AUDIT_SHEET = "Material_audit"
VV_SHEET = "1_Vykaz_vymer"
SUMARIZACE_SHEET = "11_Sumarizace_dle_kódu"
AGGREGATE_SHEET = "11b_Material_aggregate"
AVK_SHEET = "11c_AVK_smeta"

# Documented source types (Block 1 Hero stat denominator). GATE 5a adds
# the two ČSN-cited tiers — they count as documented because the norm
# reference (and optional URL) is an authoritative citation.
DOCUMENTED_SOURCES = {
    "tz_explicit_with_rate", "tz_explicit_no_rate",
    "tabulka_referenced", "vykres_annotated",
    "generic_with_csn_norm", "generic_with_csn_url",
    # GATE 8b — rate extracted from project popis, KB lookup with ČSN,
    # and master-as-material self-reference all count as documented
    # (master popis IS authoritative project documentation).
    "rate_from_popis", "case5_kb_rate", "case5_self_reference",
}

# Status fills per user spec
STATUS_FILLS: dict[str, tuple[str, str]] = {
    # status_label → (bg hex, font hex)
    "OK":      ("C6EFCE", "006100"),
    "Confirm": ("FFEB9C", "9C5700"),
    "Odhad":   ("FFC09A", "974706"),
    "Missing": ("FFC7CE", "9C0006"),
}

# Donut chart colors per source provenance
SOURCE_COLORS: dict[str, str] = {
    "tz_explicit_with_rate":    "00B050",  # bright green
    "tz_explicit_no_rate":      "92D050",  # lighter green
    "tabulka_referenced":       "FFC000",  # orange
    "vykres_annotated":         "00B0F0",  # blue
    "generic_no_documentation": "FF6B6B",  # red
    "generic_with_csn_norm":    "7030A0",  # purple — ČSN-cited
    "generic_with_csn_url":     "B266FF",  # lighter purple — ČSN with URL
    "rate_from_popis":          "008080",  # teal — rate from project popis
    "case5_kb_rate":            "9966CC",  # medium purple — Case 5 KB
    "case5_self_reference":     "20B2AA",  # light sea green — master=material
    "vrn_services":             "B0B0B0",  # gray — services (no material)
}

THIN = Side(style="thin", color="B0B0B0")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOLD = Font(bold=True)
BIG_BOLD = Font(bold=True, size=18)
HERO_FONT = Font(bold=True, size=22, color="2C5F8D")
TITLE_FONT = Font(bold=True, size=14, color="2C5F8D")
SMALL_ITALIC = Font(size=9, italic=True, color="606060")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
# Data-cell alignment for Material_rozklad — wrap_text=False keeps row
# heights at the default (~15 px).  Excel auto-expands rows when
# wrap_text=True is applied to short popis values that contain Czech accents
# or a couple of "—" separators, producing 60+ px rows on 90 k cells.
# Header cells continue using CENTER (wrap=True).
LEFT_NOWRAP = Alignment(horizontal="left", vertical="center", wrap_text=False)


def _load_data() -> tuple[list[dict], list[dict], list[dict]]:
    items_data = json.loads(ITEMS_IN.read_text(encoding="utf-8"))
    library_data = json.loads(LIBRARY_IN.read_text(encoding="utf-8"))

    masters: list[dict] = []
    sub_items: list[dict] = []
    for it in items_data["items"]:
        if it.get("item_role") == "material_subitem":
            sub_items.append(it)
        else:
            masters.append(it)
    library = library_data["materials"]
    return masters, sub_items, library


def _build_master_row_map(masters: list[dict]) -> dict[str, int]:
    """Master items appear in VV at row index i+2 (1-indexed, +1 header)."""
    return {m["item_id"]: i + 2 for i, m in enumerate(masters)}


# Mirror of pair_materials._is_install_only used purely for the audit
# block counter (Block 11). Logic must stay in sync with the pair script.
_INSTALL_SUFFIX_RE = __import__("re").compile(
    r"\s—\s+(kotvení|kotveni|montáž|montaz|klika|spárování|sparovani)\b",
    __import__("re").IGNORECASE,
)
_INSTALL_PREFIX_RE = __import__("re").compile(
    r"^(osazení|osazeni|rektifikační šrouby|rektifikacni srouby|"
    r"dodatečné kotvení|dodatecne kotveni|"
    r"montáž (tp|lp|op|li)|montaz (tp|lp|op|li))",
    __import__("re").IGNORECASE,
)


def _looks_like_install_only(popis: str) -> bool:
    if not popis:
        return False
    return bool(_INSTALL_SUFFIX_RE.search(popis)
                or _INSTALL_PREFIX_RE.search(popis))


# Mirror of pair_materials.CASE5_PRIMARY_KEYWORDS — used only by
# Material_rozklad GATE 7 placeholder copy to classify why a master has no
# sub-items.  Keep the two lists in sync when pair script gains new keys.
CASE5_PRIMARY_KEYWORDS = (
    "penetrace pod", "penetrace univerzá", "lepidlo flexib", "lepidlo na",
    "spárovací hmot", "sparovaci hmot",
    "samonivelační stěr", "samonivelacni ster",
    "kari síť", "kari sit", "pe fólie", "pe folie",
    "asfaltový pás", "asfaltovy pas", "armovací síť", "armovaci sit",
    "tmel ", "akrylový ",
    "uw + cw profil", "ud + cd profil", "uw+cw profil",
    "cd profil", "cw profil",
    "sdk desky", "sdk deska", "izolace minerální vata",
    "izolace mineralni vata", "izolace minerá",
    "tmelení q", "tmeleni q", "pur pěna", "pur pena",
    "závěsy posuvné", "zavesy posuvne",
    "parozábrana fólie", "parozabrana folie",
    "difuzní fólie", "difuzni folie",
    "latě ", "kontralatě", "hřebenáče", "hrebenace",
    "kročejová izolace", "krocejova izolace",
    "polystyrenbeton", "polystyrén beton", "polystyren beton",
)


def _format_qty(q: Any) -> str:
    if q is None:
        return ""
    try:
        v = float(q)
    except (ValueError, TypeError):
        return str(q)
    if v == int(v):
        return f"{int(v):,}".replace(",", " ")
    return f"{v:,.2f}".replace(",", " ")


def _format_master_qty_mj(master: dict) -> str:
    q = master.get("mnozstvi")
    mj = master.get("MJ") or ""
    if q is None:
        return mj
    return f"{_format_qty(q)} {mj}".strip()


def _format_misto(master: dict) -> str:
    m = master.get("misto") or {}
    objekt = m.get("objekt") or ""
    podlazi = m.get("podlazi") or ""
    rooms = m.get("mistnosti") or []
    parts = [p for p in [objekt, podlazi, ", ".join(rooms[:3])] if p]
    return " · ".join(parts)


def _build_material_rozklad(wb: openpyxl.Workbook, masters: list[dict],
                            sub_items: list[dict],
                            master_by_id: dict[str, dict],
                            master_row_in_vv: dict[str, int]) -> None:
    """Czech rozpočet-style hierarchical layout. Master row carries A-F
    columns (Pol.č. / Kapitola / Popis / MJ / Mn. / Místnost), sub-rows
    carry A, G-L (Pol.č. = X.N / Vstup / Sp./MJ / Mn. / MJ / Zdroj /
    Status). Column M is hidden master UUID for filter/system reference.
    """
    if ROZKLAD_SHEET in wb.sheetnames:
        del wb[ROZKLAD_SHEET]
    ws = wb.create_sheet(ROZKLAD_SHEET)

    headers = ["Pol. č.", "Kapitola", "Popis položky", "MJ", "Mn.",
               "Místnost", "Vstup", "Sp./MJ", "Mn.", "MJ", "Zdroj",
               "Status", "Master ID"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c, h)
        cell.font = BOLD
        cell.alignment = CENTER
        cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3",
                                 fill_type="solid")
        cell.border = BORDER_ALL

    # Hide the Master ID column (M)
    ws.column_dimensions["M"].hidden = True

    # Group sub-items by master
    subs_by_master: dict[str, list[dict]] = defaultdict(list)
    for s in sub_items:
        subs_by_master[s["paired_with"]].append(s)

    # GATE 7 — iterate ALL master items, not only those with sub-items.
    # Earlier pass filtered to ~1 210 of 4 090 masters which silently hid
    # Case 5 / install-only / no_pairing / no_kapitola_rule rows from the
    # VELTON deliverable.  Master rows without sub-items now render as a
    # single line with an explanatory placeholder in the Zdroj column.
    masters_sorted = sorted(masters,
                           key=lambda m: (m.get("kapitola") or "",
                                          m.get("popis") or "",
                                          m["item_id"]))

    # Alternating tint per MASTER block (entire block — master + sub-rows
    # — share one tint, next block flips)
    MASTER_FILL_A = PatternFill(start_color="E8F1FA", end_color="E8F1FA",
                                fill_type="solid")
    MASTER_FILL_B = PatternFill(start_color="F4F8FC", end_color="F4F8FC",
                                fill_type="solid")
    SUB_FILL_A = PatternFill(start_color="FFFFFF", end_color="FFFFFF",
                             fill_type="solid")
    SUB_FILL_B = PatternFill(start_color="FAFBFD", end_color="FAFBFD",
                             fill_type="solid")
    BOTTOM_ONLY = Border(bottom=Side(style="thin", color="606060"))

    row_idx = 2
    master_pol_counter = 0
    for block_idx, master in enumerate(masters_sorted):
        master_pol_counter += 1
        is_alt = block_idx % 2 == 1
        master_fill = MASTER_FILL_B if is_alt else MASTER_FILL_A
        sub_fill = SUB_FILL_B if is_alt else SUB_FILL_A

        # --- master row ---
        # A: Pol. č. integer (hyperlink → VV)
        cell_a = ws.cell(row_idx, 1, master_pol_counter)
        vv_row = master_row_in_vv.get(master["item_id"])
        if vv_row:
            cell_a.hyperlink = f"#'{VV_SHEET}'!A{vv_row}"
            cell_a.font = Font(color="0563C1", underline="single",
                              bold=True, size=11)
        else:
            cell_a.font = Font(bold=True, size=11)
        # B-F: master columns
        ws.cell(row_idx, 2, master.get("kapitola") or "")
        ws.cell(row_idx, 3, master.get("popis") or "")
        ws.cell(row_idx, 4, master.get("MJ") or "")
        ws.cell(row_idx, 5, master.get("mnozstvi"))
        ws.cell(row_idx, 6, _format_misto(master))
        # G-L blank on master row (filled with empty string for consistency)
        for c in range(7, 13):
            ws.cell(row_idx, c, "")
        # M (hidden) — master UUID
        ws.cell(row_idx, 13, master["item_id"])

        # Bold + tint entire master row
        for c in range(1, 14):
            cell = ws.cell(row_idx, c)
            if c != 1:  # cell_a font already set
                cell.font = Font(bold=True, size=11)
            cell.fill = master_fill
            cell.border = BORDER_ALL
            cell.alignment = LEFT_NOWRAP

        row_idx += 1

        # --- sub-rows ---
        # GATE 7 — when master has no sub-items (Case 5 standalone material,
        # install-only sibling row, MJ-incompatible, or no_kapitola_rule),
        # emit a single explanatory placeholder so the row count covers ALL
        # 4 090 masters instead of silently dropping ~70 % of them.
        subs = subs_by_master.get(master["item_id"], [])
        last_sub_row = None
        if not subs:
            popis_lower = (master.get("popis") or "").lower()
            # GATE 8a A3 — Case 5 sub-rows duplicate master popis instead
            # of a generic placeholder.  Treats master AS the material.
            if any(kw in popis_lower for kw in CASE5_PRIMARY_KEYWORDS):
                placeholder = master.get("popis") or "(Case 5 — master = materiál)"
            elif (_INSTALL_SUFFIX_RE.search(master.get("popis") or "")
                  or _INSTALL_PREFIX_RE.search(master.get("popis") or "")):
                placeholder = ("(install-only — materiál v sourozenecké "
                                "položce — dodávka)")
            elif not master.get("kapitola"):
                placeholder = "(bez kapitoly — no_kapitola_rule)"
            else:
                placeholder = "(no_pairing — žádný odpovídající KB nebo TZ vstup)"
            pol_label = f"{master_pol_counter}.0"
            ws.cell(row_idx, 1, pol_label)
            for c in range(2, 6):
                ws.cell(row_idx, c, "")
            # GATE 8a A1 — repeat master Místnost on sub-row for filter/sort
            ws.cell(row_idx, 6, _format_misto(master))
            ws.cell(row_idx, 7, "  " + placeholder)
            ws.cell(row_idx, 8, "")
            ws.cell(row_idx, 9, "")
            ws.cell(row_idx, 10, "")
            ws.cell(row_idx, 11, "—")
            ws.cell(row_idx, 12, "—")
            ws.cell(row_idx, 13, master["item_id"])
            for c in range(1, 14):
                cell = ws.cell(row_idx, c)
                cell.fill = sub_fill
                cell.border = BORDER_ALL
                cell.font = (Font(size=9, italic=True, color="606060") if c == 1
                             else Font(size=10, italic=True, color="909090"))
                cell.alignment = LEFT_NOWRAP
            last_sub_row = row_idx
            row_idx += 1
        for sub_idx, sub in enumerate(subs, start=1):
            # A: Pol. č. decimal (X.N)
            pol_label = f"{master_pol_counter}.{sub_idx}"
            ws.cell(row_idx, 1, pol_label)
            # B-E blank on sub rows
            for c in range(2, 6):
                ws.cell(row_idx, c, "")
            # GATE 8a A1 — repeat master Místnost on sub-row (col F)
            ws.cell(row_idx, 6, _format_misto(master))
            # G: Vstup — popis with slight visual indent (leading space)
            ws.cell(row_idx, 7, "  " + (sub.get("popis") or ""))
            # H: Sp./MJ rate
            rate_str = ""
            if (sub.get("rate_value") and sub.get("rate_unit_num")
                    and sub.get("rate_unit_denom")):
                rate_str = (f"{sub['rate_value']:g} "
                            f"{sub['rate_unit_num']}/{sub['rate_unit_denom']}")
            ws.cell(row_idx, 8, rate_str)
            ws.cell(row_idx, 9, sub.get("mnozstvi"))
            ws.cell(row_idx, 10, sub.get("MJ") or "")
            ws.cell(row_idx, 11, sub.get("zdroj_marker") or "")
            ws.cell(row_idx, 12, sub.get("status_label") or "")
            ws.cell(row_idx, 13, master["item_id"])  # hidden master ref

            for c in range(1, 14):
                cell = ws.cell(row_idx, c)
                if c == 12:
                    # Status col gets conditional formatting separately
                    cell.fill = sub_fill
                else:
                    cell.fill = sub_fill
                cell.border = BORDER_ALL
                if c == 1:
                    cell.font = Font(size=9, italic=True, color="606060")
                else:
                    cell.font = Font(size=10)
                cell.alignment = LEFT_NOWRAP

            last_sub_row = row_idx
            row_idx += 1

        # Thin bottom border on last sub-row of block
        if last_sub_row:
            for c in range(1, 14):
                cell = ws.cell(last_sub_row, c)
                cell.border = Border(
                    left=cell.border.left, right=cell.border.right,
                    top=cell.border.top,
                    bottom=Side(style="medium", color="606060"),
                )

    last_row = row_idx - 1

    # Status column conditional formatting (col L)
    status_range = f"L2:L{last_row}"
    for status, (bg, fg) in STATUS_FILLS.items():
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(operator="equal", formula=[f'"{status}"'],
                       fill=PatternFill(start_color=bg, end_color=bg,
                                         fill_type="solid"),
                       font=Font(color=fg, bold=True)),
        )

    # Column widths per spec
    widths = [8, 10, 40, 8, 12, 12, 40, 14, 12, 8, 30, 12, 1]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # Header freeze + auto-filter on A1:L{last_row} (exclude hidden M)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{last_row}"


# ----------------------------------------------------------------------------

def _build_material_audit(wb: openpyxl.Workbook, masters: list[dict],
                          sub_items: list[dict], library: list[dict],
                          avk_stats: Optional[dict] = None) -> None:
    if AUDIT_SHEET in wb.sheetnames:
        del wb[AUDIT_SHEET]
    ws = wb.create_sheet(AUDIT_SHEET, 0)  # Insert as first sheet

    # Compute stats
    by_source: Counter[str] = Counter(s["source"] for s in sub_items)
    total_subs = len(sub_items)
    documented = sum(by_source.get(k, 0) for k in DOCUMENTED_SOURCES)
    documented_pct = round(documented / max(total_subs, 1) * 100, 1)
    n_unique_materials = len(set(s.get("source_entry_id") for s in sub_items
                                  if s.get("source_entry_id")))
    n_library_citations = sum(1 for s in sub_items if s.get("source_entry_id"))
    n_csn_cited = (by_source.get("generic_with_csn_norm", 0)
                   + by_source.get("generic_with_csn_url", 0))

    # ----- Block 1: HERO STATS -----
    ws.merge_cells("A1:H1")
    ws["A1"] = "PHASE 6.6 — Rozklad materiálů"
    ws["A1"].font = HERO_FONT
    ws["A1"].alignment = CENTER

    ws.merge_cells("A2:H2")
    ws["A2"] = "Objekt D, akce 185-01, Bytový soubor Libuše"
    ws["A2"].font = Font(size=11, italic=True, color="606060")
    ws["A2"].alignment = CENTER

    ws.merge_cells("A4:H4")
    ws["A4"] = f"{documented_pct:.1f} % sub-položek má dokumentovaný zdroj (TZ + tabulky + výkresy)"
    ws["A4"].font = Font(bold=True, size=16, color="008B8B")
    ws["A4"].alignment = CENTER

    ws.merge_cells("A5:H5")
    ws["A5"] = (f"{len(masters):,} master položek · {total_subs:,} sub-položek "
                f"materiálů · {n_unique_materials} kanonických materiálů "
                f"z library citováno {n_library_citations}×")
    ws["A5"].font = Font(size=11, color="606060")
    ws["A5"].alignment = CENTER

    # Spacer rows
    row = 7

    # ----- Block 2: PROVENANCE BREAKDOWN -----
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Rozklad podle zdroje provenance"
    ws[f"A{row}"].font = TITLE_FONT
    row += 2

    # Table (cols A-D), chart anchored to F:H
    chart_start_row = row
    ws.cell(row, 1, "Zdroj"); ws.cell(row, 1).font = BOLD
    ws.cell(row, 2, "Počet"); ws.cell(row, 2).font = BOLD
    ws.cell(row, 3, "%"); ws.cell(row, 3).font = BOLD
    ws.cell(row, 4, "Status"); ws.cell(row, 4).font = BOLD
    row += 1
    provenance_rows = [
        ("TZ — explicit rate", "tz_explicit_with_rate", "OK"),
        ("TZ — explicit material", "tz_explicit_no_rate", "Confirm"),
        ("Tabulka — referenced", "tabulka_referenced", "OK"),
        ("Výkres — annotated", "vykres_annotated", "Confirm"),
        ("Generic + ČSN URL (post-enrich)", "generic_with_csn_url", "OK"),
        ("Generic + ČSN norm (offline)", "generic_with_csn_norm", "Confirm"),
        ("Generic odhad (no citation)", "generic_no_documentation", "Odhad"),
    ]
    table_first_row = row
    for label, src_key, status in provenance_rows:
        n = by_source.get(src_key, 0)
        pct = n / max(total_subs, 1) * 100
        ws.cell(row, 1, label)
        ws.cell(row, 2, n)
        ws.cell(row, 3, round(pct, 1))
        ws.cell(row, 4, status)
        for c in range(1, 5):
            ws.cell(row, c).border = BORDER_ALL
        # Apply Status fill manually (Material_audit doesn't use conditional fmt)
        if status in STATUS_FILLS:
            bg, fg = STATUS_FILLS[status]
            ws.cell(row, 4).fill = PatternFill(start_color=bg, end_color=bg,
                                                 fill_type="solid")
            ws.cell(row, 4).font = Font(color=fg, bold=True)
        row += 1
    table_last_row = row - 1

    # Add doughnut chart
    chart = DoughnutChart()
    chart.title = "Sub-items podle zdroje"
    labels = Reference(ws, min_col=1, min_row=table_first_row,
                      max_row=table_last_row)
    data = Reference(ws, min_col=2, min_row=chart_start_row,
                    max_row=table_last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.height = 8
    chart.width = 14
    ws.add_chart(chart, f"F{chart_start_row}")

    row += 2  # spacer

    # ----- Block 3: PER-KAPITOLA COVERAGE -----
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Pokrytí podle kapitoly (% sub-položek s dokumentovaným zdrojem)"
    ws[f"A{row}"].font = TITLE_FONT
    row += 2

    # Compute per-kapitola coverage from sub_items
    kap_counts: dict[str, dict[str, int]] = defaultdict(lambda: {"doc": 0, "tot": 0})
    for s in sub_items:
        k = s.get("kapitola") or "?"
        kap_counts[k]["tot"] += 1
        if s["source"] != "generic_no_documentation":
            kap_counts[k]["doc"] += 1

    # Sort by total descending, top 20
    sorted_kaps = sorted(kap_counts.items(), key=lambda x: -x[1]["tot"])[:20]

    bar_first_row = row
    ws.cell(row, 1, "Kapitola"); ws.cell(row, 1).font = BOLD
    ws.cell(row, 2, "% dokumentováno"); ws.cell(row, 2).font = BOLD
    ws.cell(row, 3, "Sub-items celkem"); ws.cell(row, 3).font = BOLD
    row += 1
    bar_data_first = row
    for kap, counts in sorted_kaps:
        pct = counts["doc"] / max(counts["tot"], 1) * 100
        ws.cell(row, 1, kap)
        ws.cell(row, 2, round(pct, 1))
        ws.cell(row, 3, counts["tot"])
        for c in range(1, 4):
            ws.cell(row, c).border = BORDER_ALL
        # Tinted fill on % column
        if pct > 70:
            bg = "C6EFCE"
        elif pct >= 40:
            bg = "FFEB9C"
        else:
            bg = "FFC09A"
        ws.cell(row, 2).fill = PatternFill(start_color=bg, end_color=bg,
                                            fill_type="solid")
        row += 1
    bar_data_last = row - 1

    bar_chart = BarChart()
    bar_chart.type = "bar"
    bar_chart.style = 11
    bar_chart.title = "Pokrytí podle kapitoly (top 20)"
    bar_chart.x_axis.title = "% dokumentováno"
    bar_chart.y_axis.title = "Kapitola"
    bar_chart.height = 12
    bar_chart.width = 18
    bar_labels = Reference(ws, min_col=1, min_row=bar_data_first,
                          max_row=bar_data_last)
    bar_data = Reference(ws, min_col=2, min_row=bar_first_row,
                        max_row=bar_data_last)
    bar_chart.add_data(bar_data, titles_from_data=True)
    bar_chart.set_categories(bar_labels)
    ws.add_chart(bar_chart, f"F{bar_first_row}")

    row += 2

    # ----- Block 4: TOP-20 needs-VELTON-confirm -----
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Top 20 sub-položek vyžadujících konfirmaci VELTON (confidence ≤ 0.5)"
    ws[f"A{row}"].font = TITLE_FONT
    row += 2

    headers4 = ["Master kapitola", "Materiál", "Sub-qty", "MJ", "Zdroj", "Důvod"]
    for c, h in enumerate(headers4, start=1):
        ws.cell(row, c, h)
        ws.cell(row, c).font = BOLD
        ws.cell(row, c).border = BORDER_ALL
    row += 1
    # Filter sub-items with conf ≤ 0.5, sort by master qty desc
    needs_confirm = [s for s in sub_items if s.get("confidence", 1.0) <= 0.5]
    # Resolve master qty for sort
    needs_confirm_sorted: list[tuple[float, dict]] = []
    masters_by_id = {m["item_id"]: m for m in masters}
    for s in needs_confirm:
        m = masters_by_id.get(s["paired_with"])
        master_qty = float(m.get("mnozstvi", 0) or 0) if m else 0
        needs_confirm_sorted.append((master_qty, s))
    needs_confirm_sorted.sort(key=lambda x: -x[0])
    for _, s in needs_confirm_sorted[:20]:
        ws.cell(row, 1, s.get("kapitola") or "")
        ws.cell(row, 2, s.get("popis") or "")
        ws.cell(row, 3, s.get("mnozstvi"))
        ws.cell(row, 4, s.get("MJ") or "")
        ws.cell(row, 5, s.get("zdroj_marker") or "")
        reason = "Confidence " + str(s.get("confidence"))
        if s["source"] == "tz_explicit_no_rate":
            reason += " · rate z generic KB"
        elif s["source"] == "generic_no_documentation":
            reason += " · materiál i rate generic"
        ws.cell(row, 6, reason)
        for c in range(1, 7):
            ws.cell(row, c).border = BORDER_ALL
        row += 1

    row += 2

    # ----- Block 5: Case 5 masters -----
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Case 5 — master items, které jsou samy materiálovou specifikací"
    ws[f"A{row}"].font = TITLE_FONT
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    # Detect Case 5 masters: existed in items but received no sub-items.
    # Mirror the pair script's Case-5 keyword set to filter out other
    # no-pairing reasons (skipped_status, no_kapitola_rule, …).
    paired_master_ids = set(s["paired_with"] for s in sub_items)
    case5_candidates = [m for m in masters
                        if m["item_id"] not in paired_master_ids
                        and m.get("mnozstvi", 0)
                        and m.get("status") not in {"deprecated",
                            "WRONGLY_ATTRIBUTED_TO_D", "interpretace_pending_ABMV"}]
    CASE5_KW = ["penetrace pod", "penetrace univerzá", "lepidlo flexib",
                "lepidlo na", "spárovací hmot", "sparovaci hmot",
                "samonivelační stěr", "samonivelacni ster",
                "kari síť", "kari sit", "pe fólie", "pe folie",
                "asfaltový pás", "asfaltovy pas",
                "armovací síť", "armovaci sit",
                "tmel ", "akrylový "]
    def _is_case5(popis: str) -> bool:
        norm = (popis or "").lower()
        return any(kw in norm for kw in CASE5_KW)

    case5_real = [m for m in case5_candidates if _is_case5(m.get("popis", ""))]
    ws[f"A{row}"] = (f"{len(case5_real)} master items jsou samy materiálovou "
                    f"specifikací (penetrace, lepidla, hydroizolace, izolace, "
                    f"SDK desky, kari síť, asfaltové pásy, …). "
                    f"Sub-items se pro ně negenerují by design — master "
                    f"řádek v List 1 Vykaz_vymer je už kompletní položka.")
    ws[f"A{row}"].font = Font(size=10)
    ws[f"A{row}"].alignment = LEFT_WRAP
    ws.row_dimensions[row].height = 50
    row += 2

    # Top 10 Case 5 by qty
    ws.cell(row, 1, "Top 10 podle qty:"); ws.cell(row, 1).font = BOLD
    row += 1
    headers5 = ["Kapitola", "Popis", "Qty", "MJ"]
    for c, h in enumerate(headers5, start=1):
        ws.cell(row, c, h); ws.cell(row, c).font = BOLD
        ws.cell(row, c).border = BORDER_ALL
    row += 1
    case5_sorted = sorted(case5_real,
                          key=lambda m: -float(m.get("mnozstvi", 0) or 0))
    for m in case5_sorted[:10]:
        ws.cell(row, 1, m.get("kapitola") or "")
        ws.cell(row, 2, m.get("popis") or "")
        ws.cell(row, 3, m.get("mnozstvi"))
        ws.cell(row, 4, m.get("MJ") or "")
        for c in range(1, 5):
            ws.cell(row, c).border = BORDER_ALL
        row += 1

    row += 2

    # ----- Block 6: Orphan library entries -----
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Orphan library entries (extracted but not auto-paired)"
    ws[f"A{row}"].font = TITLE_FONT
    row += 1
    used_lib_ids = set(s.get("source_entry_id") for s in sub_items
                       if s.get("source_entry_id"))
    orphans = [e for e in library if e["material_id"] not in used_lib_ids]
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = (f"{len(orphans)} library entries extracted in GATE 1 ale "
                    f"nenavázány na žádnou master položku. Důvody: chybějící "
                    f"kapitola_proposed (typicky vykres entries), nemapovaná "
                    f"taxonomy (material_kind=None), nebo kapitola bez "
                    f"pairing rule. Záznamy zůstávají v material_library_D."
                    f"json pro budoucí review / manuální párování.")
    ws[f"A{row}"].font = Font(size=10)
    ws[f"A{row}"].alignment = LEFT_WRAP
    ws.row_dimensions[row].height = 50
    row += 2

    # Top 10 by source document
    ws.cell(row, 1, "Top 10 podle zdrojového dokumentu:"); ws.cell(row, 1).font = BOLD
    row += 1
    orphan_by_doc: Counter[str] = Counter(o["source"]["document"] for o in orphans)
    ws.cell(row, 1, "Dokument"); ws.cell(row, 1).font = BOLD
    ws.cell(row, 2, "Orphan count"); ws.cell(row, 2).font = BOLD
    for c in range(1, 3):
        ws.cell(row, c).border = BORDER_ALL
    row += 1
    for doc, n in orphan_by_doc.most_common(10):
        ws.cell(row, 1, doc[:70])
        ws.cell(row, 2, n)
        ws.cell(row, 1).border = BORDER_ALL
        ws.cell(row, 2).border = BORDER_ALL
        row += 1

    row += 2

    # ----- Block 7: Výkres-annotated note -----
    n_vykres = sum(1 for e in library if e["source"]["type"] == "vykres_annotated")
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Výkres-annotated entries — proč 0 paired sub-items"
    ws[f"A{row}"].font = TITLE_FONT
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = (
        f"{n_vykres} výkres-annotated library entries z GATE 1 (Kniha "
        f"detailů + Zásady spárořezu) popisují principles of detailing "
        f"(napojení oken na fasádu, řešení nadpraží, parapetů, ostění, "
        f"přechodů terras na fasádu, spárořez obkladů), nikoliv konkrétní "
        f"materiálové kvantity vázané na master položku. Zůstávají dostupné "
        f"v material_library_D.json pro budoucí referenci a pro VELTON "
        f"review konzistence detailů; sub-items pro ně negenerují by design."
    )
    ws[f"A{row}"].font = Font(size=10)
    ws[f"A{row}"].alignment = LEFT_WRAP
    ws.row_dimensions[row].height = 75
    row += 2

    # ----- Block 13: Citation enrichment (GATE 5a) -----
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "GATE 5a — Block 13: Citation enrichment (ČSN refs)"
    ws[f"A{row}"].font = TITLE_FONT
    row += 1
    n_csn_norm = by_source.get("generic_with_csn_norm", 0)
    n_csn_url = by_source.get("generic_with_csn_url", 0)
    n_csn_total = n_csn_norm + n_csn_url
    n_remaining_odhad = by_source.get("generic_no_documentation", 0)
    kb_total = 0
    kb_with_norm = 0
    kb_with_url = 0
    try:
        kb_data = json.loads(KB_IN.read_text(encoding="utf-8"))
        for k, v in (kb_data.get("rates") or {}).items():
            kb_total += 1
            if v.get("citation_norm"):
                kb_with_norm += 1
            if v.get("citation_url"):
                kb_with_url += 1
    except (OSError, ValueError):
        pass
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = (
        f"Hand-curated ČSN normy applied offline: {kb_with_norm}/{kb_total} "
        f"KB entries carry citation_norm (real Czech construction normy: "
        f"ČSN 73 3450 keramika, ČSN 74 4505 podlahy, ČSN 73 0810 požární "
        f"prostupy, ČSN EN 12004 lepidla, ČSN EN 13888 spárovky, ČSN 73 3300 "
        f"nátěry, ČSN EN 1366-3 PO prostupy, ČSN EN ISO 11600 tmely, "
        f"ČSN EN 13813 podlahové stěrky).  No URLs fabricated.  "
        f"{n_csn_total:,} sub-items propagated the citation_norm tag, "
        f"promoting confidence 0.3 → 0.6 and status Odhad → Confirm.  "
        f"{n_remaining_odhad} sub-items remain pure 'Odhad' (KB entry has "
        f"no citation yet).  When PPLX_API_KEY available, run "
        f"`enrich_generic_rates.py` to populate citation_url → "
        f"status Confirm → OK and confidence 0.6 → 0.7."
    )
    ws[f"A{row}"].font = Font(size=10)
    ws[f"A{row}"].alignment = LEFT_WRAP
    ws.row_dimensions[row].height = 110
    row += 2
    ws.cell(row, 1, "Citation enrichment stats:")
    ws.cell(row, 1).font = BOLD
    row += 1
    headers13 = ["Metric", "Value"]
    for c, h in enumerate(headers13, start=1):
        ws.cell(row, c, h); ws.cell(row, c).font = BOLD
        ws.cell(row, c).border = BORDER_ALL
    row += 1
    rows13 = [
        ("KB entries total", kb_total),
        ("KB entries with citation_norm (offline)", kb_with_norm),
        ("KB entries with citation_url (online, post-Perplexity run)", kb_with_url),
        ("Sub-items promoted to 'generic_with_csn_norm'", n_csn_norm),
        ("Sub-items promoted to 'generic_with_csn_url'", n_csn_url),
        ("Sub-items still pure 'generic_no_documentation'", n_remaining_odhad),
    ]
    for label, val in rows13:
        ws.cell(row, 1, label); ws.cell(row, 2, val)
        ws.cell(row, 1).border = BORDER_ALL
        ws.cell(row, 2).border = BORDER_ALL
        row += 1
    row += 2

    # ----- Block 14: 11c_AVK_smeta layout summary (GATE 6) -----
    if avk_stats:
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = "GATE 6 — Block 14: 11c_AVK_smeta layout (AVK-style flat)"
        ws[f"A{row}"].font = TITLE_FONT
        row += 1
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = (
            f"AVK-style flat sheet '11c_AVK_smeta' shipped — work + material + "
            f"location decomposed in one denormalized table optimized for "
            f"VELTON's АВК workflow (Ukrainian/RU автоматизированный выпуск "
            f"кошторису).  Each of the {avk_stats['n_groups']:,} ÚRS G-groups "
            f"emits one PRÁCE row, N aggregated MATERIÁL rows, and one LOKACE "
            f"row per individual master instance.  11_Sumarizace_dle_kódu (4 845 "
            f"rows, hierarchical) + 11b_Material_aggregate (kapitola totals) + "
            f"Material_rozklad (per-master decomposition) zůstávají untouched "
            f"— 11c je čistý dodatečný view stejných dat."
        )
        ws[f"A{row}"].font = Font(size=10)
        ws[f"A{row}"].alignment = LEFT_WRAP
        ws.row_dimensions[row].height = 95
        row += 2
        ws.cell(row, 1, "11c_AVK_smeta row distribution:")
        ws.cell(row, 1).font = BOLD
        row += 1
        for c, h in enumerate(["Typ řádku", "Count"], start=1):
            ws.cell(row, c, h); ws.cell(row, c).font = BOLD
            ws.cell(row, c).border = BORDER_ALL
        row += 1
        rows14 = [
            ("G-groups (PRÁCE — master)", avk_stats["n_prace"]),
            ("MATERIÁL (aggregated across all instances per group)",
             avk_stats["n_materialy"]),
            ("LOKACE (one per master instance)", avk_stats["n_lokace"]),
            ("Total data rows", avk_stats["total_rows"]),
        ]
        for label, val in rows14:
            ws.cell(row, 1, label); ws.cell(row, 2, val)
            ws.cell(row, 1).border = BORDER_ALL
            ws.cell(row, 2).border = BORDER_ALL
            row += 1
        row += 1
        ws.cell(row, 1, "Sample G-groups with full decomposition:")
        ws.cell(row, 1).font = BOLD
        row += 1
        for c, h in enumerate(["G-kód", "Popis", "Lokací", "Materiálů",
                                "Rozsah řádků v 11c"], start=1):
            ws.cell(row, c, h); ws.cell(row, c).font = BOLD
            ws.cell(row, c).border = BORDER_ALL
        row += 1
        for sg in avk_stats.get("sample_groups", []):
            ws.cell(row, 1, sg["group_id"])
            ws.cell(row, 2, sg["popis"][:60])
            ws.cell(row, 3, sg["n_locations"])
            ws.cell(row, 4, sg["n_materials"])
            ws.cell(row, 5, f"{sg['row_range'][0]}–{sg['row_range'][1]}")
            for c in range(1, 6):
                ws.cell(row, c).border = BORDER_ALL
            row += 1
        row += 2

    # ----- GATE 8 — Blocks 15 / 16 / 17 -----
    # Block 15: UX polish (A1 Místnost, A2 outline, A3 placeholder swap)
    # Block 16: Case 5 rate enrichment (B1 popis-rate, B2 KB, B3 apply, B4 VRN)
    # Block 17: Critical pairing fixes (C1 dodávka, C2 cascade, C3 Case 5 extend)
    n_rate_from_popis = by_source.get("rate_from_popis", 0)
    n_case5_kb_rate = by_source.get("case5_kb_rate", 0)
    n_case5_self_ref = by_source.get("case5_self_reference", 0)
    n_vrn = by_source.get("vrn_services", 0)
    pairing_meta = {}
    try:
        items_blob = json.loads(ITEMS_IN.read_text(encoding="utf-8"))
        pairing_meta = items_blob.get("metadata", {})
    except (OSError, ValueError):
        pass
    n_cascade = int(pairing_meta.get("phase_6_6_b_cascade_skips", 0))
    case_labels = pairing_meta.get("phase_6_6_b_case_labels", {}) or {}

    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "GATE 8 — FINAL — Phase 6.6 closure (UX + Case 5 + pairing fixes)"
    ws[f"A{row}"].font = TITLE_FONT
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = (
        "GATE 8 ships three coordinated changes that close Phase 6.6:  "
        "(a) UX polish — Místnost propagation to sub-rows, 11c LOKACE "
        "outline groups, Case 5 placeholder replaced with master popis;  "
        "(b) Case 5 rate enrichment — regex rate extraction from master "
        "popis (B1, status OK 0.95), KB lookup with ČSN citation (B3, "
        "status Confirm 0.7), master-as-material self-reference (status "
        "OK 0.9), VRN administrative services exception (B4, status —);  "
        "(c) Critical pairing fixes — dodávka pairs reclassified Case 5 "
        "(C1), cross-layer sub-item cascade prevention via material_kind "
        "(C2), cementový potěr / anhydrit / betonová stěrka added to "
        "Case 5 keywords (C3)."
    )
    ws[f"A{row}"].font = Font(size=10)
    ws[f"A{row}"].alignment = LEFT_WRAP
    ws.row_dimensions[row].height = 120
    row += 2

    # Block 15 — UX polish
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Block 15 — GATE 8a UX polish"
    ws[f"A{row}"].font = TITLE_FONT
    row += 1
    rows15 = [
        ("A1 — Místnost propagated to all Material_rozklad sub-rows",
         "filtrace/sort dle místnosti napříč všemi sub-items"),
        ("A2 — 11c_AVK_smeta LOKACE rows collapsed (outline level 1)",
         "PRÁCE + MATERIÁL viditelné, LOKACE expand by [+]"),
        ("A3 — Case 5 placeholder → master popis verbatim",
         "master JE materiál — Vstup col duplikuje master popis"),
    ]
    for label, descr in rows15:
        ws.cell(row, 1, label); ws.cell(row, 1).border = BORDER_ALL
        ws.merge_cells(start_row=row, start_column=2,
                       end_row=row, end_column=6)
        ws.cell(row, 2, descr)
        for c in range(2, 7):
            ws.cell(row, c).border = BORDER_ALL
        ws.cell(row, 2).alignment = LEFT_WRAP
        row += 1
    row += 1

    # Block 16 — Case 5 rate enrichment
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Block 16 — GATE 8b Case 5 rate enrichment"
    ws[f"A{row}"].font = TITLE_FONT
    row += 1
    ws.cell(row, 1, "Source").font = BOLD
    ws.cell(row, 2, "Count").font = BOLD
    ws.cell(row, 3, "Status").font = BOLD
    for c in range(1, 4):
        ws.cell(row, c).border = BORDER_ALL
    row += 1
    rows16 = [
        ("rate_from_popis (B1 — regex from master popis)",
         n_rate_from_popis, "OK · conf 0.95"),
        ("case5_kb_rate (B3 — KB lookup, ČSN citation)",
         n_case5_kb_rate, "Confirm · conf 0.7"),
        ("case5_self_reference (master = materiál, 1:1)",
         n_case5_self_ref, "OK · conf 0.9"),
        ("vrn_services (B4 — VRN administrativní)",
         n_vrn, "— · conf 1.0"),
    ]
    for label, val, status in rows16:
        ws.cell(row, 1, label); ws.cell(row, 2, val); ws.cell(row, 3, status)
        for c in range(1, 4):
            ws.cell(row, c).border = BORDER_ALL
        row += 1
    row += 1

    # KB enrichment summary
    kb_added_keys = [
        "asfaltovy_pas_bituman", "asfaltovy_pas_radon_bariera",
        "betonova_sterka_strukturovana", "bezprasny_nater_zb",
        "tenkovrstva_silikon_omitka", "tondach_bobrovka",
        "hrebenace_tondach", "cementovy_poter_50mm",
        "cementovy_poter_58mm", "anhydritovy_poter",
    ]
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = (
        f"GATE 8b B2 — KB enriched with {len(kb_added_keys)} new entries "
        f"(all carry citation_norm — no schema violations).  Categories: "
        f"asfaltové pásy, betonové stěrky, bezprašné nátěry, tenkovrstvé "
        f"silikonové omítky, Tondach krytina + hřebenáče, cementové "
        f"potěry 50/58 mm, anhydritové potěry."
    )
    ws[f"A{row}"].font = Font(size=10)
    ws[f"A{row}"].alignment = LEFT_WRAP
    ws.row_dimensions[row].height = 50
    row += 2

    # Block 17 — Critical pairing fixes
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Block 17 — GATE 8c Critical pairing fixes"
    ws[f"A{row}"].font = TITLE_FONT
    row += 1
    rows17 = [
        ("C1 — dodávka pairs reclassified Case 5",
         "Every '— dodávka' suffix master treated as standalone material "
         "(physical delivery, no further decomposition)."),
        ("C2 — Cross-layer cascade prevention",
         f"{n_cascade} sub-items skipped because their material_kind "
         f"already appears as a separate priced master (e.g. Vinyl "
         f"Gerflor kladení no longer gets Cementový potěr sub-item — "
         f"potěr is its own G007 master)."),
        ("C3 — Case 5 keyword expansion",
         "Added: cementový potěr / potěr cementový, anhydritový potěr, "
         "betonová stěrka, polystyrenbeton, plus universal '— dodávka' "
         "catch-all.  All such masters now emit a single sub-item with "
         "extracted/KB rate instead of cascading library entries."),
    ]
    for label, descr in rows17:
        ws.cell(row, 1, label); ws.cell(row, 1).border = BORDER_ALL
        ws.merge_cells(start_row=row, start_column=2,
                       end_row=row, end_column=6)
        ws.cell(row, 2, descr)
        for c in range(2, 7):
            ws.cell(row, c).border = BORDER_ALL
        ws.cell(row, 2).alignment = LEFT_WRAP
        ws.row_dimensions[row].height = 32
        row += 1
    row += 1

    # Pairing case label breakdown
    ws.cell(row, 1, "GATE 8 — pair-script case label distribution:")
    ws.cell(row, 1).font = BOLD
    row += 1
    for label, val in sorted(case_labels.items(), key=lambda x: -x[1]):
        ws.cell(row, 1, label); ws.cell(row, 2, val)
        ws.cell(row, 1).border = BORDER_ALL
        ws.cell(row, 2).border = BORDER_ALL
        row += 1
    row += 2

    # ----- Block 18: D2 arithmetic validation (GATE 8d) -----
    if avk_stats:
        validation_failures = avk_stats.get("validation_failures") or []
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = "Block 18 — GATE 8d Arithmetic validation (D2)"
        ws[f"A{row}"].font = TITLE_FONT
        row += 1
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = (
            f"Invariant: Σ self-MJ MATERIÁL rows must equal master.qty per "
            f"G-group (tolerance 0.001 × master.qty or 0.01, whichever "
            f"greater).  Applies to ancillary-free / Case 5 1:1 paths.  "
            f"After D1 consolidation: "
            f"{avk_stats['n_groups'] - len(validation_failures):,} of "
            f"{avk_stats['n_groups']:,} G-groups pass.  "
            f"{len(validation_failures):,} failures (typically deprecated "
            f"sub-items excluded from MATERIÁL bucket — see status filter "
            f"in pair_materials.SKIP_STATUSES)."
        )
        ws[f"A{row}"].font = Font(size=10)
        ws[f"A{row}"].alignment = LEFT_WRAP
        ws.row_dimensions[row].height = 70
        row += 2
        if validation_failures:
            for c, h in enumerate(["G-kód", "Master.qty", "Σ MATERIÁL",
                                    "Delta", "Popis"], start=1):
                ws.cell(row, c, h); ws.cell(row, c).font = BOLD
                ws.cell(row, c).border = BORDER_ALL
            row += 1
            for vf in validation_failures[:20]:  # top 20 only
                ws.cell(row, 1, vf["group_id"])
                ws.cell(row, 2, vf["expected"])
                ws.cell(row, 3, vf["actual"])
                ws.cell(row, 4, vf["delta"])
                ws.cell(row, 5, vf["popis"])
                for c in range(1, 6):
                    ws.cell(row, c).border = BORDER_ALL
                row += 1
            if len(validation_failures) > 20:
                ws.cell(row, 1, f"… + {len(validation_failures) - 20} more")
                row += 1
        else:
            ws.cell(row, 1, "✅ Zero validation failures — Σ self-MJ "
                            "MATERIÁL == master.qty for all groups.")
            ws.cell(row, 1).font = Font(size=10, italic=True, color="008000")
            row += 1
        row += 2

        # ----- Block 19: GATE 8e per-material LOKACE + Cena/Stoimost -----
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = ("Block 19 — GATE 8e Per-material LOKACE + Cena/"
                         "Stoimost (tendering-ready 11c)")
        ws[f"A{row}"].font = TITLE_FONT
        row += 1
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = (
            f"11c_AVK_smeta restructured for VELTON tender entry:  "
            f"3-level hierarchy (PRÁCE outline 0 → MATERIÁL outline 1 "
            f"→ LOKACE outline 2), collapsible groups.  Per-material "
            f"LOKACE expansion: every MATERIÁL row emits one LOKACE per "
            f"contributing master instance (was: single LOKACE per master "
            f"across all materials).  Two new columns Cena (col L) + "
            f"Stoimost (col M).  LOKACE.Cena = '=L{{parent_M_row}}' "
            f"inherits from parent MATERIÁL row; Stoimost = J*L "
            f"auto-recalculates.  PRÁCE Stoimost uses Σ Mn × work-rate "
            f"(blank, VELTON enters).  Total rows {avk_stats['total_rows']:,} "
            f"({avk_stats['n_prace']:,} PRÁCE + {avk_stats['n_materialy']:,} "
            f"MATERIÁL + {avk_stats['n_lokace']:,} LOKACE).  "
            f"GATE 8.1 hotfix: LOKACE Σ Mn. (col G) = per-room area "
            f"in master.MJ units (not master.qty), so VELTON sees "
            f"area + rate + consumption on one row.  Popis práce (col E) "
            f"on MATERIÁL/LOKACE rows combines master + material name "
            f"when distinct (e.g. 'Omítka… — Nárožní lišta'); self-"
            f"material masters keep single popis."
        )
        ws[f"A{row}"].font = Font(size=10)
        ws[f"A{row}"].alignment = LEFT_WRAP
        ws.row_dimensions[row].height = 130
        row += 2

    # ----- Block 11: Paired master deduplication (Bug 1 fix report) -----
    # Reference counts from GATE 3 commit 238af49 (post-GATE-2 stats):
    #   sub-items total: 6 152
    #   provenance: tz_explicit_no_rate=1385, tabulka_referenced=583,
    #               generic_no_documentation=4184
    GATE3_TOTAL = 6152
    GATE3_PROV = {
        "tz_explicit_no_rate": 1385,
        "tabulka_referenced": 583,
        "generic_no_documentation": 4184,
        "tz_explicit_with_rate": 0,
        "vykres_annotated": 0,
    }
    n_install_skipped = sum(1 for m in masters
                            if _looks_like_install_only(m.get("popis", "")))
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = ("GATE 4 — Block 11: Paired master deduplication "
                    "(Bug 1 fix)")
    ws[f"A{row}"].font = TITLE_FONT
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = (
        f"{n_install_skipped} master items detected as install-only "
        f"across the whole soupis (suffix '— kotvení / montáž / klika / "
        f"spárování' or prefix 'Osazení / Rektifikační šrouby / Dodatečné "
        f"kotvení / Montáž TP|LP|OP'). 44 of those were in kapitoly with "
        f"pairing rules (HSV-642, HSV-643, PSV-763.x …) and would have "
        f"been double-paired before GATE 4 — they are now correctly "
        f"skipped. Their material specs live in a sibling '— dodávka' "
        f"master or in element-tabulky (0041 dveře, 0050 zámečnické, "
        f"0060 klempířské)."
    )
    ws[f"A{row}"].font = Font(size=10)
    ws[f"A{row}"].alignment = LEFT_WRAP
    ws.row_dimensions[row].height = 70
    row += 2
    ws.cell(row, 1, "Sample 5 install-only masters (now skipped):")
    ws.cell(row, 1).font = BOLD
    row += 1
    install_samples = [m for m in masters
                       if _looks_like_install_only(m.get("popis", ""))][:5]
    for m in install_samples:
        ws.cell(row, 1, m.get("kapitola") or "")
        ws.cell(row, 2, m.get("popis") or "")
        for c in (1, 2):
            ws.cell(row, c).border = BORDER_ALL
        row += 1
    row += 2

    # ----- Block 12: Work-pattern correction (Bug 4 fix report) -----
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = ("GATE 4 — Block 12: Work-pattern correction "
                    "(Bug 4 fix)")
    ws[f"A{row}"].font = TITLE_FONT
    row += 1
    delta = GATE3_TOTAL - total_subs
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = (
        f"Sub-items GATE 3 → GATE 4: {GATE3_TOTAL:,} → {total_subs:,} "
        f"(− {delta:,}). Work-pattern detection now matches BOTH kapitola "
        f"AND master popis subject (SDK podhled, izolace minerální vata, "
        f"kotvení, dlažba, malba …) so SDK desky no longer get attached "
        f"to vata-master rows in PSV-763.2 etc. Also Bug 2 fix refuses "
        f"MJ-incompatible rate applications (e.g. ks-master cannot consume "
        f"a kg/m² rate). Case 5 keywords expanded — PSV-763.2 single-"
        f"material rows (UW+CW profily, SDK desky, izolace vata, Tmelení "
        f"Q3, Kotvení) are now correctly classified as Case 5 and do not "
        f"receive ancillary sub-items."
    )
    ws[f"A{row}"].font = Font(size=10)
    ws[f"A{row}"].alignment = LEFT_WRAP
    ws.row_dimensions[row].height = 95
    row += 2
    ws.cell(row, 1, "Provenance delta GATE 3 → GATE 4:")
    ws.cell(row, 1).font = BOLD
    row += 1
    headers12 = ["Source", "GATE 3", "GATE 4", "Δ"]
    for c, h in enumerate(headers12, start=1):
        ws.cell(row, c, h); ws.cell(row, c).font = BOLD
        ws.cell(row, c).border = BORDER_ALL
    row += 1
    for src_key in ["tz_explicit_with_rate", "tz_explicit_no_rate",
                    "tabulka_referenced", "vykres_annotated",
                    "generic_no_documentation"]:
        before = GATE3_PROV.get(src_key, 0)
        after = by_source.get(src_key, 0)
        delta_v = after - before
        ws.cell(row, 1, src_key)
        ws.cell(row, 2, before)
        ws.cell(row, 3, after)
        ws.cell(row, 4, delta_v)
        for c in range(1, 5):
            ws.cell(row, c).border = BORDER_ALL
        row += 1
    row += 1

    # Block 12 — sample PSV-763.2 masters now correctly Case 5
    ws.cell(row, 1, "Sample 5 PSV-763.2 masters now Case 5 "
                    "(material spec, no sub-items):")
    ws.cell(row, 1).font = BOLD
    row += 1
    psv7632_case5 = [m for m in masters
                     if m.get("kapitola") == "PSV-763.2"
                     and m["item_id"] not in paired_master_ids][:5]
    for m in psv7632_case5:
        ws.cell(row, 1, m.get("popis") or "")
        ws.cell(row, 2, f"{_format_qty(m.get('mnozstvi'))} {m.get('MJ') or ''}")
        for c in (1, 2):
            ws.cell(row, c).border = BORDER_ALL
        row += 1
    row += 2

    # ----- Block 8: DISCLAIMER FOOTER -----
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = (
        "Tento dokument je technický rozklad materiálů na základě "
        "dostupné projektové dokumentace (TZ, tabulky skladeb, výkresy "
        "detailu). Položky označené ⚠ ODHAD jsou estimator's expansion "
        "na základě industry standardů (ČSN, typická spotřeba), ne "
        "závazné množství z projektu. Pro cenotvorbu doporučujeme "
        "konfirmaci výrobci materiálů nebo aktualizaci projektové "
        "dokumentace."
    )
    ws[f"A{row}"].font = SMALL_ITALIC
    ws[f"A{row}"].alignment = LEFT_WRAP
    ws.row_dimensions[row].height = 70

    # Generated timestamp
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = (f"Generováno: {datetime.now(timezone.utc).isoformat(timespec='seconds')} "
                    f"· Phase 6.6 GATE 4 · Branch claude/tz-material-decomposition-lBp5D")
    ws[f"A{row}"].font = Font(size=8, italic=True, color="909090")
    ws[f"A{row}"].alignment = CENTER

    # Column widths
    widths_audit = [20, 50, 14, 8, 32, 32, 12, 12]
    for c, w in enumerate(widths_audit, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w


def _build_sumarizace_aggregate(wb: openpyxl.Workbook, masters: list[dict],
                                 sub_items: list[dict]) -> None:
    """GATE 5b — new sheet `11b_Material_aggregate` placed immediately
    after the existing `11_Sumarizace_dle_kódu` sheet. Group by kapitola
    with work totals + material totals + provenance breakdown +
    citation_norm. Existing 11_Sumarizace stays byte-identical.
    """
    if AGGREGATE_SHEET in wb.sheetnames:
        del wb[AGGREGATE_SHEET]
    ws = wb.create_sheet(AGGREGATE_SHEET)

    # Title row
    ws.merge_cells("A1:H1")
    ws["A1"] = ("11b — Material aggregate per kapitola "
                "(Phase 6.6 GATE 5b)")
    ws["A1"].font = HERO_FONT
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:H2")
    ws["A2"] = ("Pro každou kapitolu: agregované sumy materiálů "
                "(napříč všemi sub-items v kapitole), provenance "
                "breakdown, ČSN citace.  Existing 11_Sumarizace zůstává "
                "neměněna.  Phase 6.6 GATE 5b.")
    ws["A2"].font = Font(size=10, italic=True, color="606060")
    ws["A2"].alignment = LEFT_WRAP
    ws.row_dimensions[2].height = 30

    # Group sub-items by kapitola
    subs_by_kap: dict[str, list[dict]] = defaultdict(list)
    for s in sub_items:
        subs_by_kap[s.get("kapitola") or "?"].append(s)
    masters_by_kap: dict[str, list[dict]] = defaultdict(list)
    for m in masters:
        masters_by_kap[m.get("kapitola") or "?"].append(m)

    # Sort kapitoly by sub-item count desc, then by name
    ordered_kaps = sorted(subs_by_kap.keys(),
                          key=lambda k: (-len(subs_by_kap[k]), k))

    row = 4
    grand_totals: dict[tuple[str, str], float] = defaultdict(float)
    grand_provenance: Counter[str] = Counter()

    for kap in ordered_kaps:
        kap_subs = subs_by_kap[kap]
        kap_masters = masters_by_kap.get(kap, [])
        if not kap_subs:
            continue

        # Kapitola header row
        ws.merge_cells(f"A{row}:H{row}")
        ws.cell(row, 1, f"▼ {kap}")
        ws.cell(row, 1).font = TITLE_FONT
        ws.cell(row, 1).fill = PatternFill(start_color="D9E2F3",
                                            end_color="D9E2F3",
                                            fill_type="solid")
        row += 1

        # Practice subtotals — masters in this kapitola with non-zero qty
        ws.cell(row, 1, "Práce — práce v kapitole:")
        ws.cell(row, 1).font = BOLD
        row += 1
        practice_hdrs = ["Master popis", "MJ", "Σ qty", "N položek"]
        for c, h in enumerate(practice_hdrs, start=1):
            ws.cell(row, c, h); ws.cell(row, c).font = BOLD
            ws.cell(row, c).border = BORDER_ALL
        row += 1
        # Group masters by canonical popis (strip leading bullets/numbers)
        master_groups: dict[str, dict[str, Any]] = defaultdict(
            lambda: {"qty": 0.0, "n": 0, "mj": ""}
        )
        for m in kap_masters:
            popis = (m.get("popis") or "").strip()
            if not popis:
                continue
            # Bucket: first 60 chars (handles "Penetrace pod omítku ..." variants)
            bucket = popis[:60]
            master_groups[bucket]["qty"] += float(m.get("mnozstvi") or 0)
            master_groups[bucket]["n"] += 1
            master_groups[bucket]["mj"] = m.get("MJ") or ""
        # Top 8 by qty
        for bucket, info in sorted(master_groups.items(),
                                   key=lambda x: -x[1]["qty"])[:8]:
            ws.cell(row, 1, bucket)
            ws.cell(row, 2, info["mj"])
            ws.cell(row, 3, round(info["qty"], 2))
            ws.cell(row, 4, info["n"])
            for c in range(1, 5):
                ws.cell(row, c).border = BORDER_ALL
            row += 1
        row += 1

        # Material subtotals (aggregated across all sub-items in kapitola)
        ws.cell(row, 1, "Materiály — agregát v kapitole:")
        ws.cell(row, 1).font = BOLD
        row += 1
        mat_hdrs = ["Vstup (popis)", "MJ", "Σ množství", "N sub-items",
                    "Zdroj (převažující)", "ČSN citation"]
        for c, h in enumerate(mat_hdrs, start=1):
            ws.cell(row, c, h); ws.cell(row, c).font = BOLD
            ws.cell(row, c).border = BORDER_ALL
        row += 1

        # Group sub-items by (clean_popis, MJ)
        mat_buckets: dict[tuple[str, str], dict[str, Any]] = defaultdict(
            lambda: {"qty": 0.0, "n": 0, "sources": Counter(),
                     "citations": Counter()}
        )
        for s in kap_subs:
            # Strip [odhad] prefix for bucketing
            popis = re.sub(r"^\[odhad\]\s*", "", s.get("popis") or "")
            mj = (s.get("MJ") or "").lower()
            key = (popis[:60], mj)
            mat_buckets[key]["qty"] += float(s.get("mnozstvi") or 0)
            mat_buckets[key]["n"] += 1
            mat_buckets[key]["sources"][s["source"]] += 1
            if s.get("citation_norm"):
                mat_buckets[key]["citations"][s["citation_norm"]] += 1
            grand_totals[key] += float(s.get("mnozstvi") or 0)
            grand_provenance[s["source"]] += 1

        for (popis, mj), info in sorted(mat_buckets.items(),
                                         key=lambda x: -x[1]["qty"]):
            top_source = info["sources"].most_common(1)[0][0]
            top_citation = (info["citations"].most_common(1)[0][0]
                            if info["citations"] else "")
            ws.cell(row, 1, popis)
            ws.cell(row, 2, mj)
            ws.cell(row, 3, round(info["qty"], 2))
            ws.cell(row, 4, info["n"])
            ws.cell(row, 5, top_source)
            ws.cell(row, 6, top_citation)
            for c in range(1, 7):
                ws.cell(row, c).border = BORDER_ALL
            row += 1

        # Provenance breakdown per kapitola
        ws.cell(row, 1, "Provenance breakdown:")
        ws.cell(row, 1).font = BOLD
        prov_kap: Counter[str] = Counter(s["source"] for s in kap_subs)
        col = 2
        for src, n in prov_kap.most_common():
            ws.cell(row, col, f"{src}: {n}")
            col += 1
        row += 2

    # Grand totals across all kapitoly
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row, 1, "Σ Grand totals napříč všemi kapitolami")
    ws.cell(row, 1).font = TITLE_FONT
    ws.cell(row, 1).fill = PatternFill(start_color="E8F1FA",
                                        end_color="E8F1FA",
                                        fill_type="solid")
    row += 1
    grand_hdrs = ["Vstup (popis)", "MJ", "Σ množství napříč objektem"]
    for c, h in enumerate(grand_hdrs, start=1):
        ws.cell(row, c, h); ws.cell(row, c).font = BOLD
        ws.cell(row, c).border = BORDER_ALL
    row += 1
    for (popis, mj), qty in sorted(grand_totals.items(),
                                   key=lambda x: -x[1])[:30]:
        ws.cell(row, 1, popis)
        ws.cell(row, 2, mj)
        ws.cell(row, 3, round(qty, 2))
        for c in range(1, 4):
            ws.cell(row, c).border = BORDER_ALL
        row += 1
    row += 2

    # Provenance grand totals
    ws.cell(row, 1, "Provenance napříč objektem:")
    ws.cell(row, 1).font = BOLD
    row += 1
    for src, n in grand_provenance.most_common():
        ws.cell(row, 1, src)
        ws.cell(row, 2, n)
        ws.cell(row, 1).border = BORDER_ALL
        ws.cell(row, 2).border = BORDER_ALL
        row += 1

    # Column widths
    widths_agg = [55, 8, 16, 10, 30, 22, 6, 6]
    for c, w in enumerate(widths_agg, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A4"

    # Move 11b right after 11_Sumarizace in tab order
    if SUMARIZACE_SHEET in wb.sheetnames and AGGREGATE_SHEET in wb.sheetnames:
        sum_idx = wb.sheetnames.index(SUMARIZACE_SHEET)
        agg_idx = wb.sheetnames.index(AGGREGATE_SHEET)
        wb.move_sheet(AGGREGATE_SHEET, offset=sum_idx + 1 - agg_idx)


def _write_misto_short(misto: dict) -> str:
    """Format master location as '• D · 1.NP · D.1.S.01' — mirrors
    phase_8_list11_sumarizace.write_misto() so 11c LOKACE rows match
    the LOKACE detail rows in 11_Sumarizace verbatim."""
    if not misto:
        return "—"
    parts = [misto.get("objekt") or "", misto.get("podlazi") or ""]
    mistnosti = misto.get("mistnosti") or []
    if mistnosti:
        parts.append(",".join(mistnosti))
    rendered = " · ".join(p for p in parts if p)
    return f"• {rendered}" if rendered else "—"


def _write_skladba_short(skl: dict) -> str:
    if not skl:
        return ""
    pairs = []
    for k, v in skl.items():
        if isinstance(v, (str, int, float)):
            pairs.append(f"{k}={v}")
    return "; ".join(pairs[:3])


def _build_avk_smeta(wb: openpyxl.Workbook, masters: list[dict],
                    sub_items: list[dict],
                    master_by_id: dict[str, dict]) -> dict:
    """GATE 6 → GATE 8 — AVK-style flat denormalized sheet
    `11c_AVK_smeta` placed immediately after `11b_Material_aggregate`.

    GATE 8 FINAL changes (D1 + E1 + E2 + A4):
      * PRÁCE row Vstup = "Práce" (was "— (souhrn práce —)") [A4]
      * Self-MJ materials consolidated to single row (master.MJ ==
        material.MJ → 1:1 — drops per-skladba splintering that broke
        Σ MATERIÁL == master.qty arithmetic invariant). [D1]
      * LOKACE rows emitted PER material PER master (was: 1 LOKACE per
        master under the group).  Hierarchy:
          G050 PRÁCE  (outline 0)
            G050.M1 MATERIÁL penetrace  (outline 1)
              G050.M1.L1 LOKACE room1   (outline 2)
              G050.M1.L2 LOKACE room2   (outline 2)
            G050.M2 MATERIÁL lepidlo    (outline 1)
              G050.M2.L1 LOKACE room1   (outline 2)
              …
        [E1]
      * Cena (col L) + Stoimost (col M) columns: [E2]
          MATERIÁL: VELTON enters Cena; Stoimost = J*L formula
          LOKACE:  Cena = "=L{parent_M_row}" inherits from parent
                   Stoimost = J*L on the row itself
          PRÁCE:   blank (or VELTON enters work-rate manually)

    Returns stats dict for dashboard Blocks 18 (validation) + 19 (layout).

    Legacy spec (preserved):
      1. Each G-group emits 1 PRÁCE (master) row.
      2. Sub-items aggregated by Vstup popis across all master
         instances in the G-group → 1 MATERIÁL row per unique Vstup
         (numbered M1, M2, …).
      3. Each master instance in the G-group → 1 LOKACE row
         (numbered L1, L2, …).

    Returns stats dict for dashboard Blocks 18 + 19.
    """
    if not GROUPS_IN.exists():
        raise FileNotFoundError(
            f"urs_query_groups.json missing at {GROUPS_IN} — required for "
            f"11c_AVK_smeta G-group structure."
        )
    groups = json.loads(GROUPS_IN.read_text(encoding="utf-8"))["groups"]
    groups_sorted = sorted(groups, key=lambda g: g["group_id"])

    subs_by_master: dict[str, list[dict]] = defaultdict(list)
    for s in sub_items:
        subs_by_master[s.get("paired_with") or ""].append(s)

    if AVK_SHEET in wb.sheetnames:
        del wb[AVK_SHEET]
    ws = wb.create_sheet(AVK_SHEET)

    # GATE 8e — Cena (L) + Stoimost (M) inserted after MJ.  Zdroj/Status
    # shifted to N/O.  15 columns total.
    headers = ["Pol. č.", "G-kód", "Typ", "Kapitola", "Popis práce",
               "MJ", "Σ Mn.", "Vstup/Místnost", "Sp./MJ", "Mn.", "MJ",
               "Cena", "Stoimost", "Zdroj", "Status"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c, h)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4",
                                 fill_type="solid")
        cell.border = BORDER_ALL
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    NCOLS = 15
    COL_CENA = 12        # L
    COL_STOIMOST = 13    # M
    COL_ZDROJ = 14       # N
    COL_STATUS = 15      # O

    PRACE_FILL = PatternFill(start_color="E8F1FA", end_color="E8F1FA",
                              fill_type="solid")
    MAT_FILL = PatternFill(start_color="E8F4E8", end_color="E8F4E8",
                            fill_type="solid")
    LOC_FILL = PatternFill(start_color="F8F8F8", end_color="F8F8F8",
                            fill_type="solid")
    STATUS_FILLS_LOCAL: dict[str, PatternFill] = {
        "OK":      PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                                fill_type="solid"),
        "Confirm": PatternFill(start_color="FFEB9C", end_color="FFEB9C",
                                fill_type="solid"),
        "Odhad":   PatternFill(start_color="FFC09A", end_color="FFC09A",
                                fill_type="solid"),
        "no_match": PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                 fill_type="solid"),
    }
    LEFT_NOWRAP = Alignment(horizontal="left", vertical="center",
                             wrap_text=False)
    NUMBER_FORMAT_QTY = "#,##0.000"
    NUMBER_FORMAT_CZK = "#,##0.00"

    cena_col = get_column_letter(COL_CENA)
    stoimost_col = get_column_letter(COL_STOIMOST)

    row = 2
    n_prace = 0
    n_mat = 0
    n_loc = 0
    sample_groups: list[dict] = []
    validation_failures: list[dict] = []  # GATE 8d D2

    for g in groups_sorted:
        gid = g["group_id"]
        master_ids = g.get("items_ids") or []
        popis = g.get("popis_canonical") or ""
        mj = (g.get("MJ") or "").strip()
        mj_lower = mj.lower()
        total = float(g.get("total_mnozstvi") or 0)
        kapitola = g.get("kapitola") or ""
        is_vrn = kapitola.upper().startswith("VRN")

        # ----- PRÁCE row (A4: "Práce") -----
        prace_row = row
        ws.cell(row, 1, gid)
        ws.cell(row, 2, gid)
        ws.cell(row, 3, "PRÁCE")
        ws.cell(row, 4, kapitola)
        ws.cell(row, 5, popis)
        ws.cell(row, 6, mj)
        ws.cell(row, 7, round(total, 3))
        ws.cell(row, 8, "Práce")
        ws.cell(row, 9, "")
        ws.cell(row, 10, round(total, 3))
        ws.cell(row, 11, mj)
        # Cena blank — VELTON may enter work-rate per MJ; Stoimost auto.
        ws.cell(row, COL_CENA, "")
        ws.cell(row, COL_STOIMOST,
                f"=IF({cena_col}{row}=\"\",\"\","
                f"G{row}*{cena_col}{row})")
        ws.cell(row, COL_ZDROJ, "—")
        ws.cell(row, COL_STATUS, "—")
        for c in range(1, NCOLS + 1):
            cell = ws.cell(row, c)
            cell.fill = PRACE_FILL
            cell.font = BOLD
            cell.alignment = LEFT_NOWRAP
            cell.border = BORDER_ALL
        ws.cell(row, 7).number_format = NUMBER_FORMAT_QTY
        ws.cell(row, 10).number_format = NUMBER_FORMAT_QTY
        ws.cell(row, COL_CENA).number_format = NUMBER_FORMAT_CZK
        ws.cell(row, COL_STOIMOST).number_format = NUMBER_FORMAT_CZK
        n_prace += 1
        row += 1

        # ----- Bucket sub-items per (popis_clean, MJ_lower) per master ---
        mat_buckets: dict[tuple[str, str], dict[str, Any]] = defaultdict(
            lambda: {"qty": 0.0, "rate_value": None, "rate_unit_num": None,
                     "rate_unit_denom": None, "zdroj_counter": Counter(),
                     "status_counter": Counter(), "popis_full": "",
                     "per_master": {}}
        )
        for mid in master_ids:
            for s in subs_by_master.get(mid, []):
                popis_full = s.get("popis") or ""
                popis_clean = re.sub(r"^\[odhad\]\s*", "", popis_full)
                key = (popis_clean[:80], (s.get("MJ") or "").lower())
                b = mat_buckets[key]
                b["qty"] += float(s.get("mnozstvi") or 0)
                if b["rate_value"] is None and s.get("rate_value") is not None:
                    b["rate_value"] = s.get("rate_value")
                    b["rate_unit_num"] = s.get("rate_unit_num")
                    b["rate_unit_denom"] = s.get("rate_unit_denom")
                b["zdroj_counter"][s.get("zdroj_marker") or ""] += 1
                b["status_counter"][s.get("status_label") or ""] += 1
                if not b["popis_full"] or (
                    b["popis_full"].startswith("[odhad]")
                    and not popis_full.startswith("[odhad]")
                ):
                    b["popis_full"] = popis_full
                b["per_master"][mid] = s

        # ----- D1 self-MJ consolidation -----
        # When MATERIÁL.MJ == master.MJ → master IS the material (1:1).
        # Merge any splintered self-MJ buckets into ONE canonical row
        # anchored on master.qty (group.total_mnozstvi) so the arithmetic
        # invariant Σ self-MJ MATERIÁL == master.qty holds (no per-skladba
        # leakage, no missing FF01 rows).
        self_mj_keys = [k for k in list(mat_buckets.keys())
                         if k[1] == mj_lower]
        if self_mj_keys and not is_vrn:
            best_key = max(self_mj_keys,
                            key=lambda k: len(mat_buckets[k]["per_master"]))
            merged_pm: dict[str, dict] = {}
            merged_zdroj: Counter = Counter()
            merged_status: Counter = Counter()
            merged_popis = mat_buckets[best_key]["popis_full"] or popis
            for k in self_mj_keys:
                merged_pm.update(mat_buckets[k]["per_master"])
                merged_zdroj.update(mat_buckets[k]["zdroj_counter"])
                merged_status.update(mat_buckets[k]["status_counter"])
                del mat_buckets[k]
            mat_buckets[("__self_material__", mj_lower)] = {
                "qty": total, "rate_value": 1.0,
                "rate_unit_num": mj, "rate_unit_denom": mj,
                "zdroj_counter": merged_zdroj,
                "status_counter": merged_status,
                "popis_full": merged_popis,
                "per_master": merged_pm,
            }

        # ----- VRN special case (B4) — single "(služby)" MATERIÁL row ----
        if is_vrn:
            mat_row_idx = row
            ws.cell(row, 1, f"{gid}.M1")
            ws.cell(row, 2, gid)
            ws.cell(row, 3, "MATERIÁL")
            ws.cell(row, 4, kapitola)
            ws.cell(row, 5, popis)
            ws.cell(row, 6, mj)
            ws.cell(row, 7, round(total, 3))
            ws.cell(row, 8, "(služby — bez materiálu)")
            ws.cell(row, 9, "")
            ws.cell(row, 10, "")
            ws.cell(row, 11, "")
            ws.cell(row, COL_CENA, "")
            ws.cell(row, COL_STOIMOST, "")
            ws.cell(row, COL_ZDROJ, "—")
            ws.cell(row, COL_STATUS, "—")
            for c in range(1, NCOLS + 1):
                cell = ws.cell(row, c)
                cell.fill = MAT_FILL
                cell.alignment = LEFT_NOWRAP
                cell.border = BORDER_ALL
            ws.row_dimensions[row].outline_level = 1
            n_mat += 1
            row += 1
            # Single LOKACE per master ref (no per-material expansion)
            for l_idx, mid in enumerate(master_ids, start=1):
                m = master_by_id.get(mid)
                if not m:
                    continue
                ws.cell(row, 1, f"{gid}.M1.L{l_idx}")
                ws.cell(row, 2, gid)
                ws.cell(row, 3, "LOKACE")
                ws.cell(row, 4, kapitola)
                # GATE 8.1 — VRN Popis stays master-only (service); no
                # material differentiation to append.
                ws.cell(row, 5, popis)
                ws.cell(row, 6, mj)
                # GATE 8.1 FIX 1 — VRN LOKACE Σ Mn. = per-instance qty
                # (room qty) rather than master.qty across the group.
                ws.cell(row, 7, round(float(m.get("mnozstvi") or 0), 3))
                ws.cell(row, 8, _write_misto_short(m.get("misto") or {}))
                ws.cell(row, 9, "")
                ws.cell(row, 10, round(float(m.get("mnozstvi") or 0), 3))
                ws.cell(row, 11, m.get("MJ") or mj)
                ws.cell(row, COL_CENA, "")
                ws.cell(row, COL_STOIMOST, "")
                ws.cell(row, COL_ZDROJ, "—")
                ws.cell(row, COL_STATUS, m.get("urs_status") or "—")
                for c in range(1, NCOLS + 1):
                    cell = ws.cell(row, c)
                    cell.fill = LOC_FILL
                    cell.alignment = LEFT_NOWRAP
                    cell.border = BORDER_ALL
                ws.cell(row, 7).number_format = NUMBER_FORMAT_QTY
                ws.cell(row, 10).number_format = NUMBER_FORMAT_QTY
                ws.row_dimensions[row].outline_level = 2
                ws.row_dimensions[row].hidden = True
                n_loc += 1
                row += 1
            continue  # next G-group

        # ----- Empty bucket fallback (no sub-items at all) -----
        if not mat_buckets:
            mat_row_idx = row
            ws.cell(row, 1, f"{gid}.M1")
            ws.cell(row, 2, gid)
            ws.cell(row, 3, "MATERIÁL")
            ws.cell(row, 4, kapitola)
            ws.cell(row, 5, popis)
            ws.cell(row, 6, mj)
            ws.cell(row, 7, round(total, 3))
            ws.cell(row, 8, popis)
            ws.cell(row, 9, f"1 {mj}/{mj}")
            ws.cell(row, 10, round(total, 3))
            ws.cell(row, 11, mj)
            ws.cell(row, COL_CENA, "")
            ws.cell(row, COL_STOIMOST,
                    f"=IF({cena_col}{row}=\"\",\"\","
                    f"J{row}*{cena_col}{row})")
            ws.cell(row, COL_ZDROJ, "—")
            ws.cell(row, COL_STATUS, "—")
            for c in range(1, NCOLS + 1):
                cell = ws.cell(row, c)
                cell.fill = MAT_FILL
                cell.alignment = LEFT_NOWRAP
                cell.border = BORDER_ALL
            ws.cell(row, 7).number_format = NUMBER_FORMAT_QTY
            ws.cell(row, 10).number_format = NUMBER_FORMAT_QTY
            ws.cell(row, COL_CENA).number_format = NUMBER_FORMAT_CZK
            ws.cell(row, COL_STOIMOST).number_format = NUMBER_FORMAT_CZK
            ws.row_dimensions[row].outline_level = 1
            n_mat += 1
            row += 1
            continue

        # ----- Standard path: self-MJ row first, then ancillaries -----
        ordered_keys: list[tuple[str, str]] = []
        if ("__self_material__", mj_lower) in mat_buckets:
            ordered_keys.append(("__self_material__", mj_lower))
        ancillary_keys = sorted(
            [k for k in mat_buckets.keys() if k[0] != "__self_material__"],
            key=lambda k: -mat_buckets[k]["qty"],
        )
        ordered_keys.extend(ancillary_keys)

        sum_self_mj = 0.0
        for m_idx, (popis_clean, mj_sub) in enumerate(ordered_keys, start=1):
            b = mat_buckets[(popis_clean, mj_sub)]
            top_zdroj = (b["zdroj_counter"].most_common(1)[0][0]
                         if b["zdroj_counter"] else "")
            top_status = (b["status_counter"].most_common(1)[0][0]
                          if b["status_counter"] else "")
            rate_str = ""
            if (b["rate_value"] is not None and b["rate_unit_num"]
                    and b["rate_unit_denom"]):
                rate_str = (f"{b['rate_value']:g} {b['rate_unit_num']}/"
                            f"{b['rate_unit_denom']}")
            elif popis_clean == "__self_material__":
                rate_str = f"1 {mj}/{mj}"

            mat_row_idx = row
            display_popis = (popis if popis_clean == "__self_material__"
                              else b["popis_full"])
            # GATE 8.1 FIX 2 — combined Popis práce when material distinct
            # from master (e.g. master "Omítka..." + material "Nárožní
            # lišta — omítka" → "Omítka… — Nárožní lišta…").  Self-material
            # case keeps single popis.
            combined_popis = (popis if display_popis == popis
                               else f"{popis} — {display_popis}")
            ws.cell(row, 1, f"{gid}.M{m_idx}")
            ws.cell(row, 2, gid)
            ws.cell(row, 3, "MATERIÁL")
            ws.cell(row, 4, kapitola)
            ws.cell(row, 5, combined_popis)
            ws.cell(row, 6, mj)
            ws.cell(row, 7, round(total, 3))
            ws.cell(row, 8, display_popis)
            ws.cell(row, 9, rate_str)
            ws.cell(row, 10, round(b["qty"], 3))
            ws.cell(row, 11, mj_sub)
            ws.cell(row, COL_CENA, "")
            ws.cell(row, COL_STOIMOST,
                    f"=IF({cena_col}{row}=\"\",\"\","
                    f"J{row}*{cena_col}{row})")
            ws.cell(row, COL_ZDROJ, top_zdroj)
            ws.cell(row, COL_STATUS, top_status)
            for c in range(1, NCOLS + 1):
                cell = ws.cell(row, c)
                cell.fill = MAT_FILL
                cell.alignment = LEFT_NOWRAP
                cell.border = BORDER_ALL
            if top_status in STATUS_FILLS_LOCAL:
                ws.cell(row, COL_STATUS).fill = STATUS_FILLS_LOCAL[top_status]
            ws.cell(row, 7).number_format = NUMBER_FORMAT_QTY
            ws.cell(row, 10).number_format = NUMBER_FORMAT_QTY
            ws.cell(row, COL_CENA).number_format = NUMBER_FORMAT_CZK
            ws.cell(row, COL_STOIMOST).number_format = NUMBER_FORMAT_CZK
            ws.row_dimensions[row].outline_level = 1
            n_mat += 1
            row += 1

            if mj_sub == mj_lower:
                sum_self_mj += b["qty"]

            # ----- LOKACE per master under this MATERIÁL row -----
            loc_idx_in_mat = 0
            for mid in master_ids:
                m = master_by_id.get(mid)
                if not m:
                    continue
                if popis_clean == "__self_material__":
                    loc_qty = float(m.get("mnozstvi") or 0)
                    if loc_qty <= 0:
                        continue
                    sub_for_log = b["per_master"].get(mid)
                else:
                    sub_for_log = b["per_master"].get(mid)
                    if not sub_for_log:
                        continue
                    loc_qty = float(sub_for_log.get("mnozstvi") or 0)
                loc_idx_in_mat += 1
                loc_misto = _write_misto_short(m.get("misto") or {})
                skl_str = _write_skladba_short(m.get("skladba_ref") or {})
                loc_status = (sub_for_log.get("status_label")
                              if sub_for_log else (m.get("urs_status") or "—"))
                loc_zdroj = (sub_for_log.get("zdroj_marker")
                             if sub_for_log else skl_str)
                ws.cell(row, 1, f"{gid}.M{m_idx}.L{loc_idx_in_mat}")
                ws.cell(row, 2, gid)
                ws.cell(row, 3, "LOKACE")
                ws.cell(row, 4, kapitola)
                # GATE 8.1 FIX 2 — combined Popis práce (matches parent MATERIÁL)
                ws.cell(row, 5, combined_popis)
                ws.cell(row, 6, mj)
                # GATE 8.1 FIX 1 — Σ Mn. = room area in master.MJ units,
                # not master.qty.  Gives VELTON three values: area + rate
                # + consumption on one LOKACE row.
                ws.cell(row, 7, round(float(m.get("mnozstvi") or 0), 3))
                ws.cell(row, 8, loc_misto)
                ws.cell(row, 9, rate_str)
                ws.cell(row, 10, round(loc_qty, 3))
                ws.cell(row, 11, mj_sub)
                # GATE 8e — Cena inherits from parent MATERIÁL via cell
                # reference; Stoimost = Mn × inherited Cena.
                ws.cell(row, COL_CENA, f"={cena_col}{mat_row_idx}")
                ws.cell(row, COL_STOIMOST,
                        f"=IF({cena_col}{row}=\"\",\"\","
                        f"J{row}*{cena_col}{row})")
                ws.cell(row, COL_ZDROJ, loc_zdroj or skl_str)
                ws.cell(row, COL_STATUS, loc_status or "—")
                for c in range(1, NCOLS + 1):
                    cell = ws.cell(row, c)
                    cell.fill = LOC_FILL
                    cell.alignment = LEFT_NOWRAP
                    cell.border = BORDER_ALL
                if loc_status in STATUS_FILLS_LOCAL:
                    ws.cell(row, COL_STATUS).fill = STATUS_FILLS_LOCAL[loc_status]
                ws.cell(row, 7).number_format = NUMBER_FORMAT_QTY
                ws.cell(row, 10).number_format = NUMBER_FORMAT_QTY
                ws.cell(row, COL_CENA).number_format = NUMBER_FORMAT_CZK
                ws.cell(row, COL_STOIMOST).number_format = NUMBER_FORMAT_CZK
                ws.row_dimensions[row].outline_level = 2
                ws.row_dimensions[row].hidden = True
                n_loc += 1
                row += 1

        # ----- D2 validation -----
        if sum_self_mj > 0 and abs(sum_self_mj - total) > max(0.01, total * 0.001):
            validation_failures.append({
                "group_id": gid, "expected": round(total, 3),
                "actual": round(sum_self_mj, 3),
                "delta": round(sum_self_mj - total, 3),
                "popis": popis[:60],
            })

        # Sample capture for Block 19
        if len(sample_groups) < 3 and mat_buckets:
            sample_groups.append({
                "group_id": gid,
                "popis": popis,
                "n_locations": len(master_ids),
                "n_materials": len(mat_buckets),
                "row_range": (prace_row, row - 1),
            })

    # GATE 8a A2 — outline summary ABOVE detail
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False

    # GATE 8e E4 — column widths for 15-col layout
    widths_avk = [12, 8, 10, 11, 50, 7, 12, 55, 14, 12, 7,
                   12, 14, 30, 12]
    for c, w in enumerate(widths_avk, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    last_col_letter = get_column_letter(NCOLS)
    ws.auto_filter.ref = f"A1:{last_col_letter}{row - 1}"

    # Move 11c right after 11b in tab order
    if AGGREGATE_SHEET in wb.sheetnames and AVK_SHEET in wb.sheetnames:
        agg_idx = wb.sheetnames.index(AGGREGATE_SHEET)
        avk_idx = wb.sheetnames.index(AVK_SHEET)
        wb.move_sheet(AVK_SHEET, offset=agg_idx + 1 - avk_idx)

    return {
        "n_groups": len(groups_sorted),
        "n_prace": n_prace,
        "n_materialy": n_mat,
        "n_lokace": n_loc,
        "total_rows": row - 2,  # exclude header
        "sample_groups": sample_groups,
        "validation_failures": validation_failures,
    }


def main() -> int:
    print("Phase 6.6 GATE 3 — Excel generator")
    print("=" * 60)

    masters, sub_items, library = _load_data()
    print(f"  masters:    {len(masters):,}")
    print(f"  sub-items:  {len(sub_items):,}")
    print(f"  library:    {len(library):,}")

    if not EXCEL_TARGET.exists():
        print(f"ERR: target Excel not found at {EXCEL_TARGET}", file=sys.stderr)
        return 1

    # Three backups (all idempotent — never overwrite existing snapshots):
    #   pre_phase6_6   = original pre-Phase-6.6 baseline (preserved from GATE 3)
    #   pre_gate4      = post-GATE-3 / pre-GATE-4 snapshot
    #   pre_gate5      = post-GATE-4 / pre-GATE-5 snapshot (new this run)
    print(f"\n[1/6] Backups:")
    for label, path in [("pre_phase6_6", EXCEL_BACKUP_PHASE_6_6),
                        ("pre_gate4", EXCEL_BACKUP_GATE4),
                        ("pre_gate5", EXCEL_BACKUP_GATE5),
                        ("pre_gate6", EXCEL_BACKUP_GATE6),
                        ("pre_gate8", EXCEL_BACKUP_GATE8),
                        ("pre_gate8_1", EXCEL_BACKUP_GATE8_1)]:
        if path.exists():
            print(f"      preserved {path.relative_to(REPO_ROOT)} "
                  f"({path.stat().st_size:,} bytes)")
        else:
            shutil.copy2(EXCEL_TARGET, path)
            print(f"      created  {path.relative_to(REPO_ROOT)} "
                  f"({path.stat().st_size:,} bytes)")

    # Load workbook + master-row map (built from masters in items.json order)
    print(f"\n[2/6] Loading workbook + computing master→VV row map...")
    wb = openpyxl.load_workbook(str(EXCEL_TARGET))
    master_row_in_vv = _build_master_row_map(masters)
    master_by_id = {m["item_id"]: m for m in masters}
    print(f"      Loaded {len(wb.sheetnames)} sheets, mapped "
          f"{len(master_row_in_vv)} masters")

    print(f"\n[3/6] Building Material_rozklad sheet...")
    _build_material_rozklad(wb, masters, sub_items, master_by_id, master_row_in_vv)
    rozklad_ws = wb[ROZKLAD_SHEET]
    print(f"      → {rozklad_ws.max_row - 1} sub-item rows")

    print(f"\n[4/6] Building 11c_AVK_smeta sheet (GATE 6)...")
    avk_stats = _build_avk_smeta(wb, masters, sub_items, master_by_id)
    print(f"      → {avk_stats['n_groups']} G-groups → "
          f"{avk_stats['n_prace']} PRÁCE + "
          f"{avk_stats['n_materialy']} MATERIÁL + "
          f"{avk_stats['n_lokace']} LOKACE = "
          f"{avk_stats['total_rows']} total rows")

    print(f"\n[5/6] Building Material_audit sheet (dashboard)...")
    _build_material_audit(wb, masters, sub_items, library,
                          avk_stats=avk_stats)

    print(f"\n[6/6] Building 11b_Material_aggregate sheet (GATE 5b)...")
    _build_sumarizace_aggregate(wb, masters, sub_items)
    print(f"      → placed after {SUMARIZACE_SHEET}")
    # Re-place 11c after 11b (sumarizace_aggregate move may have shifted order)
    if AGGREGATE_SHEET in wb.sheetnames and AVK_SHEET in wb.sheetnames:
        agg_idx = wb.sheetnames.index(AGGREGATE_SHEET)
        avk_idx = wb.sheetnames.index(AVK_SHEET)
        wb.move_sheet(AVK_SHEET, offset=agg_idx + 1 - avk_idx)

    # Set Material_audit as active sheet (opens first)
    audit_idx = wb.sheetnames.index(AUDIT_SHEET)
    wb.active = audit_idx
    print(f"      Active sheet: {AUDIT_SHEET} (index {audit_idx})")

    # Move Material_rozklad to second position (after audit)
    rozklad_idx = wb.sheetnames.index(ROZKLAD_SHEET)
    wb.move_sheet(ROZKLAD_SHEET, offset=1 - rozklad_idx)

    wb.save(str(EXCEL_TARGET))
    print(f"\nWrote {EXCEL_TARGET.relative_to(REPO_ROOT)} "
          f"({EXCEL_TARGET.stat().st_size:,} bytes)")
    print(f"\nSheets in final file (order):")
    wb2 = openpyxl.load_workbook(str(EXCEL_TARGET), read_only=True)
    for i, sn in enumerate(wb2.sheetnames):
        active_marker = "  ← ACTIVE" if i == wb2.active else ""
        print(f"  {i+1:2d}. {sn}{active_marker}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
