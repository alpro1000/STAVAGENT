"""
Export Service

Handles exporting project data to various formats:
- Excel (.xlsx) - Full project data with positions
- PDF - Summary report

Dependencies:
- openpyxl (Excel export)
- reportlab (PDF export)

Version: 1.0.0
Date: 2025-12-28
"""

import io
from datetime import datetime
from typing import Any, Dict, List, Optional
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.units import cm
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting project data to various formats."""

    def export_to_excel(
        self,
        project_name: str,
        positions: List[Dict[str, Any]],
        summary: Optional[Dict[str, Any]] = None,
    ) -> bytes:
        """
        Export project data to Excel format.

        Args:
            project_name: Name of the project
            positions: List of positions
            summary: Optional project summary

        Returns:
            Excel file as bytes

        Raises:
            ImportError: If openpyxl is not installed
        """
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is not installed. Install with: pip install openpyxl")

        wb = Workbook()

        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Souhrn"

        row = 1

        # Title
        ws_summary.cell(row, 1, "SOUHRN PROJEKTU")
        ws_summary.cell(row, 1).font = Font(size=16, bold=True)
        row += 2

        # Project info
        ws_summary.cell(row, 1, "Projekt:")
        ws_summary.cell(row, 1).font = Font(bold=True)
        ws_summary.cell(row, 2, project_name)
        row += 1

        ws_summary.cell(row, 1, "Datum exportu:")
        ws_summary.cell(row, 1).font = Font(bold=True)
        ws_summary.cell(row, 2, datetime.utcnow().strftime("%Y-%m-%d %H:%M"))
        row += 2

        # Summary data
        if summary:
            if summary.get('executive_summary'):
                ws_summary.cell(row, 1, "Souhrn:")
                ws_summary.cell(row, 1).font = Font(bold=True)
                row += 1
                ws_summary.cell(row, 1, summary['executive_summary'])
                ws_summary.cell(row, 1).alignment = Alignment(wrap_text=True)
                ws_summary.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                row += 2

            if summary.get('key_findings'):
                ws_summary.cell(row, 1, "Klíčová zjištění:")
                ws_summary.cell(row, 1).font = Font(bold=True)
                row += 1
                for finding in summary['key_findings']:
                    ws_summary.cell(row, 1, f"• {finding}")
                    ws_summary.cell(row, 1).alignment = Alignment(wrap_text=True)
                    ws_summary.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                    row += 1
                row += 1

            if summary.get('recommendations'):
                ws_summary.cell(row, 1, "Doporučení:")
                ws_summary.cell(row, 1).font = Font(bold=True)
                row += 1
                for rec in summary['recommendations']:
                    ws_summary.cell(row, 1, f"• {rec}")
                    ws_summary.cell(row, 1).alignment = Alignment(wrap_text=True)
                    ws_summary.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                    row += 1

        # Sheet 2: Positions
        ws_positions = wb.create_sheet(title="Pozice")

        # Headers
        headers = ["#", "Název položky", "Množství", "Jednotka", "Cena/MJ", "Celkem", "Zdrojový soubor"]
        header_fill = PatternFill(start_color="FF9F1C", end_color="FF9F1C", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws_positions.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data
        for row_idx, pos in enumerate(positions, 2):
            ws_positions.cell(row_idx, 1, row_idx - 1)
            ws_positions.cell(row_idx, 2, pos.get('item_name', 'N/A'))
            ws_positions.cell(row_idx, 3, pos.get('quantity', 0))
            ws_positions.cell(row_idx, 4, pos.get('unit', ''))
            ws_positions.cell(row_idx, 5, pos.get('unit_price', 0))
            ws_positions.cell(row_idx, 6, pos.get('total_price', 0))
            ws_positions.cell(row_idx, 7, pos.get('_source_file_name', 'N/A'))

        # Auto-width columns
        for col in range(1, len(headers) + 1):
            ws_positions.column_dimensions[get_column_letter(col)].width = 15

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"Exported {len(positions)} positions to Excel")
        return output.read()

    def export_summary_to_pdf(
        self,
        project_name: str,
        summary: Dict[str, Any],
        metadata: Optional[Dict[str, Any]] = None,
    ) -> bytes:
        """
        Export project summary to PDF format.

        Args:
            project_name: Name of the project
            summary: Project summary data
            metadata: Optional metadata (version info, etc.)

        Returns:
            PDF file as bytes

        Raises:
            ImportError: If reportlab is not installed
        """
        if not HAS_REPORTLAB:
            raise ImportError("reportlab is not installed. Install with: pip install reportlab")

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()

        # Title style
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#FF9F1C'),
            spaceAfter=30,
        )

        # Add title
        story.append(Paragraph("SOUHRN PROJEKTU", title_style))
        story.append(Spacer(1, 0.5 * cm))

        # Project info table
        info_data = [
            ["Projekt:", project_name],
            ["Datum exportu:", datetime.utcnow().strftime("%Y-%m-%d %H:%M")],
        ]

        if metadata:
            if metadata.get('version_number'):
                info_data.append(["Verze:", f"v{metadata['version_number']}"])
            if metadata.get('positions_count'):
                info_data.append(["Počet pozic:", str(metadata['positions_count'])])

        info_table = Table(info_data, colWidths=[4 * cm, 12 * cm])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 1 * cm))

        # Executive Summary
        if summary.get('executive_summary'):
            story.append(Paragraph("Souhrn", styles['Heading2']))
            story.append(Spacer(1, 0.3 * cm))
            story.append(Paragraph(summary['executive_summary'], styles['BodyText']))
            story.append(Spacer(1, 0.8 * cm))

        # Key Findings
        if summary.get('key_findings'):
            story.append(Paragraph("Klíčová zjištění", styles['Heading2']))
            story.append(Spacer(1, 0.3 * cm))
            for finding in summary['key_findings']:
                story.append(Paragraph(f"• {finding}", styles['BodyText']))
                story.append(Spacer(1, 0.2 * cm))
            story.append(Spacer(1, 0.5 * cm))

        # Recommendations
        if summary.get('recommendations'):
            story.append(Paragraph("Doporučení", styles['Heading2']))
            story.append(Spacer(1, 0.3 * cm))
            for rec in summary['recommendations']:
                story.append(Paragraph(f"• {rec}", styles['BodyText']))
                story.append(Spacer(1, 0.2 * cm))
            story.append(Spacer(1, 0.5 * cm))

        # Risk Assessment
        if summary.get('risk_assessment'):
            risk = summary['risk_assessment']
            risk_color = {
                'LOW': colors.green,
                'MEDIUM': colors.orange,
                'HIGH': colors.red,
            }.get(risk, colors.grey)

            story.append(Paragraph("Hodnocení rizika", styles['Heading2']))
            story.append(Spacer(1, 0.3 * cm))

            risk_style = ParagraphStyle(
                'RiskText',
                parent=styles['BodyText'],
                textColor=risk_color,
                fontSize=14,
                fontName='Helvetica-Bold',
            )
            story.append(Paragraph(risk, risk_style))
            story.append(Spacer(1, 0.5 * cm))

        # Cost Analysis
        if summary.get('cost_analysis'):
            cost_data = summary['cost_analysis']
            story.append(Paragraph("Analýza nákladů", styles['Heading2']))
            story.append(Spacer(1, 0.3 * cm))

            cost_table_data = []
            if cost_data.get('total_cost'):
                cost_table_data.append(["Celkové náklady:", f"{cost_data['total_cost']:,.2f} CZK"])
            if cost_data.get('labor_cost'):
                cost_table_data.append(["Pracovní náklady:", f"{cost_data['labor_cost']:,.2f} CZK"])
            if cost_data.get('material_cost'):
                cost_table_data.append(["Materiálové náklady:", f"{cost_data['material_cost']:,.2f} CZK"])

            if cost_table_data:
                cost_table = Table(cost_table_data, colWidths=[6 * cm, 8 * cm])
                cost_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ]))
                story.append(cost_table)

        # Build PDF
        doc.build(story)
        buffer.seek(0)

        logger.info(f"Exported summary to PDF for project '{project_name}'")
        return buffer.read()


# Singleton instance
_export_service: Optional[ExportService] = None


def get_export_service() -> ExportService:
    """Get the singleton export service instance."""
    global _export_service
    if _export_service is None:
        _export_service = ExportService()
    return _export_service
