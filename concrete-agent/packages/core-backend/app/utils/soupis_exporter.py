"""
Soupis prací XLSX Exporter (Tasks 34+35)

Exports assembled soupis prací to KROS-compatible xlsx format.
Follows existing pattern from excel_exporter.py (openpyxl).

Output columns (KROS compatible):
  P.č. | Typ | Kód | Popis | MJ | Množství | VV vzorec | Cen. soustava | Zdroj | Důvěra

Two sheets:
  1. Soupis prací — all positions sorted by HSV/PSV sections
  2. Rekapitulace — section summary (díly)
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

logger = logging.getLogger(__name__)

# Styles matching AuditExcelExporter conventions
HEADER_FILL = PatternFill(start_color="FF9F1C", end_color="FF9F1C", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11)
COMPANION_FILL = PatternFill(start_color="FFF8E7", end_color="FFF8E7", fill_type="solid")
BORDER = Border(
    bottom=Side(style='thin', color='E0E0E0'),
)

# Column definitions
COLUMNS = [
    ('P.č.', 6),
    ('Typ', 5),
    ('Kód', 12),
    ('Popis', 50),
    ('MJ', 6),
    ('Množství', 12),
    ('VV vzorec', 20),
    ('Cen. soustava', 14),
    ('Zdroj', 14),
    ('Důvěra', 8),
]


def generate_vv_formula(position: Dict[str, Any]) -> Optional[str]:
    """
    Generate VV (výkaz výměr) formula from position parameters.

    Task 34: Creates calculation formulas from extracted dimensions.
    E.g. "5,0*3,2*0,18" for a slab 5.0 x 3.2 x 0.18m = 2.88 m³
    """
    params = position.get('params', [])
    if not params:
        return None

    # Collect dimensional values
    dims = []
    for p in params:
        if isinstance(p, dict):
            ptype = p.get('type', '')
            value = p.get('normalized', p.get('value', ''))
        else:
            continue

        if ptype in ('volume', 'area'):
            # Already a computed value — use as-is
            return value
        elif ptype == 'thickness':
            # Extract numeric part
            num = _extract_number(value)
            if num is not None:
                # Convert mm to m if needed
                if 'mm' in str(value):
                    num = num / 1000
                dims.append(num)
        elif ptype == 'quantity':
            num = _extract_number(value)
            if num is not None:
                return str(num).replace('.', ',')

    if len(dims) >= 2:
        return '*'.join(str(d).replace('.', ',') for d in dims)
    elif len(dims) == 1:
        return str(dims[0]).replace('.', ',')

    return None


def _extract_number(value: str) -> Optional[float]:
    """Extract first number from a string like '0.18 m' or '300 mm'."""
    import re
    match = re.search(r'(\d+(?:[.,]\d+)?)', str(value))
    if match:
        return float(match.group(1).replace(',', '.'))
    return None


def export_soupis_xlsx(soupis_data: Dict[str, Any], output_path: Optional[Path] = None) -> bytes:
    """
    Export soupis prací to xlsx.

    Args:
        soupis_data: Output from soupis_to_dict() — positions, stats, warnings
        output_path: Optional file path. If None, returns bytes.

    Returns:
        bytes (xlsx content) if no output_path, else saves to file.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    positions = soupis_data.get('positions', [])

    # Sheet 1: Soupis prací
    ws = wb.create_sheet('Soupis prací')
    _write_soupis_sheet(ws, positions, soupis_data)

    # Sheet 2: Rekapitulace
    ws2 = wb.create_sheet('Rekapitulace')
    _write_rekapitulace_sheet(ws2, positions, soupis_data)

    if output_path:
        wb.save(output_path)
        logger.info(f"[SOUPIS-XLSX] Exported {len(positions)} positions to {output_path}")
        return None
    else:
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()


def _write_soupis_sheet(ws, positions: List[Dict], soupis_data: Dict):
    """Write main soupis sheet."""
    # Title row
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = 'SOUPIS PRACÍ'
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')

    # Metadata row
    ws.merge_cells('A2:J2')
    meta = ws['A2']
    stats = soupis_data.get('stats', {})
    meta.value = f"Celkem {stats.get('total_positions', 0)} položek | HSV: {stats.get('hsv_count', 0)} | PSV: {stats.get('psv_count', 0)} | Generováno: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    meta.font = Font(size=10, color='808080')

    # Header row (row 4)
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = col_width

    # Data rows
    current_section = None
    row = 5

    for pos in positions:
        # Section header
        section = pos.get('section')
        if section and section != current_section:
            current_section = section
            ws.merge_cells(f'A{row}:J{row}')
            section_cell = ws.cell(row=row, column=1, value=f'Díl: {section}')
            section_cell.fill = SECTION_FILL
            section_cell.font = SECTION_FONT
            row += 1

        # Position row
        is_companion = pos.get('source') == 'companion'
        vv = pos.get('vv_vzorec') or generate_vv_formula(pos)

        values = [
            pos.get('poradi', ''),
            pos.get('typ', 'HSV'),
            pos.get('kod', ''),
            pos.get('popis', ''),
            pos.get('mj', ''),
            pos.get('mnozstvi', ''),
            vv or '',
            pos.get('cenova_soustava', ''),
            pos.get('source', ''),
            f"{int(pos.get('confidence', 0) * 100)}%" if pos.get('confidence') else '',
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = BORDER
            if is_companion:
                cell.fill = COMPANION_FILL
            # Right-align numbers
            if col_idx in (1, 6, 10):
                cell.alignment = Alignment(horizontal='right')

        row += 1

    # Warnings
    warnings = soupis_data.get('warnings', [])
    if warnings:
        row += 1
        ws.cell(row=row, column=1, value='Upozornění:').font = Font(bold=True, color='CC0000')
        for w in warnings[:20]:
            row += 1
            ws.cell(row=row, column=1, value=f'⚠ {w}').font = Font(color='CC0000', size=9)

    # Attribution
    row += 2
    ws.merge_cells(f'A{row}:J{row}')
    attr = ws.cell(row=row, column=1)
    attr.value = soupis_data.get('attribution', 'Generováno systémem StavAgent')
    attr.font = Font(size=8, color='A0A0A0')


def _write_rekapitulace_sheet(ws, positions: List[Dict], soupis_data: Dict):
    """Write rekapitulace (section summary) sheet."""
    ws.cell(row=1, column=1, value='REKAPITULACE').font = Font(bold=True, size=14)

    # Headers
    headers = ['Díl', 'Název', 'Počet položek', 'S kódem', 'Bez kódu']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = [8, 30, 14, 10, 10][col_idx - 1]

    # Aggregate by section
    sections = {}
    for pos in positions:
        sec = pos.get('section') or pos.get('typ', 'HSV')
        if sec not in sections:
            sections[sec] = {'total': 0, 'with_code': 0, 'without_code': 0}
        sections[sec]['total'] += 1
        if pos.get('kod'):
            sections[sec]['with_code'] += 1
        else:
            sections[sec]['without_code'] += 1

    row = 4
    for sec, counts in sorted(sections.items()):
        ws.cell(row=row, column=1, value=sec)
        ws.cell(row=row, column=2, value=_section_name(sec))
        ws.cell(row=row, column=3, value=counts['total'])
        ws.cell(row=row, column=4, value=counts['with_code'])
        ws.cell(row=row, column=5, value=counts['without_code'])
        row += 1

    # Totals
    row += 1
    ws.cell(row=row, column=1, value='CELKEM').font = Font(bold=True)
    ws.cell(row=row, column=3, value=len(positions)).font = Font(bold=True)

    # Stats
    row += 2
    stats = soupis_data.get('stats', {})
    for key, val in stats.items():
        if isinstance(val, (int, float, str)):
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=str(val))
            row += 1


# HSV/PSV section names
_SECTION_NAMES = {
    '1': 'Zemní práce', '2': 'Zakládání', '3': 'Svislé konstrukce',
    '4': 'Vodorovné konstrukce', '5': 'Komunikace',
    '6': 'Úpravy povrchů, podlahy', '8': 'Trubní vedení',
    '9': 'Ostatní konstrukce, bourání',
    '711': 'Izolace proti vodě', '712': 'Povlakové krytiny',
    '713': 'Izolace tepelné', '720': 'Zdravotechnika',
    '762': 'Truhlářské konstrukce', '764': 'Klempířské práce',
    '766': 'Podlahy', '767': 'Zámečnické konstrukce',
    '783': 'Nátěry', '784': 'Malby',
    '998': 'Přesuny hmot',
    'HSV': 'Hlavní stavební výroba', 'PSV': 'Přidružená stavební výroba',
}


def _section_name(section: str) -> str:
    return _SECTION_NAMES.get(section, section)
