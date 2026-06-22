"""
Format Detector — auto-detects construction document format.

Checks: extension → signature → content markers.

Author: STAVAGENT Team
Version: 5.0.0
"""

import re
import logging
from pathlib import Path
from typing import Optional

from app.parsers.models import SourceFormat

logger = logging.getLogger(__name__)


def detect_format(file_path: str) -> SourceFormat:
    """Auto-detect file format."""
    path = Path(file_path)
    ext = path.suffix.lower()

    if ext == '.dxf':
        return SourceFormat.DXF
    if ext == '.dwg':
        return SourceFormat.DXF
    if ext == '.ifc':
        return SourceFormat.IFC

    if ext in ('.xlsx', '.xlsm', '.xls'):
        return _detect_xlsx_subtype(file_path)

    if ext == '.xml':
        return _detect_xml_subtype(file_path)

    if ext == '.pdf':
        return _detect_pdf_subtype(file_path)

    raise ValueError(f"Unknown format: {ext}")


def _detect_xlsx_subtype(file_path: str) -> SourceFormat:
    """
    Detect XLSX subtype:
    - 'Export Komplet' in A1 → XLSX_KOMPLET
    - '#RTSROZP#' in A1 → XLSX_RTSROZP
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True)
        for sheet_name in wb.sheetnames[:5]:
            ws = wb[sheet_name]
            first_cell = ws.cell(1, 1).value
            if first_cell == 'Export Komplet':
                wb.close()
                return SourceFormat.XLSX_KOMPLET
            if first_cell == '#RTSROZP#':
                wb.close()
                return SourceFormat.XLSX_RTSROZP
        wb.close()
    except Exception as e:
        logger.warning(f"XLSX detection failed: {e}")

    return SourceFormat.XLSX_KOMPLET


def _detect_xml_subtype(file_path: str) -> SourceFormat:
    """Detect XML subtype by root tag / source marker.

    Three XML families can all carry an ``<XC4>`` root and must stay distinct:
      - AspeEsticon XC4 export (soupis prací): ``<zdroj>AspeEsticon</zdroj>`` +
        ``<objekty>``/``<polozka>`` work items → ``XML_ASPE_XC4``.
      - KROS OTSKP price-list XC4: ``<CenoveSoustavy>`` price nodes → ``XML_OTSKP``.
      - KROS TSKP classification: ``<BuildingInformation>``/``<Classification>``
        → ``XML_TSKP``.
    """
    try:
        # Read the first few KB — enough for the root tag + <zdroj> marker.
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            head = f.read(4000)
        head_lower = head.lower()

        if '<xc4' in head_lower:
            # AspeEsticon export ships a <zdroj>AspeEsticon</zdroj> marker and
            # <objekty>/<polozka> work items; a genuine price-list has neither.
            if 'aspeesticon' in head_lower or '<objekty' in head_lower:
                return SourceFormat.XML_ASPE_XC4
            return SourceFormat.XML_OTSKP
        if '<buildinginformation' in head_lower or '<classification' in head_lower:
            return SourceFormat.XML_TSKP
    except Exception as e:
        logger.warning(f"XML detection failed: {e}")

    return SourceFormat.XML_OTSKP


def _detect_pdf_subtype(file_path: str) -> SourceFormat:
    """Detect PDF subtype: TZ (text) or tabular budget."""
    try:
        import pdfplumber
        with pdfplumber.open(file_path) as pdf:
            if not pdf.pages:
                return SourceFormat.PDF_TZ
            first_page = pdf.pages[0]
            tables = first_page.extract_tables()
            if tables:
                flat = str(tables)
                if re.search(r'\d{9}', flat):  # 9-digit URS code
                    return SourceFormat.PDF_ROZPOCET
    except Exception as e:
        logger.warning(f"PDF detection failed: {e}")

    return SourceFormat.PDF_TZ
