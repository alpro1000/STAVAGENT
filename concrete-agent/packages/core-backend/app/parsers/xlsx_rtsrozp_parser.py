"""
#RTSROZP# (RTS/Aspe) Parser

Format markers:
  - Cell A1 == '#RTSROZP#'
  - SO sheets: names containing 'Pol'

Column indices (0-based):
  col 0 (A) = P.c. (position number) or 'Dil:'
  col 1 (B) = URS code or chapter code
  col 2 (C) = Description
  col 3 (D) = Unit (MJ)
  col 4 (E) = Quantity
  col 5 (F) = Unit price
  col 6 (G) = Total price
  col 17 (R) = Chapter code ('800-1')
  col 18 (S) = Price source ('RTS 25/ I')
  col 32 (AG) = #TypZaznamu#: 'POL1_1', 'SPI', 'VV', 'POP'

SO identification (rows 2-5 of each sheet):
  row: col 0 = 'S:', col 1 = project_id, col 2 = project_name
  row: col 0 = 'O:', col 1 = so_id, col 2 = so_name
  row: col 0 = 'R:', col 1 = rozpocet_code, col 2 = rozpocet_name

Author: STAVAGENT Team
Version: 5.0.0
"""

import logging
from decimal import Decimal, InvalidOperation
from typing import Optional

from openpyxl import load_workbook

from app.parsers.models import (
    ParsedDocument, ParsedSO, ParsedChapter, ParsedPosition, SourceFormat,
)

logger = logging.getLogger(__name__)


def parse_xlsx_rtsrozp(file_path: str) -> ParsedDocument:
    """Parse #RTSROZP# (RTS/Aspe) format."""
    wb = load_workbook(file_path, read_only=True, data_only=True)

    doc = ParsedDocument(
        source_format=SourceFormat.XLSX_RTSROZP,
        source_file=file_path,
    )

    # Metadata from 'Stavba' sheet
    if 'Stavba' in wb.sheetnames:
        _extract_meta_stavba(doc, wb['Stavba'])

    # SO sheets — sheets containing 'Pol'
    so_sheets = [name for name in wb.sheetnames if 'Pol' in name]

    for sheet_name in so_sheets:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # SO metadata from header rows
        so_id = sheet_name
        so_name = sheet_name
        for row in rows[:6]:
            if len(row) < 3:
                continue
            if row[0] == 'O:' and row[1] and row[2]:
                so_id = str(row[1]).strip()
                so_name = str(row[2]).strip()
            if row[0] == 'S:':
                if row[1]:
                    doc.project_id = doc.project_id or str(row[1]).strip()
                if row[2]:
                    doc.project_name = doc.project_name or str(row[2]).strip()

        so = ParsedSO(
            so_id=so_id,
            so_name=so_name,
            source_sheet=sheet_name,
        )

        current_chapter: Optional[ParsedChapter] = None
        current_position: Optional[ParsedPosition] = None

        for i, row in enumerate(rows, 1):
            if i < 7 or len(row) < 5:
                continue

            row_typ = row[32] if len(row) > 32 else None

            # Chapter
            if row[0] == 'Díl:' or (isinstance(row[0], str) and row[0].startswith('Díl')):
                _flush_position(current_position, current_chapter)
                current_position = None
                if current_chapter:
                    so.chapters.append(current_chapter)

                current_chapter = ParsedChapter(
                    code=str(row[1]).strip() if row[1] is not None else '',
                    name=str(row[2]).strip() if row[2] else '',
                    so_id=so.so_id,
                )

            # Position
            elif _is_position_row(row, row_typ):
                _flush_position(current_position, current_chapter)

                current_position = ParsedPosition(
                    pc=_format_pc(row[0]),
                    chapter_code=current_chapter.code if current_chapter else None,
                    chapter_name=current_chapter.name if current_chapter else None,
                    so_id=so.so_id,
                    so_name=so.so_name,
                    code=str(row[1]).strip() if row[1] else None,
                    description=str(row[2]).strip() if row[2] else '',
                    unit=str(row[3]).strip() if row[3] else None,
                    quantity=_to_decimal(row[4]),
                    unit_price=_to_decimal(row[5]),
                    total_price=_to_decimal(row[6]) if len(row) > 6 else None,
                    price_source=str(row[18]).strip() if len(row) > 18 and row[18] else None,
                    source_format=SourceFormat.XLSX_RTSROZP,
                    source_sheet=sheet_name,
                    source_row=i,
                )

            # Specification (SPI)
            elif row_typ == 'SPI' and current_position and row[2]:
                spec = str(row[2]).strip()
                if current_position.specification:
                    current_position.specification += '\n' + spec
                else:
                    current_position.specification = spec

            # Quantity calculation (VV)
            elif row_typ == 'VV' and current_position:
                vv_text = str(row[2]).strip() if row[2] else ''
                vv_qty = _to_float(row[4])
                if vv_text or vv_qty is not None:
                    current_position.vv_lines.append({'text': vv_text, 'qty': vv_qty})

        # Flush remaining
        _flush_position(current_position, current_chapter)
        if current_chapter:
            so.chapters.append(current_chapter)

        if any(ch.positions for ch in so.chapters):
            doc.stavebni_objekty.append(so)

    wb.close()
    doc.positions_count = len(doc.all_positions)
    _calculate_coverage(doc)
    return doc


def _is_position_row(row, row_typ) -> bool:
    """Check if row is a position (work item)."""
    if row_typ == 'POL1_1':
        return True
    # Fallback: numeric PC + code-like string in col B
    if isinstance(row[0], (int, float)) and row[1] is not None:
        code = str(row[1]).strip()
        return len(code) >= 6 and any(c.isdigit() for c in code)
    return False


def _flush_position(position, chapter):
    """Add current position to chapter if both exist."""
    if position and chapter:
        chapter.positions.append(position)


def _format_pc(val) -> Optional[str]:
    if val is None:
        return None
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val)


def _extract_meta_stavba(doc: ParsedDocument, ws):
    """Extract project metadata from Stavba sheet."""
    for row in ws.iter_rows(max_row=15, values_only=True):
        if len(row) < 5:
            continue
        if row[1] == 'Stavba:' and row[4]:
            doc.project_name = str(row[4])
        if row[1] == 'Stavba:' and row[3]:
            doc.project_id = str(row[3])


def _to_decimal(val) -> Optional[Decimal]:
    if val is None:
        return None
    try:
        return Decimal(str(val))
    except (InvalidOperation, ValueError):
        return None


def _to_float(val) -> Optional[float]:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _calculate_coverage(doc: ParsedDocument):
    positions = doc.all_positions
    if not positions:
        doc.coverage_pct = 0.0
        return
    with_qty = sum(1 for p in positions if p.quantity is not None)
    doc.coverage_pct = round(with_qty / len(positions) * 100, 1)
