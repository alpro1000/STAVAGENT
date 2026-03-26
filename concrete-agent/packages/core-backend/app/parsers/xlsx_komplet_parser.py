"""
Export Komplet (KROS/RTS) Parser

Format markers:
  - Cell A1 == 'Export Komplet'
  - Row types in col D: 'D'=chapter, 'K'=position, 'PP'=spec, 'VV'=quantity calc, 'Online PSC'=URL

Column indices (0-based):
  col 2 (C) = PC (position number)
  col 3 (D) = Type: D/K/PP/VV/Online PSC
  col 4 (E) = Code (URS code or chapter code)
  col 5 (F) = Description
  col 6 (G) = Unit (MJ)
  col 7 (H) = Quantity
  col 8 (I) = Unit price
  col 9 (J) = Total price
  col 10 (K) = Price source ('CS URS 2025 02')

Author: STAVAGENT Team
Version: 5.0.0
"""

import logging
from decimal import Decimal, InvalidOperation
from typing import Optional, List

from openpyxl import load_workbook

from app.parsers.models import (
    ParsedDocument, ParsedSO, ParsedChapter, ParsedPosition, SourceFormat,
)

logger = logging.getLogger(__name__)


def parse_xlsx_komplet(file_path: str) -> ParsedDocument:
    """Parse Export Komplet format."""
    wb = load_workbook(file_path, read_only=True, data_only=True)

    doc = ParsedDocument(
        source_format=SourceFormat.XLSX_KOMPLET,
        source_file=file_path,
    )

    # Metadata from Rekapitulace stavby
    if 'Rekapitulace stavby' in wb.sheetnames:
        doc = _extract_meta(doc, wb['Rekapitulace stavby'])

    # SO sheets — all except Rekapitulace and VRN
    so_sheets = [
        name for name in wb.sheetnames
        if 'Rekapitulace' not in name and 'VRN' not in name
    ]

    for sheet_name in so_sheets:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        data_start = _find_data_start(rows)
        if data_start is None:
            doc.parser_warnings.append(f"Sheet '{sheet_name}': no data area found")
            continue

        so = ParsedSO(
            so_id=sheet_name.split(' - ')[0].strip() if ' - ' in sheet_name else sheet_name,
            so_name=sheet_name.split(' - ')[1].strip() if ' - ' in sheet_name else sheet_name,
            source_sheet=sheet_name,
        )

        current_chapter: Optional[ParsedChapter] = None
        current_position: Optional[ParsedPosition] = None

        for i, row in enumerate(rows[data_start:], start=data_start + 1):
            if len(row) < 6:
                continue
            typ = row[3]

            if typ == 'D':
                # Flush current position
                if current_position and current_chapter:
                    current_chapter.positions.append(current_position)
                    current_position = None
                if current_chapter:
                    so.chapters.append(current_chapter)

                current_chapter = ParsedChapter(
                    code=str(row[4]) if row[4] else '',
                    name=str(row[5]) if row[5] else '',
                    so_id=so.so_id,
                )

            elif typ == 'K':
                if current_position and current_chapter:
                    current_chapter.positions.append(current_position)

                current_position = ParsedPosition(
                    pc=str(row[2]) if row[2] else None,
                    chapter_code=current_chapter.code if current_chapter else None,
                    chapter_name=current_chapter.name if current_chapter else None,
                    so_id=so.so_id,
                    so_name=so.so_name,
                    code=str(row[4]) if row[4] else None,
                    description=str(row[5]) if row[5] else '',
                    unit=str(row[6]) if row[6] else None,
                    quantity=_to_decimal(row[7]),
                    unit_price=_to_decimal(row[8]) if len(row) > 8 else None,
                    total_price=_to_decimal(row[9]) if len(row) > 9 else None,
                    price_source=str(row[10]) if len(row) > 10 and row[10] else None,
                    source_format=SourceFormat.XLSX_KOMPLET,
                    source_sheet=sheet_name,
                    source_row=i,
                )

            elif typ == 'PP' and current_position:
                if row[5]:
                    spec = str(row[5])
                    if current_position.specification:
                        current_position.specification += '\n' + spec
                    else:
                        current_position.specification = spec

            elif typ == 'VV' and current_position:
                vv_text = str(row[5]) if row[5] else ''
                vv_qty = _to_float(row[7])
                if vv_text or vv_qty is not None:
                    current_position.vv_lines.append({
                        'text': vv_text, 'qty': vv_qty,
                    })

            elif typ == 'Online PSC' and current_position:
                if row[5]:
                    current_position.url = str(row[5])

        # Flush remaining
        if current_position and current_chapter:
            current_chapter.positions.append(current_position)
        if current_chapter:
            so.chapters.append(current_chapter)

        if so.chapters:
            doc.stavebni_objekty.append(so)

    wb.close()
    doc.positions_count = len(doc.all_positions)
    _calculate_coverage(doc)
    return doc


def _find_data_start(rows) -> Optional[int]:
    """Find first data row (row with 'D' or 'K' in col D, or header with 'PC')."""
    for i, row in enumerate(rows):
        if len(row) > 3 and row[3] in ('D', 'K'):
            return i
        if len(row) > 2 and row[2] == 'PČ':
            return i + 1
    return None


def _extract_meta(doc: ParsedDocument, ws) -> ParsedDocument:
    """Extract project metadata from Rekapitulace stavby sheet."""
    for row in ws.iter_rows(max_row=30, values_only=True):
        if len(row) < 11:
            continue
        label = str(row[3]) if row[3] else ''
        value = row[10]
        if 'Stavba:' in label and value:
            doc.project_name = str(value)
        elif 'Kód:' in label and value:
            doc.project_id = str(value)
        elif value and isinstance(value, str) and len(value) > 5:
            # Try to detect client name
            if not doc.client and any(kw in value for kw in ['s.r.o', 'a.s.', 'spol']):
                doc.client = value
    return doc


def _to_decimal(val) -> Optional[Decimal]:
    if val is None:
        return None
    try:
        return Decimal(str(val))
    except (InvalidOperation, ValueError):
        return None


def _to_float(val) -> Optional[float]:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _calculate_coverage(doc: ParsedDocument):
    positions = doc.all_positions
    if not positions:
        doc.coverage_pct = 0.0
        return
    with_qty = sum(1 for p in positions if p.quantity is not None)
    doc.coverage_pct = round(with_qty / len(positions) * 100, 1)
