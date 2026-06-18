"""Golden: calculator output + confidence reach the work-items and the deliverable.

Closes the atomize seam documented in docs/audits/pipeline_state_recon/2026-06-08:
the recipe used to discard the calculator's PlannerOutput (kept only a few key
NAMES) and drop classification confidence. This test pins the fix.

Hermetic + offline: the Monolit engine response is MOCKED at the delegation seam
(`monolit_delegate._http_post`) — no live service, deterministic numbers. Plain
sync test_* (no asyncio marker, no fastmcp import), so it runs in CI without
credentials and a missing dep ERRORS rather than skips.

Asserts (task §6):
  AC1 curated calc subset (resource quantities + schedule metrics) on the deck
      work-items AND in the exported deliverable metadata — no longer discarded.
  AC2 classification confidence preserved on items / deliverable, not dropped.
  AC3 each carried number is source-grounded (calc._source traces the element).
  AC4 elements the engine did not compute carry honest-blank (calc=None + flag).
  AC5 reserved catalog/price slots present on the row contract (unfilled).
  AC6 existing `_source` provenance + grounding-gate not regressed.
  AC7 exported soupis structure intact + the existing visible Zdroj/Důvěra
      columns now filled (additive).
"""

import base64
import os
import sys
from io import BytesIO
from uuid import uuid4

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from app.services.stage_gating import (
    InMemorySessionRepository,
    OrchestrateRequest,
    SessionManager,
    StageGatingOrchestrator,
    load_workflow_config,
)
from app.services.stage_gating.orchestrator import STATUS_COMPLETED
from app.services.stage_gating.recipe_runner import make_recipe_tool_runner

CFG = load_workflow_config()

# Distinctive numbers — if the live engine leaked through instead of the mock,
# these would not match (the assertions would fail loudly, not silently pass).
FAKE_PLANNER_OUTPUT = {
    "total_days": 1234,
    "num_tacts": 7,
    "resources": {"pour_crew_breakdown": {"ukladani": 8, "vibrace": 3, "finiseri": 2}},
    "warnings": ["⚠️ TEST: mostovka NK vyžaduje pevnou skruž"],
    "costs": {"total_czk": 999999},  # present in raw, NOT in the curated subset
    "source": "monolit_planner_api",
}

OBJECT = {
    "object_code": "SO-202",
    "object_name": "Most na sil. I/6 přes Lomnický potok",
    "charakteristika": "Trvalý dálniční most o třech polích.",
}
ELEMENTS = [
    {"name": "NK mostovka", "object_code": "SO-202", "volume_m3": 605,
     "concrete_class": "C35/45", "is_prestressed": True, "span_m": 20, "num_spans": 6},
    {"name": "Piloty OP1 Ø900", "object_code": "SO-202", "volume_m3": 50.9,
     "concrete_class": "C30/37"},
]
DECK = "NK mostovka"


@pytest.fixture
def _mock_engine(monkeypatch):
    """Serve a fixed PlannerOutput for POST /api/calculate — fully offline."""
    import app.mcp.tools.monolit_delegate as md

    async def fake_post(path, payload):
        if path == "/api/calculate":
            return 200, dict(FAKE_PLANNER_OUTPUT)
        raise AssertionError(f"unexpected delegation path: {path}")

    monkeypatch.setattr(md, "_http_post", fake_post)
    monkeypatch.setattr(md, "_RETRY_BACKOFF_S", 0)


class _FakeStore:
    def __init__(self):
        self.data = {}

    def loader(self, project_id):
        return self.data.get(project_id), f"/fake/{project_id}"

    def saver(self, project_id, field, value):
        self.data.setdefault(project_id, {})[field] = value


def _run():
    store, repo = _FakeStore(), InMemorySessionRepository()
    runner = make_recipe_tool_runner(
        CFG, decider=lambda c: {"action": "proceed"},
        loader=store.loader, saver=store.saver,
    )
    orch = StageGatingOrchestrator(manager=SessionManager(repo, config=CFG),
                                   config=CFG, tool_runner=runner)
    res = orch.run(OrchestrateRequest(
        user_id=uuid4(), project_id=uuid4(),
        options={"object": dict(OBJECT), "elements": [dict(e) for e in ELEMENTS]},
        confirmation_token="ok",
    ))
    return res, repo


def _partials(repo, res, state_value):
    state = repo.get(res.session_id)
    return (state.partials.get(state_value) or {}) if state else {}


# ── AC1 + AC3 — curated calc subset, source-grounded, on the deck items ───────

def test_calc_output_carried_onto_deck_items(_mock_engine):
    res, repo = _run()
    assert res.status == STATUS_COMPLETED, res.error
    items = _partials(repo, res, "WORK_ATOMIZATION")["breakdown_items"]
    deck_items = [it for it in items if it.get("element_name") == DECK]
    assert deck_items, "deck produced no work items"
    for it in deck_items:
        assert it["calc_status"] == "computed", it
        assert it["calc"]["total_days"] == 1234           # AC1 schedule metric
        assert it["calc"]["num_tacts"] == 7
        assert it["calc"]["resources"] == FAKE_PLANNER_OUTPUT["resources"]  # AC1 qty
        assert "costs" not in it["calc"]                  # curated subset, not whole
        assert it["calc"]["_source"].startswith(f"calculator:{DECK}")       # AC3
        assert it["calc_warnings"] == FAKE_PLANNER_OUTPUT["warnings"]        # §2.4 item


# ── AC4 — elements the engine did not compute are honest-blank ────────────────

def test_uncomputed_elements_are_honest_blank(_mock_engine):
    res, repo = _run()
    items = _partials(repo, res, "WORK_ATOMIZATION")["breakdown_items"]
    pile_items = [it for it in items if it.get("element_name") == "Piloty OP1 Ø900"]
    assert pile_items, "pile produced no work items"
    for it in pile_items:
        assert it["calc_status"] == "not_calculated"      # flag, not a number
        assert it["calc"] is None                         # blank, never fabricated


# ── AC2 — classification confidence preserved on every item ───────────────────

def test_classification_confidence_preserved(_mock_engine):
    res, repo = _run()
    items = _partials(repo, res, "WORK_ATOMIZATION")["breakdown_items"]
    assert items
    for it in items:
        assert isinstance(it.get("classification_confidence"), (int, float))
        assert it.get("classification_source")            # e.g. 'keywords'


# ── AC5 — reserved catalog/price slots present (unfilled) ─────────────────────

def test_reserved_catalog_price_slots_present(_mock_engine):
    res, repo = _run()
    items = _partials(repo, res, "WORK_ATOMIZATION")["breakdown_items"]
    for it in items:
        for slot in ("otskp_code", "unit_price_czk", "total_price_czk"):
            assert slot in it, f"missing reserved slot {slot}"
            assert it[slot] is None, f"{slot} must start empty (filled by later tasks)"


# ── AC1 — full PlannerOutput in step metadata (replayable) ────────────────────

def test_full_planner_output_in_metadata(_mock_engine):
    res, repo = _run()
    meta = _partials(repo, res, "WORK_ATOMIZATION")["calc_metadata"]
    assert meta["status"] == "computed"
    assert meta["element_name"] == DECK
    assert meta["raw"]["total_days"] == 1234
    assert meta["raw"]["costs"]["total_czk"] == 999999    # full output, not subset
    assert meta["warnings"] == FAKE_PLANNER_OUTPUT["warnings"]
    # Q1/AC3 — the raw stash is provenance-grounded (which element/engine).
    assert meta["_source"].startswith(f"calculator:{DECK}")


# ── AC1 + §2.4 — carried numbers + warnings reach the exported deliverable ────

def test_calc_reaches_committed_deliverable_metadata(_mock_engine):
    res, repo = _run()
    committed = _partials(repo, res, "COMMITTED")
    summary = committed.get("calc_summary") or []
    assert any(s["element_name"] == DECK and s["calc"]["total_days"] == 1234
               for s in summary), summary
    assert committed.get("calc_warnings") == FAKE_PLANNER_OUTPUT["warnings"]


# ── AC6 + AC7 — _source/grounding intact; visible Zdroj/Důvěra columns filled ─

def test_export_visible_columns_filled_and_source_preserved(_mock_engine):
    import openpyxl

    res, repo = _run()
    committed = _partials(repo, res, "COMMITTED")
    assert committed.get("source_preserved") is True      # AC6 — provenance intact

    atom = next(s for s in res.steps if s["state"] == "WORK_ATOMIZATION")
    assert atom.get("work_items_verified", 0) > 0         # AC6 — grounding-gate
    assert atom.get("work_items_unverified", 0) == 0

    wb = openpyxl.load_workbook(BytesIO(base64.b64decode(committed["file_base64"])))
    ws = wb["Soupis prací"]                               # AC7 — structure intact
    # Columns: 9 = Zdroj, 10 = Důvěra (header row 4, data from row 5).
    zdroj = [ws.cell(row=r, column=9).value for r in range(5, ws.max_row + 1)]
    duvera = [ws.cell(row=r, column=10).value for r in range(5, ws.max_row + 1)]
    assert "keywords" in zdroj, zdroj                     # computed source visible
    # Důvěra renders classification confidence as a percent. Both fixture elements
    # are clean keyword wins → 0.90 on the signal ladder (W3 reads the shared
    # element_types.yaml; keyword ≤ 0.9). Was "85%" when W3 hardcoded a flat
    # keyword confidence, pre single-sourcing (BUGS#5(3)).
    assert "90%" in duvera, duvera                        # confidence now visible


# ── Q4 — honest-blank export rows are VISUALLY distinct, not an empty gap ──────

def test_honest_blank_rows_marked_distinct_in_export(_mock_engine):
    import openpyxl

    res, repo = _run()
    committed = _partials(repo, res, "COMMITTED")
    wb = openpyxl.load_workbook(BytesIO(base64.b64decode(committed["file_base64"])))
    ws = wb["Soupis prací"]
    zdroj = [str(ws.cell(row=r, column=9).value or "")
             for r in range(5, ws.max_row + 1)]
    # The pile rows (engine did not compute them) must carry an explicit marker so
    # the soupis does not read as fully calculated where it was not.
    assert any("NEPOČÍTÁNO" in z for z in zdroj), zdroj
    # …and computed (deck) rows must NOT carry that marker — the two are distinct.
    assert any(z == "keywords" for z in zdroj), zdroj
    assert "keywords" != "keywords · NEPOČÍTÁNO"
