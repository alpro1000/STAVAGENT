"""
Golden Test: MCP tool `export_soupis` (Part B of TASK_MCP_DetectType_and_Export).

Promotes the soupis-prací assembly (today a script outside MCP) to a first MCP
deliverable tool. It wraps the SAME deterministic renderer used by the existing
REST /export-xlsx (app.utils.soupis_exporter.export_soupis_xlsx) — the render is
NOT forked. Provenance (`_source` from the work-breakdown items) is preserved in
the RESPONSE METADATA (source_preserved + source_map), NOT in the KROS soupis
columns — the KROS sheet stays clean and the format is not broken.

Skip-proof, like test_mcp_golden_so250b.py: sync test_* + asyncio.run, no
@pytest.mark.asyncio, no fastmcp/app.mcp.server import. Registration + config
assertions import their heavier deps INSIDE the test so a missing dep fails only
that test. #82 compares CELL VALUES, never file bytes (xlsx embeds timestamps).
#83 checks EXPORTED reachability at the CONFIG level (allow-list + YAML COMMITTED +
policy gateway) and renders by a direct tool call — no orchestrated session.

Criteria from §B of docs/tasks/TASK_MCP_DetectType_and_Export.md:

  #81 — given a structured item list → valid xlsx with all rows; `_source`
        preserved and traceable; no fabricated value or code.
  #82 — render is deterministic and replayable: same input → same rows/values; no LLM.
  #83 — the export terminal state is now reachable: tool registered in the right
        category and the export allow-list is filled — the formerly dead state
        declaration now has a real tool.
  #84 — registered, has a REST wrapper, is in the tool list; carries `_source`.
"""

import asyncio
import base64
import os
import sys
from io import BytesIO

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from app.mcp.tools.export import export_soupis

# Sample structured items — the shape produced by create_work_breakdown
# (work_first, code-less). Each carries a grounding `_source` (Pattern 29).
SAMPLE_ITEMS = [
    {
        "work_description": "Bednění Pilíř",
        "unit": "m²",
        "quantity": 66.0,
        "hsv_section": "HSV3",
        "element_name": "Pilíř P2",
        "element_type": "driky_piliru",
        "_source": "element:Pilíř P2 / template:Bednění {element}",
    },
    {
        "work_description": "Výztuž Pilíř z oceli B500B",
        "unit": "t",
        "quantity": 3.6,
        "hsv_section": "HSV4",
        "element_name": "Pilíř P2",
        "element_type": "driky_piliru",
        "_source": "element:Pilíř P2 / template:Výztuž {element} z oceli B500B",
    },
    {
        "work_description": "Beton Pilíř C30/37",
        "unit": "m³",
        "quantity": 24.0,
        "hsv_section": "HSV2",
        "element_name": "Pilíř P2",
        "element_type": "driky_piliru",
        "_source": "element:Pilíř P2 / template:Beton {element} {concrete_class}",
    },
]


def _load_soupis_sheet(file_base64: str):
    import openpyxl

    raw = base64.b64decode(file_base64)
    wb = openpyxl.load_workbook(BytesIO(raw))
    assert "Soupis prací" in wb.sheetnames, wb.sheetnames
    return wb, wb["Soupis prací"]


def _all_cell_strings(ws) -> list:
    out = []
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if v is not None:
                out.append(str(v))
    return out


# ── #81 — valid xlsx, all rows, _source preserved & traceable, nothing fabricated ─

def test_81_exports_valid_xlsx_with_all_rows():
    r = asyncio.run(export_soupis(items=SAMPLE_ITEMS))
    assert "error" not in r, r
    assert r["row_count"] == len(SAMPLE_ITEMS), r
    assert r["deliverable"] == "soupis_praci", r

    _, ws = _load_soupis_sheet(r["file_base64"])
    cells = _all_cell_strings(ws)
    # Every item's description appears as a row in the soupis sheet.
    for it in SAMPLE_ITEMS:
        assert it["work_description"] in cells, it["work_description"]


def test_81_source_preserved_in_metadata_not_in_kros_sheet():
    r = asyncio.run(export_soupis(items=SAMPLE_ITEMS))
    # Provenance preserved + traceable in the response metadata.
    assert r["source_preserved"] is True, r
    assert len(r["source_map"]) == len(SAMPLE_ITEMS), r["source_map"]
    for it in SAMPLE_ITEMS:
        assert it["_source"] in r["source_map"].values(), r["source_map"]

    # KROS sheet stays clean — the grounding `_source` strings must NOT appear in it.
    _, ws = _load_soupis_sheet(r["file_base64"])
    cells = _all_cell_strings(ws)
    for it in SAMPLE_ITEMS:
        assert it["_source"] not in cells, f"_source leaked into KROS sheet: {it['_source']}"


def test_81_no_fabricated_codes():
    # Input items have no catalog code → the Kód column must stay empty, never invented.
    r = asyncio.run(export_soupis(items=SAMPLE_ITEMS))
    _, ws = _load_soupis_sheet(r["file_base64"])
    # Column C = "Kód" (KROS layout). No data row may carry a fabricated code.
    kod_values = [row[2] for row in ws.iter_rows(min_row=5, values_only=True)]
    assert all((v is None or v == "") for v in kod_values), kod_values


# ── #82 — deterministic / replayable (compare CELL VALUES, not bytes) ─────────

def test_82_render_is_deterministic_by_cell_values():
    def rows(file_b64):
        _, ws = _load_soupis_sheet(file_b64)
        # Skip the metadata row (row 2) that embeds a generation timestamp.
        return [
            tuple(row)
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1)
            if i != 2
        ]

    a = asyncio.run(export_soupis(items=SAMPLE_ITEMS))
    b = asyncio.run(export_soupis(items=SAMPLE_ITEMS))
    assert rows(a["file_base64"]) == rows(b["file_base64"])


# ── #83 — EXPORTED reachable at the config level (allow-list + YAML + gateway) ─

def test_83_exported_state_reachable_via_config():
    from app.services.stage_gating.policy_gateway import (
        RE_EXPORT_ALLOW_LIST,
        evaluate_tool_policy,
    )
    from app.services.stage_gating.workflow_loader import load_workflow_config
    from app.services.stage_gating.workflow_state import WorkflowState

    # The formerly-dead export allow-list is now filled.
    assert "export_soupis" in RE_EXPORT_ALLOW_LIST

    cfg = load_workflow_config()
    # The tool is allowed in COMMITTED, and COMMITTED → EXPORTED is a real edge.
    assert "export_soupis" in cfg.tools_allowed_in(WorkflowState.COMMITTED)
    assert WorkflowState.EXPORTED in cfg.transitions[WorkflowState.COMMITTED]

    # The policy gateway lets the export tool run in the terminal COMMITTED state
    # precisely because it is on the re-export allow-list (otherwise SESSION_TERMINAL).
    decision = evaluate_tool_policy(
        tool_name="export_soupis",
        config=cfg,
        session_id="sess-1",
        current_state=WorkflowState.COMMITTED,
    )
    assert decision.allowed is True, decision


# ── #84 — registered, REST wrapper, in tool list, carries _source ─────────────

def test_84_registered_with_rest_wrapper_category_and_source():
    from app.mcp import routes as mcp_routes
    from app.mcp import auth as mcp_auth
    from app.services.stage_gating.tool_manifest import ToolCategory, get_manifest

    assert "export_soupis" in mcp_routes.TOOL_ORDER
    assert "export_soupis" in mcp_routes.TOOL_DESCRIPTIONS
    assert isinstance(mcp_auth.TOOL_COSTS.get("export_soupis"), int)
    paths = {getattr(r, "path", "") for r in mcp_routes.router.routes}
    assert any(p.endswith("/tools/export-soupis") for p in paths), sorted(paths)

    manifest = get_manifest("export_soupis")
    assert manifest is not None
    assert manifest.category == ToolCategory.RENDER

    # The tool response carries a `_source` (grounding gate).
    r = asyncio.run(export_soupis(items=SAMPLE_ITEMS))
    assert isinstance(r.get("_source"), str) and r["_source"], r
