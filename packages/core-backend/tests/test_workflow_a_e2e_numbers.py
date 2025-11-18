import asyncio
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import pytest
from openpyxl import Workbook, load_workbook

from app.parsers.excel_parser import ExcelParser
from app.services.audit_classifier import AuditClassifier
from app.utils.audit_contracts import build_audit_contract
from app.utils.excel_exporter import AuditExcelExporter
from app.validators import PositionValidator
from app.services.workflow_a import _classify_position


def _build_test_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Positions"
    ws.append(["Kód položky", "Popis", "MJ", "Množství", "Cena celkem"])

    quantities = [
        "1 200,50",
        "2\u00a0000,75",
        "3\u202f500,00",
    ]
    total_rows = 53
    for index in range(total_rows):
        code = f"VAL{index:03d}"
        quantity_token = quantities[index % len(quantities)]
        total_price = "7\u202f684,00" if index % 2 == 0 else "1\u00a0234,56"
        ws.append([code, f"Betonová pozice {index}", "m3", quantity_token, total_price])

    wb.save(path)


def test_e2e_excel_to_export(tmp_path):
    source_path = tmp_path / "eu_numbers.xlsx"
    _build_test_workbook(source_path)

    parser = ExcelParser()
    parsed = parser.parse(source_path)

    positions = parsed["positions"]
    assert len(positions) == 53
    assert parsed["diagnostics"]["normalization"]["numbers_locale"] == "EU"

    schema_validator = PositionValidator()
    schema_result = schema_validator.validate(positions)
    assert schema_result.stats["validated_total"] == 53
    assert schema_result.stats["invalid_total"] == 0

    prepared_positions = []
    for idx, position in enumerate(schema_result.positions):
        payload = dict(position)
        payload["code"] = payload.get("code") or f"VAL{idx:03d}"
        if idx < 48:
            payload["validation_status"] = "passed"
            payload["enrichment"] = {"match": "exact"}
            payload["enrichment_status"] = "matched"
            payload["unit_price"] = 120.0
        elif idx < 51:
            payload["validation_status"] = "warning"
            payload["enrichment"] = {"match": "partial"}
            payload["amber_reason"] = "partial_match"
            payload["enrichment_status"] = "partial"
            payload["unit_price"] = 80.0
        else:
            payload["validation_status"] = "failed"
            payload["enrichment"] = {"match": "none"}
            payload["notes"] = "manual review"
            payload["enrichment_status"] = "unmatched"
            payload["unit_price"] = None
        prepared_positions.append(payload)

    audit_classifier = AuditClassifier()
    audited_positions, audit_stats = audit_classifier.classify(prepared_positions)

    assert audit_stats["green"] == 48
    assert audit_stats["amber"] == 3
    assert audit_stats["red"] == 2

    enrichment_stats = {"matched": 48, "partial": 3, "unmatched": 2}
    validation_stats = {"passed": 48, "warning": 3, "failed": 2}

    contract = build_audit_contract(
        audited_positions,
        enrichment_stats=enrichment_stats,
        validation_stats=validation_stats,
        audit_stats=audit_stats,
        schema_stats=schema_result.stats,
        classify=_classify_position,
    )

    assert contract["totals"] == {"g": 48, "a": 3, "r": 2, "total": 53}

    exporter = AuditExcelExporter()
    output_path = tmp_path / "export.xlsx"
    project = {
        "project_id": "demo-project",
        "project_name": "Demo",
        "workflow": "A",
        "audit_results": contract,
    }

    exported_path = asyncio.run(exporter.export(project, output_path=output_path))
    assert exported_path.exists()

    workbook = load_workbook(exported_path)
    assert "Audit_Triage" in workbook.sheetnames
    assert "Positions" in workbook.sheetnames

    triage_sheet = workbook["Audit_Triage"]
    positions_sheet = workbook["Positions"]

    assert triage_sheet.max_row > 6
    assert positions_sheet.max_row > 2
