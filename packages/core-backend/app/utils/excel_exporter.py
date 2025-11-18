"""Excel exporter for audit results."""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.config import settings
from app.services.project_cache import _is_new_audit_format, _migrate_legacy_audit_results

logger = logging.getLogger(__name__)


class AuditExcelExporter:
    """Create a workbook with audit summary and full position listing."""

    SUMMARY_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FILL = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    CLASS_COLORS = {
        "GREEN": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "AMBER": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "RED": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }

    def __init__(self) -> None:
        self.workbook = openpyxl.Workbook()
        # Remove default sheet to control ordering explicitly
        default_sheet = self.workbook.active
        self.workbook.remove(default_sheet)

    async def export(self, project: Dict[str, Any], output_path: Path | None = None) -> Path:
        raw_audit = project.get("audit_results")
        is_new_contract = _is_new_audit_format(raw_audit)
        audit_results = _migrate_legacy_audit_results(raw_audit)

        self._create_summary_sheet(project, audit_results)
        self._create_positions_sheet(audit_results)

        if output_path is None:
            export_dir = settings.DATA_DIR / "exports" / project["project_id"]
            export_dir.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{project['project_name']}_audit_{timestamp}.xlsx"
            output_path = export_dir / filename

        self.workbook.save(output_path)

        totals = audit_results.get("totals", {})
        logger.info(
            "excel_export: positions=%d totals={g:%d,a:%d,r:%d} source_contract=%s",
            totals.get("total", 0),
            totals.get("g", 0),
            totals.get("a", 0),
            totals.get("r", 0),
            "normalized" if is_new_contract else "legacy-migrated",
        )

        return output_path

    # ------------------------------------------------------------------
    # Sheet builders
    # ------------------------------------------------------------------

    def _create_summary_sheet(self, project: Dict[str, Any], audit_results: Dict[str, Any]) -> None:
        sheet = self.workbook.create_sheet("Audit_Triage", 0)
        sheet["A1"] = "AUDIT SUMMARY"
        sheet["A1"].font = Font(bold=True, color="FFFFFF", size=15)
        sheet["A1"].fill = self.SUMMARY_FILL
        sheet.merge_cells("A1:D1")

        info_rows = [
            ("Project ID", project.get("project_id")),
            ("Project Name", project.get("project_name")),
            ("Workflow", project.get("workflow")),
            ("Generated", datetime.now().isoformat(timespec="seconds")),
            ("Enrichment", "Enabled" if project.get("enable_enrichment") else "Disabled"),
        ]

        row = 3
        for label, value in info_rows:
            sheet[f"A{row}"] = label
            sheet[f"A{row}"].font = Font(bold=True)
            sheet[f"B{row}"] = value
            row += 1

        row += 1
        sheet[f"A{row}"] = "Totals"
        sheet[f"A{row}"].font = Font(bold=True)
        sheet[f"A{row}"].fill = self.HEADER_FILL
        sheet.merge_cells(f"A{row}:D{row}")
        row += 1

        totals_block = audit_results.get("totals", {})
        totals = [
            ("Total positions", totals_block.get("total", 0), None),
            ("GREEN", totals_block.get("g", 0), self.CLASS_COLORS["GREEN"]),
            ("AMBER", totals_block.get("a", 0), self.CLASS_COLORS["AMBER"]),
            ("RED", totals_block.get("r", 0), self.CLASS_COLORS["RED"]),
        ]

        total_positions = totals_block.get("total", 0) or 1
        for label, value, fill in totals:
            sheet[f"A{row}"] = label
            sheet[f"A{row}"].font = Font(bold=True)
            sheet[f"B{row}"] = value
            if fill:
                sheet[f"B{row}"].fill = fill
            percentage = (value / total_positions) * 100 if total_positions else 0
            sheet[f"C{row}"] = f"{percentage:.1f}%"
            row += 1

        for col in ("A", "B", "C", "D"):
            sheet.column_dimensions[col].width = 24

    def _create_positions_sheet(self, audit_results: Dict[str, Any]) -> None:
        sheet = self.workbook.create_sheet("Positions")
        headers = [
            "Code",
            "Description",
            "Unit",
            "Quantity",
            "Status",
            "Issues",
            "Position ID",
            "Section",
            "Sheet",
            "Source",
        ]

        for column, title in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=column)
            cell.value = title
            cell.font = Font(bold=True, color="000000")
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        for row_index, position in enumerate(audit_results.get("items", []), start=2):
            provenance = position.get("provenance") or {}
            source = provenance.get("source")
            sheet.cell(row=row_index, column=1).value = position.get("code") or ""
            sheet.cell(row=row_index, column=2).value = position.get("description", "")
            sheet.cell(row=row_index, column=3).value = position.get("unit", "")
            sheet.cell(row=row_index, column=4).value = position.get("quantity")

            status = (position.get("status") or "").upper()
            status_cell = sheet.cell(row=row_index, column=5)
            status_cell.value = status
            status_cell.fill = self.CLASS_COLORS.get(status, PatternFill())

            issues = position.get("issues") or []
            issues_cell = sheet.cell(row=row_index, column=6)
            issues_cell.value = "\n".join(issues)
            issues_cell.alignment = Alignment(wrap_text=True, vertical="top")

            sheet.cell(row=row_index, column=7).value = provenance.get("position_id", "")
            sheet.cell(row=row_index, column=8).value = provenance.get("section", "")
            sheet.cell(row=row_index, column=9).value = provenance.get("sheet", "")
            sheet.cell(row=row_index, column=10).value = (
                ", ".join(f"{k}={v}" for k, v in source.items()) if isinstance(source, dict) else source or ""
            )

        column_widths = [14, 50, 10, 12, 12, 40, 18, 18, 20, 28]
        for index, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width

        sheet.freeze_panes = "A2"


async def export_enriched_results(project: Dict[str, Any]) -> Path:
    exporter = AuditExcelExporter()
    return await exporter.export(project)
